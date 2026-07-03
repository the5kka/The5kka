from __future__ import annotations

import copy
import random
import re
import sys
import json
import socket
import threading
import webbrowser
from dataclasses import dataclass, field
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage


APP_TITLE = "OJT EXAM MAKER - Python"
QUESTION_TYPES = ("공통", "객관식", "주관식")
DEFAULT_COUNTS = {"공통": 2, "객관식": 20, "주관식": 3}
DEFAULT_SCORES = {"공통": 2.5, "객관식": 4.0, "주관식": 5.0}

HTML = r"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>OJT EXAM MAKER - Python</title>
  <style>
    * { box-sizing: border-box; }
    body { margin:0; background:#0b1220; color:#e8eef8; font-family:"Malgun Gothic", Arial, sans-serif; }
    .app { min-height:100vh; padding:22px; display:grid; grid-template-rows:auto auto 1fr; gap:14px; }
    .top { display:flex; align-items:center; gap:16px; }
    h1 { margin:0; font-size:26px; }
    .badge { margin-left:auto; background:#102f22; color:#83ffc0; padding:10px 18px; font-weight:800; }
    .file { display:grid; grid-template-columns:90px 1fr 110px; gap:8px; align-items:center; background:#121c2e; padding:12px; border:1px solid #2a3b57; }
    input { background:#08111f; color:#fff; border:1px solid #5d6b80; padding:8px; font-weight:700; }
    button { background:#1e2d48; color:#fff; border:1px solid #dce7f7; padding:10px 14px; font-weight:800; cursor:pointer; }
    button.primary { background:#20d084; color:#04130d; border-color:#20d084; }
    .grid { display:grid; grid-template-columns:1.2fr 420px; gap:14px; min-height:0; }
    section { background:#121c2e; border:1px solid #2a3b57; padding:14px; min-height:0; }
    h2 { margin:0 0 10px; font-size:20px; }
    .tablewrap { overflow:auto; height:calc(100vh - 190px); }
    table { width:100%; border-collapse:collapse; font-size:13px; }
    th, td { border:1px solid #8391a8; padding:6px 8px; }
    th { background:#1e2d48; position:sticky; top:0; }
    tr.selected { background:#14513f; }
    .settings { display:grid; grid-template-columns:1fr 90px 90px; gap:10px; align-items:center; }
    .settings label { font-weight:800; }
    .total { margin:18px 0; padding:18px; background:#10243a; text-align:center; color:#20d084; font-size:34px; font-weight:900; }
    .status { white-space:pre-wrap; background:#07101d; padding:12px; min-height:150px; border:1px solid #2a3b57; color:#bcd1ff; }
  </style>
</head>
<body>
<div class="app">
  <div class="top"><h1>OJT EXAM MAKER - Python</h1><div id="badge" class="badge">READY</div></div>
  <div class="file">
    <div>문제은행</div>
    <input id="path" />
    <button onclick="loadBanks()">새로고침</button>
  </div>
  <div class="grid">
    <section>
      <h2>공정 선택</h2>
      <div class="tablewrap"><table><thead><tr><th>공정명</th><th>공통</th><th>객관식</th><th>주관식</th></tr></thead><tbody id="banks"></tbody></table></div>
    </section>
    <section>
      <h2>시험 조건</h2>
      <div class="settings">
        <label>공통</label><input id="count_common" value="2" oninput="updateTotal()"><input id="score_common" value="2.5" oninput="updateTotal()">
        <label>객관식</label><input id="count_choice" value="20" oninput="updateTotal()"><input id="score_choice" value="4.0" oninput="updateTotal()">
        <label>주관식</label><input id="count_subjective" value="3" oninput="updateTotal()"><input id="score_subjective" value="5.0" oninput="updateTotal()">
        <label>목표 점수</label><span></span><input id="target" value="100" oninput="updateTotal()">
      </div>
      <div id="total" class="total">TOTAL 100</div>
      <button class="primary" style="width:100%; margin-bottom:10px;" onclick="generate()">랜덤 시험지 생성</button>
      <div id="status" class="status">대기 중</div>
    </section>
  </div>
</div>
<script>
let bankList = [];
let selected = 0;
function n(id){ return Number(document.getElementById(id).value || 0); }
function badge(t, err=false){ const b=document.getElementById('badge'); b.textContent=t; b.style.background=err?'#3b1111':'#102f22'; b.style.color=err?'#ffb0b0':'#83ffc0';}
function log(t){ document.getElementById('status').textContent=t; }
function updateTotal(){
  const total=n('count_common')*n('score_common')+n('count_choice')*n('score_choice')+n('count_subjective')*n('score_subjective');
  document.getElementById('total').textContent='TOTAL '+(Number.isInteger(total)?total:total.toFixed(1));
}
function render(){
  const body=document.getElementById('banks'); body.innerHTML='';
  bankList.forEach((b,i)=>{ const tr=document.createElement('tr'); if(i===selected) tr.className='selected';
    tr.innerHTML=`<td>${b.name}</td><td>${b.counts['공통']}</td><td>${b.counts['객관식']}</td><td>${b.counts['주관식']}</td>`;
    tr.onclick=()=>{selected=i; render();}; body.appendChild(tr);
  });
}
async function loadBanks(){
  badge('LOADING'); log('문제은행 로딩 중...');
  const res=await fetch('/api/banks',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({path:document.getElementById('path').value})});
  const data=await res.json();
  if(!data.ok){ badge('ERROR',true); log(data.error); alert(data.error); return; }
  document.getElementById('path').value=data.path; bankList=data.banks; selected=0; render(); updateTotal(); badge('READY'); log(`로드 완료: ${bankList.length}개 공정`);
}
async function generate(){
  badge('RUNNING'); log('시험지 생성 중...');
  const payload={path:document.getElementById('path').value, selected,
    counts:{'공통':n('count_common'),'객관식':n('count_choice'),'주관식':n('count_subjective')},
    scores:{'공통':n('score_common'),'객관식':n('score_choice'),'주관식':n('score_subjective')},
    target:n('target')};
  const res=await fetch('/api/generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
  const data=await res.json();
  if(!data.ok){ badge('ERROR',true); log(data.error); alert(data.error); return; }
  badge('DONE'); log('생성 완료:\\n'+data.output); alert('시험지 생성 완료\\n\\n'+data.output);
}
loadBanks();
</script>
</body></html>
"""


@dataclass
class Question:
    sheet_name: str
    source_row: int
    source_no: str
    category: str
    exam_type: str
    question: str
    answer: str
    score: float
    images: list[object] = field(default_factory=list)


@dataclass
class QuestionBank:
    name: str
    questions: list[Question]
    meta: dict[str, str]

    @property
    def counts(self) -> dict[str, int]:
        return {kind: sum(1 for q in self.questions if q.exam_type == kind) for kind in QUESTION_TYPES}


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").replace("\xa0", " ")
    text = re.sub(r"[ \t]+\n", "\n", text)
    return text.strip()


def normalize(value: object) -> str:
    return re.sub(r"\s+", "", clean(value)).lower()


def safe_filename(text: str) -> str:
    text = re.sub(r'[\\/:*?"<>|]+', "_", text.strip())
    text = re.sub(r"\s+", " ", text)
    return text or "OJT"


def find_workbook() -> Path:
    candidates = [p for p in Path.cwd().glob("*.xlsm") if not p.name.startswith("~$")]
    if candidates:
        return candidates[0]
    return Path.cwd() / "OJT 시험 문제.xlsm"


def find_header(ws):
    for row in range(1, min(ws.max_row, 12) + 1):
        headers = {}
        for col in range(1, ws.max_column + 1):
            key = normalize(ws.cell(row, col).value)
            if key and key not in headers:
                headers[key] = col
        if "no" not in headers or "문제" not in headers:
            continue
        if "문제유형" not in headers and "문제유형" not in headers:
            continue
        if "답안" not in headers and "정답" not in headers:
            continue

        choice_cols = [headers[str(n)] for n in range(1, 7) if str(n) in headers]
        return {
            "row": row,
            "no": headers["no"],
            "score": headers.get("점수", 0),
            "category": headers.get("유형", 0),
            "exam_type": headers.get("문제유형", headers.get("문제유형")),
            "question": headers["문제"],
            "answer": headers.get("답안", headers.get("정답")),
            "choices": choice_cols,
        }
    return None


def build_question_text(ws, row: int, cols: dict) -> str:
    text = clean(ws.cell(row, cols["question"]).value)
    if not cols["choices"]:
        return text
    marks = ["①", "②", "③", "④", "⑤", "⑥"]
    choices = []
    for idx, col in enumerate(cols["choices"]):
        choice = clean(ws.cell(row, col).value)
        if choice:
            choices.append(f"{marks[idx]} {choice}")
    return text if not choices else text + "\n\n" + "\n".join(choices)


def load_settings(wb) -> dict[str, dict[str, str]]:
    ws = next((s for s in wb.worksheets if s.title.strip() == "시험 SETTING"), None)
    settings = {}
    if not ws:
        return settings
    for row in range(3, ws.max_row + 1):
        key = clean(ws.cell(row, 11).value)
        if not key:
            continue
        settings[key.strip()] = {
            "department": clean(ws.cell(row, 12).value),
            "evaluator": clean(ws.cell(row, 13).value),
            "job_name": clean(ws.cell(row, 14).value),
            "revision": clean(ws.cell(row, 15).value),
            "issue_date": clean(ws.cell(row, 16).value),
            "product_type": clean(ws.cell(row, 17).value),
        }
    return settings


def default_meta(name: str, settings: dict[str, dict[str, str]]) -> dict[str, str]:
    meta = dict(settings.get(name.strip(), {}))
    job_name = meta.get("job_name") or re.sub(r"\s+(일반용|전장용)\s*$", "", name).strip()
    product_type = meta.get("product_type")
    if not product_type:
        if "전장용" in name:
            product_type = "전장용"
        elif "일반용" in name or "일반" in name:
            product_type = "일반용"
        else:
            product_type = ""
    return {
        "department": meta.get("department") or "후가공",
        "evaluator": meta.get("evaluator") or "김용준",
        "job_name": job_name,
        "revision": meta.get("revision") or "0",
        "issue_date": meta.get("issue_date") or "2024.12.09",
        "product_type": product_type,
    }


def images_by_row(ws) -> dict[int, list[object]]:
    result: dict[int, list[object]] = {}
    for img in getattr(ws, "_images", []):
        try:
            row = img.anchor._from.row + 1
        except Exception:
            continue
        result.setdefault(row, []).append(img)
    return result


def load_banks(path: Path) -> list[QuestionBank]:
    wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
    settings = load_settings(wb)
    banks: list[QuestionBank] = []
    skip = {"시험 SETTING", "시험지", "답안지"}
    used_names: dict[str, int] = {}

    for ws in wb.worksheets:
        if ws.title.strip() in skip:
            continue
        cols = find_header(ws)
        if not cols:
            continue
        row_images = images_by_row(ws)
        questions: list[Question] = []
        for row in range(cols["row"] + 1, ws.max_row + 1):
            no = clean(ws.cell(row, cols["no"]).value)
            question = build_question_text(ws, row, cols)
            if not no or not question:
                continue
            exam_type = clean(ws.cell(row, cols["exam_type"]).value)
            if exam_type not in QUESTION_TYPES:
                continue
            score = ws.cell(row, cols["score"]).value if cols["score"] else None
            try:
                score_value = float(score)
            except Exception:
                score_value = DEFAULT_SCORES[exam_type]
            questions.append(
                Question(
                    sheet_name=ws.title,
                    source_row=row,
                    source_no=no,
                    category=clean(ws.cell(row, cols["category"]).value) if cols["category"] else "",
                    exam_type=exam_type,
                    question=question,
                    answer=clean(ws.cell(row, cols["answer"]).value),
                    score=score_value,
                    images=list(row_images.get(row, [])),
                )
            )
        if questions:
            name = ws.title.strip()
            used_names[name] = used_names.get(name, 0) + 1
            display = name if used_names[name] == 1 else f"{name} ({used_names[name]})"
            banks.append(QuestionBank(display, questions, default_meta(name, settings)))
    return sorted(banks, key=lambda b: b.name)


def clone_image(img) -> XLImage | None:
    try:
        data = img._data()
        return XLImage(BytesIO(data))
    except Exception:
        return None


def copy_row_format(ws, source_row: int, target_row: int, min_col: int = 2, max_col: int = 12):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(min_col, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy.copy(source.alignment)
        target.border = copy.copy(source.border)
        target.fill = copy.copy(source.fill)
        target.font = copy.copy(source.font)


def ensure_exam_capacity(ws, question_count: int):
    capacity = 45
    if question_count <= capacity:
        return
    extra = question_count - capacity
    ws.insert_rows(56, extra)
    for i in range(extra):
        row = 56 + i
        copy_row_format(ws, 55, row, 2, 12)
        try:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=11)
        except ValueError:
            pass


def clear_outputs(exam_ws, answer_ws, question_count: int):
    max_row = max(55, 10 + question_count)
    for row in range(11, max_row + 1):
        for col in range(2, 13):
            cell = exam_ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None
    for row in range(3, max(28, question_count + 4)):
        for col in range(2, 7):
            answer_ws.cell(row, col).value = None
    exam_ws._images = []


def write_settings(wb, bank: QuestionBank, counts: dict[str, int], scores: dict[str, float]):
    ws = next(s for s in wb.worksheets if s.title.strip() == "시험 SETTING")
    ws["B4"] = bank.name
    ws["C4"] = counts["공통"]
    ws["D4"] = counts["객관식"]
    ws["E4"] = counts["주관식"]
    ws["C3"] = scores["공통"]
    ws["D3"] = scores["객관식"]
    ws["E3"] = scores["주관식"]

    found = None
    for row in range(3, ws.max_row + 2):
        if clean(ws.cell(row, 11).value) == bank.name:
            found = row
            break
    if found is None:
        found = max(3, ws.max_row + 1)
    ws.cell(found, 11).value = bank.name
    ws.cell(found, 12).value = bank.meta["department"]
    ws.cell(found, 13).value = bank.meta["evaluator"]
    ws.cell(found, 14).value = bank.meta["job_name"]
    ws.cell(found, 15).value = bank.meta["revision"]
    ws.cell(found, 16).value = bank.meta["issue_date"]
    ws.cell(found, 17).value = bank.meta["product_type"]


def write_answer_sheet(answer_ws, bank: QuestionBank, questions: list[Question]):
    answer_ws["B1"] = bank.name
    for idx, q in enumerate(questions, start=1):
        if idx <= 25:
            row, no_col, answer_col = idx + 2, 2, 3
        else:
            row, no_col, answer_col = idx - 23, 5, 6
        answer_ws.cell(row, no_col).value = idx
        answer_ws.cell(row, answer_col).value = answer_for_sheet(q)


def answer_for_sheet(q: Question) -> str:
    if q.exam_type == "객관식":
        m = re.search(r"[①②③④⑤⑥⑦⑧⑨⑩1-9]", q.answer)
        if m:
            circles = "①②③④⑤⑥⑦⑧⑨⑩"
            if m.group(0) in circles:
                return m.group(0)
            try:
                return circles[int(m.group(0)) - 1]
            except Exception:
                pass
    return q.answer


def write_exam_workbook(output_path: Path, template_path: Path, bank: QuestionBank, questions: list[Question], counts: dict[str, int], scores: dict[str, float]):
    wb = openpyxl.load_workbook(template_path, data_only=False, keep_vba=True)
    exam_ws = next(s for s in wb.worksheets if s.title.strip() == "시험지")
    answer_ws = next(s for s in wb.worksheets if s.title.strip() == "답안지")
    ensure_exam_capacity(exam_ws, len(questions))
    clear_outputs(exam_ws, answer_ws, len(questions))
    write_settings(wb, bank, counts, scores)

    for idx, q in enumerate(questions, start=1):
        row = 10 + idx
        exam_ws.cell(row, 2).value = idx
        exam_ws.cell(row, 3).value = f"[{q.exam_type} / {q.score:g}점] {q.question}"
        exam_ws.cell(row, 12).value = None
        if q.exam_type == "주관식" and "(          )" not in q.question and "(       )" not in q.question:
            exam_ws.cell(row, 3).value += "\n\n(                                                   )"

        if q.images:
            exam_ws.row_dimensions[row].height = max(exam_ws.row_dimensions[row].height or 80, 120)
            start_col = 4
            for offset, source_img in enumerate(q.images[:4]):
                new_img = clone_image(source_img)
                if not new_img:
                    continue
                scale = min(1.0, 120 / max(new_img.width, 1), 90 / max(new_img.height, 1))
                new_img.width = int(new_img.width * scale)
                new_img.height = int(new_img.height * scale)
                col_letter = openpyxl.utils.get_column_letter(min(start_col + offset * 2, 10))
                exam_ws.add_image(new_img, f"{col_letter}{row}")

    write_answer_sheet(answer_ws, bank, questions)
    if hasattr(wb, "calculation"):
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def select_questions(bank: QuestionBank, counts: dict[str, int], scores: dict[str, float]) -> list[Question]:
    selected: list[Question] = []
    for kind in QUESTION_TYPES:
        candidates = [q for q in bank.questions if q.exam_type == kind]
        count = counts[kind]
        if count > len(candidates):
            raise ValueError(f"{kind} 문제가 부족합니다. 요청 {count}문항 / 보유 {len(candidates)}문항")
        picked = random.sample(candidates, count)
        for q in picked:
            copied = copy.copy(q)
            copied.images = list(q.images)
            copied.score = scores[kind]
            selected.append(copied)
    random.shuffle(selected)
    return selected


class State:
    path: Path = find_workbook()
    banks: list[QuestionBank] = []


STATE = State()


def json_response(handler: BaseHTTPRequestHandler, payload: dict):
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(200)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def read_json(handler: BaseHTTPRequestHandler) -> dict:
    length = int(handler.headers.get("Content-Length", "0"))
    if not length:
        return {}
    return json.loads(handler.rfile.read(length).decode("utf-8"))


def load_payload(path: Path) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"문제은행 파일을 찾을 수 없습니다: {path}")
    STATE.path = path
    STATE.banks = load_banks(path)
    return {
        "ok": True,
        "path": str(path),
        "banks": [{"name": b.name, "counts": b.counts} for b in STATE.banks],
    }


class Handler(BaseHTTPRequestHandler):
    def log_message(self, _format, *args):
        return

    def do_GET(self):
        if urlparse(self.path).path == "/":
            body = HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        self.send_error(404)

    def do_POST(self):
        try:
            if self.path == "/api/banks":
                data = read_json(self)
                path = Path(data.get("path") or STATE.path)
                json_response(self, load_payload(path))
                return

            if self.path == "/api/generate":
                data = read_json(self)
                path = Path(data.get("path") or STATE.path)
                if not STATE.banks or path != STATE.path:
                    load_payload(path)
                bank = STATE.banks[int(data["selected"])]
                counts = {k: int(data["counts"][k]) for k in QUESTION_TYPES}
                scores = {k: float(data["scores"][k]) for k in QUESTION_TYPES}
                total = sum(counts[k] * scores[k] for k in QUESTION_TYPES)
                target = float(data["target"])
                if abs(total - target) > 0.0001:
                    raise ValueError(f"총점이 목표 점수와 다릅니다. 현재 {total:g}점 / 목표 {target:g}점")
                questions = select_questions(bank, counts, scores)
                out_dir = path.parent / "OJT_Random_Exam_Output"
                out_path = out_dir / f"OJT_시험지_{safe_filename(bank.name)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsm"
                write_exam_workbook(out_path, path, bank, questions, counts, scores)
                json_response(self, {"ok": True, "output": str(out_path)})
                return
            self.send_error(404)
        except Exception as exc:
            json_response(self, {"ok": False, "error": str(exc)})


def free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return sock.getsockname()[1]


def run_web():
    port = free_port()
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    url = f"http://127.0.0.1:{port}/"
    threading.Timer(0.4, lambda: webbrowser.open(url)).start()
    server.serve_forever()


def self_test(path: Path):
    banks = load_banks(path)
    print(f"banks={len(banks)} questions={sum(len(b.questions) for b in banks)} images={sum(len(q.images) for b in banks for q in b.questions)}")
    for bank in banks[:8]:
        print(bank.name, bank.counts, bank.meta)


if __name__ == "__main__":
    if "--self-test" in sys.argv:
        self_test(Path(sys.argv[2]) if len(sys.argv) > 2 else find_workbook())
    else:
        run_web()
