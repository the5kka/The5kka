import json
import calendar
import ctypes
import os
import shutil
import sqlite3
import subprocess
import sys
import threading
import time
import tkinter as tk
import traceback
import unicodedata
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, simpledialog, ttk

from openpyxl import load_workbook
from openpyxl.styles import Font


# ==================================================
# 기본 설정
# ==================================================
# 현장 적용 시 기본 삭제 시간을 60초로 바꾸고 싶으면 아래 값만 수정해도 됩니다.
DNC_DELETE_SECONDS = 10
# DNC 완료 후 초품 확인 팝업이 뜨기 전 대기 시간입니다. 현장 적용 시 이 값만 바꾸면 됩니다.
FIRST_ARTICLE_WAIT_SECONDS = 5
WORK_LOG_SCHEMA_VERSION = 2
CONDITION_MASTER_SCHEMA_VERSION = 1
MASTER_SETTINGS_PASSWORD = "1"
CONDITION_MASTER_PASSWORD = "1"

APP_TITLE = "JIIN DNC Manager"
LOG_SHEET_NAME = "KCC PKG"
SINGLE_INSTANCE_MUTEX_NAME = "JIIN_DNC_Manager_Single_Instance"
ERROR_ALREADY_EXISTS = 183
SINGLE_INSTANCE_HANDLE = None
PROCESS_NAMES = ["TLB", "심텍 SPS", "심텍 HDI", "KCC PKG", "KCC HDI"]


def get_app_dir() -> Path:
    """EXE 실행 시에는 EXE가 있는 폴더, Python 실행 시에는 main.py 폴더를 반환합니다."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def acquire_single_instance_lock() -> bool:
    """프로그램이 이미 실행 중이면 두 번째 실행을 막습니다."""
    global SINGLE_INSTANCE_HANDLE
    if not sys.platform.startswith("win"):
        return True
    kernel32 = ctypes.windll.kernel32
    user32 = ctypes.windll.user32
    handle = kernel32.CreateMutexW(None, False, SINGLE_INSTANCE_MUTEX_NAME)
    if not handle:
        return True
    if kernel32.GetLastError() == ERROR_ALREADY_EXISTS:
        user32.MessageBoxW(
            None,
            "JIIN DNC Manager가 이미 실행 중입니다.\n\n현재 열린 프로그램을 사용해주세요.",
            "중복 실행 차단",
            0x40,
        )
        kernel32.CloseHandle(handle)
        return False
    SINGLE_INSTANCE_HANDLE = handle
    return True


APP_DIR = get_app_dir()
DATA_DIR = APP_DIR / "data"
LOG_DIR = DATA_DIR / "logs"
BACKUP_DIR = DATA_DIR / "backup"
CONFIG_FILE = DATA_DIR / "config.json"
LOGO_FILE = DATA_DIR / "company_logo.png"
BUNDLED_LOGO_FILE = Path(getattr(sys, "_MEIPASS", APP_DIR)) / "company_logo.png"
LEGACY_CONFIG_FILE = APP_DIR / "config.json"
KCC_PKG_DATA_DIR = DATA_DIR / "KCC_PKG"
KCC_PKG_DB_FILE = KCC_PKG_DATA_DIR / "work_log.db"
CONDITION_MASTER_DB_FILE = KCC_PKG_DATA_DIR / "condition_master.db"
LEGACY_KCC_PKG_DB_FILE = DATA_DIR / "KCC_PKG.db"
LEGACY_CONDITION_MASTER_FILE = APP_DIR / "condition_master.json"
MIGRATION_BACKUP_DONE = False

APP_BG = "#f3f6fb"
SURFACE_BG = "#ffffff"
BORDER_COLOR = "#d8e2f0"
TEXT_COLOR = "#172033"
MUTED_TEXT = "#687386"
PRIMARY = "#0f5bff"
PRIMARY_LIGHT = "#eaf2ff"
OK_COLOR = "#0f9f63"
NG_COLOR = "#dc2626"
TAB_BG = "#eaf0f8"

THEMES = {
    "MES 블루": {
        "app_bg": "#f3f6fb",
        "surface_bg": "#ffffff",
        "border_color": "#d8e2f0",
        "text_color": "#172033",
        "muted_text": "#687386",
        "primary": "#0f5bff",
        "primary_light": "#eaf2ff",
        "tab_bg": "#eaf0f8",
    },
    "KCC 민트": {
        "app_bg": "#eaf5f2",
        "surface_bg": "#fbfffe",
        "border_color": "#b7d8d2",
        "text_color": "#102a2a",
        "muted_text": "#55716f",
        "primary": "#00897b",
        "primary_light": "#dff6f1",
        "tab_bg": "#f8fffd",
    },
    "심텍 그린": {
        "app_bg": "#f0f8f1",
        "surface_bg": "#ffffff",
        "border_color": "#c8dfc9",
        "text_color": "#172417",
        "muted_text": "#617064",
        "primary": "#198754",
        "primary_light": "#e6f4ec",
        "tab_bg": "#eef6ef",
    },
    "TLB 퍼플": {
        "app_bg": "#f6f3fb",
        "surface_bg": "#ffffff",
        "border_color": "#d9cfea",
        "text_color": "#221933",
        "muted_text": "#6b617a",
        "primary": "#6f42c1",
        "primary_light": "#f0e9fb",
        "tab_bg": "#f1ecf8",
    },
    "다크 네이비": {
        "app_bg": "#eef2f6",
        "surface_bg": "#ffffff",
        "border_color": "#c9d5e2",
        "text_color": "#14213d",
        "muted_text": "#5c677d",
        "primary": "#1f3a5f",
        "primary_light": "#e7edf5",
        "tab_bg": "#e9eef5",
    },
}


def apply_theme(theme_name: str) -> None:
    """설정된 화면 색상 테마를 전역 UI 색상에 적용합니다."""
    theme = THEMES.get(theme_name, THEMES["MES 블루"])
    globals_to_update = {
        "APP_BG": theme["app_bg"],
        "SURFACE_BG": theme["surface_bg"],
        "BORDER_COLOR": theme["border_color"],
        "TEXT_COLOR": theme["text_color"],
        "MUTED_TEXT": theme["muted_text"],
        "PRIMARY": theme["primary"],
        "PRIMARY_LIGHT": theme["primary_light"],
        "TAB_BG": theme["tab_bg"],
    }
    globals().update(globals_to_update)


# ==================================================
# 로그/백업/마이그레이션 안전장치
# ==================================================
def write_log_file(filename: str, message: str) -> None:
    """현장 문제 추적용 로그를 data/logs 폴더에 남깁니다."""
    try:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_path = LOG_DIR / filename
        with log_path.open("a", encoding="utf-8") as file:
            file.write(f"[{now_text}] {message}\n")
    except Exception:
        # 로그 실패 때문에 현장 작업을 막지는 않습니다.
        pass


def log_app(message: str) -> None:
    write_log_file(f"app_{datetime.now().strftime('%Y%m%d')}.log", message)


def log_error(message: str, exc: Exception | None = None) -> None:
    detail = message
    if exc is not None:
        detail += f"\n{type(exc).__name__}: {exc}\n{traceback.format_exc()}"
    write_log_file(f"error_{datetime.now().strftime('%Y%m%d')}.log", detail)


def create_migration_backup_once() -> Path:
    """DB 마이그레이션 전에 data 주요 파일을 1회 백업합니다."""
    global MIGRATION_BACKUP_DONE
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUP_DIR / f"{timestamp}_before_migration"
    if MIGRATION_BACKUP_DONE:
        return backup_path

    try:
        backup_path.mkdir(parents=True, exist_ok=False)
        if CONFIG_FILE.exists():
            shutil.copy2(CONFIG_FILE, backup_path / "config.json")
        if LEGACY_CONFIG_FILE.exists() and not (backup_path / "config_legacy.json").exists():
            shutil.copy2(LEGACY_CONFIG_FILE, backup_path / "config_legacy.json")

        kcc_backup = backup_path / "KCC_PKG"
        kcc_backup.mkdir(parents=True, exist_ok=True)
        for source, target_name in (
            (KCC_PKG_DB_FILE, "work_log.db"),
            (CONDITION_MASTER_DB_FILE, "condition_master.db"),
            (LEGACY_KCC_PKG_DB_FILE, "legacy_KCC_PKG.db"),
            (LEGACY_CONDITION_MASTER_FILE, "legacy_condition_master.json"),
        ):
            if source.exists():
                shutil.copy2(source, kcc_backup / target_name)
        MIGRATION_BACKUP_DONE = True
        log_app(f"마이그레이션 백업 생성: {backup_path}")
        return backup_path
    except Exception as exc:
        log_error("마이그레이션 백업 실패", exc)
        raise RuntimeError("DB 마이그레이션 전 백업에 실패했습니다.\n기존 data 보호를 위해 실행을 중단합니다.") from exc


def table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone()
    return row is not None


def get_existing_columns(conn: sqlite3.Connection, table_name: str) -> set[str]:
    if not table_exists(conn, table_name):
        return set()
    return {row[1] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()}


def ensure_schema_version_table(conn: sqlite3.Connection, db_name: str, version: int) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS schema_version (
            db_name TEXT PRIMARY KEY,
            version INTEGER NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        INSERT INTO schema_version (db_name, version, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(db_name) DO UPDATE SET
            version=excluded.version,
            updated_at=excluded.updated_at
        """,
        (db_name, version, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    )


def migrate_table_columns(conn: sqlite3.Connection, table_name: str, expected_columns: dict[str, str]) -> list[str]:
    """없는 컬럼만 ALTER TABLE로 추가해서 중복 추가 오류를 막습니다."""
    added_columns: list[str] = []
    existing_columns = get_existing_columns(conn, table_name)
    for column_name, column_sql in expected_columns.items():
        if column_name in existing_columns:
            continue
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_sql}")
        added_columns.append(column_name)
    return added_columns


# ==================================================
# 공통 파일/설정 함수
# ==================================================
def get_desktop_path() -> Path:
    """현재 Windows 사용자의 바탕화면 경로를 반환합니다."""
    return Path.home() / "Desktop"


def get_default_config() -> dict:
    """config.json이 없거나 값이 비어 있을 때 사용할 기본값입니다."""
    desktop = get_desktop_path()
    source_folders = {
        "TLB": str(desktop / "TLB"),
        "심텍 SPS": str(desktop / "SIMMTECH_SPS"),
        "심텍 HDI": str(desktop / "SIMMTECH_HDI"),
        "KCC PKG": str(desktop / "KCC_PKG"),
        "KCC HDI": str(desktop / "KCC_HDI"),
    }
    return {
        "config_version": 1,
        "excel_file": "",
        "source_dnc_folder": str(desktop / "KCC_PKG"),
        "source_dnc_folders": source_folders,
        "transfer_dnc_folder": str(desktop / "DNC"),
        "dnc_delete_seconds": DNC_DELETE_SECONDS,
        "first_article_wait_seconds": FIRST_ARTICLE_WAIT_SECONDS,
        "auto_export_after_dnc": True,
        "auto_shift_group_enabled": True,
        "a_group_day_start_date": "2026-05-17",
        "master_password": MASTER_SETTINGS_PASSWORD,
        "condition_master_password": CONDITION_MASTER_PASSWORD,
        "clear_common_after_normal": False,
        "machine": "트리밍 1호기",
        "theme": "MES 블루",
    }


def normalize_machine_name(machine: str) -> str:
    """화면 표시용 호기명을 작업일보 기록용 짧은 이름으로 바꿉니다."""
    text = str(machine or "").strip()
    for number in ("1", "2", "3"):
        if number in text:
            return f"{number}호기"
    return text


def get_work_period(now: datetime | None = None) -> dict[str, str]:
    """08:30 기준 작업일자와 08:30/20:30 기준 근무를 계산합니다."""
    now = now or datetime.now()
    day_start = now.replace(hour=8, minute=30, second=0, microsecond=0)
    night_start = now.replace(hour=20, minute=30, second=0, microsecond=0)
    if now < day_start:
        work_date = (now - timedelta(days=1)).strftime("%Y-%m-%d")
        shift = "야간"
    elif now < night_start:
        work_date = now.strftime("%Y-%m-%d")
        shift = "주간"
    else:
        work_date = now.strftime("%Y-%m-%d")
        shift = "야간"
    return {
        "work_date": work_date,
        "shift": shift,
        "period_key": f"{work_date}_{shift}",
    }


def get_auto_shift_group(work_date_text: str, shift: str, a_group_day_start_date: str) -> str:
    """A조 주간 시작일을 기준으로 4근 2휴 조 패턴에서 현재 근무 조를 계산합니다."""
    try:
        work_date = datetime.strptime(work_date_text, "%Y-%m-%d").date()
        base_date = datetime.strptime(a_group_day_start_date, "%Y-%m-%d").date()
    except ValueError:
        return ""
    day_offset = (work_date - base_date).days
    # 각 조는 12일 주기로 주간4/휴2/야간4/휴2 패턴입니다.
    # B조, C조는 A조 기준에서 각각 4일, 8일 밀린 패턴으로 계산합니다.
    group_offsets = {"A": 0, "B": 4, "C": 8}
    for group, offset in group_offsets.items():
        cycle_day = (day_offset + offset) % 12
        if shift == "주간" and 0 <= cycle_day <= 3:
            return group
        if shift == "야간" and 6 <= cycle_day <= 9:
            return group
    return ""


def load_config() -> dict:
    """config.json을 읽고, 없는 값은 기본값으로 채웁니다."""
    config = get_default_config()
    changed = False
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        if not CONFIG_FILE.exists() and LEGACY_CONFIG_FILE.exists():
            shutil.copy2(LEGACY_CONFIG_FILE, CONFIG_FILE)
            log_app("기존 config.json을 data/config.json으로 복사")
        if CONFIG_FILE.exists():
            saved = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            if isinstance(saved, dict):
                config.update({key: value for key, value in saved.items() if value not in (None, "")})
                for key, value in get_default_config().items():
                    if key not in saved:
                        config[key] = value
                        changed = True
        else:
            changed = True
    except Exception:
        # 설정 파일이 손상되어도 프로그램은 기본값으로 실행되게 합니다.
        log_error("config.json 로드 실패 - 기본값 사용")
        pass
    default_sources = get_default_config()["source_dnc_folders"]
    saved_sources = config.get("source_dnc_folders")
    if not isinstance(saved_sources, dict):
        saved_sources = {}
        changed = True
    for process_name, default_path in default_sources.items():
        if not saved_sources.get(process_name):
            saved_sources[process_name] = config.get("source_dnc_folder", default_path) if process_name == "KCC PKG" else default_path
            changed = True
    config["source_dnc_folders"] = saved_sources
    config["source_dnc_folder"] = saved_sources.get("KCC PKG", config.get("source_dnc_folder", default_sources["KCC PKG"]))
    if changed:
        try:
            save_config(config)
            log_app("config.json 누락 항목 자동 보정")
        except Exception as exc:
            log_error("config.json 자동 보정 저장 실패", exc)
    return config


def save_config(config: dict) -> None:
    """현재 설정을 config.json에 저장합니다."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    CONFIG_FILE.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")


def select_excel_file(parent, config: dict, excel_var: tk.StringVar) -> None:
    """설정 탭에서 작업일보 Excel 파일을 선택합니다."""
    file_path = filedialog.askopenfilename(
        parent=parent,
        title="작업일보 Excel 파일 선택",
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    if file_path:
        config["excel_file"] = file_path
        excel_var.set(file_path)
        save_config(config)


def ensure_excel_file_selected(parent, config: dict, excel_var: tk.StringVar | None = None) -> bool:
    """작업일보 파일이 없을 때만 사용자에게 한 번 선택을 요청합니다."""
    current_path = ""
    if excel_var is not None:
        current_path = excel_var.get().strip()
    if not current_path:
        current_path = config.get("excel_file", "").strip()
    if not current_path and CONFIG_FILE.exists():
        try:
            saved = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            current_path = str(saved.get("excel_file", "")).strip()
        except Exception:
            current_path = ""

    if current_path and Path(current_path).exists():
        config["excel_file"] = current_path
        if excel_var is not None:
            excel_var.set(current_path)
        save_config(config)
        return True

    messagebox.showwarning(
        "작업일보 선택 필요",
        "작업일보 Excel 파일이 선택되어 있지 않거나 파일을 찾을 수 없습니다.\n\n"
        "DNC 이력을 저장할 작업일보 Excel 파일을 선택해주세요.",
        parent=parent,
    )
    file_path = filedialog.askopenfilename(
        parent=parent,
        title="작업일보 Excel 파일 선택",
        initialdir=str(Path(current_path).parent) if current_path else str(get_desktop_path()),
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    if not file_path:
        return False

    config["excel_file"] = file_path
    if excel_var is not None:
        excel_var.set(file_path)
    save_config(config)
    return True


def open_path(path: Path) -> None:
    """파일 또는 폴더를 Windows 기본 프로그램으로 엽니다."""
    if sys.platform.startswith("win"):
        os.startfile(str(path))
    else:
        subprocess.Popen(["xdg-open", str(path)])


def open_log_excel(config: dict) -> None:
    """선택된 작업일보 Excel 파일을 엽니다."""
    if not ensure_excel_file_selected(None, config):
        return
    excel_file = Path(config.get("excel_file", ""))
    try:
        open_path(excel_file)
    except Exception as exc:
        messagebox.showerror("작업일보 열기 실패", str(exc))


# ==================================================
# DNC 파일 처리 함수
# ==================================================
def delete_existing_dnc_txt(transfer_folder: Path) -> None:
    """DNC 전송 폴더에 남아 있는 txt 파일을 실행 전에 모두 삭제합니다."""
    transfer_folder.mkdir(parents=True, exist_ok=True)
    for txt_file in transfer_folder.glob("*.txt"):
        txt_file.unlink()


def search_condition_file(condition_name: str, source_folder: Path) -> list[Path]:
    """원본 DNC 폴더와 하위 폴더에서 조건명과 일치하는 txt 파일을 찾습니다."""
    source_folder.mkdir(parents=True, exist_ok=True)

    def normalize_name(value: str) -> str:
        text = unicodedata.normalize("NFC", str(value or "")).strip()
        if text.lower().endswith(".txt"):
            text = text[:-4]
        return text.casefold()

    normalized = normalize_name(condition_name)

    matches = []
    for file_path in source_folder.rglob("*.txt"):
        if normalize_name(file_path.stem) == normalized:
            matches.append(file_path)
    return matches


def copy_dnc_file(source_file: Path, transfer_folder: Path) -> Path:
    """조건 txt 파일을 DNC 전송 폴더로 복사합니다."""
    transfer_folder.mkdir(parents=True, exist_ok=True)
    copied_file = transfer_folder / source_file.name
    shutil.copy2(source_file, copied_file)
    return copied_file


def delete_after_delay(copied_file: Path, seconds: int, status_callback=None) -> None:
    """지정된 초만큼 기다린 후 복사된 DNC txt 파일을 삭제합니다."""
    for remaining in range(seconds, 0, -1):
        if status_callback:
            status_callback(f"DNC 삭제 대기중 ({remaining})")
        time.sleep(1)
    if copied_file.exists():
        copied_file.unlink()
    if status_callback:
        status_callback("DNC 완료")


# ==================================================
# 검증 함수
# ==================================================
def check_mes_core(lot_no: str, process_code: str) -> bool:
    """LOT No 9번째 문자부터 4자리와 공정코드 끝 2자리로 MES Core 일치화를 판정합니다."""
    lot_no = lot_no.strip()
    process_code = process_code.strip().upper()
    if len(lot_no) < 12 or len(process_code) < 2:
        return False

    lot_core = lot_no[8:12]
    process_tail = process_code[-2:]
    return (lot_core == "0000" and process_tail == "T1") or (lot_core == "0205" and process_tail == "11")


def get_mes_core_message(lot_no: str, process_code: str) -> tuple[bool, str]:
    """MES Core 판정 결과와 작업자가 확인할 상세 사유를 함께 반환합니다."""
    lot_no = lot_no.strip()
    process_code = process_code.strip().upper()
    if len(lot_no) < 12:
        return False, f"LOT No가 너무 짧습니다. 현재 LOT No: {lot_no or '빈칸'}"
    if len(process_code) < 2:
        return False, f"공정코드가 너무 짧습니다. 현재 공정코드: {process_code or '빈칸'}"

    lot_core = lot_no[8:12]
    process_tail = process_code[-2:]
    if lot_core == "0000" and process_tail == "T1":
        return True, "OK - LOT 중간 4자리 0000 / 공정코드 끝자리 T1"
    if lot_core == "0205" and process_tail == "11":
        return True, "OK - LOT 중간 4자리 0205 / 공정코드 끝자리 11"
    return (
        False,
        "MES Core 기준이 맞지 않습니다. "
        f"LOT 중간 4자리: {lot_core}, 공정코드 끝 2자리: {process_tail}. "
        "허용 기준은 0000+T1 또는 0205+11 입니다.",
    )


def check_condition_ok(lot1: dict, lot2: dict | None = None) -> bool:
    """작업조건 KCC_ 시작 여부와 지그/2LOT 조건 일치 여부를 확인합니다."""
    if not lot1.get("condition", "").strip().upper().startswith("KCC_"):
        return False
    if not lot1.get("jig", "").strip():
        return False
    if lot2:
        return (
            lot1.get("condition", "").strip() == lot2.get("condition", "").strip()
            and lot1.get("jig", "").strip() == lot2.get("jig", "").strip()
        )
    return True


def get_single_condition_message(lot: dict) -> tuple[bool, str]:
    """LOT 한 개의 작업조건/지그 기본 조건을 상세하게 확인합니다."""
    condition = lot.get("condition", "").strip()
    jig = lot.get("jig", "").strip()
    if not condition:
        return False, "작업조건이 비어 있습니다. [작업조건 / 지그 불러오기]를 눌러주세요."
    if not condition.upper().startswith("KCC_"):
        return False, f"작업조건이 KCC_로 시작하지 않습니다. 현재 작업조건: {condition}"
    if not jig:
        return False, "지그가 비어 있습니다. [작업조건 / 지그 불러오기]를 눌러주세요."
    return True, "OK - 작업조건 KCC_ 시작 / 지그 입력 완료"


def get_lot_match_message(lot1: dict, lot2: dict) -> tuple[bool, str]:
    """2LOT 작업 시 LOT 1/LOT 2의 작업조건과 지그 일치 여부를 작업자용으로 간단히 확인합니다."""
    lot1_no = lot1.get("lot_no", "").strip()
    lot2_no = lot2.get("lot_no", "").strip()
    lot1_condition = lot1.get("condition", "").strip()
    lot2_condition = lot2.get("condition", "").strip()
    lot1_jig = lot1.get("jig", "").strip()
    lot2_jig = lot2.get("jig", "").strip()
    messages = []
    if lot1_no and lot2_no and lot1_no == lot2_no:
        messages.append("LOT No가 같습니다")
    if lot1_condition != lot2_condition:
        if not lot1_condition or not lot2_condition:
            messages.append("작업조건 미조회")
        else:
            messages.append("작업조건 다름")
    if lot1_jig != lot2_jig:
        if not lot1_jig or not lot2_jig:
            messages.append("지그 미조회")
        else:
            messages.append("지그 다름")
    if messages:
        return False, "확인 필요 - " + ", ".join(dict.fromkeys(messages))
    if not lot1_condition or not lot1_jig:
        return False, "확인 필요 - 작업조건/지그 미조회"
    return True, "OK - 2LOT 조건 일치"


def validate_positive_number(value: str, field_name: str, required: bool = True) -> tuple[bool, str]:
    """매수/Stack처럼 0보다 큰 숫자만 허용하는 필드 검증입니다."""
    text = value.strip()
    if not text:
        return (not required, "" if not required else f"{field_name}은(는) 필수입니다.")
    if not text.isdigit():
        return False, f"{field_name}은(는) 숫자만 입력 가능합니다."
    if int(text) <= 0:
        return False, f"{field_name}은(는) 0보다 커야 합니다."
    return True, ""


def validate_zero_or_positive_number(value: str, field_name: str) -> tuple[bool, str]:
    """신규 모델 검증 매수처럼 0 이상 숫자를 반드시 입력해야 하는 필드 검증입니다."""
    text = value.strip()
    if not text:
        return True, ""
    if not text.isdigit():
        return False, f"{field_name}은(는) 숫자만 입력 가능합니다."
    return True, ""


def validate_lot_required(lot: dict, lot_name: str, require_qty: bool) -> list[str]:
    """LOT 입력 필수값을 확인합니다."""
    errors = []
    required_fields = [
        ("step", "STEP"),
        ("round", "차수"),
        ("manage_no", "관리번호"),
        ("lot_no", "LOT No"),
        ("process_code", "공정코드"),
        ("condition", "작업조건"),
        ("jig", "지그"),
    ]
    if require_qty:
        required_fields.insert(4, ("qty", "매수"))

    for key, label in required_fields:
        if not lot.get(key, "").strip():
            errors.append(f"{lot_name} {label}을(를) 입력하세요.")

    if require_qty:
        ok, message = validate_positive_number(lot.get("qty", ""), f"{lot_name} 매수", required=True)
        if not ok:
            errors.append(message)
    elif lot.get("qty", "").strip():
        ok, message = validate_positive_number(lot.get("qty", ""), f"{lot_name} 매수", required=False)
        if not ok:
            errors.append(message)

    if lot.get("condition", "").strip() and not lot.get("condition", "").strip().upper().startswith("KCC_"):
        errors.append(f"{lot_name} 작업조건은 반드시 KCC_ 로 시작해야 합니다.")
    return errors


def validate_normal_dnc(common: dict, lot1: dict, lot2: dict | None) -> tuple[bool, list[str]]:
    """일반 DNC 입력값 전체 검증입니다."""
    errors = []
    for key, label in (("work_date", "작업일자"), ("machine", "트리밍 호기"), ("shift_group", "조"), ("shift", "근무"), ("worker", "작업자")):
        if not common.get(key, "").strip():
            errors.append(f"{label}을(를) 입력하세요.")

    errors.extend(validate_lot_required(lot1, "LOT 1", require_qty=True))
    if lot2:
        errors.extend(validate_lot_required(lot2, "LOT 2", require_qty=True))

    lot1_mes_ok, lot1_mes_message = get_mes_core_message(lot1.get("lot_no", ""), lot1.get("process_code", ""))
    if not lot1_mes_ok:
        errors.append(f"LOT 1 MES Core 불일치: {lot1_mes_message}")
    lot1_condition_ok, lot1_condition_message = get_single_condition_message(lot1)
    if not lot1_condition_ok:
        errors.append(f"LOT 1 조건 적용 불가: {lot1_condition_message}")

    if lot2:
        lot2_mes_ok, lot2_mes_message = get_mes_core_message(lot2.get("lot_no", ""), lot2.get("process_code", ""))
        if not lot2_mes_ok:
            errors.append(f"LOT 2 MES Core 불일치: {lot2_mes_message}")
        lot2_condition_ok, lot2_condition_message = get_single_condition_message(lot2)
        if not lot2_condition_ok:
            errors.append(f"LOT 2 조건 적용 불가: {lot2_condition_message}")
        lot_match_ok, lot_match_message = get_lot_match_message(lot1, lot2)
        if not lot_match_ok:
            errors.append(f"2LOT 조건 불일치: {lot_match_message}")
    return len(errors) == 0, errors


def validate_new_model_dnc(common: dict, lot: dict) -> tuple[bool, list[str]]:
    """신규 모델 검증 DNC 입력값 전체 검증입니다."""
    errors = []
    for key, label in (("work_date", "작업일자"), ("machine", "트리밍 호기"), ("shift_group", "조"), ("shift", "근무"), ("worker", "작업자")):
        if not common.get(key, "").strip():
            errors.append(f"{label}을(를) 입력하세요.")

    errors.extend(validate_lot_required(lot, "신규 모델", require_qty=False))
    ok, message = validate_zero_or_positive_number(lot.get("qty", ""), "신규 모델 매수")
    if not ok:
        errors.append(message)
    mes_ok, mes_message = get_mes_core_message(lot.get("lot_no", ""), lot.get("process_code", ""))
    if not mes_ok:
        errors.append(f"MES Core 불일치: {mes_message}")
    condition_ok, condition_message = get_single_condition_message(lot)
    if not condition_ok:
        errors.append(f"조건 적용 불가: {condition_message}")
    return len(errors) == 0, errors


# ==================================================
# Excel 저장 함수
# ==================================================
def get_next_empty_row(ws) -> int:
    """A:Q 범위 기준으로 8행부터 첫 빈 행을 찾습니다."""
    row = 8
    while True:
        has_value = any(ws.cell(row=row, column=col).value not in (None, "") for col in range(1, 18))
        if not has_value:
            return row
        row += 1


def write_process_code_backup(ws, row: int, process_code: str) -> None:
    """조건 마스터 복구용으로 작업일보 AD열에 공정코드를 백업합니다."""
    # A열에 호기가 추가되어 기존 A:AC 양식이 한 칸씩 밀렸습니다.
    # AD열만 프로그램 복구용 공정코드 백업 칸으로 사용합니다.
    if not ws.cell(row=6, column=30).value:
        ws.cell(row=6, column=30).value = "공정코드"
    ws.cell(row=row, column=30).value = excel_upper_value(process_code)


def excel_upper_value(value):
    """작업일보에서 작업 P/G를 제외한 문자 값은 보기 좋게 대문자로 저장합니다."""
    if isinstance(value, str):
        return value.upper()
    return value


def ensure_log_sheet_machine_column(ws) -> None:
    """작업일보에 A열 '호기'가 없으면 자동으로 한 칸 삽입합니다.

    기존 양식은 A열이 작업일자였으므로 새 EXE가 바로 쓰면 날짜를 덮어쓸 수 있습니다.
    A열에 호기가 없고 작업일자 양식으로 보이면 A열을 추가해 기존 데이터를 오른쪽으로 보존합니다.
    """
    first_header = str(ws.cell(row=6, column=1).value or "").strip()
    top_header = str(ws.cell(row=2, column=1).value or "").strip()
    if first_header == "호기" or top_header == "호기":
        return
    if first_header == "작업일자" or top_header == "작업일자":
        ws.insert_cols(1)
        ws.cell(row=6, column=1).value = "호기"
        ws.cell(row=2, column=1).value = "호기"


def open_log_workbook(config: dict):
    """작업일보 파일과 KCC PKG 시트를 열고 기본 오류를 메시지로 변환합니다."""
    excel_file = config.get("excel_file", "")
    if not excel_file:
        raise FileNotFoundError("작업일보 Excel 파일을 선택해주세요.")
    path = Path(excel_file)
    if not path.exists():
        raise FileNotFoundError("작업일보 Excel 파일을 선택해주세요.")

    try:
        if not zipfile.is_zipfile(path):
            raise ValueError("작업일보 Excel 파일 형식이 손상되었거나 올바른 Excel 파일이 아닙니다.")
        workbook = load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm")
    except PermissionError:
        raise PermissionError("작업일보 Excel 파일이 열려 있어 저장할 수 없습니다.\n파일을 닫고 다시 실행해주세요.")
    except zipfile.BadZipFile:
        raise ValueError("작업일보 Excel 파일 형식이 손상되었거나 올바른 Excel 파일이 아닙니다.")

    if LOG_SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise KeyError("작업일보 파일에 KCC PKG 시트가 없습니다.")
    ws = workbook[LOG_SHEET_NAME]
    ensure_log_sheet_machine_column(ws)
    return workbook, ws, path


def save_workbook_safely(workbook, path: Path) -> None:
    """Excel 저장 실패 시 사용자가 이해할 수 있는 메시지를 발생시킵니다."""
    while True:
        try:
            workbook.save(path)
            return
        except PermissionError:
            retry = ask_excel_save_retry()
            if not retry:
                raise PermissionError("작업일보 Excel 파일이 열려 있어 현재 저장을 중단했습니다.\n나중에 [작업일보 반영]을 눌러 다시 저장해주세요.")


def ask_excel_save_retry() -> bool:
    """작업일보가 열려 있을 때 현장 작업자가 이해하기 쉬운 문구로 재시도 여부를 묻습니다."""
    dialog = tk.Toplevel()
    dialog.title("작업일보 저장 대기")
    dialog.resizable(False, False)
    dialog.configure(bg=SURFACE_BG)
    result = {"retry": False}

    body = tk.Frame(dialog, bg=SURFACE_BG, padx=22, pady=18)
    body.pack(fill=tk.BOTH, expand=True)
    tk.Label(
        body,
        text="작업일보 Excel 파일이 열려 있어 저장할 수 없습니다.",
        bg=SURFACE_BG,
        fg=TEXT_COLOR,
        font=("맑은 고딕", 10, "bold"),
        anchor="w",
        justify=tk.LEFT,
    ).pack(fill=tk.X, pady=(0, 10))
    tk.Label(
        body,
        text=(
            "Excel에서 작업일보를 닫은 뒤 [재 시도]를 눌러주세요.\n"
            "[다음에 저장]을 누르면 현재 저장을 중단합니다."
        ),
        bg=SURFACE_BG,
        fg=TEXT_COLOR,
        font=("맑은 고딕", 10),
        anchor="w",
        justify=tk.LEFT,
    ).pack(fill=tk.X)

    buttons = tk.Frame(dialog, bg=APP_BG, padx=14, pady=12)
    buttons.pack(fill=tk.X)

    def choose_retry() -> None:
        result["retry"] = True
        dialog.destroy()

    def choose_later() -> None:
        result["retry"] = False
        dialog.destroy()

    ttk.Button(buttons, text="재 시도", command=choose_retry, style="Primary.TButton", width=16).pack(side=tk.RIGHT, padx=(8, 0))
    ttk.Button(buttons, text="다음에 저장", command=choose_later, width=16).pack(side=tk.RIGHT)
    dialog.protocol("WM_DELETE_WINDOW", choose_later)
    dialog.update_idletasks()
    width = dialog.winfo_reqwidth()
    height = dialog.winfo_reqheight()
    screen_width = dialog.winfo_screenwidth()
    screen_height = dialog.winfo_screenheight()
    dialog.geometry(f"{width}x{height}+{(screen_width - width) // 2}+{(screen_height - height) // 2}")
    dialog.grab_set()
    dialog.lift()
    dialog.focus_force()
    dialog.attributes("-topmost", True)
    dialog.wait_window()
    return result["retry"]


def ask_numeric_input(parent, title: str, prompt: str) -> str | None:
    """숫자만 입력 가능한 간단한 팝업입니다."""
    dialog = tk.Toplevel(parent)
    dialog.title(title)
    dialog.resizable(False, False)
    dialog.configure(bg=SURFACE_BG)
    result = {"value": None}
    value_var = tk.StringVar()

    body = tk.Frame(dialog, bg=SURFACE_BG, padx=18, pady=16)
    body.pack(fill=tk.BOTH, expand=True)
    tk.Label(body, text=prompt, bg=SURFACE_BG, fg=TEXT_COLOR, font=("맑은 고딕", 10, "bold"), anchor="w").pack(fill=tk.X, pady=(0, 8))

    vcmd = (dialog.register(lambda text: text.isdigit() or text == ""), "%P")
    entry = ttk.Entry(body, textvariable=value_var, style="Wide.TEntry", width=18, validate="key", validatecommand=vcmd)
    entry.pack(fill=tk.X)

    buttons = tk.Frame(dialog, bg=APP_BG, padx=14, pady=12)
    buttons.pack(fill=tk.X)

    def confirm() -> None:
        value = value_var.get().strip()
        if not value:
            messagebox.showwarning(title, "숫자를 입력하세요.", parent=dialog)
            return
        result["value"] = value
        dialog.destroy()

    ttk.Button(buttons, text="확인", command=confirm, style="Primary.TButton", width=12).pack(side=tk.RIGHT, padx=(8, 0))
    ttk.Button(buttons, text="취소", command=dialog.destroy, width=12).pack(side=tk.RIGHT)
    dialog.transient(parent)
    dialog.grab_set()
    dialog.update_idletasks()
    width = dialog.winfo_reqwidth()
    height = dialog.winfo_reqheight()
    x = parent.winfo_rootx() + max((parent.winfo_width() - width) // 2, 0)
    y = parent.winfo_rooty() + max((parent.winfo_height() - height) // 2, 0)
    dialog.geometry(f"{width}x{height}+{x}+{y}")
    dialog.lift()
    entry.focus_set()
    dialog.wait_window()
    return result["value"]


# ==================================================
# KCC PKG DB 저장 함수
# ==================================================
WORK_LOG_COLUMNS = {
    "customer_process": "TEXT NOT NULL DEFAULT 'KCC PKG'",
    "dnc_type": "TEXT NOT NULL DEFAULT '일반'",
    "status": "TEXT NOT NULL DEFAULT '완료'",
    "machine": "TEXT",
    "work_date": "TEXT",
    "shift_group": "TEXT",
    "shift_name": "TEXT",
    "worker": "TEXT",
    "step": "TEXT",
    "round_no": "TEXT",
    "manage_no": "TEXT",
    "lot_no": "TEXT",
    "qty_text": "TEXT",
    "qty_number": "INTEGER",
    "result_value": "REAL",
    "process_code": "TEXT",
    "condition_name": "TEXT",
    "jig": "TEXT",
    "stack": "TEXT",
    "model_change_text": "TEXT",
    "burr_result": "TEXT",
    "record_time": "TEXT",
    "first_axis_1": "TEXT",
    "first_axis_2": "TEXT",
    "first_axis_3": "TEXT",
    "first_axis_4": "TEXT",
    "first_axis_5": "TEXT",
    "first_axis_6": "TEXT",
    "jig_axis_1": "TEXT",
    "jig_axis_2": "TEXT",
    "jig_axis_3": "TEXT",
    "jig_axis_4": "TEXT",
    "jig_axis_5": "TEXT",
    "jig_axis_6": "TEXT",
    "exported": "INTEGER NOT NULL DEFAULT 0",
    "exported_at": "TEXT",
    "created_at": "TEXT",
}

CONDITION_MASTER_COLUMNS = {
    "step": "TEXT",
    "round_no": "TEXT",
    "manage_no": "TEXT",
    "process_code": "TEXT",
    "lot_no": "TEXT",
    "condition_name": "TEXT",
    "jig": "TEXT",
    "source": "TEXT",
    "updated_at": "TEXT",
}


def get_kcc_pkg_db_path() -> Path:
    """KCC PKG 원본 이력을 저장하는 SQLite DB 파일 경로를 반환합니다."""
    KCC_PKG_DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not KCC_PKG_DB_FILE.exists() and LEGACY_KCC_PKG_DB_FILE.exists():
        shutil.copy2(LEGACY_KCC_PKG_DB_FILE, KCC_PKG_DB_FILE)
    return KCC_PKG_DB_FILE


def get_kcc_pkg_connection() -> sqlite3.Connection:
    """KCC PKG DB 연결을 만들고 필요한 테이블을 자동 생성합니다."""
    db_path = get_kcc_pkg_db_path()
    existing_db = db_path.exists()
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        need_backup = False
        if existing_db:
            need_backup = not table_exists(conn, "schema_version")
            if table_exists(conn, "dnc_logs"):
                missing_columns = set(WORK_LOG_COLUMNS) - get_existing_columns(conn, "dnc_logs")
                need_backup = need_backup or bool(missing_columns)
        if need_backup:
            create_migration_backup_once()

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS dnc_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_process TEXT NOT NULL DEFAULT 'KCC PKG',
                dnc_type TEXT NOT NULL DEFAULT '일반',
                status TEXT NOT NULL DEFAULT '완료',
                machine TEXT,
                work_date TEXT,
                shift_group TEXT,
                shift_name TEXT,
                worker TEXT,
                step TEXT,
                round_no TEXT,
                manage_no TEXT,
                lot_no TEXT,
                qty_text TEXT,
                qty_number INTEGER,
                result_value REAL,
                process_code TEXT,
                condition_name TEXT,
                jig TEXT,
                stack TEXT,
                model_change_text TEXT,
                burr_result TEXT,
                record_time TEXT,
                first_axis_1 TEXT,
                first_axis_2 TEXT,
                first_axis_3 TEXT,
                first_axis_4 TEXT,
                first_axis_5 TEXT,
                first_axis_6 TEXT,
                jig_axis_1 TEXT,
                jig_axis_2 TEXT,
                jig_axis_3 TEXT,
                jig_axis_4 TEXT,
                jig_axis_5 TEXT,
                jig_axis_6 TEXT,
                exported INTEGER NOT NULL DEFAULT 0,
                exported_at TEXT,
                created_at TEXT
            )
            """
        )
        added = migrate_table_columns(conn, "dnc_logs", WORK_LOG_COLUMNS)
        ensure_schema_version_table(conn, "work_log", WORK_LOG_SCHEMA_VERSION)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dnc_logs_export ON dnc_logs(exported, status, id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dnc_logs_created ON dnc_logs(created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dnc_logs_lot ON dnc_logs(lot_no)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dnc_logs_work_date ON dnc_logs(work_date)")
        conn.commit()
        if added:
            log_app(f"work_log.db 컬럼 자동 추가: {', '.join(added)}")
    except Exception as exc:
        conn.rollback()
        conn.close()
        log_error("work_log.db 마이그레이션 실패", exc)
        raise
    return conn


def get_condition_master_connection() -> sqlite3.Connection:
    """KCC PKG 조건 마스터 DB 연결을 만들고 필요한 테이블을 자동 생성합니다."""
    KCC_PKG_DATA_DIR.mkdir(parents=True, exist_ok=True)
    existing_db = CONDITION_MASTER_DB_FILE.exists()
    conn = sqlite3.connect(CONDITION_MASTER_DB_FILE)
    conn.row_factory = sqlite3.Row
    try:
        need_backup = False
        if existing_db:
            need_backup = not table_exists(conn, "schema_version")
            if table_exists(conn, "condition_master"):
                missing_columns = set(CONDITION_MASTER_COLUMNS) - get_existing_columns(conn, "condition_master")
                need_backup = need_backup or bool(missing_columns)
        if need_backup:
            create_migration_backup_once()

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS condition_master (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                step TEXT,
                round_no TEXT,
                manage_no TEXT,
                process_code TEXT,
                lot_no TEXT,
                condition_name TEXT,
                jig TEXT,
                source TEXT,
                updated_at TEXT
            )
            """
        )
        added = migrate_table_columns(conn, "condition_master", CONDITION_MASTER_COLUMNS)
        ensure_schema_version_table(conn, "condition_master", CONDITION_MASTER_SCHEMA_VERSION)
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_condition_master_key
            ON condition_master(step, round_no, process_code, manage_no)
            """
        )
        conn.commit()
        if added:
            log_app(f"condition_master.db 컬럼 자동 추가: {', '.join(added)}")
    except Exception as exc:
        conn.rollback()
        conn.close()
        log_error("condition_master.db 마이그레이션 실패", exc)
        raise
    return conn


def calculate_result_value(qty_number: int) -> float | None:
    """매수 기준 실적을 계산합니다. 더미는 실적을 비워 둡니다."""
    if qty_number <= 0:
        return None
    return round(qty_number * 0.2, 1)


def insert_normal_dnc_db(common: dict, lots: list[dict], stack: str, model_change: bool) -> list[int]:
    """일반 DNC 이력을 Excel 대신 KCC_PKG.db에 먼저 저장합니다."""
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ids: list[int] = []
    conn = get_kcc_pkg_connection()
    try:
        for index, lot in enumerate(lots):
            qty_number = int(lot["qty"])
            cursor = conn.execute(
                """
                INSERT INTO dnc_logs (
                    dnc_type, status, machine, work_date, shift_group, shift_name, worker,
                    step, round_no, manage_no, lot_no, qty_text, qty_number, result_value,
                    process_code, condition_name, jig, stack, model_change_text, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "일반",
                    "DNC 진행",
                    normalize_machine_name(common["machine"]),
                    common["work_date"],
                    common["shift_group"],
                    common["shift"],
                    common["worker"],
                    lot["step"],
                    lot["round"],
                    lot["manage_no"],
                    lot["lot_no"],
                    lot["qty"],
                    qty_number,
                    calculate_result_value(qty_number),
                    lot["process_code"],
                    lot["condition"],
                    lot["jig"],
                    stack,
                    "기종교체" if model_change and index == 0 else "",
                    now_text,
                ),
            )
            ids.append(int(cursor.lastrowid))
        conn.commit()
        log_app(f"일반 DNC DB 저장: {len(ids)} LOT / ids={ids}")
    finally:
        conn.close()
    return ids


def update_normal_frequent_check_db(log_ids: list[int], model_change: bool, frequent_check: list[str]) -> None:
    """초품 4Point와 하부 Pin 확인 결과를 KCC_PKG.db에 반영합니다."""
    conn = get_kcc_pkg_connection()
    try:
        for index, log_id in enumerate(log_ids):
            first_values = frequent_check[:6]
            jig_values = frequent_check[6:] if model_change and index == 0 else [""] * 6
            conn.execute(
                """
                UPDATE dnc_logs
                   SET first_axis_1=?, first_axis_2=?, first_axis_3=?,
                       first_axis_4=?, first_axis_5=?, first_axis_6=?,
                       jig_axis_1=?, jig_axis_2=?, jig_axis_3=?,
                       jig_axis_4=?, jig_axis_5=?, jig_axis_6=?
                 WHERE id=?
                """,
                (*first_values, *jig_values, log_id),
            )
        conn.commit()
    finally:
        conn.close()


def update_normal_burr_db(log_ids: list[int], burr_ok: bool) -> None:
    """Burr 결과와 작업 시간을 KCC_PKG.db에 저장합니다."""
    result = "이상 없음" if burr_ok else "Burr 발생"
    now_text = datetime.now().strftime("%H:%M:%S")
    conn = get_kcc_pkg_connection()
    try:
        conn.executemany(
            "UPDATE dnc_logs SET status='완료', burr_result=?, record_time=? WHERE id=?",
            [(result, now_text, log_id) for log_id in log_ids],
        )
        conn.commit()
        log_app(f"일반 DNC 완료 처리: ids={log_ids}, Burr={result}")
    finally:
        conn.close()


def insert_new_model_db(common: dict, lot: dict, leader_name: str) -> int:
    """신규 모델 검증 DNC 이력을 KCC_PKG.db에 저장합니다."""
    qty_text = lot.get("qty", "").strip()
    qty_number = int(qty_text) if qty_text else 0
    display_qty = "더미" if qty_number == 0 else qty_text
    conn = get_kcc_pkg_connection()
    try:
        cursor = conn.execute(
            """
            INSERT INTO dnc_logs (
                dnc_type, status, machine, work_date, shift_group, shift_name, worker,
                step, round_no, manage_no, lot_no, qty_text, qty_number, result_value,
                process_code, condition_name, jig, model_change_text, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "신규 검증",
                "DNC 진행",
                normalize_machine_name(common["machine"]),
                common["work_date"],
                common["shift_group"],
                common["shift"],
                leader_name,
                lot["step"],
                lot["round"],
                lot["manage_no"],
                lot["lot_no"],
                display_qty,
                qty_number,
                calculate_result_value(qty_number),
                lot["process_code"],
                lot["condition"],
                lot["jig"],
                "신규 검증",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        conn.commit()
        log_id = int(cursor.lastrowid)
        log_app(f"신규 모델 DNC DB 저장: id={log_id}, STEP={lot['step']}, LOT={lot['lot_no']}")
        return log_id
    finally:
        conn.close()


def update_new_model_db(log_id: int, condition_name: str, first_article_ok: bool) -> None:
    """신규 모델 초도품 확인 결과를 KCC_PKG.db에 저장합니다."""
    final_condition = condition_name if first_article_ok else f"[검증 NG 발생] {condition_name}"
    conn = get_kcc_pkg_connection()
    try:
        conn.execute(
            "UPDATE dnc_logs SET status='완료', condition_name=?, record_time=? WHERE id=?",
            (final_condition, datetime.now().strftime("%H:%M:%S"), log_id),
        )
        conn.commit()
        log_app(f"신규 모델 DNC 완료 처리: id={log_id}, 초도품={'OK' if first_article_ok else 'NG'}")
    finally:
        conn.close()


def sync_condition_master_from_completed_logs() -> int:
    """완료된 DB 이력 중 조건 마스터에 반영 가능한 조건을 동기화합니다."""
    conn = get_kcc_pkg_connection()
    updated = 0
    try:
        rows = conn.execute(
            """
            SELECT dnc_type, step, round_no, manage_no, lot_no, process_code,
                   condition_name, jig, created_at
              FROM dnc_logs
             WHERE status='완료'
               AND step IS NOT NULL
               AND round_no IS NOT NULL
               AND manage_no IS NOT NULL
               AND process_code IS NOT NULL
               AND condition_name IS NOT NULL
               AND jig IS NOT NULL
             ORDER BY id
            """
        ).fetchall()
    finally:
        conn.close()

    for row in rows:
        condition = str(row["condition_name"] or "").strip()
        jig = str(row["jig"] or "").strip()
        if not condition or condition.startswith("[검증 NG 발생]") or not jig:
            continue
        lot = {
            "step": str(row["step"] or "").strip(),
            "round": str(row["round_no"] or "").strip(),
            "manage_no": str(row["manage_no"] or "").strip(),
            "lot_no": str(row["lot_no"] or "").strip(),
            "process_code": str(row["process_code"] or "").strip(),
        }
        source = "신규 검증 DB" if row["dnc_type"] == "신규 검증" else "DNC 완료 DB"
        before = len(load_condition_master())
        upsert_condition_master(lot, condition, jig, source)
        after = len(load_condition_master())
        updated += 1 if after >= before else 0
    return updated


def get_unexported_kcc_pkg_count() -> int:
    """Excel 작업일보에 아직 반영되지 않은 KCC PKG DB 이력 수를 반환합니다."""
    conn = get_kcc_pkg_connection()
    try:
        row = conn.execute("SELECT COUNT(*) AS count FROM dnc_logs WHERE exported=0 AND status='완료'").fetchone()
        return int(row["count"])
    finally:
        conn.close()


def get_incomplete_kcc_pkg_count() -> int:
    """DNC 중간 종료 등으로 완료 처리되지 않은 미반영 이력 수를 반환합니다."""
    conn = get_kcc_pkg_connection()
    try:
        row = conn.execute("SELECT COUNT(*) AS count FROM dnc_logs WHERE exported=0 AND status!='완료'").fetchone()
        return int(row["count"])
    finally:
        conn.close()


def load_work_history(limit: int = 500, only_unexported: bool = False, only_incomplete: bool = False, keyword: str = "") -> list[sqlite3.Row]:
    """작업 이력 보기 창에서 사용할 최근 DNC 이력을 조회합니다."""
    conn = get_kcc_pkg_connection()
    try:
        where = []
        params: list[object] = []
        if only_unexported:
            where.append("exported=0 AND status='완료'")
        if only_incomplete:
            where.append("status!='완료'")
        search = keyword.strip()
        if search:
            where.append("(lot_no LIKE ? OR manage_no LIKE ? OR condition_name LIKE ? OR worker LIKE ?)")
            like = f"%{search}%"
            params.extend([like, like, like, like])
        sql = """
            SELECT id, dnc_type, status, machine, work_date, shift_group, shift_name, worker,
                   step, round_no, manage_no, lot_no, qty_text, result_value,
                   condition_name, jig, stack, model_change_text, burr_result,
                   record_time, exported, exported_at, created_at
              FROM dnc_logs
        """
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY id DESC LIMIT ?"
        params.append(max(1, min(int(limit), 5000)))
        return conn.execute(sql, params).fetchall()
    finally:
        conn.close()


def fetch_unexported_kcc_pkg_logs() -> list[sqlite3.Row]:
    """Excel로 내보낼 미반영 KCC PKG 이력을 오래된 순서로 조회합니다."""
    conn = get_kcc_pkg_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM dnc_logs WHERE exported=0 AND status='완료' ORDER BY id"
        ).fetchall()
        return rows
    finally:
        conn.close()


def mark_kcc_pkg_logs_exported(log_ids: list[int]) -> None:
    """Excel 반영이 끝난 DB 이력을 반영 완료 처리합니다."""
    if not log_ids:
        return
    conn = get_kcc_pkg_connection()
    try:
        conn.executemany(
            "UPDATE dnc_logs SET exported=1, exported_at=? WHERE id=?",
            [(datetime.now().strftime("%Y-%m-%d %H:%M:%S"), log_id) for log_id in log_ids],
        )
        conn.commit()
    finally:
        conn.close()


def write_db_log_row_to_excel(ws, row: int, log: sqlite3.Row) -> None:
    """KCC_PKG.db 한 건을 기존 KCC PKG 작업일보 양식에 맞춰 기록합니다."""
    values = [
        log["machine"] or "",
        log["work_date"],
        log["shift_group"],
        log["shift_name"],
        log["worker"],
        log["step"],
        log["round_no"],
        log["manage_no"],
        log["lot_no"],
        log["qty_text"],
        "" if log["result_value"] is None else log["result_value"],
        log["condition_name"],
        log["burr_result"] or "",
        log["stack"] or "",
        log["jig"],
        log["model_change_text"] or "",
        log["record_time"] or "",
    ]
    for col, value in enumerate(values, start=1):
        # L열 작업 P/G는 조건 파일명과 같은 원문을 유지하고, 나머지 문자 값만 대문자로 정리합니다.
        ws.cell(row=row, column=col).value = value if col == 12 else excel_upper_value(value)
    write_process_code_backup(ws, row, log["process_code"] or "")
    if log["burr_result"] == "Burr 발생":
        ws.cell(row=row, column=13).font = Font(color="FF0000", bold=False)
    if str(log["condition_name"] or "").startswith("[검증 NG 발생]"):
        ws.cell(row=row, column=12).font = Font(name="맑은 고딕", size=10, color="FF0000", bold=False)

    frequent_values = [
        log["first_axis_1"],
        log["first_axis_2"],
        log["first_axis_3"],
        log["first_axis_4"],
        log["first_axis_5"],
        log["first_axis_6"],
        log["jig_axis_1"],
        log["jig_axis_2"],
        log["jig_axis_3"],
        log["jig_axis_4"],
        log["jig_axis_5"],
        log["jig_axis_6"],
    ]
    for offset, value in enumerate(frequent_values, start=18):
        ws.cell(row=row, column=offset).value = value or ""


def export_kcc_pkg_db_to_excel(config: dict) -> int:
    """KCC_PKG.db의 미반영 이력을 Excel KCC PKG 시트로 내보냅니다."""
    logs = fetch_unexported_kcc_pkg_logs()
    if not logs:
        return 0

    exported_ids: list[int] = []
    log_app(f"작업일보 반영 시작: {len(logs)}건")
    workbook = None
    try:
        workbook, ws, path = open_log_workbook(config)
        start_row = get_next_empty_row(ws)
        for index, log in enumerate(logs):
            write_db_log_row_to_excel(ws, start_row + index, log)
            exported_ids.append(int(log["id"]))
        save_workbook_safely(workbook, path)
    except Exception as exc:
        log_error("작업일보 반영 실패 - exported 상태 변경 안 함", exc)
        raise
    finally:
        if workbook is not None:
            workbook.close()
    mark_kcc_pkg_logs_exported(exported_ids)
    log_app(f"작업일보 반영 완료: {len(exported_ids)}건")
    return len(exported_ids)


def export_process_db_to_excel(process_name: str, config: dict) -> int:
    """공정별 미반영 이력을 작업일보에 반영합니다. 현재는 KCC PKG만 실제 구현되어 있습니다."""
    if process_name == "KCC PKG":
        return export_kcc_pkg_db_to_excel(config)
    return 0


def get_unexported_process_count(process_name: str) -> int:
    """공정별 Excel 미반영 수를 반환합니다. 미구현 공정은 0건으로 둡니다."""
    if process_name == "KCC PKG":
        return get_unexported_kcc_pkg_count()
    return 0


def get_incomplete_process_count(process_name: str) -> int:
    """공정별 미완료 수를 반환합니다. 미구현 공정은 0건으로 둡니다."""
    if process_name == "KCC PKG":
        return get_incomplete_kcc_pkg_count()
    return 0


def export_all_processes_to_excel(config: dict) -> dict[str, int]:
    """수동 작업일보 반영 버튼에서 5개 공정을 순서대로 확인하고 반영합니다."""
    result: dict[str, int] = {}
    for process_name in PROCESS_NAMES:
        result[process_name] = export_process_db_to_excel(process_name, config)
    return result


def make_condition_key(step: str, round_no: str, manage_no: str, process_code: str) -> str:
    """조건 마스터에서 중복을 제거하기 위한 기준 키를 만듭니다.

    작업조건/지그는 STEP, 차수, 공정코드가 맞을 때만 불러와야 하므로
    공정코드까지 중복 판단 기준에 포함합니다.
    """
    return "|".join(
        [
            step.strip(),
            round_no.strip(),
            process_code.strip(),
            manage_no.strip(),
        ]
    )


def get_condition_source_priority(source: str) -> int:
    """조건 마스터 출처별 신뢰 우선순위를 반환합니다.

    신규 검증으로 확정한 조건은 작업일보 이력보다 우선합니다.
    현장에서 직접 수정한 값은 의도적으로 고친 값이므로 가장 높게 둡니다.
    """
    source_text = str(source or "")
    if "사용자 수정" in source_text:
        return 40
    if "신규 검증" in source_text:
        return 30
    if "DNC 완료 DB" in source_text:
        return 20
    if "작업일보" in source_text:
        return 10
    return 0


def should_replace_condition_record(current: dict, incoming: dict) -> bool:
    """조건 마스터 중복 키 병합 시 새 기록으로 교체해도 되는지 판단합니다."""
    current_priority = get_condition_source_priority(current.get("source", ""))
    incoming_priority = get_condition_source_priority(incoming.get("source", ""))
    if incoming_priority < current_priority:
        return False
    return True


def merge_condition_records(records: list[dict]) -> list[dict]:
    """같은 STEP/차수/관리번호 조건은 한 줄로 합치되 출처 우선순위를 보호합니다."""
    merged: dict[str, dict] = {}
    for record in records:
        if not str(record.get("process_code", "")).strip():
            continue
        key = make_condition_key(
            record.get("step", ""),
            record.get("round", ""),
            record.get("manage_no", ""),
            record.get("process_code", ""),
        )
        if not key.replace("|", "").strip():
            continue
        if key not in merged:
            merged[key] = dict(record)
            continue

        current = merged[key]
        if not should_replace_condition_record(current, record):
            continue
        for field in ("step", "round", "manage_no", "condition", "jig", "source", "updated_at", "lot_no"):
            value = str(record.get(field, "")).strip()
            if value:
                current[field] = value
        process_code = str(record.get("process_code", "")).strip()
        if process_code:
            current["process_code"] = process_code

    return sorted(merged.values(), key=lambda item: (item.get("step", ""), item.get("round", ""), item.get("process_code", ""), item.get("manage_no", "")))


def load_condition_master() -> list[dict]:
    """KCC PKG 조건 마스터 DB를 읽습니다. 예전 JSON 파일이 있으면 1회 가져옵니다."""
    conn = get_condition_master_connection()
    try:
        count = int(conn.execute("SELECT COUNT(*) FROM condition_master").fetchone()[0])
        if count == 0 and LEGACY_CONDITION_MASTER_FILE.exists():
            try:
                data = json.loads(LEGACY_CONDITION_MASTER_FILE.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    records = merge_condition_records(data)
                    for record in records:
                        conn.execute(
                            """
                            INSERT OR REPLACE INTO condition_master (
                                step, round_no, manage_no, process_code, lot_no,
                                condition_name, jig, source, updated_at
                            )
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                record.get("step", ""),
                                record.get("round", ""),
                                record.get("manage_no", ""),
                                record.get("process_code", ""),
                                record.get("lot_no", ""),
                                record.get("condition", ""),
                                record.get("jig", ""),
                                record.get("source", ""),
                                record.get("updated_at", ""),
                            ),
                        )
                    conn.commit()
            except Exception:
                pass
        rows = conn.execute(
            """
            SELECT step, round_no, manage_no, process_code, lot_no,
                   condition_name, jig, source, updated_at
              FROM condition_master
             ORDER BY step, round_no, process_code, manage_no
            """
        ).fetchall()
        return [
            {
                "step": row["step"] or "",
                "round": row["round_no"] or "",
                "manage_no": row["manage_no"] or "",
                "process_code": row["process_code"] or "",
                "lot_no": row["lot_no"] or "",
                "condition": row["condition_name"] or "",
                "jig": row["jig"] or "",
                "source": row["source"] or "",
                "updated_at": row["updated_at"] or "",
            }
            for row in rows
        ]
    finally:
        conn.close()


def save_condition_master(records: list[dict]) -> None:
    """조건 마스터 DB를 저장합니다."""
    records = merge_condition_records(records)
    conn = get_condition_master_connection()
    try:
        conn.execute("DELETE FROM condition_master")
        for record in records:
            conn.execute(
                """
                INSERT OR REPLACE INTO condition_master (
                    step, round_no, manage_no, process_code, lot_no,
                    condition_name, jig, source, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    record.get("step", ""),
                    record.get("round", ""),
                    record.get("manage_no", ""),
                    record.get("process_code", ""),
                    record.get("lot_no", ""),
                    record.get("condition", ""),
                    record.get("jig", ""),
                    record.get("source", ""),
                    record.get("updated_at", ""),
                ),
            )
        conn.commit()
    finally:
        conn.close()


def upsert_condition_master(lot: dict, condition: str, jig: str, source: str) -> None:
    """현재 입력값과 불러온 조건/지그를 조건 마스터에 추가 또는 갱신합니다."""
    step = lot.get("step", "").strip()
    round_no = lot.get("round", "").strip()
    manage_no = lot.get("manage_no", "").strip()
    process_code = lot.get("process_code", "").strip()
    if not (step and round_no and process_code and manage_no and condition and jig):
        return

    key = make_condition_key(step, round_no, manage_no, process_code)
    records = load_condition_master()
    for record in records:
        record_key = make_condition_key(
            record.get("step", ""),
            record.get("round", ""),
            record.get("manage_no", ""),
            record.get("process_code", ""),
        )
        if record_key == key:
            incoming = {"source": source}
            if not should_replace_condition_record(record, incoming):
                log_app(
                    "조건 마스터 갱신 건너뜀: "
                    f"기존 출처={record.get('source', '')}, 새 출처={source}, "
                    f"STEP={step}, 차수={round_no}, 관리번호={manage_no}, 공정코드={process_code}"
                )
                return
            record.update(
                {
                    "lot_no": lot.get("lot_no", "").strip(),
                    "condition": condition,
                    "jig": jig,
                    "source": source,
                    "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
            )
            save_condition_master(records)
            return

    records.append(
        {
            "step": step,
            "round": round_no,
            "manage_no": manage_no,
            "process_code": process_code,
            "lot_no": lot.get("lot_no", "").strip(),
            "condition": condition,
            "jig": jig,
            "source": source,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
    )
    save_condition_master(records)


def rebuild_condition_master_from_log(config: dict) -> int:
    """작업일보 KCC PKG 시트의 최신 이력을 기존 마스터에 병합합니다.

    작업일보에서 사라진 모델도 조건 마스터에서는 보존합니다.
    """
    workbook, ws, _path = open_log_workbook(config)
    records = load_condition_master()
    try:
        for row in range(8, ws.max_row + 1):
            step = str(ws.cell(row=row, column=6).value or "").strip()
            round_no = str(ws.cell(row=row, column=7).value or "").strip()
            manage_no = str(ws.cell(row=row, column=8).value or "").strip()
            # AD열은 프로그램 복구용 공정코드 백업 칸입니다.
            # AD가 비어 있는 예전 이력은 정확한 조건 키를 만들 수 없어 건너뜁니다.
            process_code = str(ws.cell(row=row, column=30).value or "").strip()
            condition = str(ws.cell(row=row, column=12).value or "").strip()
            jig = str(ws.cell(row=row, column=15).value or "").strip()
            lot_no = str(ws.cell(row=row, column=9).value or "").strip()
            if condition.startswith("[검증 NG 발생]"):
                continue
            if not (step and round_no and manage_no and process_code and condition and jig):
                continue
            records.append(
                {
                    "step": step,
                    "round": round_no,
                    "manage_no": manage_no,
                    "process_code": process_code,
                    "lot_no": lot_no,
                    "condition": condition,
                    "jig": jig,
                    "source": f"작업일보 {row}행",
                    "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
            )
    finally:
        workbook.close()

    save_condition_master(records)
    return len(load_condition_master())


def lookup_condition_jig_from_master(lot: dict) -> tuple[str, str, str]:
    """조건 마스터에서 작업조건/지그를 먼저 찾습니다."""
    records = load_condition_master()
    step = lot.get("step", "").strip()
    round_no = lot.get("round", "").strip()
    manage_no = lot.get("manage_no", "").strip()
    process_code = lot.get("process_code", "").strip()
    if not (step and round_no and process_code):
        return "", "", ""

    for record in reversed(records):
        if (
            record.get("step", "") == step
            and record.get("round", "") == round_no
            and record.get("process_code", "") == process_code
            and (not manage_no or record.get("manage_no", "") == manage_no)
        ):
            return record.get("condition", ""), record.get("jig", ""), f"조건 마스터({record.get('source', '저장값')})"

    return "", "", ""


def describe_condition_lookup_mismatch(lot: dict) -> str:
    """조건 마스터에서 왜 조건을 찾지 못했는지 작업자용 문구로 설명합니다."""
    records = load_condition_master()
    step = lot.get("step", "").strip()
    round_no = lot.get("round", "").strip()
    manage_no = lot.get("manage_no", "").strip()
    process_code = lot.get("process_code", "").strip()

    if not records:
        return "조건 마스터에 저장된 조건이 없습니다."

    candidates = [
        record
        for record in records
        if record.get("step", "") == step
        and record.get("round", "") == round_no
        and (not manage_no or record.get("manage_no", "") == manage_no)
    ]
    if candidates:
        lines = ["가장 가까운 조건은 찾았지만 아래 항목이 다릅니다."]
        for record in candidates[:5]:
            differences = []
            for field, label, current_value in (
                ("process_code", "공정코드", process_code),
                ("manage_no", "관리번호", manage_no),
                ("step", "STEP", step),
                ("round", "차수", round_no),
            ):
                saved_value = str(record.get(field, "")).strip()
                if current_value and saved_value != current_value:
                    differences.append(f"{label}: 입력 [{current_value}] / 마스터 [{saved_value or '빈칸'}]")
            if differences:
                lines.append("- " + " / ".join(differences))
        return "\n".join(lines)

    step_candidates = [record for record in records if record.get("step", "") == step]
    if step_candidates:
        examples = []
        for record in step_candidates[:5]:
            examples.append(
                f"- 차수 [{record.get('round', '')}], 관리번호 [{record.get('manage_no', '')}], 공정코드 [{record.get('process_code', '')}]"
            )
        return "STEP은 같지만 차수/관리번호/공정코드가 일치하는 조건이 없습니다.\n" + "\n".join(examples)

    return "STEP부터 일치하는 조건이 조건 마스터에 없습니다."


def lookup_condition_jig_from_history(config: dict, lot: dict) -> tuple[str, str, str]:
    """기존 작업일보 이력에서 작업조건(K열)과 지그(N열)를 찾아옵니다.

    작업조건/지그는 작업자가 직접 입력하는 값이 아니라 기존 진행 이력을 참고해야 하므로,
    KCC PKG 시트 8행 이후를 아래 우선순위로 뒤에서부터 검색합니다.

    1순위: 조건 마스터에서 STEP + 차수 + 공정코드 일치
    2순위: 작업일보에서 STEP + 차수 + 관리번호 일치

    작업일보에는 공정코드가 저장되지 않으므로, 공정코드가 입력된 경우에는
    조건 마스터를 우선 사용하고 작업일보 이력은 보조로만 사용합니다.
    """
    # The work log does not store process_code, so using it as a fallback can
    # bring the wrong condition when STEP/round are similar. Use only the local
    # master, whose key includes STEP + round + process_code.
    if not (
        lot.get("step", "").strip()
        and lot.get("round", "").strip()
        and lot.get("process_code", "").strip()
    ):
        return "", "", ""

    return lookup_condition_jig_from_master(lot)


def write_common_lot_row(ws, row: int, common: dict, lot: dict, stack: str, model_change: str, frequent_check: list[str] | None = None) -> None:
    """일반 DNC 작업일보 한 줄을 기록합니다."""
    qty = int(lot["qty"])
    result = round(qty * 0.2, 1)
    values = [
        normalize_machine_name(common.get("machine", "")),
        common["work_date"],
        common["shift_group"],
        common["shift"],
        common["worker"],
        lot["step"],
        lot["round"],
        lot["manage_no"],
        lot["lot_no"],
        qty,
        result,
        lot["condition"],
        "",
        stack,
        lot["jig"],
        model_change,
        "",
    ]
    for col, value in enumerate(values, start=1):
        # L열 작업 P/G는 원문 그대로, 나머지는 작업일보 가독성을 위해 대문자로 저장합니다.
        ws.cell(row=row, column=col).value = value if col == 12 else excel_upper_value(value)
    write_process_code_backup(ws, row, lot.get("process_code", ""))
    if frequent_check:
        for offset, value in enumerate(frequent_check, start=18):
            ws.cell(row=row, column=offset).value = value


def save_normal_dnc_log(config: dict, common: dict, lots: list[dict], stack: str, model_change: bool, frequent_check: list[str] | None = None) -> tuple[Path, list[int]]:
    """일반 DNC 내용을 작업일보에 저장하고 저장된 행 번호를 반환합니다."""
    workbook, ws, path = open_log_workbook(config)
    rows = []
    try:
        start_row = get_next_empty_row(ws)
        for index, lot in enumerate(lots):
            row = start_row + index
            rows.append(row)
            row_frequent_check = None
            if frequent_check:
                # 초품 4Point(Q:V)는 모든 LOT 저장행에 기록합니다.
                # 지그교체 하부핀(W:AB)은 기종교체 표시가 들어가는 첫 저장행에만 기록합니다.
                row_frequent_check = frequent_check[:6] + (frequent_check[6:] if model_change and index == 0 else [""] * 6)
            write_common_lot_row(
                ws,
                row,
                common,
                lot,
                stack,
                "기종교체" if model_change and index == 0 else "",
                row_frequent_check,
            )
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()
    return path, rows


def update_normal_burr_result(config: dict, rows: list[int], burr_ok: bool) -> None:
    """DNC 완료 후 Burr 결과와 기록시간을 일반 DNC 저장행에 반영합니다."""
    workbook, ws, path = open_log_workbook(config)
    try:
        result = "이상 없음" if burr_ok else "Burr 발생"
        now_text = datetime.now().strftime("%H:%M:%S")
        for row in rows:
            ws.cell(row=row, column=13).value = result
            ws.cell(row=row, column=17).value = now_text
            if not burr_ok:
                ws.cell(row=row, column=13).font = Font(color="FF0000", bold=False)
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()


def update_normal_frequent_check_result(config: dict, rows: list[int], model_change: bool, frequent_check: list[str]) -> None:
    """DNC 완료 후 확인한 초품/하부 Pin 결과를 저장된 작업일보 행에 반영합니다."""
    workbook, ws, path = open_log_workbook(config)
    try:
        for index, row in enumerate(rows):
            # 초품 4Point(Q:V)는 모든 LOT 저장행에 기록합니다.
            for offset, value in enumerate(frequent_check[:6], start=18):
                ws.cell(row=row, column=offset).value = value
            # 하부 Pin 3개 확인(W:AB)은 기종교체 표시가 들어가는 첫 저장행에만 기록합니다.
            for offset, value in enumerate(frequent_check[6:] if model_change and index == 0 else [""] * 6, start=24):
                ws.cell(row=row, column=offset).value = value
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()


def save_new_model_log(config: dict, common: dict, lot: dict, leader_name: str) -> tuple[Path, int]:
    """신규 모델 검증 DNC 내용을 작업일보에 저장합니다."""
    workbook, ws, path = open_log_workbook(config)
    try:
        row = get_next_empty_row(ws)
        qty_text = lot.get("qty", "").strip()
        qty_number = int(qty_text) if qty_text else 0
        qty_value = "더미" if qty_number == 0 else qty_number
        result_value = "" if qty_number == 0 else round(qty_number * 0.2, 1)
        values = [
            normalize_machine_name(common.get("machine", "")),
            common["work_date"],
            common["shift_group"],
            common["shift"],
            leader_name,
            lot["step"],
            lot["round"],
            lot["manage_no"],
            lot["lot_no"],
            qty_value,
            result_value,
            lot["condition"],
            "",
            "",
            lot["jig"],
            "신규 검증",
            "",
        ]
        for col, value in enumerate(values, start=1):
            # L열 작업 P/G는 원문 그대로, 나머지는 작업일보 가독성을 위해 대문자로 저장합니다.
            ws.cell(row=row, column=col).value = value if col == 12 else excel_upper_value(value)
        write_process_code_backup(ws, row, lot.get("process_code", ""))
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()
    return path, row


def update_new_model_result(config: dict, row: int, condition_name: str, first_article_ok: bool) -> None:
    """신규 모델 DNC 완료 후 초도품 확인 결과와 기록시간을 저장합니다."""
    workbook, ws, path = open_log_workbook(config)
    try:
        k_cell = ws.cell(row=row, column=12)
        if first_article_ok:
            k_cell.value = condition_name
        else:
            k_cell.value = f"[검증 NG 발생] {condition_name}"
            k_cell.font = Font(name="맑은 고딕", size=10, color="FF0000", bold=False)
        ws.cell(row=row, column=17).value = datetime.now().strftime("%H:%M:%S")
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()


# ==================================================
# GUI
# ==================================================
class LabeledEntry(ttk.Frame):
    """라벨과 입력칸을 한 줄로 만드는 작은 공용 위젯입니다."""

    def __init__(
        self,
        parent,
        label: str,
        width: int = 18,
        style: str = "Wide.TEntry",
        readonly: bool = False,
        on_change=None,
        uppercase: bool = False,
        numeric_only: bool = False,
    ):
        super().__init__(parent)
        self.var = tk.StringVar()
        self.on_change = on_change
        self.uppercase = uppercase
        self.numeric_only = numeric_only
        self._normalizing = False
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        self.entry = ttk.Entry(
            self,
            textvariable=self.var,
            width=width,
            style=style,
            state="readonly" if readonly else "normal",
        )
        self.entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        if not readonly and (uppercase or numeric_only):
            self.var.trace_add("write", self.normalize_value)
        if on_change and not readonly:
            self.entry.bind("<KeyRelease>", lambda _event: on_change())
            self.entry.bind("<FocusOut>", lambda _event: on_change())

    def normalize_value(self, *_args) -> None:
        if self._normalizing:
            return
        value = self.var.get()
        normalized = value
        if self.numeric_only:
            normalized = "".join(ch for ch in normalized if ch.isdigit())
        if self.uppercase:
            normalized = normalized.upper()
        if normalized != value:
            self._normalizing = True
            self.var.set(normalized)
            self._normalizing = False

    def get(self) -> str:
        return self.var.get().strip()

    def set(self, value: str) -> None:
        text = str(value or "")
        if self.numeric_only:
            text = "".join(ch for ch in text if ch.isdigit())
        if self.uppercase:
            text = text.upper()
        self.var.set(text)

    def clear(self) -> None:
        self.var.set("")


class DateField(ttk.Frame):
    """날짜 형식을 통일하기 위한 선택식 날짜 입력 위젯입니다."""

    def __init__(self, parent, label: str, on_change=None):
        super().__init__(parent)
        self.var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.on_change = on_change
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        self.entry = ttk.Entry(self, textvariable=self.var, width=14, style="Wide.TEntry", state="readonly")
        self.entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(self, text="선택", width=6, command=self.open_picker).pack(side=tk.LEFT, padx=(6, 0))

    def get(self) -> str:
        return self.var.get().strip()

    def set(self, value: str) -> None:
        self.var.set(value)

    def clear(self) -> None:
        self.var.set(datetime.now().strftime("%Y-%m-%d"))

    def open_picker(self) -> None:
        picker = tk.Toplevel(self)
        picker.title("작업일자 선택")
        picker.configure(bg=APP_BG)
        picker.resizable(False, False)

        today = datetime.now()
        try:
            selected = datetime.strptime(self.var.get(), "%Y-%m-%d")
        except ValueError:
            selected = today

        year_var = tk.IntVar(value=selected.year)
        month_var = tk.IntVar(value=selected.month)
        day_var = tk.IntVar(value=selected.day)

        top = ttk.Frame(picker, padding=(12, 12, 12, 6))
        top.pack(fill=tk.X)
        ttk.Spinbox(top, from_=today.year - 5, to=today.year + 5, textvariable=year_var, width=8).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Spinbox(top, from_=1, to=12, textvariable=month_var, width=5).pack(side=tk.LEFT, padx=(0, 6))

        days_frame = ttk.Frame(picker, padding=(12, 6, 12, 12))
        days_frame.pack()

        def refresh_days() -> None:
            for child in days_frame.winfo_children():
                child.destroy()
            last_day = calendar.monthrange(year_var.get(), month_var.get())[1]
            for day in range(1, last_day + 1):
                button = ttk.Button(
                    days_frame,
                    text=str(day),
                    width=4,
                    command=lambda value=day: select_day(value),
                )
                button.grid(row=(day - 1) // 7, column=(day - 1) % 7, padx=2, pady=2)

        def select_day(day: int) -> None:
            day_var.set(day)
            self.var.set(f"{year_var.get():04d}-{month_var.get():02d}-{day_var.get():02d}")
            if self.on_change:
                self.on_change()
            picker.destroy()

        ttk.Button(top, text="변경", command=refresh_days).pack(side=tk.LEFT)
        refresh_days()


class SegmentedField(ttk.Frame):
    """조/근무처럼 정해진 값만 선택하게 하는 버튼형 입력 위젯입니다."""

    def __init__(self, parent, label: str, options: list[str], initial: str | None = None, button_width: int | None = None, allow_empty: bool = False, on_change=None):
        super().__init__(parent)
        self.var = tk.StringVar(value=initial if initial in options else ("" if allow_empty else (options[0] if options else "")))
        self.on_change = on_change
        self.buttons: list[tk.Button] = []
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        wrap = tk.Frame(self, bg=APP_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        wrap.pack(side=tk.LEFT, fill=tk.X, expand=True)
        for option in options:
            button = tk.Button(
                wrap,
                text=option,
                command=lambda value=option: self.set_user(value),
                relief=tk.FLAT,
                bd=0,
                width=button_width or 0,
                padx=18,
                pady=6,
                cursor="hand2",
                font=("맑은 고딕", 10),
            )
            button.pack(side=tk.LEFT, fill=tk.X, expand=True)
            self.buttons.append(button)
        self.update_buttons()

    def get(self) -> str:
        return self.var.get().strip()

    def set(self, value: str) -> None:
        self.var.set(value)
        self.update_buttons()

    def set_user(self, value: str) -> None:
        self.set(value)
        if self.on_change:
            self.on_change()

    def clear(self) -> None:
        self.var.set("")
        self.update_buttons()

    def update_buttons(self) -> None:
        for button in self.buttons:
            selected = button.cget("text") == self.var.get()
            button.configure(bg=PRIMARY_LIGHT if selected else SURFACE_BG, fg=PRIMARY if selected else TEXT_COLOR)


class ComboField(ttk.Frame):
    """거의 고정으로 쓰는 값은 실수 클릭을 줄이기 위해 드롭다운으로 선택합니다."""

    def __init__(self, parent, label: str, options: list[str], initial: str | None = None, width: int = 12):
        super().__init__(parent)
        self.var = tk.StringVar(value=initial if initial in options else (options[0] if options else ""))
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        self.combo = ttk.Combobox(
            self,
            textvariable=self.var,
            values=options,
            state="readonly",
            width=width,
            style="White.TCombobox",
            font=("맑은 고딕", 10),
        )
        self.combo.pack(side=tk.LEFT, fill=tk.X, expand=True)

    def get(self) -> str:
        return self.var.get().strip()

    def set(self, value: str) -> None:
        self.var.set(value)

    def clear(self) -> None:
        pass


class RoundField(ttk.Frame):
    """차수를 1차~8차 버튼으로 선택하게 하는 작은 입력 위젯입니다."""

    OPTIONS = [f"{number}차" for number in range(1, 9)]

    def __init__(self, parent, label: str):
        super().__init__(parent)
        self.var = tk.StringVar()
        self.buttons: list[tk.Button] = []
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        wrap = tk.Frame(self, bg=APP_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        wrap.pack(side=tk.LEFT, fill=tk.X, expand=True)
        for option in self.OPTIONS:
            button = tk.Button(
                wrap,
                text=option,
                command=lambda value=option: self.toggle(value),
                relief=tk.FLAT,
                bd=0,
                width=3,
                padx=1,
                pady=6,
                cursor="hand2",
                font=("맑은 고딕", 9),
            )
            button.pack(side=tk.LEFT, fill=tk.X, expand=True)
            self.buttons.append(button)
        self.update_buttons()

    def get(self) -> str:
        return self.var.get().strip()

    def toggle(self, value: str) -> None:
        """작업자가 같은 차수 버튼을 한 번 더 누르면 선택을 취소합니다."""
        self.var.set("" if self.var.get() == value else value)
        self.update_buttons()

    def set(self, value: str) -> None:
        self.var.set(value)
        self.update_buttons()

    def clear(self) -> None:
        self.var.set("")
        self.update_buttons()

    def update_buttons(self) -> None:
        for button in self.buttons:
            selected = button.cget("text") == self.var.get()
            button.configure(bg=PRIMARY_LIGHT if selected else SURFACE_BG, fg=PRIMARY if selected else TEXT_COLOR)


class SimpleTabNotebook(ttk.Frame):
    """Tkinter 기본 Notebook 대신 선택 테두리가 선명한 간단 탭 UI입니다."""

    def __init__(self, parent):
        super().__init__(parent)
        self.selected_index = -1
        self.pages: list[tk.Frame] = []
        self.labels: list[tk.Label] = []
        self.tab_bar = tk.Frame(self, bg=APP_BG)
        self.tab_bar.pack(fill=tk.X)
        self.page_area = tk.Frame(self, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        self.page_area.pack(fill=tk.BOTH, expand=True)

    def add(self, page: tk.Frame, text: str) -> None:
        index = len(self.pages)
        self.pages.append(page)
        label = tk.Label(
            self.tab_bar,
            text=text,
            bg=TAB_BG,
            fg=MUTED_TEXT,
            width=12,
            padx=0,
            pady=12,
            font=("맑은 고딕", 10),
            cursor="hand2",
            highlightthickness=1,
            highlightbackground=BORDER_COLOR,
            bd=0,
        )
        label.pack(side=tk.LEFT)
        label.bind("<Button-1>", lambda _event, i=index: self.select(i))
        label.bind("<Enter>", lambda _event, i=index: self.hover(i))
        label.bind("<Leave>", lambda _event, i=index: self.update_label(i))
        self.labels.append(label)
        page.pack_forget()
        if index == 0:
            self.select(0)
        else:
            self.update_label(index)

    def select(self, index: int) -> None:
        for page in self.pages:
            page.pack_forget()
        self.selected_index = index
        self.pages[index].pack(fill=tk.BOTH, expand=True)
        for i in range(len(self.labels)):
            self.update_label(i)

    def hover(self, index: int) -> None:
        if index != self.selected_index:
            self.labels[index].configure(bg="#f7faff", fg=TEXT_COLOR)

    def update_label(self, index: int) -> None:
        if index == self.selected_index:
            self.labels[index].configure(bg=SURFACE_BG, fg=PRIMARY, highlightbackground=PRIMARY)
            self.page_area.configure(highlightbackground=PRIMARY)
        else:
            self.labels[index].configure(bg=TAB_BG, fg=MUTED_TEXT, highlightbackground=BORDER_COLOR)


def validate_frequent_check_values(values: list[str], check_mode: str = "first") -> tuple[bool, str]:
    """초품 4Point 또는 하부 Pin 3개 확인 값을 검증합니다."""
    if len(values) != 12:
        return False, "확인 데이터가 올바르지 않습니다."
    first_count = sum(1 for value in values[:6] if value == "OK")
    jig_count = sum(1 for value in values[6:] if value == "OK")
    if check_mode == "jig":
        if jig_count == 0:
            return False, "하부 Pin 3개 확인은 1개 축 이상 선택해야 합니다."
        return True, f"하부 Pin {jig_count}개 축 확인 완료"
    if check_mode == "first" and first_count == 0:
        return False, "초품 4Point 확인은 1개 축 이상 선택해야 합니다."
    return True, f"초품 {first_count}개 축 확인 완료"


def count_frequent_check_axes(values: list[str]) -> int:
    """자주검사에서 실제 사용하는 축 수를 반환합니다.

    좌/우 그룹은 저장 전에 같은 개수로 검증되므로, 앞쪽 초품 확인 축 수를 기준으로 계산합니다.
    """
    return sum(1 for value in values[:6] if value == "OK")


def validate_frequent_check_capacity(lots: list[dict], stack: str, values: list[str]) -> tuple[bool, str]:
    """선택한 축 수 x Stack 수가 LOT 총 매수보다 작을 때만 차단합니다.

    Stack은 축당 최대 투입 가능 매수입니다. 예를 들어 10Stack에서 12매는
    2축으로 10매 + 2매 투입이 가능하므로 OK입니다.
    """
    axis_count = count_frequent_check_axes(values)
    stack_count = int(stack)
    total_qty = sum(int(lot.get("qty", "0")) for lot in lots)
    max_qty = axis_count * stack_count

    if total_qty > max_qty:
        shortage = total_qty - max_qty
        return (
            False,
            "초품 확인 축 수가 LOT 총 매수보다 부족합니다.\n\n"
            f"LOT 총 매수: {total_qty}매\n"
            f"선택 가능 수량: {axis_count}축 x {stack_count}Stack = 최대 {max_qty}매\n\n"
            f"부족 수량: {shortage}매\n\n"
            "초품 확인 축 수를 늘리거나 LOT 매수 / Stack 수를 확인해주세요.",
        )
    spare_qty = max_qty - total_qty
    return True, f"초품 확인 OK: LOT {total_qty}매 / {axis_count}축 x {stack_count}Stack = 최대 {max_qty}매 / 여유 {spare_qty}매"


class JiinDncManager:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1420x820")
        self.root.minsize(1180, 700)
        self.config = load_config()
        apply_theme(self.config.get("theme", "MES 블루"))
        self.root.configure(bg=APP_BG)
        try:
            sync_condition_master_from_completed_logs()
        except Exception:
            # 조건 마스터 동기화가 실패해도 현장 작업 화면은 열리게 둡니다.
            pass
        self.is_running = False

        self.common_entries: dict[str, LabeledEntry] = {}
        self.lot1_entries: dict[str, LabeledEntry] = {}
        self.lot2_entries: dict[str, LabeledEntry] = {}
        self.normal_buttons: list[ttk.Button] = []
        self.new_model_button: ttk.Button | None = None
        self.status_labels: dict[str, tk.Label] = {}
        self.lot_status_labels: dict[str, tk.Label] = {}
        self.log_text: scrolledtext.ScrolledText | None = None
        self.logo_image: tk.PhotoImage | None = None
        self.frequent_check_values: list[str] = [""] * 12
        self.work_axis_values: list[str] = [""] * 6
        self.lot_condition_keys: dict[int, str] = {1: "", 2: ""}
        self.current_work_period_key = ""
        self.last_common_manual_change_at: datetime | None = None

        self.setup_style()
        self.create_layout()
        self.apply_work_time_defaults(initial=True)
        self.update_status_checks()

    def setup_style(self) -> None:
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TFrame", background=APP_BG)
        style.configure("Header.TFrame", background=SURFACE_BG)
        style.configure("Panel.TFrame", background=SURFACE_BG)
        style.configure("TLabel", background=APP_BG, foreground=TEXT_COLOR, font=("맑은 고딕", 10))
        style.configure("Header.TLabel", background=SURFACE_BG, foreground=TEXT_COLOR, font=("맑은 고딕", 18, "bold"))
        style.configure("Hint.TLabel", background=APP_BG, foreground=MUTED_TEXT, font=("맑은 고딕", 10))
        style.configure("Panel.TLabel", background=SURFACE_BG, foreground=TEXT_COLOR, font=("맑은 고딕", 11, "bold"))
        style.configure("TButton", font=("맑은 고딕", 10), padding=(12, 8), background=SURFACE_BG, foreground=TEXT_COLOR)
        style.map("TButton", background=[("active", PRIMARY_LIGHT)])
        style.configure("Primary.TButton", font=("맑은 고딕", 11, "bold"), padding=(16, 10), background=PRIMARY_LIGHT, foreground=PRIMARY)
        style.configure("Side.TButton", font=("맑은 고딕", 10), padding=(12, 8), background=SURFACE_BG, foreground=TEXT_COLOR)
        style.configure("SidePrimary.TButton", font=("맑은 고딕", 10, "bold"), padding=(12, 8), background=PRIMARY_LIGHT, foreground=PRIMARY)
        style.configure("Wide.TEntry", padding=(8, 5), fieldbackground=SURFACE_BG)
        style.configure("White.TCombobox", padding=(8, 5), fieldbackground=SURFACE_BG, background=SURFACE_BG, foreground=TEXT_COLOR)
        style.map(
            "White.TCombobox",
            fieldbackground=[("readonly", SURFACE_BG), ("!disabled", SURFACE_BG)],
            background=[("readonly", SURFACE_BG), ("!disabled", SURFACE_BG)],
            foreground=[("readonly", TEXT_COLOR), ("!disabled", TEXT_COLOR)],
        )
        style.configure("Lookup.TEntry", padding=(8, 5), fieldbackground="#eef4fb", foreground=PRIMARY)
        style.map(
            "Lookup.TEntry",
            fieldbackground=[("readonly", "#eef4fb")],
            foreground=[("readonly", PRIMARY)],
        )
        style.configure("TCheckbutton", background=APP_BG, foreground=TEXT_COLOR, font=("맑은 고딕", 10))

    def create_layout(self) -> None:
        header = ttk.Frame(self.root, style="Header.TFrame", padding=(20, 12, 20, 12))
        header.pack(fill=tk.X, padx=12, pady=(12, 0))
        ttk.Label(header, text=APP_TITLE, style="Header.TLabel").pack(side=tk.LEFT)
        logo_path = BUNDLED_LOGO_FILE if BUNDLED_LOGO_FILE.exists() else LOGO_FILE
        if logo_path.exists():
            try:
                logo = tk.PhotoImage(file=str(logo_path))
                scale = max(1, logo.height() // 54)
                self.logo_image = logo.subsample(scale, scale)
                tk.Label(header, image=self.logo_image, bg=SURFACE_BG, bd=0).pack(side=tk.RIGHT)
            except tk.TclError:
                self.logo_image = None

        self.notebook = SimpleTabNotebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=12, pady=(10, 8))

        for tab_name in ["TLB", "심텍 SPS", "심텍 HDI"]:
            self.notebook.add(self.create_placeholder_tab(tab_name), tab_name)
        self.kcc_pkg_page = tk.Frame(self.notebook.page_area, bg=APP_BG)
        self.notebook.add(self.kcc_pkg_page, "KCC PKG")
        self.notebook.add(self.create_placeholder_tab("KCC HDI"), "KCC HDI")
        self.settings_page = tk.Frame(self.notebook.page_area, bg=APP_BG)
        self.notebook.add(self.settings_page, "설정")

        self.create_kcc_pkg_tab()
        self.create_settings_tab()
        self.notebook.select(3)

    def create_placeholder_tab(self, name: str) -> tk.Frame:
        page = tk.Frame(self.notebook.page_area, bg=APP_BG)
        tk.Label(page, text=f"{name}\n추후 개발 예정", bg=APP_BG, fg=MUTED_TEXT, font=("맑은 고딕", 22, "bold")).pack(expand=True)
        return page

    def create_kcc_pkg_tab(self) -> None:
        self.kcc_pkg_page.columnconfigure(0, weight=1)
        self.kcc_pkg_page.rowconfigure(2, weight=1)

        title_wrap = tk.Frame(self.kcc_pkg_page, bg=PRIMARY_LIGHT)
        title_wrap.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 8))
        title_wrap.columnconfigure(0, weight=1)
        title = tk.Label(
            title_wrap,
            text="KCC PKG 일반 DNC",
            bg=PRIMARY_LIGHT,
            fg=PRIMARY,
            font=("맑은 고딕", 14, "bold"),
            height=2,
        )
        title.grid(row=0, column=0, sticky="ew")
        title_buttons = tk.Frame(title_wrap, bg=PRIMARY_LIGHT)
        title_buttons.grid(row=0, column=1, sticky="e", padx=(8, 10))
        self.add_normal_button(title_buttons, "일반 DNC 실행", self.run_normal_dnc, "Primary.TButton").grid(row=0, column=0, padx=4, pady=4)
        self.add_normal_button(title_buttons, "입력 초기화", self.clear_normal_inputs).grid(row=0, column=1, padx=4, pady=4)

        title_legacy = tk.Label(
            self.kcc_pkg_page,
            text="KCC PKG 일반 DNC",
            bg=PRIMARY_LIGHT,
            fg=PRIMARY,
            font=("맑은 고딕", 14, "bold"),
            height=2,
        )

        common = self.create_panel(self.kcc_pkg_page, "공통 입력")
        common.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 8))
        common_widgets = [
            ("machine", ComboField(common, "설비 호기", ["트리밍 1호기", "트리밍 2호기", "트리밍 3호기"], initial=self.config.get("machine", "트리밍 1호기"), width=12)),
            ("work_date", DateField(common, "작업일자", on_change=self.mark_common_manual_change)),
            ("shift_group", SegmentedField(common, "조", ["A", "B", "C"], allow_empty=True, on_change=self.mark_common_manual_change)),
            ("shift", SegmentedField(common, "근무", ["주간", "야간"], on_change=self.mark_common_manual_change)),
            ("worker", LabeledEntry(common, "작업자", width=12, on_change=self.mark_common_manual_change)),
        ]
        for index, (key, entry) in enumerate(common_widgets):
            entry.grid(row=1, column=index, sticky="ew", padx=8, pady=8)
            self.common_entries[key] = entry
        common.columnconfigure(0, weight=0, minsize=320)
        common.columnconfigure(1, weight=1, minsize=430)
        common.columnconfigure(2, weight=1, minsize=330)
        common.columnconfigure(3, weight=1, minsize=360)
        common.columnconfigure(4, weight=0, minsize=320)

        lots = ttk.Frame(self.kcc_pkg_page)
        lots.grid(row=2, column=0, sticky="nsew", padx=14, pady=0)
        lots.columnconfigure(0, weight=1)
        lots.columnconfigure(1, weight=1)

        lot1 = self.create_lot_panel(lots, "LOT 1 입력", self.lot1_entries, 1)
        lot1.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=0)
        lot2 = self.create_lot_panel(lots, "LOT 2 입력 (선택)", self.lot2_entries, 2)
        lot2.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=0)

        bottom = ttk.Frame(self.kcc_pkg_page)
        bottom.grid(row=3, column=0, sticky="ew", padx=14, pady=(8, 14))
        bottom.columnconfigure(0, weight=1)

        status_panel = tk.Frame(bottom, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        status_panel.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        status_panel.columnconfigure(1, weight=1)
        status_panel.columnconfigure(2, weight=0)
        tk.Label(status_panel, text="2LOT 조건 일치 확인", bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 11, "bold"), width=22, height=2).grid(row=0, column=0, sticky="nsw")
        match_label = tk.Label(status_panel, text="LOT 2 미사용", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 12, "bold"), anchor="w")
        match_label.grid(row=0, column=1, sticky="ew", padx=14)
        self.status_labels["lot_match"] = match_label
        tk.Label(status_panel, text="DNC 진행 상태", bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 11, "bold"), width=22, height=2).grid(row=1, column=0, sticky="nsw")
        dnc_label = tk.Label(status_panel, text="대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 12, "bold"), anchor="w")
        dnc_label.grid(row=1, column=1, columnspan=2, sticky="ew", padx=14)
        self.status_labels["dnc"] = dnc_label
        tk.Label(status_panel, text="작업일보 반영", bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 11, "bold"), width=22, height=2).grid(row=2, column=0, sticky="nsw")
        excel_label = tk.Label(status_panel, text="대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 12, "bold"), anchor="w")
        excel_label.grid(row=2, column=1, columnspan=2, sticky="ew", padx=14)
        self.status_labels["excel"] = excel_label

        log_panel = tk.Frame(bottom, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        log_panel.grid(row=1, column=0, sticky="ew", padx=(0, 10), pady=(8, 0))
        log_panel.columnconfigure(0, weight=1)
        tk.Label(log_panel, text="DNC 작업 로그", bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 10, "bold"), height=1).grid(row=0, column=0, sticky="ew")
        self.log_text = scrolledtext.ScrolledText(
            log_panel,
            height=5,
            wrap=tk.WORD,
            state="disabled",
            bg=SURFACE_BG,
            fg=TEXT_COLOR,
            font=("맑은 고딕", 10),
            relief=tk.FLAT,
            padx=10,
            pady=8,
        )
        self.log_text.grid(row=1, column=0, sticky="ew")

        button_panel = ttk.Frame(bottom)
        button_panel.grid(row=0, column=1, rowspan=2, sticky="ne")
        for column in range(2):
            button_panel.columnconfigure(column, weight=1, uniform="side_buttons")
        self.new_model_button = self.add_side_button(button_panel, "신규 모델 검증 DNC", self.open_new_model_popup, "SidePrimary.TButton")
        self.new_model_button.grid(row=0, column=0, sticky="nsew", padx=4, pady=4)
        self.add_side_button(button_panel, "조건 마스터 관리", self.open_condition_master_popup, "SidePrimary.TButton").grid(row=0, column=1, sticky="nsew", padx=4, pady=4)
        self.add_side_button(button_panel, "작업일보 반영", self.export_kcc_pkg_to_excel_from_ui).grid(row=1, column=0, sticky="nsew", padx=4, pady=4)
        self.add_side_button(button_panel, "작업일보 열기", self.open_log_excel_from_ui).grid(row=1, column=1, sticky="nsew", padx=4, pady=4)

        for entry in list(self.lot1_entries.values()) + list(self.lot2_entries.values()):
            entry.var.trace_add("write", lambda *_args: self.update_status_checks())
        for lot_number, entries in ((1, self.lot1_entries), (2, self.lot2_entries)):
            for key in ("step", "round", "manage_no", "process_code"):
                entries[key].var.trace_add("write", lambda *_args, n=lot_number: self.handle_lot_key_change(n))

    def create_panel(self, parent, title: str) -> tk.Frame:
        panel = tk.Frame(parent, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        tk.Label(panel, text=title, bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 10, "bold"), height=2).grid(row=0, column=0, columnspan=8, sticky="ew")
        return panel

    def create_lot_panel(self, parent, title: str, target: dict[str, LabeledEntry], lot_number: int) -> tk.Frame:
        panel = self.create_panel(parent, title)
        fields = [
            ("step", "STEP"),
            ("round", "차수"),
            ("manage_no", "관리번호"),
            ("lot_no", "LOT No"),
            ("qty", "매수"),
            ("process_code", "공정코드"),
            ("condition", "조건(조회)"),
            ("jig", "지그(조회)"),
        ]
        for index, (key, label) in enumerate(fields):
            row = index // 2 + 1
            col = index % 2
            if key == "round":
                entry = RoundField(panel, label)
            elif key in {"condition", "jig"}:
                entry = LabeledEntry(panel, label, width=24, style="Lookup.TEntry", readonly=True)
            elif key in {"step", "qty"}:
                entry = LabeledEntry(panel, label, width=24, numeric_only=True)
            else:
                entry = LabeledEntry(panel, label, width=24, uppercase=True)
            entry.grid(row=row, column=col, sticky="ew", padx=10, pady=8)
            panel.columnconfigure(col, weight=1)
            target[key] = entry
        load_button = ttk.Button(
            panel,
            text="조건 / 지그 조회",
            command=lambda: self.load_condition_jig_for_lot(lot_number),
            style="Primary.TButton",
        )
        load_button.grid(row=5, column=0, columnspan=2, sticky="ew", padx=10, pady=(8, 6))

        status = tk.Frame(panel, bg=SURFACE_BG)
        status.grid(row=6, column=0, columnspan=2, sticky="ew", padx=10, pady=(0, 12))
        status.columnconfigure(0, weight=1)
        status.columnconfigure(1, weight=1)
        mes_label = self.create_judgement_card(status, "MES Core")
        mes_label.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        condition_label = self.create_judgement_card(status, "조건 적용")
        condition_label.grid(row=0, column=1, sticky="ew", padx=(6, 0))
        self.lot_status_labels[f"lot{lot_number}_mes"] = mes_label
        self.lot_status_labels[f"lot{lot_number}_condition"] = condition_label
        return panel

    def create_judgement_card(self, parent, title: str) -> tk.Label:
        """LOT별 핵심 판정 상태를 크게 보여주는 카드형 라벨입니다."""
        return tk.Label(
            parent,
            text=f"{title}\n대기중",
            bg="#f8fafc",
            fg=MUTED_TEXT,
            font=("맑은 고딕", 12, "bold"),
            height=3,
            relief=tk.SOLID,
            bd=1,
            justify=tk.CENTER,
            anchor="center",
        )

    def add_normal_button(self, parent, text: str, command, style: str = "TButton") -> ttk.Button:
        button = ttk.Button(parent, text=text, command=command, style=style, width=18)
        self.normal_buttons.append(button)
        return button

    def add_side_button(self, parent, text: str, command, style: str = "Side.TButton") -> ttk.Button:
        button = ttk.Button(parent, text=text, command=command, style=style, width=18)
        self.normal_buttons.append(button)
        return button

    def create_settings_tab(self) -> None:
        page = self.settings_page
        page.columnconfigure(0, weight=1)
        panel = self.create_panel(page, "설정")
        panel.grid(row=0, column=0, sticky="ew", padx=14, pady=14)
        panel.columnconfigure(1, weight=1)

        self.excel_var = tk.StringVar(value=self.config.get("excel_file", ""))
        source_folders = self.config.get("source_dnc_folders", {})
        self.source_var = tk.StringVar(value=source_folders.get("KCC PKG", self.config.get("source_dnc_folder", str(get_desktop_path() / "KCC_PKG"))))
        self.source_vars = {
            process_name: tk.StringVar(value=source_folders.get(process_name, ""))
            for process_name in PROCESS_NAMES
        }
        self.transfer_var = tk.StringVar(value=self.config.get("transfer_dnc_folder", str(get_desktop_path() / "DNC")))
        self.delete_seconds_var = tk.StringVar(value=str(self.config.get("dnc_delete_seconds", DNC_DELETE_SECONDS)))
        self.first_article_wait_var = tk.StringVar(value=str(self.config.get("first_article_wait_seconds", FIRST_ARTICLE_WAIT_SECONDS)))
        self.clear_common_var = tk.BooleanVar(value=bool(self.config.get("clear_common_after_normal", False)))
        self.auto_shift_group_var = tk.BooleanVar(value=bool(self.config.get("auto_shift_group_enabled", True)))
        self.a_group_day_start_var = tk.StringVar(value=str(self.config.get("a_group_day_start_date", "2026-05-17")))

        common_rows = [
            ("작업일보 경로", self.excel_var, lambda: select_excel_file(self.root, self.config, self.excel_var)),
            ("DNC 전송 폴더", self.transfer_var, lambda: self.select_folder_to_var(self.transfer_var, save_after=True)),
        ]
        for row, (label, var, command) in enumerate(common_rows, start=1):
            ttk.Label(panel, text=label, background=SURFACE_BG, width=16).grid(row=row, column=0, sticky="e", padx=10, pady=8)
            entry = ttk.Entry(panel, textvariable=var, style="Wide.TEntry")
            entry.grid(row=row, column=1, sticky="ew", padx=8, pady=8)
            entry.bind("<FocusOut>", lambda _event: self.save_settings_from_ui_silent(show_error=False))
            ttk.Button(panel, text="선택", command=command, width=20).grid(row=row, column=2, padx=8, pady=8)

        ttk.Label(panel, text="공정별 원본 DNC 폴더", background=PRIMARY_LIGHT, foreground=PRIMARY, anchor="center", font=("맑은 고딕", 10, "bold")).grid(
            row=3, column=0, columnspan=3, sticky="ew", padx=10, pady=(16, 6)
        )
        process_bg = ["#eef6ff", "#ecfdf5", "#f5f3ff", "#f8fafc", "#ecfeff"]
        for index, process_name in enumerate(PROCESS_NAMES):
            row = 4 + index
            label_bg = process_bg[index % len(process_bg)]
            ttk.Label(panel, text=process_name, background=label_bg, foreground=TEXT_COLOR, width=16, anchor="center").grid(row=row, column=0, sticky="ew", padx=10, pady=5)
            entry = ttk.Entry(panel, textvariable=self.source_vars[process_name], style="Wide.TEntry")
            entry.grid(row=row, column=1, sticky="ew", padx=8, pady=5)
            entry.bind("<FocusOut>", lambda _event: self.save_settings_from_ui_silent(show_error=False))
            ttk.Button(panel, text="폴더 선택", command=lambda name=process_name: self.select_source_folder(name), width=20).grid(row=row, column=2, padx=8, pady=5)

        shift_row = 9
        ttk.Label(panel, text="근무표 자동 조", background=PRIMARY_LIGHT, foreground=PRIMARY, anchor="center", font=("맑은 고딕", 10, "bold")).grid(
            row=shift_row, column=0, columnspan=3, sticky="ew", padx=10, pady=(16, 6)
        )
        auto_wrap = tk.Frame(panel, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        auto_wrap.grid(row=shift_row + 1, column=0, columnspan=3, sticky="ew", padx=10, pady=5)
        auto_wrap.columnconfigure(2, weight=1)
        tk.Label(auto_wrap, text="조 자동 선택", bg=SURFACE_BG, fg=TEXT_COLOR, width=16, anchor="e").grid(row=0, column=0, sticky="e", padx=(10, 6), pady=8)
        toggle_wrap = tk.Frame(auto_wrap, bg=APP_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        toggle_wrap.grid(row=0, column=1, sticky="w", padx=(0, 18), pady=8)
        self.auto_shift_buttons = {}
        for value, text in ((True, "사용"), (False, "미사용")):
            button = tk.Button(
                toggle_wrap,
                text=text,
                command=lambda selected=value: self.set_auto_shift_group_enabled(selected),
                relief=tk.FLAT,
                bd=0,
                width=10,
                padx=10,
                pady=5,
                cursor="hand2",
                font=("맑은 고딕", 10, "bold"),
            )
            button.pack(side=tk.LEFT, fill=tk.X, expand=True)
            self.auto_shift_buttons[value] = button
        self.update_auto_shift_buttons()
        tk.Label(auto_wrap, text="A조 주간 시작 기준일", bg=PRIMARY_LIGHT, fg=TEXT_COLOR, width=18, anchor="center").grid(row=0, column=2, sticky="e", padx=(0, 8), pady=8)
        a_entry = ttk.Entry(auto_wrap, textvariable=self.a_group_day_start_var, style="Wide.TEntry", width=14)
        a_entry.grid(row=0, column=3, sticky="w", padx=(0, 8), pady=8)
        tk.Label(auto_wrap, text="예: 2026-05-17", bg=SURFACE_BG, fg=MUTED_TEXT).grid(row=0, column=4, sticky="w", padx=(0, 10), pady=8)
        apply_button = tk.Button(
            auto_wrap,
            text="적용",
            command=self.apply_auto_shift_settings,
            relief=tk.FLAT,
            bd=0,
            width=10,
            padx=10,
            pady=5,
            cursor="hand2",
            bg=PRIMARY_LIGHT,
            fg=PRIMARY,
            activebackground=PRIMARY_LIGHT,
            activeforeground=PRIMARY,
            font=("맑은 고딕", 10, "bold"),
        )
        apply_button.grid(row=0, column=5, sticky="e", padx=(4, 10), pady=8)

        ttk.Button(panel, text="마스터 설정", command=self.open_master_settings_popup, style="Primary.TButton", width=20).grid(row=shift_row + 2, column=2, padx=8, pady=(18, 14))

    def select_folder_to_var(self, var: tk.StringVar, save_after: bool = False) -> None:
        folder = filedialog.askdirectory(initialdir=var.get() or str(get_desktop_path()))
        if folder:
            var.set(folder)
            if save_after:
                self.save_settings_from_ui_silent(show_error=False)

    def select_source_folder(self, process_name: str) -> None:
        var = self.source_vars[process_name]
        folder = filedialog.askdirectory(initialdir=var.get() or str(get_desktop_path()))
        if folder:
            var.set(folder)
            self.save_settings_from_ui_silent(show_error=False)

    def set_auto_shift_group_enabled(self, enabled: bool) -> None:
        self.auto_shift_group_var.set(enabled)
        self.update_auto_shift_buttons()

    def apply_auto_shift_settings(self) -> None:
        if self.save_settings_from_ui_silent(show_error=True):
            self.apply_work_time_defaults(initial=False, schedule_next=False)
            messagebox.showinfo("근무표 자동 조", "근무표 자동 조 설정을 적용했습니다.", parent=self.root)

    def update_auto_shift_buttons(self) -> None:
        if not hasattr(self, "auto_shift_buttons"):
            return
        for value, button in self.auto_shift_buttons.items():
            selected = bool(self.auto_shift_group_var.get()) == value
            button.configure(
                bg=PRIMARY_LIGHT if selected else SURFACE_BG,
                fg=PRIMARY if selected else TEXT_COLOR,
                activebackground=PRIMARY_LIGHT,
                activeforeground=PRIMARY,
            )

    def save_settings_from_ui(self) -> None:
        self.save_settings_from_ui_silent(show_error=True)

    def rebuild_condition_master(self) -> None:
        self.save_settings_from_ui_silent()
        try:
            count = rebuild_condition_master_from_log(self.config)
        except Exception as exc:
            messagebox.showerror("조건 마스터 갱신 실패", str(exc))
            return
        messagebox.showinfo("조건 마스터 갱신 완료", f"{count}개 조건을 저장했습니다.")

    def open_work_history_popup(self) -> None:
        WorkHistoryPopup(self)

    def open_master_settings_popup(self) -> None:
        password = simpledialog.askstring("마스터 설정", "비밀번호를 입력하세요.", show="*", parent=self.root)
        if password is None:
            return
        if password != str(self.config.get("master_password", MASTER_SETTINGS_PASSWORD)):
            messagebox.showwarning("비밀번호 확인", "비밀번호가 맞지 않습니다.")
            return
        MasterSettingsPopup(self)

    def get_common_data(self) -> dict:
        return {key: entry.get() for key, entry in self.common_entries.items()}

    def get_lot_data(self, entries: dict[str, LabeledEntry]) -> dict:
        return {key: entry.get() for key, entry in entries.items()}

    def lot_has_any_value(self, lot: dict) -> bool:
        # LOT 2는 선택 입력이므로 차수 버튼만 눌린 상태는 사용으로 보지 않습니다.
        return any(value.strip() for key, value in lot.items() if key != "round")

    def has_new_model_target(self) -> bool:
        """조건/지그가 없는 신규 LOT가 있거나 입력 LOT가 없을 때만 신규 검증을 허용합니다."""
        used_lot_count = 0
        for entries in (self.lot1_entries, self.lot2_entries):
            lot = self.get_lot_data(entries)
            if not self.lot_has_any_value(lot):
                continue
            used_lot_count += 1
            condition_ok, _message = get_single_condition_message(lot)
            if not condition_ok:
                return True
        return used_lot_count == 0

    def update_new_model_button_state(self) -> None:
        if self.new_model_button is None:
            return
        state = "normal" if self.has_new_model_target() and not self.is_running else "disabled"
        self.new_model_button.configure(state=state)

    def mark_common_manual_change(self) -> None:
        self.last_common_manual_change_at = datetime.now()
        if not self.common_entries:
            return
        if self.common_entries["shift_group"].get() and self.common_entries["worker"].get():
            dnc_label = self.status_labels.get("dnc")
            if dnc_label and "근무" in str(dnc_label.cget("text")):
                self.set_status("dnc", "대기중", None)

    def get_work_prep_start(self, period: dict[str, str]) -> datetime:
        work_date = datetime.strptime(period["work_date"], "%Y-%m-%d")
        if period["shift"] == "주간":
            return work_date.replace(hour=8, minute=0, second=0, microsecond=0)
        return work_date.replace(hour=20, minute=0, second=0, microsecond=0)

    def has_prepared_next_shift(self, period: dict[str, str]) -> bool:
        changed_at = self.last_common_manual_change_at
        if changed_at is None:
            return False
        if changed_at < self.get_work_prep_start(period):
            return False
        return bool(self.common_entries["shift_group"].get() and self.common_entries["worker"].get())

    def apply_work_time_defaults(self, initial: bool = False, schedule_next: bool = True) -> None:
        """작업일자와 근무를 시간 기준으로 자동 적용하고 근무 전환 시 조/작업자를 다시 입력하게 합니다."""
        if not self.common_entries:
            return
        period = get_work_period()
        period_changed = bool(self.current_work_period_key and self.current_work_period_key != period["period_key"])
        prepared_next_shift = period_changed and self.has_prepared_next_shift(period)
        self.current_work_period_key = period["period_key"]
        self.common_entries["work_date"].set(period["work_date"])
        self.common_entries["shift"].set(period["shift"])
        auto_group = ""
        if self.config.get("auto_shift_group_enabled", True):
            auto_group = get_auto_shift_group(
                period["work_date"],
                period["shift"],
                str(self.config.get("a_group_day_start_date", "2026-05-17")),
            )
            if auto_group:
                self.common_entries["shift_group"].set(auto_group)
        if period_changed and not prepared_next_shift:
            if not auto_group:
                self.common_entries["shift_group"].clear()
            self.common_entries["worker"].clear()
            self.set_status("dnc", "근무 전환 확인 필요", False)
            if not initial:
                messagebox.showwarning(
                    "근무 전환 확인",
                    "작업일자/근무가 자동 변경되었습니다.\n\n"
                    "작업자를 다시 입력해야 DNC를 진행할 수 있습니다.",
                    parent=self.root,
                )
        elif period_changed and prepared_next_shift:
            self.set_status("dnc", "근무 전환 확인 완료", True)
        if schedule_next:
            self.root.after(60000, self.apply_work_time_defaults)

    def ensure_work_period_ready(self) -> bool:
        self.apply_work_time_defaults(initial=False, schedule_next=False)
        missing = []
        if not self.common_entries["shift_group"].get():
            missing.append("조")
        if not self.common_entries["worker"].get():
            missing.append("작업자")
        if missing:
            shift_notice = "08:30 / 20:30 근무 전환 후에는 작업자를 다시 확인해주세요."
            if not self.config.get("auto_shift_group_enabled", True):
                shift_notice = "08:30 / 20:30 근무 전환 후에는 조와 작업자를 다시 확인해주세요."
            messagebox.showwarning(
                "근무 정보 확인",
                " / ".join(missing) + "을(를) 입력해야 DNC를 진행할 수 있습니다.\n\n"
                f"{shift_notice}",
                parent=self.root,
            )
            self.set_status("dnc", "근무 정보 확인 필요", False)
            return False
        dnc_label = self.status_labels.get("dnc")
        if dnc_label and "근무" in str(dnc_label.cget("text")):
            self.set_status("dnc", "대기중", None)
        return True

    def set_status(self, key: str, text: str, ok: bool | None = None) -> None:
        color = MUTED_TEXT if ok is None else (OK_COLOR if ok else NG_COLOR)
        self.status_labels[key].configure(text=text, fg=color)
        if key == "dnc":
            self.append_log(text)

    def set_dnc_status(self, text: str) -> None:
        self.root.after(0, lambda: self.set_status("dnc", text, None))

    def append_log(self, text: str) -> None:
        if self.log_text is None:
            return
        if text.startswith("DNC 삭제 대기중"):
            return
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert(tk.END, f"[{timestamp}] {text}\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state="disabled")

    def update_status_checks(self) -> None:
        lot1 = self.get_lot_data(self.lot1_entries)
        lot2 = self.get_lot_data(self.lot2_entries)
        lot2_used = self.lot_has_any_value(lot2)
        self.update_lot_detail_status("lot1", lot1, waiting_when_empty=True)
        if lot2_used:
            self.update_lot_detail_status("lot2", lot2, waiting_when_empty=True)
            lot_match, lot_match_message = get_lot_match_message(lot1, lot2)
            if lot_match:
                self.set_status("lot_match", lot_match_message, True)
            else:
                self.set_status("lot_match", f"NG - {lot_match_message}", False)
        else:
            self.set_lot_status("lot2_mes", "대기중", None)
            self.set_lot_status("lot2_condition", "대기중", None)
            self.set_status("lot_match", "LOT 2 미사용", None)
        self.update_new_model_button_state()

    def update_lot_detail_status(self, prefix: str, lot: dict, waiting_when_empty: bool) -> None:
        lot_no = lot.get("lot_no", "").strip()
        process_code = lot.get("process_code", "").strip()
        condition = lot.get("condition", "").strip()
        jig = lot.get("jig", "").strip()

        if waiting_when_empty and not lot_no and not process_code:
            self.set_lot_status(f"{prefix}_mes", "대기중", None)
        else:
            mes_ok, mes_message = get_mes_core_message(lot_no, process_code)
            if mes_ok:
                self.set_lot_status(f"{prefix}_mes", "OK", True)
            else:
                self.set_lot_status(f"{prefix}_mes", "NG", False)

        if waiting_when_empty and not condition and not jig:
            self.set_lot_status(f"{prefix}_condition", "대기중", None)
        else:
            condition_ok, condition_message = get_single_condition_message(lot)
            self.set_lot_status(f"{prefix}_condition", "OK" if condition_ok else "NG", condition_ok)

    def set_lot_status(self, key: str, text: str, ok: bool | None = None) -> None:
        if key not in self.lot_status_labels:
            return
        label = self.lot_status_labels[key]
        title = "MES Core" if key.endswith("_mes") else "조건 적용"
        if ok is True:
            bg = "#dcfce7"
            fg = OK_COLOR
            border = OK_COLOR
        elif ok is False:
            bg = "#fee2e2"
            fg = NG_COLOR
            border = NG_COLOR
        else:
            bg = "#f8fafc"
            fg = MUTED_TEXT
            border = BORDER_COLOR
        label.configure(text=f"{title}\n{text}", fg=fg, bg=bg, highlightthickness=2, highlightbackground=border)

    def set_running(self, running: bool) -> None:
        self.is_running = running
        state = "disabled" if running else "normal"
        for button in self.normal_buttons:
            button.configure(state=state)
        if not running:
            self.update_new_model_button_state()

    def validate_condition_file(self, condition_name: str) -> Path | None:
        self.set_dnc_status("조건 파일 검색중")
        matches = search_condition_file(condition_name, Path(self.config["source_dnc_folder"]))
        if len(matches) == 0:
            messagebox.showerror("조건 파일 없음", f"복사할 조건 파일이 없습니다.\n조건명: {condition_name}\n관리자 확인 필요.")
            return None
        if len(matches) >= 2:
            messagebox.showerror(
                "동일 프로그램 차단",
                f"동일 프로그램 확인으로 차단!\n관리자 확인 필요!\n조건명: {condition_name}\n검색 수량: {len(matches)}개",
            )
            return None
        return matches[0]

    def make_lot_lookup_key(self, lot: dict) -> str:
        """조건/지그를 불러온 기준 키입니다. 이 값이 바뀌면 기존 조건을 비웁니다."""
        return "|".join(
            [
                lot.get("step", "").strip(),
                lot.get("round", "").strip(),
                lot.get("process_code", "").strip(),
                lot.get("manage_no", "").strip(),
            ]
        )

    def handle_lot_key_change(self, lot_number: int) -> None:
        entries = self.lot1_entries if lot_number == 1 else self.lot2_entries
        loaded_key = self.lot_condition_keys.get(lot_number, "")
        if not loaded_key:
            return
        current_key = self.make_lot_lookup_key(self.get_lot_data(entries))
        if current_key == loaded_key:
            return
        entries["condition"].clear()
        entries["jig"].clear()
        self.lot_condition_keys[lot_number] = ""
        self.set_status("dnc", f"LOT {lot_number} 기준값 변경 - 작업조건/지그 초기화", None)

    def load_condition_jig_for_lot(self, lot_number: int) -> bool:
        """작업일보 이력에서 선택 LOT의 작업조건/지그를 불러와 화면에 채웁니다."""
        self.save_settings_from_ui_silent()
        entries = self.lot1_entries if lot_number == 1 else self.lot2_entries
        lot = self.get_lot_data(entries)
        missing = [
            label
            for key, label in (("step", "STEP"), ("round", "차수"), ("process_code", "공정코드"))
            if not lot.get(key, "").strip()
        ]
        if missing:
            messagebox.showwarning(
                "이력 조회",
                f"LOT {lot_number} {' / '.join(missing)} 입력 후 불러오세요.\n\n"
                "작업조건/지그는 STEP, 차수, 공정코드가 모두 맞을 때만 불러옵니다.",
            )
            return False
        try:
            condition, jig, source = lookup_condition_jig_from_history(self.config, lot)
        except Exception as exc:
            messagebox.showerror("이력 조회 실패", str(exc))
            return False
        if not condition or not jig:
            detail = describe_condition_lookup_mismatch(lot)
            log_app(
                "조건 조회 실패: "
                f"LOT {lot_number}, STEP={lot.get('step')}, 차수={lot.get('round')}, "
                f"관리번호={lot.get('manage_no')}, 공정코드={lot.get('process_code')} / {detail}"
            )
            messagebox.showwarning(
                "이력 없음",
                "STEP, 차수, 공정코드가 일치하는 작업조건/지그를 찾을 수 없습니다.\n\n"
                f"LOT {lot_number}\n"
                f"STEP: {lot.get('step') or '빈칸'}\n"
                f"차수: {lot.get('round') or '빈칸'}\n"
                f"공정코드: {lot.get('process_code') or '빈칸'}\n"
                f"관리번호: {lot.get('manage_no') or '빈칸'}\n\n"
                f"{detail}",
            )
            return False
        entries["condition"].set(condition)
        entries["jig"].set(jig)
        refreshed_lot = self.get_lot_data(entries)
        self.lot_condition_keys[lot_number] = self.make_lot_lookup_key(refreshed_lot)
        self.update_status_checks()
        current_condition = lot.get("condition", "").strip()
        current_jig = lot.get("jig", "").strip()
        if current_condition != condition or current_jig != jig:
            self.set_status("dnc", f"LOT {lot_number} 조건/지그 불러옴: {source}", True)
        return True

    def run_normal_dnc(self) -> None:
        if self.is_running:
            messagebox.showwarning("진행 중", "DNC 실행중입니다.\n작업 완료 후 다시 실행해주세요.")
            return
        if not self.ensure_work_period_ready():
            return
        if not self.save_settings_from_ui_silent():
            return
        self.set_status("dnc", "입력값 확인중", None)
        common = self.get_common_data()
        lot1 = self.get_lot_data(self.lot1_entries)
        lot2_data = self.get_lot_data(self.lot2_entries)
        lot2 = lot2_data if self.lot_has_any_value(lot2_data) else None

        # Normal DNC conditions are lookup values, not operator-entered values.
        # Refresh them right before running so a stale condition cannot remain
        # after STEP / round / process code was changed on screen.
        if not self.load_condition_jig_for_lot(1):
            return
        lot1 = self.get_lot_data(self.lot1_entries)
        if lot2:
            if not self.load_condition_jig_for_lot(2):
                return
            lot2 = self.get_lot_data(self.lot2_entries)
        ok, errors = validate_normal_dnc(common, lot1, lot2)
        if not ok:
            messagebox.showwarning("입력값 확인", "\n".join(errors))
            self.set_status("dnc", "입력값 NG", False)
            return
        model_change = messagebox.askyesno("기종교체 확인", "기종교체 입니까?", parent=self.root)
        self.frequent_check_values = [""] * 12
        self.work_axis_values = [""] * 6
        if model_change:
            self.set_status("dnc", "하부 Pin 확인 대기중", None)
            if not self.open_frequent_check_popup("jig"):
                self.set_status("dnc", "하부 Pin 확인 미완료", False)
                return
        stack = ask_numeric_input(self.root, "Stack 수 입력", "Stack 수를 입력 하세요.")
        ok, message = validate_positive_number(stack or "", "Stack 수", required=True)
        if not ok:
            messagebox.showwarning("Stack 수 확인", message)
            return
        lots = [lot1] + ([lot2] if lot2 else [])
        if model_change:
            capacity_values = ["OK" if value == "OK" else "" for value in self.frequent_check_values[6:]] + [""] * 6
            self.work_axis_values = capacity_values[:6]
        else:
            self.set_status("dnc", "작업 축 수 확인 대기중", None)
            self.frequent_check_values = [""] * 12
            if not self.open_frequent_check_popup("capacity"):
                self.set_status("dnc", "작업 축 수 확인 미완료", False)
                return
            capacity_values = self.frequent_check_values[:]
            self.work_axis_values = capacity_values[:6]
        capacity_ok, capacity_message = validate_frequent_check_capacity(lots, stack, capacity_values)
        if not capacity_ok:
            messagebox.showwarning("작업 수량 확인", capacity_message)
            self.set_status("dnc", "수량/Stack NG", False)
            return
        if model_change:
            self.set_status("dnc", "작업 수량 확인 완료", True)
        condition_file = self.validate_condition_file(lot1["condition"])
        if not condition_file:
            self.set_status("dnc", "조건 파일 NG", False)
            return

        log_app(f"일반 DNC 시작: LOT수={len(lots)}, 조건={lot1['condition']}, 기종교체={model_change}")
        self.set_running(True)
        threading.Thread(target=self.normal_worker, args=(common, lots, stack, model_change, condition_file), daemon=True).start()

    def normal_worker(self, common: dict, lots: list[dict], stack: str, model_change: bool, condition_file: Path) -> None:
        try:
            self.set_dnc_status("DB 저장중")
            log_ids = insert_normal_dnc_db(common, lots, stack, model_change)
            self.set_dnc_status("DB 저장 완료")
            log_app(f"일반 DNC 파일 처리 시작: ids={log_ids}")
            delete_existing_dnc_txt(Path(self.config["transfer_dnc_folder"]))
            log_app("DNC 전송 폴더 기존 txt 삭제 완료")
            copied_file = copy_dnc_file(condition_file, Path(self.config["transfer_dnc_folder"]))
            self.set_dnc_status("DNC 파일 복사 완료")
            log_app(f"DNC 파일 복사 완료: {copied_file}")
            delete_thread = threading.Thread(
                target=delete_after_delay,
                args=(copied_file, int(self.config["dnc_delete_seconds"]), self.set_dnc_status),
                daemon=True,
            )
            delete_thread.start()
            wait_seconds = int(self.config.get("first_article_wait_seconds", FIRST_ARTICLE_WAIT_SECONDS))
            log_app(f"초품 확인 대기 시작: {wait_seconds}초")
            for remain in range(wait_seconds, 0, -1):
                self.set_dnc_status(f"초품 확인 대기중 ({remain}초)")
                time.sleep(1)
            log_app("초품 확인 팝업 호출")
            self.root.after(0, lambda: self.finish_normal_dnc(log_ids, lots, stack, model_change, delete_thread))
        except Exception as exc:
            self.root.after(0, lambda error=exc: self.handle_run_error(error))

    def release_running_after_delete(self, delete_thread: threading.Thread | None) -> None:
        if delete_thread and delete_thread.is_alive():
            self.set_status("dnc", "DNC 삭제 완료 대기중", None)

            def wait_and_release() -> None:
                delete_thread.join()
                self.root.after(0, lambda: self.set_running(False))

            threading.Thread(target=wait_and_release, daemon=True).start()
            return
        self.set_running(False)

    def finish_normal_dnc(self, log_ids: list[int], lots: list[dict], stack: str, model_change: bool, delete_thread: threading.Thread | None = None) -> None:
        try:
            while True:
                allowed_axes = None
                if model_change:
                    allowed_axes = [
                        index
                        for index, value in enumerate(self.frequent_check_values[6:])
                        if value == "OK"
                    ]
                    for index in range(6):
                        self.frequent_check_values[index] = ""
                else:
                    allowed_axes = [
                        index
                        for index, value in enumerate(self.work_axis_values)
                        if value == "OK"
                    ]
                    self.frequent_check_values = [""] * 12
                if not self.open_frequent_check_popup("first", allowed_axes=allowed_axes):
                    self.set_status("dnc", "초품 확인 미완료", False)
                    return
                if model_change:
                    first_axes = [
                        index
                        for index, value in enumerate(self.frequent_check_values[:6])
                        if value == "OK"
                    ]
                    if first_axes != allowed_axes:
                        messagebox.showwarning(
                            "초품 4Point 확인",
                            "하부 Pin 확인 축과 초품 확인 축이 같아야 합니다.\n\n"
                            f"하부 Pin 확인 축: {', '.join(str(axis + 1) + '축' for axis in allowed_axes)}\n"
                            f"초품 확인 축: {', '.join(str(axis + 1) + '축' for axis in first_axes) or '없음'}",
                        )
                        self.set_status("dnc", "초품 축 확인 NG", False)
                        continue
                ok, message = validate_frequent_check_capacity(lots, stack, self.frequent_check_values)
                if ok:
                    self.set_status("dnc", message, True)
                    break
                messagebox.showwarning("초품 수량 확인", message)
                self.set_status("dnc", "초품 수량 NG", False)
            update_normal_frequent_check_db(log_ids, model_change, self.frequent_check_values)
            burr_ok = messagebox.askyesno("Burr 확인", "4면 Burr 이상 없습니까?")
            update_normal_burr_db(log_ids, burr_ok)
            pending_count = get_unexported_kcc_pkg_count()
            self.set_status("dnc", "DNC 완료", True)
            self.set_status("excel", f"DB 저장 완료 / Excel 미반영 {pending_count}건", True)
            log_app(f"일반 DNC 완료: ids={log_ids}, Excel 미반영={pending_count}건")
            self.auto_export_kcc_pkg_to_excel(parent=self.root)
            self.clear_normal_inputs(after_done=True)
        except Exception as exc:
            self.handle_run_error(exc)
        finally:
            self.release_running_after_delete(delete_thread)

    def clear_normal_inputs(self, after_done: bool = False) -> None:
        for entry in self.lot1_entries.values():
            entry.clear()
        for entry in self.lot2_entries.values():
            entry.clear()
        self.frequent_check_values = [""] * 12
        self.work_axis_values = [""] * 6
        self.lot_condition_keys = {1: "", 2: ""}
        self.update_status_checks()
        if after_done:
            self.set_status("dnc", "DNC 완료", True)
        else:
            self.set_status("dnc", "대기중", None)

    def open_new_model_popup(self) -> None:
        if not self.has_new_model_target():
            messagebox.showinfo(
                "신규 모델 검증 DNC",
                "현재 입력된 LOT는 조건/지그가 이미 조회되어 신규 검증 대상이 아닙니다.\n\n일반 DNC 실행을 사용해주세요.",
                parent=self.root,
            )
            return
        if not self.ensure_work_period_ready():
            return
        open_new_model_popup(self)

    def open_log_excel_from_ui(self) -> None:
        self.save_settings_from_ui_silent()
        if ensure_excel_file_selected(self.root, self.config, self.excel_var if hasattr(self, "excel_var") else None):
            open_log_excel(self.config)

    def auto_export_kcc_pkg_to_excel(self, parent=None) -> bool:
        """DNC 완료 후 작업일보 반영을 자동으로 시도합니다. 실패해도 DB 미반영 상태는 유지됩니다."""
        pending_before = get_unexported_kcc_pkg_count()
        if pending_before == 0:
            self.set_status("excel", "Excel 미반영 0건", True)
            return True
        if not self.config.get("auto_export_after_dnc", True):
            self.set_status("excel", f"자동 반영 꺼짐 / Excel 미반영 {pending_before}건", False)
            return False
        excel_path = Path(self.config.get("excel_file", ""))
        if not excel_path.exists():
            self.set_status("excel", f"작업일보 경로 필요 / Excel 미반영 {pending_before}건", False)
            self.append_log(f"작업일보 자동 반영 실패: 작업일보 파일 없음 / 미반영 {pending_before}건")
            return False
        try:
            exported_count = export_kcc_pkg_db_to_excel(self.config)
        except Exception as exc:
            pending_after = get_unexported_kcc_pkg_count()
            self.set_status("excel", f"자동 반영 실패 / Excel 미반영 {pending_after}건", False)
            self.append_log(f"작업일보 자동 반영 실패: {exc}")
            messagebox.showwarning(
                "작업일보 자동 반영 실패",
                "DNC 이력은 DB에 저장되어 있습니다.\n\n"
                "작업일보 Excel 파일이 열려 있거나 경로가 맞지 않아 자동 반영하지 못했습니다.\n"
                "Excel 파일을 닫은 뒤 [작업일보 반영]을 눌러주세요.",
                parent=parent or self.root,
            )
            return False
        pending_after = get_unexported_kcc_pkg_count()
        self.set_status("excel", f"작업일보 자동 반영 완료 / Excel 미반영 {pending_after}건", True)
        self.append_log(f"작업일보 자동 반영 완료: {exported_count}건 / 미반영 {pending_after}건")
        return True

    def export_kcc_pkg_to_excel_from_ui(self) -> None:
        """전체 공정 DB에 저장된 미반영 이력을 Excel 작업일보로 내보냅니다."""
        self.save_settings_from_ui_silent()
        if not ensure_excel_file_selected(self.root, self.config, self.excel_var if hasattr(self, "excel_var") else None):
            self.set_status("excel", "작업일보 반영 취소", False)
            return
        try:
            result = export_all_processes_to_excel(self.config)
        except FileNotFoundError as exc:
            messagebox.showerror("작업일보 파일 없음", str(exc))
            self.set_status("excel", "작업일보 파일 없음", False)
            return
        except PermissionError as exc:
            messagebox.showerror("작업일보 저장 실패", str(exc))
            self.set_status("excel", "작업일보 저장 실패", False)
            return
        except KeyError as exc:
            messagebox.showerror("작업일보 시트 없음", str(exc))
            self.set_status("excel", "작업일보 시트 없음", False)
            return
        except ValueError as exc:
            messagebox.showerror("작업일보 파일 오류", str(exc))
            self.set_status("excel", "작업일보 파일 오류", False)
            return
        except Exception as exc:
            messagebox.showerror("작업일보 반영 실패", str(exc))
            self.set_status("excel", "작업일보 반영 실패", False)
            return
        total_count = sum(result.values())
        pending_total = sum(get_unexported_process_count(name) for name in PROCESS_NAMES)
        if total_count == 0:
            messagebox.showinfo("작업일보 반영", "Excel에 반영할 DB 이력이 없습니다.")
            self.set_status("excel", "Excel 미반영 0건", True)
            return
        detail = "\n".join(f"{name}: {count}건" for name, count in result.items() if count)
        messagebox.showinfo("작업일보 반영 완료", f"작업일보에 총 {total_count}건을 반영했습니다.\n\n{detail}")
        self.set_status("excel", f"작업일보 반영 완료 / Excel 미반영 {pending_total}건", True)

    def open_condition_master_popup(self) -> None:
        password = simpledialog.askstring("조건 마스터 관리", "비밀번호를 입력하세요.", show="*", parent=self.root)
        if password is None:
            return
        if password != str(self.config.get("condition_master_password", CONDITION_MASTER_PASSWORD)):
            messagebox.showwarning("비밀번호 확인", "비밀번호가 맞지 않습니다.")
            return
        ConditionMasterPopup(self)

    def open_frequent_check_popup(self, mode: str, allowed_axes: list[int] | None = None) -> bool:
        if mode == "first":
            self.set_status("dnc", "초품 확인창 응답 대기중", None)
        elif mode == "jig":
            self.set_status("dnc", "하부 Pin 확인창 응답 대기중", None)
        elif mode == "capacity":
            self.set_status("dnc", "작업 축 수 확인창 응답 대기중", None)
        popup = FrequentCheckPopup(self, mode=mode, allowed_axes=allowed_axes)
        self.root.wait_window(popup.window)
        return popup.saved

    def has_frequent_check_completed(self, require_jig_check: bool = False) -> bool:
        return validate_frequent_check_values(self.frequent_check_values, check_mode="first")[0]

    def save_settings_from_ui_silent(self, show_error: bool = True) -> bool:
        if hasattr(self, "excel_var"):
            ok, message = validate_positive_number(self.delete_seconds_var.get(), "삭제 대기 시간", required=True)
            if not ok:
                if show_error:
                    messagebox.showwarning("설정 확인", message)
                return False
            ok, message = validate_positive_number(self.first_article_wait_var.get(), "초품 알람 시간", required=True)
            if not ok:
                if show_error:
                    messagebox.showwarning("설정 확인", message)
                return False
            excel_path = self.excel_var.get().strip()
            if excel_path:
                self.config["excel_file"] = excel_path
            elif self.config.get("excel_file", ""):
                self.excel_var.set(self.config["excel_file"])
            source_folders = dict(self.config.get("source_dnc_folders", {}))
            if hasattr(self, "source_vars"):
                for process_name, var in self.source_vars.items():
                    source_folders[process_name] = var.get().strip()
            self.config["source_dnc_folders"] = source_folders
            self.config["source_dnc_folder"] = source_folders.get("KCC PKG", self.source_var.get().strip())
            self.config["transfer_dnc_folder"] = self.transfer_var.get().strip()
            self.config["dnc_delete_seconds"] = int(self.delete_seconds_var.get().strip())
            self.config["first_article_wait_seconds"] = int(self.first_article_wait_var.get().strip())
            self.config["clear_common_after_normal"] = self.clear_common_var.get()
            if hasattr(self, "auto_shift_group_var"):
                self.config["auto_shift_group_enabled"] = bool(self.auto_shift_group_var.get())
            if hasattr(self, "a_group_day_start_var"):
                base_date = self.a_group_day_start_var.get().strip()
                try:
                    datetime.strptime(base_date, "%Y-%m-%d")
                    self.config["a_group_day_start_date"] = base_date
                except ValueError:
                    if show_error:
                        messagebox.showwarning("설정 확인", "A조 주간 시작 기준일은 YYYY-MM-DD 형식으로 입력해주세요.")
                    return False
            if "machine" in self.common_entries:
                self.config["machine"] = self.common_entries["machine"].get()
            save_config(self.config)
        return True

    def handle_run_error(self, exc: Exception) -> None:
        log_error("DNC 실행 오류", exc)
        messagebox.showerror("오류", str(exc))
        self.set_status("dnc", "오류", False)
        self.set_running(False)


class FrequentCheckPopup:
    """Q:AB에 기록할 초품/하부 Pin 확인 값을 클릭으로 입력하는 창입니다."""

    LABELS = [
        "1축",
        "2축",
        "3축",
        "4축",
        "5축",
        "6축",
        "1축",
        "2축",
        "3축",
        "4축",
        "5축",
        "6축",
    ]
    FIRST_OK_COLOR = "#0ea5e9"
    JIG_OK_COLOR = "#10b981"

    def __init__(self, app: JiinDncManager, mode: str, allowed_axes: list[int] | None = None):
        self.app = app
        self.mode = mode
        self.allowed_axes = set(allowed_axes) if allowed_axes is not None else None
        self.saved = False
        self.values = app.frequent_check_values[:]
        if self.mode == "jig":
            self.values[:6] = [""] * 6
        self.buttons: list[tk.Button] = []
        self.window = tk.Toplevel(app.root)
        self.window.title(self.get_title())
        self.window.geometry("560x300")
        self.window.configure(bg=APP_BG)
        self.window.resizable(False, False)
        self.create_ui()
        self.window.transient(app.root)
        self.window.grab_set()
        self.window.lift()
        self.window.focus_force()
        self.window.attributes("-topmost", True)
        self.center_on_parent()

    def center_on_parent(self) -> None:
        self.window.update_idletasks()
        parent_x = self.app.root.winfo_rootx()
        parent_y = self.app.root.winfo_rooty()
        parent_w = self.app.root.winfo_width()
        parent_h = self.app.root.winfo_height()
        width = self.window.winfo_width()
        height = self.window.winfo_height()
        x = parent_x + max((parent_w - width) // 2, 0)
        y = parent_y + max((parent_h - height) // 2, 0)
        self.window.geometry(f"{width}x{height}+{x}+{y}")

    def get_title(self) -> str:
        if self.mode == "capacity":
            return "작업 축 수 확인"
        return "초품 4Point 확인" if self.mode == "first" else "하부 Pin 3개 확인"

    def get_value_index(self, axis_index: int) -> int:
        return axis_index if self.mode in ("first", "capacity") else axis_index + 6

    def create_ui(self) -> None:
        title = tk.Label(
            self.window,
            text=self.get_title(),
            bg=PRIMARY_LIGHT,
            fg=PRIMARY,
            font=("맑은 고딕", 14, "bold"),
            height=2,
        )
        title.pack(fill=tk.X, padx=14, pady=(14, 8))

        body = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        body.pack(fill=tk.BOTH, expand=True, padx=14, pady=8)

        tk.Label(body, text=self.get_title(), bg="#d99a9a", fg="#111827", font=("맑은 고딕", 10, "bold"), height=2).grid(row=0, column=0, columnspan=6, sticky="ew")

        for axis_index, label in enumerate(self.LABELS[:6]):
            value_index = self.get_value_index(axis_index)
            disabled_axis = self.allowed_axes is not None and axis_index not in self.allowed_axes
            button = tk.Button(
                body,
                text=f"{label}\n{'미사용' if disabled_axis else ('OK' if self.values[value_index] == 'OK' else '클릭')}",
                command=lambda i=axis_index: self.toggle(i),
                bg="#f3f4f6" if disabled_axis else (self.get_ok_color() if self.values[value_index] == "OK" else SURFACE_BG),
                fg=MUTED_TEXT if disabled_axis else ("#ffffff" if self.values[value_index] == "OK" else TEXT_COLOR),
                relief=tk.SOLID,
                bd=1,
                width=8,
                height=3,
                cursor="arrow" if disabled_axis else "hand2",
                font=("맑은 고딕", 9, "bold"),
                highlightthickness=1,
                highlightbackground="#cbd5e1",
                state=tk.DISABLED if disabled_axis else tk.NORMAL,
            )
            button.grid(row=1, column=axis_index, sticky="nsew", padx=2, pady=8)
            body.columnconfigure(axis_index, weight=1)
            self.buttons.append(button)

        bottom = ttk.Frame(self.window, padding=(14, 4, 14, 14))
        bottom.pack(fill=tk.X)
        ttk.Button(bottom, text="초기화", command=self.clear).pack(side=tk.LEFT)
        ttk.Button(bottom, text="저장", command=self.save, style="Primary.TButton").pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(bottom, text="취소", command=self.window.destroy).pack(side=tk.RIGHT)

    def toggle(self, axis_index: int) -> None:
        if self.allowed_axes is not None and axis_index not in self.allowed_axes:
            return
        value_index = self.get_value_index(axis_index)
        self.values[value_index] = "" if self.values[value_index] == "OK" else "OK"
        self.refresh_button(axis_index)

    def refresh_button(self, axis_index: int) -> None:
        if self.allowed_axes is not None and axis_index not in self.allowed_axes:
            self.buttons[axis_index].configure(text=f"{self.LABELS[axis_index]}\n미사용", bg="#f3f4f6", fg=MUTED_TEXT)
            return
        value_index = self.get_value_index(axis_index)
        ok = self.values[value_index] == "OK"
        self.buttons[axis_index].configure(
            text=f"{self.LABELS[axis_index]}\n{'OK' if ok else '클릭'}",
            bg=self.get_ok_color() if ok else SURFACE_BG,
            fg="#ffffff" if ok else TEXT_COLOR,
        )

    def get_ok_color(self) -> str:
        return self.FIRST_OK_COLOR if self.mode in ("first", "capacity") else self.JIG_OK_COLOR

    def clear(self) -> None:
        target_range = range(0, 6) if self.mode in ("first", "capacity") else range(6, 12)
        for index in target_range:
            self.values[index] = ""
        for axis_index in range(6):
            self.refresh_button(axis_index)

    def save(self) -> None:
        check_mode = "first" if self.mode == "capacity" else self.mode
        ok, message = validate_frequent_check_values(self.values, check_mode=check_mode)
        if not ok:
            messagebox.showwarning(self.get_title(), message, parent=self.window)
            return
        self.app.frequent_check_values = self.values[:]
        if self.mode == "capacity":
            self.app.set_status("dnc", "작업 축 수 확인 완료", True)
        else:
            self.app.set_status("dnc", message, True)
        self.saved = True
        self.window.destroy()


class WorkHistoryPopup:
    """DB Browser 없이 프로그램 안에서 KCC PKG 작업 이력을 확인하는 보기 전용 창입니다."""

    def __init__(self, app: JiinDncManager):
        self.app = app
        self.window = tk.Toplevel(app.root)
        self.window.title("KCC PKG 작업 이력 보기")
        self.window.geometry("1320x660")
        self.window.configure(bg=APP_BG)
        self.keyword_var = tk.StringVar()
        self.limit_var = tk.StringVar(value="500")
        self.only_unexported_var = tk.BooleanVar(value=False)
        self.only_incomplete_var = tk.BooleanVar(value=False)
        self.summary_var = tk.StringVar(value="")
        self.create_ui()
        self.refresh()

    def create_ui(self) -> None:
        top = ttk.Frame(self.window, padding=(12, 12, 12, 8))
        top.pack(fill=tk.X)
        ttk.Label(top, text="검색", width=6).pack(side=tk.LEFT)
        ttk.Entry(top, textvariable=self.keyword_var, style="Wide.TEntry", width=28).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Checkbutton(top, text="Excel 미반영만", variable=self.only_unexported_var, command=self.refresh).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Checkbutton(top, text="미완료만", variable=self.only_incomplete_var, command=self.refresh).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(top, text="표시 수").pack(side=tk.LEFT, padx=(10, 4))
        ttk.Combobox(top, textvariable=self.limit_var, values=["100", "500", "1000", "3000", "5000"], state="readonly", width=8).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(top, text="조회", command=self.refresh, style="Primary.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(top, text="닫기", command=self.window.destroy).pack(side=tk.RIGHT)

        ttk.Label(self.window, textvariable=self.summary_var, background=APP_BG, foreground=MUTED_TEXT).pack(fill=tk.X, padx=14, pady=(0, 6))

        body = ttk.Frame(self.window, padding=(12, 0, 12, 12))
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        columns = (
            "id", "dnc_type", "status", "exported", "machine", "work_date", "shift_group", "shift_name",
            "worker", "step", "round_no", "manage_no", "lot_no", "qty_text", "result_value",
            "condition_name", "jig", "stack", "model_change_text", "burr_result", "record_time", "created_at",
        )
        self.tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="browse")
        headings = {
            "id": "ID",
            "dnc_type": "구분",
            "status": "상태",
            "exported": "Excel",
            "machine": "호기",
            "work_date": "작업일자",
            "shift_group": "조",
            "shift_name": "근무",
            "worker": "작업자",
            "step": "STEP",
            "round_no": "차수",
            "manage_no": "관리번호",
            "lot_no": "LOT",
            "qty_text": "매수",
            "result_value": "실적",
            "condition_name": "작업조건",
            "jig": "지그",
            "stack": "Stack",
            "model_change_text": "기종/검증",
            "burr_result": "Burr",
            "record_time": "작업시간",
            "created_at": "DB저장시간",
        }
        widths = {
            "id": 60,
            "dnc_type": 80,
            "status": 80,
            "exported": 80,
            "machine": 70,
            "work_date": 100,
            "shift_group": 50,
            "shift_name": 70,
            "worker": 90,
            "step": 70,
            "round_no": 70,
            "manage_no": 130,
            "lot_no": 150,
            "qty_text": 70,
            "result_value": 70,
            "condition_name": 240,
            "jig": 80,
            "stack": 70,
            "model_change_text": 90,
            "burr_result": 90,
            "record_time": 90,
            "created_at": 150,
        }
        for column in columns:
            self.tree.heading(column, text=headings[column])
            self.tree.column(column, width=widths[column], anchor="center" if column not in ("condition_name", "lot_no", "manage_no") else "w")
        self.tree.grid(row=0, column=0, sticky="nsew")
        y_scroll = ttk.Scrollbar(body, orient=tk.VERTICAL, command=self.tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(body, orient=tk.HORIZONTAL, command=self.tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

    def refresh(self) -> None:
        try:
            limit = int(self.limit_var.get())
        except ValueError:
            limit = 500
            self.limit_var.set("500")
        rows = load_work_history(
            limit=limit,
            only_unexported=self.only_unexported_var.get(),
            only_incomplete=self.only_incomplete_var.get(),
            keyword=self.keyword_var.get(),
        )
        self.tree.delete(*self.tree.get_children())
        for row in rows:
            exported_text = "반영" if int(row["exported"] or 0) == 1 else "미반영"
            values = (
                row["id"],
                row["dnc_type"] or "",
                row["status"] or "",
                exported_text,
                row["machine"] or "",
                row["work_date"] or "",
                row["shift_group"] or "",
                row["shift_name"] or "",
                row["worker"] or "",
                row["step"] or "",
                row["round_no"] or "",
                row["manage_no"] or "",
                row["lot_no"] or "",
                row["qty_text"] or "",
                "" if row["result_value"] is None else row["result_value"],
                row["condition_name"] or "",
                row["jig"] or "",
                row["stack"] or "",
                row["model_change_text"] or "",
                row["burr_result"] or "",
                row["record_time"] or "",
                row["created_at"] or "",
            )
            self.tree.insert("", "end", values=values)
        pending_count = get_unexported_kcc_pkg_count()
        incomplete_count = get_incomplete_kcc_pkg_count()
        self.summary_var.set(
            f"최근 {len(rows)}건 표시 / Excel 미반영 {pending_count}건 / 미완료 {incomplete_count}건"
        )


class MasterSettingsPopup:
    """작업자가 실수로 만지면 위험한 점검/관리 기능을 모아둔 관리자 전용 창입니다."""

    def __init__(self, app: JiinDncManager):
        self.app = app
        self.window = tk.Toplevel(app.root)
        self.window.title("마스터 설정")
        self.window.geometry("760x430")
        self.window.minsize(760, 430)
        self.window.configure(bg=APP_BG)
        self.delete_seconds_var = tk.StringVar(value=str(app.config.get("dnc_delete_seconds", DNC_DELETE_SECONDS)))
        self.first_article_wait_var = tk.StringVar(value=str(app.config.get("first_article_wait_seconds", FIRST_ARTICLE_WAIT_SECONDS)))
        self.create_ui()
        self.window.protocol("WM_DELETE_WINDOW", self.close)

    def create_ui(self) -> None:
        panel = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        panel.pack(fill=tk.BOTH, expand=True, padx=14, pady=14)
        panel.columnconfigure(1, weight=1)
        tk.Label(
            panel,
            text="마스터 설정",
            bg=PRIMARY_LIGHT,
            fg=PRIMARY,
            font=("맑은 고딕", 14, "bold"),
            height=2,
        ).grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 12))

        ttk.Label(panel, text="삭제 대기 시간", background=SURFACE_BG, width=16).grid(row=1, column=0, sticky="e", padx=10, pady=8)
        ttk.Entry(panel, textvariable=self.delete_seconds_var, style="Wide.TEntry").grid(row=1, column=1, sticky="ew", padx=8, pady=8)
        ttk.Label(panel, text="초", background=SURFACE_BG).grid(row=1, column=2, sticky="w", padx=(0, 10), pady=8)

        ttk.Label(panel, text="초품 알람 시간", background=SURFACE_BG, width=16).grid(row=2, column=0, sticky="e", padx=10, pady=8)
        ttk.Entry(panel, textvariable=self.first_article_wait_var, style="Wide.TEntry").grid(row=2, column=1, sticky="ew", padx=8, pady=8)
        ttk.Label(panel, text="초", background=SURFACE_BG).grid(row=2, column=2, sticky="w", padx=(0, 10), pady=8)

        buttons = ttk.Frame(panel)
        buttons.grid(row=3, column=0, columnspan=3, sticky="ew", padx=10, pady=(18, 8))
        buttons.columnconfigure((0, 1), weight=1)
        ttk.Button(buttons, text="작업 이력 보기", command=self.app.open_work_history_popup).grid(row=0, column=0, sticky="ew", padx=4, pady=4)
        ttk.Button(buttons, text="마스터 복구", command=self.rebuild_condition_master).grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        ttk.Button(buttons, text="마스터 설정 비번 변경", command=lambda: self.change_password("master_password", "마스터 설정", MASTER_SETTINGS_PASSWORD)).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
        ttk.Button(buttons, text="조건 마스터 비번 변경", command=lambda: self.change_password("condition_master_password", "조건 마스터 관리", CONDITION_MASTER_PASSWORD)).grid(row=1, column=1, sticky="ew", padx=4, pady=4)

        bottom = ttk.Frame(panel)
        bottom.grid(row=4, column=0, columnspan=3, sticky="ew", padx=10, pady=(22, 10))
        bottom.columnconfigure(0, weight=1)
        ttk.Button(bottom, text="닫기", command=self.close, width=18).grid(row=0, column=1, sticky="e")

    def save_master_settings(self, show_message: bool = False) -> bool:
        ok, message = validate_positive_number(self.delete_seconds_var.get(), "삭제 대기 시간", required=True)
        if not ok:
            messagebox.showwarning("마스터 설정", message, parent=self.window)
            return False
        ok, message = validate_positive_number(self.first_article_wait_var.get(), "초품 알람 시간", required=True)
        if not ok:
            messagebox.showwarning("마스터 설정", message, parent=self.window)
            return False
        self.app.config["dnc_delete_seconds"] = int(self.delete_seconds_var.get().strip())
        self.app.config["first_article_wait_seconds"] = int(self.first_article_wait_var.get().strip())
        if hasattr(self.app, "delete_seconds_var"):
            self.app.delete_seconds_var.set(self.delete_seconds_var.get().strip())
        if hasattr(self.app, "first_article_wait_var"):
            self.app.first_article_wait_var.set(self.first_article_wait_var.get().strip())
        save_config(self.app.config)
        if show_message:
            messagebox.showinfo("적용 완료", "마스터 설정을 적용했습니다.", parent=self.window)
        return True

    def close(self) -> None:
        if self.save_master_settings(show_message=False):
            self.window.destroy()

    def rebuild_condition_master(self) -> None:
        if not self.save_master_settings(show_message=False):
            return
        try:
            count = rebuild_condition_master_from_log(self.app.config)
        except Exception as exc:
            messagebox.showerror("조건 복구 실패", str(exc), parent=self.window)
            return
        messagebox.showinfo("조건 복구 완료", f"{count}개 조건을 저장했습니다.", parent=self.window)

    def change_password(self, config_key: str, title: str, default_password: str) -> None:
        current_password = str(self.app.config.get(config_key, default_password))
        old_password = simpledialog.askstring(f"{title} 비밀번호 변경", "현재 비밀번호를 입력하세요.", show="*", parent=self.window)
        if old_password is None:
            return
        if old_password != current_password:
            messagebox.showwarning("비밀번호 확인", "현재 비밀번호가 맞지 않습니다.", parent=self.window)
            return
        new_password = simpledialog.askstring(f"{title} 비밀번호 변경", "새 비밀번호를 입력하세요.", show="*", parent=self.window)
        if new_password is None:
            return
        new_password = new_password.strip()
        if not new_password:
            messagebox.showwarning("비밀번호 확인", "새 비밀번호를 입력하세요.", parent=self.window)
            return
        confirm_password = simpledialog.askstring(f"{title} 비밀번호 변경", "새 비밀번호를 한 번 더 입력하세요.", show="*", parent=self.window)
        if confirm_password is None:
            return
        if new_password != confirm_password.strip():
            messagebox.showwarning("비밀번호 확인", "새 비밀번호가 서로 다릅니다.", parent=self.window)
            return
        self.app.config[config_key] = new_password
        save_config(self.app.config)
        messagebox.showinfo("변경 완료", f"{title} 비밀번호를 변경했습니다.", parent=self.window)


class ConditionMasterPopup:
    """작업일보에서 추출한 조건 마스터를 보고 수정하는 창입니다."""

    def __init__(self, app: JiinDncManager):
        self.app = app
        self.records = load_condition_master()
        self.search_var = tk.StringVar()
        self.window = tk.Toplevel(app.root)
        self.window.title("KCC PKG 조건 마스터 관리")
        self.window.geometry("980x620")
        self.window.configure(bg=APP_BG)
        self.create_ui()
        self.refresh_tree()

    def create_ui(self) -> None:
        top = ttk.Frame(self.window, padding=(12, 12, 12, 8))
        top.pack(fill=tk.X)
        ttk.Button(top, text="작업일보 이력으로 갱신", command=self.rebuild_from_log, style="Primary.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(top, text="선택 수정 저장", command=self.save_selected_edit).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(top, text="선택 삭제", command=self.delete_selected_record).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(top, text="닫기", command=self.window.destroy).pack(side=tk.RIGHT)

        search = ttk.Frame(self.window, padding=(12, 0, 12, 8))
        search.pack(fill=tk.X)
        ttk.Label(search, text="조회").pack(side=tk.LEFT, padx=(0, 6))
        search_entry = ttk.Entry(search, textvariable=self.search_var, style="Wide.TEntry", width=42)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(search, text="조회", command=self.refresh_tree).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(search, text="전체 보기", command=self.clear_search).pack(side=tk.LEFT)
        self.search_var.trace_add("write", lambda *_args: self.refresh_tree())

        body = ttk.Frame(self.window, padding=(12, 0, 12, 8))
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        columns = ("step", "round", "manage_no", "process_code", "condition", "jig", "source")
        self.tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="browse")
        headings = {
            "step": "STEP",
            "round": "차수",
            "manage_no": "관리번호",
            "process_code": "공정코드",
            "condition": "작업조건",
            "jig": "지그",
            "source": "출처",
        }
        widths = {
            "step": 80,
            "round": 70,
            "manage_no": 150,
            "process_code": 110,
            "condition": 220,
            "jig": 140,
            "source": 140,
        }
        for column in columns:
            self.tree.heading(column, text=headings[column])
            self.tree.column(column, width=widths[column], anchor="w")
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        scroll = ttk.Scrollbar(body, orient=tk.VERTICAL, command=self.tree.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scroll.set)

        edit = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        edit.pack(fill=tk.X, padx=12, pady=(0, 12))
        self.edit_vars = {
            "step": tk.StringVar(),
            "round": tk.StringVar(),
            "manage_no": tk.StringVar(),
            "process_code": tk.StringVar(),
            "condition": tk.StringVar(),
            "jig": tk.StringVar(),
        }
        edit_fields = [
            ("step", "STEP"),
            ("round", "차수"),
            ("manage_no", "관리번호"),
            ("process_code", "공정코드"),
            ("condition", "작업조건"),
            ("jig", "지그"),
        ]
        for index, (key, label) in enumerate(edit_fields):
            ttk.Label(edit, text=label, background=SURFACE_BG).grid(row=index // 3, column=(index % 3) * 2, sticky="e", padx=(10, 4), pady=8)
            ttk.Entry(edit, textvariable=self.edit_vars[key], style="Wide.TEntry", width=24).grid(row=index // 3, column=(index % 3) * 2 + 1, sticky="ew", padx=(0, 10), pady=8)
            edit.columnconfigure((index % 3) * 2 + 1, weight=1)

    def refresh_tree(self) -> None:
        self.tree.delete(*self.tree.get_children())
        keyword = self.search_var.get().strip().lower()
        for index, record in enumerate(self.records):
            if keyword and not self.record_matches_keyword(record, keyword):
                continue
            self.tree.insert(
                "",
                "end",
                iid=str(index),
                values=(
                    record.get("step", ""),
                    record.get("round", ""),
                    record.get("manage_no", ""),
                    record.get("process_code", ""),
                    record.get("condition", ""),
                    record.get("jig", ""),
                    record.get("source", ""),
                ),
            )

    def clear_search(self) -> None:
        self.search_var.set("")
        self.refresh_tree()

    def record_matches_keyword(self, record: dict, keyword: str) -> bool:
        searchable = [
            record.get("step", ""),
            record.get("round", ""),
            record.get("manage_no", ""),
            record.get("process_code", ""),
            record.get("condition", ""),
            record.get("jig", ""),
            record.get("source", ""),
        ]
        return any(keyword in str(value).lower() for value in searchable)

    def on_select(self, _event=None) -> None:
        selected = self.tree.selection()
        if not selected:
            return
        record = self.records[int(selected[0])]
        for key, var in self.edit_vars.items():
            var.set(record.get(key, ""))

    def save_selected_edit(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("선택 필요", "수정할 조건을 먼저 선택하세요.", parent=self.window)
            return
        index = int(selected[0])
        record = self.records[index]
        for key, var in self.edit_vars.items():
            record[key] = var.get().strip()
        record["source"] = "사용자 수정"
        record["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_condition_master(self.records)
        self.refresh_tree()
        self.tree.selection_set(str(index))
        messagebox.showinfo("저장 완료", "조건 마스터를 수정했습니다.", parent=self.window)

    def delete_selected_record(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("선택 필요", "삭제할 조건을 먼저 선택하세요.", parent=self.window)
            return
        index = int(selected[0])
        record = self.records[index]
        ok = messagebox.askyesno(
            "조건 삭제 확인",
            "선택한 조건을 조건 마스터에서 삭제하시겠습니까?\n\n"
            f"STEP: {record.get('step', '')}\n"
            f"차수: {record.get('round', '')}\n"
            f"관리번호: {record.get('manage_no', '')}\n"
            f"공정코드: {record.get('process_code', '')}\n"
            f"작업조건: {record.get('condition', '')}\n"
            f"지그: {record.get('jig', '')}",
            parent=self.window,
        )
        if not ok:
            return
        del self.records[index]
        save_condition_master(self.records)
        self.records = load_condition_master()
        self.refresh_tree()
        messagebox.showinfo("삭제 완료", "선택한 조건을 삭제했습니다.", parent=self.window)

    def rebuild_from_log(self) -> None:
        self.app.save_settings_from_ui_silent()
        try:
            count = rebuild_condition_master_from_log(self.app.config)
        except Exception as exc:
            messagebox.showerror("갱신 실패", str(exc), parent=self.window)
            return
        self.records = load_condition_master()
        self.refresh_tree()
        messagebox.showinfo(
            "갱신 완료",
            f"조건 마스터를 최신 이력으로 정리했습니다.\n\n"
            f"현재 보관 조건: {count}개\n"
            "작업일보에서 삭제된 모델도 마스터에서는 보존됩니다.",
            parent=self.window,
        )


class NewModelPopup:
    def __init__(self, app: JiinDncManager):
        self.app = app
        self.window = tk.Toplevel(app.root)
        self.window.title("KCC PKG 신규 모델 검증 DNC")
        self.window.geometry("1280x760")
        self.window.minsize(1180, 730)
        self.window.configure(bg=APP_BG)
        self.entries: dict[str, LabeledEntry] = {}
        self.both_entries: dict[str, dict[str, LabeledEntry]] = {}
        self.buttons: list[ttk.Button] = []
        self.mode_buttons: dict[str, tk.Button] = {}
        self.is_loading_fields = False
        self.selected_lot = tk.StringVar(value="lot1")
        self.run_mode = tk.StringVar(value="lot1")
        empty_lot = {
            key: ""
            for key in ["step", "round", "manage_no", "lot_no", "qty", "process_code", "condition", "jig"]
        }
        self.lot_drafts = {
            "lot1": dict(empty_lot),
            "lot2": dict(empty_lot),
        }
        self.target_lots = self.get_new_model_target_lots()
        if not self.target_lots:
            messagebox.showinfo(
                "신규 모델 검증 DNC",
                "조건 적용이 필요한 LOT가 없습니다.\n\n조건/지그가 없는 LOT가 있을 때 신규 모델 검증 DNC를 실행해주세요.",
                parent=app.root,
            )
            self.window.destroy()
            return
        self.selected_lot.set(self.target_lots[0])
        self.run_mode.set(self.target_lots[0])
        self.is_running = False
        self.create_ui()
        self.load_lot_drafts_from_main()
        self.load_selected_lot()
        self.load_both_lot_panels()
        self.refresh_input_mode()
        self.update_checks()

    def get_new_model_target_lots(self) -> list[str]:
        """메인 화면에서 조건/지그가 없는 LOT만 신규 검증 대상으로 반환합니다."""
        targets: list[str] = []
        used_lot_count = 0
        for lot_key, entries in (("lot1", self.app.lot1_entries), ("lot2", self.app.lot2_entries)):
            lot = self.app.get_lot_data(entries)
            if not self.app.lot_has_any_value(lot):
                continue
            used_lot_count += 1
            condition_ok, _message = get_single_condition_message(lot)
            if not condition_ok:
                targets.append(lot_key)
        if targets:
            return targets
        if used_lot_count == 0:
            return ["lot1", "lot2"]
        return []

    def create_ui(self) -> None:
        title = tk.Label(self.window, text="KCC PKG 신규 모델 검증 DNC", bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 14, "bold"), height=2)
        title.pack(fill=tk.X, padx=14, pady=(14, 8))

        lot_select = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        lot_select.pack(fill=tk.X, padx=14, pady=(0, 8))
        if len(self.target_lots) == 1:
            lot_name = "LOT 1" if self.target_lots[0] == "lot1" else "LOT 2"
            tk.Label(
                lot_select,
                text=f"{lot_name} 입력",
                bg=SURFACE_BG,
                fg=PRIMARY,
                font=("맑은 고딕", 10, "bold"),
            ).pack(side=tk.LEFT, padx=14, pady=8)
        else:
            choices = [("lot1", "LOT 1 입력"), ("lot2", "LOT 2 입력"), ("both", "LOT 1 + LOT 2")]
            for value, label in choices:
                self.create_mode_button(lot_select, label, value).pack(side=tk.LEFT, padx=4, pady=8)

        self.input_area = tk.Frame(self.window, bg=APP_BG)
        self.input_area.pack(fill=tk.BOTH, expand=True, padx=14, pady=8)

        self.single_panel = tk.Frame(self.input_area, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        self.entries = self.create_lot_input_fields(self.single_panel, columns_per_row=2)

        self.both_panel = tk.Frame(self.input_area, bg=APP_BG)
        self.both_panel.columnconfigure(0, weight=1)
        self.both_panel.columnconfigure(1, weight=1)
        for column, (lot_key, title_text) in enumerate((("lot1", "LOT 1 신규 입력"), ("lot2", "LOT 2 신규 입력"))):
            lot_panel = tk.Frame(self.both_panel, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
            lot_panel.grid(row=0, column=column, sticky="nsew", padx=(0, 8) if column == 0 else (8, 0))
            tk.Label(lot_panel, text=title_text, bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 11, "bold"), height=2).pack(fill=tk.X)
            fields_frame = tk.Frame(lot_panel, bg=SURFACE_BG)
            fields_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            self.both_entries[lot_key] = self.create_lot_input_fields(fields_frame, columns_per_row=1)

        status = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        status.pack(fill=tk.X, padx=14, pady=8)
        self.mes_label = tk.Label(status, text="MES Core 일치화: 대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 10, "bold"))
        self.mes_label.pack(side=tk.LEFT, padx=16, pady=10)
        self.condition_label = tk.Label(status, text="조건 적용 확인: 대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 10, "bold"))
        self.condition_label.pack(side=tk.LEFT, padx=16, pady=10)
        self.dnc_label = tk.Label(status, text="DNC 진행 상태: 대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 10, "bold"))
        self.dnc_label.pack(side=tk.LEFT, padx=16, pady=10)
        self.excel_label = tk.Label(status, text="작업일보 반영: 대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 10, "bold"))
        self.excel_label.pack(side=tk.LEFT, padx=16, pady=10)

        buttons = ttk.Frame(self.window)
        buttons.pack(fill=tk.X, padx=14, pady=(4, 14))
        self.add_button(buttons, "신규 모델 DNC 실행", self.run_new_model_dnc, "Primary.TButton").pack(side=tk.LEFT, padx=4)
        self.add_button(buttons, "입력 초기화", self.clear_inputs).pack(side=tk.LEFT, padx=4)
        self.add_button(buttons, "닫기", self.window.destroy).pack(side=tk.RIGHT, padx=4)

    def add_button(self, parent, text, command, style="TButton") -> ttk.Button:
        button = ttk.Button(parent, text=text, command=command, style=style, width=20)
        self.buttons.append(button)
        return button

    def create_mode_button(self, parent, text: str, value: str) -> tk.Button:
        button = tk.Button(
            parent,
            text=text,
            command=lambda selected=value: self.set_run_mode(selected),
            relief=tk.FLAT,
            bd=0,
            padx=16,
            pady=6,
            cursor="hand2",
            font=("맑은 고딕", 10, "bold"),
        )
        self.mode_buttons[value] = button
        self.update_mode_buttons()
        return button

    def update_mode_buttons(self) -> None:
        for value, button in self.mode_buttons.items():
            selected = value == self.run_mode.get()
            button.configure(
                bg=PRIMARY_LIGHT if selected else SURFACE_BG,
                fg=PRIMARY if selected else TEXT_COLOR,
                highlightthickness=1,
                highlightbackground=PRIMARY if selected else BORDER_COLOR,
            )

    def create_lot_input_fields(self, parent, columns_per_row: int = 2) -> dict[str, LabeledEntry]:
        fields = [
            ("step", "STEP"),
            ("round", "차수"),
            ("manage_no", "관리번호"),
            ("lot_no", "LOT No"),
            ("qty", "매수"),
            ("process_code", "공정코드"),
            ("condition", "작업조건"),
            ("jig", "지그"),
        ]
        entries: dict[str, LabeledEntry] = {}
        for index, (key, label) in enumerate(fields):
            if key == "round":
                entry = RoundField(parent, label)
            elif key in {"step", "qty"}:
                entry = LabeledEntry(parent, label, width=24, numeric_only=True)
            elif key == "condition":
                entry = LabeledEntry(parent, label, width=24)
            else:
                entry = LabeledEntry(parent, label, width=24, uppercase=True)
            row = index // columns_per_row
            column = index % columns_per_row
            entry.grid(row=row, column=column, sticky="ew", padx=10, pady=8)
            parent.columnconfigure(column, weight=1)
            entries[key] = entry
            if hasattr(entry, "var"):
                entry.var.trace_add("write", lambda *_args: self.update_checks())
        return entries

    def read_entry_group(self, entries: dict[str, LabeledEntry]) -> dict:
        return {
            key: entries[key].get()
            for key in ["step", "round", "manage_no", "lot_no", "qty", "process_code", "condition", "jig"]
        }

    def fill_entry_group(self, entries: dict[str, LabeledEntry], draft: dict) -> None:
        self.is_loading_fields = True
        try:
            for key, entry in entries.items():
                entry.set(draft.get(key, ""))
        finally:
            self.is_loading_fields = False

    def save_current_lot_draft(self) -> None:
        if not self.entries or self.is_loading_fields:
            return
        if self.run_mode.get() == "both":
            for lot_key, entries in self.both_entries.items():
                self.lot_drafts[lot_key] = self.read_entry_group(entries)
            return
        self.lot_drafts[self.selected_lot.get()] = self.read_entry_group(self.entries)

    def load_selected_lot(self) -> None:
        draft = self.lot_drafts.get(self.selected_lot.get(), {})
        self.fill_entry_group(self.entries, draft)

    def load_both_lot_panels(self) -> None:
        for lot_key, entries in self.both_entries.items():
            self.fill_entry_group(entries, self.lot_drafts.get(lot_key, {}))

    def load_lot_drafts_from_main(self) -> None:
        """메인 일반 DNC 화면에 입력된 LOT 값을 신규 검증 팝업 초기값으로 가져옵니다.

        신규 검증에서는 매수/작업조건/지그를 다시 판단해야 하므로 비워둡니다.
        """
        for lot_key, entries in (("lot1", self.app.lot1_entries), ("lot2", self.app.lot2_entries)):
            lot = self.app.get_lot_data(entries)
            if not self.app.lot_has_any_value(lot):
                continue
            draft = dict(self.lot_drafts[lot_key])
            for key in ("step", "round", "manage_no", "lot_no", "process_code"):
                draft[key] = lot.get(key, "")
            draft["qty"] = ""
            draft["condition"] = ""
            draft["jig"] = ""
            self.lot_drafts[lot_key] = draft
        if not self.app.lot_has_any_value(self.app.get_lot_data(self.app.lot1_entries)):
            if self.app.lot_has_any_value(self.app.get_lot_data(self.app.lot2_entries)):
                self.selected_lot.set("lot2")

    def set_run_mode(self, mode: str) -> None:
        self.save_current_lot_draft()
        self.run_mode.set(mode)
        next_lot = "lot1" if mode == "both" else mode
        self.selected_lot.set(next_lot)
        self.load_selected_lot()
        self.load_both_lot_panels()
        self.refresh_input_mode()

    def refresh_input_mode(self) -> None:
        self.update_mode_buttons()
        if self.run_mode.get() == "both":
            self.single_panel.pack_forget()
            self.both_panel.pack(fill=tk.BOTH, expand=True)
        else:
            self.both_panel.pack_forget()
            self.single_panel.pack(fill=tk.BOTH, expand=True)
        self.update_checks()

    def get_data(self) -> tuple[dict, dict]:
        self.save_current_lot_draft()
        common = self.app.get_common_data()
        lot = dict(self.lot_drafts[self.selected_lot.get()])
        return common, lot

    def get_run_lots(self) -> tuple[dict, list[tuple[str, dict]]]:
        """신규 검증 실행 시 저장해야 할 LOT 목록을 반환합니다.

        둘 다 신규일 때만 작업자가 LOT 1만/LOT 2만/LOT 1+LOT 2 중 선택합니다.
        """
        self.save_current_lot_draft()
        common = self.app.get_common_data()
        mode = self.run_mode.get()
        run_keys = ["lot1", "lot2"] if mode == "both" else [mode]
        lots = [(key, dict(self.lot_drafts[key])) for key in run_keys]
        return common, lots

    def update_checks(self) -> None:
        if self.is_loading_fields:
            return
        _common, lot_items = self.get_run_lots()
        lots = [lot for _lot_key, lot in lot_items]

        if all(not lot.get("lot_no", "").strip() and not lot.get("process_code", "").strip() for lot in lots):
            self.mes_label.configure(text="MES Core 일치화: 대기중", fg=MUTED_TEXT)
        else:
            mes_results = [
                get_mes_core_message(lot.get("lot_no", ""), lot.get("process_code", ""))[0]
                for lot in lots
            ]
            if all(mes_results):
                self.mes_label.configure(text="MES Core 일치화: OK", fg=OK_COLOR)
            else:
                self.mes_label.configure(text="MES Core 일치화: NG", fg=NG_COLOR)

        if all(not lot.get("condition", "").strip() and not lot.get("jig", "").strip() for lot in lots):
            self.condition_label.configure(text="조건 적용 확인: 대기중", fg=MUTED_TEXT)
        else:
            condition_results = [get_single_condition_message(lot)[0] for lot in lots]
            self.condition_label.configure(
                text=f"조건 적용 확인: {'OK' if all(condition_results) else 'NG'}",
                fg=OK_COLOR if all(condition_results) else NG_COLOR,
            )

    def set_running(self, running: bool) -> None:
        self.is_running = running
        for button in self.buttons:
            button.configure(state="disabled" if running else "normal")

    def set_dnc_status(self, text: str) -> None:
        self.window.after(0, lambda: self.dnc_label.configure(text=f"DNC 진행 상태: {text}", fg=MUTED_TEXT))

    def run_new_model_dnc(self) -> None:
        run_new_model_dnc(self)

    def clear_inputs(self) -> None:
        run_keys = ["lot1", "lot2"] if self.run_mode.get() == "both" else [self.selected_lot.get()]
        for lot_key in run_keys:
            self.lot_drafts[lot_key] = {
                key: ""
                for key in ["step", "round", "manage_no", "lot_no", "qty", "process_code", "condition", "jig"]
            }
        self.load_selected_lot()
        self.load_both_lot_panels()
        self.update_checks()
        self.dnc_label.configure(text="DNC 진행 상태: 대기중", fg=MUTED_TEXT)

    def clear_after_done(self) -> None:
        _common, lot_items = self.get_run_lots()
        for lot_key, draft in lot_items:
            for key in ["round", "qty", "condition", "jig"]:
                draft[key] = ""
            self.lot_drafts[lot_key] = draft
        self.load_selected_lot()
        self.load_both_lot_panels()
        self.update_checks()


def open_new_model_popup(app: JiinDncManager) -> None:
    """KCC PKG 신규 모델 검증 DNC 팝업창을 엽니다."""
    NewModelPopup(app)


def run_new_model_dnc(popup: NewModelPopup) -> None:
    """신규 모델 DNC 실행 버튼 흐름입니다."""
    if popup.is_running or popup.app.is_running:
        messagebox.showwarning("진행 중", "DNC 실행중입니다.\n작업 완료 후 다시 실행해주세요.")
        return
    if not popup.app.save_settings_from_ui_silent():
        return
    common, lot_items = popup.get_run_lots()
    lots = [lot for _lot_key, lot in lot_items]
    all_errors = []
    for lot_key, lot in lot_items:
        lot_label = "LOT 1" if lot_key == "lot1" else "LOT 2"
        ok, errors = validate_new_model_dnc(common, lot)
        if not ok:
            all_errors.extend([f"{lot_label} - {error}" for error in errors])
    condition_names = {lot.get("condition", "").strip() for lot in lots if lot.get("condition", "").strip()}
    jig_names = {lot.get("jig", "").strip() for lot in lots if lot.get("jig", "").strip()}
    if len(lots) > 1 and len(condition_names) > 1:
        all_errors.append("LOT 1 / LOT 2 작업조건이 서로 다릅니다. 신규 DNC는 같은 작업조건일 때만 같이 적용할 수 있습니다.")
    if len(lots) > 1 and len(jig_names) > 1:
        all_errors.append("LOT 1 / LOT 2 지그가 서로 다릅니다. 신규 DNC는 같은 지그일 때만 같이 적용할 수 있습니다.")
    if all_errors:
        messagebox.showwarning("입력값 확인", "\n".join(all_errors), parent=popup.window)
        popup.dnc_label.configure(text="DNC 진행 상태: 입력값 NG", fg=NG_COLOR)
        return
    leader_name = simpledialog.askstring("조장명 입력", "신규 모델 검증 조장님 성함을 기재하세요", parent=popup.window)
    if not leader_name or not leader_name.strip():
        popup.dnc_label.configure(text="DNC 진행 상태: 취소", fg=MUTED_TEXT)
        return
    condition_file = popup.app.validate_condition_file(lots[0]["condition"])
    if not condition_file:
        popup.dnc_label.configure(text="DNC 진행 상태: 조건 파일 NG", fg=NG_COLOR)
        return
    log_app(f"신규 모델 DNC 시작: {len(lots)} LOT, 조건={lots[0]['condition']}")
    popup.set_running(True)
    popup.app.set_running(True)
    threading.Thread(target=new_model_worker, args=(popup, common, lots, leader_name.strip(), condition_file), daemon=True).start()


def run_normal_dnc(app: JiinDncManager) -> None:
    """요청 함수명 보존용 래퍼입니다. 실제 일반 DNC 실행은 JiinDncManager.run_normal_dnc에서 처리합니다."""
    app.run_normal_dnc()


def new_model_worker(popup: NewModelPopup, common: dict, lots: list[dict], leader_name: str, condition_file: Path) -> None:
    """신규 모델 DNC 백그라운드 작업입니다."""
    try:
        popup.set_dnc_status("DB 저장중")
        log_ids = [insert_new_model_db(common, lot, leader_name) for lot in lots]
        popup.set_dnc_status("DB 저장 완료")
        delete_existing_dnc_txt(Path(popup.app.config["transfer_dnc_folder"]))
        copied_file = copy_dnc_file(condition_file, Path(popup.app.config["transfer_dnc_folder"]))
        popup.set_dnc_status("DNC 파일 복사 완료")
        delete_after_delay(copied_file, int(popup.app.config["dnc_delete_seconds"]), popup.set_dnc_status)
        popup.window.after(0, lambda: finish_new_model_dnc(popup, log_ids, lots))
    except Exception as exc:
        popup.window.after(0, lambda error=exc: handle_popup_error(popup, error))


def finish_new_model_dnc(popup: NewModelPopup, log_ids: list[int], lots: list[dict]) -> None:
    """신규 모델 DNC 완료 후 초도품 확인 결과를 저장합니다."""
    try:
        first_article_ok = messagebox.askyesno("초도품 확인", "초도품 이상 없습니까?", parent=popup.window)
        for log_id, lot in zip(log_ids, lots):
            condition_name = lot["condition"]
            update_new_model_db(log_id, condition_name, first_article_ok)
            if first_article_ok:
                upsert_condition_master(lot, condition_name, lot["jig"], "신규 검증 완료")
        pending_count = get_unexported_kcc_pkg_count()
        popup.dnc_label.configure(text="DNC 진행 상태: DNC 완료", fg=OK_COLOR)
        popup.excel_label.configure(text=f"작업일보 반영: Excel 미반영 {pending_count}건", fg=OK_COLOR)
        log_app(f"신규 모델 DNC 완료: ids={log_ids}, Excel 미반영={pending_count}건")
        if popup.app.auto_export_kcc_pkg_to_excel(parent=popup.window):
            popup.excel_label.configure(text="작업일보 반영: 자동 반영 완료", fg=OK_COLOR)
        else:
            popup.excel_label.configure(text=f"작업일보 반영: Excel 미반영 {get_unexported_kcc_pkg_count()}건", fg=NG_COLOR)
        popup.clear_after_done()
    except Exception as exc:
        handle_popup_error(popup, exc)
    finally:
        popup.set_running(False)
        popup.app.set_running(False)


def handle_popup_error(popup: NewModelPopup, exc: Exception) -> None:
    log_error("신규 모델 DNC 오류", exc)
    messagebox.showerror("오류", str(exc), parent=popup.window)
    popup.dnc_label.configure(text="DNC 진행 상태: 오류", fg=NG_COLOR)
    popup.set_running(False)
    popup.app.set_running(False)


def main() -> None:
    if not acquire_single_instance_lock():
        return
    root = tk.Tk()
    app = JiinDncManager(root)
    root.mainloop()


if __name__ == "__main__":
    main()
