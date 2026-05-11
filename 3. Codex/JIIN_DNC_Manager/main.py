import json
import calendar
import os
import shutil
import subprocess
import sys
import threading
import time
import tkinter as tk
import zipfile
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, simpledialog, ttk

from openpyxl import load_workbook
from openpyxl.styles import Font


# ==================================================
# 기본 설정
# ==================================================
# 현장 적용 시 기본 삭제 시간을 60초로 바꾸고 싶으면 아래 값만 수정해도 됩니다.
DNC_DELETE_SECONDS = 10

APP_TITLE = "JIIN DNC Manager"
LOG_SHEET_NAME = "KCC PKG"


def get_app_dir() -> Path:
    """EXE 실행 시에는 EXE가 있는 폴더, Python 실행 시에는 main.py 폴더를 반환합니다."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


APP_DIR = get_app_dir()
CONFIG_FILE = APP_DIR / "config.json"
CONDITION_MASTER_FILE = APP_DIR / "condition_master.json"

APP_BG = "#f3f6fb"
SURFACE_BG = "#ffffff"
BORDER_COLOR = "#d8e2f0"
TEXT_COLOR = "#172033"
MUTED_TEXT = "#687386"
PRIMARY = "#0f5bff"
PRIMARY_LIGHT = "#eaf2ff"
OK_COLOR = "#0f9f63"
NG_COLOR = "#dc2626"
TAB_BG = "#eaf0f8"

THEMES = {
    "MES 블루": {
        "app_bg": "#f3f6fb",
        "surface_bg": "#ffffff",
        "border_color": "#d8e2f0",
        "text_color": "#172033",
        "muted_text": "#687386",
        "primary": "#0f5bff",
        "primary_light": "#eaf2ff",
        "tab_bg": "#eaf0f8",
    },
    "KCC 민트": {
        "app_bg": "#eaf5f2",
        "surface_bg": "#fbfffe",
        "border_color": "#b7d8d2",
        "text_color": "#102a2a",
        "muted_text": "#55716f",
        "primary": "#00897b",
        "primary_light": "#dff6f1",
        "tab_bg": "#f8fffd",
    },
    "심텍 그린": {
        "app_bg": "#f0f8f1",
        "surface_bg": "#ffffff",
        "border_color": "#c8dfc9",
        "text_color": "#172417",
        "muted_text": "#617064",
        "primary": "#198754",
        "primary_light": "#e6f4ec",
        "tab_bg": "#eef6ef",
    },
    "TLB 퍼플": {
        "app_bg": "#f6f3fb",
        "surface_bg": "#ffffff",
        "border_color": "#d9cfea",
        "text_color": "#221933",
        "muted_text": "#6b617a",
        "primary": "#6f42c1",
        "primary_light": "#f0e9fb",
        "tab_bg": "#f1ecf8",
    },
    "다크 네이비": {
        "app_bg": "#eef2f6",
        "surface_bg": "#ffffff",
        "border_color": "#c9d5e2",
        "text_color": "#14213d",
        "muted_text": "#5c677d",
        "primary": "#1f3a5f",
        "primary_light": "#e7edf5",
        "tab_bg": "#e9eef5",
    },
}


def apply_theme(theme_name: str) -> None:
    """설정된 화면 색상 테마를 전역 UI 색상에 적용합니다."""
    theme = THEMES.get(theme_name, THEMES["MES 블루"])
    globals_to_update = {
        "APP_BG": theme["app_bg"],
        "SURFACE_BG": theme["surface_bg"],
        "BORDER_COLOR": theme["border_color"],
        "TEXT_COLOR": theme["text_color"],
        "MUTED_TEXT": theme["muted_text"],
        "PRIMARY": theme["primary"],
        "PRIMARY_LIGHT": theme["primary_light"],
        "TAB_BG": theme["tab_bg"],
    }
    globals().update(globals_to_update)


# ==================================================
# 공통 파일/설정 함수
# ==================================================
def get_desktop_path() -> Path:
    """현재 Windows 사용자의 바탕화면 경로를 반환합니다."""
    return Path.home() / "Desktop"


def get_default_config() -> dict:
    """config.json이 없거나 값이 비어 있을 때 사용할 기본값입니다."""
    desktop = get_desktop_path()
    return {
        "excel_file": "",
        "source_dnc_folder": str(desktop / "KCC_PKG"),
        "transfer_dnc_folder": str(desktop / "DNC"),
        "dnc_delete_seconds": DNC_DELETE_SECONDS,
        "clear_common_after_normal": False,
        "theme": "MES 블루",
    }


def load_config() -> dict:
    """config.json을 읽고, 없는 값은 기본값으로 채웁니다."""
    config = get_default_config()
    try:
        if CONFIG_FILE.exists():
            saved = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            if isinstance(saved, dict):
                config.update({key: value for key, value in saved.items() if value not in (None, "")})
    except Exception:
        # 설정 파일이 손상되어도 프로그램은 기본값으로 실행되게 합니다.
        pass
    return config


def save_config(config: dict) -> None:
    """현재 설정을 config.json에 저장합니다."""
    CONFIG_FILE.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")


def select_excel_file(parent, config: dict, excel_var: tk.StringVar) -> None:
    """설정 탭에서 작업일보 Excel 파일을 선택합니다."""
    file_path = filedialog.askopenfilename(
        parent=parent,
        title="작업일보 Excel 파일 선택",
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    if file_path:
        config["excel_file"] = file_path
        excel_var.set(file_path)
        save_config(config)


def ensure_excel_file_selected(parent, config: dict, excel_var: tk.StringVar | None = None) -> bool:
    """작업일보 파일이 없을 때만 사용자에게 한 번 선택을 요청합니다."""
    current_path = ""
    if excel_var is not None:
        current_path = excel_var.get().strip()
    if not current_path:
        current_path = config.get("excel_file", "").strip()
    if not current_path and CONFIG_FILE.exists():
        try:
            saved = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            current_path = str(saved.get("excel_file", "")).strip()
        except Exception:
            current_path = ""

    if current_path and Path(current_path).exists():
        config["excel_file"] = current_path
        if excel_var is not None:
            excel_var.set(current_path)
        save_config(config)
        return True

    messagebox.showwarning(
        "작업일보 선택 필요",
        "작업일보 Excel 파일이 선택되어 있지 않거나 파일을 찾을 수 없습니다.\n\n"
        "DNC 이력을 저장할 작업일보 Excel 파일을 선택해주세요.",
        parent=parent,
    )
    file_path = filedialog.askopenfilename(
        parent=parent,
        title="작업일보 Excel 파일 선택",
        initialdir=str(Path(current_path).parent) if current_path else str(get_desktop_path()),
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    if not file_path:
        return False

    config["excel_file"] = file_path
    if excel_var is not None:
        excel_var.set(file_path)
    save_config(config)
    return True


def open_path(path: Path) -> None:
    """파일 또는 폴더를 Windows 기본 프로그램으로 엽니다."""
    if sys.platform.startswith("win"):
        os.startfile(str(path))
    else:
        subprocess.Popen(["xdg-open", str(path)])


def open_log_excel(config: dict) -> None:
    """선택된 작업일보 Excel 파일을 엽니다."""
    if not ensure_excel_file_selected(None, config):
        return
    excel_file = Path(config.get("excel_file", ""))
    try:
        open_path(excel_file)
    except Exception as exc:
        messagebox.showerror("작업일보 열기 실패", str(exc))


# ==================================================
# DNC 파일 처리 함수
# ==================================================
def delete_existing_dnc_txt(transfer_folder: Path) -> None:
    """DNC 전송 폴더에 남아 있는 txt 파일을 실행 전에 모두 삭제합니다."""
    transfer_folder.mkdir(parents=True, exist_ok=True)
    for txt_file in transfer_folder.glob("*.txt"):
        txt_file.unlink()


def search_condition_file(condition_name: str, source_folder: Path) -> list[Path]:
    """원본 DNC 폴더와 하위 폴더에서 조건명과 일치하는 txt 파일을 찾습니다."""
    source_folder.mkdir(parents=True, exist_ok=True)
    normalized = condition_name.strip()
    if normalized.lower().endswith(".txt"):
        normalized = normalized[:-4]

    matches = []
    for file_path in source_folder.rglob("*.txt"):
        if file_path.stem == normalized or file_path.name == condition_name.strip():
            matches.append(file_path)
    return matches


def copy_dnc_file(source_file: Path, transfer_folder: Path) -> Path:
    """조건 txt 파일을 DNC 전송 폴더로 복사합니다."""
    transfer_folder.mkdir(parents=True, exist_ok=True)
    copied_file = transfer_folder / source_file.name
    shutil.copy2(source_file, copied_file)
    return copied_file


def delete_after_delay(copied_file: Path, seconds: int, status_callback=None) -> None:
    """지정된 초만큼 기다린 후 복사된 DNC txt 파일을 삭제합니다."""
    if status_callback:
        status_callback(f"DNC 삭제 대기중 ({seconds}초)")
    time.sleep(seconds)
    if copied_file.exists():
        copied_file.unlink()
    if status_callback:
        status_callback("DNC 완료")


# ==================================================
# 검증 함수
# ==================================================
def check_mes_core(lot_no: str, process_code: str) -> bool:
    """LOT No 9번째 문자부터 4자리와 공정코드 끝 2자리로 MES Core 일치화를 판정합니다."""
    lot_no = lot_no.strip()
    process_code = process_code.strip().upper()
    if len(lot_no) < 12 or len(process_code) < 2:
        return False

    lot_core = lot_no[8:12]
    process_tail = process_code[-2:]
    return (lot_core == "0000" and process_tail == "T1") or (lot_core == "0205" and process_tail == "11")


def get_mes_core_message(lot_no: str, process_code: str) -> tuple[bool, str]:
    """MES Core 판정 결과와 작업자가 확인할 상세 사유를 함께 반환합니다."""
    lot_no = lot_no.strip()
    process_code = process_code.strip().upper()
    if len(lot_no) < 12:
        return False, f"LOT No가 너무 짧습니다. 현재 LOT No: {lot_no or '빈칸'}"
    if len(process_code) < 2:
        return False, f"공정코드가 너무 짧습니다. 현재 공정코드: {process_code or '빈칸'}"

    lot_core = lot_no[8:12]
    process_tail = process_code[-2:]
    if lot_core == "0000" and process_tail == "T1":
        return True, "OK - LOT 중간 4자리 0000 / 공정코드 끝자리 T1"
    if lot_core == "0205" and process_tail == "11":
        return True, "OK - LOT 중간 4자리 0205 / 공정코드 끝자리 11"
    return (
        False,
        "MES Core 기준이 맞지 않습니다. "
        f"LOT 중간 4자리: {lot_core}, 공정코드 끝 2자리: {process_tail}. "
        "허용 기준은 0000+T1 또는 0205+11 입니다.",
    )


def check_condition_ok(lot1: dict, lot2: dict | None = None) -> bool:
    """작업조건 KCC_ 시작 여부와 지그/2LOT 조건 일치 여부를 확인합니다."""
    if not lot1.get("condition", "").strip().startswith("KCC_"):
        return False
    if not lot1.get("jig", "").strip():
        return False
    if lot2:
        return (
            lot1.get("condition", "").strip() == lot2.get("condition", "").strip()
            and lot1.get("jig", "").strip() == lot2.get("jig", "").strip()
        )
    return True


def get_single_condition_message(lot: dict) -> tuple[bool, str]:
    """LOT 한 개의 작업조건/지그 기본 조건을 상세하게 확인합니다."""
    condition = lot.get("condition", "").strip()
    jig = lot.get("jig", "").strip()
    if not condition:
        return False, "작업조건이 비어 있습니다. [작업조건 / 지그 불러오기]를 눌러주세요."
    if not condition.startswith("KCC_"):
        return False, f"작업조건이 KCC_로 시작하지 않습니다. 현재 작업조건: {condition}"
    if not jig:
        return False, "지그가 비어 있습니다. [작업조건 / 지그 불러오기]를 눌러주세요."
    return True, "OK - 작업조건 KCC_ 시작 / 지그 입력 완료"


def get_lot_match_message(lot1: dict, lot2: dict) -> tuple[bool, str]:
    """2LOT 작업 시 LOT 1/LOT 2의 작업조건과 지그 일치 여부를 상세하게 확인합니다."""
    lot1_no = lot1.get("lot_no", "").strip()
    lot2_no = lot2.get("lot_no", "").strip()
    lot1_condition = lot1.get("condition", "").strip()
    lot2_condition = lot2.get("condition", "").strip()
    lot1_jig = lot1.get("jig", "").strip()
    lot2_jig = lot2.get("jig", "").strip()
    messages = []
    if lot1_no and lot2_no and lot1_no == lot2_no:
        messages.append(f"LOT No 중복: LOT 1 [{lot1_no}] / LOT 2 [{lot2_no}]")
    if lot1_condition != lot2_condition:
        messages.append(f"작업조건 불일치: LOT 1 [{lot1_condition or '빈칸'}] / LOT 2 [{lot2_condition or '빈칸'}]")
    if lot1_jig != lot2_jig:
        messages.append(f"지그 불일치: LOT 1 [{lot1_jig or '빈칸'}] / LOT 2 [{lot2_jig or '빈칸'}]")
    if messages:
        return False, " / ".join(messages)
    if not lot1_condition or not lot1_jig:
        return False, "작업조건 또는 지그가 비어 있습니다."
    return True, "OK - LOT 1 / LOT 2 작업조건과 지그가 같습니다."


def validate_positive_number(value: str, field_name: str, required: bool = True) -> tuple[bool, str]:
    """매수/Stack처럼 0보다 큰 숫자만 허용하는 필드 검증입니다."""
    text = value.strip()
    if not text:
        return (not required, "" if not required else f"{field_name}은(는) 필수입니다.")
    if not text.isdigit():
        return False, f"{field_name}은(는) 숫자만 입력 가능합니다."
    if int(text) <= 0:
        return False, f"{field_name}은(는) 0보다 커야 합니다."
    return True, ""


def validate_zero_or_positive_number(value: str, field_name: str) -> tuple[bool, str]:
    """신규 모델 검증 매수처럼 0 이상 숫자를 반드시 입력해야 하는 필드 검증입니다."""
    text = value.strip()
    if not text:
        return False, f"{field_name}은(는) 필수입니다. 더미 작업이면 0을 입력하세요."
    if not text.isdigit():
        return False, f"{field_name}은(는) 숫자만 입력 가능합니다."
    return True, ""


def validate_lot_required(lot: dict, lot_name: str, require_qty: bool) -> list[str]:
    """LOT 입력 필수값을 확인합니다."""
    errors = []
    required_fields = [
        ("step", "STEP"),
        ("round", "차수"),
        ("manage_no", "관리번호"),
        ("lot_no", "LOT No"),
        ("process_code", "공정코드"),
        ("condition", "작업조건"),
        ("jig", "지그"),
    ]
    if require_qty:
        required_fields.insert(4, ("qty", "매수"))

    for key, label in required_fields:
        if not lot.get(key, "").strip():
            errors.append(f"{lot_name} {label}을(를) 입력하세요.")

    if require_qty:
        ok, message = validate_positive_number(lot.get("qty", ""), f"{lot_name} 매수", required=True)
        if not ok:
            errors.append(message)
    elif lot.get("qty", "").strip():
        ok, message = validate_positive_number(lot.get("qty", ""), f"{lot_name} 매수", required=False)
        if not ok:
            errors.append(message)

    if lot.get("condition", "").strip() and not lot.get("condition", "").strip().startswith("KCC_"):
        errors.append(f"{lot_name} 작업조건은 반드시 KCC_ 로 시작해야 합니다.")
    return errors


def validate_normal_dnc(common: dict, lot1: dict, lot2: dict | None) -> tuple[bool, list[str]]:
    """일반 DNC 입력값 전체 검증입니다."""
    errors = []
    for key, label in (("work_date", "작업일자"), ("shift_group", "조"), ("shift", "근무"), ("worker", "작업자")):
        if not common.get(key, "").strip():
            errors.append(f"{label}을(를) 입력하세요.")

    errors.extend(validate_lot_required(lot1, "LOT 1", require_qty=True))
    if lot2:
        errors.extend(validate_lot_required(lot2, "LOT 2", require_qty=True))

    lot1_mes_ok, lot1_mes_message = get_mes_core_message(lot1.get("lot_no", ""), lot1.get("process_code", ""))
    if not lot1_mes_ok:
        errors.append(f"LOT 1 MES Core 불일치: {lot1_mes_message}")
    lot1_condition_ok, lot1_condition_message = get_single_condition_message(lot1)
    if not lot1_condition_ok:
        errors.append(f"LOT 1 조건 적용 불가: {lot1_condition_message}")

    if lot2:
        lot2_mes_ok, lot2_mes_message = get_mes_core_message(lot2.get("lot_no", ""), lot2.get("process_code", ""))
        if not lot2_mes_ok:
            errors.append(f"LOT 2 MES Core 불일치: {lot2_mes_message}")
        lot2_condition_ok, lot2_condition_message = get_single_condition_message(lot2)
        if not lot2_condition_ok:
            errors.append(f"LOT 2 조건 적용 불가: {lot2_condition_message}")
        lot_match_ok, lot_match_message = get_lot_match_message(lot1, lot2)
        if not lot_match_ok:
            errors.append(f"2LOT 조건 불일치: {lot_match_message}")
    return len(errors) == 0, errors


def validate_new_model_dnc(common: dict, lot: dict) -> tuple[bool, list[str]]:
    """신규 모델 검증 DNC 입력값 전체 검증입니다."""
    errors = []
    for key, label in (("work_date", "작업일자"), ("shift_group", "조"), ("shift", "근무"), ("worker", "작업자")):
        if not common.get(key, "").strip():
            errors.append(f"{label}을(를) 입력하세요.")

    errors.extend(validate_lot_required(lot, "신규 모델", require_qty=False))
    ok, message = validate_zero_or_positive_number(lot.get("qty", ""), "신규 모델 매수")
    if not ok:
        errors.append(message)
    mes_ok, mes_message = get_mes_core_message(lot.get("lot_no", ""), lot.get("process_code", ""))
    if not mes_ok:
        errors.append(f"MES Core 불일치: {mes_message}")
    condition_ok, condition_message = get_single_condition_message(lot)
    if not condition_ok:
        errors.append(f"조건 적용 불가: {condition_message}")
    return len(errors) == 0, errors


# ==================================================
# Excel 저장 함수
# ==================================================
def get_next_empty_row(ws) -> int:
    """A:P 범위 기준으로 8행부터 첫 빈 행을 찾습니다."""
    row = 8
    while True:
        has_value = any(ws.cell(row=row, column=col).value not in (None, "") for col in range(1, 17))
        if not has_value:
            return row
        row += 1


def open_log_workbook(config: dict):
    """작업일보 파일과 KCC PKG 시트를 열고 기본 오류를 메시지로 변환합니다."""
    excel_file = config.get("excel_file", "")
    if not excel_file:
        raise FileNotFoundError("작업일보 Excel 파일을 선택해주세요.")
    path = Path(excel_file)
    if not path.exists():
        raise FileNotFoundError("작업일보 Excel 파일을 선택해주세요.")

    try:
        if not zipfile.is_zipfile(path):
            raise ValueError("작업일보 Excel 파일 형식이 손상되었거나 올바른 Excel 파일이 아닙니다.")
        workbook = load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm")
    except PermissionError:
        raise PermissionError("작업일보 Excel 파일이 열려 있어 저장할 수 없습니다.\n파일을 닫고 다시 실행해주세요.")
    except zipfile.BadZipFile:
        raise ValueError("작업일보 Excel 파일 형식이 손상되었거나 올바른 Excel 파일이 아닙니다.")

    if LOG_SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise KeyError("작업일보 파일에 KCC PKG 시트가 없습니다.")
    return workbook, workbook[LOG_SHEET_NAME], path


def save_workbook_safely(workbook, path: Path) -> None:
    """Excel 저장 실패 시 사용자가 이해할 수 있는 메시지를 발생시킵니다."""
    while True:
        try:
            workbook.save(path)
            return
        except PermissionError:
            retry = messagebox.askretrycancel(
                "작업일보 저장 대기",
                "작업일보 Excel 파일이 열려 있어 저장할 수 없습니다.\n\n"
                "Excel에서 작업일보를 닫은 뒤 [다시 시도]를 눌러주세요.\n"
                "취소를 누르면 DNC 작업이 중단됩니다.",
            )
            if not retry:
                raise PermissionError("작업일보 Excel 파일이 열려 있어 저장할 수 없습니다.\n파일을 닫고 다시 실행해주세요.")


def make_condition_key(step: str, round_no: str, manage_no: str, process_code: str) -> str:
    """조건 마스터에서 중복을 제거하기 위한 기준 키를 만듭니다.

    작업일보에는 공정코드가 저장되지 않는 행도 있어서, 마스터 중복 판단은
    STEP/차수/관리번호 기준으로 합니다. 공정코드는 기록용으로 보존합니다.
    """
    return "|".join(
        [
            step.strip(),
            round_no.strip(),
            manage_no.strip(),
        ]
    )


def merge_condition_records(records: list[dict]) -> list[dict]:
    """같은 STEP/차수/관리번호 조건은 한 줄로 합치고, 뒤에 들어온 최신 값을 우선합니다."""
    merged: dict[str, dict] = {}
    for record in records:
        key = make_condition_key(
            record.get("step", ""),
            record.get("round", ""),
            record.get("manage_no", ""),
            record.get("process_code", ""),
        )
        if not key.replace("|", "").strip():
            continue
        if key not in merged:
            merged[key] = dict(record)
            continue

        current = merged[key]
        for field in ("step", "round", "manage_no", "condition", "jig", "source", "updated_at", "lot_no"):
            value = str(record.get(field, "")).strip()
            if value:
                current[field] = value
        process_code = str(record.get("process_code", "")).strip()
        if process_code:
            current["process_code"] = process_code

    return sorted(merged.values(), key=lambda item: (item.get("step", ""), item.get("round", ""), item.get("manage_no", "")))


def load_condition_master() -> list[dict]:
    """작업일보에서 추출해 둔 조건 마스터 JSON을 읽습니다."""
    try:
        if CONDITION_MASTER_FILE.exists():
            data = json.loads(CONDITION_MASTER_FILE.read_text(encoding="utf-8"))
            if isinstance(data, list):
                return merge_condition_records(data)
    except Exception:
        pass
    return []


def save_condition_master(records: list[dict]) -> None:
    """조건 마스터 JSON을 저장합니다."""
    records = merge_condition_records(records)
    CONDITION_MASTER_FILE.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def upsert_condition_master(lot: dict, condition: str, jig: str, source: str) -> None:
    """현재 입력값과 불러온 조건/지그를 조건 마스터에 추가 또는 갱신합니다."""
    step = lot.get("step", "").strip()
    round_no = lot.get("round", "").strip()
    manage_no = lot.get("manage_no", "").strip()
    process_code = lot.get("process_code", "").strip()
    if not (step and round_no and manage_no and condition and jig):
        return

    key = make_condition_key(step, round_no, manage_no, process_code)
    records = load_condition_master()
    for record in records:
        record_key = make_condition_key(
            record.get("step", ""),
            record.get("round", ""),
            record.get("manage_no", ""),
            record.get("process_code", ""),
        )
        if record_key == key:
            record.update(
                {
                    "lot_no": lot.get("lot_no", "").strip(),
                    "condition": condition,
                    "jig": jig,
                    "source": source,
                    "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
            )
            save_condition_master(records)
            return

    records.append(
        {
            "step": step,
            "round": round_no,
            "manage_no": manage_no,
            "process_code": process_code,
            "lot_no": lot.get("lot_no", "").strip(),
            "condition": condition,
            "jig": jig,
            "source": source,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
    )
    save_condition_master(records)


def rebuild_condition_master_from_log(config: dict) -> int:
    """작업일보 KCC PKG 시트의 최신 이력을 기존 마스터에 병합합니다.

    작업일보에서 사라진 모델도 조건 마스터에서는 보존합니다.
    """
    workbook, ws, _path = open_log_workbook(config)
    records = load_condition_master()
    try:
        for row in range(8, ws.max_row + 1):
            step = str(ws.cell(row=row, column=5).value or "").strip()
            round_no = str(ws.cell(row=row, column=6).value or "").strip()
            manage_no = str(ws.cell(row=row, column=7).value or "").strip()
            # 작업일보에는 공정코드 저장 열이 없어서, 같은 키 구조 유지를 위해 빈 값으로 둡니다.
            process_code = ""
            condition = str(ws.cell(row=row, column=11).value or "").strip()
            jig = str(ws.cell(row=row, column=14).value or "").strip()
            lot_no = str(ws.cell(row=row, column=8).value or "").strip()
            if not (step and round_no and manage_no and condition and jig):
                continue
            records.append(
                {
                    "step": step,
                    "round": round_no,
                    "manage_no": manage_no,
                    "process_code": process_code,
                    "lot_no": lot_no,
                    "condition": condition,
                    "jig": jig,
                    "source": f"작업일보 {row}행",
                    "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
            )
    finally:
        workbook.close()

    save_condition_master(records)
    return len(load_condition_master())


def lookup_condition_jig_from_master(lot: dict) -> tuple[str, str, str]:
    """조건 마스터에서 작업조건/지그를 먼저 찾습니다."""
    records = load_condition_master()
    step = lot.get("step", "").strip()
    round_no = lot.get("round", "").strip()
    manage_no = lot.get("manage_no", "").strip()
    process_code = lot.get("process_code", "").strip()
    lot_no = lot.get("lot_no", "").strip()

    for record in reversed(records):
        if (
            step
            and round_no
            and manage_no
            and record.get("step", "") == step
            and record.get("round", "") == round_no
            and record.get("manage_no", "") == manage_no
            and (not process_code or not record.get("process_code") or record.get("process_code") == process_code)
        ):
            return record.get("condition", ""), record.get("jig", ""), f"조건 마스터({record.get('source', '저장값')})"

    # STEP/차수가 입력되어 있으면 관리번호만으로 fallback 하지 않습니다.
    # 같은 관리번호 안에서 1차/2차 조건이 달라질 수 있기 때문입니다.
    if not (step and round_no):
        for record in reversed(records):
            if manage_no and record.get("manage_no") == manage_no:
                return record.get("condition", ""), record.get("jig", ""), f"조건 마스터({record.get('source', '관리번호')})"

    if not (step and round_no):
        for record in reversed(records):
            if lot_no and record.get("lot_no") == lot_no:
                return record.get("condition", ""), record.get("jig", ""), f"조건 마스터({record.get('source', 'LOT No')})"

    return "", "", ""


def lookup_condition_jig_from_history(config: dict, lot: dict) -> tuple[str, str, str]:
    """기존 작업일보 이력에서 작업조건(K열)과 지그(N열)를 찾아옵니다.

    작업조건/지그는 작업자가 직접 입력하는 값이 아니라 기존 진행 이력을 참고해야 하므로,
    KCC PKG 시트 8행 이후를 아래 우선순위로 뒤에서부터 검색합니다.

    1순위: STEP + 차수 + 관리번호 모두 일치
    2순위: 관리번호 일치
    3순위: LOT No 일치
    """
    workbook, ws, _path = open_log_workbook(config)
    try:
        step = lot.get("step", "").strip()
        round_no = lot.get("round", "").strip()
        manage_no = lot.get("manage_no", "").strip()
        lot_no = lot.get("lot_no", "").strip()

        candidates = []
        for row in range(ws.max_row, 7, -1):
            row_step = str(ws.cell(row=row, column=5).value or "").strip()
            row_round = str(ws.cell(row=row, column=6).value or "").strip()
            row_manage_no = str(ws.cell(row=row, column=7).value or "").strip()
            row_lot_no = str(ws.cell(row=row, column=8).value or "").strip()
            row_condition = str(ws.cell(row=row, column=11).value or "").strip()
            row_jig = str(ws.cell(row=row, column=14).value or "").strip()

            if not row_condition or not row_jig:
                continue
            candidates.append((row, row_step, row_round, row_manage_no, row_lot_no, row_condition, row_jig))

        for row, row_step, row_round, row_manage_no, _row_lot_no, condition, jig in candidates:
            if step and round_no and manage_no and row_step == step and row_round == round_no and row_manage_no == manage_no:
                return condition, jig, f"{row}행 이력(STEP+차수+관리번호)"

        condition, jig, source = lookup_condition_jig_from_master(lot)
        if condition and jig:
            return condition, jig, source

        if not (step and round_no):
            for row, _row_step, _row_round, row_manage_no, _row_lot_no, condition, jig in candidates:
                if manage_no and row_manage_no == manage_no:
                    return condition, jig, f"{row}행 이력(관리번호)"

        if not (step and round_no):
            for row, _row_step, _row_round, _row_manage_no, row_lot_no, condition, jig in candidates:
                if lot_no and row_lot_no == lot_no:
                    return condition, jig, f"{row}행 이력(LOT No)"

        return "", "", ""
    finally:
        workbook.close()


def write_common_lot_row(ws, row: int, common: dict, lot: dict, stack: str, model_change: str, frequent_check: list[str] | None = None) -> None:
    """일반 DNC 작업일보 한 줄을 기록합니다."""
    qty = int(lot["qty"])
    result = round(qty * 0.2, 1)
    values = [
        common["work_date"],
        common["shift_group"],
        common["shift"],
        common["worker"],
        lot["step"],
        lot["round"],
        lot["manage_no"],
        lot["lot_no"],
        qty,
        result,
        lot["condition"],
        "",
        stack,
        lot["jig"],
        model_change,
        "",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(row=row, column=col).value = value
    if frequent_check:
        for offset, value in enumerate(frequent_check, start=17):
            ws.cell(row=row, column=offset).value = value


def save_normal_dnc_log(config: dict, common: dict, lots: list[dict], stack: str, model_change: bool, frequent_check: list[str] | None = None) -> tuple[Path, list[int]]:
    """일반 DNC 내용을 작업일보에 저장하고 저장된 행 번호를 반환합니다."""
    workbook, ws, path = open_log_workbook(config)
    rows = []
    try:
        start_row = get_next_empty_row(ws)
        for index, lot in enumerate(lots):
            row = start_row + index
            rows.append(row)
            row_frequent_check = None
            if frequent_check:
                # 초품 4Point(Q:V)는 모든 LOT 저장행에 기록합니다.
                # 지그교체 하부핀(W:AB)은 기종교체 표시가 들어가는 첫 저장행에만 기록합니다.
                row_frequent_check = frequent_check[:6] + (frequent_check[6:] if model_change and index == 0 else [""] * 6)
            write_common_lot_row(
                ws,
                row,
                common,
                lot,
                stack,
                "기종교체" if model_change and index == 0 else "",
                row_frequent_check,
            )
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()
    return path, rows


def update_normal_burr_result(config: dict, rows: list[int], burr_ok: bool) -> None:
    """DNC 완료 후 Burr 결과와 기록시간을 일반 DNC 저장행에 반영합니다."""
    workbook, ws, path = open_log_workbook(config)
    try:
        result = "이상 없음" if burr_ok else "Burr 발생"
        now_text = datetime.now().strftime("%H:%M:%S")
        for row in rows:
            ws.cell(row=row, column=12).value = result
            ws.cell(row=row, column=16).value = now_text
            if not burr_ok:
                ws.cell(row=row, column=12).font = Font(color="FF0000", bold=False)
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()


def save_new_model_log(config: dict, common: dict, lot: dict, leader_name: str) -> tuple[Path, int]:
    """신규 모델 검증 DNC 내용을 작업일보에 저장합니다."""
    workbook, ws, path = open_log_workbook(config)
    try:
        row = get_next_empty_row(ws)
        qty_text = lot.get("qty", "").strip()
        qty_number = int(qty_text)
        qty_value = "더미" if qty_number == 0 else qty_number
        result_value = "" if qty_number == 0 else round(qty_number * 0.2, 1)
        values = [
            common["work_date"],
            common["shift_group"],
            common["shift"],
            leader_name,
            lot["step"],
            lot["round"],
            lot["manage_no"],
            lot["lot_no"],
            qty_value,
            result_value,
            lot["condition"],
            "",
            "",
            lot["jig"],
            "신규 검증",
            "",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col).value = value
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()
    return path, row


def update_new_model_result(config: dict, row: int, condition_name: str, first_article_ok: bool) -> None:
    """신규 모델 DNC 완료 후 초도품 확인 결과와 기록시간을 저장합니다."""
    workbook, ws, path = open_log_workbook(config)
    try:
        k_cell = ws.cell(row=row, column=11)
        if first_article_ok:
            k_cell.value = condition_name
        else:
            k_cell.value = f"[검증 NG 발생] {condition_name}"
            k_cell.font = Font(name="맑은 고딕", size=10, color="FF0000", bold=False)
        ws.cell(row=row, column=16).value = datetime.now().strftime("%H:%M:%S")
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()


# ==================================================
# GUI
# ==================================================
class LabeledEntry(ttk.Frame):
    """라벨과 입력칸을 한 줄로 만드는 작은 공용 위젯입니다."""

    def __init__(self, parent, label: str, width: int = 18):
        super().__init__(parent)
        self.var = tk.StringVar()
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        self.entry = ttk.Entry(self, textvariable=self.var, width=width, style="Wide.TEntry")
        self.entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

    def get(self) -> str:
        return self.var.get().strip()

    def set(self, value: str) -> None:
        self.var.set(value)

    def clear(self) -> None:
        self.var.set("")


class DateField(ttk.Frame):
    """날짜 형식을 통일하기 위한 선택식 날짜 입력 위젯입니다."""

    def __init__(self, parent, label: str):
        super().__init__(parent)
        self.var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        self.entry = ttk.Entry(self, textvariable=self.var, width=14, style="Wide.TEntry", state="readonly")
        self.entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(self, text="선택", width=6, command=self.open_picker).pack(side=tk.LEFT, padx=(6, 0))

    def get(self) -> str:
        return self.var.get().strip()

    def set(self, value: str) -> None:
        self.var.set(value)

    def clear(self) -> None:
        self.var.set(datetime.now().strftime("%Y-%m-%d"))

    def open_picker(self) -> None:
        picker = tk.Toplevel(self)
        picker.title("작업일자 선택")
        picker.configure(bg=APP_BG)
        picker.resizable(False, False)

        today = datetime.now()
        try:
            selected = datetime.strptime(self.var.get(), "%Y-%m-%d")
        except ValueError:
            selected = today

        year_var = tk.IntVar(value=selected.year)
        month_var = tk.IntVar(value=selected.month)
        day_var = tk.IntVar(value=selected.day)

        top = ttk.Frame(picker, padding=(12, 12, 12, 6))
        top.pack(fill=tk.X)
        ttk.Spinbox(top, from_=today.year - 5, to=today.year + 5, textvariable=year_var, width=8).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Spinbox(top, from_=1, to=12, textvariable=month_var, width=5).pack(side=tk.LEFT, padx=(0, 6))

        days_frame = ttk.Frame(picker, padding=(12, 6, 12, 12))
        days_frame.pack()

        def refresh_days() -> None:
            for child in days_frame.winfo_children():
                child.destroy()
            last_day = calendar.monthrange(year_var.get(), month_var.get())[1]
            for day in range(1, last_day + 1):
                button = ttk.Button(
                    days_frame,
                    text=str(day),
                    width=4,
                    command=lambda value=day: select_day(value),
                )
                button.grid(row=(day - 1) // 7, column=(day - 1) % 7, padx=2, pady=2)

        def select_day(day: int) -> None:
            day_var.set(day)
            self.var.set(f"{year_var.get():04d}-{month_var.get():02d}-{day_var.get():02d}")
            picker.destroy()

        ttk.Button(top, text="변경", command=refresh_days).pack(side=tk.LEFT)
        refresh_days()


class SegmentedField(ttk.Frame):
    """조/근무처럼 정해진 값만 선택하게 하는 버튼형 입력 위젯입니다."""

    def __init__(self, parent, label: str, options: list[str]):
        super().__init__(parent)
        self.var = tk.StringVar(value=options[0] if options else "")
        self.buttons: list[tk.Button] = []
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        wrap = tk.Frame(self, bg=APP_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        wrap.pack(side=tk.LEFT, fill=tk.X, expand=True)
        for option in options:
            button = tk.Button(
                wrap,
                text=option,
                command=lambda value=option: self.set(value),
                relief=tk.FLAT,
                bd=0,
                padx=18,
                pady=6,
                cursor="hand2",
                font=("맑은 고딕", 10),
            )
            button.pack(side=tk.LEFT, fill=tk.X, expand=True)
            self.buttons.append(button)
        self.update_buttons()

    def get(self) -> str:
        return self.var.get().strip()

    def set(self, value: str) -> None:
        self.var.set(value)
        self.update_buttons()

    def clear(self) -> None:
        self.update_buttons()

    def update_buttons(self) -> None:
        for button in self.buttons:
            selected = button.cget("text") == self.var.get()
            button.configure(bg=PRIMARY_LIGHT if selected else SURFACE_BG, fg=PRIMARY if selected else TEXT_COLOR)


class RoundField(ttk.Frame):
    """차수를 1차~8차 버튼으로 선택하게 하는 작은 입력 위젯입니다."""

    OPTIONS = [f"{number}차" for number in range(1, 9)]

    def __init__(self, parent, label: str):
        super().__init__(parent)
        self.var = tk.StringVar()
        self.buttons: list[tk.Button] = []
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        wrap = tk.Frame(self, bg=APP_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        wrap.pack(side=tk.LEFT, fill=tk.X, expand=True)
        for option in self.OPTIONS:
            button = tk.Button(
                wrap,
                text=option,
                command=lambda value=option: self.toggle(value),
                relief=tk.FLAT,
                bd=0,
                width=3,
                padx=1,
                pady=6,
                cursor="hand2",
                font=("맑은 고딕", 9),
            )
            button.pack(side=tk.LEFT, fill=tk.X, expand=True)
            self.buttons.append(button)
        self.update_buttons()

    def get(self) -> str:
        return self.var.get().strip()

    def toggle(self, value: str) -> None:
        """작업자가 같은 차수 버튼을 한 번 더 누르면 선택을 취소합니다."""
        self.var.set("" if self.var.get() == value else value)
        self.update_buttons()

    def set(self, value: str) -> None:
        self.var.set(value)
        self.update_buttons()

    def clear(self) -> None:
        self.var.set("")
        self.update_buttons()

    def update_buttons(self) -> None:
        for button in self.buttons:
            selected = button.cget("text") == self.var.get()
            button.configure(bg=PRIMARY_LIGHT if selected else SURFACE_BG, fg=PRIMARY if selected else TEXT_COLOR)


class SimpleTabNotebook(ttk.Frame):
    """Tkinter 기본 Notebook 대신 선택 테두리가 선명한 간단 탭 UI입니다."""

    def __init__(self, parent):
        super().__init__(parent)
        self.selected_index = -1
        self.pages: list[tk.Frame] = []
        self.labels: list[tk.Label] = []
        self.tab_bar = tk.Frame(self, bg=APP_BG)
        self.tab_bar.pack(fill=tk.X)
        self.page_area = tk.Frame(self, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        self.page_area.pack(fill=tk.BOTH, expand=True)

    def add(self, page: tk.Frame, text: str) -> None:
        index = len(self.pages)
        self.pages.append(page)
        label = tk.Label(
            self.tab_bar,
            text=text,
            bg=TAB_BG,
            fg=MUTED_TEXT,
            padx=28,
            pady=12,
            font=("맑은 고딕", 10),
            cursor="hand2",
            highlightthickness=1,
            highlightbackground=BORDER_COLOR,
            bd=0,
        )
        label.pack(side=tk.LEFT)
        label.bind("<Button-1>", lambda _event, i=index: self.select(i))
        label.bind("<Enter>", lambda _event, i=index: self.hover(i))
        label.bind("<Leave>", lambda _event, i=index: self.update_label(i))
        self.labels.append(label)
        page.pack_forget()
        if index == 0:
            self.select(0)
        else:
            self.update_label(index)

    def select(self, index: int) -> None:
        for page in self.pages:
            page.pack_forget()
        self.selected_index = index
        self.pages[index].pack(fill=tk.BOTH, expand=True)
        for i in range(len(self.labels)):
            self.update_label(i)

    def hover(self, index: int) -> None:
        if index != self.selected_index:
            self.labels[index].configure(bg="#f7faff", fg=TEXT_COLOR)

    def update_label(self, index: int) -> None:
        if index == self.selected_index:
            self.labels[index].configure(bg=SURFACE_BG, fg=PRIMARY, highlightbackground=PRIMARY)
            self.page_area.configure(highlightbackground=PRIMARY)
        else:
            self.labels[index].configure(bg=TAB_BG, fg=MUTED_TEXT, highlightbackground=BORDER_COLOR)


def validate_frequent_check_values(values: list[str], require_jig_check: bool = False) -> tuple[bool, str]:
    """초품 4Point는 매번, 지그 하부핀 확인은 기종교체일 때만 검증합니다."""
    if len(values) != 12:
        return False, "자주검사 데이터가 올바르지 않습니다."
    first_count = sum(1 for value in values[:6] if value == "OK")
    jig_count = sum(1 for value in values[6:] if value == "OK")
    if first_count == 0:
        return False, "초품 4Point 확인은 1개 축 이상 선택해야 합니다."
    if require_jig_check:
        if jig_count == 0:
            return False, "기종교체 시에는 지그교체 하부핀 확인도 1개 축 이상 선택해야 합니다."
        if first_count != jig_count:
            return False, f"초품 확인 축 수({first_count})와 지그교체 확인 축 수({jig_count})가 같아야 합니다."
        return True, f"초품 {first_count}개 축 / 지그교체 {jig_count}개 축 자주검사 완료"
    if jig_count:
        return False, "기종교체가 아니면 지그교체 하부핀 확인은 선택하지 않습니다."
    return True, f"초품 {first_count}개 축 자주검사 완료"


def count_frequent_check_axes(values: list[str]) -> int:
    """자주검사에서 실제 사용하는 축 수를 반환합니다.

    좌/우 그룹은 저장 전에 같은 개수로 검증되므로, 앞쪽 초품 확인 축 수를 기준으로 계산합니다.
    """
    return sum(1 for value in values[:6] if value == "OK")


def validate_frequent_check_capacity(lots: list[dict], stack: str, values: list[str]) -> tuple[bool, str]:
    """자주검사 축 수 x Stack 수가 실제 LOT 총 매수와 정확히 맞는지 확인합니다."""
    axis_count = count_frequent_check_axes(values)
    stack_count = int(stack)
    total_qty = sum(int(lot.get("qty", "0")) for lot in lots)
    checked_qty = axis_count * stack_count

    if checked_qty != total_qty:
        diff = total_qty - checked_qty
        diff_message = f"{abs(diff)}매 {'부족합니다' if diff > 0 else '많습니다'}"
        return (
            False,
            "자주검사 수량과 LOT 총 매수가 맞지 않습니다.\n\n"
            f"LOT 총 매수: {total_qty}매\n"
            f"자주검사 계산: {axis_count}축 x {stack_count}Stack = {checked_qty}매\n\n"
            f"차이: {diff_message}\n\n"
            "LOT 매수, Stack 수, 자주검사 축 수를 확인해주세요.",
        )
    return True, f"자주검사 계산 OK: {axis_count}축 x {stack_count}Stack = {checked_qty}매 / LOT {total_qty}매"


class JiinDncManager:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1420x820")
        self.root.minsize(1180, 700)
        self.config = load_config()
        apply_theme(self.config.get("theme", "MES 블루"))
        self.root.configure(bg=APP_BG)
        self.is_running = False

        self.common_entries: dict[str, LabeledEntry] = {}
        self.lot1_entries: dict[str, LabeledEntry] = {}
        self.lot2_entries: dict[str, LabeledEntry] = {}
        self.normal_buttons: list[ttk.Button] = []
        self.status_labels: dict[str, tk.Label] = {}
        self.lot_status_labels: dict[str, tk.Label] = {}
        self.log_text: scrolledtext.ScrolledText | None = None
        self.frequent_check_values: list[str] = [""] * 12

        self.setup_style()
        self.create_layout()
        self.update_status_checks()

    def setup_style(self) -> None:
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TFrame", background=APP_BG)
        style.configure("Header.TFrame", background=SURFACE_BG)
        style.configure("Panel.TFrame", background=SURFACE_BG)
        style.configure("TLabel", background=APP_BG, foreground=TEXT_COLOR, font=("맑은 고딕", 10))
        style.configure("Header.TLabel", background=SURFACE_BG, foreground=TEXT_COLOR, font=("맑은 고딕", 18, "bold"))
        style.configure("Hint.TLabel", background=APP_BG, foreground=MUTED_TEXT, font=("맑은 고딕", 10))
        style.configure("Panel.TLabel", background=SURFACE_BG, foreground=TEXT_COLOR, font=("맑은 고딕", 11, "bold"))
        style.configure("TButton", font=("맑은 고딕", 10), padding=(12, 8), background=SURFACE_BG, foreground=TEXT_COLOR)
        style.map("TButton", background=[("active", PRIMARY_LIGHT)])
        style.configure("Primary.TButton", font=("맑은 고딕", 11, "bold"), padding=(16, 10), background=PRIMARY_LIGHT, foreground=PRIMARY)
        style.configure("Wide.TEntry", padding=(8, 5), fieldbackground=SURFACE_BG)
        style.configure("TCheckbutton", background=APP_BG, foreground=TEXT_COLOR, font=("맑은 고딕", 10))

    def create_layout(self) -> None:
        header = ttk.Frame(self.root, style="Header.TFrame", padding=(20, 12, 20, 12))
        header.pack(fill=tk.X, padx=12, pady=(12, 0))
        ttk.Label(header, text=APP_TITLE, style="Header.TLabel").pack(side=tk.LEFT)
        ttk.Label(header, text="KCC PKG DNC 작업 자동화", style="Header.TLabel", font=("맑은 고딕", 10)).pack(side=tk.RIGHT, pady=(8, 0))

        self.notebook = SimpleTabNotebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=12, pady=(10, 8))

        for tab_name in ["TLB", "심텍 SPS", "심텍 HDI"]:
            self.notebook.add(self.create_placeholder_tab(tab_name), tab_name)
        self.kcc_pkg_page = tk.Frame(self.notebook.page_area, bg=APP_BG)
        self.notebook.add(self.kcc_pkg_page, "KCC PKG")
        self.notebook.add(self.create_placeholder_tab("KCC HDI"), "KCC HDI")
        self.settings_page = tk.Frame(self.notebook.page_area, bg=APP_BG)
        self.notebook.add(self.settings_page, "설정")

        self.create_kcc_pkg_tab()
        self.create_settings_tab()
        self.notebook.select(3)

    def create_placeholder_tab(self, name: str) -> tk.Frame:
        page = tk.Frame(self.notebook.page_area, bg=APP_BG)
        tk.Label(page, text=f"{name}\n추후 개발 예정", bg=APP_BG, fg=MUTED_TEXT, font=("맑은 고딕", 22, "bold")).pack(expand=True)
        return page

    def create_kcc_pkg_tab(self) -> None:
        self.kcc_pkg_page.columnconfigure(0, weight=1)
        self.kcc_pkg_page.rowconfigure(2, weight=1)

        title_wrap = tk.Frame(self.kcc_pkg_page, bg=PRIMARY_LIGHT)
        title_wrap.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 8))
        title_wrap.columnconfigure(0, weight=1)
        title = tk.Label(
            title_wrap,
            text="KCC PKG 일반 DNC",
            bg=PRIMARY_LIGHT,
            fg=PRIMARY,
            font=("맑은 고딕", 14, "bold"),
            height=2,
        )
        title.grid(row=0, column=0, sticky="ew")
        title_buttons = tk.Frame(title_wrap, bg=PRIMARY_LIGHT)
        title_buttons.grid(row=0, column=1, sticky="e", padx=(8, 10))
        self.add_normal_button(title_buttons, "일반 DNC 실행", self.run_normal_dnc, "Primary.TButton").grid(row=0, column=0, padx=4, pady=4)
        self.add_normal_button(title_buttons, "입력 초기화", self.clear_normal_inputs).grid(row=0, column=1, padx=4, pady=4)

        title_legacy = tk.Label(
            self.kcc_pkg_page,
            text="KCC PKG 일반 DNC",
            bg=PRIMARY_LIGHT,
            fg=PRIMARY,
            font=("맑은 고딕", 14, "bold"),
            height=2,
        )

        common = self.create_panel(self.kcc_pkg_page, "공통 입력")
        common.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 8))
        common_widgets = [
            ("work_date", DateField(common, "작업일자")),
            ("shift_group", SegmentedField(common, "조", ["A", "B", "C"])),
            ("shift", SegmentedField(common, "근무", ["주간", "야간"])),
            ("worker", LabeledEntry(common, "작업자")),
        ]
        for index, (key, entry) in enumerate(common_widgets):
            entry.grid(row=1, column=index, sticky="ew", padx=8, pady=8)
            self.common_entries[key] = entry
            common.columnconfigure(index, weight=1)

        lots = ttk.Frame(self.kcc_pkg_page)
        lots.grid(row=2, column=0, sticky="nsew", padx=14, pady=0)
        lots.columnconfigure(0, weight=1)
        lots.columnconfigure(1, weight=1)

        lot1 = self.create_lot_panel(lots, "LOT 1 입력", self.lot1_entries, 1)
        lot1.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=0)
        lot2 = self.create_lot_panel(lots, "LOT 2 입력 (선택)", self.lot2_entries, 2)
        lot2.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=0)

        bottom = ttk.Frame(self.kcc_pkg_page)
        bottom.grid(row=3, column=0, sticky="ew", padx=14, pady=(8, 14))
        bottom.columnconfigure(0, weight=1)

        status_panel = tk.Frame(bottom, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        status_panel.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        status_panel.columnconfigure(1, weight=1)
        status_panel.columnconfigure(2, weight=0)
        tk.Label(status_panel, text="2LOT 조건 일치 확인", bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 11, "bold"), width=22, height=2).grid(row=0, column=0, sticky="nsw")
        match_label = tk.Label(status_panel, text="LOT 2 미사용", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 12, "bold"), anchor="w")
        match_label.grid(row=0, column=1, sticky="ew", padx=14)
        self.status_labels["lot_match"] = match_label
        self.add_normal_button(status_panel, "자주검사", self.open_frequent_check_popup, "Primary.TButton").grid(row=0, column=2, sticky="e", padx=(0, 12), pady=6)
        tk.Label(status_panel, text="DNC 진행 상태", bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 11, "bold"), width=22, height=2).grid(row=1, column=0, sticky="nsw")
        dnc_label = tk.Label(status_panel, text="대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 12, "bold"), anchor="w")
        dnc_label.grid(row=1, column=1, columnspan=2, sticky="ew", padx=14)
        self.status_labels["dnc"] = dnc_label

        log_panel = tk.Frame(bottom, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        log_panel.grid(row=1, column=0, sticky="ew", padx=(0, 10), pady=(8, 0))
        log_panel.columnconfigure(0, weight=1)
        tk.Label(log_panel, text="DNC 작업 로그", bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 10, "bold"), height=1).grid(row=0, column=0, sticky="ew")
        self.log_text = scrolledtext.ScrolledText(
            log_panel,
            height=5,
            wrap=tk.WORD,
            state="disabled",
            bg=SURFACE_BG,
            fg=TEXT_COLOR,
            font=("맑은 고딕", 10),
            relief=tk.FLAT,
            padx=10,
            pady=8,
        )
        self.log_text.grid(row=1, column=0, sticky="ew")

        button_panel = ttk.Frame(bottom)
        button_panel.grid(row=0, column=1, rowspan=2, sticky="ne")
        self.add_normal_button(button_panel, "신규 모델 검증 DNC", self.open_new_model_popup).grid(row=0, column=0, padx=4, pady=4)
        self.add_normal_button(button_panel, "작업일보 열기", self.open_log_excel_from_ui).grid(row=0, column=1, padx=4, pady=4)
        self.add_normal_button(button_panel, "조건 마스터 관리", self.open_condition_master_popup).grid(row=1, column=0, columnspan=2, sticky="ew", padx=4, pady=4)

        for entry in list(self.lot1_entries.values()) + list(self.lot2_entries.values()):
            entry.var.trace_add("write", lambda *_args: self.update_status_checks())

    def create_panel(self, parent, title: str) -> tk.Frame:
        panel = tk.Frame(parent, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        tk.Label(panel, text=title, bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 10, "bold"), height=2).grid(row=0, column=0, columnspan=8, sticky="ew")
        return panel

    def create_lot_panel(self, parent, title: str, target: dict[str, LabeledEntry], lot_number: int) -> tk.Frame:
        panel = self.create_panel(parent, title)
        fields = [
            ("step", "STEP"),
            ("round", "차수"),
            ("manage_no", "관리번호"),
            ("lot_no", "LOT No"),
            ("qty", "매수"),
            ("process_code", "공정코드"),
            ("condition", "작업조건"),
            ("jig", "지그"),
        ]
        for index, (key, label) in enumerate(fields):
            row = index // 2 + 1
            col = index % 2
            entry = RoundField(panel, label) if key == "round" else LabeledEntry(panel, label, width=24)
            entry.grid(row=row, column=col, sticky="ew", padx=10, pady=8)
            panel.columnconfigure(col, weight=1)
            target[key] = entry
            if key in {"condition", "jig"}:
                entry.entry.configure(state="readonly")
        load_button = ttk.Button(
            panel,
            text="작업조건 / 지그 불러오기",
            command=lambda: self.load_condition_jig_for_lot(lot_number),
            style="Primary.TButton",
        )
        load_button.grid(row=5, column=0, columnspan=2, sticky="ew", padx=10, pady=(8, 6))

        status = tk.Frame(panel, bg=SURFACE_BG)
        status.grid(row=6, column=0, columnspan=2, sticky="ew", padx=10, pady=(0, 12))
        status.columnconfigure(1, weight=1)
        status.columnconfigure(3, weight=1)
        tk.Label(status, text="MES Core", bg=SURFACE_BG, fg=TEXT_COLOR, font=("맑은 고딕", 10)).grid(row=0, column=0, sticky="e", padx=(0, 6))
        mes_label = tk.Label(status, text="대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 10, "bold"), width=14)
        mes_label.grid(row=0, column=1, sticky="w")
        tk.Label(status, text="조건 적용", bg=SURFACE_BG, fg=TEXT_COLOR, font=("맑은 고딕", 10)).grid(row=0, column=2, sticky="e", padx=(14, 6))
        condition_label = tk.Label(status, text="대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 10, "bold"), width=14)
        condition_label.grid(row=0, column=3, sticky="w")
        self.lot_status_labels[f"lot{lot_number}_mes"] = mes_label
        self.lot_status_labels[f"lot{lot_number}_condition"] = condition_label
        return panel

    def add_normal_button(self, parent, text: str, command, style: str = "TButton") -> ttk.Button:
        button = ttk.Button(parent, text=text, command=command, style=style, width=18)
        self.normal_buttons.append(button)
        return button

    def create_settings_tab(self) -> None:
        page = self.settings_page
        page.columnconfigure(0, weight=1)
        panel = self.create_panel(page, "설정")
        panel.grid(row=0, column=0, sticky="ew", padx=14, pady=14)
        panel.columnconfigure(1, weight=1)

        self.excel_var = tk.StringVar(value=self.config.get("excel_file", ""))
        self.source_var = tk.StringVar(value=self.config.get("source_dnc_folder", str(get_desktop_path() / "KCC_PKG")))
        self.transfer_var = tk.StringVar(value=self.config.get("transfer_dnc_folder", str(get_desktop_path() / "DNC")))
        self.delete_seconds_var = tk.StringVar(value=str(self.config.get("dnc_delete_seconds", DNC_DELETE_SECONDS)))
        self.clear_common_var = tk.BooleanVar(value=bool(self.config.get("clear_common_after_normal", False)))
        self.theme_var = tk.StringVar(value=self.config.get("theme", "MES 블루"))

        rows = [
            ("작업일보 경로", self.excel_var),
            ("원본 DNC 폴더", self.source_var),
            ("DNC 전송 폴더", self.transfer_var),
            ("삭제 대기 시간", self.delete_seconds_var),
        ]
        for row, (label, var) in enumerate(rows, start=1):
            ttk.Label(panel, text=label, background=SURFACE_BG, width=16).grid(row=row, column=0, sticky="e", padx=10, pady=8)
            ttk.Entry(panel, textvariable=var, style="Wide.TEntry").grid(row=row, column=1, sticky="ew", padx=8, pady=8)

        ttk.Button(panel, text="작업일보 Excel 선택", command=lambda: select_excel_file(self.root, self.config, self.excel_var), width=20).grid(row=1, column=2, padx=8, pady=8)
        ttk.Button(panel, text="원본 폴더 선택", command=lambda: self.select_folder_to_var(self.source_var), width=20).grid(row=2, column=2, padx=8, pady=8)
        ttk.Button(panel, text="전송 폴더 선택", command=lambda: self.select_folder_to_var(self.transfer_var), width=20).grid(row=3, column=2, padx=8, pady=8)
        ttk.Label(panel, text="화면 색상", background=SURFACE_BG, width=16).grid(row=5, column=0, sticky="e", padx=10, pady=8)
        ttk.Combobox(panel, textvariable=self.theme_var, values=list(THEMES.keys()), state="readonly").grid(row=5, column=1, sticky="ew", padx=8, pady=8)
        ttk.Label(panel, text="색상 변경은 저장 후 프로그램을 다시 켜면 적용됩니다.", background=SURFACE_BG, foreground=MUTED_TEXT).grid(row=6, column=1, sticky="w", padx=8, pady=4)
        ttk.Button(panel, text="조건 마스터 갱신", command=self.rebuild_condition_master, width=20).grid(row=6, column=2, padx=8, pady=8)
        ttk.Button(panel, text="설정 저장", command=self.save_settings_from_ui, style="Primary.TButton", width=20).grid(row=7, column=2, padx=8, pady=14)

    def select_folder_to_var(self, var: tk.StringVar) -> None:
        folder = filedialog.askdirectory(initialdir=var.get() or str(get_desktop_path()))
        if folder:
            var.set(folder)

    def save_settings_from_ui(self) -> None:
        ok, message = validate_positive_number(self.delete_seconds_var.get(), "삭제 대기 시간", required=True)
        if not ok:
            messagebox.showwarning("설정 확인", message)
            return
        self.config["excel_file"] = self.excel_var.get().strip()
        self.config["source_dnc_folder"] = self.source_var.get().strip()
        self.config["transfer_dnc_folder"] = self.transfer_var.get().strip()
        self.config["dnc_delete_seconds"] = int(self.delete_seconds_var.get().strip())
        self.config["clear_common_after_normal"] = self.clear_common_var.get()
        self.config["theme"] = self.theme_var.get().strip() or "MES 블루"
        save_config(self.config)
        messagebox.showinfo("저장 완료", "설정을 저장했습니다.\n\n화면 색상은 프로그램을 다시 켜면 적용됩니다.")

    def rebuild_condition_master(self) -> None:
        self.save_settings_from_ui_silent()
        try:
            count = rebuild_condition_master_from_log(self.config)
        except Exception as exc:
            messagebox.showerror("조건 마스터 갱신 실패", str(exc))
            return
        messagebox.showinfo("조건 마스터 갱신 완료", f"{count}개 조건을 저장했습니다.")

    def get_common_data(self) -> dict:
        return {key: entry.get() for key, entry in self.common_entries.items()}

    def get_lot_data(self, entries: dict[str, LabeledEntry]) -> dict:
        return {key: entry.get() for key, entry in entries.items()}

    def lot_has_any_value(self, lot: dict) -> bool:
        # LOT 2는 선택 입력이므로 차수 버튼만 눌린 상태는 사용으로 보지 않습니다.
        return any(value.strip() for key, value in lot.items() if key != "round")

    def set_status(self, key: str, text: str, ok: bool | None = None) -> None:
        color = MUTED_TEXT if ok is None else (OK_COLOR if ok else NG_COLOR)
        self.status_labels[key].configure(text=text, fg=color)
        if key == "dnc":
            self.append_log(text)

    def set_dnc_status(self, text: str) -> None:
        self.root.after(0, lambda: self.set_status("dnc", text, None))

    def append_log(self, text: str) -> None:
        if self.log_text is None:
            return
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert(tk.END, f"[{timestamp}] {text}\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state="disabled")

    def update_status_checks(self) -> None:
        lot1 = self.get_lot_data(self.lot1_entries)
        lot2 = self.get_lot_data(self.lot2_entries)
        lot2_used = self.lot_has_any_value(lot2)
        self.update_lot_detail_status("lot1", lot1, waiting_when_empty=True)
        if lot2_used:
            self.update_lot_detail_status("lot2", lot2, waiting_when_empty=True)
            lot_match, lot_match_message = get_lot_match_message(lot1, lot2)
            if lot_match:
                self.set_status("lot_match", lot_match_message, True)
            else:
                self.set_status("lot_match", f"NG - {lot_match_message}", False)
        else:
            self.set_lot_status("lot2_mes", "대기중", None)
            self.set_lot_status("lot2_condition", "대기중", None)
            self.set_status("lot_match", "LOT 2 미사용", None)

    def update_lot_detail_status(self, prefix: str, lot: dict, waiting_when_empty: bool) -> None:
        lot_no = lot.get("lot_no", "").strip()
        process_code = lot.get("process_code", "").strip()
        condition = lot.get("condition", "").strip()
        jig = lot.get("jig", "").strip()

        if waiting_when_empty and not lot_no and not process_code:
            self.set_lot_status(f"{prefix}_mes", "대기중", None)
        else:
            mes_ok, mes_message = get_mes_core_message(lot_no, process_code)
            if mes_ok:
                self.set_lot_status(f"{prefix}_mes", "OK", True)
            else:
                self.set_lot_status(f"{prefix}_mes", "NG", False)

        if waiting_when_empty and not condition and not jig:
            self.set_lot_status(f"{prefix}_condition", "대기중", None)
        else:
            condition_ok, condition_message = get_single_condition_message(lot)
            self.set_lot_status(f"{prefix}_condition", "OK" if condition_ok else "NG", condition_ok)

    def set_lot_status(self, key: str, text: str, ok: bool | None = None) -> None:
        if key not in self.lot_status_labels:
            return
        color = MUTED_TEXT if ok is None else (OK_COLOR if ok else NG_COLOR)
        self.lot_status_labels[key].configure(text=text, fg=color)

    def set_running(self, running: bool) -> None:
        self.is_running = running
        state = "disabled" if running else "normal"
        for button in self.normal_buttons:
            button.configure(state=state)

    def validate_condition_file(self, condition_name: str) -> Path | None:
        self.set_dnc_status("조건 파일 검색중")
        matches = search_condition_file(condition_name, Path(self.config["source_dnc_folder"]))
        if len(matches) == 0:
            messagebox.showerror("조건 파일 없음", f"복사할 조건 파일이 없습니다.\n조건명: {condition_name}\n관리자 확인 필요.")
            return None
        if len(matches) >= 2:
            messagebox.showerror(
                "동일 프로그램 차단",
                f"동일 프로그램 확인으로 차단!\n관리자 확인 필요!\n조건명: {condition_name}\n검색 수량: {len(matches)}개",
            )
            return None
        return matches[0]

    def load_condition_jig_for_lot(self, lot_number: int) -> bool:
        """작업일보 이력에서 선택 LOT의 작업조건/지그를 불러와 화면에 채웁니다."""
        self.save_settings_from_ui_silent()
        entries = self.lot1_entries if lot_number == 1 else self.lot2_entries
        lot = self.get_lot_data(entries)
        if not (lot.get("manage_no") or lot.get("lot_no")):
            messagebox.showwarning("이력 조회", f"LOT {lot_number} 관리번호 또는 LOT No를 입력한 뒤 불러오세요.")
            return False
        try:
            condition, jig, source = lookup_condition_jig_from_history(self.config, lot)
        except Exception as exc:
            messagebox.showerror("이력 조회 실패", str(exc))
            return False
        if not condition or not jig:
            messagebox.showwarning(
                "이력 없음",
                f"작업일보 이력에서 작업조건/지그를 찾을 수 없습니다.\nLOT {lot_number} 관리번호: {lot.get('manage_no')}\nLOT No: {lot.get('lot_no')}",
            )
            return False
        entries["condition"].set(condition)
        entries["jig"].set(jig)
        refreshed_lot = self.get_lot_data(entries)
        upsert_condition_master(refreshed_lot, condition, jig, source)
        self.update_status_checks()
        self.set_status("dnc", f"LOT {lot_number} 조건/지그 불러옴: {source}", True)
        return True

    def run_normal_dnc(self) -> None:
        if self.is_running:
            messagebox.showwarning("진행 중", "DNC 실행중입니다.\n작업 완료 후 다시 실행해주세요.")
            return
        if not self.save_settings_from_ui_silent():
            return
        if not ensure_excel_file_selected(self.root, self.config, self.excel_var if hasattr(self, "excel_var") else None):
            self.set_status("dnc", "작업일보 선택 취소", False)
            return
        self.set_status("dnc", "입력값 확인중", None)
        common = self.get_common_data()
        lot1 = self.get_lot_data(self.lot1_entries)
        lot2_data = self.get_lot_data(self.lot2_entries)
        lot2 = lot2_data if self.lot_has_any_value(lot2_data) else None

        if not lot1.get("condition") or not lot1.get("jig"):
            if not self.load_condition_jig_for_lot(1):
                return
            lot1 = self.get_lot_data(self.lot1_entries)
        if lot2 and (not lot2.get("condition") or not lot2.get("jig")):
            if not self.load_condition_jig_for_lot(2):
                return
            lot2 = self.get_lot_data(self.lot2_entries)
        ok, errors = validate_normal_dnc(common, lot1, lot2)
        if not ok:
            messagebox.showwarning("입력값 확인", "\n".join(errors))
            self.set_status("dnc", "입력값 NG", False)
            return
        if not messagebox.askyesno("작업일보 기록", "작업일보를 기록하시겠습니까?"):
            self.set_status("dnc", "취소", None)
            return
        model_change = messagebox.askyesno("기종교체 확인", "기종교체 입니까?")
        messagebox.showinfo(
            "자주검사 필요",
            "초품 4Point 확인은 매 DNC마다 입력해야 합니다.\n\n"
            + (
                "기종교체이므로 지그교체 하부핀 3개 확인도 같이 입력해야 합니다.\n"
                "초품 확인과 지그교체 확인의 축 개수는 같아야 합니다."
                if model_change
                else "기종교체가 아니므로 지그교체 하부핀 확인은 입력하지 않습니다."
            ),
        )
        self.frequent_check_values = [""] * 12
        if not self.open_frequent_check_popup(require_jig_check=model_change) or not self.has_frequent_check_completed(require_jig_check=model_change):
            self.set_status("dnc", "자주검사 미완료", False)
            return
        stack = simpledialog.askstring("Stack 수 입력", "Stack 수를 입력 하세요.", parent=self.root)
        ok, message = validate_positive_number(stack or "", "Stack 수", required=True)
        if not ok:
            messagebox.showwarning("Stack 수 확인", message)
            return
        lots = [lot1] + ([lot2] if lot2 else [])
        ok, message = validate_frequent_check_capacity(lots, stack, self.frequent_check_values)
        if not ok:
            messagebox.showwarning("자주검사 수량 확인", message)
            self.set_status("dnc", "자주검사 수량 NG", False)
            return
        self.set_status("dnc", message, True)
        condition_file = self.validate_condition_file(lot1["condition"])
        if not condition_file:
            self.set_status("dnc", "조건 파일 NG", False)
            return

        self.set_running(True)
        frequent_check = self.frequent_check_values[:]
        threading.Thread(target=self.normal_worker, args=(common, lots, stack, model_change, condition_file, frequent_check), daemon=True).start()

    def normal_worker(self, common: dict, lots: list[dict], stack: str, model_change: bool, condition_file: Path, frequent_check: list[str] | None) -> None:
        try:
            self.set_dnc_status("작업일보 저장중")
            _path, rows = save_normal_dnc_log(self.config, common, lots, stack, model_change, frequent_check)
            self.set_dnc_status("작업일보 저장 완료")
            delete_existing_dnc_txt(Path(self.config["transfer_dnc_folder"]))
            copied_file = copy_dnc_file(condition_file, Path(self.config["transfer_dnc_folder"]))
            self.set_dnc_status("DNC 파일 복사 완료")
            delete_after_delay(copied_file, int(self.config["dnc_delete_seconds"]), self.set_dnc_status)
            self.root.after(0, lambda: self.finish_normal_dnc(rows))
        except Exception as exc:
            self.root.after(0, lambda error=exc: self.handle_run_error(error))

    def finish_normal_dnc(self, rows: list[int]) -> None:
        try:
            burr_ok = messagebox.askyesno("Burr 확인", "4면 Burr 이상 없습니까?")
            update_normal_burr_result(self.config, rows, burr_ok)
            self.set_status("dnc", "DNC 완료", True)
            self.clear_normal_inputs(after_done=True)
        except Exception as exc:
            self.handle_run_error(exc)
        finally:
            self.set_running(False)

    def clear_normal_inputs(self, after_done: bool = False) -> None:
        for entry in self.lot1_entries.values():
            entry.clear()
        for entry in self.lot2_entries.values():
            entry.clear()
        self.frequent_check_values = [""] * 12
        self.update_status_checks()
        self.set_status("dnc", "대기중", None)

    def open_new_model_popup(self) -> None:
        open_new_model_popup(self)

    def open_log_excel_from_ui(self) -> None:
        self.save_settings_from_ui_silent()
        if ensure_excel_file_selected(self.root, self.config, self.excel_var if hasattr(self, "excel_var") else None):
            open_log_excel(self.config)

    def open_condition_master_popup(self) -> None:
        ConditionMasterPopup(self)

    def open_frequent_check_popup(self, require_jig_check: bool = False) -> bool:
        popup = FrequentCheckPopup(self, require_jig_check=require_jig_check)
        self.root.wait_window(popup.window)
        return popup.saved

    def has_frequent_check_completed(self, require_jig_check: bool = False) -> bool:
        return validate_frequent_check_values(self.frequent_check_values, require_jig_check=require_jig_check)[0]

    def save_settings_from_ui_silent(self) -> bool:
        if hasattr(self, "excel_var"):
            ok, message = validate_positive_number(self.delete_seconds_var.get(), "삭제 대기 시간", required=True)
            if not ok:
                messagebox.showwarning("설정 확인", message)
                return False
            excel_path = self.excel_var.get().strip()
            if excel_path:
                self.config["excel_file"] = excel_path
            elif self.config.get("excel_file", ""):
                self.excel_var.set(self.config["excel_file"])
            self.config["source_dnc_folder"] = self.source_var.get().strip()
            self.config["transfer_dnc_folder"] = self.transfer_var.get().strip()
            self.config["dnc_delete_seconds"] = int(self.delete_seconds_var.get().strip())
            self.config["clear_common_after_normal"] = self.clear_common_var.get()
            self.config["theme"] = self.theme_var.get().strip() or self.config.get("theme", "MES 블루")
            save_config(self.config)
        return True

    def handle_run_error(self, exc: Exception) -> None:
        messagebox.showerror("오류", str(exc))
        self.set_status("dnc", "오류", False)
        self.set_running(False)


class FrequentCheckPopup:
    """Q:AB에 기록할 자주검사 값을 클릭으로 입력하는 창입니다."""

    LABELS = [
        "1축",
        "2축",
        "3축",
        "4축",
        "5축",
        "6축",
        "1축",
        "2축",
        "3축",
        "4축",
        "5축",
        "6축",
    ]
    FIRST_OK_COLOR = "#0ea5e9"
    JIG_OK_COLOR = "#10b981"

    def __init__(self, app: JiinDncManager, require_jig_check: bool = False):
        self.app = app
        self.require_jig_check = require_jig_check
        self.saved = False
        self.values = app.frequent_check_values[:]
        if not self.require_jig_check:
            self.values[6:] = [""] * 6
        self.buttons: list[tk.Button] = []
        self.window = tk.Toplevel(app.root)
        self.window.title("자주검사 입력")
        self.window.geometry("860x380")
        self.window.configure(bg=APP_BG)
        self.window.resizable(False, False)
        self.create_ui()

    def create_ui(self) -> None:
        title = tk.Label(
            self.window,
            text="자주검사 입력",
            bg=PRIMARY_LIGHT,
            fg=PRIMARY,
            font=("맑은 고딕", 14, "bold"),
            height=2,
        )
        title.pack(fill=tk.X, padx=14, pady=(14, 8))

        body = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        body.pack(fill=tk.BOTH, expand=True, padx=14, pady=8)

        tk.Label(body, text="초품 4 Point 확인 (매 DNC)", bg="#d99a9a", fg="#111827", font=("맑은 고딕", 10, "bold"), height=2).grid(row=0, column=0, columnspan=6, sticky="ew", padx=(0, 8))
        separator = tk.Frame(body, bg="#6b7280", width=3)
        separator.grid(row=0, column=6, rowspan=2, sticky="ns", padx=8, pady=0)
        jig_header_bg = "#d99a9a" if self.require_jig_check else "#e5e7eb"
        jig_header_text = "지그 교체시 하부면 3개 확인" if self.require_jig_check else "지그 교체시 하부면 3개 확인 (기종교체시만)"
        tk.Label(body, text=jig_header_text, bg=jig_header_bg, fg="#111827", font=("맑은 고딕", 10, "bold"), height=2).grid(row=0, column=7, columnspan=6, sticky="ew", padx=(8, 0))

        for index, label in enumerate(self.LABELS):
            grid_column = index if index < 6 else index + 1
            disabled_jig = index >= 6 and not self.require_jig_check
            button = tk.Button(
                body,
                text=f"{label}\n{'미사용' if disabled_jig else ('OK' if self.values[index] == 'OK' else '클릭')}",
                command=lambda i=index: self.toggle(i),
                bg="#f3f4f6" if disabled_jig else (self.get_ok_color(index) if self.values[index] == "OK" else SURFACE_BG),
                fg=MUTED_TEXT if disabled_jig else ("#ffffff" if self.values[index] == "OK" else TEXT_COLOR),
                relief=tk.SOLID,
                bd=1,
                width=8,
                height=3,
                cursor="arrow" if disabled_jig else "hand2",
                font=("맑은 고딕", 9, "bold"),
                highlightthickness=1,
                highlightbackground="#cbd5e1",
                state=tk.DISABLED if disabled_jig else tk.NORMAL,
            )
            button.grid(row=1, column=grid_column, sticky="nsew", padx=2, pady=8)
            body.columnconfigure(grid_column, weight=1)
            self.buttons.append(button)

        bottom = ttk.Frame(self.window, padding=(14, 4, 14, 14))
        bottom.pack(fill=tk.X)
        ttk.Button(bottom, text="전체 OK", command=self.select_all, style="Primary.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(bottom, text="초기화", command=self.clear).pack(side=tk.LEFT)
        ttk.Button(bottom, text="저장", command=self.save, style="Primary.TButton").pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(bottom, text="취소", command=self.window.destroy).pack(side=tk.RIGHT)

    def toggle(self, index: int) -> None:
        if index >= 6 and not self.require_jig_check:
            return
        self.values[index] = "" if self.values[index] == "OK" else "OK"
        self.refresh_button(index)

    def refresh_button(self, index: int) -> None:
        if index >= 6 and not self.require_jig_check:
            self.buttons[index].configure(text=f"{self.LABELS[index]}\n미사용", bg="#f3f4f6", fg=MUTED_TEXT)
            return
        ok = self.values[index] == "OK"
        self.buttons[index].configure(
            text=f"{self.LABELS[index]}\n{'OK' if ok else '클릭'}",
            bg=self.get_ok_color(index) if ok else SURFACE_BG,
            fg="#ffffff" if ok else TEXT_COLOR,
        )

    def get_ok_color(self, index: int) -> str:
        return self.FIRST_OK_COLOR if index < 6 else self.JIG_OK_COLOR

    def select_all(self) -> None:
        self.values = ["OK"] * (12 if self.require_jig_check else 6) + ([] if self.require_jig_check else [""] * 6)
        for index in range(12):
            self.refresh_button(index)

    def clear(self) -> None:
        self.values = [""] * 12
        for index in range(12):
            self.refresh_button(index)

    def save(self) -> None:
        if not self.require_jig_check:
            self.values[6:] = [""] * 6
        ok, message = validate_frequent_check_values(self.values, require_jig_check=self.require_jig_check)
        if not ok:
            messagebox.showwarning("자주검사 확인", message, parent=self.window)
            return
        self.app.frequent_check_values = self.values[:]
        self.app.set_status("dnc", message, True)
        self.saved = True
        self.window.destroy()


class ConditionMasterPopup:
    """작업일보에서 추출한 조건 마스터를 보고 수정하는 창입니다."""

    def __init__(self, app: JiinDncManager):
        self.app = app
        self.records = load_condition_master()
        self.window = tk.Toplevel(app.root)
        self.window.title("KCC PKG 조건 마스터 관리")
        self.window.geometry("980x620")
        self.window.configure(bg=APP_BG)
        self.create_ui()
        self.refresh_tree()

    def create_ui(self) -> None:
        top = ttk.Frame(self.window, padding=(12, 12, 12, 8))
        top.pack(fill=tk.X)
        ttk.Button(top, text="작업일보 이력으로 갱신", command=self.rebuild_from_log, style="Primary.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(top, text="선택 수정 저장", command=self.save_selected_edit).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(top, text="닫기", command=self.window.destroy).pack(side=tk.RIGHT)

        body = ttk.Frame(self.window, padding=(12, 0, 12, 8))
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        columns = ("step", "round", "manage_no", "process_code", "condition", "jig", "source")
        self.tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="browse")
        headings = {
            "step": "STEP",
            "round": "차수",
            "manage_no": "관리번호",
            "process_code": "공정코드",
            "condition": "작업조건",
            "jig": "지그",
            "source": "출처",
        }
        widths = {
            "step": 80,
            "round": 70,
            "manage_no": 150,
            "process_code": 110,
            "condition": 220,
            "jig": 140,
            "source": 140,
        }
        for column in columns:
            self.tree.heading(column, text=headings[column])
            self.tree.column(column, width=widths[column], anchor="w")
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        scroll = ttk.Scrollbar(body, orient=tk.VERTICAL, command=self.tree.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scroll.set)

        edit = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        edit.pack(fill=tk.X, padx=12, pady=(0, 12))
        self.edit_vars = {
            "step": tk.StringVar(),
            "round": tk.StringVar(),
            "manage_no": tk.StringVar(),
            "process_code": tk.StringVar(),
            "condition": tk.StringVar(),
            "jig": tk.StringVar(),
        }
        edit_fields = [
            ("step", "STEP"),
            ("round", "차수"),
            ("manage_no", "관리번호"),
            ("process_code", "공정코드"),
            ("condition", "작업조건"),
            ("jig", "지그"),
        ]
        for index, (key, label) in enumerate(edit_fields):
            ttk.Label(edit, text=label, background=SURFACE_BG).grid(row=index // 3, column=(index % 3) * 2, sticky="e", padx=(10, 4), pady=8)
            ttk.Entry(edit, textvariable=self.edit_vars[key], style="Wide.TEntry", width=24).grid(row=index // 3, column=(index % 3) * 2 + 1, sticky="ew", padx=(0, 10), pady=8)
            edit.columnconfigure((index % 3) * 2 + 1, weight=1)

    def refresh_tree(self) -> None:
        self.tree.delete(*self.tree.get_children())
        for index, record in enumerate(self.records):
            self.tree.insert(
                "",
                "end",
                iid=str(index),
                values=(
                    record.get("step", ""),
                    record.get("round", ""),
                    record.get("manage_no", ""),
                    record.get("process_code", ""),
                    record.get("condition", ""),
                    record.get("jig", ""),
                    record.get("source", ""),
                ),
            )

    def on_select(self, _event=None) -> None:
        selected = self.tree.selection()
        if not selected:
            return
        record = self.records[int(selected[0])]
        for key, var in self.edit_vars.items():
            var.set(record.get(key, ""))

    def save_selected_edit(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("선택 필요", "수정할 조건을 먼저 선택하세요.", parent=self.window)
            return
        index = int(selected[0])
        record = self.records[index]
        for key, var in self.edit_vars.items():
            record[key] = var.get().strip()
        record["source"] = "사용자 수정"
        record["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_condition_master(self.records)
        self.refresh_tree()
        self.tree.selection_set(str(index))
        messagebox.showinfo("저장 완료", "조건 마스터를 수정했습니다.", parent=self.window)

    def rebuild_from_log(self) -> None:
        self.app.save_settings_from_ui_silent()
        try:
            count = rebuild_condition_master_from_log(self.app.config)
        except Exception as exc:
            messagebox.showerror("갱신 실패", str(exc), parent=self.window)
            return
        self.records = load_condition_master()
        self.refresh_tree()
        messagebox.showinfo(
            "갱신 완료",
            f"조건 마스터를 최신 이력으로 정리했습니다.\n\n"
            f"현재 보관 조건: {count}개\n"
            "작업일보에서 삭제된 모델도 마스터에서는 보존됩니다.",
            parent=self.window,
        )


class NewModelPopup:
    def __init__(self, app: JiinDncManager):
        self.app = app
        self.window = tk.Toplevel(app.root)
        self.window.title("KCC PKG 신규 모델 검증 DNC")
        self.window.geometry("800x560")
        self.window.configure(bg=APP_BG)
        self.entries: dict[str, LabeledEntry] = {}
        self.buttons: list[ttk.Button] = []
        self.selected_lot = tk.StringVar(value="lot1")
        self.lot_drafts = {
            "lot1": self.app.get_lot_data(self.app.lot1_entries),
            "lot2": self.app.get_lot_data(self.app.lot2_entries),
        }
        self.lot_drafts["lot1"]["qty"] = ""
        self.lot_drafts["lot2"]["qty"] = ""
        self.is_running = False
        self.create_ui()
        self.load_selected_lot()
        self.update_checks()

    def create_ui(self) -> None:
        title = tk.Label(self.window, text="KCC PKG 신규 모델 검증 DNC", bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 14, "bold"), height=2)
        title.pack(fill=tk.X, padx=14, pady=(14, 8))

        lot_select = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        lot_select.pack(fill=tk.X, padx=14, pady=(0, 8))
        tk.Label(lot_select, text="대상 LOT", bg=SURFACE_BG, fg=TEXT_COLOR, font=("맑은 고딕", 10, "bold")).pack(side=tk.LEFT, padx=(14, 8), pady=8)
        ttk.Radiobutton(lot_select, text="LOT 1 입력값 사용", value="lot1", variable=self.selected_lot, command=self.change_selected_lot).pack(side=tk.LEFT, padx=8, pady=8)
        ttk.Radiobutton(lot_select, text="LOT 2 입력값 사용", value="lot2", variable=self.selected_lot, command=self.change_selected_lot).pack(side=tk.LEFT, padx=8, pady=8)

        panel = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        panel.pack(fill=tk.BOTH, expand=True, padx=14, pady=8)
        fields = [
            ("step", "STEP"),
            ("round", "차수"),
            ("manage_no", "관리번호"),
            ("lot_no", "LOT No"),
            ("qty", "매수"),
            ("process_code", "공정코드"),
            ("condition", "작업조건"),
            ("jig", "지그"),
        ]
        for index, (key, label) in enumerate(fields):
            entry = RoundField(panel, label) if key == "round" else LabeledEntry(panel, label, width=26)
            entry.grid(row=index // 2, column=index % 2, sticky="ew", padx=10, pady=8)
            panel.columnconfigure(index % 2, weight=1)
            self.entries[key] = entry
            if hasattr(entry, "var"):
                entry.var.trace_add("write", lambda *_args: self.update_checks())

        status = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground=BORDER_COLOR, bd=0)
        status.pack(fill=tk.X, padx=14, pady=8)
        self.mes_label = tk.Label(status, text="MES Core 일치화: 대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 10, "bold"))
        self.mes_label.pack(side=tk.LEFT, padx=16, pady=10)
        self.condition_label = tk.Label(status, text="조건 적용 확인: 대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 10, "bold"))
        self.condition_label.pack(side=tk.LEFT, padx=16, pady=10)
        self.dnc_label = tk.Label(status, text="DNC 진행 상태: 대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 10, "bold"))
        self.dnc_label.pack(side=tk.LEFT, padx=16, pady=10)

        buttons = ttk.Frame(self.window)
        buttons.pack(fill=tk.X, padx=14, pady=(4, 14))
        self.add_button(buttons, "신규 모델 DNC 실행", self.run_new_model_dnc, "Primary.TButton").pack(side=tk.LEFT, padx=4)
        self.add_button(buttons, "입력 초기화", self.clear_inputs).pack(side=tk.LEFT, padx=4)
        self.add_button(buttons, "닫기", self.window.destroy).pack(side=tk.RIGHT, padx=4)

    def add_button(self, parent, text, command, style="TButton") -> ttk.Button:
        button = ttk.Button(parent, text=text, command=command, style=style, width=20)
        self.buttons.append(button)
        return button

    def save_current_lot_draft(self) -> None:
        if not self.entries:
            return
        self.lot_drafts[self.selected_lot.get()] = {
            key: self.entries[key].get()
            for key in ["step", "round", "manage_no", "lot_no", "qty", "process_code", "condition", "jig"]
        }

    def load_selected_lot(self) -> None:
        draft = self.lot_drafts.get(self.selected_lot.get(), {})
        for key, entry in self.entries.items():
            entry.set(draft.get(key, ""))
        self.update_checks()

    def change_selected_lot(self) -> None:
        previous = "lot2" if self.selected_lot.get() == "lot1" else "lot1"
        if self.entries:
            self.lot_drafts[previous] = {
                key: self.entries[key].get()
                for key in ["step", "round", "manage_no", "lot_no", "qty", "process_code", "condition", "jig"]
            }
        self.load_selected_lot()

    def get_data(self) -> tuple[dict, dict]:
        self.save_current_lot_draft()
        common = self.app.get_common_data()
        lot = dict(self.lot_drafts[self.selected_lot.get()])
        return common, lot

    def update_checks(self) -> None:
        _common, lot = self.get_data()
        lot_no = lot.get("lot_no", "").strip()
        process_code = lot.get("process_code", "").strip()
        condition = lot.get("condition", "").strip()
        jig = lot.get("jig", "").strip()

        if not lot_no and not process_code:
            self.mes_label.configure(text="MES Core 일치화: 대기중", fg=MUTED_TEXT)
        else:
            mes_ok, mes_message = get_mes_core_message(lot_no, process_code)
            if mes_ok:
                self.mes_label.configure(text="MES Core 일치화: OK", fg=OK_COLOR)
            else:
                self.mes_label.configure(text="MES Core 일치화: NG", fg=NG_COLOR)

        if not condition and not jig:
            self.condition_label.configure(text="조건 적용 확인: 대기중", fg=MUTED_TEXT)
        else:
            condition_ok, condition_message = get_single_condition_message(lot)
            self.condition_label.configure(
                text=f"조건 적용 확인: {'OK' if condition_ok else 'NG'}",
                fg=OK_COLOR if condition_ok else NG_COLOR,
            )

    def set_running(self, running: bool) -> None:
        self.is_running = running
        for button in self.buttons:
            button.configure(state="disabled" if running else "normal")

    def set_dnc_status(self, text: str) -> None:
        self.window.after(0, lambda: self.dnc_label.configure(text=f"DNC 진행 상태: {text}", fg=MUTED_TEXT))

    def run_new_model_dnc(self) -> None:
        run_new_model_dnc(self)

    def clear_inputs(self) -> None:
        for entry in self.entries.values():
            entry.clear()
        self.lot_drafts[self.selected_lot.get()] = {
            key: ""
            for key in ["step", "round", "manage_no", "lot_no", "qty", "process_code", "condition", "jig"]
        }
        self.update_checks()
        self.dnc_label.configure(text="DNC 진행 상태: 대기중", fg=MUTED_TEXT)

    def clear_after_done(self) -> None:
        for key in ["round", "qty", "condition", "jig"]:
            self.entries[key].clear()
        self.save_current_lot_draft()
        self.update_checks()


def open_new_model_popup(app: JiinDncManager) -> None:
    """KCC PKG 신규 모델 검증 DNC 팝업창을 엽니다."""
    NewModelPopup(app)


def run_new_model_dnc(popup: NewModelPopup) -> None:
    """신규 모델 DNC 실행 버튼 흐름입니다."""
    if popup.is_running or popup.app.is_running:
        messagebox.showwarning("진행 중", "DNC 실행중입니다.\n작업 완료 후 다시 실행해주세요.")
        return
    if not popup.app.save_settings_from_ui_silent():
        return
    if not ensure_excel_file_selected(popup.window, popup.app.config, popup.app.excel_var if hasattr(popup.app, "excel_var") else None):
        popup.dnc_label.configure(text="DNC 진행 상태: 작업일보 선택 취소", fg=NG_COLOR)
        return
    common, lot = popup.get_data()
    ok, errors = validate_new_model_dnc(common, lot)
    if not ok:
        messagebox.showwarning("입력값 확인", "\n".join(errors), parent=popup.window)
        popup.dnc_label.configure(text="DNC 진행 상태: 입력값 NG", fg=NG_COLOR)
        return
    leader_name = simpledialog.askstring("조장명 입력", "신규 모델 검증 조장님 성함을 기재하세요", parent=popup.window)
    if not leader_name or not leader_name.strip():
        popup.dnc_label.configure(text="DNC 진행 상태: 취소", fg=MUTED_TEXT)
        return
    condition_file = popup.app.validate_condition_file(lot["condition"])
    if not condition_file:
        popup.dnc_label.configure(text="DNC 진행 상태: 조건 파일 NG", fg=NG_COLOR)
        return
    popup.set_running(True)
    popup.app.set_running(True)
    threading.Thread(target=new_model_worker, args=(popup, common, lot, leader_name.strip(), condition_file), daemon=True).start()


def run_normal_dnc(app: JiinDncManager) -> None:
    """요청 함수명 보존용 래퍼입니다. 실제 일반 DNC 실행은 JiinDncManager.run_normal_dnc에서 처리합니다."""
    app.run_normal_dnc()


def new_model_worker(popup: NewModelPopup, common: dict, lot: dict, leader_name: str, condition_file: Path) -> None:
    """신규 모델 DNC 백그라운드 작업입니다."""
    try:
        popup.set_dnc_status("작업일보 저장중")
        _path, row = save_new_model_log(popup.app.config, common, lot, leader_name)
        popup.set_dnc_status("작업일보 저장 완료")
        delete_existing_dnc_txt(Path(popup.app.config["transfer_dnc_folder"]))
        copied_file = copy_dnc_file(condition_file, Path(popup.app.config["transfer_dnc_folder"]))
        popup.set_dnc_status("DNC 파일 복사 완료")
        delete_after_delay(copied_file, int(popup.app.config["dnc_delete_seconds"]), popup.set_dnc_status)
        popup.window.after(0, lambda: finish_new_model_dnc(popup, row, lot["condition"]))
    except Exception as exc:
        popup.window.after(0, lambda error=exc: handle_popup_error(popup, error))


def finish_new_model_dnc(popup: NewModelPopup, row: int, condition_name: str) -> None:
    """신규 모델 DNC 완료 후 초도품 확인 결과를 저장합니다."""
    try:
        first_article_ok = messagebox.askyesno("초도품 확인", "초도품 이상 없습니까?", parent=popup.window)
        update_new_model_result(popup.app.config, row, condition_name, first_article_ok)
        popup.dnc_label.configure(text="DNC 진행 상태: DNC 완료", fg=OK_COLOR)
        popup.clear_after_done()
    except Exception as exc:
        handle_popup_error(popup, exc)
    finally:
        popup.set_running(False)
        popup.app.set_running(False)


def handle_popup_error(popup: NewModelPopup, exc: Exception) -> None:
    messagebox.showerror("오류", str(exc), parent=popup.window)
    popup.dnc_label.configure(text="DNC 진행 상태: 오류", fg=NG_COLOR)
    popup.set_running(False)
    popup.app.set_running(False)


def main() -> None:
    root = tk.Tk()
    app = JiinDncManager(root)
    root.mainloop()


if __name__ == "__main__":
    main()
