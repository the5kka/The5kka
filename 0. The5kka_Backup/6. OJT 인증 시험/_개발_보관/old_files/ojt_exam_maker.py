from __future__ import annotations

import random
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
import tkinter as tk

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "openpyxl 라이브러리가 필요합니다.\n"
        "명령 프롬프트에서 다음 명령을 실행하세요: pip install openpyxl"
    ) from exc


APP_TITLE = "OJT 랜덤 시험지 생성기"
DEFAULT_SCORES = {"공통": 2.5, "객관식": 4.0, "주관식": 5.0}
QUESTION_TYPES = ("공통", "객관식", "주관식")


@dataclass
class Question:
    source_sheet: str
    source_no: object
    category: str
    exam_type: str
    question: str
    answer: str
    score: float


@dataclass
class QuestionBank:
    display_name: str
    sheet_name: str
    questions: list[Question]
    counts: dict[str, int]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+\n", "\n", text)
    return text.strip()


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", clean_text(value)).lower()


def find_header_columns(ws) -> tuple[int, dict[str, int]] | None:
    for row in range(1, min(ws.max_row, 12) + 1):
        headers = {
            normalize_header(ws.cell(row=row, column=col).value): col
            for col in range(1, ws.max_column + 1)
        }
        question_col = headers.get("문제")
        answer_col = headers.get("답안")
        exam_type_col = headers.get("문제유형") or headers.get("문제type")
        no_col = headers.get("no")
        if question_col and answer_col and exam_type_col and no_col:
            return row, {
                "no": no_col,
                "category": headers.get("유형", 0),
                "exam_type": exam_type_col,
                "question": question_col,
                "answer": answer_col,
                "score": headers.get("점수", 0),
            }
    return None


def score_for_type(exam_type: str, score_cell: object | None = None) -> float:
    if isinstance(score_cell, (int, float)):
        return float(score_cell)
    return DEFAULT_SCORES.get(exam_type, 0.0)


def load_question_banks(path: Path) -> list[QuestionBank]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    banks: list[QuestionBank] = []
    used_display_names: dict[str, int] = {}

    for ws in workbook.worksheets:
        if ws.title.strip() in {"시험 SETTING", "시험지", "답안지"}:
            continue

        header_info = find_header_columns(ws)
        if not header_info:
            continue

        header_row, cols = header_info
        questions: list[Question] = []

        for row in range(header_row + 1, ws.max_row + 1):
            question_text = clean_text(ws.cell(row=row, column=cols["question"]).value)
            source_no = ws.cell(row=row, column=cols["no"]).value
            if not question_text or source_no in (None, ""):
                continue

            exam_type = clean_text(ws.cell(row=row, column=cols["exam_type"]).value)
            if exam_type not in QUESTION_TYPES:
                continue

            category = clean_text(ws.cell(row=row, column=cols["category"]).value) if cols["category"] else ""
            answer = clean_text(ws.cell(row=row, column=cols["answer"]).value)
            score_cell = ws.cell(row=row, column=cols["score"]).value if cols["score"] else None

            questions.append(
                Question(
                    source_sheet=ws.title,
                    source_no=source_no,
                    category=category,
                    exam_type=exam_type,
                    question=question_text,
                    answer=answer,
                    score=score_for_type(exam_type, score_cell),
                )
            )

        if not questions:
            continue

        display = ws.title.strip()
        used_display_names[display] = used_display_names.get(display, 0) + 1
        if used_display_names[display] > 1:
            display = f"{display} ({used_display_names[display]})"

        counts = {kind: sum(1 for q in questions if q.exam_type == kind) for kind in QUESTION_TYPES}
        banks.append(QuestionBank(display, ws.title, questions, counts))

    banks.sort(key=lambda bank: bank.display_name)
    return banks


def parse_nonnegative_int(value: str, label: str) -> int:
    try:
        number = int(value)
    except ValueError as exc:
        raise ValueError(f"{label} 문항 수는 숫자로 입력해야 합니다.") from exc
    if number < 0:
        raise ValueError(f"{label} 문항 수는 0 이상이어야 합니다.")
    return number


def parse_positive_float(value: str, label: str) -> float:
    try:
        number = float(value)
    except ValueError as exc:
        raise ValueError(f"{label} 점수는 숫자로 입력해야 합니다.") from exc
    if number <= 0:
        raise ValueError(f"{label} 점수는 0보다 커야 합니다.")
    return number


def calculate_total(counts: dict[str, int], scores: dict[str, float]) -> float:
    return sum(counts[kind] * scores[kind] for kind in QUESTION_TYPES)


def select_questions(bank: QuestionBank, counts: dict[str, int], scores: dict[str, float]) -> list[Question]:
    selected: list[Question] = []
    for kind in QUESTION_TYPES:
        candidates = [q for q in bank.questions if q.exam_type == kind]
        requested = counts[kind]
        if requested > len(candidates):
            raise ValueError(f"{kind} 문제가 부족합니다. 요청 {requested}문항 / 보유 {len(candidates)}문항")
        picked = random.sample(candidates, requested)
        for question in picked:
            question.score = scores[kind]
        selected.extend(picked)
    random.shuffle(selected)
    return selected


def safe_filename(text: str) -> str:
    text = re.sub(r'[\\/:*?"<>|]+', "_", text.strip())
    text = re.sub(r"\s+", " ", text)
    return text or "OJT"


def style_cell_range(ws, cell_range: str, fill: str | None = None, bold: bool = False):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="맑은 고딕", size=10, bold=bold, color="FFFFFF" if fill else "000000")


def add_borders(ws, min_row: int, max_row: int, min_col: int, max_col: int):
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def estimate_row_height(text: str) -> int:
    lines = max(1, text.count("\n") + 1)
    length_lines = max(1, len(text) // 70)
    return min(180, max(32, (lines + length_lines) * 15))


def write_exam_workbook(
    output_path: Path,
    process_name: str,
    questions: list[Question],
    counts: dict[str, int],
    scores: dict[str, float],
    target_score: float,
    department: str = "",
    evaluator: str = "",
    employee_name: str = "",
) -> None:
    wb = Workbook()
    exam_ws = wb.active
    exam_ws.title = "시험지"
    answer_ws = wb.create_sheet("답안지")
    summary_ws = wb.create_sheet("생성정보")

    total_score = calculate_total(counts, scores)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    exam_ws.sheet_view.showGridLines = False
    exam_ws["A1"] = "OJT 교육 평가서"
    exam_ws["A1"].font = Font(name="맑은 고딕", size=16, bold=True)
    exam_ws.merge_cells("A1:E1")
    exam_ws["A2"] = "공정명"
    exam_ws["B2"] = process_name
    exam_ws["D2"] = "총점"
    exam_ws["E2"] = total_score
    exam_ws["A3"] = "부서명"
    exam_ws["B3"] = department
    exam_ws["D3"] = "평가자"
    exam_ws["E3"] = evaluator
    exam_ws["A4"] = "성명"
    exam_ws["B4"] = employee_name
    exam_ws["D4"] = "생성일"
    exam_ws["E4"] = generated_at
    for col_idx, value in enumerate(["No", "점수", "문제유형", "문제", "답안"], start=1):
        exam_ws.cell(row=6, column=col_idx, value=value)
    style_cell_range(exam_ws, "A6:E6", fill="1F4E78", bold=True)

    for idx, question in enumerate(questions, start=1):
        row = idx + 6
        exam_ws.cell(row=row, column=1, value=idx)
        exam_ws.cell(row=row, column=2, value=question.score)
        exam_ws.cell(row=row, column=3, value=question.exam_type)
        exam_ws.cell(row=row, column=4, value=question.question)
        exam_ws.cell(row=row, column=5, value="")
        exam_ws.row_dimensions[row].height = estimate_row_height(question.question)

    exam_ws.freeze_panes = "A7"
    exam_ws.column_dimensions["A"].width = 6
    exam_ws.column_dimensions["B"].width = 8
    exam_ws.column_dimensions["C"].width = 12
    exam_ws.column_dimensions["D"].width = 82
    exam_ws.column_dimensions["E"].width = 18
    add_borders(exam_ws, 2, len(questions) + 6, 1, 5)
    for row in range(2, 5):
        for col in (1, 4):
            exam_ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="D9EAF7")
            exam_ws.cell(row=row, column=col).font = Font(name="맑은 고딕", size=10, bold=True)

    answer_ws.sheet_view.showGridLines = False
    answer_ws["A1"] = "OJT 답안지"
    answer_ws["A1"].font = Font(name="맑은 고딕", size=16, bold=True)
    answer_ws.merge_cells("A1:F1")
    answer_ws["A2"] = "공정명"
    answer_ws["B2"] = process_name
    answer_ws["D2"] = "총점"
    answer_ws["E2"] = total_score
    for col_idx, value in enumerate(["No", "점수", "문제유형", "원본 No", "답안", "문제"], start=1):
        answer_ws.cell(row=4, column=col_idx, value=value)
    style_cell_range(answer_ws, "A4:F4", fill="8064A2", bold=True)

    for idx, question in enumerate(questions, start=1):
        row = idx + 4
        answer_ws.cell(row=row, column=1, value=idx)
        answer_ws.cell(row=row, column=2, value=question.score)
        answer_ws.cell(row=row, column=3, value=question.exam_type)
        answer_ws.cell(row=row, column=4, value=question.source_no)
        answer_ws.cell(row=row, column=5, value=question.answer)
        answer_ws.cell(row=row, column=6, value=question.question)
        answer_ws.row_dimensions[row].height = estimate_row_height(question.answer + "\n" + question.question)

    answer_ws.freeze_panes = "A5"
    for col, width in {"A": 6, "B": 8, "C": 12, "D": 10, "E": 42, "F": 70}.items():
        answer_ws.column_dimensions[col].width = width
    add_borders(answer_ws, 2, len(questions) + 4, 1, 6)

    summary_ws.sheet_view.showGridLines = False
    summary_rows = [
        ["항목", "값"],
        ["공정명", process_name],
        ["생성일", generated_at],
        ["목표 점수", target_score],
        ["생성 점수", total_score],
        ["총 문항 수", len(questions)],
        ["공통 문항/점수", f"{counts['공통']}문항 x {scores['공통']}점"],
        ["객관식 문항/점수", f"{counts['객관식']}문항 x {scores['객관식']}점"],
        ["주관식 문항/점수", f"{counts['주관식']}문항 x {scores['주관식']}점"],
    ]
    for row_idx, row_values in enumerate(summary_rows, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            summary_ws.cell(row=row_idx, column=col_idx, value=value)
    style_cell_range(summary_ws, "A1:B1", fill="548235", bold=True)
    add_borders(summary_ws, 1, len(summary_rows), 1, 2)
    summary_ws.column_dimensions["A"].width = 24
    summary_ws.column_dimensions["B"].width = 44

    for ws in (exam_ws, answer_ws, summary_ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.font = cell.font.copy(name="맑은 고딕")

    wb.save(output_path)


class ExamMakerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("760x560")
        self.minsize(720, 520)

        self.workbook_path = tk.StringVar(value=str(self.default_workbook_path()))
        self.output_dir = tk.StringVar(value=str(self.default_output_dir()))
        self.process_var = tk.StringVar()
        self.target_score = tk.StringVar(value="100")
        self.count_vars = {kind: tk.StringVar(value=str(default)) for kind, default in {"공통": 2, "객관식": 20, "주관식": 3}.items()}
        self.score_vars = {kind: tk.StringVar(value=str(DEFAULT_SCORES[kind]).rstrip("0").rstrip(".")) for kind in QUESTION_TYPES}
        self.department_var = tk.StringVar()
        self.evaluator_var = tk.StringVar()
        self.employee_var = tk.StringVar()
        self.total_var = tk.StringVar(value="100")
        self.status_var = tk.StringVar(value="문제은행 엑셀을 불러오세요.")
        self.banks: list[QuestionBank] = []

        self.configure(bg="#F5F7FA")
        self.create_widgets()

        if Path(self.workbook_path.get()).exists():
            self.load_workbook()

    @staticmethod
    def default_workbook_path() -> Path:
        desktop = Path.home() / "Desktop"
        candidate = desktop / "OJT 시험 문제.xlsm"
        return candidate if candidate.exists() else Path.cwd() / "OJT 시험 문제.xlsm"

    @staticmethod
    def default_output_dir() -> Path:
        desktop = Path.home() / "Desktop"
        return desktop if desktop.exists() else Path.cwd()

    def create_widgets(self):
        root = ttk.Frame(self, padding=16)
        root.pack(fill="both", expand=True)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TFrame", background="#F5F7FA")
        style.configure("TLabelframe", background="#F5F7FA")
        style.configure("TLabelframe.Label", background="#F5F7FA", font=("맑은 고딕", 10, "bold"))
        style.configure("TLabel", background="#F5F7FA", font=("맑은 고딕", 10))
        style.configure("TButton", font=("맑은 고딕", 10))
        style.configure("Accent.TButton", font=("맑은 고딕", 11, "bold"))

        file_frame = ttk.LabelFrame(root, text="문제은행")
        file_frame.pack(fill="x")
        ttk.Entry(file_frame, textvariable=self.workbook_path).pack(side="left", fill="x", expand=True, padx=8, pady=8)
        ttk.Button(file_frame, text="찾기", command=self.pick_workbook).pack(side="left", padx=4)
        ttk.Button(file_frame, text="불러오기", command=self.load_workbook).pack(side="left", padx=8)

        option_frame = ttk.LabelFrame(root, text="시험 설정")
        option_frame.pack(fill="x", pady=12)
        option_frame.columnconfigure(1, weight=1)
        ttk.Label(option_frame, text="공정").grid(row=0, column=0, sticky="w", padx=8, pady=8)
        self.process_combo = ttk.Combobox(option_frame, textvariable=self.process_var, state="readonly")
        self.process_combo.grid(row=0, column=1, sticky="ew", padx=8, pady=8)
        self.process_combo.bind("<<ComboboxSelected>>", lambda _: self.refresh_process_status())

        ttk.Label(option_frame, text="목표 점수").grid(row=0, column=2, sticky="e", padx=8)
        ttk.Entry(option_frame, textvariable=self.target_score, width=10).grid(row=0, column=3, padx=8)

        count_frame = ttk.Frame(option_frame)
        count_frame.grid(row=1, column=0, columnspan=4, sticky="ew", padx=8, pady=6)
        for idx, kind in enumerate(QUESTION_TYPES):
            ttk.Label(count_frame, text=f"{kind} 문항").grid(row=0, column=idx * 2, padx=(0, 4), pady=4)
            ttk.Entry(count_frame, textvariable=self.count_vars[kind], width=8).grid(row=0, column=idx * 2 + 1, padx=(0, 16), pady=4)
            ttk.Label(count_frame, text=f"{kind} 점수").grid(row=1, column=idx * 2, padx=(0, 4), pady=4)
            ttk.Entry(count_frame, textvariable=self.score_vars[kind], width=8).grid(row=1, column=idx * 2 + 1, padx=(0, 16), pady=4)

        for var in list(self.count_vars.values()) + list(self.score_vars.values()) + [self.target_score]:
            var.trace_add("write", lambda *_: self.update_total())

        meta_frame = ttk.LabelFrame(root, text="표기 정보")
        meta_frame.pack(fill="x")
        for col in range(6):
            meta_frame.columnconfigure(col, weight=1)
        ttk.Label(meta_frame, text="부서명").grid(row=0, column=0, padx=8, pady=8, sticky="w")
        ttk.Entry(meta_frame, textvariable=self.department_var).grid(row=0, column=1, padx=8, pady=8, sticky="ew")
        ttk.Label(meta_frame, text="평가자").grid(row=0, column=2, padx=8, pady=8, sticky="w")
        ttk.Entry(meta_frame, textvariable=self.evaluator_var).grid(row=0, column=3, padx=8, pady=8, sticky="ew")
        ttk.Label(meta_frame, text="성명").grid(row=0, column=4, padx=8, pady=8, sticky="w")
        ttk.Entry(meta_frame, textvariable=self.employee_var).grid(row=0, column=5, padx=8, pady=8, sticky="ew")

        output_frame = ttk.LabelFrame(root, text="저장 위치")
        output_frame.pack(fill="x", pady=12)
        ttk.Entry(output_frame, textvariable=self.output_dir).pack(side="left", fill="x", expand=True, padx=8, pady=8)
        ttk.Button(output_frame, text="찾기", command=self.pick_output_dir).pack(side="left", padx=8)

        action_frame = ttk.Frame(root)
        action_frame.pack(fill="x")
        self.total_label = ttk.Label(action_frame, textvariable=self.total_var, font=("맑은 고딕", 18, "bold"))
        self.total_label.pack(side="left", padx=4)
        ttk.Button(action_frame, text="시험지 생성", style="Accent.TButton", command=self.generate_exam).pack(side="right", padx=4)

        log_frame = ttk.LabelFrame(root, text="상태")
        log_frame.pack(fill="both", expand=True, pady=12)
        self.status_text = tk.Text(log_frame, height=8, wrap="word", font=("맑은 고딕", 10), bg="white")
        self.status_text.pack(fill="both", expand=True, padx=8, pady=8)
        self.log(self.status_var.get())

    def log(self, message: str):
        self.status_text.insert("end", f"{message}\n")
        self.status_text.see("end")
        self.status_var.set(message)

    def pick_workbook(self):
        path = filedialog.askopenfilename(
            title="문제은행 엑셀 선택",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if path:
            self.workbook_path.set(path)

    def pick_output_dir(self):
        path = filedialog.askdirectory(title="저장 위치 선택")
        if path:
            self.output_dir.set(path)

    def load_workbook(self):
        path = Path(self.workbook_path.get())
        if not path.exists():
            messagebox.showerror(APP_TITLE, "문제은행 엑셀 파일을 찾을 수 없습니다.")
            return
        try:
            self.banks = load_question_banks(path)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"문제은행을 불러오지 못했습니다.\n\n{exc}")
            return

        if not self.banks:
            messagebox.showwarning(APP_TITLE, "사용 가능한 문제은행 시트를 찾지 못했습니다.")
            return

        self.process_combo["values"] = [bank.display_name for bank in self.banks]
        self.process_var.set(self.banks[0].display_name)
        self.log(f"문제은행 로드 완료: {len(self.banks)}개 공정")
        self.refresh_process_status()

    def get_selected_bank(self) -> QuestionBank:
        selected = self.process_var.get()
        for bank in self.banks:
            if bank.display_name == selected:
                return bank
        raise ValueError("공정을 선택하세요.")

    def update_total(self):
        try:
            counts = {kind: parse_nonnegative_int(self.count_vars[kind].get(), kind) for kind in QUESTION_TYPES}
            scores = {kind: parse_positive_float(self.score_vars[kind].get(), kind) for kind in QUESTION_TYPES}
            total = calculate_total(counts, scores)
            self.total_var.set(f"현재 총점: {total:g}점")
            target = float(self.target_score.get())
            color = "#188038" if abs(total - target) < 0.0001 else "#C5221F"
            self.total_label.configure(foreground=color)
        except Exception:
            self.total_var.set("현재 총점: 입력 확인 필요")
            self.total_label.configure(foreground="#C5221F")

    def refresh_process_status(self):
        if not self.banks:
            return
        try:
            bank = self.get_selected_bank()
        except ValueError:
            return
        self.log(
            f"{bank.display_name}: 공통 {bank.counts['공통']} / "
            f"객관식 {bank.counts['객관식']} / 주관식 {bank.counts['주관식']} 보유"
        )
        self.update_total()

    def generate_exam(self):
        if not self.banks:
            messagebox.showwarning(APP_TITLE, "먼저 문제은행을 불러오세요.")
            return

        try:
            bank = self.get_selected_bank()
            counts = {kind: parse_nonnegative_int(self.count_vars[kind].get(), kind) for kind in QUESTION_TYPES}
            scores = {kind: parse_positive_float(self.score_vars[kind].get(), kind) for kind in QUESTION_TYPES}
            target = parse_positive_float(self.target_score.get(), "목표")
            total = calculate_total(counts, scores)
            if abs(total - target) >= 0.0001:
                raise ValueError(f"총점이 {target:g}점이 아닙니다. 현재 총점: {total:g}점")

            selected = select_questions(bank, counts, scores)
            output_dir = Path(self.output_dir.get())
            output_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"OJT_시험지_{safe_filename(bank.display_name)}_{timestamp}.xlsx"
            output_path = output_dir / filename

            write_exam_workbook(
                output_path=output_path,
                process_name=bank.display_name,
                questions=selected,
                counts=counts,
                scores=scores,
                target_score=target,
                department=self.department_var.get(),
                evaluator=self.evaluator_var.get(),
                employee_name=self.employee_var.get(),
            )
        except Exception as exc:
            messagebox.showerror(APP_TITLE, str(exc))
            self.log(f"생성 실패: {exc}")
            return

        self.log(f"생성 완료: {output_path}")
        messagebox.showinfo(APP_TITLE, f"시험지 생성 완료\n\n{output_path}")


def main():
    app = ExamMakerApp()
    app.mainloop()


if __name__ == "__main__":
    main()
