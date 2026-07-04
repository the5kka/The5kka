from __future__ import annotations

import random
import re
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


APP_TITLE = "OJT 랜덤 시험지 생성기"
QUESTION_TYPES = ("공통", "객관식", "주관식")
DEFAULT_COUNTS = {"공통": 2, "객관식": 20, "주관식": 3}
DEFAULT_SCORES = {"공통": 2.5, "객관식": 4.0, "주관식": 5.0}


@dataclass
class Question:
    source_sheet: str
    source_no: object
    category: str
    exam_type: str
    question: str
    answer: str
    score: float


@dataclass
class QuestionBank:
    display_name: str
    sheet_name: str
    questions: list[Question]
    counts: dict[str, int]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+\n", "\n", text)
    return text.strip()


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", clean_text(value)).lower()


def find_header_columns(ws) -> tuple[int, dict[str, int]] | None:
    for row in range(1, min(ws.max_row, 12) + 1):
        headers = {
            normalize_header(ws.cell(row=row, column=col).value): col
            for col in range(1, ws.max_column + 1)
        }
        if headers.get("문제") and headers.get("답안") and headers.get("문제유형") and headers.get("no"):
            return row, {
                "no": headers["no"],
                "category": headers.get("유형", 0),
                "exam_type": headers["문제유형"],
                "question": headers["문제"],
                "answer": headers["답안"],
                "score": headers.get("점수", 0),
            }
    return None


def score_for_type(exam_type: str, score_cell: object | None = None) -> float:
    if isinstance(score_cell, (int, float)):
        return float(score_cell)
    return DEFAULT_SCORES.get(exam_type, 0.0)


def load_question_banks(path: Path) -> list[QuestionBank]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    banks: list[QuestionBank] = []
    used_names: dict[str, int] = {}

    for ws in workbook.worksheets:
        if ws.title.strip() in {"시험 SETTING", "시험지", "답안지"}:
            continue
        header_info = find_header_columns(ws)
        if not header_info:
            continue
        header_row, cols = header_info
        questions: list[Question] = []

        for row in range(header_row + 1, ws.max_row + 1):
            question_text = clean_text(ws.cell(row=row, column=cols["question"]).value)
            source_no = ws.cell(row=row, column=cols["no"]).value
            if not question_text or source_no in (None, ""):
                continue
            exam_type = clean_text(ws.cell(row=row, column=cols["exam_type"]).value)
            if exam_type not in QUESTION_TYPES:
                continue
            category = clean_text(ws.cell(row=row, column=cols["category"]).value) if cols["category"] else ""
            answer = clean_text(ws.cell(row=row, column=cols["answer"]).value)
            score_cell = ws.cell(row=row, column=cols["score"]).value if cols["score"] else None
            questions.append(
                Question(ws.title, source_no, category, exam_type, question_text, answer, score_for_type(exam_type, score_cell))
            )

        if questions:
            display_name = ws.title.strip()
            used_names[display_name] = used_names.get(display_name, 0) + 1
            if used_names[display_name] > 1:
                display_name = f"{display_name} ({used_names[display_name]})"
            counts = {kind: sum(1 for q in questions if q.exam_type == kind) for kind in QUESTION_TYPES}
            banks.append(QuestionBank(display_name, ws.title, questions, counts))

    return sorted(banks, key=lambda bank: bank.display_name)


def calculate_total(counts: dict[str, int], scores: dict[str, float]) -> float:
    return sum(counts[kind] * scores[kind] for kind in QUESTION_TYPES)


def select_questions(bank: QuestionBank, counts: dict[str, int], scores: dict[str, float]) -> list[Question]:
    selected: list[Question] = []
    for kind in QUESTION_TYPES:
        candidates = [q for q in bank.questions if q.exam_type == kind]
        requested = counts[kind]
        if requested > len(candidates):
            raise ValueError(f"{kind} 문제가 부족합니다. 요청 {requested}문항 / 보유 {len(candidates)}문항")
        picked = random.sample(candidates, requested)
        for q in picked:
            q.score = scores[kind]
        selected.extend(picked)
    random.shuffle(selected)
    return selected


def safe_filename(text: str) -> str:
    text = re.sub(r'[\\/:*?"<>|]+', "_", text.strip())
    text = re.sub(r"\s+", " ", text)
    return text or "OJT"


def add_borders(ws, min_row: int, max_row: int, min_col: int, max_col: int):
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def estimate_row_height(text: str) -> int:
    lines = max(1, text.count("\n") + 1)
    length_lines = max(1, len(text) // 70)
    return min(180, max(32, (lines + length_lines) * 15))


def write_header(ws, row: int, values: list[str], fill: str):
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_exam_workbook(
    output_path: Path,
    process_name: str,
    questions: list[Question],
    counts: dict[str, int],
    scores: dict[str, float],
    target_score: float,
    template_path: Path | None = None,
):
    if template_path is not None and template_path.exists():
        write_template_workbook(output_path, template_path, process_name, questions, counts)
        return

    wb = Workbook()
    exam_ws = wb.active
    exam_ws.title = "시험지"
    answer_ws = wb.create_sheet("답안지")
    summary_ws = wb.create_sheet("생성정보")

    total_score = calculate_total(counts, scores)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    exam_ws.sheet_view.showGridLines = False
    exam_ws["A1"] = "OJT 교육 평가서"
    exam_ws["A1"].font = Font(name="맑은 고딕", size=16, bold=True)
    exam_ws.merge_cells("A1:E1")
    exam_ws["A2"], exam_ws["B2"], exam_ws["D2"], exam_ws["E2"] = "공정명", process_name, "총점", total_score
    exam_ws["A3"], exam_ws["B3"], exam_ws["D3"], exam_ws["E3"] = "생성일", generated_at, "목표점수", target_score
    write_header(exam_ws, 5, ["No", "점수", "문제유형", "문제", "답안"], "1F4E78")

    for idx, q in enumerate(questions, start=1):
        row = idx + 5
        exam_ws.cell(row, 1, idx)
        exam_ws.cell(row, 2, q.score)
        exam_ws.cell(row, 3, q.exam_type)
        exam_ws.cell(row, 4, q.question)
        exam_ws.cell(row, 5, "")
        exam_ws.row_dimensions[row].height = estimate_row_height(q.question)
    for col, width in {"A": 6, "B": 8, "C": 12, "D": 82, "E": 18}.items():
        exam_ws.column_dimensions[col].width = width
    add_borders(exam_ws, 2, len(questions) + 5, 1, 5)
    exam_ws.freeze_panes = "A6"

    answer_ws.sheet_view.showGridLines = False
    answer_ws["A1"] = "OJT 답안지"
    answer_ws["A1"].font = Font(name="맑은 고딕", size=16, bold=True)
    answer_ws.merge_cells("A1:F1")
    answer_ws["A2"], answer_ws["B2"], answer_ws["D2"], answer_ws["E2"] = "공정명", process_name, "총점", total_score
    write_header(answer_ws, 4, ["No", "점수", "문제유형", "원본 No", "답안", "문제"], "8064A2")
    for idx, q in enumerate(questions, start=1):
        row = idx + 4
        answer_ws.cell(row, 1, idx)
        answer_ws.cell(row, 2, q.score)
        answer_ws.cell(row, 3, q.exam_type)
        answer_ws.cell(row, 4, q.source_no)
        answer_ws.cell(row, 5, q.answer)
        answer_ws.cell(row, 6, q.question)
        answer_ws.row_dimensions[row].height = estimate_row_height(q.answer + "\n" + q.question)
    for col, width in {"A": 6, "B": 8, "C": 12, "D": 10, "E": 42, "F": 70}.items():
        answer_ws.column_dimensions[col].width = width
    add_borders(answer_ws, 2, len(questions) + 4, 1, 6)
    answer_ws.freeze_panes = "A5"

    rows = [
        ["항목", "값"],
        ["공정명", process_name],
        ["생성일", generated_at],
        ["목표 점수", target_score],
        ["생성 점수", total_score],
        ["총 문항 수", len(questions)],
        ["공통 문항/점수", f"{counts['공통']}문항 x {scores['공통']}점"],
        ["객관식 문항/점수", f"{counts['객관식']}문항 x {scores['객관식']}점"],
        ["주관식 문항/점수", f"{counts['주관식']}문항 x {scores['주관식']}점"],
    ]
    for r, row_values in enumerate(rows, start=1):
        for c, value in enumerate(row_values, start=1):
            summary_ws.cell(r, c, value)
    write_header(summary_ws, 1, ["항목", "값"], "548235")
    summary_ws.column_dimensions["A"].width = 24
    summary_ws.column_dimensions["B"].width = 44
    add_borders(summary_ws, 1, len(rows), 1, 2)

    wb.save(output_path)


def copy_row_format(ws, source_row: int, target_row: int, min_col: int = 1, max_col: int = 12):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(min_col, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def ensure_exam_rows(ws, required_count: int):
    existing_capacity = 45
    if required_count <= existing_capacity:
        return

    extra = required_count - existing_capacity
    ws.insert_rows(56, extra)
    for offset in range(extra):
        row = 56 + offset
        copy_row_format(ws, 55, row, 2, 12)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=11)


def clear_template_outputs(exam_ws, answer_ws, question_count: int):
    max_exam_row = max(55, 10 + question_count)
    for row in range(11, max_exam_row + 1):
        exam_ws.cell(row, 2).value = None
        exam_ws.cell(row, 3).value = None
        exam_ws.cell(row, 12).value = None

    for row in range(3, 28):
        answer_ws.cell(row, 2).value = None
        answer_ws.cell(row, 3).value = None
        answer_ws.cell(row, 5).value = None
        answer_ws.cell(row, 6).value = None


def write_template_workbook(
    output_path: Path,
    template_path: Path,
    process_name: str,
    questions: list[Question],
    counts: dict[str, int],
):
    wb = openpyxl.load_workbook(template_path, keep_vba=True, data_only=False)
    setting_ws = wb["시험 SETTING"]
    exam_ws = wb["시험지"]
    answer_ws = wb["답안지"]

    ensure_exam_rows(exam_ws, len(questions))
    clear_template_outputs(exam_ws, answer_ws, len(questions))

    setting_ws["B4"] = process_name
    setting_ws["C4"] = counts["공통"]
    setting_ws["D4"] = counts["객관식"]
    setting_ws["E4"] = counts["주관식"]

    for idx, q in enumerate(questions, start=1):
        row = 10 + idx
        exam_ws.cell(row, 2).value = idx
        exam_ws.cell(row, 3).value = q.question
        exam_ws.cell(row, 12).value = None

    for idx, q in enumerate(questions, start=1):
        if idx <= 25:
            row = idx + 2
            answer_ws.cell(row, 2).value = idx
            answer_ws.cell(row, 3).value = q.answer
        elif idx <= 50:
            row = idx - 23
            answer_ws.cell(row, 5).value = idx
            answer_ws.cell(row, 6).value = q.answer
        else:
            row = 28 + (idx - 51)
            if row > answer_ws.max_row:
                answer_ws.insert_rows(row)
                copy_row_format(answer_ws, 27, row, 2, 6)
            answer_ws.cell(row, 2).value = idx
            answer_ws.cell(row, 3).value = q.answer

    if hasattr(wb, "calculation"):
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True

    wb.save(output_path)


def ask_int(prompt: str, default: int) -> int:
    value = input(f"{prompt} [{default}]: ").strip()
    if not value:
        return default
    number = int(value)
    if number < 0:
        raise ValueError("문항 수는 0 이상이어야 합니다.")
    return number


def ask_float(prompt: str, default: float) -> float:
    value = input(f"{prompt} [{default:g}]: ").strip()
    if not value:
        return default
    number = float(value)
    if number <= 0:
        raise ValueError("점수는 0보다 커야 합니다.")
    return number


def default_desktop() -> Path:
    desktop = Path.home() / "Desktop"
    return desktop if desktop.exists() else Path.cwd()


def find_default_workbook() -> Path | None:
    candidates = [
        Path.home() / "Desktop" / "OJT 시험 문제.xlsm",
        Path.cwd() / "OJT 시험 문제.xlsm",
    ]
    users_dir = Path("C:/Users")
    if users_dir.exists():
        for user_dir in users_dir.iterdir():
            candidates.append(user_dir / "Desktop" / "OJT 시험 문제.xlsm")

    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def find_output_desktop(workbook_path: Path) -> Path:
    if workbook_path.parent.name.lower() == "desktop":
        return workbook_path.parent
    return default_desktop()


def main():
    print("=" * 60)
    print(APP_TITLE)
    print("=" * 60)

    workbook_path = find_default_workbook()
    if workbook_path is None:
        entered = input("문제은행 엑셀 경로를 입력하세요: ").strip().strip('"')
        workbook_path = Path(entered)
    if not workbook_path.exists():
        raise FileNotFoundError(f"문제은행 파일을 찾을 수 없습니다: {workbook_path}")
    desktop = find_output_desktop(workbook_path)

    banks = load_question_banks(workbook_path)
    if not banks:
        raise RuntimeError("사용 가능한 문제은행 시트를 찾지 못했습니다.")

    print(f"\n문제은행: {workbook_path}")
    print("\n공정을 선택하세요.")
    for idx, bank in enumerate(banks, start=1):
        print(f"{idx:>2}. {bank.display_name}  (공통 {bank.counts['공통']}, 객관식 {bank.counts['객관식']}, 주관식 {bank.counts['주관식']})")

    selected_no = int(input("\n번호 입력: ").strip())
    if selected_no < 1 or selected_no > len(banks):
        raise ValueError("잘못된 번호입니다.")
    bank = banks[selected_no - 1]

    print(f"\n선택: {bank.display_name}")
    counts = {kind: ask_int(f"{kind} 문항 수", DEFAULT_COUNTS[kind]) for kind in QUESTION_TYPES}
    scores = {kind: ask_float(f"{kind} 점수", DEFAULT_SCORES[kind]) for kind in QUESTION_TYPES}
    target_score = ask_float("목표 점수", 100.0)
    total_score = calculate_total(counts, scores)

    print(f"\n현재 총점: {total_score:g}점")
    if abs(total_score - target_score) >= 0.0001:
        raise ValueError(f"총점이 목표 점수({target_score:g}점)와 다릅니다.")

    selected = select_questions(bank, counts, scores)
    output_dir = desktop / "OJT_Random_Exam_Output"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"OJT_시험지_{safe_filename(bank.display_name)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsm"
    write_exam_workbook(output_path, bank.display_name, selected, counts, scores, target_score, template_path=workbook_path)

    print("\n생성 완료!")
    print(output_path)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print("\n[오류]")
        print(exc)
    finally:
        input("\nEnter 키를 누르면 종료합니다...")
