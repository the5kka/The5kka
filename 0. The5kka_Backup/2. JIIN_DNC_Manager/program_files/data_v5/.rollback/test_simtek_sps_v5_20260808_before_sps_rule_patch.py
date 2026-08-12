import tempfile
import unittest
import zipfile
from pathlib import Path

from simtek_sps import (
    SPS_PROGRAM_HISTORY,
    SpsLot,
    SpsNewModelPopup,
    SpsRepository,
    SimtekSpsController,
    build_sps_program_catalog,
)
from simtek_sps_rules import (
    Recommendation,
    SpsRuleBook,
    extract_size_key,
    extract_program_shift,
    extract_program_size_key,
    find_exact_txt_files,
    normalize_oms_text,
    parse_condition,
    validate_program_condition,
    validate_steps,
)


RULE_FILE = Path(__file__).resolve().parent / "data_v5" / "sps_rules_default.xlsx"
WORKLOG_FILE = Path(r"D:\QC\6. Ongoing\기타 진행 현황\심텍 SPS DNC\2026 ST-PKG 1호기 작업일보 V15.xlsm")


class DummyVar:
    def __init__(self):
        self.value = ""

    def set(self, value):
        self.value = value

    def get(self):
        return self.value


class SimtekSpsRuleTests(unittest.TestCase):
    def test_oms_normalization_ignores_spacing_only(self):
        left = "▷후가공 Size 411 512 mm"
        right = "  후가공   Size 411 512 mm  "
        self.assertEqual(normalize_oms_text(left), normalize_oms_text(right))

    def test_size_variants(self):
        self.assertEqual(extract_size_key("▷후가공 Size 411 512 mm"), "411-512")
        self.assertEqual(extract_size_key("CUT SIZE : 415X516mm"), "415-516")
        self.assertEqual(extract_size_key("▷후가공 413-513 mm"), "413-513")

    def test_oms2_keeps_multiple_features(self):
        parsed = parse_condition(
            "▷후가공 Size 411 512 mm",
            "▷스텍 핀 : 판넬 센터 ▷방향홀 우측 Stacking 必, Edge 좌측 : 5R, 우측 : 2R 가공 ▷뽕따기 금지 X -5.0mm Y 3mm",
        )
        self.assertEqual(parsed.features.pin_position, "CENTER")
        self.assertTrue(parsed.features.direction_hole)
        self.assertTrue(parsed.features.stacking)
        self.assertTrue(parsed.features.punch_prohibit)
        self.assertEqual(parsed.features.shift_x, -5.0)
        self.assertEqual(parsed.features.shift_y, 3.0)
        self.assertEqual(parsed.features.left_edge_r, "5R")
        self.assertEqual(parsed.features.right_edge_r, "2R")

    def test_right_10mm_and_shift_labels(self):
        parsed = parse_condition(
            "후가공 Size 411X512mm",
            "방향홀 / 우측 : 10.0mm / X SHIFT : -2.5mm / Y SHIFT = 1mm",
        )
        self.assertTrue(parsed.features.right_10mm)
        self.assertEqual(parsed.features.shift_x, -2.5)
        self.assertEqual(parsed.features.shift_y, 1.0)

    def test_program_size_and_shift_parsing(self):
        self.assertEqual(extract_program_size_key("ST-06-1-500+10(413-513)2R"), "413-513")
        self.assertEqual(extract_program_shift("ST-06-1-500+10(413-513)2R"), 10.0)
        self.assertEqual(extract_program_shift("ST-06-1-500(+10)(413-513)2R"), 10.0)
        self.assertIsNone(extract_program_shift("ST-3Layer-500+4Hole(411-512)_2R"))

    def test_historical_program_catalog_is_grouped_by_size(self):
        catalog = build_sps_program_catalog(SPS_PROGRAM_HISTORY)
        self.assertEqual(len(catalog), 17)
        self.assertIn("413-513", {option.size_key for option in catalog})
        shifted = [option for option in catalog if option.shift is not None]
        self.assertEqual({option.shift for option in shifted}, {10.0})
        self.assertTrue(any(option.program == "ST-06-1-500+10(413-513)2R" for option in shifted))

    def test_plus_4hole_catalog_entry_is_not_shifted(self):
        catalog = build_sps_program_catalog(SPS_PROGRAM_HISTORY)
        option = next(item for item in catalog if "+4Hole" in item.program)
        self.assertIsNone(option.shift)

    def test_right_10mm_requires_plus_10_program(self):
        parsed = parse_condition(
            "▷후가공 Size 413 513 mm",
            "(Y500 역삽X105) ▷스텍핀 위치 우측 10mm",
        )
        ok = validate_program_condition(parsed, "ST-06-1-500+10(413-513)2R")
        self.assertTrue(ok.ok)
        missing = validate_program_condition(parsed, "ST-BOC-500(413-513)2R")
        self.assertFalse(missing.ok)
        self.assertIn("시프트 누락", missing.summary)

    def test_program_size_mismatch_is_blocked(self):
        parsed = parse_condition("▷후가공 Size 413 513 mm", "스텍핀 판넬 센터")
        result = validate_program_condition(parsed, "ST-BOC-500(411-512)2R")
        self.assertFalse(result.ok)
        self.assertIn("사이즈 불일치", result.summary)

    def test_unrequested_shift_program_is_blocked(self):
        parsed = parse_condition("▷후가공 Size 413 513 mm", "스텍핀 판넬 센터")
        result = validate_program_condition(parsed, "ST-06-1-500+10(413-513)2R")
        self.assertFalse(result.ok)
        self.assertIn("OMS 시프트 없음", result.summary)

    def test_step_rule_from_v15_vba(self):
        self.assertTrue(validate_steps("180", "240").ok)
        self.assertTrue(validate_steps("180", "250").ok)
        self.assertFalse(validate_steps("180", "230").ok)

    def test_rule_book_management_and_multiple_candidate(self):
        rules = SpsRuleBook(RULE_FILE)
        center = parse_condition("▷후가공 Size 411 512 mm", "판넬 CENTER")
        result = rules.recommend("FLC10340B00-007", "2차", center)
        self.assertEqual(result.programs, ("06-OUT_(411-512)2R(New)",))
        self.assertEqual(result.candidates[0].source, "관리번호 전용 규칙")

        no_detail = parse_condition("▷후가공 Size 411 512 mm", "기타 확인 조건")
        result = rules.recommend("FLC99999A00-001", "2차", no_detail)
        self.assertEqual(len(result.programs), 2)
        self.assertIn("자동 DNC 금지", " ".join(result.warnings))

    def test_input_recommendation_is_used_when_oms2_is_blank(self):
        rules = SpsRuleBook(RULE_FILE)
        parsed = parse_condition("▷후가공 Size 411 512 mm", "")
        result = rules.recommend("FLC10340B00-005", "2차", parsed)
        self.assertEqual(result.programs, ("ST-BOC-500(411-512)2R",))
        self.assertEqual(result.basis, "INPUT 추천 Trim P/G 정확 일치")

    def test_input_dnc_block_cannot_fall_back_to_any_rule(self):
        rules = SpsRuleBook(RULE_FILE)
        parsed = parse_condition(
            "▷후가공 Size 415 515 mm",
            "▷ 뽕따기 금지 !! ▷스텍 핀 : 판넬 센터 ▷방향홀 우측 Stacking 必, Edge 좌측 : 5R, 우측 : 2R 가공",
        )
        result = rules.recommend("MCP21815C00-013", "2차", parsed)
        self.assertEqual(result.programs, ())
        self.assertEqual(result.basis, "INPUT 최종판정 DNC 차단")
        self.assertIn("조건 미등록", " ".join(result.warnings))

    def test_input_oms2_meaning_match_accepts_spacing_variant(self):
        rules = SpsRuleBook(RULE_FILE)
        parsed = parse_condition(
            "▷후가공 Size 411 512 mm",
            "▷스텍핀:판넬 센터 ▷방향홀 우측 Stacking 必, Edge 좌측:5R, 우측:2R 가공",
        )
        result = rules.recommend("MCP22465B00-005", "2차", parsed)
        self.assertEqual(
            result.programs,
            ("ST-3Layer-500+4Hole(411-512)_2R_2_2024 T01 T02 test",),
        )
        self.assertEqual(result.basis, "INPUT 추천 Trim P/G OMS2 의미 일치")

    def test_input_oms2_meaning_match_keeps_blocked_edge_condition(self):
        rules = SpsRuleBook(RULE_FILE)
        parsed = parse_condition(
            "▷후가공 Size 415 515 mm",
            "▷뽕따기 금지 ▷스텍핀:판넬 센터 ▷방향홀 우측 Stacking 必, Edge 좌측:5R, 우측:2R 가공",
        )
        result = rules.recommend("MCP21815C00-013", "2차", parsed)
        self.assertEqual(result.programs, ())
        self.assertEqual(result.basis, "INPUT 최종판정 DNC 차단")

    def test_clean_text_removes_invisible_format_characters(self):
        rules = SpsRuleBook(RULE_FILE)
        parsed = parse_condition(
            "▷후가공\u200b Size 411 512 mm",
            "▷스텍 핀 : 판넬 센터\ufeff ▷방향홀 우측 Stacking 必, Edge 좌측 : 5R, 우측 : 2R 가공",
        )
        result = rules.recommend("MCP22465B00-005", "2차", parsed)
        self.assertEqual(
            result.programs,
            ("ST-3Layer-500+4Hole(411-512)_2R_2_2024 T01 T02 test",),
        )

    def test_exact_txt_file_only(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / "ABC.txt").write_text("ok", encoding="utf-8")
            (root / "ABC_REV1.txt").write_text("no", encoding="utf-8")
            (root / "ABC.dnc").write_text("no", encoding="utf-8")
            self.assertEqual([path.name for path in find_exact_txt_files(root, "ABC")], ["ABC.txt"])


class SimtekSpsRepositoryTests(unittest.TestCase):
    def _lot(self):
        parsed = parse_condition("▷후가공 Size 411 512 mm", "판넬 CENTER 방향홀")
        return SpsLot(1, "MCP00001A00-001", "LOT-001", 180, 240, 1, 60, parsed.oms1_raw, parsed.oms2_raw, parsed)

    def test_master_exclude_and_reregister(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            repo = SpsRepository(Path(temp_dir) / "sps.db")
            lot = self._lot()
            first_id = repo.register_master(lot, "ST-BOC-500(411-512)2R", "조장", "시험")
            self.assertEqual(len(repo.find_active_master(lot)), 1)
            repo.set_master_active(first_id, False, "오등록")
            self.assertEqual(repo.find_active_master(lot), [])
            second_id = repo.register_master(lot, "ST-BOC-500(411-512)2R", "조장", "재등록")
            self.assertNotEqual(first_id, second_id)

    def test_history_manage_number_ignores_hyphens(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            repo = SpsRepository(Path(temp_dir) / "sps.db")
            lot = self._lot()
            lot.manage_no = "MCP-00001A00-001"
            lot.program = "ST-BOC-500(411-512)2R"
            ids = repo.insert_logs(
                {"machine": "1호기", "work_date": "2026-08-07", "shift_group": "A", "shift": "주간", "worker": "시험자"},
                [lot], "일반", "", "시험",
            )
            repo.finish_logs(ids, True, True)
            lookup = self._lot()
            lookup.manage_no = "MCP00001A00001"
            self.assertEqual(repo.history_candidates(lookup), [("ST-BOC-500(411-512)2R", 1)])

    def test_two_lot_same_lot_number_is_blocked(self):
        controller = SimtekSpsController.__new__(SimtekSpsController)
        controller.master_by_lot = {}
        base = {
            "manage_no": "MCP00001A00-001", "lot_no": "LOT-001",
            "start_step": "180", "current_step": "240", "round": "1", "qty": "60",
            "oms1": "후가공 Size 411 512 mm", "oms2": "판넬 CENTER",
            "program": "", "detail": "",
        }
        controller.lot_vars = {
            1: {key: DummyVar() for key in base},
            2: {key: DummyVar() for key in base},
        }
        for lot_number in (1, 2):
            for key, value in base.items():
                controller.lot_vars[lot_number][key].set(value)
        controller.lot_vars[2]["manage_no"].set("MCP00001A00-002")
        with self.assertRaisesRegex(ValueError, "동일 LOT"):
            controller._collect_lots()

    def test_new_model_without_recommendation_shows_simple_message(self):
        class DummyController:
            @staticmethod
            def _recommend(_lot):
                return Recommendation("관리자 확인 필요", (), "추천 규칙 없음", ())

        popup = SpsNewModelPopup.__new__(SpsNewModelPopup)
        popup.controller = DummyController()
        popup.lots = [self._lot()]
        popup.program_var = DummyVar()
        popup.recommendation_basis = DummyVar()
        popup.recommendation_text = {}
        popup._prepare_recommendations()
        self.assertFalse(popup.can_run)
        self.assertEqual(popup.recommendation_text[1], "추천 규칙 조건 없음")

    def test_new_model_requires_confirmed_program(self):
        oms1 = "▷후가공 Size 413 513 mm"
        oms2 = "(Y500 역삽X105)▷스텍핀 위치 우측 10mm(후가공 주의할 것)"
        lot = SpsLot(
            1,
            "MCP24317A00-008",
            "1111",
            80,
            150,
            1,
            60,
            oms1,
            oms2,
            parse_condition(oms1, oms2),
        )
        popup = SpsNewModelPopup.__new__(SpsNewModelPopup)
        popup.lots = [lot]
        popup.can_run = False
        popup.confirmed_program = ""
        popup.run_button = None
        with self.assertRaisesRegex(ValueError, "조건 확인·확정"):
            popup.require_confirmed_program("ST-06-1-500+10(413-513)2R")
        popup.can_run = True
        popup.confirmed_program = "ST-06-1-500+10(413-513)2R"
        self.assertEqual(
            popup.require_confirmed_program("ST-06-1-500+10(413-513)2R"),
            "ST-06-1-500+10(413-513)2R",
        )

    def test_worklog_export_writes_condition2_and_time(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            target = temp_root / "worklog.xlsx"
            from openpyxl import Workbook, load_workbook
            seed = Workbook()
            ws = seed.active
            ws.title = "ST-PKG"
            for column, header in enumerate((
                "작업일자", "조", "근무", "작업자", "STEP", "관리번호", "LOT NO", "수량",
                "실적", "차수", "OMS 조건1", "OMS 조건2", "작업 P/G", "자주검사", "Trimming 확인",
                "", "", "작업 시작시간",
            ), start=1):
                ws.cell(8, column).value = header
            seed.save(target)
            seed.close()
            repo = SpsRepository(temp_root / "sps.db")
            lot = self._lot()
            lot.program = "ST-BOC-500(411-512)2R"
            ids = repo.insert_logs(
                {"machine": "트리밍 1호기", "work_date": "2026-08-07", "shift_group": "A", "shift": "주간", "worker": "시험자"},
                [lot], "일반", "", "시험",
            )
            repo.finish_logs(ids, True, True)
            controller = SimtekSpsController.__new__(SimtekSpsController)
            controller.repo = repo
            controller.config = {"simtek_sps_excel_file": str(target)}
            controller.acquire_excel_lock_callback = None
            controller.release_excel_lock_callback = None
            controller.save_workbook_callback = lambda workbook, path: workbook.save(path)
            controller.excel_status_var = DummyVar()
            controller._append_log = lambda _text: None
            controller._alert = lambda *_args, **_kwargs: None
            self.assertEqual(controller.export_logs_to_excel(), 1)

            workbook = load_workbook(target, read_only=True, data_only=False)
            ws = workbook["ST-PKG"]
            markers = [ws.cell(row=row, column=33).value for row in range(9, ws.max_row + 1)]
            row = markers.index(f"SPS:{ids[0]}") + 9
            self.assertEqual(ws.cell(row, 11).value, lot.oms1)
            self.assertEqual(ws.cell(row, 12).value, lot.oms2)
            self.assertEqual(ws.cell(row, 13).value, lot.program)
            self.assertIsNotNone(ws.cell(row, 18).value)
            self.assertEqual(ws.cell(row, 18).number_format, "hh:mm:ss")
            workbook.close()

    def test_worklog_and_db_master_must_match(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            target = temp_root / "worklog.xlsx"
            from openpyxl import Workbook, load_workbook

            lot = self._lot()
            program = "ST-BOC-500(411-512)2R"
            workbook = Workbook()
            ws = workbook.active
            ws.title = "ST-PKG"
            headers = (
                "작업일자", "조", "근무", "작업자", "STEP", "관리번호", "LOT", "수량",
                "실적", "차수", "OMS Trim Size", "OMS Trim Size (2)", "작업 P/G",
                "자주 검사", "트리밍 확인란",
            )
            for column, header in enumerate(headers, start=1):
                ws.cell(8, column).value = header
            values = {
                5: lot.current_step, 6: lot.manage_no, 7: lot.lot_no, 8: lot.qty,
                10: lot.round_no, 11: lot.oms1, 12: lot.oms2, 13: program,
                14: "이상없음", 15: "OK",
            }
            for column, value in values.items():
                ws.cell(9, column).value = value
            workbook.save(target)
            workbook.close()

            repo = SpsRepository(temp_root / "sps.db")
            repo.register_master(lot, program, "조장", "신규 검증")
            controller = SimtekSpsController.__new__(SimtekSpsController)
            controller.repo = repo
            controller.config = {"simtek_sps_excel_file": str(target)}
            master, selected, reason = controller._verified_existing_condition(lot)
            self.assertIsNotNone(master)
            self.assertEqual(selected, program)
            self.assertIn("일치", reason)

            workbook = load_workbook(target)
            workbook["ST-PKG"].cell(9, 13).value = "DIFFERENT-PROGRAM"
            workbook.save(target)
            workbook.close()
            master, selected, reason = controller._verified_existing_condition(lot)
            self.assertIsNone(master)
            self.assertEqual(selected, "")
            self.assertIn("불일치", reason)

    @unittest.skipUnless(WORKLOG_FILE.exists(), "현장 참조 작업일보가 없음")
    def test_reference_worklog_contains_macro(self):
        with zipfile.ZipFile(WORKLOG_FILE) as archive:
                self.assertIn("xl/vbaProject.bin", archive.namelist())


if __name__ == "__main__":
    unittest.main(verbosity=2)
