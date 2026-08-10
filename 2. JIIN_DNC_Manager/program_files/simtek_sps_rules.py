from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ALLOWED_STEP_DELTAS = (60, 70)


def clean_text(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = "".join(character for character in text if unicodedata.category(character) != "Cf")
    text = text.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    text = text.replace("：", ":").replace("，", ",")
    return re.sub(r"\s+", " ", text).strip()


def normalize_oms_text(value: object) -> str:
    text = clean_text(value).casefold()
    text = re.sub(r"[▷▶►]+", " ", text)
    text = re.sub(r"\s*([:(),])\s*", r"\1", text)
    # Size separators are normalized only when they are between numbers.
    # A standalone X/Y in OMS2 is a shift label and its boundary must remain.
    text = re.sub(r"(?<=\d)\s*[x×*]\s*(?=\d)", "x", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_manage_no(value: object) -> str:
    return re.sub(r"[\s-]+", "", clean_text(value)).upper()


def parse_round(value: object) -> int | None:
    match = re.search(r"\d+", clean_text(value))
    return int(match.group()) if match else None


def extract_size_key(oms1: object) -> str:
    text = normalize_oms_text(oms1)
    patterns = (
        r"(?:size|cut\s*size)[^0-9]{0,20}(\d{3})\s*(?:x|-|\s)\s*(\d{3})",
        r"(?:후가공|trim)[^0-9]{0,30}(\d{3})\s*(?:x|-|\s)\s*(\d{3})",
        r"\b(\d{3})\s*(?:x|-)\s*(\d{3})\b",
        r"\b(\d{3})\s+(\d{3})\s*mm\b",
    )
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return f"{int(match.group(1))}-{int(match.group(2))}"
    return ""


def _extract_number_after(label_pattern: str, text: str) -> float | None:
    match = re.search(label_pattern + r"\s*[:=]?\s*([+-]?\d+(?:\.\d+)?)\s*(?:mm)?", text, re.IGNORECASE)
    return float(match.group(1)) if match else None


def _extract_edge_radius(side: str, text: str) -> str:
    patterns = (
        rf"{side}[^,;▷]*?([0-9]+(?:\.[0-9]+)?)\s*r",
        rf"{side}\s*edge[^,;▷]*?([0-9]+(?:\.[0-9]+)?)\s*r",
    )
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            number = float(match.group(1))
            return f"{number:g}R"
    return ""


@dataclass(frozen=True)
class Oms2Features:
    pin_position: str = ""
    direction_hole: bool = False
    right_10mm: bool = False
    process_pass: int | None = None
    punch_prohibit: bool = False
    size_note: bool = False
    stacking: bool = False
    shift_x: float | None = None
    shift_y: float | None = None
    left_edge_r: str = ""
    right_edge_r: str = ""

    def to_json(self) -> str:
        return json.dumps(asdict(self), ensure_ascii=False, sort_keys=True)

    def summary(self) -> str:
        labels: list[str] = []
        if self.pin_position:
            labels.append(f"PIN={self.pin_position}")
        if self.direction_hole:
            labels.append("방향홀")
        if self.right_10mm:
            labels.append("우측 10mm")
        if self.process_pass:
            labels.append(f"후가공 {self.process_pass}차")
        if self.punch_prohibit:
            labels.append("뽕따기 금지")
        if self.size_note:
            labels.append("SIZE 주의")
        if self.stacking:
            labels.append("STACKING")
        if self.shift_x is not None:
            labels.append(f"X={self.shift_x:g}")
        if self.shift_y is not None:
            labels.append(f"Y={self.shift_y:g}")
        if self.left_edge_r:
            labels.append(f"좌측={self.left_edge_r}")
        if self.right_edge_r:
            labels.append(f"우측={self.right_edge_r}")
        return ", ".join(labels) if labels else "추출 항목 없음"


def extract_oms2_features(oms2: object) -> Oms2Features:
    text = normalize_oms_text(oms2)
    process_match = re.search(r"후가공\s*\(?\s*(\d+)\s*차\s*\)?", text)
    pin_position = "CENTER" if ("센터" in text or "center" in text) else ""
    right_10 = bool(re.search(r"우측\s*[:=]?\s*10(?:\.0+)?\s*mm", text))
    return Oms2Features(
        pin_position=pin_position,
        direction_hole="방향홀" in text,
        right_10mm=right_10,
        process_pass=int(process_match.group(1)) if process_match else None,
        punch_prohibit=("뽕따기" in text and "금지" in text),
        size_note=("size" in text and any(word in text for word in ("주의", "유의", "확인"))),
        stacking=("스텍" in text or "stacking" in text or "stack" in text),
        shift_x=_extract_number_after(r"(?<![0-9a-z])x(?:\s*shift)?", text),
        shift_y=_extract_number_after(r"(?<![0-9a-z])y(?:\s*shift)?", text),
        left_edge_r=_extract_edge_radius("좌측", text),
        right_edge_r=_extract_edge_radius("우측", text),
    )


def get_legacy_oms2_key(features: Oms2Features, oms2: object) -> str:
    if features.right_10mm:
        return "우측10mm"
    if features.direction_hole:
        return "방향홀"
    if features.process_pass in {1, 2, 3}:
        return f"후가공({features.process_pass}차)"
    if features.pin_position == "CENTER":
        return "CENTER"
    if features.size_note:
        return "SIZE 유의"
    # The reference workbook intentionally treats a blank OMS2 as
    # "OMS2 확인" so the ANY rule can still provide a recommendation.
    return "OMS2 확인"


@dataclass(frozen=True)
class ParsedCondition:
    oms1_raw: str
    oms2_raw: str
    oms1_norm: str
    oms2_norm: str
    size_key: str
    oms2_key: str
    features: Oms2Features

    def exact_key(self) -> tuple[str, str]:
        return self.oms1_norm, self.oms2_norm


def parse_condition(oms1: object, oms2: object) -> ParsedCondition:
    oms1_raw = clean_text(oms1)
    oms2_raw = clean_text(oms2)
    features = extract_oms2_features(oms2_raw)
    return ParsedCondition(
        oms1_raw=oms1_raw,
        oms2_raw=oms2_raw,
        oms1_norm=normalize_oms_text(oms1_raw),
        oms2_norm=normalize_oms_text(oms2_raw),
        size_key=extract_size_key(oms1_raw),
        oms2_key=get_legacy_oms2_key(features, oms2_raw),
        features=features,
    )


def extract_program_size_key(program: object) -> str:
    text = clean_text(program)
    for match in re.finditer(r"\((\d{3})\s*[-x×]\s*(\d{3})(?:-[^)]+)?\)", text, re.IGNORECASE):
        return f"{int(match.group(1))}-{int(match.group(2))}"
    return ""


def extract_program_shift(program: object) -> float | None:
    text = clean_text(program)
    patterns = (
        r"\(\s*([+-]\s*\d+(?:\.\d+)?)\s*\)",
        r"500\s*([+-]\s*\d+(?:\.\d+)?)(?!\s*hole)",
    )
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return float(match.group(1).replace(" ", ""))
    return None


def extract_oms_program_shift(oms2: object) -> float | None:
    text = normalize_oms_text(oms2)
    for direction, sign in (("우측", 1), ("좌측", -1)):
        match = re.search(rf"{direction}\s*[:=]?\s*(\d+(?:\.\d+)?)\s*mm", text)
        if match:
            return sign * float(match.group(1))
    match = re.search(r"(?:shift|시프트)\s*[:=]?\s*([+-]\d+(?:\.\d+)?)\s*(?:mm)?", text)
    return float(match.group(1)) if match else None


def format_shift(value: float | None) -> str:
    if value is None:
        return "없음"
    return f"{value:+g}"


@dataclass(frozen=True)
class ProgramConditionCheck:
    ok: bool
    expected_size: str
    program_size: str
    expected_shift: float | None
    program_shift: float | None
    errors: tuple[str, ...]

    @property
    def summary(self) -> str:
        if self.ok:
            return (
                f"사이즈 {self.expected_size} 일치 / "
                f"시프트 {format_shift(self.expected_shift)} 일치"
            )
        return " / ".join(self.errors)


def validate_program_condition(parsed: ParsedCondition, program: object) -> ProgramConditionCheck:
    program_name = clean_text(program)
    expected_size = parsed.size_key
    program_size = extract_program_size_key(program_name)
    expected_shift = extract_oms_program_shift(parsed.oms2_raw)
    program_shift = extract_program_shift(program_name)
    errors: list[str] = []

    if not program_name:
        errors.append("최종 Trim P/G를 선택하거나 입력하세요.")
    if not expected_size:
        errors.append("OMS 조건1에서 사이즈를 확인할 수 없습니다.")
    elif not program_size:
        errors.append("Trim P/G 이름에서 (가로-세로) 사이즈를 확인할 수 없습니다.")
    elif expected_size != program_size:
        errors.append(f"사이즈 불일치: OMS {expected_size} / P/G {program_size}")

    if expected_shift is None and program_shift is not None:
        errors.append(f"시프트 불일치: OMS 시프트 없음 / P/G {format_shift(program_shift)}")
    elif expected_shift is not None and program_shift is None:
        errors.append(f"시프트 누락: OMS {format_shift(expected_shift)} / P/G 시프트 없음")
    elif expected_shift is not None and program_shift is not None and abs(expected_shift - program_shift) > 0.001:
        errors.append(
            f"시프트 불일치: OMS {format_shift(expected_shift)} / P/G {format_shift(program_shift)}"
        )

    return ProgramConditionCheck(
        not errors,
        expected_size,
        program_size,
        expected_shift,
        program_shift,
        tuple(errors),
    )


@dataclass(frozen=True)
class StepCheck:
    ok: bool
    start_step: int | None
    current_step: int | None
    delta: int | None
    message: str


def validate_steps(start_value: object, current_value: object) -> StepCheck:
    try:
        start_step = int(clean_text(start_value))
        current_step = int(clean_text(current_value))
    except ValueError:
        return StepCheck(False, None, None, None, "시작 STEP과 현 STEP을 숫자로 입력하세요.")
    delta = current_step - start_step
    if delta not in ALLOWED_STEP_DELTAS:
        allowed = " 또는 ".join(f"+{value}" for value in ALLOWED_STEP_DELTAS)
        return StepCheck(False, start_step, current_step, delta, f"STEP 차이 {delta}: 현재 기준은 {allowed}만 허용합니다.")
    return StepCheck(True, start_step, current_step, delta, f"STEP 확인 OK (+{delta})")


@dataclass(frozen=True)
class RuleCandidate:
    program: str
    source: str
    priority: int
    note: str = ""
    rule_key: str = ""
    status: str = "CONFIRMED"


@dataclass(frozen=True)
class Recommendation:
    status: str
    candidates: tuple[RuleCandidate, ...]
    basis: str
    warnings: tuple[str, ...]

    @property
    def programs(self) -> tuple[str, ...]:
        return tuple(candidate.program for candidate in self.candidates)


def _split_program_candidates(value: object) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    if text.startswith("후보:"):
        text = text.split(":", 1)[1]
    return [part.strip() for part in re.split(r"\s+/\s+", text) if part.strip()]


def normalize_oms2_rule_key(value: object) -> str:
    text = clean_text(value)
    match = re.fullmatch(r"후가공\s*\(?\s*(\d+)\s*차\s*\)?", text)
    if match:
        return f"후가공({int(match.group(1))}차)"
    return text


def normalize_rule_status(value: object, default: str = "CONFIRMED") -> str:
    text = clean_text(value).upper()
    if not text:
        return default
    if text in {"확정", "등록", "예외"}:
        return "CONFIRMED"
    if text in {"후보", "검토", "추천"}:
        return "REVIEW"
    if text in {"비활성", "자동 적용 해제", "사용 안 함"}:
        return "INACTIVE"
    return text


def is_active_rule(row: dict[str, object]) -> bool:
    return normalize_rule_status(row.get("상태")) not in {"INACTIVE", "LEGACY", "DISABLED"}


class SpsRuleBook:
    def __init__(self, path: Path):
        self.path = Path(path)
        self.input_rows: list[dict[str, object]] = []
        self.main_rules: list[dict[str, object]] = []
        self.management_rules: list[dict[str, object]] = []
        self.registered_programs: set[str] = set()
        self._load()

    @staticmethod
    def _sheet_rows(workbook, sheet_name: str, header_row: int = 1) -> list[dict[str, object]]:
        if sheet_name not in workbook.sheetnames:
            return []
        rows = list(workbook[sheet_name].iter_rows(min_row=header_row, values_only=True))
        if not rows:
            return []
        headers = [clean_text(value) for value in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:] if any(value not in (None, "") for value in row)]

    def _load(self) -> None:
        if not self.path.exists():
            raise FileNotFoundError(f"SPS 규칙 파일 없음: {self.path}")
        workbook = load_workbook(self.path, read_only=True, data_only=True, keep_links=False)
        try:
            self.input_rows = self._sheet_rows(workbook, "INPUT", header_row=3)
            self.main_rules = self._sheet_rows(workbook, "RULE_MAIN")
            self.management_rules = self._sheet_rows(workbook, "MGMT_RULE")
            for row in self._sheet_rows(workbook, "PG_LIST"):
                program = clean_text(row.get("실제 DNC Program 파일명"))
                status = clean_text(row.get("등록상태"))
                if program and (not status or status == "등록"):
                    self.registered_programs.add(program)
        finally:
            workbook.close()

    def recommend(self, manage_no: object, round_value: object, parsed: ParsedCondition) -> Recommendation:
        warnings: list[str] = []
        manage_key = normalize_manage_no(manage_no)
        round_no = parse_round(round_value)
        product_family = manage_key[:3]

        if not parsed.size_key:
            warnings.append("OMS 조건1에서 SIZE를 추출하지 못했습니다.")
        if round_no is None:
            warnings.append("차수를 확인할 수 없습니다.")
        if warnings:
            return Recommendation("관리자 확인 필요", (), "입력 조건 부족", tuple(warnings))

        management_matches: list[RuleCandidate] = []
        for row in self.management_rules:
            if not is_active_rule(row):
                continue
            if normalize_manage_no(row.get("관리번호")) != manage_key:
                continue
            if parse_round(row.get("차수")) != round_no:
                continue
            if clean_text(row.get("SIZE_KEY")) != parsed.size_key:
                continue
            rule_oms2 = normalize_oms2_rule_key(row.get("OMS2_KEY"))
            if rule_oms2 not in {parsed.oms2_key, "ANY", "OMS2 확인"}:
                continue
            for program in _split_program_candidates(row.get("Trim P/G")):
                note = clean_text(row.get("비고"))
                if rule_oms2 == "OMS2 확인":
                    note = (note + " / OMS2 상세는 작업자 확인 필요").strip(" /")
                management_matches.append(RuleCandidate(
                    program,
                    "관리번호 전용 규칙",
                    1,
                    note,
                    clean_text(row.get("관리번호_KEY")),
                    normalize_rule_status(row.get("상태")),
                ))
        if management_matches:
            return self._result(management_matches, "관리번호 + 차수 + SIZE + OMS2 규칙", warnings)

        exact_matches: list[RuleCandidate] = []
        any_matches: list[RuleCandidate] = []
        for row in self.main_rules:
            if not is_active_rule(row):
                continue
            if clean_text(row.get("제품군")).upper() != product_family:
                continue
            if parse_round(row.get("차수")) != round_no:
                continue
            if clean_text(row.get("SIZE_KEY")) != parsed.size_key:
                continue
            rule_oms2 = normalize_oms2_rule_key(row.get("OMS2_KEY"))
            target = exact_matches if rule_oms2 == parsed.oms2_key else any_matches if rule_oms2 == "ANY" else None
            if target is None:
                continue
            priority = 2 if target is exact_matches else 3
            source = "제품군 상세 규칙" if target is exact_matches else "제품군 일반 규칙"
            row_status = normalize_rule_status(row.get("상태"))
            if clean_text(row.get("Trim P/G")).startswith("후보:"):
                row_status = "REVIEW"
            for program in _split_program_candidates(row.get("Trim P/G")):
                target.append(RuleCandidate(
                    program,
                    source,
                    priority,
                    clean_text(row.get("비고")),
                    clean_text(row.get("RULE_KEY")),
                    row_status,
                ))
        matches = exact_matches or any_matches
        if matches:
            return self._result(matches, "제품군 + 차수 + SIZE + OMS2 추천", warnings)

        # INPUT is historical reference data, so it is only a candidate after
        # confirmed management/exact/any rules have all failed to match.
        input_base_rows = [
            row for row in self.input_rows
            if normalize_manage_no(row.get("관리번호")) == manage_key
            and parse_round(row.get("차수")) == round_no
            and normalize_oms_text(row.get("OMS 1")) == parsed.oms1_norm
        ]
        input_matches = [
            row for row in input_base_rows
            if normalize_oms_text(row.get("OMS 2")) == parsed.oms2_norm
        ]
        input_match_basis = "INPUT 추천 Trim P/G 정확 일치"
        if not input_matches and (parsed.oms2_raw or parsed.features != Oms2Features()):
            input_matches = [
                row for row in input_base_rows
                if parse_condition(row.get("OMS 1"), row.get("OMS 2")).features == parsed.features
            ]
            input_match_basis = "INPUT 추천 Trim P/G OMS2 의미 일치"
        if input_matches:
            blocked = [row for row in input_matches if clean_text(row.get("최종판정")) != "DNC 가능"]
            if blocked:
                details = []
                for row in blocked:
                    verdict = clean_text(row.get("최종판정")) or "판정 없음"
                    note = clean_text(row.get("비고"))
                    details.append(f"{verdict}{f' - {note}' if note else ''}")
                return Recommendation("BLOCKED", (), "INPUT 최종판정 DNC 차단", tuple(dict.fromkeys(details)))

            input_candidates: list[RuleCandidate] = []
            for row in input_matches:
                source = clean_text(row.get("적용구분")) or "INPUT"
                note = clean_text(row.get("비고"))
                for program in _split_program_candidates(row.get("추천 Trim P/G")):
                    if program == "관리자 확인":
                        continue
                    input_candidates.append(RuleCandidate(
                        program,
                        f"INPUT {source}",
                        4,
                        note,
                        f"INPUT:{manage_key}|{round_no}|{parsed.size_key}|{parsed.oms2_key}",
                        "REVIEW",
                    ))
            if input_candidates:
                return self._result(input_candidates, input_match_basis, warnings)
            return Recommendation("BLOCKED", (), "INPUT 추천 Trim P/G 없음", ("추천 Trim P/G를 확인하세요.",))

        warnings.append("일치하는 추천 규칙이 없습니다.")
        return Recommendation("BLOCKED", (), "추천 규칙 없음", tuple(warnings))

    def _result(self, candidates: Iterable[RuleCandidate], basis: str, warnings: list[str]) -> Recommendation:
        unique: dict[str, RuleCandidate] = {}
        for candidate in candidates:
            unique.setdefault(candidate.program.casefold(), candidate)
        values = tuple(unique.values())
        for candidate in values:
            if self.registered_programs and candidate.program not in self.registered_programs:
                warnings.append(f"PG_LIST 미등록: {candidate.program}")
        if len(values) > 1:
            warnings.append(f"동일 조건에서 Program {len(values)}개 발견 - 자동 DNC 금지")
        pg_missing = any(message.startswith("PG_LIST 미등록:") for message in warnings)
        if not values or pg_missing:
            status = "BLOCKED"
        elif len(values) == 1 and values[0].status == "CONFIRMED" and not warnings:
            status = "CONFIRMED"
        else:
            status = "REVIEW"
        return Recommendation(status, values, basis, tuple(dict.fromkeys(warnings)))


def find_exact_txt_files(folder: Path, program_name: object) -> list[Path]:
    root = Path(folder)
    name = clean_text(program_name)
    if not root.is_dir() or not name:
        return []
    expected = name.casefold()
    matches = [
        path
        for path in root.rglob("*.txt")
        if path.is_file() and path.stem.casefold() == expected
    ]
    return sorted(matches, key=lambda item: str(item).casefold())
