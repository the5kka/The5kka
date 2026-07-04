import io
import json
import base64
import calendar
import ctypes
import os
import shutil
import sqlite3
import socket
import subprocess
import sys
import threading
import time
import tkinter as tk
import traceback
import unicodedata
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, simpledialog, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from PIL import Image, ImageDraw, ImageTk

from dnc_rules import (
    format_duplicate_condition_files,
    search_condition_file_exact_txt,
    validate_process_paths,
)


# ==================================================
# 기본 설정
# ==================================================
# 현장 적용 시 기본 삭제 시간을 60초로 바꾸고 싶으면 아래 값만 수정해도 됩니다.
DNC_DELETE_SECONDS = 10
# DNC 완료 후 초품 확인 팝업이 뜨기 전 대기 시간입니다. 현장 적용 시 이 값만 바꾸면 됩니다.
FIRST_ARTICLE_WAIT_SECONDS = 5
EXCEL_LOCK_STALE_SECONDS = 10 * 60
WORK_LOG_SCHEMA_VERSION = 2
CONDITION_MASTER_SCHEMA_VERSION = 2
MASTER_SETTINGS_PASSWORD = "1"
CONDITION_MASTER_PASSWORD = "1"
LICENSE_PASSWORD = "1"
DEFAULT_MASTER_PC_NAME = "KUKJIN"
DEFAULT_ALLOWED_IP_PREFIXES = ["121.155.196"]

APP_TITLE = "JIIN DNC Manager"
APP_VERSION_TEXT = "JIIN_DNC_Manager V3.0"
LOG_SHEET_NAME = "KCC PKG"
EXCEL_PROCESS_CODE_COLUMN = 30
EXCEL_EXPORT_ID_COLUMN = 31
SINGLE_INSTANCE_MUTEX_NAME = "JIIN_DNC_Manager_Single_Instance"
ERROR_ALREADY_EXISTS = 183
SINGLE_INSTANCE_HANDLE = None
PROCESS_NAMES = ["TLB", "심텍 SPS", "심텍 HDI", "KCC PKG", "KCC HDI"]
TLB_LOOKUP_GEAR_PNG_BASE64 = "iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAYAAAD0eNT6AAAACXBIWXMAAA7DAAAOwwHHb6hkAAAAGXRFWHRTb2Z0d2FyZQB3d3cuaW5rc2NhcGUub3Jnm+48GgAAIABJREFUeJzs3Xm8rlP9//HXPiPO4TjmWYbMSsgQDWQokiJUir4SpaRZJVJpUFJKhUaKCg0KkbEMyTzPZD4OzjnOwJn3/v3xOftn2/Zw39f6rOG67vfz8ViPqO77/qx13/tan2tdawAR6SRdwLuAPwKPAnOBnsLLHOB/wOnADv5NIiIi0myrA9eSv0MPLX8BlnZuGxERkUZaFXic/J23V7kVWNK1hURERBron+TvtL3L6a4tJCIi0jDbkr+zjlEWAhs5tpNIRxmROwARiW7P3AFEMgJ4T+4gROpKCYBI862fO4CINskdgEhdKQEQab4mz5hfNncAInWlBEBERKQDKQEQERHpQEoAREREOpASABERkQ6kBEBERKQDjcodgIgUayG25W5sSwLrJfgcEelDCYCIDOZFYMsEn7MjcFmCzxGRPvQIQEREpAMpARAREelAegQgEm5VYA9gQ2AWcC9wATA1Z1AiIiISx1rAWdhkuf4n1b0IfBsYmy26l1xBtdP2ZiSKb8eK8fVQxtyBscDHgSuxpG8BMBm4EPgAGmkVEWmUfbEOcrgO6hry78WvBCCeTYAHGTrGK7CVDiIiUnP7AvNpvZO6FhifJVKjBCCOrYHpg8Q10G9gqTxhioiIh3cD82i/o7oaGJchXlACEMNmwJQWY+ybBGgkQESkht4GzKF6Z/VP8swJUALg67XAcxXjvQYlASIitbILMJvqHVVv+QvpV98oAfCzAfB0QLw92GhQzkdCIiLSop3w6fx7yznAyITxKwHwsT4wKSDWvuUqlASIiBRte2xtv1fn31t+Q7rlYUoAwr0aeDIgzoHKv1ESICJSpO2Amfh3/r3lx4nqoQQgzLrAEwExDlUuARZPUAcREWnRG2htnX9oOT5iHbqA1wH3VIytDgnAXdiM/K5Isa0FPBYQXyvlYmCxSPGLiEgbtqL19d0e5RjH2JcG9gF+BTwVGFcdEoDe8iTwC2Bv/GbZrwH8zyG2Vso/UBIgIpJVlfXdHuXIgJjHYGcRnAG84BhTnRKAvmU2cDbWJlVXXKzG8Dv8eZeLUBIgIpJFrs6/B+gGDmsz3q2An0WMua4JQN/yKPB17MCmVq0C3B8xpqHK37GETkREEvFY3+2RBHxkmDiXAg4BbkoQTxMSgN4yDxsV2GaYWFbA5hXk/B1cSBmHSImINN6G2OltOS/6vWUBsN8AMW6DPdePsSRxsDKl3Yas6E0J69SDrYp4+wBxLA/cmTiWwcpfgNHtNqSIiLRuPcIny3mXecCe2IS+TwC3Z4rjtoB2bcd6CevUt9wAvBNbQbAscGumOAYr56IkQEQkinWIt747tMwBXswcw/erN21busibhN2KJTu5v/OBytmk3zpaRKTR1sImiHlerO8BDsB32+BcZe6iNkrla5HqkbrcAxxI2KFR/csfUBIgIuIixvruu4CVFr3/blgHmrszCilfqNa0lY0HHnKKPVe5H1tBAPYb8EwCziLt+REiIo2zKv7ru+/AZo/3tSf2LD93p1SlnES8nfWGUtJkzHbLQ8Dq/erjnQT8DiUBIiKVrAzch++F/1Zs9vhA9gbmO39ezDIZe4SR0+rApeRvi3bKI8CrBqmPdxJwOukOkRIRaYQVgbvxvfDfjM0eH8p7saV9uTupwcocbC/6g/HbQtfDm4DT8D+Bz7s8jk0mHYp3EvBrlASIiLRkOWyY3vPCfxPDd/699qWsJOAZbNvgfYAJLdYhp42xrZIvoazHKk9jjy1asSu+k0N/hZIAEZEhLYP/+u7rsDX67TgIWOgcRzvlCeDbwJbkeb7vZRlgf2y3vJxJVTudfy/vkYDTqPd3KSISzQTgenwv/Ndg2/FWcRC23W+qTmoO4YfilGxl4AjSb+DzLLBpxZi9RwJ+gUYCREReZgLwX3wv/FcR/pz8cOInATcDh9L+KEWdbQ2cgu9piAOVKdihUSG8RwJ+gkYCREQAGAf8C98L/7/xmyT3KefYesvV2N1+J3cGywHHAs/h377PA693itN7JECPA0Sk4y2BHfbieeH/F7ZRjafPOMW2EDtC1qtjaopx2OMBr90eZwFvdI7ROwn4oXN8IiK1sQRwOb6d/z+BxSPFe3RAXAuwu741IsXWFKOBDwGPUb2tZwLbRYrP+3HAiZHiFBEp1hjgAnw7/4uI1/n3+kaFuC4BXhM5rqYZg40IPE97bf0CsEPk2LxHApQEiEjHGAGch2/nfz4wNlH8R9PaxMAbid8ZNd0K2KS5VnZonAK8OVFc3iMB30wUt4hIVsfi2/n/FbtjTGlX4M5B4rkXeD+a5OVpA2yJ5EB7M8wHfk/6xyveScCBacOXutEFRepuVeAB/Ibq/4xt3zvf6f3atRWwDXa+wBRsKeN/sREC8bcq8Fbs6OOFwMPYPJKnM8WzG/Yb9Bh9moZtUzzN4b1ERIrjNZu+B7sjHJ02fJFX8BwJ+HTi2EVEkvkTPhfKs2jmbnlST15JwEWpAxcRSeXfhF8kp2OTw0RKcjrhv+27k0cttaE9pKXupjq8x1LYev/lHd5LxMNRwAEO7zPH4T1ERIr0RfzmANwBrJg2fJFXOAq/3/SZiWMXEUlmLWAufhfMe4BVktZA5CWenX8PsHfa8EVE0joB34vm/cBqSWsg4t/534oe84pIw43G/wCgh4A1U1ZCOpp35z8d2DhpDUREMlkSnxUBfcsjwNoJ6yCdybvzT3F+gYhIUZbADsrxvJhOAjZKWQnpKJ/Hv/PfMWkNREQKsTi2pM/zovo0Gk4Vf+r8RUScjcX/dMDJwKYpKyGNps5fRCSSMcBf8L3ITgW2TFkJaSR1/iIikY3G76yA3jINO7FPpAp1/iIiiYwEfovvRfd57MhekXao8xcRSWwUtiWqdxLw+pSVkFr7HOr8pTBduQMQSWQk8Ct8Dljp9QTwWnwOJKqbFYH1gAnAuEX/ueSif56HdVAvADOxZGkytsNiJx5OsyXwH/yOm34R2AO43On9REQabwTwC3zvxE5OWoP0xgDbA1/CHqVcj3XoVdpqIfAw8A/gB8AH6IxzFy5Cd/4iItl1Aafgd0F+EZiYtAbxbQV8GdtP4QV8E6aByj3AT7CDa8YnqF9KqwPdqPMXESlCF/Bj/DqwJpy4tgZwJHAf8Tv8ocps4GxsiNtryDynD6LOX0SkKF3AifhcnI9MHLuX0cCBwDX43aV6lseB47D5BnV1DOr8RUSKdDzhF+hPJ486zFhsMuQD5O/kWylzgFOx4fS6ORZ1/iIiRZpI+HPuupy8thg2WjGZ/J16lTIXSwTqNHHwAMLq/I/0IYuIdIbvEHaBfpB6PKt+KzbZLncn7lFmYXfWYzwbKJJVsdUPVevaDWyRPGoRkYZbhfC7//2SR92eVYEzyN9pxyi3Adv5NVU0fyCsnuenD1lEpNlClwP+h7I30zqENMv4cpZubE+BkkcD1sRGLULqqa2nRUScrIvtWBfS8WybPOrWLAmcRf7OOWW5EVjHo/Ei+RJh9bsgfcgiIs10GmEX5D+mD7klm1Of2f3eZTqwb3gTRjEGuJfqdesGNkwetYhIw6yEbThT9WI8lzLvNvfDlszl7ohzlm5s7X2J3kVY3U5JH7KISLN8j7AL8fHpQx7WJwibbd608hPs7IfSXEX1Or0ILJc+ZBGRZpgIzKD6RXgqsHTyqAfXRfhmM00tf8H2PijJmwir05fThyxNVfIMZpEYPgucEPD6bwJfcYolVBfwM+DQxJ87BbgbO973fuBRXjr+dwZ29PJSi8p4YH3s6OD1gQ2wLYhTuQQ7V2Buws8czjXAGyq+9nFgLWy0R0RE2nAX1e++5gArpw95UN8mzZ30dOBc7DHDxoExjwN2xTZgujFR/OdgSUkp3k1YfXZJH7KISL29kbAL76npQx7UJ4jbaS7A7p4PwDrtWNbHHmE8FLk+JU2g68JGUKrWpdQVKCIixQrZEa+kZVgHEO/0vlnAD0l/8M4I7FjlmwJiH658NVlthncw1esxF1g+fcgiIvW0JDaLuupF9y/pQx7Q1oRtYDRYmQ18izJmmb8duIU4ScDeCesxlLHAU1SvR91OoBQRyWYvwjqO7dOH/AoTgYfx7xQvxSbnlWQENtLxHL51nQasnbAeQzmW6vW4KX24IiL19BuqX2wfJP+KmS7gb/h2hs9iE9JKtiL+9b4BuwPPbS3CHuVskj5kEZF6GQE8TfUL7bHJI36lz+HbCV4JrJayAgG6sCHvufjV/6SkNRjc1VSvw7cyxCsNkvuuRqS/JYAV8N3FbRPgvIDXr4+td89lPeB2/O5ajweOon5rybcC/orPUswebFOeqx3eK8QhVF9d8hiwg2MsYEnWM8B85/cVERnQq7CtW5/A9y7Xo1wfr9otuxCfuiwADkscu7dXEXaoTt9yB2k3JRrIRMLOpYhRurE9Gj6LJeQiIlF8lLIPsDkiXtVb8l586jGXcmbAh1oeS8w82qWE2fRnk/93Plh5BNgsWs1FpGMdTv4L3FBlIXZyYC5LAU8OEFe7ZQHlHpNb1QR8lgrOAFZNHHt/oStUYpfphO8AKSLy/22MPWfMfXEbquQe/v8G4XXoxjadaaJVgf8R3kanpQ68nyXxneAYo9xOWdspi0iNnUX+i9pw5bhotR/eBGzNemgdjk0cd2rrY3eoIW00l/Q7H/Z3Ofl/78OV/aLVXkQ6xhjCjuRNVd4cqwFa8JUh4mq1XEpn3LXtQ3hb5V4W6L3MM0bRGQQiEmw98l/MhivzyDcDehy2FCsk/sfprP3if0RYe80m70mPGw8SV0nlrmi1F5GOsS35L2bDlRuj1X54RwwRV6tl5+RR5zUWuIewNsu9sc4j5P/dD1UmR6u5ZOG52YpIq+rwu7sm42f/X+Drz8SO8u0kc4FDsY6qqgPJ+8jkwoyf3Yo6/N2KSOG2I//dzHBlm2i1H9pmbcQ4UHmevEPZuYUc+dwD7JI+5P9v+yHiKqE8G6/qItIpSk8Act6JnThEXK2UL6QPuSirEraz3u/Sh/wylxD2/SsBEJGilZwAPEW+jWFGAZNaiHGw8hy2przT/YTqbfgCedtwdexZe+6/AyUAIhJFqQnAA9i68lx2GCSuVstR6UMu0uqEbayTe9fEDYGHyP/3oASg4TSpQ8SOCv4a9vz9voxx7Bjw2tnYna/YEsizA14f8j14uAd4LbYT5DOZY5EG03HAksN2VD+G9U78DnDpxobce0+Xy+0a4A0VX3sWsL9jLHW3E9VXQtxP3pGgvkYAG2BnUnjdsJ0EbFThdc/RWXtLiEgEIY8ArsoQbwrjsc2HqrZLztnrJRoBPEr19lwjfcjJXIMeAQh6BCBSijdR/Wz6p4HLHGNpgm7g9wGv38ErEJFSKQEQKcPrA157CXZ0sbzcxQGv3cItCpFCKQEQKUPIM2fd/Q/sGmxZXxUbeAYiUiIlACJlWC/gtZe7RdEs86i+pXMpkwBFolECIJJfF9U7nKexZW8ysKqHOq2OTcwUaSwlACL5rUr1zuZez0AaqOq+Dl3Aup6BiJRGCYBIfqsEvFYJwNBC2ifXltAiSSgBEMkvZO/5B92iaKb7A16rcxWk0ZQAiOQX0tFMc4uimWZgewJUoQRAGk0JgEh+IR3NLLcomqmb6ksBlQBIoykBEMkvZLb5TLcomqtqGykBkEZTAiCS36iA1y5wi6K55ld8XdWtmUVqQQmASH5Vh6hBa9VbUfVOXqMr0mhKAETyC+loNEw9vKpJkhIAaTQlACL5KQGIZywwpuJrNcFSGk0JgEh+IR2NNqsZ2moBr53hFoVIgZQAiOQXspZfh9YMLeSQpefdohApkBIAkfwepvpmNSEdXCcIOdb3AbcoRAqkBEAkv9nAYxVfuy7Vn3F3gqoJwCzgKc9AREqjBECkDFX3rF8M2NIzkIZ5Y8XX3Q/0eAYiUholACJlCDm0Zge3KJplZaqPAIR8HyK1oARApAx3B7xWCcDAdgC6Kr72Hs9AREqkBECkDFcFvHZ7YBmvQBpk94DXhnwfIrWgBECkDHcBkyu+diywj2MsTTAO2LPia+cA1znGIlKkkENIRHJYDr/Obh4wCbiF6gfGeOkB/k31un0QONUvnNp7D5YEVPEfbGVGbmsAhwE7Yxs+ea320O6RIpLNdliHV0qZDpwILB2z0i34GNXr0E3YmvemuZzqbXlMhnj7+zQ2EpH7b6NveTZqjUWkI5SWAPSWx4HXRqz3cNYdJK5Wyy/Th1ykrQhrx23Th/wy3yUsfiUAIlKsUhOAHuA5YK14VR/W9YPE1UqZhw0bd7q/Ur0NH6H6ygEP7xsgplKKEoCG0SRAkZdblrx30r8NeO1o4HNegdTUpsA7A17/W6yzy2Ec8P1Mny0ikkTJIwC9peoOcqGWx+7kq8Y9H+sEO9VFhH3vOedRHDBEXCUUjQA0jEYARAa2b6bPfRa4OOD1o4CTyTuMnct7gV0DXn89cK9TLFVoKaeINN625L+bGa7cFK32w3vHEHG1Wt6XPOq8lgSeIKzNDk4e9UvGAS8OElcppeo+FSIi/9865L+YDVfmAUvEaoBhdAE3txDjUGUa8KrEced0BmHt9Th5T1V89yBxlVTuiFZ7EekYo4HnyX9BG67kmgcAPrPB/4u1ddMdRHhbHZE86pf7FeF1iF3OjFZ7EekovyH/BW248rVYlW/BSF46kjakfDd14IltCrxAWBs9Q77RHrDv+pkB4iqtvCtWA4hIZ3k1MJf8F7WhyrXRat+a/8OnHoelDjyR1YDHCG+fL6UOvJ86rIq5gc6cWCoikRyEbWGb++I2WJmP7QuQywhsGD+0HguxvfGbZAJwK+Ft8wh57/4Bfkz+3/pQ5Tls3o6IiKv3AzPJf5EbrBwUr+ot2QrrwEPrMRvYI3HssUzERmc8vt/cbTIW62Bz/84HK3ejMyZEJKIVgW9jF5v55L/o9S3/iFjvVp2CT10WkD+hCbUSPnf+pXy3+5D/N96/vAhcARxCZ0wi7Vh6piOlGYEN73ragOrP8+djR7Hm3AVtGWyDmuUd3qsHOBI4YdE/18mGwIX4LG+cDWwCPOzwXiEuAHar+NrHgM0cYwEbbZrh/J4iIlndQvU7os9kiLe/3fGdL3EeNpReF3vhu3T042nDH9BKhI14nZA+ZBGR+vk01S+0d2WIdyAn4jvU+yDw+qQ1aN8SwM/wrfe5SWswuC8QVo/cxxaLiNRC6N1Wzk2Beo3Gb/Jbb1kInAoslbAerdodG6L3rO+j2COV3EYRVrfH0CNcEZGWnU/1C+5fMsQ7kLWwbX49O8Ue4EngA9imNLmtC/wJ/zrOBbZOWI+h7EdYXXRssIhIG/Yk7E65lOVQb8YmsXl3kD3YXekh2B1qamtjoxExVoJ0Awemq8qwriOsPq9JH7KISH2Nwg59qXrRPSV9yIPaC1vWFyMJ6MGGyr9G/A1gxmAH4ZwXuT6fi1yPdryZsLr8J33IIiL1dyzVL7zzsCH4UhxKvA6z753zVcDngc2xZZqhlsY6/Z+RZhOc0obLQx5F9QAfTh+yiEj9rUbYEPPP04c8pC8RvwPtW54D/gp8CxtS3wpYhVdOIhyJLTNcB9gF+CTwU2xr45h3+v3Lb/BJWrxsRNhyzhnA+ORRi4g0xNlUvwDPwzakKclHSdupDlWex3aSyx1HD7ZssrSZ8n8krE4/Sh+yiEhzbEvYRfjS9CEPa0/K6Xhzl27sUU9ptiDs7n8BOpRHRCTYfwjrZEo8Xe8txFkiWKcyj3LPO/gnYXX7ffqQRUSaJ3Qd9uOU+Sx2Dfw3C6pLeRzYLrwJo3gz4fUrZQ8DEZFaC92JrQebCFeiUcB38D07oPRyKXaqZIm6sMmPofUTEREnHyXsojwHeHXyqFu3J3aKYe7OOWaZC3yR8ib79fUewuupff9FRBwthm2BG3JhPid51O2ZCJxEOasEPMuVwMZuLRXHMsAkwup5fvKoRUQ6wGcIuzgvxNbBl25zwrefLaVMAg6g7Lv+XkcSVtdu4HXJoxYR6QDjgVmEXaQ/kDzqakZiM+QfIH8nXqVMB74JTPBumIguJazOF6QPWUSk+bqAHxPeMX0ldeCBRgB7ADeSv1NvteP/DmUc49uu2wmr+xxgt+RRi4g02AjscB+PDuqoxLF76cImCl5AnBP4Qstd2CE+S8ZqgAQuJ7wdlASIiDgZAfwSv47qfWnDj2IZ7Bjgq8nb6U/FjgTePm51k/kaPu2iJEBEJNBI4Az8OqwFwMpJaxDfesAngD8R/7S+Bdga+W8DOwOjE9QvpdXx26JZSYCISEWjgLPw7cD+kLQG6Y0AXgscjp3mdynwGNXaajZwK3YY0zeAd1KvCX1VHYHf701JgLipwzIaEQ+jgTOBfRzfczJ2wMuTju9ZF+OAtYHlsJUU47AjgSdg+/HPwo6unb7onycBj2LL2jrRt7ENizzMBfYCLnR6PxGRxhoD/AXfO/8p2Np6kVadiEYCRESSGQP8Fd/Ofxrw+pSVkEbwWnbaW+Zij1FERKSfscDfUOcv5egCfoKSABGRaJYg/Px1df4SQxd+e1AoCRAR6WMJwrdgVecvMY3AdzmqkgAR6Xjj8Nl9TZ2/xDYS+B1KAkREgk0ArkWdv9THSOD3KAkQEalsafyPu1XnLymMxnelipIAEekYiwH/wbfz1zp/SWkMcD5+v985wHZJayAiksHJqPOX+lsMuBi/3/HT2MiYiEgjrYvvUbbq/CWnxYHL8Ps9H5c2fBGRdI7G72KpZ/5SgiXwW8nyCDrzRUQaymunP3X+UpIlgH/h89tePnHsUhMjcgcgEsjrGec/gRud3ksk1IvAr5zea6LT+4iIFOUc/B4B/AwNl0oZ9sKOVfb4XWsioIg00ifxSwB6gJ+jkTHJy7Pzvy1x7CIiyawIzMQ3Ccg1EjAOOAK4GngOeBbb3OjLwLIZ4ukE2wK/AB4ApgJPAGcDb84Uj2fn34P9nkREGutIfBOAHEnAdtiM7cHimQ4chU0Ok3CvYfiNd35I2t+Ad+d/J3YktohIY3UBf6C+ScAewOwWY3oS+DC2h7y0by3gt8BCWmvvH5HmN+Dd+U8G1kkQt4hIdt6HqvSW2HMC3o5t29puXPcA+0SMq2mWBb5D64lW33IqcZOAGJ3/phHjFREpzmjgT9RnJOCd2MEtIbFdAewQIbamWB74OjCDsHaONRLg3flPAjaKEKeISPHqMhJQ9c5/sHILcAAwyjHGOlsHOAl4Ab829h4J8O78nwY2doxPRKR2Sk8CvDv/vuV/2KTICQ5x1tEWwBnAAuK0r1cSoDt/EZFISn0csAfhw/6tlCnYHfCWAbHWxTLAx4D/Er9de7B2DfkNeHf+TwCvDohHRKRxShsJiHnnP1S5BzgWmwHfFCOBnbC7fc9h/lZL1ZEA787/MexETBER6aeUkYBUd/5DlYXApcDHqWensRSwJ3AKNsKRsy17aH8kwLvzfxQt9RMRGVLukYBcd/7DlYexO9l9KHO/+JHYM/0jgUvIn0ANVE6htSRAnb+ISCa5RgJKuPNvpcwDLgB2ba05o1oPOA3bmjd3u7RShhsJ8O78HwZe1VpTiogIpE8C6tL59y/nAEu20a6evoBvZ5k7CfDu/B8C1mirRUVEBEiXBNS18+8t1wCLtd26Yb7pFHuu0j8J8O78HwRWr9CuIiKySOwkoO6df2/5btUGrmBnoDtSPVKW3iTAu/O/H1itcuuKiMj/FysJuIBmdP492L75y1Vt4DZdnahOKcr5+Hb+9wGrVm9aERHpL1YS4FHuAfYFbswcxwcqt27rliff3f8C4CzKWFY42O9g5epNKyIigykxCXiQl+74eoeT78wUy7HVmrUt22So10LgTGD9RTFsTnlJwN3ASlUbVUREhldSEvAYAy/xGgG8H7g1cTzfa68pK9kxYX0WYHtCDLRvfklJwJ3Aiu02pIiItC/WZkHtlMnABi3Euj3wd9IMmzclAZiBTdB71TCxbAY8lyCeocrdaNhfRCSpnCMBz9L+Ua4bAycT96617gnAPcDhtLevQc6RgNuwOREiIpJYjpGA5wk7tW8MtvTwbPxXINQxAZiNtcVOVD+1L8dIwK2kW3UhIiIDSDkSMBPY1jH2ZbG5Amfi04HVJQFYCPwLOBg7NMhDypGAm7HvTkREMkuRBLwIvCViHUYCOwTGWHoCcCfwJeJtj5siCbgRWCZS/CIiUkHMxwFzgd0T1GF8YJylJwApjjWO+TjgJtT5i4gUKUYSsAA7gjcFJQA+YiQBuvOXrEblDkCkcAt5aTe89zq8XzdwAHbintTHrdikwkvxeVZ/E3b+wTSH92rXaCwxzOVFbARMRKQWPEYCuoFDEsetEQBfHiMB1+A3UbFVb8V+v5Mo49Cl54DzgPdQfaWGiEgyoUnAZ9OHrAQggpAk4GrSdv7jKWeXy8HKVWjjIxGpgapJwBdzBIsSgFiqJAFX0d6GRKHGYqMNuTv4VspDaA8EEamBdpcIfjVPmIASgJjaWSJ4JTAucXwntBhbKeXPcZpBRMTXSOD7DP08dTbpn/n3pwQgrvWw7YaHivFcYPHEca2I/f5yd+rtlpAdMUVEktoCO0/+eV66iD0J/ARYM2NcvZQAxDcG+ChwCfA4MBW4F/gNVrccPkL+zrxK+U6MxpCBaRmgSJibsK13AZYG5gMv5AtHMpgHnLKolGKz3AFUVNe4a0kJgIif53MHILJIXSfUrZA7gE4yIncAIiLirq7X9rrGXUsaARCRoSwNrB35M7QGXCQDJQAiMpSDFxURaRgNt4iIiHQgJQAiIiIdSI8AZChd2MYc22Ibi4zMG46LKcDtwOXYkj0RkY6kBEAGswtwIrBx7kAieRo4BvgFtgGJSAxLA+/CNowaCTwI/G3Rf0p6Y7HjvffCrm1jsE2b/gT8Cts9UaSjfYEyjgxNUc7C9vZvqtCdAEsvJewEOJBRWII5k1fG3I3te78HqOtjAAAgAElEQVRmxM8/Z4DPrUO5NUZjLPIW4LEhPvsRYOuIn18cjQBIfwcAx+cOIqH3AbPIv2e/NMd44K/AWwf537uAdwNvBPYErk0UVyvux+6EQ6wEfMohFk/7Y1szD9XnrYkd2vQu4OL4IYmUZSK2j3nuu4AcZVeH9ivRSGAB+ds3VlnDr6lcLIHNL2k1/lnAmyPEUXUE4CKHz35Nxc/uIc4IwP7YfJ9WY5gD7BYhjuJoFYD0dTCWBHSiL+UOIJKF2LBnE3UDk3MH0cfiwN+BHdp4zTjgAuIkAWKHIp1Be6PdY7ETHAcbwWkMJQDSV0dkvYN4I7Bs7iAiuTB3AJHcCMzNHcQivZ1GldP/lATE8THgVKr1c4tjkzXbSeZqRwmA9LVW7gAyGkH8LW9z+RHNXPIY+qzayxhsFnlIAq0kwNfngZ9i8y2qWgI4H5s8KNJ4z5H/mW7O0uQhv2PI376e5QbKmMQ8Fhth8arXDOANDnH9reLn554DcDfh+40cHfD5A5WZwHaBMYkUTwlAc3UBJ5O/jT3KXZRxgNAY7A7Ru35Vk4ClsGHvWwI+O3cC0IPN6zgNeBvtJ3nHBX72YGU6tiGaSGMpAWi+9zH0WuiSy4vA97FldrmNxpb6xaprO6sDtsCedQ+050C7pYQEoG+Zsqhur23hc2N1/r1lOg3bJyDk+Yg0z3M0dyJcK3YCLssdRAIjsUmPmwMrZI6lFTOwoeErsItwbqOBP2Jr+WOaid0FD7RPwOLA+4FPAJs5fubFiz4zxGuA2xxi6e9fwE+wxKvvnJYu4CTg8Aif2d/zwM7YBFSRRtEIgMjQRpF2l73+jwOWA75FvL/V0kYABipPYnNaVsYm754a+fP6l6lY8izSKEoARAY3CrvzT/27nIFtVHUsNgIS87PqkAD0lrmEzXcIKVPwHXnJooRZtCIipRsJ/BrYN8NnL4lPx9w0Y8jXCS+D7fi4E3BzphiCaR8AEZGhjcA6/w/kDkSKMhFLzDbNHUhVGgEQD9sB9+QOAjvEo5TNYaQZRmC/qQ/mDiSRBYW8R10sD1yC7QB5d+ZY2qYEQDzMAKblDgJ4IXcA0igjgJ8DB+YOJKGnHd7jGYf3qJMVsdVDOwD3Zo6lLXoEICLySl3AKcBBuQNJ7DqH93gOeMDhfepkJWxOwHq5A2mHEgARkZfrwvaR/0juQBKbhR1o5OEMp/epk5WxJGDd3IG0SgmAiMhLuoAfAx91ft9ngMed39PbcdhGNx5+SPnHUMc4IGtVLAlo6sFi0mBV9wHYJEewA9iXsLW92gegs/XuKOe9ZnwysDHWKZS6DfPZ+N8QbkH8fQuqllux009viPT+jwKvCm5BqY1XAx8GjsTuHuq4U5QSgHpZF3tGfSR2CMyWaHvvECfi3xH0dv69SksC5gDfIPwEvsFshG2bm7uefctd2Ox9sPX8sTYT+h+wZmD7SeE2xYZ8BvoB3I4tD6kLJQD1sAVwJQPX4V7gndkiq6/v4d8B9O/8e+VOAhYA/wGOAlav2mBtGAG8HVtOmTv5uY9XniS5HHatjvF5D5GmjSWDXRn+BK4F+D9PjEUJQPk+iN21DVeX49FoQKu+jf+Ff7DOv1fqJOAp4JfY38gyVRrJ0auxa+LZ2Ja6qdrgYWC1QWJaEVvHH+NzHxjic6Wm3gbMpvUfwZF5wmyLEoCyfRhYSOv1+Tma6DucGEfIDtf594qdBMzFOtm3EW+IP9QY7FTFvwLziNcWj2HP/IeyMjZCEOPz78eSDGmAdjv/3lJ6EqAEoFztdv695TQ0EjCYw/C/0Lfa+feKkQQ8CByBDW3XyfJY3N6d8FPYqEMrVsOG7b1/Fz3AVZSbiEmLqnb+vaXkJEAJQJmqdv69RUnAK22A/x3nZKr9LXglAXcB+1P/TmYk8H7gTny+k43a/Pw1gUccPnugckCbsUhBQjv/3lJqEqAEoDyhnX9v+SFKAvo6E98Le9XOv1dIEnAbsBfNe9wzAqvXHVRrlynY0cRVrI3t2+CdAFxVMR7JzKvzLzkJUAJQFq/Ov7f8IG34xRqDnV9RSuffa21s/Xirn/sMcCj1v+Mfzijgk9h5I622zfPYstgQ62GPDzwTgPnY709qxLvzLzUJUAJQDu/OX0nAS9alvM6/1yoMvsSzt8wFTgAmOH5uHayIHcvczdDt8yCwmdNnboh9x55/g6s6xSYJxOr8S0wClACUIVbn31tOTFeVIm1KmZ1/X7sBv8M6s6nAJOBq4Gi0rnxrbO+V/n8jT2LtM8758zYFnsXnN9MDLO0cn0QSu/MvLQlQApBf7M6/t3w/VYUKNJHh7yJbKW9PHbi8zMrALsDuWCcd81HIZvjsWfBkxBjF0e60tuGKR+nGlr/kMhGbnfoi1eJvSgJwMq0vF4rhEHw6plbLt9JUq0jXE95+DwFrpA5cspgA3ET4b+anqQOX9qW68+9fUo4ELId1+n/HnimGxN2UBKC33AUcS/vLh0KkuvPvX76bonIF2h+f9lMS0HwTgOsI/63MZvjNiCSzlHf+/UvskYBlgcOBa/HtbJqWAPQtdwJfJu5Wnqnv/PuXr0esW6lGAP/Ep/2UBDSXV+ffA/xf4tilTbnu/PsXz5GAkcBOwBlUH+IfrjQ5AegtC4FLsFGTJRxjznXn378c61inuphI9TXm/YuSgObx7Pw/lzh2aVPOO//+xWMkYB1sqdCkBPF2QgLQt0wBfoTtJhci951//3JMYH3qaHn8ToJ7BA3xNsUE7MREj9/FlxLHLm0q5c6/f6kyErAFdrc/P2GcnZYA9JZubFRgD9rfZa+UO//+5eg269EESgKkL3X+HaSkO/+BOphWRgJGAe/FZ2ZzldLO4Scx7UO+7+pm4APA6BbiLO3Ov3/pxIuWkgABdf4dpeTOv7cMlQSMwJ5J/y9zjKsM29JpvJX839djwEEMvibZs/PvBj4O/ClCPUrZmyIlJQGdTZ1/B4nR+c/H7gS9L8YDJQHvwO9iFVKeppxDZpYh7aOPocrd2DnnfdvGu/M/dNH7jiZOEvD5llq9WZbHd2Jgk3btm4ithlkbS/qbtKOdJvx1kFid/77AYsCFzu/de8E/AtgOO1HK+/2rlpPbbPvYziF/m/Qt1wFvIV7n3ytWEvDZVhq9YTyTgDuBJdOGX1kX9jjvIOB44C9YIjuLoes4E2uvPwPfwea3bEI5NwbDGY8tjfb4vtX5Fy5G5z8P2LvPZ4zFNtjxvhiX9tx4JuUtfVoL2zs9d9vE+u4G6vx7xUoCPj10kzeSZxJQ8tkLK2G/pz/if+jNM8C5wCeIu49GqF/jU191/oWL0fnPxYZ6+xuDZdDeF+NSSjc26a1E21H9XIOSy1Cdf68YSUDurapz8UoCZgFLJY59KIthK1fOxm5eUvx2F2KHGR2B7UJaivXxSc478XFZrcTq/N85xGeOprwhaY8yh/J3tVoROAl4nPzt5VFa6fx7xUoCDm/x85vEKwnYLXXgA1geOA6YRt7f8gvYPhprxq1uS44mvD7q/AsXo/Ofg03EG84o4A/On52rzAHOwrLmOlkCeBPwFeAfwAzyt2U7pZ3Ov1esJODjbcbRBB5JwMeSR/2SVYAfYh1v7t9y3zIPOB1YL17Vh3XuAHG1U65NH7K0I0bnP5v2MvqR2Nneuf/g2i0LsROwvgfsiv8527mMBnbA6nUX+dt5qFKl8+9bzxhJQM7OLJeLCGu3g9OHzChsyH16hXhTlnnYaF2OyZKhE7a7gdclj1paEmOHvznY87N2jcRvsknMMhvb3e4IYNUK9ayjNbH6Xk3+9u9/cTkssG4jsVGb0uKqk9cR3mZvShzzdsBtDnGnLE9gK6lSOs0h7vMSxywtiHHn/yKwc0BMI4BfOMfkUeYBF2CT+sYH1K8JNsKOyPWeEd1uCbnz7y/WSMBHnOIr3e8Ja6sp2KTgFLqwTZwWBMacs5xBuuuQx/bh3dgW7FKIGJ3/C9hOc6FGAKc4x1a13II90y1pVm4pxmBLOy8m/fJLz86/V4wkYCG25rvJViJ8pvzxiWJdgfBHFaWU+0gztD4auN8h3r8liFVaEKPzn4Vt5uKlC9tAJ8cf1ixsFGIrx/o03UbAqcQ7TrlvidH594qVBJS+IiTEUYS1z0ysY47tdcBTgbGWVmbz8v1VYtnPKd7XJ4hVhhCj859JnOd3XdjM3FR/TI9jQ4MTI9SlUywLfJV4Gw3F7Px7xUoCDowcdw4jCD9n4+sJ4nwL8HxgnKWWbuAzbi01sC7gRodYz44cpwwhRuc/A9g+ctwnOMfcv9wBvJ/WTqmT1iwFfBl4Ft8L3UcTxR8jCVgAfDBR/Km8mbA2mUr8hHtvyj/QzKN8w6vBBrGLQ4zz6ZyJ00WJ0flPB96QKP7vOMfeA9yKXRxGJKpDJxqPjQiE7iuQsvPvFSsJ2D9lJSILnSH+hcjx7U66nfxKKEf5NNugLnWI8ZjIMUo/MTr/54FtUlYCy3A9Yr8P6/jrchBHE6yA7Ww2l/a/rxydf69YScD7UlYikjHY7P2q7TCFuPtmbMXwh/Q0scTcjXJ7h/gejBif9BNjnf/zwNYpK9HHkS3GOFB5dtHrUy03kldaFzif1r+zEnbWi7FPQBNGAvYkrA1iDllvSFhy0k5ZiM2DuAjbzOwUbFXDqYv+/WLg4UX/vxTxLADeFd6Eg/qPQ4ybR4xPFolx5z8V2DJlJQbwZdqLeR62br3T1++XZA/sTmC4C1kpS+hijAT0Ho9dV2dTve5zgJUjxTUOO6I3Zid7A5bA7IBto92KxYA3Ynvr/5u4CcF0LNmOYR+H+L4VKTZ3dR0m3h27YI11fM8p2ESQmx3fs6rDsCNEh6vfldgd5N2xA5K2jQU+C3wK20u+r1uBTwJXpQ5qCKOxMyv2cnzP3scB5zq+53CWx+4QN8EOhKo6B2YPrFOr4lfES+5+Q5wVF9OAny96/3sc3m8NbPLxYcDqDu/X3y3YHK05zu87EngAO1a8qvnY45mqr30Sm4/wS+yRrvSxG/53/s8Cm6WsRAvWxw7KmMnLY+0GrkPP+etiDHZn9GHgAGDTvOEMKcZIwDx8k4rBjAW+jf8jwSplk0h1PDBCrFOwUcdYxxWPXhT3QxFi/0mkmI+IEGuVMh+bID4yUj1rZzv8N2N5BnhNykq0aQz2TGknbD+ClfKGIw0XIwmYA+wYMeYlsWHn3BfsHuD6SHVcGd+1/t3Y3X6KTYrARlOOxvdEwm7i7NGyLGUtrTwT3eyxGOGbcvQvkyn7jkwkhxgTA58m3mlvf3GONaTEmqV+pmOMz2Cne+awIb6HFN1BnD1OQuaAxCiddPjWgD6Mb4NOwrZ4FZFXijES8IkIce7hHGNImccr53t42AG/8yiuIf9mNYthow9e7f65CDHu5hifR3mG6nNSGuHP+DXmU8AGacMXqR3vJOAfEWK8xDG+0BLjWNgu4Han+P4OLB4hxiq6gOPwqdcMYBnn+EZik/Fy/6b6lt2d61grt+LTiE9iE+xEZHij8dknvQf/Gc1jKWsnvA851w/C9yPoLWcDoyLEFyr0sKXecmyE2E50is2rfDVCHWvjJsIbcDLw6tSBi9TYx/AbfvZerrqmU1xeZQ3n+gH81yGuSyh7gzCPjnYK/nNM3ugQl2f5mXP9auX3hDdgE3YoE0nFs/PvAf7qHN+GjrGFlgec6wbwVoe47iHe5EsvI4ALCa/r5yPENckhLq9ymnP9auW9+DSikgCR4Xl3/j34D5GXlACc6lw3CJ9/MZvy9jcZzDKEr/J6AP/lcr8IjMmzdHQCMBK/yTBKAkQGF6Pzvw//5VolJQD7OddtIuFr0T/pHFNsuxL+PXgf4vYOh5i8SkcnAOB7CEYTzy8XCfUR/Pdxn06cO9FSEoB5+M9C/2hgTDdSzx3kQtffn+wcz+KUsbtkD0oAADsG02tHLI0EiLwkxp3/TOyY1RhKSQB+EaFuVwXGtG2EmFJYg2rHaveWZ/FPfH4ZEI9nUQKwyOZoJEDEU4w7/1nAWyLGXEIC8Ci2daynCdh1qWpMFznHk9pphH0n3ie6rkQZkwGVAPShJEDERx07f8ifANxPnGNpQ587vzFCTCmtQ9jv0Xs1ANgBT48ExKQEIAIlASJh6tr5Q74E4AnsYJslItUrZF38XZFiSu1yqrfBhZFiWhL4Gvb9NyIBKHFnqHbcDOyMbXQROglnJPDrRf/828D3EqmDj2DL1zyXTr2A3cFe6fieMZwN/LzN18zBthF/2D+clwk54e50tyjyOh07A6GK7bE1/N1+4QA2n+Wri8qa2JkK7SaBuwGfdo6r42kkQKQ9H8F/wl+qO/9eISMAxyeMsx0jCDvyfK30IUcxkbB5EGumD7klH6GgEYAR3m+YSe9IwFSH9+odCVASIE3VyXf+pVuD6gf2PIxtptME07DrelXreQXSZE1JAEBJgEgr1PmXLaTjutIriEJcHvBaHfjWgiYlAKAkQGQo6vzLF5IA3O4WRRlC6qND31rQtAQAlASIDESdfz0sF/Dae92iKMM9Aa/13pmxkZqYAICSAJG+1PnXx/iA1z7iFUQhHgp4beknIBahqQkAWBLwdmwf8lC9SYC2DZa6ORT/zn8W8DbU+ccQ0nF5XOtKMgOYX/G1IYlUx2hyAgBwPbALfknA6SgJkPp4DfAj/Dv/twNXO76nvCSk43rBLYpyTKv4Oo0AtKDuGwG1ojcJ+Ce2x3aIkcBvgOeAiwPfq9ONBJbCvpP+iejzWEczL3VQDfMNYIzj+6nzj68ndwANsSB3AHXQCQkAWBKwIz47Bo4CzgQ2wBIBGdzqwKaLyjrAqxb9dyvQ2vcwE2vjJ7Dnm49gW53eiZ0vrz/ywS0L7O74fi8Ae6DOP7aZAa9dMvD1JVqq4uued42ioTolAYCX5gR4jAQsCxyObQkpL1kP275zR2xHuBUC32/JRWUtXnnAyRzgOuCKReW/aMSgrzfgdyyq7vzTmRXw2gnYVsVNMQZYrOJrlQC0oOlzAPrznBOwj8N71N1YYC9sbsTj2F35KcC+hHf+w1kMSzK+Bvwbe1b4T+wksNUjf3YdrOr0Pur80wpJAF7lFUQhQrY1rjp3oKN0WgIAfknA+sDo8HBqaQvgJKzT/xNwALBa1ojsUI6dge9ijwquBg6h+hBi3VWdPd2XOv/0ng147QZuUZQhZDe/yW5RNFgnJgBgScD3At/Dc2Z1HUwAjsQ2G7kR+CSwfNaIBjcC2A5b/vY0cNaif+8k9wW+Xp1/HvcHvPY1blGUIaQ+IZsIScMtBjxG2MlnTTl0YzjLAsdimyp5nhyXo1yNTWTrhORtNGEnZM7AvvuSNfE0wDWoXqdH0ocb1eVUb4uNMsTbCp0GWICPEf6c+O8egRRsBazjfwib7DgxazQ+tgP+BtyCPbbwmiRXovnYfIyqlgQ+5RSLtO5x7DjgKtakOccBLw5sW/G184EHHWNprE64E+pvPNaphUxSmwtsQjN/ZKOBw4DjaP5uWrdgdb0udyCRLI09slmx4uunYxPLSp1RvSFwd8XX3kf7h80swJ4tXwOcj61EieEGYMuKr/08cIJjLLm8Bzin4mtvB17rGEtfY4HdgO2xv6t299lYG5tDVcXPsXlNEuAowoeSv5k86jTejK2zzz1Un7J0A2cQdghLyT5IWPsckz7kloU8AggtTwMfiFSvEwLiujNSTKn9jeptcGKkmPbHvvdcvzn3RwCdZmnCn2U/CoxLHXhkK2GdYDf5fty5yxTgCJr3WKyLsGepUyh3W9WcCUBviXEz8I7AmLaJEFNKq2F7elStv+cGWL2+HhCPV1ECEOgrhH8JeySPOq49CJss1rRyGbByUIuWZxNs+LpqmxyZPuSWlJAA9AB7O9drAmHf13nO8aR2EtXrPhf/R5d7BMTjWZQABBgLTCLsC6j7H1Zfo7BJfgvJ/8MurUzG9opoklOp3h6PU+aeF6UkAPfiP5/q6oB4uoHNnONJZWVsEmTVul8aIaZbAuLxLEoAAhxEWOPPw7a6bYLVsYlMuX/QJZdu4Ds0Z6XACtikvqrtsV/6kIdVSgLQA2zsXLfDAuO5inpO8j6TsHof4BzPmoHxeBYlABV1YTNDQxr/R8mjjmNnyhzyn0rYsGescjnNWAIJNqGvajtcmyHe4axP/t9Hb/mQc92WxYazQ2Ly7gxj25GweUgz8J+f9bGAeLzLyc516xi7ENbw02jGLPH3En5RqVIexvZN+B5wMHawz9rYXWn/P9jFsLZeG3g9NtP6OOBsLInLkSTcSf6tjj2MA56hejtslT7kIS1L/otyb/lNhPqdFxjTNOqzL8BEbHO1kPr+KkJcfw+MybN8KUL9OsKFhDX8UelDdncY6Z73T8I67EPwP6BkPLATNjx/Y8I6PUkztlo9kupt8LsM8Q6lCzv9LveFuQdbHeRtN4e4/kv7a9VT68Kno329c1yLYcdg5/5t9ZaqGyN1tDUI6ySepdxlUK0Kuei3WqZhSwl3Iu2zx9V46YyC2HWcSv3PFBhP9VGAuZR3/sNPyH9h7i3rRKjfTQ5xnUXZy1tD9j3oLRdGiOvtDnF5ldso+zss1tGENXypS6BadSJxf5j/xnbtGpuqQkPYBtstK+ZjjhewDZPq7EtUr/+nM8Q7lDWA2eS/QPcQZ5e2/Zxii7U5TqjP41O/N0SI7TSn2ELLAmCHCPVrvC5su96qDf8M9d7058vE+1H2HqxTohWxRwSxhu+mA5snq42/icBMqtW93e1zUwjd7dCrXBKhbiOxbYs94juZsu4ij6Xcdl8c2wI792+qGzg8Qv06wlsIa/yjk0fs5yDi7Ox3NeVNBhvMCtidz3z822ESNlGxrn5E9bpX3cs8pv3J/7x2AdXPXRhK6M6AfcvZ5L+pGYuN1HnUZx620ZW39znFF1KeAvaKULeOcTrVG/8F6jvz/x34d3pTqe9WuetjG4R4/4E+hG2jXEdrUX1FRanLkVYFfkj4Ud8hJdbdWsje+P3LPcCmkeIczprAf1qIsdXy3UhxXhQQ0yzs2lClXI8laQfR/MPYohqHfRFVv8RSL3LD2ZqwnbQGKr+m/LPhh9OFrdX23gPhZmCJdNVwVXXm9VRshnTJxmMjNFXLo1Rrm2si1WctfP+uX8DmgqRaITAKmz8yw7EOjxGnk1yNsOXGn4kQk7Rpb6p/gd3Aq9OHHGxpwtfS9i0zgfcnrUF8qxO2zepA5ZdJa+DnnVSv87szxJvS76jeNrGWi34mIKbByt3Ydxlr5U4X9jsL3Yitf1kI7Bop5uMD4ppL2FHz4uQsqn+JF2WIN1QX8Gd8Lwze25uWIsY5CAemrICTUdj+BlXqe2aGeFM6hOq/hZ9HislrvfxA5TbsN+w1P2A8thPhrZHiPc4pzoHiDjkx9pxIcUkbxhI2g/Nd6UMOdgR+f1xnY7Ngm+6d+E0cmwVslDZ8F9+kWn2nU/5jgBArU30Y+EXizR9aFjucKUan2oON+p2BTYJrd0LjKtjOnb8j7PHrcOVK4p3REXodfVukuKQNITtoPUb9DoDZEr+17z+lnhP9qtoK2+zJo+3upH7zATamen33zBBvSpdQvW1ibtu6Lf7zfAYq3cAD2KjD97E9UT6GjY58fNG/nwT8A9vuO3Y8PdjcjFjHdY8kbNn449Sv72ikX1D9S/xGhnhDjMVvF7yvJY69FBvhN3P8pMSxe7iDanX9bY5gEzqQ6r+DJ4i7OdYexFneWnJ5FtjAo/EG8f7A+I6JGFsUKbZsHQVsj+3TnGom+UeAZSq+dgNs4426OAqf52FfwjbP6VRrYEeorhH4Pgux3/otwRGl8xWqJb7TgX2wi19ss4FHsDkLqYzF7jirru3/ODaiFstH6JwjYl/ETjKNdSrlSGwEr2qC8SJ27smzXgE1wf5UX06To/w3TjNEswY+z9p+kjrwQr0amEx4e15PvR6jbED+v71Wyx1YxzcqSku80jcCYn2K+HNpPkOcDb9KKjOIvxXuQYExlrrVcjYhO43lKnXbctFjc5DSDwpJ7fVU3ya3bzk4deCBHiL/31875d+kWW61MmHzaz6VIMYPYDvi5f5OYpSnib/t9hjClk/PphnHhbvxOuAhZenGdhKriz0Ir/OVlH9UaA67E75E8DnqtZPkyeT/G2y33Eya3dJODYjxmUQxvoP8WyF7lwdIs932YYFxHp8gxtpYhXr+EGM9W4phBLadZ0h9J2PflQzsOMJ/U99PHnV1u5P/b7BK+WGMxuhnLcIm3H0+QYxgk1nvDIizpHIe1edxtWMxwpZVTksUZ20cRf4fT5XyhRiNEUnvxKuqJeYuWk0xAriMsHaeBSyfOvCKxlPPWeXziHMAT39nBsT4JOnmLIwPjDV3mY8tL0wxQR1snlpIvPeiBOBlriD/j6hKWS9GY0TQhQ19htT128mjrqdVsSHckLaOtWNZDDeQ/++wSjkkRmP0sQVhu8P1EOe8+qEciN/+FqnKTaQ/afKXTnErCVjkAfL/kNotD0dpiThCNjnqAe4n7vrkptmPsPaejp3RUAc/IP/fYpXygxiNsYhH598DfDhijIOZiO1LEXKwTYoyDduBL8cmOv+oEO9ARUnAIk+R/wfVbvlZlJaI4yrC6vrW9CHX3oWEtflR6UOuJOQArZwl1oZEXp1/DzbRLJetCDveNlaZjU0+TfEIZzB/GCCuqkVJAPVMAOqy9/82hNXz9+lDboT1gTlUb/fJwOjkUbdvSdJsMetdYiQAnp1/D3buRG6vxfb5zz0iMAsbmShh1dWn8a1bxycBdUsA5gFLRWkJfz8l7I9Os/6rq3poTkkdQCvquBzQOwHw7vznkm4X1Fash81NeYS039O12EjIxOg1bN2K2EiEZz07OgmoWwJQl6N/x2Bry6vWUztVhZmAPaus2v51OSZ0OWwf+9x/l+0UzwTAu/PvAX7jGJ+nLuDN2KZtd+K/o+BcbMOmozb8Vx8AABsGSURBVLFdNksVY+VaxyYBdUoAFgLbxWkGdyHPZ+dQxnBb3YWMApR2FziU1xGWbKYuXglAjM7/aeKdXudtRWzS6/eBC7AJ3a0uDZ0D3IYlul8HdgHGpQ2/shHAufj/LjsyCahTAvDZSG0Qw1+pXs86TXIs2bKEbRN8aPqQK1sTOJ/8f6OtFI8EYHP8O/9p2NbSdTYG2whpM+xQt7dh+5DsjB1LvCmwOvXfTnwsfisC+paik4AYmy08RfkZ72Rs2ckfcwfSomWxdq2ybW83NvxWp6WOJfsB1fd3vwa7iNbJesCO2AhSzImMY6nerr8DPhjw2ZsBl+I7QjMN6yyvd3xPiWtxLOnd0fl9bwV2AqY4v2+wkhKAh7ADGWKZg+3I9S/gL4v+vS72w5arVHE5WvrnaUPg7oqvXYjtDDjNL5zGmAA8X/G1IQnA5ljn7zkx7RnsDvl2x/eUNMZh88K8E/WbscRiuvP7FqfqI4BjcgRbE6dQfQjqQ+nDbbyQHfPqsuQ0tQmkfwSwCf7D/k9iSaLU11LY0fDejwPOTVmJVtT9uU2nqHoW9ovAnzwDESDsmXPsc82lNUtjz3w97/wfw2bV3+P4npLeDOzxzS3O77s3NjJUDCUA5VuV6ucUnIdNWhNfv8dmR1ehxzFl+CK+Z7g/jHX+Dzq+p+QzDVvJcKfz+x7u/H5BlACUL2RCysVuUUhfzwL/qfjajYCVHGOR9o0gbNJgf/dhnf8jju8p+T2HTd67z/E9d6SgfreYQGRQbwp47eVuUUh/Vdu2C3ijZyDSttXx2xXzHuyi/oTT+0lZJmOjdg85vd84bLOtIigBKN/rKr7uAeBxz0DkZa4IeO1r3aKQKrxOZ7wVu/N/yun9pExPYknAY7kD8TYqdwAypJHYkHEVuvuP6zpskuUSFV67qXMs0p5nnN7nWOxxkOQxGtvjZHFsNUfMZeQLsN08Q83EHi0UQSMAZVsb+3FXcbNnIPIK86i+ycvGnoFI2ybhM1nvd8AbHN5nIJtgW/Jej43kPYRtz3sIsFikz6yLlbCD0Z4F7gJuxCZhPoyd7ue9YdWq2Iifx3kGl2KbszWW9gHwsyvV15yGzB2Q1vyMat/NfDT61l/qfQA+H/B5fcvzwFYVPn8wiwOnYZtGDbXXwEF03g3cGOBz2GY6Q30nN+A3x2NV4P5hPq+d8hanuIqlBMDPoVT/oa2YId5O8ymqfz9rZoi3ZKkTgLHYM3yPi/pUqs/V6WsCtl10q597C52zrHQ3bDZ+q23zMLBG4Gd6d/5nBsZTC0oA/BxHtbasuqWqtOftVL8YaITm5XLsBLgK1lF4XNyfA15TMQ6w+l9X8bP/ju8oREm2wh59VGmXkCTAu/O/DliyYiy1ogTAz6lUa0s9/09jbapfEPbOEG/JciQAYB2EVxLwDNXmd0zA9pUI/fyrgT2Ic8ZLattjiU1omzyKnWbYjpWw8z68Ov+bKfhEQG9KAPz8iWpteUWOYDvQeKpfFA7JEG/JciUAYB3EowGf37c8RXs7d4bc+Q/V4fwf9vuskyWxs0tCztoYqDxE6yMB3nf+N+K73XTxlAD4uYxqbfn3HMF2oC6Gnqw1VPlihnhLljMBAJuT8b+AGPqWJ4B1WvhMrzv/wcqLwNnYbnaljgqMwO72T8WWyMVqi1ZGArzv/G/F94jpWlAC4KedCUF9S0dMNinEcDOSBytfzRFswXInAGDLvJ4MiKN/h/OqIT4rxp3/UOVB4HvYLpQjK7WOn/HYqZin4tferZShRgK87/w7svMHJQCeqg6FnZoj2A71BNW+o2/lCLZgJSQAAOtT/Ro2UIcz0IFDqTv//uVZ4NfAO7GldSmsCXwWWwc/N0Kd2vlO+icB3p3/HcDyQa1VY0oA/FRdpqQEIJ12lib1LSfkCLZgpSQAABsCTwfE07fcz8vXpOfu/PuXSdieArGsDPwR2/wmd117S98kwLvzvwtYIbzZ0ui0jSTqZkHF13X6TmEpjav4unmuUYine7D19R5bBr8am8uzItb5Xwxs7fC+XlYCfoltauU9T2AD4CZg3wjvHWJtbKL0tvjt8AdwL3YwlNdW09EpAShb1b2nq+xPL9VUXdurBKBsd2GT5zz2+t8AG/ourfPv66PAJx3fbwlsMvLKju/paW3gWvw6//uwzn+y0/sloQSgbHMqvq7q+QHSni6qjwB4HCwicd0B7AxMcXivTSi38+/1dWApp/c6HFjX6b1K9wA2YjQpdyDtUgJQtmkVX1e1U5L2LEH1GdXarbEebsOSgKm5A0lgKWAvp/f6kNP7lO4h7M7/ydyBVKEEoGxV7zxWd41CBhMy09fjrlLSuAXYhc5I2rZ1eI+J2GqKpvsf1vk/kTuQqpQAlK3q88c10WlzKbSz41t/nXBH2SQ3YUnA9Eyf/0Kiz/F4Zr8i8Sf9LQTOxeYuvBj5swbyKLAD8FiGz3ajBKBsj1d83Sh02lwKIXc5Vb9byecG7I6v6qO5qu7Ffms7E/+cD499AWLuLTAT+BE2v2AfbMnzrsCsiJ/Z3+PY7+DRhJ8ZhRKAsj0S8NpWtiKVMFVHAHqo+Z1DB7sZeAfWEaVwO/Bm7BnzpcDrgf2A/yb6/FLcCRyBrds/gpdfG3sPQUoxEvAEduf/cILPik4JQNkeCnjtRm5RyGA2qPi6ycBsz0AkqWux8+lj33XezCvXlXdj+/tvA2wHnEVzf0t3Acdi17JNsTv/wRKvK7HELGYS8CT2fYRclxtPOwH6GYFdZKq057kZ4u0ko6h+DsDlGeItXUk7AbbqTVT/+xyuXAcs3WIcE4CDsQ2H5gd+7kXtN8MrvCbg86dj52RUvYHZAZsv4f19PEVnTGwMpgTA1/VUa89nKGv3rabZhuoXkx9liLd0dUwAIE6HcxXV1+MvAxyArU2vYwJwq8Pne38nT1N9tK9oegRQvlsqvm55lLHGtEPAa+9yi0JyuwJ4G36z9P+NPV6YUfH1U4Ez8OlI66r3O/F4RPMstiPkvQ7vVRwlAOX7V8Br3+QWhfQXkgBc6xaFlOAq4N1U37mz18VYx5VqgmGTXUX4PI1nsR3+7nSJqEBKAMp3RcBrvXb1kpdbGjtXvYrnaPAFpYNdAuxJ9STgQuBdNHdCXw4hScDzwNux7aAbSwlA+SZRffjprdimHOJrH6qfuHgl9lxRmuefVBsJOB9L1kNHEOSVqiQBz2N7LtwUJaKCKAGoh6qzxkcB7/EMRAD4QMBrQx7pSPkuwtbpt3ra4x+wpEGHQ8VzFfBOWksCpmDP/G+MGlEhlADUQ8hjgPe5RSFgOyxuH/D6K53iSGEN4AfAPdj6as+Z7v1Lk/bZ/xt21znUeQ8LgW9iyeSCFEF1uCuwfRNuH+L/cxmwJR1w59+rpP3iD8WytFi6sUkddwHnYNt61sVl2B3C2AqvfQO2LGeoH7607mCqJ86Tqc8KgAOAU9DR0lVdhq3C+Tzwfl46oOtZLEE4Ebg7T2gd63Zgc2yuxbuw3VLnYCc+ngtcky+0PEpKAFZZVGLbDfuj/DOWdDyX4DNDTcOeE+5d4bVdWH0/6BpRZ1oK+HjA68+mHs//PwT8Cu0jEWoK8MVFZQJ2dHTph0Atiw2Bh1jbI5BIFgJ/WlQkgqobAeUo95Em6fCwJ9XrOR8dDuThS4T93rZOH3LbNiD+cL93ybkRUKnOIf/3UqV08v4FyXX6HID1sLuykbkDacGFVB+tGAV8xjGWTrQE8KmA1z+E7epYuuPQsL9IR+j0BABsYsiHcgfRgvnYjOGqPgKs5RRLJ/oUsELA68/E7nBKtgI20iQiHUAJgAm5s0vp9IDXLo5NPJL2rQkcFfgeZ3oEEtmulDUvSEQiipEA1HE96ybYjNDS3UjYOvJ3YRd5ac9J2COAqi4H7neKJaaquxuKSA3FSACeivCeKeyYO4AWfSvw9SdRbTlhp9qd8GHxb3sEkkBdz46IeQa8pFX6Y7JGiZEA1HWns5DDXVL6J2G7VK1PfTqk3JYHTg18jxuASx1iiW1FbFJsHT2SO4AC1bUjnZQ7AAmzPjZhLfdyknbLJOqz7nkvwuraTdxNl5pgBLata+jv6t2pA6/og+T/G6xaXhehPeruJ+T/XqqUk2I0hqT1Q/L/kKqU18dojAhGYDvKhdT1GeqzB0IOXyb893Q39Zlo+3vy//1VKf+O0RgN8DHyfzdVyh4xGkPSGo3tXJf7x9Ru+UaMxohkd3wunlVPtWuynfEZxXpv6sArGontWpf776/dMgvYNEJ7NMFq2IFEub+jdsozaA+KxhgJfBV4gfw/rFbLbVFaIp6/El7n86jHRkipbAnMILxdr6I+j5TeRP6/vXbLTOy8dhncT8n/PbVTPhunGWQwKS5QvZuLbAOsBIyJ/HnjgG0DXr8O8LBTLLGtgQ0zjwt8n1OwIcNOtw52IMiKge+zANiC+hzAdDJhZxykdhnwSXSYznDGA1cDr80dSAuuwEbeFuYOpJPU5Q6lHWOwE7eWqvj6r2DHdNbFUdj2raGOoV6PQLytjq1g8dgt8YfApx3eJ4WRwJNUS3p6gF8s+s/Y5gGPAhcDdyT4vKZYFvgj8NbcgQzhQuxx2czcgUgz/JHqw1B1u6sYix1q5DEE912amRQOZ2PgcXzacBJ28ltd7ET1uurglnroAt6DJU9zyD/U34PNsbkc2JfOvOYUoanbfp6H/bCq2BAbMqvLfIC52BKuqwh/vPJ5YDns3IBOGYrbHjuffaLT+x0OTHd6rxT2D3jtZW5RSEw92Hn352Kd7YqE7WwZag424W9BxhikwZYmbAbsCelDDvZZ/LLz86j+CKVO9sP36NufpQ0/2FLYTPqq9dUkPBEp0qVUv7A9R/22y+3COm6vzuwR6nF+fRVjsQ1HPIc076B+S5gOpXp9Z6IlpCJSqMMIu6Dvlz7kYMtgE6W8OrU5wBFJaxDfesAt+Hb+s7BHR3VzPdXrfHaGeEVEWrIMMJvqF7i6Pt98I/4TfS4A1k5ZiQhGY3McQoa8BysHJqyHl23ovARZRDpI6Pamdd1jfF9sEp9nJ/cicCz1HPZ9C3An/h1/D+GnM+YSslJmCvX8HYhIBwlZ4tQD/DZ9yG6OIE6Hdx+2ZrcOuwe+BjiHOO3QA/yGei5hWoOwrY5/lD5kEZH2jMB29at6oZuHXSzr6njidX4PY0lGiZMlN8eeUXcTr/4XUN9ltCcQVve6joyJSIc5mrCL3Q/Sh+ymCziDeJ1gDzbp8BjyzxFYCvg/bDe/mPXtAa4jfPvlXJbFZvBXrftN6UMWEalmFWyznKoXvJnYhMK6GgmcSvxOsRvbd/xQ7MyHFMYB78SeZ3uu5x+qXEu9fw/fJKz+n0gfsohIdb8k7KJ3VPqQXXVhk9VSdJC95S7gx8Be2IFQHpbE9jQ/Dks2Uh93ej55d1ALNRHbpbBq/Wdgm2yJSAPUcQJTFRtiHVLV+k4GXoUtr6uzzwLfI8/3PhW4H5tEeD+2Feh0bIRlFrZkcwx2gtnSWGe/DPBqbO3+ethoTi6nAwdT7+1Lvwl8OeD1J6IjW0WkhkJ3yftM+pCj+BBhj0Q6sRxP/ZPllQnbA6HuE2JFpINtTVgnMJV6P/vtawvgIfJ3rKWXmYQdllOS0Hkgp6cPWUTEz9WEXQS/kz7kaCZgp4Pl7mRLLXdhxwQ3wQaErfvvxvZUEBGprdCNgV4EVksedTxd2Fr+1JPpSi9nUO/Jfv39g7D2+Ef6kEVE/IWcEtgD/Cp9yNFthf8BOXUsTwJ7B7ZlafYkvF3emDxqEZEItiRsh7iFwGbJo45vFDYaMIP8HXHqMh87Hnip4FYsyxhsxUVI21ycPGoRkYj+RNhF8fz0ISezCvF3Dyyp/AvYxKXlyvNlwttn2+RRi4hEtBG2njvkwvim5FGntQtwDfk76FjlDuzUxLov7xvMBMKPPtazfxFpnC7gBsIujn9LHnUe2xM+b6KkcjNwAHZQVJMdSHhbbZ08ahGRiLqAnxF+cZxNfU+Dq+ItwD+xORC5O/F2SzdwCbCjd6MU7GTC2uxpmjcnQkQ6mFfn31vWTBt+EVYDjiR8clmKcg9wLLBujIYo3OmEt9+1KAkQkQbw7vx7yH8Ebk5d2OOBX2D7+ufu7HvL49ie9VvEq3otfAef9lQSICK1FqPznw+MTVmJwq0NHAKcDTxPug7/6UWfeQTW6Td1Ul+7dsGvjZUEiEgtxej8e7DlYzKwMdjowBHYUczXA1MIT7j+B1yEnWj4IWyLWxnYSOA2lASIyCCafrfUBfwU+GiE994fOCvC+zbZ0tiJcisDywHLAosv+t8mYpsQLcQ6+6lY0jAVeALbpa/OR/HmsAXwb/y2Nf4P8DbsexIRKVasO/8e4Aqav4xMmmE3YA4aCRCRDtFF+DKowcrdwErpqiIS7G3YslWvv4EbsREbEZGiqPMXeSUlASLSaOr8RQanJEBEGkmdv8jwlASISKOo8xdpnZIAEWkEdf4i7VMSICK1ps5fpDolASJSS+r8RcIpCRCRWlHnL+JnV5QEiEgNqPMX8ackQESKFnN739uB5dNVRaQ42jZYRIr1ZdT5i8QUIwkYl7QGItI4m2KnwanzF4nLOwn4TdLoRaRxzkKdv0gq3knANmnDF5GmGAlM5/+1d7ehclR3HMe/xmiiMd4kelON5oUFpbUNsVKrlCqKIEXwCaqtgfoEra+aFkpB+6KFShWVioqIgoLUQmir8QGNMZaiGNq0VsXnIjaPYmpTjY0xqTfeXF8cr27W3Xtnd8+cszPz/cB5ldzd/8zsnd//zs45Y/hLKcVsAu5NXLukmliE4S/lEKsJ2Jq6cEn1sBjDX8phBHiJOL97sxPXLqkGDgB2EeckdB8wM235UiWNAOuI83u3FzgwbfmS6uIh4l0F+D02AdJUYob/BLAlbfmS6uRbxDsZ2QRI3cUO/wng7qRbIKl2Yi8BbBMg7auM8N8LLE25EZLqZybwMDYBqp5DGf4V8coI/wng1pQbIam+ZgGP0Jwm4HDgUuBa4NfAFcDCrBWpqDMJn61tfPZZexO4E/hyxro6KSv8Hyf8zkpSFLOAVdS7CZgD/IbOc7DHgJuBedmq01S+AvyJqT9ve4CrchXYpszwd+qfpOhmA49RzyZgFHiG6evdCizLVKM+bz5wCyHci37mbsxS6WcMf0mVNJtwoqlTEzBK7wuvPAEcl6NYAWF++3L2vdTfy8jVBBj+kiptFvAocU9gfyRPEzBKWKmwn5rHCN8tu8phOvsBFwJvMPhn7qbEtY8Af41Qd6fwPyjhdkhquIOY/jvXXkfqKwH9/OXfabwD/ARXXSvTDOAi+m/Wuo1UVwL8y19SrRwM/JlqNgGxwr91bAR+SFhGWXHMBL4PvEb88EzVBBj+kmppDvAk1WoCygj/1rEB+AFOxRrEPOBnwCbKO04pmgDDX1KtzQGeohpNQNnh3zq2Aj8HFpSwHXW1lLD65PukOUZlNgGGv6RGOAR4muFuAlKGf+vYCdyOS7N2MwJcCfyd9MemrCbA8JfUKHOBtQxnE5Ar/NvHOsLKgnMjbFOVzQG+BzwA7Cb/cYnZBBj+khppLvAXhqsJGJbwbx27gD8AF9Cck/rRhJskHwA+IP8xKKMJMPwlNdqhxJ/v3G8TMIzh3z7+R1gH4TLq9cyBIwhz9m8l/vS9YWwCDH9JopyTYa9NQBXCv32MA/8gPJPgXMJSt1UyAlxN9fZ7t1G0CTD8JanFPOLf2FW0Cahi+Hca45/swyqsL3Au8B/y77PYY7omwPCXpA7mUewhO72M6ZqAuoR/+1hHuKw+jC4jNCu591FZ44Yu2234S9IU5gPPEvcE2a0JiBn+HwKXEE7+uyLX3+94kbAC4zA5kbCvcu+bybGTcq5EtDcBhr8kFbAAeI64J8r2JiB2+J/X8tpHER7808tjZ8sauR9n2241+ffJ5DG7jXCV5EuEhZhiv8dkE2D4S1IPDgOeJ+4Jc7IJKDP8Wx0LrAD2Rt6OXsZOwkyLYTBK/kv/48BvgWPaajseeLuE97sZw1+SenY48AJxT5z3kSb8Wy0hhM5Y5G0pOs4pUGMKZ5Jn+ycIwb8S+OoU9S0BtmWssegw/CU1wijDOS+8aPi3Wkx4rvyOxLX+uMc6y3Ih6Y/T/4G7CJf5i1gK/DdDnUWH4S+pUYbtTv1+wr/VfMKDf7YkqvfqAWqN6SLSHaP3gOuBRX3UeSLwbsJaDX9JmsJC4BXyn4QHDf9W+wNnE76WKPPO+CY1AC8Ayxn8voevA9sT1Gv4S1IBXwBepR7h324h8FPKaXLq3gC8B9wBnBS53pMJSy8b/pI0BI4AXqNe4d/ueOAqwt3jMWYQ1LEB2A08Slh7ocy1Dr5J+ns2DH9J6uJI4J/UM/zbLQKuZLB7IOrSALwL3At8BzgkYd2nEqZTGv6SNAQWAa9T7/BvdS3NbgB+wWCPeR7U6aR9LLHhLzXcjNwFDLG3gNMIXweUZQ/wXeChEt9DxWwAPsr4/k8C3yY0AWVbA5xPmMIoqaFsAKb2b+As4F8lvPYewl+sD5bw2qqmp4ELKDeYJ8N/d4nvIakCbACm9yZwBnGbAMNf3TxB+EqojCbA8Jf0KRuAYrYQmoD1EV7L8Nd01hCuBHwY+TUNf0mfsgEoLkYTYPirqNXEawIMf0mfYwPQm82EJmBDHz9r+KtXjwHLCJ+dfhn+kjqyAejdZsKUrY09/Izhr36tBC6mvxkKhr+krmwA+rOZ8OjZlwv83+2Ek7Dhr37dT1iRcLyHn1lDuJnQ8JfUkQ1A/9YD3wB+CbzT4d/HgHuAE4BV6cpSTa0gXEXaUeD//o7yZhJIqomcK5/VwW7gV4RV9E4BvggcQFhEaC3wfr7SVEMrgWcJj3pexr7LFU8AfwOuAx5OX5qkqrEBiOMjQuCvzV2Iam8T4dkNPwKWEp5bsYOwYuXbGeuSVDE2AFI1jQHP5C5CUnV5D4AkSQ1kAyBJUgPZAEiS1EA2AJIkNZANgCRJDeQsAMVwMfC13EUAi3MXIElVYQOgGJZ8MiRJFeFXAJIkNZANgCRJDWQDIElSA9kASJLUQDYAkiQ1kA2AJo3nLkCSlI4NgCZty11AZh/kLkCSUrIB0KTncheQ2Ru5C5AkKYf9gU3ARAPHBmC/wXehJEnVdAn5wzjHuDzGzpMkqcruIn8gpxwr8K9/SZKYAVwDjJE/nMsc48At+DwMSZL2cRxwE/ASsJP8gR1j7AJeB+4AToi3qySpej4G5sfrOqGsHvEAAAAASUVORK5CYII="


def get_app_dir() -> Path:
    """EXE 실행 시에는 EXE가 있는 폴더, Python 실행 시에는 main.py 폴더를 반환합니다."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def bring_existing_app_to_front() -> None:
    """?? ?? ?? JIIN DNC Manager ?? ?? ??? ????."""
    if not sys.platform.startswith("win"):
        return
    try:
        user32 = ctypes.windll.user32
        hwnd = user32.FindWindowW(None, APP_TITLE)
        if hwnd:
            user32.ShowWindow(hwnd, 9)  # SW_RESTORE
            user32.SetForegroundWindow(hwnd)
    except Exception:
        pass


def acquire_single_instance_lock() -> bool:
    """????? ?? ?? ??? ? ?? ??? ?? ?? ?? ??? ????."""
    global SINGLE_INSTANCE_HANDLE
    if not sys.platform.startswith("win"):
        return True
    kernel32 = ctypes.windll.kernel32
    handle = kernel32.CreateMutexW(None, False, SINGLE_INSTANCE_MUTEX_NAME)
    if not handle:
        return True
    if kernel32.GetLastError() == ERROR_ALREADY_EXISTS:
        kernel32.CloseHandle(handle)
        bring_existing_app_to_front()
        return False
    SINGLE_INSTANCE_HANDLE = handle
    return True


APP_DIR = get_app_dir()
DATA_DIR = APP_DIR / "data"
LOG_DIR = DATA_DIR / "logs"
BACKUP_DIR = DATA_DIR / "backup"
EXPORT_DIR = DATA_DIR / "export"
AUTO_BACKUP_DIR = DATA_DIR / "auto_backup"
CONFIG_FILE = DATA_DIR / "config.json"
CONFIG_BACKUP_FILE = DATA_DIR / "config.json.bak"
LOGO_FILE = DATA_DIR / "company_logo.png"
BUNDLED_LOGO_FILE = Path(getattr(sys, "_MEIPASS", APP_DIR)) / "company_logo.png"
KCC_LOGO_FILE = DATA_DIR / "korea_circuit_logo.png"
BUNDLED_KCC_LOGO_FILE = Path(getattr(sys, "_MEIPASS", APP_DIR)) / "korea_circuit_logo.png"
TLB_LOGO_FILE = DATA_DIR / "tlb_logo.png"
BUNDLED_TLB_LOGO_FILE = Path(getattr(sys, "_MEIPASS", APP_DIR)) / "tlb_logo.png"
LEGACY_CONFIG_FILE = APP_DIR / "config.json"
PROCESS_DATA_FOLDER_NAMES = {
    "KCC PKG": "KCC_PKG",
    "TLB": "TLB",
    "KCC HDI": "KCC_HDI",
}
KCC_PKG_DATA_DIR = DATA_DIR / "KCC_PKG"
KCC_PKG_DB_FILE = KCC_PKG_DATA_DIR / "work_log.db"
TLB_DATA_DIR = DATA_DIR / "TLB"
TLB_DB_FILE = TLB_DATA_DIR / "work_log.db"
KCC_HDI_DATA_DIR = DATA_DIR / "KCC_HDI"
KCC_HDI_DB_FILE = KCC_HDI_DATA_DIR / "work_log.db"
CONDITION_MASTER_DB_FILE = KCC_PKG_DATA_DIR / "condition_master.db"
LEGACY_KCC_PKG_DB_FILE = DATA_DIR / "KCC_PKG.db"
LEGACY_CONDITION_MASTER_FILE = APP_DIR / "condition_master.json"
MIGRATION_BACKUP_DONE = False

APP_BG = "#f3f6fb"
SURFACE_BG = "#ffffff"
BORDER_COLOR = "#d8e2f0"
TEXT_COLOR = "#172033"
MUTED_TEXT = "#687386"
PRIMARY = "#0f5bff"
PRIMARY_LIGHT = "#eaf2ff"
OK_COLOR = "#0f9f63"
NG_COLOR = "#dc2626"
TAB_BG = "#eaf0f8"
# 사용자 지정 팔레트 번호 기준
# TLB=2, 심텍 SPS=1, 심텍 HDI=3, KCC PKG=7, KCC HDI=4
PROCESS_COLORS = {
    "TLB": {"bg": "#f4f5f6", "light": "#e8ecef", "primary": "#475569", "border": "#cbd5e1"},
    "심텍 SPS": {"bg": "#f3f6fb", "light": "#eaf2ff", "primary": "#0f5bff", "border": "#b8cdfa"},
    "심텍 HDI": {"bg": "#f1f4f8", "light": "#e3e9f1", "primary": "#26364a", "border": "#b8c4d3"},
    "KCC PKG": {"bg": "#f3f5fb", "light": "#e6ecfb", "primary": "#3b5bdb", "border": "#b8c7f2"},
    "KCC HDI": {"bg": "#f8f3f5", "light": "#f0e4e9", "primary": "#7f1d36", "border": "#d7b7c1"},
}

THEMES = {
    "MES 블루": {
        "app_bg": "#f3f6fb",
        "surface_bg": "#ffffff",
        "border_color": "#d8e2f0",
        "text_color": "#172033",
        "muted_text": "#687386",
        "primary": "#0f5bff",
        "primary_light": "#eaf2ff",
        "tab_bg": "#eaf0f8",
    },
    "KCC 민트": {
        "app_bg": "#eaf5f2",
        "surface_bg": "#fbfffe",
        "border_color": "#b7d8d2",
        "text_color": "#102a2a",
        "muted_text": "#55716f",
        "primary": "#00897b",
        "primary_light": "#dff6f1",
        "tab_bg": "#f8fffd",
    },
    "심텍 그린": {
        "app_bg": "#f0f8f1",
        "surface_bg": "#ffffff",
        "border_color": "#c8dfc9",
        "text_color": "#172417",
        "muted_text": "#617064",
        "primary": "#198754",
        "primary_light": "#e6f4ec",
        "tab_bg": "#eef6ef",
    },
    "TLB 퍼플": {
        "app_bg": "#f6f3fb",
        "surface_bg": "#ffffff",
        "border_color": "#d9cfea",
        "text_color": "#221933",
        "muted_text": "#6b617a",
        "primary": "#6f42c1",
        "primary_light": "#f0e9fb",
        "tab_bg": "#f1ecf8",
    },
    "다크 네이비": {
        "app_bg": "#eef2f6",
        "surface_bg": "#ffffff",
        "border_color": "#c9d5e2",
        "text_color": "#14213d",
        "muted_text": "#5c677d",
        "primary": "#1f3a5f",
        "primary_light": "#e7edf5",
        "tab_bg": "#e9eef5",
    },
}


def apply_theme(theme_name: str) -> None:
    """설정된 화면 색상 테마를 전역 UI 색상에 적용합니다."""
    theme = THEMES.get(theme_name, THEMES["MES 블루"])
    globals_to_update = {
        "APP_BG": theme["app_bg"],
        "SURFACE_BG": theme["surface_bg"],
        "BORDER_COLOR": theme["border_color"],
        "TEXT_COLOR": theme["text_color"],
        "MUTED_TEXT": theme["muted_text"],
        "PRIMARY": theme["primary"],
        "PRIMARY_LIGHT": theme["primary_light"],
        "TAB_BG": theme["tab_bg"],
    }
    globals().update(globals_to_update)


# ==================================================
# 로그/백업/마이그레이션 안전장치
# ==================================================
def write_log_file(filename: str, message: str) -> None:
    """현장 문제 추적용 로그를 data/logs 폴더에 남깁니다."""
    try:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_path = LOG_DIR / filename
        with log_path.open("a", encoding="utf-8") as file:
            file.write(f"[{now_text}] {message}\n")
    except Exception:
        # 로그 실패 때문에 현장 작업을 막지는 않습니다.
        pass


def log_app(message: str) -> None:
    write_log_file(f"app_{datetime.now().strftime('%Y%m%d')}.log", message)


def log_error(message: str, exc: Exception | None = None) -> None:
    detail = message
    if exc is not None:
        detail += f"\n{type(exc).__name__}: {exc}\n{traceback.format_exc()}"
    write_log_file(f"error_{datetime.now().strftime('%Y%m%d')}.log", detail)


def create_migration_backup_once() -> Path:
    """DB 마이그레이션 전에 data 주요 파일을 1회 백업합니다."""
    global MIGRATION_BACKUP_DONE
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUP_DIR / f"{timestamp}_before_migration"
    if MIGRATION_BACKUP_DONE:
        return backup_path

    try:
        backup_path.mkdir(parents=True, exist_ok=False)
        if CONFIG_FILE.exists():
            shutil.copy2(CONFIG_FILE, backup_path / "config.json")
        if LEGACY_CONFIG_FILE.exists() and not (backup_path / "config_legacy.json").exists():
            shutil.copy2(LEGACY_CONFIG_FILE, backup_path / "config_legacy.json")

        kcc_backup = backup_path / "KCC_PKG"
        kcc_backup.mkdir(parents=True, exist_ok=True)
        for source, target_name in (
            (KCC_PKG_DB_FILE, "work_log.db"),
            (CONDITION_MASTER_DB_FILE, "condition_master.db"),
            (LEGACY_KCC_PKG_DB_FILE, "legacy_KCC_PKG.db"),
            (LEGACY_CONDITION_MASTER_FILE, "legacy_condition_master.json"),
        ):
            if source.exists():
                shutil.copy2(source, kcc_backup / target_name)
        MIGRATION_BACKUP_DONE = True
        log_app(f"마이그레이션 백업 생성: {backup_path}")
        return backup_path
    except Exception as exc:
        log_error("마이그레이션 백업 실패", exc)
        raise RuntimeError("DB 마이그레이션 전 백업에 실패했습니다.\n기존 data 보호를 위해 실행을 중단합니다.") from exc


def table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone()
    return row is not None


def get_existing_columns(conn: sqlite3.Connection, table_name: str) -> set[str]:
    if not table_exists(conn, table_name):
        return set()
    return {row[1] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()}


def ensure_schema_version_table(conn: sqlite3.Connection, db_name: str, version: int) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS schema_version (
            db_name TEXT PRIMARY KEY,
            version INTEGER NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        INSERT INTO schema_version (db_name, version, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(db_name) DO UPDATE SET
            version=excluded.version,
            updated_at=excluded.updated_at
        """,
        (db_name, version, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    )


def migrate_table_columns(conn: sqlite3.Connection, table_name: str, expected_columns: dict[str, str]) -> list[str]:
    """없는 컬럼만 ALTER TABLE로 추가해서 중복 추가 오류를 막습니다."""
    added_columns: list[str] = []
    existing_columns = get_existing_columns(conn, table_name)
    for column_name, column_sql in expected_columns.items():
        if column_name in existing_columns:
            continue
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_sql}")
        added_columns.append(column_name)
    return added_columns


# ==================================================
# 공통 파일/설정 함수
# ==================================================
def get_desktop_path() -> Path:
    """현재 Windows 사용자의 바탕화면 경로를 반환합니다."""
    return Path.home() / "Desktop"


def get_default_config() -> dict:
    """config.json이 없거나 값이 비어 있을 때 사용할 기본값입니다."""
    source_folders = {
        process_name: ""
        for process_name in PROCESS_NAMES
    }
    return {
        "config_version": 1,
        "excel_file": "",
        "source_dnc_folder": "",
        "source_dnc_folders": source_folders,
        "transfer_dnc_folder": "",
        "dnc_delete_seconds": DNC_DELETE_SECONDS,
        "first_article_wait_seconds": FIRST_ARTICLE_WAIT_SECONDS,
        "auto_export_after_dnc": True,
        "auto_shift_group_enabled": True,
        "a_group_day_start_date": "2026-05-17",
        "master_password": MASTER_SETTINGS_PASSWORD,
        "condition_master_password": CONDITION_MASTER_PASSWORD,
        "license_password": LICENSE_PASSWORD,
        "license_master_pc_name": DEFAULT_MASTER_PC_NAME,
        "license_allowed_ip_prefixes": DEFAULT_ALLOWED_IP_PREFIXES[:],
        "clear_common_after_normal": False,
        "machine": "트리밍 1호기",
        "theme": "MES 블루",
        "last_master_auto_backup_date": "",
        "tlb_condition_sheet": "",
        "kcc_hdi_condition_sheet": "",
    }


def normalize_saved_source_folders(config: dict) -> dict[str, str]:
    """공정별 DNC 조건 시트 폴더 설정을 보정합니다.

    예전 V2에서는 아직 만들지 않은 공정에도 바탕화면 기본 경로를 자동으로
    넣었습니다. 현장에서는 지정하지 않은 공정이 빈칸이어야 하므로,
    그 자동 기본값만 빈칸으로 되돌립니다.
    """
    desktop = get_desktop_path()
    old_generated_defaults = {
        "심텍 SPS": str(desktop / "SIMMTECH_SPS"),
        "심텍 HDI": str(desktop / "SIMMTECH_HDI"),
        "KCC HDI": str(desktop / "KCC_HDI"),
    }
    raw_sources = config.get("source_dnc_folders")
    if not isinstance(raw_sources, dict):
        raw_sources = {}

    normalized: dict[str, str] = {}
    for process_name in PROCESS_NAMES:
        value = str(raw_sources.get(process_name, "") or "").strip()
        old_default = old_generated_defaults.get(process_name, "")
        if old_default and Path(value.replace("\\", "/")).as_posix().lower() == Path(old_default).as_posix().lower():
            value = ""
        normalized[process_name] = value

    legacy_kcc = str(config.get("source_dnc_folder", "") or "").strip()
    if legacy_kcc and not normalized.get("KCC PKG"):
        normalized["KCC PKG"] = legacy_kcc

    return normalized


def normalize_machine_name(machine: str) -> str:
    """화면 표시용 호기명을 작업일보 기록용 짧은 이름으로 바꿉니다."""
    text = str(machine or "").strip()
    for number in ("1", "2", "3"):
        if number in text:
            return f"{number}호기"
    return text


def get_machine_axis_count(machine: str) -> int:
    """설비 호기별 실제 사용 가능한 축 수를 반환합니다."""
    normalized = normalize_machine_name(machine)
    if normalized == "3호기":
        return 4
    return 6


def get_machine_allowed_axes(machine: str) -> list[int]:
    """설비 호기 기준으로 팝업에서 클릭 가능한 축 index 목록을 반환합니다."""
    return list(range(get_machine_axis_count(machine)))


def get_work_period(now: datetime | None = None) -> dict[str, str]:
    """08:30 기준 작업일자와 08:30/20:30 기준 근무를 계산합니다."""
    now = now or datetime.now()
    day_start = now.replace(hour=8, minute=30, second=0, microsecond=0)
    night_start = now.replace(hour=20, minute=30, second=0, microsecond=0)
    if now < day_start:
        work_date = (now - timedelta(days=1)).strftime("%Y-%m-%d")
        shift = "야간"
    elif now < night_start:
        work_date = now.strftime("%Y-%m-%d")
        shift = "주간"
    else:
        work_date = now.strftime("%Y-%m-%d")
        shift = "야간"
    return {
        "work_date": work_date,
        "shift": shift,
        "period_key": f"{work_date}_{shift}",
    }


def get_auto_shift_group(work_date_text: str, shift: str, a_group_day_start_date: str) -> str:
    """A조 주간 시작일을 기준으로 4근 2휴 조 패턴에서 현재 근무 조를 계산합니다."""
    try:
        work_date = datetime.strptime(work_date_text, "%Y-%m-%d").date()
        base_date = datetime.strptime(a_group_day_start_date, "%Y-%m-%d").date()
    except ValueError:
        return ""
    day_offset = (work_date - base_date).days
    # 각 조는 12일 주기로 주간4/휴2/야간4/휴2 패턴입니다.
    # B조, C조는 A조 기준에서 각각 4일, 8일 밀린 패턴으로 계산합니다.
    group_offsets = {"A": 0, "B": 4, "C": 8}
    for group, offset in group_offsets.items():
        cycle_day = (day_offset + offset) % 12
        if shift == "주간" and 0 <= cycle_day <= 3:
            return group
        if shift == "야간" and 6 <= cycle_day <= 9:
            return group
    return ""


def load_config() -> dict:
    """config.json을 읽고, 없는 값은 기본값으로 채웁니다."""
    config = get_default_config()
    changed = False
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        if not CONFIG_FILE.exists() and LEGACY_CONFIG_FILE.exists():
            shutil.copy2(LEGACY_CONFIG_FILE, CONFIG_FILE)
            log_app("기존 config.json을 data/config.json으로 복사")
        if CONFIG_FILE.exists():
            try:
                saved = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            except Exception as exc:
                log_error("config.json 로드 실패 - 백업 복구 시도", exc)
                if CONFIG_BACKUP_FILE.exists():
                    saved = json.loads(CONFIG_BACKUP_FILE.read_text(encoding="utf-8"))
                    shutil.copy2(CONFIG_BACKUP_FILE, CONFIG_FILE)
                    log_app("config.json.bak에서 설정 복구")
                    changed = True
                else:
                    raise
            if isinstance(saved, dict):
                config.update({key: value for key, value in saved.items() if value not in (None, "")})
                for key, value in get_default_config().items():
                    if key not in saved:
                        config[key] = value
                        changed = True
        else:
            changed = True
    except Exception:
        # 설정 파일이 손상되어도 프로그램은 기본값으로 실행되게 합니다.
        log_error("config.json 로드 실패 - 기본값 사용")
        pass
    saved_sources = normalize_saved_source_folders(config)
    if saved_sources != config.get("source_dnc_folders"):
        changed = True
    config["source_dnc_folders"] = saved_sources
    config["source_dnc_folder"] = saved_sources.get("KCC PKG", "")
    tlb_sheet = str(config.get("tlb_condition_sheet", "") or "").strip()
    kcc_hdi_sheet = str(config.get("kcc_hdi_condition_sheet", "") or "").strip()
    if tlb_sheet and not kcc_hdi_sheet:
        config["kcc_hdi_condition_sheet"] = tlb_sheet
        changed = True
    elif kcc_hdi_sheet and not tlb_sheet:
        config["tlb_condition_sheet"] = kcc_hdi_sheet
        changed = True
    if changed:
        try:
            save_config(config)
            log_app("config.json 누락 항목 자동 보정")
        except Exception as exc:
            log_error("config.json 자동 보정 저장 실패", exc)
    return config


def save_config(config: dict) -> None:
    """현재 설정을 config.json에 저장합니다."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if CONFIG_FILE.exists():
        try:
            shutil.copy2(CONFIG_FILE, CONFIG_BACKUP_FILE)
        except Exception as exc:
            log_error("config.json.bak 저장 실패", exc)
    CONFIG_FILE.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    try:
        shutil.copy2(CONFIG_FILE, CONFIG_BACKUP_FILE)
    except Exception as exc:
        log_error("config.json.bak 갱신 실패", exc)


def get_current_pc_name() -> str:
    """Windows 장치 이름을 가져옵니다. 라이선스 마스터 PC 판단에 사용합니다."""
    return socket.gethostname().strip().upper()


def get_current_ip_addresses() -> list[str]:
    """현재 PC의 IPv4 주소 목록을 빠르게 가져옵니다. 외부 인터넷 접속은 하지 않습니다."""
    addresses: list[str] = []
    try:
        host_name = socket.gethostname()
        for address in socket.gethostbyname_ex(host_name)[2]:
            if address and "." in address and not address.startswith("127."):
                addresses.append(address)
    except Exception as exc:
        log_error("라이선스 IP 조회 실패", exc)
    return list(dict.fromkeys(addresses))


def normalize_ip_prefixes(value) -> list[str]:
    """config에 저장된 IP 대역을 리스트로 정리합니다."""
    if isinstance(value, list):
        raw_items = value
    else:
        raw_items = str(value or "").replace(";", "\n").replace(",", "\n").splitlines()
    prefixes: list[str] = []
    for item in raw_items:
        prefix = str(item or "").strip()
        if prefix:
            prefixes.append(prefix)
    return list(dict.fromkeys(prefixes))


def is_license_allowed(config: dict) -> bool:
    """프로그램 시작 시 PC 이름/IP 기준으로 실행 권한을 확인합니다."""
    pc_name = get_current_pc_name()
    master_pc_name = str(config.get("license_master_pc_name", DEFAULT_MASTER_PC_NAME)).strip().upper()
    if master_pc_name and pc_name == master_pc_name:
        log_app(f"라이선스 허용: 마스터 PC {pc_name}")
        return True

    allowed_prefixes = normalize_ip_prefixes(config.get("license_allowed_ip_prefixes", DEFAULT_ALLOWED_IP_PREFIXES))
    if not allowed_prefixes:
        log_app(f"라이선스 허용: IP 제한 없음, PC={pc_name}")
        return True

    current_ips = get_current_ip_addresses()
    for address in current_ips:
        if any(address.startswith(prefix) for prefix in allowed_prefixes):
            log_app(f"라이선스 허용: PC={pc_name}, IP={address}")
            return True

    log_app(f"라이선스 불일치: PC={pc_name}, IP={current_ips}, 허용대역={allowed_prefixes}")
    return False


def show_license_block_message() -> None:
    """Tk 창이 뜨기 전에도 표시할 수 있는 라이선스 차단 메시지입니다."""
    message = "라이선스가 불일치 합니다.\n관리자에게 문의 바랍니다."
    if sys.platform.startswith("win"):
        ctypes.windll.user32.MessageBoxW(None, message, "라이선스 확인", 0x10)
    else:
        print(message)


def select_excel_file(parent, config: dict, excel_var: tk.StringVar) -> None:
    """설정 탭에서 작업일보 Excel 파일을 선택합니다."""
    file_path = filedialog.askopenfilename(
        parent=parent,
        title="작업일보 Excel 파일 선택",
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    if file_path:
        config["excel_file"] = file_path
        excel_var.set(file_path)
        save_config(config)


def ensure_excel_file_selected(parent, config: dict, excel_var: tk.StringVar | None = None) -> bool:
    """작업일보 파일이 없을 때만 사용자에게 한 번 선택을 요청합니다."""
    current_path = ""
    if excel_var is not None:
        current_path = excel_var.get().strip()
    if not current_path:
        current_path = config.get("excel_file", "").strip()
    if not current_path and CONFIG_FILE.exists():
        try:
            saved = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            current_path = str(saved.get("excel_file", "")).strip()
        except Exception:
            current_path = ""

    if current_path and Path(current_path).exists():
        config["excel_file"] = current_path
        if excel_var is not None:
            excel_var.set(current_path)
        save_config(config)
        return True

    show_operator_alert(parent, "작업일보 확인", "작업일보 파일 선택 필요")
    file_path = filedialog.askopenfilename(
        parent=parent,
        title="작업일보 Excel 파일 선택",
        initialdir=str(Path(current_path).parent) if current_path else str(get_desktop_path()),
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    if not file_path:
        return False

    config["excel_file"] = file_path
    if excel_var is not None:
        excel_var.set(file_path)
    save_config(config)
    return True


def open_path(path: Path) -> None:
    """파일 또는 폴더를 Windows 기본 프로그램으로 엽니다."""
    if sys.platform.startswith("win"):
        os.startfile(str(path))
    else:
        subprocess.Popen(["xdg-open", str(path)])


def open_log_excel(config: dict) -> None:
    """선택된 작업일보 Excel 파일을 엽니다."""
    if not ensure_excel_file_selected(None, config):
        return
    excel_file = Path(config.get("excel_file", ""))
    try:
        open_path(excel_file)
    except Exception as exc:
        show_operator_alert(None, "작업일보 열기 실패", str(exc), "error")


# ==================================================
# DNC 파일 처리 함수
# ==================================================
def delete_existing_dnc_txt(transfer_folder: Path) -> None:
    """DNC 전송 폴더에 남아 있는 txt 파일을 실행 전에 모두 삭제합니다."""
    if not transfer_folder.exists() or not transfer_folder.is_dir():
        raise FileNotFoundError(f"DNC 전송 폴더 확인 필요: {transfer_folder}")
    for txt_file in transfer_folder.glob("*.txt"):
        txt_file.unlink()


def search_condition_file(condition_name: str, source_folder: Path) -> list[Path]:
    """원본 DNC 폴더와 하위 폴더에서 .txt 조건 파일명 완전일치만 찾습니다."""
    return search_condition_file_exact_txt(condition_name, source_folder)


def copy_dnc_file(source_file: Path, transfer_folder: Path) -> Path:
    """조건 txt 파일을 DNC 전송 폴더로 복사합니다."""
    if not transfer_folder.exists() or not transfer_folder.is_dir():
        raise FileNotFoundError(f"DNC 전송 폴더 확인 필요: {transfer_folder}")
    copied_file = transfer_folder / source_file.name
    shutil.copy2(source_file, copied_file)
    return copied_file


def delete_after_delay(copied_file: Path, seconds: int, status_callback=None) -> None:
    """지정된 초만큼 기다린 후 복사된 DNC txt 파일을 삭제합니다."""
    for remaining in range(seconds, 0, -1):
        if status_callback:
            status_callback(f"DNC 삭제 대기중 ({remaining})")
        time.sleep(1)
    for attempt in range(1, 31):
        if not copied_file.exists():
            break
        try:
            copied_file.unlink()
            log_app(f"DNC 복사 파일 삭제 완료: {copied_file}")
            break
        except PermissionError as exc:
            log_error(f"DNC 복사 파일 사용중 - 삭제 재시도 {attempt}/30: {copied_file}", exc)
            if status_callback:
                status_callback(f"DNC 파일 사용중 - 삭제 재시도 ({attempt})")
            time.sleep(1)
        except OSError as exc:
            log_error(f"DNC 복사 파일 삭제 실패 - 재시도 {attempt}/30: {copied_file}", exc)
            if status_callback:
                status_callback(f"DNC 삭제 재시도 ({attempt})")
            time.sleep(1)
    if copied_file.exists():
        log_error(f"DNC 복사 파일 삭제 실패: {copied_file}")
        if status_callback:
            status_callback("DNC 삭제 실패")
        return
    if status_callback:
        status_callback("DNC 파일 삭제 완료")


# ==================================================
# 검증 함수
# ==================================================
def check_mes_core(lot_no: str, process_code: str) -> bool:
    """LOT No 9번째 문자부터 4자리와 공정코드 끝 2자리로 MES Core 일치화를 판정합니다."""
    lot_no = lot_no.strip()
    process_code = process_code.strip().upper()
    if len(lot_no) < 12 or len(process_code) < 2:
        return False

    lot_core = lot_no[8:12]
    process_tail = process_code[-2:]
    return (lot_core == "0000" and process_tail == "T1") or (lot_core == "0205" and process_tail == "11")


def get_mes_core_message(lot_no: str, process_code: str) -> tuple[bool, str]:
    """MES Core 판정 결과를 작업자용 짧은 문구로 반환합니다."""
    lot_no = lot_no.strip()
    process_code = process_code.strip().upper()
    if len(lot_no) < 12:
        return False, "LOT No 확인 필요"
    if len(process_code) < 2:
        return False, "공정코드 확인 필요"

    lot_core = lot_no[8:12]
    process_tail = process_code[-2:]
    if lot_core == "0000" and process_tail == "T1":
        return True, "OK"
    if lot_core == "0205" and process_tail == "11":
        return True, "OK"
    log_app(f"MES Core NG: LOT={lot_no}, LOT중간4자리={lot_core}, 공정코드={process_code}, 끝2자리={process_tail}")
    return False, "MES Core NG"


def check_condition_ok(lot1: dict, lot2: dict | None = None) -> bool:
    """작업조건 KCC_ 시작 여부와 지그/2LOT 조건 일치 여부를 확인합니다."""
    if not lot1.get("condition", "").strip().upper().startswith("KCC_"):
        return False
    if not lot1.get("jig", "").strip():
        return False
    if lot2:
        return (
            lot1.get("condition", "").strip() == lot2.get("condition", "").strip()
            and lot1.get("jig", "").strip() == lot2.get("jig", "").strip()
        )
    return True


def get_single_condition_message(lot: dict) -> tuple[bool, str]:
    """LOT 한 개의 작업조건/지그 기본 조건을 작업자용 짧은 문구로 확인합니다."""
    condition = lot.get("condition", "").strip()
    jig = lot.get("jig", "").strip()
    if not condition:
        return False, "작업조건 없음"
    if not condition.upper().startswith("KCC_"):
        return False, "작업조건 확인 필요"
    if not jig:
        return False, "지그 없음"
    return True, "OK"


def get_lot_match_message(lot1: dict, lot2: dict) -> tuple[bool, str]:
    """2LOT 작업 시 LOT 1/LOT 2의 작업조건과 지그 일치 여부를 작업자용으로 간단히 확인합니다."""
    lot1_no = lot1.get("lot_no", "").strip()
    lot2_no = lot2.get("lot_no", "").strip()
    lot1_condition = lot1.get("condition", "").strip()
    lot2_condition = lot2.get("condition", "").strip()
    lot1_jig = lot1.get("jig", "").strip()
    lot2_jig = lot2.get("jig", "").strip()
    messages = []
    if lot1_no and lot2_no and lot1_no == lot2_no:
        messages.append("LOT No가 같습니다")
    if lot1_condition != lot2_condition:
        if not lot1_condition or not lot2_condition:
            messages.append("작업조건 미조회")
        else:
            messages.append("작업조건 다름")
    if lot1_jig != lot2_jig:
        if not lot1_jig or not lot2_jig:
            messages.append("지그 미조회")
        else:
            messages.append("지그 다름")
    if messages:
        return False, "확인 필요 - " + ", ".join(dict.fromkeys(messages))
    if not lot1_condition or not lot1_jig:
        return False, "확인 필요 - 작업조건/지그 미조회"
    return True, "OK - 2LOT 조건 일치"


def validate_positive_number(value: str, field_name: str, required: bool = True) -> tuple[bool, str]:
    """매수/Stack처럼 0보다 큰 숫자만 허용하는 필드 검증입니다."""
    text = value.strip()
    if not text:
        return (not required, "" if not required else f"{field_name} 입력 필요")
    if not text.isdigit():
        return False, f"{field_name} 숫자만 입력"
    if int(text) <= 0:
        return False, f"{field_name} 0 초과 필요"
    return True, ""


def validate_zero_or_positive_number(value: str, field_name: str) -> tuple[bool, str]:
    """신규 모델 검증 매수처럼 0 이상 숫자를 반드시 입력해야 하는 필드 검증입니다."""
    text = value.strip()
    if not text:
        return True, ""
    if not text.isdigit():
        return False, f"{field_name} 숫자만 입력"
    return True, ""


def validate_lot_required(lot: dict, lot_name: str, require_qty: bool) -> list[str]:
    """LOT 입력 필수값을 확인합니다."""
    errors = []
    required_fields = [
        ("step", "STEP"),
        ("round", "차수"),
        ("manage_no", "관리번호"),
        ("lot_no", "LOT No"),
        ("process_code", "공정코드"),
        ("condition", "작업조건"),
        ("jig", "지그"),
    ]
    if require_qty:
        required_fields.insert(4, ("qty", "매수"))

    for key, label in required_fields:
        if not lot.get(key, "").strip():
            errors.append(f"{lot_name} {label} 입력 필요")

    if require_qty:
        ok, message = validate_positive_number(lot.get("qty", ""), f"{lot_name} 매수", required=True)
        if not ok:
            errors.append(message)
    elif lot.get("qty", "").strip():
        ok, message = validate_positive_number(lot.get("qty", ""), f"{lot_name} 매수", required=False)
        if not ok:
            errors.append(message)

    if lot.get("condition", "").strip() and not lot.get("condition", "").strip().upper().startswith("KCC_"):
        errors.append(f"{lot_name} 작업조건 확인 필요")
    return errors


def missing_common_message(label: str) -> str:
    """현장 알람용 공통 입력 누락 문구입니다."""
    return f"{label} 입력 없음"


def validate_normal_dnc(common: dict, lot1: dict, lot2: dict | None) -> tuple[bool, list[str]]:
    """일반 DNC 입력값 전체 검증입니다."""
    errors = []
    for key, label in (("work_date", "작업일자"), ("machine", "트리밍 호기"), ("shift_group", "조"), ("shift", "근무"), ("worker", "작업자")):
        if not common.get(key, "").strip():
            errors.append(missing_common_message(label))

    errors.extend(validate_lot_required(lot1, "LOT 1", require_qty=True))
    if lot2:
        errors.extend(validate_lot_required(lot2, "LOT 2", require_qty=True))

    lot1_mes_ok, lot1_mes_message = get_mes_core_message(lot1.get("lot_no", ""), lot1.get("process_code", ""))
    if not lot1_mes_ok:
        errors.append(f"LOT 1 {lot1_mes_message}")
    lot1_condition_ok, lot1_condition_message = get_single_condition_message(lot1)
    if not lot1_condition_ok:
        errors.append(f"LOT 1 {lot1_condition_message}")

    if lot2:
        lot2_mes_ok, lot2_mes_message = get_mes_core_message(lot2.get("lot_no", ""), lot2.get("process_code", ""))
        if not lot2_mes_ok:
            errors.append(f"LOT 2 {lot2_mes_message}")
        lot2_condition_ok, lot2_condition_message = get_single_condition_message(lot2)
        if not lot2_condition_ok:
            errors.append(f"LOT 2 {lot2_condition_message}")
        lot_match_ok, lot_match_message = get_lot_match_message(lot1, lot2)
        if not lot_match_ok:
            errors.append(f"2LOT {lot_match_message}")
    return len(errors) == 0, errors


def validate_new_model_dnc(common: dict, lot: dict) -> tuple[bool, list[str]]:
    """신규 모델 검증 DNC 입력값 전체 검증입니다."""
    errors = []
    for key, label in (("work_date", "작업일자"), ("machine", "트리밍 호기"), ("shift_group", "조"), ("shift", "근무"), ("worker", "작업자")):
        if not common.get(key, "").strip():
            errors.append(missing_common_message(label))

    errors.extend(validate_lot_required(lot, "신규 모델", require_qty=False))
    ok, message = validate_zero_or_positive_number(lot.get("qty", ""), "신규 모델 매수")
    if not ok:
        errors.append(message)
    mes_ok, mes_message = get_mes_core_message(lot.get("lot_no", ""), lot.get("process_code", ""))
    if not mes_ok:
        errors.append(mes_message)
    condition_ok, condition_message = get_single_condition_message(lot)
    if not condition_ok:
        errors.append(condition_message)
    return len(errors) == 0, errors


def format_operator_errors(errors: list[str]) -> str:
    """작업자 팝업에는 중복 없는 핵심 오류만 짧게 보여줍니다."""
    clean_errors = []
    for error in errors:
        text = str(error).strip()
        text = text.replace("을(를) 입력해야 DNC를 진행할 수 있습니다.", " 입력 없음")
        text = text.replace("을(를) 입력하세요.", " 입력 없음")
        text = text.replace("입력해야 DNC를 진행할 수 있습니다.", "입력 없음")
        if text and text not in clean_errors:
            clean_errors.append(text)
    if not clean_errors:
        return "입력값을 확인해주세요."
    if len(clean_errors) <= 8:
        return "\n".join(clean_errors)
    visible = clean_errors[:8]
    visible.append(f"외 {len(clean_errors) - 8}건 추가 확인 필요")
    return "\n".join(visible)


def format_excel_error_for_operator(exc: Exception, context: str = "Excel 파일") -> str:
    """openpyxl 원문 오류 대신 작업자가 이해할 수 있는 한국어 안내를 보여줍니다."""
    message = str(exc).strip()
    lowered = message.lower()
    if "openpyxl does not support" in lowered or "supported formats are" in lowered:
        return (
            f"{context} 형식 확인 필요\n"
            "지원 형식: .xlsx / .xlsm\n"
            "Excel에서 파일을 열어 다른 이름으로 저장 후 다시 선택하세요."
        )
    if isinstance(exc, PermissionError):
        return f"{context} 열림\nExcel 파일을 닫고 다시 시도하세요."
    if message:
        return message
    return f"{context} 확인 필요"


def keep_modal_on_top(dialog: tk.Toplevel, parent=None, focus_widget=None) -> None:
    """모달창이 다른 Windows 창 뒤로 숨어 프로그램이 멈춘 것처럼 보이는 상황을 막습니다."""
    target = focus_widget or dialog

    def raise_dialog(count: int = 0) -> None:
        try:
            if not dialog.winfo_exists():
                return
            dialog.deiconify()
            dialog.attributes("-topmost", True)
            dialog.lift()
            target.focus_force()
            if count < 12:
                dialog.after(500, lambda: raise_dialog(count + 1))
        except tk.TclError:
            return

    if parent is not None:
        try:
            dialog.transient(parent)
        except tk.TclError:
            pass
    dialog.bind("<Visibility>", lambda _event: raise_dialog(0))
    raise_dialog()


def show_operator_alert(parent, title: str, message: str, kind: str = "warning") -> None:
    """작업자용 큰 글씨 알람입니다. 핵심 문장만 크게 보여줍니다."""
    dialog = tk.Toplevel(parent) if parent else tk.Toplevel()
    # Keep the dialog hidden until its final centered geometry is ready.
    dialog.withdraw()
    dialog.title(title)
    dialog.configure(bg=SURFACE_BG)
    dialog.resizable(False, False)
    dialog.transient(parent if parent else None)

    accent = NG_COLOR if kind in {"warning", "error"} else OK_COLOR
    icon = "⚠" if kind in {"warning", "error"} else "✓"
    lines = format_operator_errors(message.splitlines()).splitlines()
    icon_message = "\n".join(f"{icon}  {line}" for line in lines if line.strip())
    body = tk.Frame(dialog, bg=SURFACE_BG, padx=28, pady=22)
    body.pack(fill=tk.BOTH, expand=True)
    tk.Label(
        body,
        text=title,
        bg=SURFACE_BG,
        fg=accent,
        font=("맑은 고딕", 16, "bold"),
        anchor="center",
        justify=tk.CENTER,
    ).pack(fill=tk.X, pady=(0, 14))
    tk.Label(
        body,
        text=icon_message,
        bg=SURFACE_BG,
        fg=TEXT_COLOR,
        font=("맑은 고딕", 13, "bold"),
        justify=tk.CENTER,
        anchor="center",
        wraplength=520,
    ).pack(fill=tk.X)

    footer = tk.Frame(dialog, bg=APP_BG, padx=18, pady=14)
    footer.pack(fill=tk.X)
    ttk.Button(footer, text="확인", command=dialog.destroy, style="Primary.TButton", width=14).pack(anchor="center")

    dialog.update_idletasks()
    width = max(460, dialog.winfo_reqwidth())
    height = max(220, dialog.winfo_reqheight())
    if parent:
        x = parent.winfo_rootx() + max((parent.winfo_width() - width) // 2, 0)
        y = parent.winfo_rooty() + max((parent.winfo_height() - height) // 2, 0)
    else:
        x = (dialog.winfo_screenwidth() - width) // 2
        y = (dialog.winfo_screenheight() - height) // 2
    dialog.geometry(f"{width}x{height}+{x}+{y}")
    dialog.deiconify()
    dialog.grab_set()
    keep_modal_on_top(dialog, parent)
    dialog.wait_window()


def center_dialog(dialog: tk.Toplevel, parent=None, min_width: int = 460, min_height: int = 220) -> None:
    dialog.update_idletasks()
    width = max(min_width, dialog.winfo_reqwidth())
    height = max(min_height, dialog.winfo_reqheight())
    if parent:
        x = parent.winfo_rootx() + max((parent.winfo_width() - width) // 2, 0)
        y = parent.winfo_rooty() + max((parent.winfo_height() - height) // 2, 0)
    else:
        x = (dialog.winfo_screenwidth() - width) // 2
        y = (dialog.winfo_screenheight() - height) // 2
    dialog.geometry(f"{width}x{height}+{x}+{y}")


def ask_system_input(parent, title: str, prompt: str, show: str | None = None, numeric_only: bool = False, initial: str = "") -> str | None:
    """기본 simpledialog 대신 사용하는 시스템 스타일 입력창입니다."""
    dialog = tk.Toplevel(parent) if parent else tk.Toplevel()
    dialog.title(title)
    dialog.resizable(False, False)
    dialog.configure(bg=SURFACE_BG)
    dialog.transient(parent if parent else None)
    result = {"value": None}
    value_var = tk.StringVar(value=initial)

    body = tk.Frame(dialog, bg=SURFACE_BG, padx=30, pady=24)
    body.pack(fill=tk.BOTH, expand=True)
    tk.Label(body, text=title, bg=SURFACE_BG, fg=PRIMARY, font=("맑은 고딕", 16, "bold"), anchor="center").pack(fill=tk.X, pady=(0, 14))
    tk.Label(body, text=prompt, bg=SURFACE_BG, fg=TEXT_COLOR, font=("맑은 고딕", 12, "bold"), anchor="center", justify=tk.CENTER).pack(fill=tk.X, pady=(0, 12))
    vcmd = (dialog.register(lambda text: text.isdigit() or text == ""), "%P") if numeric_only else None
    entry = ttk.Entry(
        body,
        textvariable=value_var,
        style="Wide.TEntry",
        show=show or "",
        width=34,
        validate="key" if numeric_only else "none",
        validatecommand=vcmd,
        font=("맑은 고딕", 12),
    )
    entry.pack(fill=tk.X)

    footer = tk.Frame(dialog, bg=APP_BG, padx=18, pady=14)
    footer.pack(fill=tk.X)

    def confirm() -> None:
        # Korean IME may still be composing the last syllable when the button is
        # clicked. Move focus once and read shortly after so the final character
        # is committed before we capture the value.
        dialog.focus_set()
        dialog.after(80, finalize_confirm)

    def finalize_confirm() -> None:
        result["value"] = value_var.get().strip()
        dialog.destroy()

    ttk.Button(footer, text="확인", command=confirm, style="Primary.TButton", width=14).pack(side=tk.LEFT, expand=True, padx=(60, 8))
    ttk.Button(footer, text="취소", command=dialog.destroy, width=14).pack(side=tk.LEFT, expand=True, padx=(8, 60))
    dialog.bind("<Return>", lambda _event: confirm())
    dialog.bind("<Escape>", lambda _event: dialog.destroy())
    dialog.protocol("WM_DELETE_WINDOW", dialog.destroy)
    center_dialog(dialog, parent, 500, 250)
    dialog.grab_set()
    keep_modal_on_top(dialog, parent, entry)
    dialog.wait_window()
    return result["value"]


def ask_system_yes_no(parent, title: str, message: str) -> bool:
    """기본 askyesno 대신 사용하는 시스템 스타일 확인창입니다."""
    dialog = tk.Toplevel(parent) if parent else tk.Toplevel()
    dialog.title(title)
    dialog.resizable(False, False)
    dialog.configure(bg=SURFACE_BG)
    dialog.transient(parent if parent else None)
    result = {"ok": False}

    body = tk.Frame(dialog, bg=SURFACE_BG, padx=30, pady=24)
    body.pack(fill=tk.BOTH, expand=True)
    tk.Label(body, text=title, bg=SURFACE_BG, fg=PRIMARY, font=("맑은 고딕", 16, "bold"), anchor="center").pack(fill=tk.X, pady=(0, 14))
    tk.Label(body, text=f"⚙  {message}", bg=SURFACE_BG, fg=TEXT_COLOR, font=("맑은 고딕", 13, "bold"), justify=tk.CENTER, anchor="center", wraplength=520).pack(fill=tk.X)

    footer = tk.Frame(dialog, bg=APP_BG, padx=18, pady=14)
    footer.pack(fill=tk.X)

    def choose(value: bool) -> None:
        result["ok"] = value
        dialog.destroy()

    button_wrap = tk.Frame(footer, bg=APP_BG)
    button_wrap.pack(anchor="center")
    for column, (text, value) in enumerate((("예", True), ("아니오", False))):
        button = tk.Button(
            button_wrap,
            text=text,
            command=lambda selected=value: choose(selected),
            width=16,
            height=2,
            bg=PRIMARY_LIGHT,
            fg=PRIMARY,
            activebackground=PRIMARY_LIGHT,
            activeforeground=PRIMARY,
            relief=tk.SOLID,
            bd=1,
            highlightthickness=1,
            highlightbackground=BORDER_COLOR,
            cursor="hand2",
            font=("맑은 고딕", 11, "bold"),
            takefocus=0,
        )
        button.grid(row=0, column=column, padx=8, sticky="nsew")
    dialog.bind("<Return>", lambda _event: choose(True))
    dialog.bind("<Escape>", lambda _event: choose(False))
    dialog.protocol("WM_DELETE_WINDOW", lambda: choose(False))
    center_dialog(dialog, parent, 500, 250)
    dialog.grab_set()
    keep_modal_on_top(dialog, parent)
    dialog.wait_window()
    return bool(result["ok"])


def ask_incomplete_action(parent, count: int) -> str:
    """시작 시 미완료 이력을 작업자가 바로 처리할 수 있게 묻습니다."""
    dialog = tk.Toplevel(parent) if parent else tk.Toplevel()
    dialog.title("미완료 이력 확인")
    dialog.resizable(False, False)
    dialog.configure(bg=SURFACE_BG)
    dialog.transient(parent if parent else None)
    result = {"action": "later"}

    body = tk.Frame(dialog, bg=SURFACE_BG, padx=30, pady=24)
    body.pack(fill=tk.BOTH, expand=True)
    tk.Label(body, text="미완료 이력 있음", bg=SURFACE_BG, fg=NG_COLOR, font=("맑은 고딕", 16, "bold"), anchor="center").pack(fill=tk.X, pady=(0, 14))
    tk.Label(
        body,
        text=f"완료 처리되지 않은 이력 {count}건\n처리 방법을 선택하세요.",
        bg=SURFACE_BG,
        fg=TEXT_COLOR,
        font=("맑은 고딕", 13, "bold"),
        justify=tk.CENTER,
        anchor="center",
        wraplength=520,
    ).pack(fill=tk.X)

    footer = tk.Frame(dialog, bg=APP_BG, padx=18, pady=14)
    footer.pack(fill=tk.X)

    def choose(action: str) -> None:
        result["action"] = action
        dialog.destroy()

    ttk.Button(footer, text="삭제", command=lambda: choose("delete"), width=14).pack(side=tk.LEFT, expand=True, padx=(20, 6))
    ttk.Button(footer, text="완료처리", command=lambda: choose("complete"), style="Primary.TButton", width=14).pack(side=tk.LEFT, expand=True, padx=6)
    ttk.Button(footer, text="나중에", command=lambda: choose("later"), width=14).pack(side=tk.LEFT, expand=True, padx=(6, 20))
    dialog.bind("<Escape>", lambda _event: choose("later"))
    dialog.protocol("WM_DELETE_WINDOW", lambda: choose("later"))
    center_dialog(dialog, parent, 520, 260)
    dialog.grab_set()
    keep_modal_on_top(dialog, parent)
    dialog.wait_window()
    return result["action"]


# ==================================================
# Excel 저장 함수
# ==================================================
def get_log_header_row(ws) -> int:
    """작업일보 로그 헤더 행을 찾습니다.

    KCC PKG는 상단 DNC 입력 영역 때문에 로그 헤더가 아래쪽에 있고,
    TLB는 시트 상단에 로그 헤더가 있습니다. 그래서 고정 행 대신
    작업일자/차수/관리번호/LOT/매수/작업P/G 헤더가 있는 행을 찾습니다.
    """
    required_headers = {"작업일자", "차수", "관리번호", "LOT", "매수"}
    preferred_headers = {"작업P/G", "실적"}
    best_row = 8
    best_score = 0
    for row in range(1, min(ws.max_row, 30) + 1):
        row_values = {
            str(ws.cell(row=row, column=col).value or "").strip()
            for col in range(1, 20)
        }
        score = len(required_headers & row_values) * 10 + len(preferred_headers & row_values)
        if score > best_score:
            best_row = row
            best_score = score
    return best_row if best_score >= 40 else 8


def get_next_empty_row(ws) -> int:
    """작업일보 로그 헤더 바로 아래부터 첫 빈 행을 찾습니다."""
    row = get_log_header_row(ws) + 1
    while True:
        has_value = any(ws.cell(row=row, column=col).value not in (None, "") for col in range(1, 18))
        if not has_value:
            return row
        row += 1


def write_process_code_backup(ws, row: int, process_code: str) -> None:
    """조건 마스터 복구용으로 작업일보 AD열에 공정코드를 백업합니다."""
    # A열에 호기가 추가되어 기존 A:AC 양식이 한 칸씩 밀렸습니다.
    # AD열만 프로그램 복구용 공정코드 백업 칸으로 사용합니다.
    if not ws.cell(row=6, column=EXCEL_PROCESS_CODE_COLUMN).value:
        ws.cell(row=6, column=EXCEL_PROCESS_CODE_COLUMN).value = "공정코드"
    ws.cell(row=row, column=EXCEL_PROCESS_CODE_COLUMN).value = excel_upper_value(process_code)


def write_export_id_backup(ws, row: int, log_id: int) -> None:
    """Excel 저장 직후 강제 종료되어도 중복 반영을 막기 위한 DB ID 백업 칸입니다."""
    header_row = get_log_header_row(ws)
    if not ws.cell(row=header_row, column=EXCEL_EXPORT_ID_COLUMN).value:
        ws.cell(row=header_row, column=EXCEL_EXPORT_ID_COLUMN).value = "DNC_LOG_ID"
    ws.cell(row=row, column=EXCEL_EXPORT_ID_COLUMN).value = int(log_id)


def get_excel_exported_log_ids(ws) -> set[int]:
    """작업일보에 이미 기록된 DB ID를 읽어 중복 반영을 방지합니다."""
    existing_ids: set[int] = set()
    for row in range(get_log_header_row(ws) + 1, ws.max_row + 1):
        value = ws.cell(row=row, column=EXCEL_EXPORT_ID_COLUMN).value
        if value in (None, ""):
            continue
        try:
            existing_ids.add(int(value))
        except (TypeError, ValueError):
            continue
    return existing_ids


def excel_upper_value(value):
    """작업일보에서 작업 P/G를 제외한 문자 값은 보기 좋게 대문자로 저장합니다."""
    if isinstance(value, str):
        return value.upper()
    return value


def ensure_log_sheet_machine_column(ws) -> None:
    """작업일보에 A열 '호기'가 없으면 자동으로 한 칸 삽입합니다.

    기존 양식은 A열이 작업일자였으므로 새 EXE가 바로 쓰면 날짜를 덮어쓸 수 있습니다.
    A열에 호기가 없고 작업일자 양식으로 보이면 A열을 추가해 기존 데이터를 오른쪽으로 보존합니다.
    """
    first_header = str(ws.cell(row=6, column=1).value or "").strip()
    top_header = str(ws.cell(row=2, column=1).value or "").strip()
    if first_header == "호기" or top_header == "호기":
        return
    if first_header == "작업일자" or top_header == "작업일자":
        ws.insert_cols(1)
        ws.cell(row=6, column=1).value = "호기"
        ws.cell(row=2, column=1).value = "호기"


def open_log_workbook(config: dict, sheet_name: str = LOG_SHEET_NAME):
    """작업일보 파일과 지정 시트를 열고 기본 오류를 메시지로 변환합니다."""
    excel_file = config.get("excel_file", "")
    if not excel_file:
        raise FileNotFoundError("작업일보 파일 선택 필요")
    path = Path(excel_file)
    if not path.exists():
        raise FileNotFoundError("작업일보 파일 없음")

    try:
        if not zipfile.is_zipfile(path):
            raise ValueError("작업일보 파일 오류")
        workbook = load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm")
    except PermissionError:
        raise PermissionError("작업일보 파일 열림")
    except zipfile.BadZipFile:
        raise ValueError("작업일보 파일 오류")

    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise KeyError(f"{sheet_name} 시트 없음")
    ws = workbook[sheet_name]
    ensure_log_sheet_machine_column(ws)
    return workbook, ws, path


def get_excel_lock_path(path: Path) -> Path:
    """작업일보와 같은 폴더에 PC 간 Excel 반영 잠금 파일 경로를 만듭니다."""
    return path.with_name(f"{path.name}.lock")


def acquire_excel_export_lock(path: Path) -> Path:
    """여러 PC가 같은 작업일보를 동시에 저장하지 못하게 잠금 파일을 생성합니다."""
    lock_path = get_excel_lock_path(path)
    now = datetime.now()
    if lock_path.exists():
        try:
            lock_age = now - datetime.fromtimestamp(lock_path.stat().st_mtime)
        except OSError:
            lock_age = timedelta(seconds=0)
        if lock_age.total_seconds() < EXCEL_LOCK_STALE_SECONDS:
            raise PermissionError("다른 PC 작업일보 반영 중")
        try:
            lock_path.unlink()
            log_app(f"오래된 작업일보 lock 자동 삭제: {lock_path}")
        except OSError as exc:
            log_error("오래된 작업일보 lock 삭제 실패", exc)
            raise PermissionError("작업일보 반영 대기") from exc

    lock_text = {
        "created_at": now.strftime("%Y-%m-%d %H:%M:%S"),
        "pc_name": socket.gethostname(),
        "pid": os.getpid(),
        "excel_file": str(path),
    }
    try:
        fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        with os.fdopen(fd, "w", encoding="utf-8") as file:
            json.dump(lock_text, file, ensure_ascii=False, indent=2)
    except FileExistsError as exc:
        raise PermissionError("다른 PC 작업일보 반영 중") from exc
    except OSError as exc:
        log_error("작업일보 lock 생성 실패", exc)
        raise PermissionError("작업일보 반영 대기") from exc
    return lock_path


def release_excel_export_lock(lock_path: Path | None) -> None:
    """작업일보 반영 성공/실패와 관계없이 잠금 파일을 정리합니다."""
    if lock_path is None:
        return
    try:
        if lock_path.exists():
            lock_path.unlink()
    except OSError as exc:
        log_error("작업일보 lock 삭제 실패", exc)


def save_workbook_safely(workbook, path: Path) -> None:
    """Excel 저장 실패 시 사용자가 이해할 수 있는 메시지를 발생시킵니다."""
    while True:
        try:
            workbook.save(path)
            return
        except PermissionError:
            retry = ask_excel_save_retry()
            if not retry:
                raise PermissionError("작업일보 저장 보류")


def ask_excel_save_retry() -> bool:
    """작업일보가 열려 있을 때 현장 작업자가 이해하기 쉬운 문구로 재시도 여부를 묻습니다."""
    dialog = tk.Toplevel()
    dialog.title("작업일보 저장 대기")
    dialog.resizable(False, False)
    dialog.configure(bg=SURFACE_BG)
    result = {"retry": False}

    body = tk.Frame(dialog, bg=SURFACE_BG, padx=22, pady=18)
    body.pack(fill=tk.BOTH, expand=True)
    tk.Label(
        body,
        text="작업일보 파일 열림",
        bg=SURFACE_BG,
        fg=TEXT_COLOR,
        font=("맑은 고딕", 10, "bold"),
        anchor="w",
        justify=tk.LEFT,
    ).pack(fill=tk.X, pady=(0, 10))
    tk.Label(
        body,
        text=(
            "파일을 닫고 [재 시도]\n"
            "나중에 하려면 [다음에 저장]"
        ),
        bg=SURFACE_BG,
        fg=TEXT_COLOR,
        font=("맑은 고딕", 10),
        anchor="w",
        justify=tk.LEFT,
    ).pack(fill=tk.X)

    buttons = tk.Frame(dialog, bg=APP_BG, padx=14, pady=12)
    buttons.pack(fill=tk.X)

    def choose_retry() -> None:
        result["retry"] = True
        dialog.destroy()

    def choose_later() -> None:
        result["retry"] = False
        dialog.destroy()

    ttk.Button(buttons, text="재 시도", command=choose_retry, style="Primary.TButton", width=16).pack(side=tk.RIGHT, padx=(8, 0))
    ttk.Button(buttons, text="다음에 저장", command=choose_later, width=16).pack(side=tk.RIGHT)
    dialog.protocol("WM_DELETE_WINDOW", choose_later)
    dialog.update_idletasks()
    width = dialog.winfo_reqwidth()
    height = dialog.winfo_reqheight()
    screen_width = dialog.winfo_screenwidth()
    screen_height = dialog.winfo_screenheight()
    dialog.geometry(f"{width}x{height}+{(screen_width - width) // 2}+{(screen_height - height) // 2}")
    dialog.grab_set()
    keep_modal_on_top(dialog)
    dialog.wait_window()
    return result["retry"]


def ask_numeric_input(parent, title: str, prompt: str) -> str | None:
    """숫자만 입력 가능한 간단한 팝업입니다."""
    value = ask_system_input(parent, title, prompt, numeric_only=True)
    if value == "":
        show_operator_alert(parent, title, "숫자 입력 필요")
        return None
    return value


# ==================================================
# KCC PKG DB 저장 함수
# ==================================================
WORK_LOG_COLUMNS = {
    "customer_process": "TEXT NOT NULL DEFAULT 'KCC PKG'",
    "dnc_type": "TEXT NOT NULL DEFAULT '일반'",
    "status": "TEXT NOT NULL DEFAULT '완료'",
    "machine": "TEXT",
    "work_date": "TEXT",
    "shift_group": "TEXT",
    "shift_name": "TEXT",
    "worker": "TEXT",
    "step": "TEXT",
    "round_no": "TEXT",
    "manage_no": "TEXT",
    "lot_no": "TEXT",
    "qty_text": "TEXT",
    "qty_number": "INTEGER",
    "result_value": "REAL",
    "process_code": "TEXT",
    "condition_name": "TEXT",
    "jig": "TEXT",
    "stack": "TEXT",
    "model_change_text": "TEXT",
    "burr_result": "TEXT",
    "record_time": "TEXT",
    "first_axis_1": "TEXT",
    "first_axis_2": "TEXT",
    "first_axis_3": "TEXT",
    "first_axis_4": "TEXT",
    "first_axis_5": "TEXT",
    "first_axis_6": "TEXT",
    "jig_axis_1": "TEXT",
    "jig_axis_2": "TEXT",
    "jig_axis_3": "TEXT",
    "jig_axis_4": "TEXT",
    "jig_axis_5": "TEXT",
    "jig_axis_6": "TEXT",
    "exported": "INTEGER NOT NULL DEFAULT 0",
    "exported_at": "TEXT",
    "created_at": "TEXT",
}

CONDITION_MASTER_COLUMNS = {
    "step": "TEXT",
    "round_no": "TEXT",
    "manage_no": "TEXT",
    "process_code": "TEXT",
    "lot_no": "TEXT",
    "condition_name": "TEXT",
    "jig": "TEXT",
    "source": "TEXT",
    "updated_at": "TEXT",
}


def get_kcc_pkg_db_path() -> Path:
    """KCC PKG 원본 이력을 저장하는 SQLite DB 파일 경로를 반환합니다."""
    KCC_PKG_DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not KCC_PKG_DB_FILE.exists() and LEGACY_KCC_PKG_DB_FILE.exists():
        shutil.copy2(LEGACY_KCC_PKG_DB_FILE, KCC_PKG_DB_FILE)
    return KCC_PKG_DB_FILE


def get_kcc_pkg_connection() -> sqlite3.Connection:
    """KCC PKG DB 연결을 만들고 필요한 테이블을 자동 생성합니다."""
    db_path = get_kcc_pkg_db_path()
    existing_db = db_path.exists()
    conn = sqlite3.connect(db_path, timeout=10)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA busy_timeout=10000")
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        need_backup = False
        if existing_db:
            need_backup = not table_exists(conn, "schema_version")
            if table_exists(conn, "dnc_logs"):
                missing_columns = set(WORK_LOG_COLUMNS) - get_existing_columns(conn, "dnc_logs")
                need_backup = need_backup or bool(missing_columns)
        if need_backup:
            create_migration_backup_once()

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS dnc_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_process TEXT NOT NULL DEFAULT 'KCC PKG',
                dnc_type TEXT NOT NULL DEFAULT '일반',
                status TEXT NOT NULL DEFAULT '완료',
                machine TEXT,
                work_date TEXT,
                shift_group TEXT,
                shift_name TEXT,
                worker TEXT,
                step TEXT,
                round_no TEXT,
                manage_no TEXT,
                lot_no TEXT,
                qty_text TEXT,
                qty_number INTEGER,
                result_value REAL,
                process_code TEXT,
                condition_name TEXT,
                jig TEXT,
                stack TEXT,
                model_change_text TEXT,
                burr_result TEXT,
                record_time TEXT,
                first_axis_1 TEXT,
                first_axis_2 TEXT,
                first_axis_3 TEXT,
                first_axis_4 TEXT,
                first_axis_5 TEXT,
                first_axis_6 TEXT,
                jig_axis_1 TEXT,
                jig_axis_2 TEXT,
                jig_axis_3 TEXT,
                jig_axis_4 TEXT,
                jig_axis_5 TEXT,
                jig_axis_6 TEXT,
                exported INTEGER NOT NULL DEFAULT 0,
                exported_at TEXT,
                created_at TEXT
            )
            """
        )
        added = migrate_table_columns(conn, "dnc_logs", WORK_LOG_COLUMNS)
        ensure_schema_version_table(conn, "work_log", WORK_LOG_SCHEMA_VERSION)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dnc_logs_export ON dnc_logs(exported, status, id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dnc_logs_created ON dnc_logs(created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dnc_logs_lot ON dnc_logs(lot_no)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dnc_logs_work_date ON dnc_logs(work_date)")
        conn.commit()
        if added:
            log_app(f"work_log.db 컬럼 자동 추가: {', '.join(added)}")
    except Exception as exc:
        conn.rollback()
        conn.close()
        log_error("work_log.db 마이그레이션 실패", exc)
        raise
    return conn




def get_process_data_dir(process_name: str) -> Path:
    folder_name = PROCESS_DATA_FOLDER_NAMES.get(process_name, process_name.replace(" ", "_"))
    return DATA_DIR / folder_name


def get_process_db_file(process_name: str) -> Path:
    if process_name == "KCC PKG":
        return KCC_PKG_DB_FILE
    return get_process_data_dir(process_name) / "work_log.db"


def initialize_work_log_connection(db_path: Path, backup_on_migration: bool = False) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    existing_db = db_path.exists()
    conn = sqlite3.connect(db_path, timeout=10)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA busy_timeout=10000")
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        need_backup = False
        if existing_db:
            need_backup = not table_exists(conn, "schema_version")
            if table_exists(conn, "dnc_logs"):
                missing_columns = set(WORK_LOG_COLUMNS) - get_existing_columns(conn, "dnc_logs")
                need_backup = need_backup or bool(missing_columns)
        if backup_on_migration and need_backup:
            create_migration_backup_once()

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS dnc_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_process TEXT NOT NULL DEFAULT 'KCC PKG',
                dnc_type TEXT NOT NULL DEFAULT '\uc77c\ubc18',
                status TEXT NOT NULL DEFAULT '\uc644\ub8cc',
                machine TEXT,
                work_date TEXT,
                shift_group TEXT,
                shift_name TEXT,
                worker TEXT,
                step TEXT,
                round_no TEXT,
                manage_no TEXT,
                lot_no TEXT,
                qty_text TEXT,
                qty_number INTEGER,
                result_value REAL,
                process_code TEXT,
                condition_name TEXT,
                jig TEXT,
                stack TEXT,
                model_change_text TEXT,
                burr_result TEXT,
                record_time TEXT,
                first_axis_1 TEXT,
                first_axis_2 TEXT,
                first_axis_3 TEXT,
                first_axis_4 TEXT,
                first_axis_5 TEXT,
                first_axis_6 TEXT,
                jig_axis_1 TEXT,
                jig_axis_2 TEXT,
                jig_axis_3 TEXT,
                jig_axis_4 TEXT,
                jig_axis_5 TEXT,
                jig_axis_6 TEXT,
                exported INTEGER NOT NULL DEFAULT 0,
                exported_at TEXT,
                created_at TEXT
            )
            """
        )
        added = migrate_table_columns(conn, "dnc_logs", WORK_LOG_COLUMNS)
        ensure_schema_version_table(conn, "work_log", WORK_LOG_SCHEMA_VERSION)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dnc_logs_export ON dnc_logs(exported, status, id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dnc_logs_created ON dnc_logs(created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dnc_logs_lot ON dnc_logs(lot_no)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dnc_logs_work_date ON dnc_logs(work_date)")
        conn.commit()
        if added:
            log_app(f"{db_path.parent.name}/work_log.db columns added: {', '.join(added)}")
    except Exception as exc:
        conn.rollback()
        conn.close()
        log_error(f"{db_path} migration failed", exc)
        raise
    return conn


def migrate_process_rows_from_kcc_pkg(process_name: str, target_conn: sqlite3.Connection) -> None:
    if process_name == "KCC PKG" or not KCC_PKG_DB_FILE.exists():
        return
    try:
        target_count = target_conn.execute("SELECT COUNT(*) AS count FROM dnc_logs").fetchone()["count"]
        if int(target_count or 0) > 0:
            return
        source_conn = sqlite3.connect(KCC_PKG_DB_FILE, timeout=10)
        source_conn.row_factory = sqlite3.Row
        try:
            if not table_exists(source_conn, "dnc_logs"):
                return
            columns = ["id"] + list(WORK_LOG_COLUMNS)
            select_cols = ", ".join(columns)
            rows = source_conn.execute(
                f"SELECT {select_cols} FROM dnc_logs WHERE customer_process=? ORDER BY id",
                (process_name,),
            ).fetchall()
            if not rows:
                return
            placeholders = ", ".join("?" for _ in columns)
            target_conn.executemany(
                f"INSERT OR IGNORE INTO dnc_logs ({select_cols}) VALUES ({placeholders})",
                [[row[col] for col in columns] for row in rows],
            )
            target_conn.commit()
            log_app(f"{process_name} migrated rows copied: {len(rows)}")
        finally:
            source_conn.close()
    except Exception as exc:
        log_error(f"{process_name} migrated rows copy failed", exc)


def get_process_connection(process_name: str) -> sqlite3.Connection:
    if process_name == "KCC PKG":
        return get_kcc_pkg_connection()
    conn = initialize_work_log_connection(get_process_db_file(process_name), backup_on_migration=False)
    migrate_process_rows_from_kcc_pkg(process_name, conn)
    return conn

def get_condition_master_connection() -> sqlite3.Connection:
    """KCC PKG 조건 마스터 DB 연결을 만들고 필요한 테이블을 자동 생성합니다."""
    KCC_PKG_DATA_DIR.mkdir(parents=True, exist_ok=True)
    existing_db = CONDITION_MASTER_DB_FILE.exists()
    conn = sqlite3.connect(CONDITION_MASTER_DB_FILE, timeout=10)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA busy_timeout=10000")
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        need_backup = False
        if existing_db:
            need_backup = not table_exists(conn, "schema_version")
            if table_exists(conn, "condition_master"):
                missing_columns = set(CONDITION_MASTER_COLUMNS) - get_existing_columns(conn, "condition_master")
                need_backup = need_backup or bool(missing_columns)
        if need_backup:
            create_migration_backup_once()

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS condition_master (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                step TEXT,
                round_no TEXT,
                manage_no TEXT,
                process_code TEXT,
                lot_no TEXT,
                condition_name TEXT,
                jig TEXT,
                source TEXT,
                updated_at TEXT
            )
            """
        )
        added = migrate_table_columns(conn, "condition_master", CONDITION_MASTER_COLUMNS)
        ensure_schema_version_table(conn, "condition_master", CONDITION_MASTER_SCHEMA_VERSION)
        conn.execute("DROP INDEX IF EXISTS idx_condition_master_key")
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_condition_master_lookup
            ON condition_master(step, round_no, process_code, manage_no)
            """
        )
        conn.commit()
        if added:
            log_app(f"condition_master.db 컬럼 자동 추가: {', '.join(added)}")
    except Exception as exc:
        conn.rollback()
        conn.close()
        log_error("condition_master.db 마이그레이션 실패", exc)
        raise
    return conn


def calculate_result_value(qty_number: int) -> float | None:
    """매수 기준 실적을 계산합니다. 더미는 실적을 비워 둡니다."""
    if qty_number <= 0:
        return None
    return round(qty_number * 0.2, 1)


def insert_normal_dnc_db(common: dict, lots: list[dict], stack: str, model_change: bool) -> list[int]:
    """일반 DNC 이력을 Excel 대신 KCC_PKG.db에 먼저 저장합니다."""
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ids: list[int] = []
    conn = get_kcc_pkg_connection()
    try:
        for index, lot in enumerate(lots):
            qty_number = int(lot["qty"])
            cursor = conn.execute(
                """
                INSERT INTO dnc_logs (
                    dnc_type, status, machine, work_date, shift_group, shift_name, worker,
                    step, round_no, manage_no, lot_no, qty_text, qty_number, result_value,
                    process_code, condition_name, jig, stack, model_change_text, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "일반",
                    "DNC 진행",
                    normalize_machine_name(common["machine"]),
                    common["work_date"],
                    common["shift_group"],
                    common["shift"],
                    common["worker"],
                    lot["step"],
                    lot["round"],
                    lot["manage_no"],
                    lot["lot_no"],
                    lot["qty"],
                    qty_number,
                    calculate_result_value(qty_number),
                    lot["process_code"],
                    lot["condition"],
                    lot["jig"],
                    stack,
                    "기종교체" if model_change and index == 0 else "",
                    now_text,
                ),
            )
            ids.append(int(cursor.lastrowid))
        conn.commit()
        log_app(f"일반 DNC DB 저장: {len(ids)} LOT / ids={ids}")
    finally:
        conn.close()
    return ids


def update_normal_frequent_check_db(log_ids: list[int], model_change: bool, frequent_check: list[str], process_name: str = "KCC PKG") -> None:
    """?? 4Point? ?? Pin ?? ??? ??? work_log.db? ?????."""
    conn = get_process_connection(process_name)
    try:
        for index, log_id in enumerate(log_ids):
            first_values = frequent_check[:6]
            jig_values = frequent_check[6:] if model_change and index == 0 else [""] * 6
            conn.execute(
                """
                UPDATE dnc_logs
                   SET first_axis_1=?, first_axis_2=?, first_axis_3=?,
                       first_axis_4=?, first_axis_5=?, first_axis_6=?,
                       jig_axis_1=?, jig_axis_2=?, jig_axis_3=?,
                       jig_axis_4=?, jig_axis_5=?, jig_axis_6=?
                 WHERE id=?
                """,
                (*first_values, *jig_values, log_id),
            )
        conn.commit()
    finally:
        conn.close()


def update_normal_burr_db(log_ids: list[int], burr_ok: bool, process_name: str = "KCC PKG") -> None:
    """Burr 확인 결과를 공정별 work_log.db에 저장합니다."""
    result = "이상 없음" if burr_ok else "Burr 발생"
    now_text = datetime.now().strftime("%H:%M:%S")
    conn = get_process_connection(process_name)
    try:
        conn.executemany(
            "UPDATE dnc_logs SET status='\uc644\ub8cc', burr_result=?, record_time=? WHERE id=?",
            [(result, now_text, log_id) for log_id in log_ids],
        )
        conn.commit()
        log_app(f"{process_name} DNC Burr 확인: ids={log_ids}, Burr={result}")
    finally:
        conn.close()

def insert_new_model_db(common: dict, lot: dict, leader_name: str) -> int:
    """신규 모델 검증 DNC 이력을 KCC_PKG.db에 저장합니다."""
    qty_text = lot.get("qty", "").strip()
    qty_number = int(qty_text) if qty_text else 0
    display_qty = "더미" if qty_number == 0 else qty_text
    conn = get_kcc_pkg_connection()
    try:
        cursor = conn.execute(
            """
            INSERT INTO dnc_logs (
                dnc_type, status, machine, work_date, shift_group, shift_name, worker,
                step, round_no, manage_no, lot_no, qty_text, qty_number, result_value,
                process_code, condition_name, jig, model_change_text, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "신규 검증",
                "DNC 진행",
                normalize_machine_name(common["machine"]),
                common["work_date"],
                common["shift_group"],
                common["shift"],
                leader_name,
                lot["step"],
                lot["round"],
                lot["manage_no"],
                lot["lot_no"],
                display_qty,
                qty_number,
                calculate_result_value(qty_number),
                lot["process_code"],
                lot["condition"],
                lot["jig"],
                "신규 검증",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        conn.commit()
        log_id = int(cursor.lastrowid)
        log_app(f"신규 모델 DNC DB 저장: id={log_id}, STEP={lot['step']}, LOT={lot['lot_no']}")
        return log_id
    finally:
        conn.close()


def update_new_model_db(log_id: int, condition_name: str, first_article_ok: bool) -> None:
    """신규 모델 초도품 확인 결과를 KCC_PKG.db에 저장합니다."""
    final_condition = condition_name if first_article_ok else f"[검증 NG 발생] {condition_name}"
    conn = get_kcc_pkg_connection()
    try:
        conn.execute(
            "UPDATE dnc_logs SET status='완료', condition_name=?, record_time=? WHERE id=?",
            (final_condition, datetime.now().strftime("%H:%M:%S"), log_id),
        )
        conn.commit()
        log_app(f"신규 모델 DNC 완료 처리: id={log_id}, 초도품={'OK' if first_article_ok else 'NG'}")
    finally:
        conn.close()


def sync_condition_master_from_completed_logs() -> int:
    """완료된 DB 이력 중 조건 마스터에 반영 가능한 조건을 동기화합니다."""
    conn = get_kcc_pkg_connection()
    updated = 0
    try:
        rows = conn.execute(
            """
            SELECT dnc_type, step, round_no, manage_no, lot_no, process_code,
                   condition_name, jig, created_at
              FROM dnc_logs
             WHERE customer_process='KCC PKG'
               AND status='완료'
               AND step IS NOT NULL
               AND round_no IS NOT NULL
               AND manage_no IS NOT NULL
               AND process_code IS NOT NULL
               AND condition_name IS NOT NULL
               AND jig IS NOT NULL
             ORDER BY id
            """
        ).fetchall()
    finally:
        conn.close()

    for row in rows:
        condition = str(row["condition_name"] or "").strip()
        jig = str(row["jig"] or "").strip()
        if not condition or condition.startswith("[검증 NG 발생]") or not jig:
            continue
        lot = {
            "step": str(row["step"] or "").strip(),
            "round": str(row["round_no"] or "").strip(),
            "manage_no": str(row["manage_no"] or "").strip(),
            "lot_no": str(row["lot_no"] or "").strip(),
            "process_code": str(row["process_code"] or "").strip(),
        }
        source = "신규 검증 DB" if row["dnc_type"] == "신규 검증" else "DNC 완료 DB"
        before = len(load_condition_master())
        upsert_condition_master(lot, condition, jig, source)
        after = len(load_condition_master())
        updated += 1 if after >= before else 0
    return updated


def get_unexported_kcc_pkg_count() -> int:
    """Excel 작업일보에 아직 반영되지 않은 KCC PKG DB 이력 수를 반환합니다."""
    conn = get_kcc_pkg_connection()
    try:
        row = conn.execute("SELECT COUNT(*) AS count FROM dnc_logs WHERE customer_process='KCC PKG' AND exported=0 AND status='완료'").fetchone()
        return int(row["count"])
    finally:
        conn.close()


def get_incomplete_kcc_pkg_count() -> int:
    """DNC 중간 종료 등으로 완료 처리되지 않은 미반영 이력 수를 반환합니다."""
    conn = get_kcc_pkg_connection()
    try:
        row = conn.execute("SELECT COUNT(*) AS count FROM dnc_logs WHERE customer_process='KCC PKG' AND exported=0 AND status!='완료'").fetchone()
        return int(row["count"])
    finally:
        conn.close()




def get_unexported_process_log_count(process_name: str) -> int:
    conn = get_process_connection(process_name)
    try:
        row = conn.execute(
            "SELECT COUNT(*) AS count FROM dnc_logs WHERE customer_process=? AND exported=0 AND status='완료'",
            (process_name,),
        ).fetchone()
        return int(row["count"])
    finally:
        conn.close()


def get_incomplete_process_log_count(process_name: str) -> int:
    conn = get_process_connection(process_name)
    try:
        row = conn.execute(
            "SELECT COUNT(*) AS count FROM dnc_logs WHERE customer_process=? AND exported=0 AND status!='완료'",
            (process_name,),
        ).fetchone()
        return int(row["count"])
    finally:
        conn.close()


def format_excel_cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_round_key(value) -> str:
    text = format_excel_cell_value(value).upper().replace(" ", "")
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits or text


def normalize_condition_sheet_process(value) -> str:
    text = format_excel_cell_value(value).upper().replace(" ", "")
    if text in {"KCC", "KCCHDI", "KCC_HDI"}:
        return "KCC"
    if text == "TLB":
        return "TLB"
    return text


def lookup_tlb_condition_record_from_sheet(config: dict, tool_no: str, round_no: str, process_filter: str = "TLB") -> dict:
    sheet_path = Path(str(config.get("tlb_condition_sheet", "")).strip())
    if not sheet_path:
        raise FileNotFoundError("TLB 조건 시트 선택 필요")
    if not sheet_path.exists():
        raise FileNotFoundError("TLB 조건 시트 열림")
    try:
        workbook = load_workbook(sheet_path, read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(format_excel_error_for_operator(exc, "TLB 조건 시트")) from exc
    target_tool = tool_no.strip().upper()
    target_round = normalize_round_key(round_no)
    target_process = normalize_condition_sheet_process(process_filter)
    matches: list[dict] = []
    try:
        if "Database" not in workbook.sheetnames:
            raise KeyError("Database 시트 없음")
        ws = workbook["Database"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_process = normalize_condition_sheet_process(row[0] if len(row) > 0 else "")
            if target_process and row_process != target_process:
                continue
            row_tool = format_excel_cell_value(row[1] if len(row) > 1 else "").upper()
            row_round = normalize_round_key(row[2] if len(row) > 2 else "")
            if row_tool == target_tool and row_round == target_round:
                trim_program = normalize_tlb_condition_name(format_excel_cell_value(row[13] if len(row) > 13 else ""))
                jig_name = format_excel_cell_value(row[14] if len(row) > 14 else "")
                jig_value = format_excel_cell_value(row[17] if len(row) > 17 else "")
                jig = f"[{jig_name}] {jig_value}".strip() if jig_name else jig_value
                matches.append(
                    {
                        "condition": trim_program,
                        "jig": jig,
                        "product": format_excel_cell_value(row[3] if len(row) > 3 else ""),
                        "stack": format_excel_cell_value(row[6] if len(row) > 6 else ""),
                        "ukp": format_excel_cell_value(row[7] if len(row) > 7 else ""),
                        "top_left": format_excel_cell_value(row[15] if len(row) > 15 else ""),
                        "top": format_excel_cell_value(row[16] if len(row) > 16 else ""),
                        "center": format_excel_cell_value(row[17] if len(row) > 17 else ""),
                        "bottom_left": format_excel_cell_value(row[18] if len(row) > 18 else ""),
                        "bottom": format_excel_cell_value(row[19] if len(row) > 19 else ""),
                        "right_top": format_excel_cell_value(row[20] if len(row) > 20 else ""),
                        "left_mid": format_excel_cell_value(row[21] if len(row) > 21 else ""),
                        "left_gap": format_excel_cell_value(row[22] if len(row) > 22 else ""),
                        "right_mid": format_excel_cell_value(row[23] if len(row) > 23 else ""),
                        "right_bottom": format_excel_cell_value(row[24] if len(row) > 24 else ""),
                        "hole_shift": format_excel_cell_value(row[26] if len(row) > 26 else ""),
                    }
                )
    finally:
        workbook.close()
    unique = []
    for item in matches:
        if item not in unique:
            unique.append(item)
    if not unique:
        process_text = "KCC" if target_process == "KCC" else target_process
        raise LookupError(f"조건 시트에서 {process_text} / Tool No / 차수를 찾을 수 없습니다.\n기술 문의 바랍니다.")
    if len(unique) > 1:
        raise ValueError("조건 시트에 동일 조건이 2개 이상입니다.\n기술 문의 바랍니다.")
    condition = unique[0]["condition"]
    jig = unique[0]["jig"]
    if not condition or not jig:
        raise LookupError("조건 시트에 Trim 조건 또는 지그가 없습니다.\n기술 문의 바랍니다.")
    return unique[0]


def normalize_tlb_condition_name(value: str) -> str:
    """TLB 조건시트의 '3차 : 조건명' 표시에서 실제 DNC 조건명만 분리합니다."""
    text = str(value or "").strip()
    for delimiter in (":", "："):
        if delimiter in text:
            prefix, suffix = text.split(delimiter, 1)
            if "차" in prefix:
                return suffix.strip()
    return text


def lookup_tlb_condition_from_sheet(config: dict, tool_no: str, round_no: str) -> tuple[str, str]:
    record = lookup_tlb_condition_record_from_sheet(config, tool_no, round_no, process_filter="TLB")
    condition = record["condition"]
    jig = record["jig"]
    return condition, jig


def lookup_kcc_hdi_condition_record_from_sheet(config: dict, tool_no: str, round_no: str) -> dict:
    hdi_config = dict(config)
    hdi_config["tlb_condition_sheet"] = str(config.get("kcc_hdi_condition_sheet", "")).strip()
    try:
        return lookup_tlb_condition_record_from_sheet(hdi_config, tool_no, round_no, process_filter="KCC")
    except Exception as exc:
        message = format_excel_error_for_operator(exc, "KCC HDI 조건 시트")
        message = message.replace("TLB", "KCC HDI").replace("Tool No", "\uad00\ub9ac\ubc88\ud638")
        raise type(exc)(message) from exc


def lookup_kcc_hdi_condition_from_sheet(config: dict, tool_no: str, round_no: str) -> tuple[str, str]:
    record = lookup_kcc_hdi_condition_record_from_sheet(config, tool_no, round_no)
    return record["condition"], record["jig"]


def calculate_tlb_result_value(qty_number: int) -> float | None:
    if qty_number <= 0:
        return None
    return round(qty_number * 0.3, 1)


def insert_tlb_dnc_db(common: dict, lots: list[dict], stack: str, model_change: bool) -> list[int]:
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_process_connection("TLB")
    log_ids: list[int] = []
    try:
        for index, lot in enumerate(lots):
            qty_number = int(lot["qty"])
            cursor = conn.execute(
                """
                INSERT INTO dnc_logs (
                    customer_process, dnc_type, status, machine, work_date, shift_group, shift_name, worker,
                    step, round_no, manage_no, lot_no, qty_text, qty_number, result_value,
                    process_code, condition_name, jig, stack, model_change_text, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "TLB", "일반", "DNC 진행", normalize_machine_name(common["machine"]),
                    common["work_date"], common["shift_group"], common["shift"], common["worker"],
                    "", lot["round"], lot["manage_no"], lot["lot_no"], lot["qty"], qty_number,
                    calculate_tlb_result_value(qty_number), "", lot["condition"], lot["jig"], stack,
                    "기종교체" if model_change and index == 0 else "",
                    now_text,
                ),
            )
            log_ids.append(int(cursor.lastrowid))
        conn.commit()
        log_app(f"TLB DNC DB 저장: ids={log_ids}")
        return log_ids
    finally:
        conn.close()


def insert_kcc_hdi_dnc_db(common: dict, lots: list[dict], stack: str, model_change: bool) -> list[int]:
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_process_connection("KCC HDI")
    log_ids: list[int] = []
    try:
        for index, lot in enumerate(lots):
            qty_number = int(lot["qty"])
            cursor = conn.execute(
                """
                INSERT INTO dnc_logs (
                    customer_process, dnc_type, status, machine, work_date, shift_group, shift_name, worker,
                    step, round_no, manage_no, lot_no, qty_text, qty_number, result_value,
                    process_code, condition_name, jig, stack, model_change_text, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "KCC HDI", "일반", "DNC 진행", normalize_machine_name(common["machine"]),
                    common["work_date"], common["shift_group"], common["shift"], common["worker"],
                    "", lot["round"], lot["manage_no"], lot["lot_no"], lot["qty"], qty_number,
                    calculate_tlb_result_value(qty_number), "", lot["condition"], lot["jig"], stack,
                    "기종교체" if model_change and index == 0 else "",
                    now_text,
                ),
            )
            log_ids.append(int(cursor.lastrowid))
        conn.commit()
        log_app(f"KCC HDI DNC DB 저장: ids={log_ids}")
        return log_ids
    finally:
        conn.close()


def fetch_unexported_process_logs(process_name: str) -> list[sqlite3.Row]:
    conn = get_process_connection(process_name)
    try:
        return conn.execute(
            "SELECT * FROM dnc_logs WHERE customer_process=? AND exported=0 AND status='완료' ORDER BY id",
            (process_name,),
        ).fetchall()
    finally:
        conn.close()


def export_process_logs_to_excel(config: dict, process_name: str, sheet_name: str) -> int:
    logs = fetch_unexported_process_logs(process_name)
    if not logs:
        return 0
    exported_ids: list[int] = []
    workbook = None
    lock_path = None
    try:
        excel_path = Path(config.get("excel_file", ""))
        if not excel_path:
            raise FileNotFoundError("작업일보 경로 선택 필요")
        if not excel_path.exists():
            raise FileNotFoundError("작업일보 파일 없음")
        lock_path = acquire_excel_export_lock(excel_path)
        workbook, ws, path = open_log_workbook(config, sheet_name)
        existing_excel_ids = get_excel_exported_log_ids(ws)
        start_row = get_next_empty_row(ws)
        written_count = 0
        for log in logs:
            log_id = int(log["id"])
            if log_id in existing_excel_ids:
                exported_ids.append(log_id)
                continue
            write_process_log_row_to_excel(ws, start_row + written_count, log)
            exported_ids.append(log_id)
            written_count += 1
        if written_count:
            save_workbook_safely(workbook, path)
    except Exception as exc:
        log_error(f"{process_name} 작업일보 반영 실패", exc)
        raise
    finally:
        if workbook is not None:
            workbook.close()
        release_excel_export_lock(lock_path)
    mark_process_logs_exported(process_name, exported_ids)
    log_app(f"{process_name} 작업일보 반영 완료: {len(exported_ids)}건")
    return len(exported_ids)

def fetch_incomplete_kcc_pkg_logs() -> list[sqlite3.Row]:
    """앱 시작 시 안내할 미완료 이력을 조회합니다."""
    conn = get_kcc_pkg_connection()
    try:
        return conn.execute(
            """
            SELECT id, dnc_type, status, work_date, worker, step, round_no,
                   manage_no, lot_no, condition_name, created_at
              FROM dnc_logs
             WHERE status!='완료'
             ORDER BY id
            """
        ).fetchall()
    finally:
        conn.close()


def delete_incomplete_kcc_pkg_logs(log_ids: list[int]) -> None:
    """미완료 이력을 삭제합니다. 시작 알림에서 작업자가 선택했을 때만 사용합니다."""
    if not log_ids:
        return
    conn = get_kcc_pkg_connection()
    try:
        conn.executemany("DELETE FROM dnc_logs WHERE id=? AND status!='완료'", [(log_id,) for log_id in log_ids])
        conn.commit()
        log_app(f"미완료 이력 삭제: ids={log_ids}")
    finally:
        conn.close()


def complete_incomplete_kcc_pkg_logs(log_ids: list[int]) -> None:
    """미완료 이력을 완료 처리합니다. 작업일보 반영 대상은 아니게 exported=1로 처리합니다."""
    if not log_ids:
        return
    conn = get_kcc_pkg_connection()
    try:
        now_text = datetime.now().strftime("%H:%M:%S")
        exported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.executemany(
            """
            UPDATE dnc_logs
               SET status='완료',
                   record_time=COALESCE(NULLIF(record_time, ''), ?),
                   exported=1,
                   exported_at=COALESCE(NULLIF(exported_at, ''), ?)
             WHERE id=? AND status!='완료'
            """,
            [(now_text, exported_at, log_id) for log_id in log_ids],
        )
        conn.commit()
        log_app(f"미완료 이력 완료 처리: ids={log_ids}")
    finally:
        conn.close()


def load_work_history(limit: int = 500, only_unexported: bool = False, only_incomplete: bool = False, keyword: str = "", process_name: str = "") -> list[sqlite3.Row]:
    """?? ?? ?? ??? ??? ?? DNC ??? ?????."""
    selected_process = process_name.strip()
    conn = get_process_connection(selected_process) if selected_process else get_kcc_pkg_connection()
    try:
        where = []
        params: list[object] = []
        if selected_process:
            where.append("customer_process=?")
            params.append(selected_process)
        if only_unexported:
            where.append("exported=0 AND status='\uc644\ub8cc'")
        if only_incomplete:
            where.append("status!='\uc644\ub8cc'")
        search = keyword.strip()
        if search:
            where.append("(lot_no LIKE ? OR manage_no LIKE ? OR condition_name LIKE ? OR worker LIKE ?)")
            like = f"%{search}%"
            params.extend([like, like, like, like])
        sql = """
            SELECT id, dnc_type, status, machine, work_date, shift_group, shift_name, worker,
                   step, round_no, manage_no, lot_no, qty_text, result_value,
                   condition_name, jig, stack, model_change_text, burr_result,
                   record_time, exported, exported_at, created_at
              FROM dnc_logs
        """
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY id DESC LIMIT ?"
        params.append(max(1, min(int(limit), 5000)))
        return conn.execute(sql, params).fetchall()
    finally:
        conn.close()

def autosize_excel_columns(ws) -> None:
    """내보낸 엑셀 파일을 바로 보기 좋게 열 너비를 맞춥니다."""
    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 3, 10), 45)


def save_rows_to_excel(path: Path, sheet_name: str, headers: list[str], rows: list[list[object]]) -> None:
    """조회 결과를 새 Excel 파일로 저장합니다."""
    workbook = Workbook()
    ws = workbook.active
    ws.title = sheet_name
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    autosize_excel_columns(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def export_kcc_pkg_work_history_excel(path: Path) -> Path:
    """KCC PKG 작업 이력을 확인용 Excel 파일로 내보냅니다."""
    conn = get_kcc_pkg_connection()
    try:
        records = conn.execute(
            """
            SELECT id, dnc_type, status, exported, machine, work_date, shift_group, shift_name,
                   worker, step, round_no, manage_no, lot_no, qty_text, result_value,
                   process_code, condition_name, jig, stack, model_change_text, burr_result,
                   record_time, exported_at, created_at
              FROM dnc_logs
              WHERE customer_process='KCC PKG'
             ORDER BY id
            """
        ).fetchall()
    finally:
        conn.close()

    headers = [
        "ID", "구분", "상태", "Excel", "호기", "작업일자", "조", "근무", "작업자",
        "STEP", "차수", "관리번호", "LOT", "매수", "실적", "공정코드",
        "작업P/G", "지그", "Stack", "기종교체", "Burr", "DNC 시작시간", "DNC 완료시간", "엑셀 반영시간",
    ]
    rows = [
        [
            record["id"],
            record["dnc_type"],
            record["status"],
            "반영" if int(record["exported"] or 0) else "미반영",
            record["machine"],
            record["work_date"],
            record["shift_group"],
            record["shift_name"],
            record["worker"],
            record["step"],
            record["round_no"],
            record["manage_no"],
            record["lot_no"],
            record["qty_text"],
            record["result_value"],
            record["process_code"],
            record["condition_name"],
            record["jig"],
            record["stack"],
            record["model_change_text"],
            record["burr_result"],
            record["created_at"],
            record["record_time"],
            record["exported_at"],
        ]
        for record in records
    ]
    save_rows_to_excel(path, "KCC PKG 작업이력", headers, rows)
    log_app(f"KCC PKG 작업 이력 Excel 내보내기: {path}")
    return path



def export_process_work_history_excel(process_name: str, path: Path) -> Path:
    """선택한 공정의 작업 이력을 확인용 Excel 파일로 내보냅니다."""
    conn = get_process_connection(process_name)
    try:
        records = conn.execute(
            """
            SELECT id, dnc_type, status, exported, machine, work_date, shift_group, shift_name,
                   worker, step, round_no, manage_no, lot_no, qty_text, result_value,
                   process_code, condition_name, jig, stack, model_change_text, burr_result,
                   record_time, exported_at, created_at
              FROM dnc_logs
             WHERE customer_process=?
             ORDER BY id
            """,
            (process_name,),
        ).fetchall()
    finally:
        conn.close()

    headers = [
        "ID", "구분", "상태", "Excel", "호기", "작업일자", "조", "근무", "작업자",
        "STEP", "차수", "관리번호", "LOT", "매수", "실적", "공정코드",
        "작업P/G", "지그", "Stack", "기종교체", "Burr", "DNC 시작시간", "DNC 완료시간", "엑셀 반영시간",
    ]
    rows = [
        [
            record["id"],
            record["dnc_type"],
            record["status"],
            "반영" if int(record["exported"] or 0) else "미반영",
            record["machine"],
            record["work_date"],
            record["shift_group"],
            record["shift_name"],
            record["worker"],
            record["step"],
            record["round_no"],
            record["manage_no"],
            record["lot_no"],
            record["qty_text"],
            record["result_value"],
            record["process_code"],
            record["condition_name"],
            record["jig"],
            record["stack"],
            record["model_change_text"],
            record["burr_result"],
            record["created_at"],
            record["record_time"],
            record["exported_at"],
        ]
        for record in records
    ]
    sheet_title = f"{process_name} 작업이력"[:31]
    save_rows_to_excel(path, sheet_title, headers, rows)
    log_app(f"{process_name} 작업 이력 Excel 내보내기: {path}")
    return path


def export_tlb_condition_sheet_excel(config: dict, path: Path, process_name: str = "TLB", config_key: str = "tlb_condition_sheet") -> Path:
    """TLB 계열 기술 조건시트(Database)를 확인용 Excel 파일로 내보냅니다."""
    sheet_path = Path(str(config.get(config_key, "")).strip())
    if not sheet_path:
        raise FileNotFoundError(f"{process_name} 조건 시트 선택 필요")
    if not sheet_path.exists():
        raise FileNotFoundError(f"{process_name} 조건 시트 없음")
    workbook = load_workbook(sheet_path, read_only=True, data_only=True)
    try:
        if "Database" not in workbook.sheetnames:
            raise KeyError("Database 시트 없음")
        ws = workbook["Database"]
        key_header = "관리번호" if process_name == "KCC HDI" else "Tool No"
        headers = [key_header, "차수", "Trim Program", "적용 지그", "지그", "표시 지그", "제품명", "Stack", "UKP"]
        rows: list[list[object]] = []
        seen: set[tuple[str, str, str, str]] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            tool_no = format_excel_cell_value(row[1] if len(row) > 1 else "").upper()
            round_no = format_excel_cell_value(row[2] if len(row) > 2 else "")
            condition = normalize_tlb_condition_name(format_excel_cell_value(row[13] if len(row) > 13 else ""))
            applied_jig = format_excel_cell_value(row[14] if len(row) > 14 else "")
            jig_value = format_excel_cell_value(row[17] if len(row) > 17 else "")
            display_jig = f"[{applied_jig}] {jig_value}".strip() if applied_jig else jig_value
            if not tool_no or not round_no or not condition:
                continue
            key = (tool_no, normalize_round_key(round_no), condition, display_jig)
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                [
                    tool_no,
                    round_no,
                    condition,
                    applied_jig,
                    jig_value,
                    display_jig,
                    format_excel_cell_value(row[3] if len(row) > 3 else ""),
                    format_excel_cell_value(row[6] if len(row) > 6 else ""),
                    format_excel_cell_value(row[7] if len(row) > 7 else ""),
                ]
            )
    finally:
        workbook.close()
    save_rows_to_excel(path, f"{process_name} 조건시트"[:31], headers, rows)
    log_app(f"{process_name} 조건시트 Excel 내보내기: {path}")
    return path


def export_process_condition_master_excel(process_name: str, config: dict, path: Path) -> Path:
    """공정별 작업조건/지그 정보를 확인용 Excel 파일로 내보냅니다."""
    if process_name == "KCC PKG":
        return export_kcc_pkg_condition_master_excel(path)
    if process_name == "TLB":
        return export_tlb_condition_sheet_excel(config, path)
    if process_name == "KCC HDI":
        return export_tlb_condition_sheet_excel(config, path, "KCC HDI", "kcc_hdi_condition_sheet")
    raise NotImplementedError(f"{process_name} 준비중")
def export_kcc_pkg_condition_master_excel(path: Path) -> Path:
    """KCC PKG 작업조건/지그 마스터를 확인용 Excel 파일로 내보냅니다."""
    records = load_condition_master()
    headers = ["STEP", "차수", "관리번호", "공정코드", "LOT", "작업조건", "지그", "출처", "수정시간"]
    rows = [
        [
            record.get("step", ""),
            record.get("round", ""),
            record.get("manage_no", ""),
            record.get("process_code", ""),
            record.get("lot_no", ""),
            record.get("condition", ""),
            record.get("jig", ""),
            record.get("source", ""),
            record.get("updated_at", ""),
        ]
        for record in records
    ]
    save_rows_to_excel(path, "KCC PKG 조건마스터", headers, rows)
    log_app(f"KCC PKG 작업조건/지그 Excel 내보내기: {path}")
    return path


def create_full_data_backup_zip(path: Path) -> Path:
    """현장 data 폴더 전체를 zip으로 백업합니다. export 폴더는 중복 백업 방지를 위해 제외합니다."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as backup_zip:
        for item in DATA_DIR.rglob("*"):
            if not item.is_file():
                continue
            try:
                item.relative_to(EXPORT_DIR)
                continue
            except ValueError:
                pass
            backup_zip.write(item, item.relative_to(DATA_DIR))
    log_app(f"전체 data 백업 생성: {path}")
    return path


def cleanup_old_text_logs(days: int = 90) -> None:
    """오래된 텍스트 로그만 조용히 정리합니다. DB/작업일보/마스터는 건드리지 않습니다."""
    try:
        if not LOG_DIR.exists():
            return
        cutoff = datetime.now() - timedelta(days=days)
        for log_file in LOG_DIR.glob("*.log"):
            try:
                modified = datetime.fromtimestamp(log_file.stat().st_mtime)
                if modified < cutoff:
                    log_file.unlink()
            except OSError:
                continue
    except Exception as exc:
        log_error("오래된 로그 정리 실패", exc)


def create_daily_master_backup(config: dict) -> None:
    """조건 마스터와 설정만 하루 1회 자동 백업합니다."""
    today = datetime.now().strftime("%Y-%m-%d")
    if config.get("last_master_auto_backup_date") == today:
        return
    try:
        AUTO_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        path = AUTO_BACKUP_DIR / f"{datetime.now().strftime('%Y%m%d')}_master_backup.zip"
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as backup_zip:
            if CONFIG_FILE.exists():
                backup_zip.write(CONFIG_FILE, "config.json")
            if CONFIG_BACKUP_FILE.exists():
                backup_zip.write(CONFIG_BACKUP_FILE, "config.json.bak")
            if CONDITION_MASTER_DB_FILE.exists():
                backup_zip.write(CONDITION_MASTER_DB_FILE, "KCC_PKG/condition_master.db")
            wal_file = CONDITION_MASTER_DB_FILE.with_name(CONDITION_MASTER_DB_FILE.name + "-wal")
            shm_file = CONDITION_MASTER_DB_FILE.with_name(CONDITION_MASTER_DB_FILE.name + "-shm")
            if wal_file.exists():
                backup_zip.write(wal_file, "KCC_PKG/condition_master.db-wal")
            if shm_file.exists():
                backup_zip.write(shm_file, "KCC_PKG/condition_master.db-shm")
        config["last_master_auto_backup_date"] = today
        save_config(config)
        log_app(f"마스터 자동 백업 생성: {path}")
    except Exception as exc:
        log_error("마스터 자동 백업 실패", exc)


def fetch_unexported_kcc_pkg_logs() -> list[sqlite3.Row]:
    """Excel로 내보낼 미반영 KCC PKG 이력을 오래된 순서로 조회합니다."""
    conn = get_kcc_pkg_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM dnc_logs WHERE customer_process='KCC PKG' AND exported=0 AND status='완료' ORDER BY id"
        ).fetchall()
        return rows
    finally:
        conn.close()


def mark_process_logs_exported(process_name: str, log_ids: list[int]) -> None:
    """Excel ??? ?? ??? DB ??? ?? ?? ?????."""
    if not log_ids:
        return
    conn = get_process_connection(process_name)
    try:
        conn.executemany(
            "UPDATE dnc_logs SET exported=1, exported_at=? WHERE id=?",
            [(datetime.now().strftime("%Y-%m-%d %H:%M:%S"), log_id) for log_id in log_ids],
        )
        conn.commit()
    finally:
        conn.close()


def mark_kcc_pkg_logs_exported(log_ids: list[int]) -> None:
    mark_process_logs_exported("KCC PKG", log_ids)

def write_tlb_log_row_to_excel(ws, row: int, log: sqlite3.Row) -> None:
    """TLB 이력 한 건을 TLB 작업일보 양식에 맞춰 기록합니다.

    KCC PKG 저장 함수와 분리해 KCC 배포본 양식 변경 위험을 줄입니다.
    TLB는 STEP/공정코드를 사용하지 않으므로 작업일보에는 빈칸으로 남깁니다.
    """
    values = [
        log["machine"] or "",
        log["work_date"],
        log["shift_group"],
        log["shift_name"],
        log["worker"],
        "",
        log["round_no"],
        log["manage_no"],
        log["lot_no"],
        log["qty_text"],
        "" if log["result_value"] is None else log["result_value"],
        log["condition_name"],
        log["burr_result"] or "",
        log["stack"] or "",
        log["jig"],
        log["model_change_text"] or "",
        log["record_time"] or "",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(row=row, column=col).value = value if col == 12 else excel_upper_value(value)
    if log["burr_result"] == "Burr 발생":
        ws.cell(row=row, column=13).font = Font(color="FF0000", bold=False)

    frequent_values = [
        log["first_axis_1"],
        log["first_axis_2"],
        log["first_axis_3"],
        log["first_axis_4"],
        log["first_axis_5"],
        log["first_axis_6"],
        log["jig_axis_1"],
        log["jig_axis_2"],
        log["jig_axis_3"],
        log["jig_axis_4"],
        log["jig_axis_5"],
        log["jig_axis_6"],
    ]
    for offset, value in enumerate(frequent_values, start=18):
        ws.cell(row=row, column=offset).value = value or ""
    write_export_id_backup(ws, row, int(log["id"]))


def write_process_log_row_to_excel(ws, row: int, log: sqlite3.Row) -> None:
    """공정별 작업일보 저장 라우터입니다. 현재 TLB만 KCC writer와 분리합니다."""
    process_name = str(log["customer_process"] or "")
    if process_name in {"TLB", "KCC HDI"}:
        write_tlb_log_row_to_excel(ws, row, log)
        return
    write_db_log_row_to_excel(ws, row, log)

def write_db_log_row_to_excel(ws, row: int, log: sqlite3.Row) -> None:
    """KCC_PKG.db 한 건을 기존 KCC PKG 작업일보 양식에 맞춰 기록합니다."""
    values = [
        log["machine"] or "",
        log["work_date"],
        log["shift_group"],
        log["shift_name"],
        log["worker"],
        log["step"],
        log["round_no"],
        log["manage_no"],
        log["lot_no"],
        log["qty_text"],
        "" if log["result_value"] is None else log["result_value"],
        log["condition_name"],
        log["burr_result"] or "",
        log["stack"] or "",
        log["jig"],
        log["model_change_text"] or "",
        log["record_time"] or "",
    ]
    for col, value in enumerate(values, start=1):
        # L열 작업 P/G는 조건 파일명과 같은 원문을 유지하고, 나머지 문자 값만 대문자로 정리합니다.
        ws.cell(row=row, column=col).value = value if col == 12 else excel_upper_value(value)
    write_process_code_backup(ws, row, log["process_code"] or "")
    if log["burr_result"] == "Burr 발생":
        ws.cell(row=row, column=13).font = Font(color="FF0000", bold=False)
    if str(log["condition_name"] or "").startswith("[검증 NG 발생]"):
        ws.cell(row=row, column=12).font = Font(name="맑은 고딕", size=10, color="FF0000", bold=False)

    frequent_values = [
        log["first_axis_1"],
        log["first_axis_2"],
        log["first_axis_3"],
        log["first_axis_4"],
        log["first_axis_5"],
        log["first_axis_6"],
        log["jig_axis_1"],
        log["jig_axis_2"],
        log["jig_axis_3"],
        log["jig_axis_4"],
        log["jig_axis_5"],
        log["jig_axis_6"],
    ]
    for offset, value in enumerate(frequent_values, start=18):
        ws.cell(row=row, column=offset).value = value or ""
    write_export_id_backup(ws, row, int(log["id"]))


def export_kcc_pkg_db_to_excel(config: dict) -> int:
    """KCC_PKG.db의 미반영 이력을 Excel KCC PKG 시트로 내보냅니다."""
    logs = fetch_unexported_kcc_pkg_logs()
    if not logs:
        return 0

    exported_ids: list[int] = []
    log_app(f"작업일보 반영 시작: {len(logs)}건")
    workbook = None
    lock_path = None
    try:
        excel_path = Path(config.get("excel_file", ""))
        if not excel_path:
            raise FileNotFoundError("작업일보 파일 선택 필요")
        if not excel_path.exists():
            raise FileNotFoundError("작업일보 파일 없음")
        lock_path = acquire_excel_export_lock(excel_path)
        workbook, ws, path = open_log_workbook(config)
        existing_excel_ids = get_excel_exported_log_ids(ws)
        start_row = get_next_empty_row(ws)
        written_count = 0
        for log in logs:
            log_id = int(log["id"])
            if log_id in existing_excel_ids:
                exported_ids.append(log_id)
                log_app(f"작업일보 중복 반영 방지: 이미 Excel에 있는 DB id={log_id}")
                continue
            write_db_log_row_to_excel(ws, start_row + written_count, log)
            exported_ids.append(log_id)
            written_count += 1
        if written_count:
            save_workbook_safely(workbook, path)
    except Exception as exc:
        log_error("작업일보 반영 실패 - exported 상태 변경 안 함", exc)
        raise
    finally:
        if workbook is not None:
            workbook.close()
        release_excel_export_lock(lock_path)
    mark_kcc_pkg_logs_exported(exported_ids)
    log_app(f"작업일보 반영 완료: {len(exported_ids)}건")
    return len(exported_ids)


def export_process_db_to_excel(process_name: str, config: dict) -> int:
    """공정별 미반영 이력을 작업일보에 반영합니다. 현재는 KCC PKG만 실제 구현되어 있습니다."""
    if process_name == "KCC PKG":
        return export_kcc_pkg_db_to_excel(config)
    if process_name == "TLB":
        return export_process_logs_to_excel(config, "TLB", "TLB")
    if process_name == "KCC HDI":
        return export_process_logs_to_excel(config, "KCC HDI", "KCC HDI")
    return 0


def get_unexported_process_count(process_name: str) -> int:
    """공정별 Excel 미반영 수를 반환합니다. 미구현 공정은 0건으로 둡니다."""
    if process_name == "KCC PKG":
        return get_unexported_kcc_pkg_count()
    if process_name == "TLB":
        return get_unexported_process_log_count("TLB")
    if process_name == "KCC HDI":
        return get_unexported_process_log_count("KCC HDI")
    return 0


def get_incomplete_process_count(process_name: str) -> int:
    """공정별 미완료 수를 반환합니다. 미구현 공정은 0건으로 둡니다."""
    if process_name == "KCC PKG":
        return get_incomplete_kcc_pkg_count()
    if process_name == "TLB":
        return get_incomplete_process_log_count("TLB")
    if process_name == "KCC HDI":
        return get_incomplete_process_log_count("KCC HDI")
    return 0


def export_all_processes_to_excel(config: dict) -> dict[str, int]:
    """수동 작업일보 반영 버튼에서 5개 공정을 순서대로 확인하고 반영합니다."""
    result: dict[str, int] = {}
    for process_name in PROCESS_NAMES:
        result[process_name] = export_process_db_to_excel(process_name, config)
    return result


def make_condition_key(step: str, round_no: str, manage_no: str, process_code: str) -> str:
    """조건 마스터에서 중복을 제거하기 위한 기준 키를 만듭니다.

    작업조건/지그는 STEP, 차수, 공정코드가 맞을 때만 불러와야 하므로
    공정코드까지 중복 판단 기준에 포함합니다.
    """
    return "|".join(
        [
            step.strip(),
            round_no.strip(),
            process_code.strip(),
            manage_no.strip(),
        ]
    )


def make_condition_record_key(record: dict) -> str:
    """조건 마스터의 조회 기준 키입니다. 4개 항목이 모두 같아야 같은 조건입니다."""
    return make_condition_key(
        record.get("step", ""),
        record.get("round", ""),
        record.get("manage_no", ""),
        record.get("process_code", ""),
    )


def make_condition_value_key(record: dict) -> str:
    """동일 조회 키 안에서 작업조건/지그까지 같은 완전 동일 행인지 확인합니다."""
    return "|".join(
        [
            make_condition_record_key(record),
            str(record.get("condition", "")).strip(),
            str(record.get("jig", "")).strip(),
        ]
    )


def get_duplicate_condition_keys(records: list[dict]) -> set[str]:
    """같은 STEP/차수/관리번호/공정코드에 조건이 2개 이상인 키를 찾습니다."""
    counts: dict[str, set[str]] = {}
    for record in records:
        key = make_condition_record_key(record)
        if not key.replace("|", "").strip():
            continue
        counts.setdefault(key, set()).add(make_condition_value_key(record))
    return {key for key, values in counts.items() if len(values) >= 2}


def get_condition_source_priority(source: str) -> int:
    """조건 마스터 출처별 신뢰 우선순위를 반환합니다.

    신규 검증으로 확정한 조건은 작업일보 이력보다 우선합니다.
    현장에서 직접 수정한 값은 의도적으로 고친 값이므로 가장 높게 둡니다.
    """
    source_text = str(source or "")
    if "사용자 수정" in source_text:
        return 40
    if "신규 검증" in source_text:
        return 30
    if "DNC 완료 DB" in source_text:
        return 20
    if "작업일보" in source_text:
        return 10
    return 0


def should_replace_condition_record(current: dict, incoming: dict) -> bool:
    """조건 마스터 중복 키 병합 시 새 기록으로 교체해도 되는지 판단합니다."""
    current_priority = get_condition_source_priority(current.get("source", ""))
    incoming_priority = get_condition_source_priority(incoming.get("source", ""))
    if incoming_priority < current_priority:
        return False
    if incoming_priority == current_priority:
        current_updated = str(current.get("updated_at", ""))
        incoming_updated = str(incoming.get("updated_at", ""))
        if current_updated and incoming_updated and incoming_updated < current_updated:
            return False
    return True


def find_manage_no_conflict(records: list[dict], incoming: dict) -> dict | None:
    """동일 관리번호가 다른 조건 키로 존재하는지 확인합니다."""
    # 관리번호가 같아도 STEP/차수/공정코드가 다르면 실제 작업 조건이 다를 수 있습니다.
    # 중복 판단은 make_condition_key(STEP+차수+관리번호+공정코드) 기준으로만 처리합니다.
    return None


def merge_condition_records(records: list[dict]) -> list[dict]:
    """?? ???? ?????.

    STEP/??/????/????? ??? ???? ?? ??? ???
    ??? ? ? ?????. ? ?? ?? ? ?? ???? ????,
    ??? ?? ????? ??? ????? ???.
    """
    merged: dict[str, dict] = {}
    for record in records:
        if not str(record.get("process_code", "")).strip():
            continue
        lookup_key = make_condition_record_key(record)
        if not lookup_key.replace("|", "").strip():
            continue
        value_key = make_condition_value_key(record)
        if value_key not in merged:
            merged[value_key] = dict(record)
            continue

        current = merged[value_key]
        if not should_replace_condition_record(current, record):
            continue
        for field in ("step", "round", "manage_no", "condition", "jig", "source", "updated_at", "lot_no"):
            value = str(record.get(field, "")).strip()
            if value:
                current[field] = value
        process_code = str(record.get("process_code", "")).strip()
        if process_code:
            current["process_code"] = process_code

    return sorted(
        merged.values(),
        key=lambda item: (
            item.get("manage_no", ""),
            item.get("round", ""),
            item.get("step", ""),
            item.get("process_code", ""),
            item.get("condition", ""),
            item.get("jig", ""),
        ),
    )

def load_condition_master() -> list[dict]:
    """KCC PKG 조건 마스터 DB를 읽습니다. 예전 JSON 파일이 있으면 1회 가져옵니다."""
    conn = get_condition_master_connection()
    try:
        count = int(conn.execute("SELECT COUNT(*) FROM condition_master").fetchone()[0])
        if count == 0 and LEGACY_CONDITION_MASTER_FILE.exists():
            try:
                data = json.loads(LEGACY_CONDITION_MASTER_FILE.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    records = merge_condition_records(data)
                    for record in records:
                        conn.execute(
                            """
                            INSERT OR REPLACE INTO condition_master (
                                step, round_no, manage_no, process_code, lot_no,
                                condition_name, jig, source, updated_at
                            )
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                record.get("step", ""),
                                record.get("round", ""),
                                record.get("manage_no", ""),
                                record.get("process_code", ""),
                                record.get("lot_no", ""),
                                record.get("condition", ""),
                                record.get("jig", ""),
                                record.get("source", ""),
                                record.get("updated_at", ""),
                            ),
                        )
                    conn.commit()
            except Exception:
                pass
        rows = conn.execute(
            """
            SELECT step, round_no, manage_no, process_code, lot_no,
                   condition_name, jig, source, updated_at
              FROM condition_master
             ORDER BY manage_no, round_no, step, process_code
            """
        ).fetchall()
        return [
            {
                "step": row["step"] or "",
                "round": row["round_no"] or "",
                "manage_no": row["manage_no"] or "",
                "process_code": row["process_code"] or "",
                "lot_no": row["lot_no"] or "",
                "condition": row["condition_name"] or "",
                "jig": row["jig"] or "",
                "source": row["source"] or "",
                "updated_at": row["updated_at"] or "",
            }
            for row in rows
        ]
    finally:
        conn.close()


def save_condition_master(records: list[dict]) -> None:
    """조건 마스터 DB를 저장합니다."""
    records = merge_condition_records(records)
    conn = get_condition_master_connection()
    try:
        conn.execute("DELETE FROM condition_master")
        for record in records:
            conn.execute(
                """
                INSERT OR REPLACE INTO condition_master (
                    step, round_no, manage_no, process_code, lot_no,
                    condition_name, jig, source, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    record.get("step", ""),
                    record.get("round", ""),
                    record.get("manage_no", ""),
                    record.get("process_code", ""),
                    record.get("lot_no", ""),
                    record.get("condition", ""),
                    record.get("jig", ""),
                    record.get("source", ""),
                    record.get("updated_at", ""),
                ),
            )
        conn.commit()
    finally:
        conn.close()


def upsert_condition_master(lot: dict, condition: str, jig: str, source: str) -> None:
    """?? ???? ??/??? ?? ???? ?? ?? ?????.

    ?? ??? STEP + ?? + ???? + ???????.
    ?? ?? ??? ????/??? ??? ?? ??? ? ? ?????.
    """
    step = lot.get("step", "").strip()
    round_no = lot.get("round", "").strip()
    manage_no = lot.get("manage_no", "").strip()
    process_code = lot.get("process_code", "").strip()
    condition = str(condition or "").strip()
    jig = str(jig or "").strip()
    if not (step and round_no and process_code and manage_no and condition and jig):
        return

    key = make_condition_key(step, round_no, manage_no, process_code)
    records = load_condition_master()
    incoming_record = {
        "step": step,
        "round": round_no,
        "manage_no": manage_no,
        "process_code": process_code,
        "lot_no": lot.get("lot_no", "").strip(),
        "condition": condition,
        "jig": jig,
        "source": source,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    incoming_value_key = make_condition_value_key(incoming_record)

    same_lookup_count = 0
    for record in records:
        record_key = make_condition_record_key(record)
        if record_key != key:
            continue
        same_lookup_count += 1
        if make_condition_value_key(record) != incoming_value_key:
            continue
        incoming = {"source": source}
        if not should_replace_condition_record(record, incoming):
            log_app(
                "조건 마스터 갱신 건너뜀: "
                f"기존 출처={record.get('source', '')}, 새 출처={source}, "
                f"STEP={step}, 차수={round_no}, 관리번호={manage_no}, 공정코드={process_code}"
            )
            return
        record.update(
            {
                "lot_no": lot.get("lot_no", "").strip(),
                "condition": condition,
                "jig": jig,
                "source": source,
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
        )
        save_condition_master(records)
        return

    if same_lookup_count:
        log_app(
            "조건 마스터 중복 후보 추가: "
            f"STEP={step}, 차수={round_no}, 관리번호={manage_no}, 공정코드={process_code}"
        )
    records.append(incoming_record)
    save_condition_master(records)

def rebuild_condition_master_from_log(config: dict) -> int:
    """작업일보 KCC PKG 시트의 최신 이력을 기존 마스터에 병합합니다.

    작업일보에서 사라진 모델도 조건 마스터에서는 보존합니다.
    """
    # 작업일보에 아직 반영되지 않은 신규 검증/완료 DB 이력도 먼저 마스터에 반영합니다.
    # 현장에서 Excel이 늦게 반영되어도 신규 검증 조건을 바로 조회할 수 있게 하기 위함입니다.
    sync_condition_master_from_completed_logs()
    workbook, ws, _path = open_log_workbook(config)
    records = load_condition_master()
    before_keys = {make_condition_value_key(record) for record in records}
    try:
        for row in range(8, ws.max_row + 1):
            step = str(ws.cell(row=row, column=6).value or "").strip()
            round_no = str(ws.cell(row=row, column=7).value or "").strip()
            manage_no = str(ws.cell(row=row, column=8).value or "").strip()
            # AD열은 프로그램 복구용 공정코드 백업 칸입니다.
            # AD가 비어 있는 예전 이력은 정확한 조건 키를 만들 수 없어 건너뜁니다.
            process_code = str(ws.cell(row=row, column=30).value or "").strip()
            condition = str(ws.cell(row=row, column=12).value or "").strip()
            jig = str(ws.cell(row=row, column=15).value or "").strip()
            lot_no = str(ws.cell(row=row, column=9).value or "").strip()
            if condition.startswith("[검증 NG 발생]"):
                continue
            if not (step and round_no and manage_no and process_code and condition and jig):
                continue
            records.append(
                {
                    "step": step,
                    "round": round_no,
                    "manage_no": manage_no,
                    "process_code": process_code,
                    "lot_no": lot_no,
                    "condition": condition,
                    "jig": jig,
                    "source": f"작업일보 {row}행",
                    "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
            )
    finally:
        workbook.close()

    save_condition_master(records)
    after_records = load_condition_master()
    after_keys = {make_condition_value_key(record) for record in after_records}
    return len(after_keys - before_keys)


def lookup_condition_jig_from_master(lot: dict) -> tuple[str, str, str]:
    """?? ????? ????/??? ????.

    STEP + ?? + ???? + ????? ?? ???? ?????.
    ?? ?? ??/??? 2? ???? ???? ?? ???? ??? ?????.
    """
    records = load_condition_master()
    step = lot.get("step", "").strip()
    round_no = lot.get("round", "").strip()
    manage_no = lot.get("manage_no", "").strip()
    process_code = lot.get("process_code", "").strip()
    if not (step and round_no and manage_no and process_code):
        return "", "", ""

    matches = [
        record
        for record in records
        if record.get("step", "") == step
        and record.get("round", "") == round_no
        and record.get("manage_no", "") == manage_no
        and record.get("process_code", "") == process_code
    ]
    distinct: dict[str, dict] = {}
    for record in matches:
        distinct[make_condition_value_key(record)] = record
    matches = list(distinct.values())
    if len(matches) >= 2:
        details = " / ".join(
            f"{record.get('condition', '')}, 지그 {record.get('jig', '')}"
            for record in matches[:3]
        )
        raise ValueError(f"중복 조건\n조건 마스터에서 하나만 남기세요.\n{details}")
    if len(matches) == 1:
        record = matches[0]
        return record.get("condition", ""), record.get("jig", ""), f"조건 마스터({record.get('source', '저장값')})"

    return "", "", ""

def describe_condition_lookup_mismatch(lot: dict) -> str:
    """조건 마스터에서 왜 조건을 찾지 못했는지 작업자용 문구로 설명합니다."""
    records = load_condition_master()
    step = lot.get("step", "").strip()
    round_no = lot.get("round", "").strip()
    manage_no = lot.get("manage_no", "").strip()
    process_code = lot.get("process_code", "").strip()

    if not records:
        return "조건 마스터에 저장된 조건이 없습니다."

    candidates = [
        record
        for record in records
        if record.get("step", "") == step
        and record.get("round", "") == round_no
        and (not manage_no or record.get("manage_no", "") == manage_no)
    ]
    if candidates:
        lines = ["가장 가까운 조건은 찾았지만 아래 항목이 다릅니다."]
        for record in candidates[:5]:
            differences = []
            for field, label, current_value in (
                ("process_code", "공정코드", process_code),
                ("manage_no", "관리번호", manage_no),
                ("step", "STEP", step),
                ("round", "차수", round_no),
            ):
                saved_value = str(record.get(field, "")).strip()
                if current_value and saved_value != current_value:
                    differences.append(f"{label}: 입력 [{current_value}] / 마스터 [{saved_value or '빈칸'}]")
            if differences:
                lines.append("- " + " / ".join(differences))
        return "\n".join(lines)

    step_candidates = [record for record in records if record.get("step", "") == step]
    if step_candidates:
        examples = []
        for record in step_candidates[:5]:
            examples.append(
                f"- 차수 [{record.get('round', '')}], 관리번호 [{record.get('manage_no', '')}], 공정코드 [{record.get('process_code', '')}]"
            )
        return "STEP은 같지만 차수/관리번호/공정코드가 일치하는 조건이 없습니다.\n" + "\n".join(examples)

    return "STEP부터 일치하는 조건이 조건 마스터에 없습니다."


def lookup_condition_jig_from_history(config: dict, lot: dict) -> tuple[str, str, str]:
    """기존 작업일보 이력에서 작업조건(K열)과 지그(N열)를 찾아옵니다.

    작업조건/지그는 작업자가 직접 입력하는 값이 아니라 기존 진행 이력을 참고해야 하므로,
    KCC PKG 시트 8행 이후를 아래 우선순위로 뒤에서부터 검색합니다.

    1순위: 조건 마스터에서 STEP + 차수 + 공정코드 일치
    2순위: 작업일보에서 STEP + 차수 + 관리번호 일치

    작업일보에는 공정코드가 저장되지 않으므로, 공정코드가 입력된 경우에는
    조건 마스터를 우선 사용하고 작업일보 이력은 보조로만 사용합니다.
    """
    # The work log does not store process_code, so using it as a fallback can
    # bring the wrong condition when STEP/round are similar. Use only the local
    # master, whose key includes STEP + round + process_code.
    if not (
        lot.get("step", "").strip()
        and lot.get("round", "").strip()
        and lot.get("process_code", "").strip()
    ):
        return "", "", ""

    return lookup_condition_jig_from_master(lot)


def write_common_lot_row(ws, row: int, common: dict, lot: dict, stack: str, model_change: str, frequent_check: list[str] | None = None) -> None:
    """일반 DNC 작업일보 한 줄을 기록합니다."""
    qty = int(lot["qty"])
    result = round(qty * 0.2, 1)
    values = [
        normalize_machine_name(common.get("machine", "")),
        common["work_date"],
        common["shift_group"],
        common["shift"],
        common["worker"],
        lot["step"],
        lot["round"],
        lot["manage_no"],
        lot["lot_no"],
        qty,
        result,
        lot["condition"],
        "",
        stack,
        lot["jig"],
        model_change,
        "",
    ]
    for col, value in enumerate(values, start=1):
        # L열 작업 P/G는 원문 그대로, 나머지는 작업일보 가독성을 위해 대문자로 저장합니다.
        ws.cell(row=row, column=col).value = value if col == 12 else excel_upper_value(value)
    write_process_code_backup(ws, row, lot.get("process_code", ""))
    if frequent_check:
        for offset, value in enumerate(frequent_check, start=18):
            ws.cell(row=row, column=offset).value = value


def save_normal_dnc_log(config: dict, common: dict, lots: list[dict], stack: str, model_change: bool, frequent_check: list[str] | None = None) -> tuple[Path, list[int]]:
    """일반 DNC 내용을 작업일보에 저장하고 저장된 행 번호를 반환합니다."""
    workbook, ws, path = open_log_workbook(config)
    rows = []
    try:
        start_row = get_next_empty_row(ws)
        for index, lot in enumerate(lots):
            row = start_row + index
            rows.append(row)
            row_frequent_check = None
            if frequent_check:
                # 초품 4Point(Q:V)는 모든 LOT 저장행에 기록합니다.
                # 지그교체 하부핀(W:AB)은 기종교체 표시가 들어가는 첫 저장행에만 기록합니다.
                row_frequent_check = frequent_check[:6] + (frequent_check[6:] if model_change and index == 0 else [""] * 6)
            write_common_lot_row(
                ws,
                row,
                common,
                lot,
                stack,
                "기종교체" if model_change and index == 0 else "",
                row_frequent_check,
            )
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()
    return path, rows


def update_normal_burr_result(config: dict, rows: list[int], burr_ok: bool) -> None:
    """DNC 완료 후 Burr 결과와 기록시간을 일반 DNC 저장행에 반영합니다."""
    workbook, ws, path = open_log_workbook(config)
    try:
        result = "이상 없음" if burr_ok else "Burr 발생"
        now_text = datetime.now().strftime("%H:%M:%S")
        for row in rows:
            ws.cell(row=row, column=13).value = result
            ws.cell(row=row, column=17).value = now_text
            if not burr_ok:
                ws.cell(row=row, column=13).font = Font(color="FF0000", bold=False)
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()


def update_normal_frequent_check_result(config: dict, rows: list[int], model_change: bool, frequent_check: list[str]) -> None:
    """DNC 완료 후 확인한 초품/하부 Pin 결과를 저장된 작업일보 행에 반영합니다."""
    workbook, ws, path = open_log_workbook(config)
    try:
        for index, row in enumerate(rows):
            # 초품 4Point(Q:V)는 모든 LOT 저장행에 기록합니다.
            for offset, value in enumerate(frequent_check[:6], start=18):
                ws.cell(row=row, column=offset).value = value
            # 하부 Pin 3개 확인(W:AB)은 기종교체 표시가 들어가는 첫 저장행에만 기록합니다.
            for offset, value in enumerate(frequent_check[6:] if model_change and index == 0 else [""] * 6, start=24):
                ws.cell(row=row, column=offset).value = value
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()


def save_new_model_log(config: dict, common: dict, lot: dict, leader_name: str) -> tuple[Path, int]:
    """신규 모델 검증 DNC 내용을 작업일보에 저장합니다."""
    workbook, ws, path = open_log_workbook(config)
    try:
        row = get_next_empty_row(ws)
        qty_text = lot.get("qty", "").strip()
        qty_number = int(qty_text) if qty_text else 0
        qty_value = "더미" if qty_number == 0 else qty_number
        result_value = "" if qty_number == 0 else round(qty_number * 0.2, 1)
        values = [
            normalize_machine_name(common.get("machine", "")),
            common["work_date"],
            common["shift_group"],
            common["shift"],
            leader_name,
            lot["step"],
            lot["round"],
            lot["manage_no"],
            lot["lot_no"],
            qty_value,
            result_value,
            lot["condition"],
            "",
            "",
            lot["jig"],
            "신규 검증",
            "",
        ]
        for col, value in enumerate(values, start=1):
            # L열 작업 P/G는 원문 그대로, 나머지는 작업일보 가독성을 위해 대문자로 저장합니다.
            ws.cell(row=row, column=col).value = value if col == 12 else excel_upper_value(value)
        write_process_code_backup(ws, row, lot.get("process_code", ""))
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()
    return path, row


def update_new_model_result(config: dict, row: int, condition_name: str, first_article_ok: bool) -> None:
    """신규 모델 DNC 완료 후 초도품 확인 결과와 기록시간을 저장합니다."""
    workbook, ws, path = open_log_workbook(config)
    try:
        k_cell = ws.cell(row=row, column=12)
        if first_article_ok:
            k_cell.value = condition_name
        else:
            k_cell.value = f"[검증 NG 발생] {condition_name}"
            k_cell.font = Font(name="맑은 고딕", size=10, color="FF0000", bold=False)
        ws.cell(row=row, column=17).value = datetime.now().strftime("%H:%M:%S")
        save_workbook_safely(workbook, path)
    finally:
        workbook.close()


# ==================================================
# GUI
# ==================================================
class LabeledEntry(ttk.Frame):
    """라벨과 입력칸을 한 줄로 만드는 작은 공용 위젯입니다."""

    def __init__(
        self,
        parent,
        label: str,
        width: int = 18,
        style: str = "Wide.TEntry",
        readonly: bool = False,
        on_change=None,
        uppercase: bool = False,
        numeric_only: bool = False,
        live_change: bool = True,
    ):
        super().__init__(parent)
        self.var = tk.StringVar()
        self.on_change = on_change
        self.uppercase = uppercase
        self.numeric_only = numeric_only
        self._normalizing = False
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        self.entry = ttk.Entry(
            self,
            textvariable=self.var,
            width=width,
            style=style,
            state="readonly" if readonly else "normal",
        )
        self.entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        if not readonly and (uppercase or numeric_only):
            self.var.trace_add("write", self.normalize_value)
        if on_change and not readonly:
            if live_change:
                self.entry.bind("<KeyRelease>", lambda _event: on_change())
            self.entry.bind("<Return>", lambda _event: on_change())
            self.entry.bind("<FocusOut>", lambda _event: on_change())

    def set_readonly(self, readonly: bool = True) -> None:
        self.entry.configure(state="readonly" if readonly else "normal")

    def normalize_value(self, *_args) -> None:
        if self._normalizing:
            return
        value = self.var.get()
        normalized = value
        if self.numeric_only:
            normalized = "".join(ch for ch in normalized if ch.isdigit())
        if self.uppercase:
            normalized = normalized.upper()
        if normalized != value:
            self._normalizing = True
            self.var.set(normalized)
            self._normalizing = False

    def get(self) -> str:
        return self.var.get().strip()

    def set(self, value: str) -> None:
        text = str(value or "")
        if self.numeric_only:
            text = "".join(ch for ch in text if ch.isdigit())
        if self.uppercase:
            text = text.upper()
        self.var.set(text)

    def clear(self) -> None:
        self.var.set("")


class DateField(ttk.Frame):
    """날짜 형식을 통일하기 위한 선택식 날짜 입력 위젯입니다."""

    def __init__(self, parent, label: str, on_change=None):
        super().__init__(parent)
        self.var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.on_change = on_change
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        self.entry = ttk.Entry(self, textvariable=self.var, width=14, style="Wide.TEntry", state="readonly")
        self.entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(self, text="선택", width=6, command=self.open_picker).pack(side=tk.LEFT, padx=(6, 0))

    def get(self) -> str:
        return self.var.get().strip()

    def set(self, value: str) -> None:
        self.var.set(value)

    def clear(self) -> None:
        self.var.set(datetime.now().strftime("%Y-%m-%d"))

    def open_picker(self) -> None:
        picker = tk.Toplevel(self)
        picker.title("작업일자 선택")
        picker.configure(bg=APP_BG)
        picker.resizable(False, False)

        today = datetime.now()
        try:
            selected = datetime.strptime(self.var.get(), "%Y-%m-%d")
        except ValueError:
            selected = today

        year_var = tk.IntVar(value=selected.year)
        month_var = tk.IntVar(value=selected.month)
        day_var = tk.IntVar(value=selected.day)

        top = ttk.Frame(picker, padding=(12, 12, 12, 6))
        top.pack(fill=tk.X)
        ttk.Spinbox(top, from_=today.year - 5, to=today.year + 5, textvariable=year_var, width=8).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Spinbox(top, from_=1, to=12, textvariable=month_var, width=5).pack(side=tk.LEFT, padx=(0, 6))

        days_frame = ttk.Frame(picker, padding=(12, 6, 12, 12))
        days_frame.pack()

        def refresh_days() -> None:
            for child in days_frame.winfo_children():
                child.destroy()
            last_day = calendar.monthrange(year_var.get(), month_var.get())[1]
            for day in range(1, last_day + 1):
                button = ttk.Button(
                    days_frame,
                    text=str(day),
                    width=4,
                    command=lambda value=day: select_day(value),
                )
                button.grid(row=(day - 1) // 7, column=(day - 1) % 7, padx=2, pady=2)

        def select_day(day: int) -> None:
            day_var.set(day)
            self.var.set(f"{year_var.get():04d}-{month_var.get():02d}-{day_var.get():02d}")
            if self.on_change:
                self.on_change()
            picker.destroy()

        ttk.Button(top, text="변경", command=refresh_days).pack(side=tk.LEFT)
        refresh_days()


class SegmentedField(ttk.Frame):
    """조/근무처럼 정해진 값만 선택하게 하는 버튼형 입력 위젯입니다."""

    def __init__(self, parent, label: str, options: list[str], initial: str | None = None, button_width: int | None = None, allow_empty: bool = False, on_change=None):
        super().__init__(parent)
        self.var = tk.StringVar(value=initial if initial in options else ("" if allow_empty else (options[0] if options else "")))
        self.on_change = on_change
        self.buttons: list[tk.Button] = []
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        wrap = tk.Frame(self, bg=APP_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        wrap.pack(side=tk.LEFT, fill=tk.X, expand=True)
        for option in options:
            button = tk.Button(
                wrap,
                text=option,
                command=lambda value=option: self.set_user(value),
                relief=tk.FLAT,
                bd=0,
                width=button_width or 0,
                padx=18,
                pady=6,
                cursor="hand2",
                font=("맑은 고딕", 10),
            )
            button.pack(side=tk.LEFT, fill=tk.X, expand=True)
            self.buttons.append(button)
        self.update_buttons()

    def get(self) -> str:
        return self.var.get().strip()

    def set(self, value: str) -> None:
        self.var.set(value)
        self.update_buttons()

    def set_user(self, value: str) -> None:
        self.set(value)
        if self.on_change:
            self.on_change()

    def clear(self) -> None:
        self.var.set("")
        self.update_buttons()

    def update_buttons(self) -> None:
        for button in self.buttons:
            selected = button.cget("text") == self.var.get()
            button.configure(bg=PRIMARY_LIGHT if selected else SURFACE_BG, fg=PRIMARY if selected else TEXT_COLOR)


class ComboField(ttk.Frame):
    """거의 고정으로 쓰는 값은 실수 클릭을 줄이기 위해 드롭다운으로 선택합니다."""

    def __init__(self, parent, label: str, options: list[str], initial: str | None = None, width: int = 12):
        super().__init__(parent)
        self.var = tk.StringVar(value=initial if initial in options else (options[0] if options else ""))
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        self.combo = ttk.Combobox(
            self,
            textvariable=self.var,
            values=options,
            state="readonly",
            width=width,
            style="White.TCombobox",
            font=("맑은 고딕", 10),
        )
        self.combo.pack(side=tk.LEFT, fill=tk.X, expand=True)

    def get(self) -> str:
        return self.var.get().strip()

    def set(self, value: str) -> None:
        self.var.set(value)

    def clear(self) -> None:
        pass


class RoundField(ttk.Frame):
    """차수를 1차~8차 버튼으로 선택하게 하는 작은 입력 위젯입니다."""

    OPTIONS = [f"{number}차" for number in range(1, 9)]

    def __init__(self, parent, label: str):
        super().__init__(parent)
        self.var = tk.StringVar()
        self.buttons: list[tk.Button] = []
        self.readonly = False
        ttk.Label(self, text=label, width=9, anchor="e").pack(side=tk.LEFT, padx=(0, 6))
        wrap = tk.Frame(self, bg=APP_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        wrap.pack(side=tk.LEFT, fill=tk.X, expand=True)
        for option in self.OPTIONS:
            button = tk.Button(
                wrap,
                text=option,
                command=lambda value=option: self.toggle(value),
                relief=tk.FLAT,
                bd=0,
                width=3,
                padx=1,
                pady=6,
                cursor="hand2",
                font=("맑은 고딕", 9),
            )
            button.pack(side=tk.LEFT, fill=tk.X, expand=True)
            self.buttons.append(button)
        self.update_buttons()

    def get(self) -> str:
        return self.var.get().strip()

    def toggle(self, value: str) -> None:
        """작업자가 같은 차수 버튼을 한 번 더 누르면 선택을 취소합니다."""
        if self.readonly:
            return
        self.var.set("" if self.var.get() == value else value)
        self.update_buttons()

    def set(self, value: str) -> None:
        self.var.set(value)
        self.update_buttons()

    def clear(self) -> None:
        self.var.set("")
        self.update_buttons()

    def set_readonly(self, readonly: bool = True) -> None:
        self.readonly = readonly
        self.update_buttons()

    def update_buttons(self) -> None:
        for button in self.buttons:
            selected = button.cget("text") == self.var.get()
            button.configure(
                bg=PRIMARY_LIGHT if selected else SURFACE_BG,
                fg=PRIMARY if selected else TEXT_COLOR,
                cursor="arrow" if self.readonly else "hand2",
            )


class SimpleTabNotebook(ttk.Frame):
    """Tkinter 기본 Notebook 대신 선택 테두리가 선명한 간단 탭 UI입니다."""

    def __init__(self, parent):
        super().__init__(parent)
        self.selected_index = -1
        self.pages: list[tk.Frame] = []
        self.labels: list[tk.Label] = []
        self.tab_bar = tk.Frame(self, bg=APP_BG)
        self.tab_bar.pack(fill=tk.X)
        self.page_area = tk.Frame(self, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        self.page_area.pack(fill=tk.BOTH, expand=True)

    def add(self, page: tk.Frame, text: str) -> None:
        index = len(self.pages)
        self.pages.append(page)
        label = tk.Label(
            self.tab_bar,
            text=text,
            bg=TAB_BG,
            fg=MUTED_TEXT,
            width=12,
            padx=0,
            pady=12,
            font=("맑은 고딕", 10),
            cursor="hand2",
            highlightthickness=1,
            highlightbackground="#93c5fd",
            bd=0,
        )
        label.pack(side=tk.LEFT)
        label.bind("<Button-1>", lambda _event, i=index: self.select(i))
        label.bind("<Enter>", lambda _event, i=index: self.hover(i))
        label.bind("<Leave>", lambda _event, i=index: self.update_label(i))
        self.labels.append(label)
        page.pack_forget()
        if index == 0:
            self.select(0)
        else:
            self.update_label(index)

    def select(self, index: int) -> None:
        for page in self.pages:
            page.pack_forget()
        self.selected_index = index
        self.pages[index].pack(fill=tk.BOTH, expand=True)
        for i in range(len(self.labels)):
            self.update_label(i)

    def hover(self, index: int) -> None:
        if index != self.selected_index:
            self.labels[index].configure(bg="#f7faff", fg=TEXT_COLOR)

    def update_label(self, index: int) -> None:
        if index == self.selected_index:
            self.labels[index].configure(bg=SURFACE_BG, fg=PRIMARY, highlightbackground=PRIMARY)
            self.page_area.configure(highlightbackground=PRIMARY)
        else:
            self.labels[index].configure(bg=TAB_BG, fg=MUTED_TEXT, highlightbackground=BORDER_COLOR)


def validate_frequent_check_values(values: list[str], check_mode: str = "first") -> tuple[bool, str]:
    """초품 4Point 또는 하부 Pin 3개 확인 값을 검증합니다."""
    if len(values) != 12:
        return False, "확인 데이터가 올바르지 않습니다."
    first_count = sum(1 for value in values[:6] if value == "OK")
    jig_count = sum(1 for value in values[6:] if value == "OK")
    if check_mode == "jig":
        if jig_count == 0:
            return False, "하부 Pin 축 선택 필요"
        return True, f"하부 Pin {jig_count}축 확인"
    if check_mode == "first" and first_count == 0:
        return False, "초품 축 선택 필요"
    return True, f"초품 {first_count}축 확인"


def count_frequent_check_axes(values: list[str]) -> int:
    """자주검사에서 실제 사용하는 축 수를 반환합니다.

    좌/우 그룹은 저장 전에 같은 개수로 검증되므로, 앞쪽 초품 확인 축 수를 기준으로 계산합니다.
    """
    return sum(1 for value in values[:6] if value == "OK")


def validate_frequent_check_capacity(lots: list[dict], stack: str, values: list[str]) -> tuple[bool, str]:
    """선택한 축 수 x Stack 수가 LOT 총 매수보다 작을 때만 차단합니다.

    Stack은 축당 최대 투입 가능 매수입니다. 예를 들어 10Stack에서 12매는
    2축으로 10매 + 2매 투입이 가능하므로 OK입니다.
    """
    axis_count = count_frequent_check_axes(values)
    stack_count = int(stack)
    total_qty = sum(int(lot.get("qty", "0")) for lot in lots)
    max_qty = axis_count * stack_count

    if total_qty > max_qty:
        shortage = total_qty - max_qty
        return (
            False,
            f"LOT 총 매수: {total_qty}매\n"
            f"확인 가능: {max_qty}매\n"
            f"부족 수량: {shortage}매\n\n"
            "초품 축 수 / Stack 수 확인 필요",
        )
    spare_qty = max_qty - total_qty
    return True, f"초품 확인 OK: LOT {total_qty}매 / {axis_count}축 x {stack_count}Stack = 최대 {max_qty}매 / 여유 {spare_qty}매"


class JiinDncManager:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1420x820")
        self.root.minsize(1180, 700)
        self.config = load_config()
        apply_theme(self.config.get("theme", "MES 블루"))
        self.root.configure(bg=APP_BG)
        try:
            sync_condition_master_from_completed_logs()
        except Exception:
            # 조건 마스터 동기화가 실패해도 현장 작업 화면은 열리게 둡니다.
            pass
        self.is_running = False
        self.is_exporting_excel = False

        self.common_entries: dict[str, LabeledEntry] = {}
        self.lot1_entries: dict[str, LabeledEntry] = {}
        self.lot2_entries: dict[str, LabeledEntry] = {}
        self.tlb_common_entries: dict[str, LabeledEntry] = {}
        self.tlb_entries: dict[str, LabeledEntry] = {}
        self.tlb_lot2_entries: dict[str, LabeledEntry] = {}
        self.tlb_condition_records: dict[int, dict] = {}
        self.tlb_cycle_state: dict | None = None
        self.tlb_status_labels: dict[str, tk.Label] = {}
        self.tlb_log_text: scrolledtext.ScrolledText | None = None
        self.tlb_preview_canvas: tk.Canvas | None = None
        self.tlb_run_button: tk.Canvas | None = None
        self.kcc_hdi_common_entries: dict[str, LabeledEntry] = {}
        self.kcc_hdi_entries: dict[str, LabeledEntry] = {}
        self.kcc_hdi_lot2_entries: dict[str, LabeledEntry] = {}
        self.kcc_hdi_condition_records: dict[int, dict] = {}
        self.kcc_hdi_cycle_state: dict | None = None
        self.kcc_hdi_status_labels: dict[str, tk.Label] = {}
        self.kcc_hdi_log_text: scrolledtext.ScrolledText | None = None
        self.kcc_hdi_preview_canvas: tk.Canvas | None = None
        self.kcc_hdi_run_button: tk.Canvas | None = None
        self.kcc_run_button: tk.Canvas | None = None
        self.normal_buttons: list[ttk.Button] = []
        self.new_model_button: ttk.Button | None = None
        self.status_labels: dict[str, tk.Label] = {}
        self.lot_match_frame: tk.Frame | None = None
        self.lot_status_labels: dict[str, tk.Label] = {}
        self.log_text: scrolledtext.ScrolledText | None = None
        self.tlb_logo_image: tk.PhotoImage | None = None
        self.kcc_hdi_logo_image: tk.PhotoImage | None = None
        self.logo_image: tk.PhotoImage | None = None
        self.kcc_logo_image: tk.PhotoImage | None = None
        self.frequent_check_values: list[str] = [""] * 12
        self.work_axis_values: list[str] = [""] * 6
        self.lot_condition_keys: dict[int, str] = {1: "", 2: ""}
        self.current_work_period_key = ""
        self.last_common_manual_change_at: datetime | None = None
        self.master_settings_popup = None
        self.condition_master_popup = None

        self.setup_style()
        self.create_layout()
        self.apply_work_time_defaults(initial=True)
        self.update_status_checks()
        cleanup_old_text_logs(days=90)
        create_daily_master_backup(self.config)
        self.root.after(500, self.handle_startup_incomplete_logs)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def on_close(self) -> None:
        """작업표시줄 닫기/X 버튼은 숨은 알람창 때문에 막히지 않도록 항상 종료합니다."""
        try:
            if self.is_running:
                log_app("프로그램 종료 요청: DNC 진행 중 강제 종료")
            if self.is_exporting_excel:
                log_app("프로그램 종료 요청: 작업일보 반영 중 강제 종료")
            for child in list(self.root.winfo_children()):
                try:
                    if isinstance(child, tk.Toplevel) and child.winfo_exists():
                        child.grab_release()
                        child.destroy()
                except tk.TclError:
                    pass
            self.root.quit()
            self.root.destroy()
        except Exception as exc:
            log_error("프로그램 종료 처리 실패", exc)
            os._exit(0)

    def handle_startup_incomplete_logs(self) -> None:
        """앱 시작 시 미완료 이력이 있으면 현장에서 바로 처리할 수 있게 안내합니다."""
        try:
            rows = fetch_incomplete_kcc_pkg_logs()
        except Exception as exc:
            log_error("미완료 이력 조회 실패", exc)
            return
        if not rows:
            return
        log_ids = [int(row["id"]) for row in rows]
        action = ask_incomplete_action(self.root, len(log_ids))
        try:
            if action == "delete":
                delete_incomplete_kcc_pkg_logs(log_ids)
                self.set_status("excel", "미완료 이력 삭제 완료", True)
            elif action == "complete":
                complete_incomplete_kcc_pkg_logs(log_ids)
                self.set_status("excel", "미완료 이력 완료처리", True)
        except Exception as exc:
            log_error("미완료 이력 처리 실패", exc)
            show_operator_alert(self.root, "미완료 처리 실패", "처리 실패", "error")

    def setup_style(self) -> None:
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TFrame", background=APP_BG)
        style.configure("Header.TFrame", background=SURFACE_BG)
        style.configure("Panel.TFrame", background=SURFACE_BG)
        style.configure("TLabel", background=APP_BG, foreground=TEXT_COLOR, font=("맑은 고딕", 10))
        style.configure("Header.TLabel", background=SURFACE_BG, foreground=TEXT_COLOR, font=("맑은 고딕", 18, "bold"))
        style.configure("Hint.TLabel", background=APP_BG, foreground=MUTED_TEXT, font=("맑은 고딕", 10))
        style.configure("Panel.TLabel", background=SURFACE_BG, foreground=TEXT_COLOR, font=("맑은 고딕", 11, "bold"))
        style.configure("TButton", font=("맑은 고딕", 10), padding=(12, 8), background=SURFACE_BG, foreground=TEXT_COLOR)
        style.map("TButton", background=[("active", PRIMARY_LIGHT), ("pressed", "#dbeafe")])
        style.configure("Primary.TButton", font=("맑은 고딕", 11, "bold"), padding=(16, 10), background=PRIMARY_LIGHT, foreground=PRIMARY, relief="solid", borderwidth=1)
        style.map("Primary.TButton", background=[("active", "#dbeafe"), ("pressed", "#bfdbfe")], foreground=[("active", PRIMARY)])
        style.configure("Side.TButton", font=("맑은 고딕", 10), padding=(12, 8), background=SURFACE_BG, foreground=TEXT_COLOR)
        style.map("Side.TButton", background=[("active", "#f1f5f9"), ("pressed", "#e2e8f0")], foreground=[("active", TEXT_COLOR)])
        style.configure("SidePrimary.TButton", font=("맑은 고딕", 10, "bold"), padding=(12, 8), background=PRIMARY_LIGHT, foreground=PRIMARY, relief="solid", borderwidth=1)
        style.map("SidePrimary.TButton", background=[("active", "#dbeafe"), ("pressed", "#bfdbfe")], foreground=[("active", PRIMARY)])
        style.configure("SideDanger.TButton", font=("맑은 고딕", 10, "bold"), padding=(12, 8), background="#fee2e2", foreground=NG_COLOR)
        style.map("SideDanger.TButton", background=[("active", "#fecaca")], foreground=[("active", NG_COLOR)])
        style.configure("Wide.TEntry", padding=(8, 5), fieldbackground=SURFACE_BG)
        style.configure("White.TCombobox", padding=(8, 5), fieldbackground=SURFACE_BG, background=SURFACE_BG, foreground=TEXT_COLOR)
        style.map(
            "White.TCombobox",
            fieldbackground=[("readonly", SURFACE_BG), ("!disabled", SURFACE_BG)],
            background=[("readonly", SURFACE_BG), ("!disabled", SURFACE_BG)],
            foreground=[("readonly", TEXT_COLOR), ("!disabled", TEXT_COLOR)],
        )
        style.configure("Lookup.TEntry", padding=(8, 5), fieldbackground="#eef4fb", foreground=PRIMARY)
        style.map(
            "Lookup.TEntry",
            fieldbackground=[("readonly", "#eef4fb")],
            foreground=[("readonly", PRIMARY)],
        )
        style.configure("TCheckbutton", background=APP_BG, foreground=TEXT_COLOR, font=("맑은 고딕", 10))

    def create_layout(self) -> None:
        header = ttk.Frame(self.root, style="Header.TFrame", padding=(20, 12, 20, 12))
        header.pack(fill=tk.X, padx=12, pady=(12, 0))
        ttk.Label(header, text=APP_TITLE, style="Header.TLabel").pack(side=tk.LEFT)
        logo_path = BUNDLED_LOGO_FILE if BUNDLED_LOGO_FILE.exists() else LOGO_FILE
        if logo_path.exists():
            try:
                logo = tk.PhotoImage(file=str(logo_path))
                scale = max(1, logo.height() // 54)
                self.logo_image = logo.subsample(scale, scale)
                tk.Label(header, image=self.logo_image, bg=SURFACE_BG, bd=0).pack(side=tk.RIGHT)
            except tk.TclError:
                self.logo_image = None

        self.notebook = SimpleTabNotebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=12, pady=(10, 4))

        footer = tk.Frame(self.root, bg=APP_BG)
        footer.pack(fill=tk.X, padx=16, pady=(0, 6))
        tk.Label(
            footer,
            text=APP_VERSION_TEXT,
            bg=APP_BG,
            fg=MUTED_TEXT,
            font=("맑은 고딕", 8),
        ).pack(side=tk.RIGHT)

        self.tlb_page = tk.Frame(self.notebook.page_area, bg=APP_BG)
        self.notebook.add(self.tlb_page, "TLB")
        for tab_name in ["심텍 SPS", "심텍 HDI"]:
            self.notebook.add(self.create_placeholder_tab(tab_name), tab_name)
        self.kcc_pkg_page = tk.Frame(self.notebook.page_area, bg=APP_BG)
        self.notebook.add(self.kcc_pkg_page, "KCC PKG")
        self.kcc_hdi_page = tk.Frame(self.notebook.page_area, bg=APP_BG)
        self.notebook.add(self.kcc_hdi_page, "KCC HDI")
        self.settings_page = tk.Frame(self.notebook.page_area, bg=APP_BG)
        self.notebook.add(self.settings_page, "설정")

        self.create_tlb_tab()
        self.create_kcc_hdi_tab()
        self.create_kcc_pkg_tab()
        self.create_settings_tab()
        self.notebook.select(0)

    def create_placeholder_tab(self, name: str) -> tk.Frame:
        process_theme = PROCESS_COLORS.get(name, PROCESS_COLORS["KCC PKG"])
        page = tk.Frame(self.notebook.page_area, bg=process_theme["bg"])
        tk.Label(page, text=f"{name}\n추후 개발 예정", bg=process_theme["bg"], fg=process_theme["primary"], font=("맑은 고딕", 22, "bold")).pack(expand=True)
        return page

    def create_tlb_tab(self) -> None:
        font_name = "맑은 고딕"
        tlb_theme = PROCESS_COLORS["TLB"]
        tlb_bg = tlb_theme["bg"]
        tlb_light = tlb_theme["light"]
        tlb_primary = tlb_theme["primary"]
        tlb_border = tlb_theme["border"]
        self.tlb_page.configure(bg=tlb_bg)
        self.tlb_page.columnconfigure(0, weight=1)
        self.tlb_page.rowconfigure(2, weight=0)
        self.tlb_page.rowconfigure(3, weight=0)
        title_wrap = tk.Frame(self.tlb_page, bg=tlb_light)
        title_wrap.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 8))
        title_wrap.columnconfigure(0, minsize=260)
        title_wrap.columnconfigure(1, weight=1)
        title_wrap.columnconfigure(2, minsize=360)
        logo_slot = tk.Frame(title_wrap, bg=tlb_light, width=220, height=1)
        logo_slot.grid(row=0, column=0, sticky="w", padx=(14, 8))
        logo_slot.grid_propagate(False)
        tlb_logo_path = BUNDLED_TLB_LOGO_FILE if BUNDLED_TLB_LOGO_FILE.exists() else TLB_LOGO_FILE
        if tlb_logo_path.exists():
            try:
                tlb_logo = tk.PhotoImage(file=str(tlb_logo_path))
                scale = max(1, tlb_logo.height() // 32)
                self.tlb_logo_image = tlb_logo.subsample(scale, scale)
                tk.Label(logo_slot, image=self.tlb_logo_image, bg=tlb_light, bd=0).pack(side=tk.LEFT, anchor="w")
            except tk.TclError:
                self.tlb_logo_image = None
        tk.Label(title_wrap, text="TLB HDI DNC", bg=tlb_light, fg=tlb_primary, font=(font_name, 14, "bold"), height=2).grid(row=0, column=1, sticky="ew")
        title_buttons = tk.Frame(title_wrap, bg=tlb_light)
        title_buttons.grid(row=0, column=2, sticky="e", padx=(8, 10))
        self.tlb_run_button = self.create_tlb_gradient_run_button(title_buttons)
        self.tlb_run_button.grid(row=0, column=0, padx=4, pady=4)
        self.add_normal_button(title_buttons, "입력 초기화", self.clear_tlb_inputs).grid(row=0, column=1, padx=4, pady=4)

        common = self.create_panel(self.tlb_page, "공통 입력", tlb_theme)
        common.configure(highlightbackground=tlb_border)
        common.winfo_children()[0].configure(bg=tlb_light, fg=tlb_primary)
        common.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 8))
        common_widgets = [
            ("machine", ComboField(common, "설비 호기", ["트리밍 1호기", "트리밍 2호기", "트리밍 3호기"], initial=self.config.get("machine", "트리밍 1호기"), width=12)),
            ("work_date", DateField(common, "작업일자", on_change=lambda key="work_date": self.handle_common_change(key, "tlb"))),
            ("shift_group", SegmentedField(common, "조", ["A", "B", "C"], allow_empty=True, on_change=lambda key="shift_group": self.handle_common_change(key, "tlb"))),
            ("shift", SegmentedField(common, "근무", ["주간", "야간"], on_change=lambda key="shift": self.handle_common_change(key, "tlb"))),
            ("worker", LabeledEntry(common, "작업자", width=12, on_change=lambda key="worker": self.handle_common_change(key, "tlb"), live_change=False)),
        ]
        for index, (key, entry) in enumerate(common_widgets):
            entry.grid(row=1, column=index, sticky="ew", padx=8, pady=8)
            self.tlb_common_entries[key] = entry
            if key == "machine":
                entry.combo.bind("<<ComboboxSelected>>", lambda _event, field=key: self.handle_common_change(field, "tlb"))
        common.columnconfigure(0, weight=0, minsize=320)
        common.columnconfigure(1, weight=1, minsize=430)
        common.columnconfigure(2, weight=1, minsize=330)
        common.columnconfigure(3, weight=1, minsize=360)
        common.columnconfigure(4, weight=0, minsize=320)

        body = tk.Frame(self.tlb_page, bg=tlb_bg)
        body.grid(row=2, column=0, sticky="nsew", padx=14, pady=0)
        body.columnconfigure(0, weight=1, uniform="tlb_body")
        body.columnconfigure(1, weight=1, uniform="tlb_body")
        fields = [
            ("manage_no", "Tool No"),
            ("round", "차수"),
            ("lot_no", "LOT No"),
            ("qty", "매수"),
            ("condition", "조건(조회)"),
            ("jig", "지그(조회)"),
        ]

        def build_tlb_lot_panel(parent: tk.Frame, title: str, entries: dict, status_key: str, column: int) -> None:
            panel = self.create_panel(parent, title, tlb_theme)
            panel.configure(highlightbackground=tlb_border)
            panel.winfo_children()[0].configure(bg=tlb_light, fg=tlb_primary)
            panel.grid(row=0, column=column, sticky="nsew", padx=(0, 8) if column == 0 else (8, 0))
            lot_no = 1 if column == 0 else 2
            for index, (key, label) in enumerate(fields):
                row = index // 2 + 1
                col = index % 2
                if key == "round":
                    entry = RoundField(panel, label)
                elif key in {"condition", "jig"}:
                    entry = LabeledEntry(panel, label, width=32, style="Lookup.TEntry", readonly=True)
                elif key == "qty":
                    entry = LabeledEntry(panel, label, width=32, numeric_only=True)
                else:
                    entry = LabeledEntry(panel, label, width=32, uppercase=True)
                entry.grid(row=row, column=col, sticky="ew", padx=14, pady=9)
                panel.columnconfigure(col, weight=1)
                entries[key] = entry
                if key in {"manage_no", "round", "lot_no", "qty"}:
                    entry.var.trace_add("write", lambda *_args, lot_no=lot_no: self.handle_tlb_input_changed(lot_no))
            self.create_mes_lookup_button(
                panel,
                command=lambda lot_no=lot_no: self.load_tlb_condition_jig(lot_no),
                scheme="tlb",
                ready_check=lambda lot_no=lot_no: self.is_tlb_lookup_ready(lot_no),
            ).grid(row=4, column=0, columnspan=2, sticky="ew", padx=14, pady=(12, 8))
            status = tk.Frame(panel, bg=SURFACE_BG)
            status.grid(row=5, column=0, columnspan=2, sticky="ew", padx=14, pady=(0, 12))
            status.columnconfigure(0, weight=1, uniform="tlb_lot_status")
            status.columnconfigure(1, weight=1, uniform="tlb_lot_status")
            self.tlb_status_labels[status_key] = self.create_judgement_card(status, "조건 조회")
            self.tlb_status_labels[status_key].grid(row=0, column=0, sticky="ew", padx=(0, 6))
            self.hide_judgement_card(self.tlb_status_labels[status_key])
            if column == 0:
                self.tlb_status_labels["cycle"] = self.create_judgement_card(status, "완료 사이클")
                self.tlb_status_labels["cycle"].grid(row=0, column=1, sticky="ew", padx=(6, 0))
                self.hide_judgement_card(self.tlb_status_labels["cycle"])

        build_tlb_lot_panel(body, "LOT 1 입력", self.tlb_entries, "condition1", 0)
        build_tlb_lot_panel(body, "LOT 2 입력 (선택)", self.tlb_lot2_entries, "condition2", 1)
        self.tlb_status_labels["condition"] = self.tlb_status_labels["condition1"]

        bottom = tk.Frame(self.tlb_page, bg=tlb_bg)
        bottom.grid(row=3, column=0, sticky="ew", padx=14, pady=(8, 14))
        bottom.columnconfigure(0, weight=1)
        status_panel = tk.Frame(bottom, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        status_panel.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        status_panel.columnconfigure(1, weight=1)
        tk.Label(status_panel, text="2LOT 조건 일치 확인", bg=tlb_light, fg=tlb_primary, font=(font_name, 11, "bold"), width=22, height=2).grid(row=0, column=0, sticky="nsw")
        match_label = tk.Label(status_panel, text="LOT 2 미사용", bg=SURFACE_BG, fg=MUTED_TEXT, font=(font_name, 12, "bold"), anchor="w")
        match_label.grid(row=0, column=1, sticky="ew", padx=14)
        self.tlb_status_labels["match"] = match_label
        tk.Label(status_panel, text="DNC 진행 상태", bg=tlb_light, fg=tlb_primary, font=(font_name, 11, "bold"), width=22, height=2).grid(row=1, column=0, sticky="nsw")
        dnc_label = tk.Label(status_panel, text="대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=(font_name, 12, "bold"), anchor="w")
        dnc_label.grid(row=1, column=1, sticky="ew", padx=14)
        self.tlb_status_labels["dnc"] = dnc_label
        tk.Label(status_panel, text="작업일보 반영", bg=tlb_light, fg=tlb_primary, font=(font_name, 11, "bold"), width=22, height=2).grid(row=2, column=0, sticky="nsw")
        excel_label = tk.Label(status_panel, text="대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=(font_name, 12, "bold"), anchor="w")
        excel_label.grid(row=2, column=1, sticky="ew", padx=14)
        self.tlb_status_labels["excel"] = excel_label
        log_panel = tk.Frame(bottom, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        log_panel.grid(row=1, column=0, sticky="ew", padx=(0, 10), pady=(8, 0))
        log_panel.columnconfigure(0, weight=1)
        tk.Label(log_panel, text="TLB HDI DNC 작업 로그", bg=tlb_light, fg=tlb_primary, font=(font_name, 10, "bold"), height=1).grid(row=0, column=0, sticky="ew")
        self.tlb_log_text = scrolledtext.ScrolledText(log_panel, height=5, wrap=tk.WORD, state="disabled", bg=SURFACE_BG, fg=TEXT_COLOR, font=(font_name, 10), relief=tk.FLAT, padx=10, pady=8)
        self.tlb_log_text.grid(row=1, column=0, sticky="ew")
        button_panel = tk.Frame(bottom, bg=tlb_bg)
        button_panel.grid(row=0, column=1, rowspan=2, sticky="ne")
        button_panel.columnconfigure((0, 1), weight=1, uniform="tlb_side")
        self.add_side_button(button_panel, "조건 시트 선택", self.select_tlb_condition_sheet, "SidePrimary.TButton").grid(row=0, column=0, sticky="nsew", padx=4, pady=4)
        self.add_side_button(button_panel, "작업일보 반영", self.export_tlb_to_excel_from_ui).grid(row=1, column=0, sticky="nsew", padx=4, pady=4)
        self.add_side_button(button_panel, "작업일보 열기", self.open_log_excel_from_ui).grid(row=1, column=1, sticky="nsew", padx=4, pady=4)

    def get_tlb_common_data(self) -> dict:
        return {key: entry.get() for key, entry in self.tlb_common_entries.items()}

    def get_tlb_lot_data(self) -> dict:
        return {key: entry.get() for key, entry in self.tlb_entries.items()}

    def get_tlb_lot2_data(self) -> dict:
        return {key: entry.get() for key, entry in self.tlb_lot2_entries.items()}

    def is_tlb_lookup_ready(self, lot_no: int) -> bool:
        entries = self.tlb_entries if lot_no == 1 else self.tlb_lot2_entries
        required_keys = ("manage_no", "round", "lot_no", "qty")
        return all(entries[key].get().strip() for key in required_keys)

    def is_tlb_lot_used(self, lot: dict) -> bool:
        return any(str(lot.get(key, "")).strip() for key in ("manage_no", "round", "lot_no", "qty", "condition", "jig"))

    def handle_tlb_input_changed(self, lot_no: int) -> None:
        """TLB 입력값이 바뀌면 기존 조건 조회 결과를 즉시 무효화합니다."""
        entries = self.tlb_entries if lot_no == 1 else self.tlb_lot2_entries
        for key in ("condition", "jig"):
            if key in entries:
                entries[key].clear()
        self.tlb_condition_records.pop(lot_no, None)
        status_key = "condition1" if lot_no == 1 else "condition2"
        label = self.tlb_status_labels.get(status_key)
        if label is not None:
            self.hide_judgement_card(label)
        self.hide_tlb_cycle_status()
        self.set_tlb_run_gradient(False)
        self.update_tlb_match_status(self.get_tlb_used_lots())

    def get_tlb_used_lots(self) -> list[dict]:
        lots = [self.get_tlb_lot_data()]
        lot2 = self.get_tlb_lot2_data()
        if self.is_tlb_lot_used(lot2):
            lots.append(lot2)
        return lots

    def is_tlb_run_ready_by_condition(self) -> bool:
        lot1 = self.get_tlb_lot_data()
        if not self.is_tlb_lot_condition_confirmed(1, lot1):
            return False
        lot2 = self.get_tlb_lot2_data()
        if self.is_tlb_lot_used(lot2):
            return self.is_tlb_lot_condition_confirmed(2, lot2)
        return True

    def is_tlb_lot_condition_confirmed(self, lot_no: int, lot: dict) -> bool:
        """조건 시트 조회 버튼으로 확정된 LOT인지 확인합니다."""
        record = self.tlb_condition_records.get(lot_no)
        if not record:
            return False
        return (
            lot.get("condition", "").strip() == str(record.get("condition", "")).strip()
            and lot.get("jig", "").strip() == str(record.get("jig", "")).strip()
        )

    def ensure_tlb_condition_confirmed(self, lots: list[dict]) -> bool:
        missing_lots = [
            f"LOT {index}"
            for index, lot in enumerate(lots, start=1)
            if not self.is_tlb_lot_condition_confirmed(index, lot)
        ]
        if missing_lots:
            show_operator_alert(self.root, "DNC 조건 조회", "DNC 조건 조회 필요")
            self.set_tlb_status("dnc", "DNC 조건 조회 필요", False)
            return False
        return True

    def calculate_tlb_cycle_count(
        self,
        common: dict,
        lots: list[dict],
        stack: str,
        axis_values: list[str] | None = None,
    ) -> tuple[int, str]:
        total_qty = sum(int(lot.get("qty", "0") or 0) for lot in lots)
        axis_count = count_frequent_check_axes(axis_values) if axis_values is not None else get_machine_axis_count(common.get("machine", ""))
        stack_text = str(stack or "").strip()
        ok, message = validate_positive_number(stack_text, "Stack", required=True)
        if not ok:
            raise ValueError(message)
        stack_count = int(stack_text)
        capacity = axis_count * stack_count
        if capacity <= 0:
            raise ValueError("설비 축 수 / Stack 확인 필요")
        cycle_count = max(1, (total_qty + capacity - 1) // capacity)
        detail = f"총 {cycle_count}사이클 / {total_qty}매 / {axis_count}축 x {stack_count}Stack"
        return cycle_count, detail

    def make_tlb_cycle_signature(self, common: dict, lots: list[dict]) -> tuple:
        return (
            common.get("machine", "").strip(),
            common.get("work_date", "").strip(),
            common.get("shift_group", "").strip(),
            common.get("shift", "").strip(),
            common.get("worker", "").strip(),
            tuple(
                (
                    lot.get("manage_no", "").strip(),
                    lot.get("round", "").strip(),
                    lot.get("lot_no", "").strip(),
                    lot.get("qty", "").strip(),
                    lot.get("condition", "").strip(),
                    lot.get("jig", "").strip(),
                )
                for lot in lots
            ),
        )

    def update_tlb_cycle_status(self, cycle_count: int, remaining: int | None = None) -> None:
        label = self.tlb_status_labels.get("cycle")
        if label is None:
            return
        if cycle_count <= 1 or remaining is None or remaining <= 0:
            self.hide_tlb_cycle_status()
            return
        completed_cycle = cycle_count - remaining
        if completed_cycle <= 0:
            self.hide_tlb_cycle_status()
            return
        completed_cycle = min(cycle_count - 1, completed_cycle)
        text = f"완료 사이클\n{completed_cycle} / {cycle_count}회"
        condition_label = self.tlb_status_labels.get("condition1")
        if condition_label is not None:
            condition_label.grid_configure(row=0, column=0, columnspan=1, sticky="ew", padx=(0, 6))
        label.configure(text=text, fg="#075985", bg="#e0f2fe", highlightthickness=2, highlightbackground="#38bdf8")
        label.grid(row=0, column=1, sticky="ew", padx=(6, 0))

    def set_tlb_last_cycle_running(self) -> None:
        label = self.tlb_status_labels.get("cycle")
        if label is None:
            return
        condition_label = self.tlb_status_labels.get("condition1")
        if condition_label is not None:
            condition_label.grid_configure(row=0, column=0, columnspan=1, sticky="ew", padx=(0, 6))
        label.configure(text="마지막 Cycle\n진행 중", fg="#075985", bg="#e0f2fe", highlightthickness=2, highlightbackground="#38bdf8")
        label.grid(row=0, column=1, sticky="ew", padx=(6, 0))

    def hide_tlb_cycle_status(self) -> None:
        label = self.tlb_status_labels.get("cycle")
        if label is not None:
            self.hide_judgement_card(label)
            label.grid(row=0, column=1, sticky="ew", padx=(6, 0))
        condition_label = self.tlb_status_labels.get("condition1")
        if condition_label is not None:
            condition_label.grid_configure(row=0, column=0, columnspan=1, sticky="ew", padx=(0, 6))

    def hide_tlb_condition_cards(self) -> None:
        """TLB 조건 조회 결과 카드는 조회/실행 직후에만 보이게 초기화합니다."""
        for key in ("condition1", "condition2"):
            label = self.tlb_status_labels.get(key)
            if label is not None:
                self.hide_judgement_card(label)

    def refresh_tlb_cycle_preview(self) -> None:
        # TLB 사이클은 작업자가 DNC 실행 시 입력하는 Stack 기준으로 계산합니다.
        # 조건 조회 단계에서는 실제 Stack을 아직 모르므로 사이클을 미리 표시하지 않습니다.
        return

    def update_tlb_match_status(self, lots: list[dict]) -> None:
        label = self.tlb_status_labels.get("match")
        if label is None:
            return
        if len(lots) < 2:
            label.configure(text="LOT 2 미사용", fg=MUTED_TEXT)
            return
        lot1, lot2 = lots[0], lots[1]
        if lot1.get("condition", "").strip() != lot2.get("condition", "").strip():
            label.configure(text="NG - 조건 불일치", fg=NG_COLOR)
            return
        if lot1.get("jig", "").strip() != lot2.get("jig", "").strip():
            label.configure(text="NG - 지그 불일치", fg=NG_COLOR)
            return
        label.configure(text="OK - 2LOT 조건 일치", fg=OK_COLOR)

    def set_tlb_status(self, key: str, text: str, ok: bool | None = None) -> None:
        label = self.tlb_status_labels.get(key)
        if label is None:
            return
        fg = OK_COLOR if ok is True else NG_COLOR if ok is False else MUTED_TEXT
        label.configure(text=text, fg=fg)
        if key == "dnc":
            self.append_tlb_log(text)

    def draw_tlb_condition_preview(self, record: dict | None) -> None:
        canvas = self.tlb_preview_canvas
        if canvas is None:
            return
        canvas.delete("all")
        width = max(canvas.winfo_width(), 520)
        height = max(canvas.winfo_height(), 250)
        if not record:
            canvas.create_text(width / 2, height / 2, text="조건 시트 조회 후 표시", fill=MUTED_TEXT, font=("맑은 고딕", 12, "bold"))
            return

        line_color = "#2563eb"
        guide_color = "#93c5fd"
        value_bg = "#dff3f8"
        text_color = TEXT_COLOR
        red = "#ef4444"

        x1, y1 = width * 0.25, height * 0.22
        x2, y2 = width * 0.78, height * 0.78
        mid_y = height * 0.50
        hole_x = width * 0.245
        right_hole_x = width * 0.79

        def fmt(value: str) -> str:
            if value in ("", None):
                return ""
            try:
                return f"{float(value):.3f}"
            except Exception:
                return str(value)

        def value_box(x: float, y: float, value: str, fill: str = value_bg) -> None:
            text = fmt(value)
            if not text:
                return
            canvas.create_rectangle(x - 48, y - 14, x + 48, y + 14, fill=fill, outline="")
            canvas.create_text(x, y, text=text, fill=red if str(value).startswith("-") else text_color, font=("맑은 고딕", 10))

        canvas.create_rectangle(x1, y1, x2, y2, outline=line_color, width=2)
        canvas.create_line(x1, mid_y, x2 + 54, mid_y, fill=line_color, dash=(2, 2))
        canvas.create_line(hole_x, y1 - 26, hole_x, y2 + 26, fill=line_color, dash=(2, 2))
        canvas.create_line(right_hole_x, y1 - 26, right_hole_x, y2 + 26, fill=line_color, dash=(2, 2))
        canvas.create_line(x1 - 50, mid_y + 22, x2 + 68, mid_y + 22, fill=guide_color)
        canvas.create_line(x1 - 44, y1 - 14, x2 + 48, y1 - 14, fill=guide_color, dash=(2, 2))
        canvas.create_line(x1 - 44, y2 + 14, x2 + 48, y2 + 14, fill=guide_color, dash=(2, 2))

        for x, y in ((hole_x, mid_y), (hole_x, mid_y + 52), (right_hole_x, mid_y)):
            canvas.create_oval(x - 8, y - 8, x + 8, y + 8, outline="#64748b", width=3)
            canvas.create_oval(x - 3, y - 3, x + 3, y + 3, fill="#64748b", outline="")

        value_box((x1 + x2) / 2, y1 - 38, record.get("top", ""))
        value_box((x1 + x2) / 2, mid_y - 28, record.get("center", ""))
        value_box((x1 + x2) / 2, mid_y + 58, record.get("bottom_left", ""))
        value_box((x1 + x2) / 2, y2 + 38, record.get("bottom", ""))
        value_box(x1 - 92, mid_y - 40, record.get("left_mid", ""))

        left_gap = ""
        try:
            left_gap = f"{float(record.get('left_gap', '')) - float(record.get('left_mid', '')):.3f}"
        except Exception:
            left_gap = record.get("left_gap", "")
        value_box(x1 - 92, mid_y + 52, left_gap)
        value_box(x2 + 64, mid_y + 58, record.get("right_mid", ""))
        value_box(x2 + 88, mid_y, record.get("hole_shift", ""), "#dff3f8")

        product = record.get("product", "")
        stack = record.get("stack", "")
        if product or stack:
            canvas.create_text(14, 18, text=f"{product}  Stack {stack}".strip(), anchor="w", fill=MUTED_TEXT, font=("맑은 고딕", 9, "bold"))

    def append_tlb_log(self, text: str) -> None:
        if self.tlb_log_text is None or text.startswith("DNC 삭제 대기중"):
            return
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.tlb_log_text.configure(state="normal")
        self.tlb_log_text.insert(tk.END, f"[{timestamp}] {text}\n")
        self.tlb_log_text.see(tk.END)
        self.tlb_log_text.configure(state="disabled")

    def select_tlb_condition_sheet(self) -> None:
        path = filedialog.askopenfilename(parent=self.root, title="TLB 조건 시트 선택", filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if not path:
            return
        self.config["tlb_condition_sheet"] = path
        self.config["kcc_hdi_condition_sheet"] = path
        if hasattr(self, "tlb_condition_sheet_var"):
            self.tlb_condition_sheet_var.set(path)
        if hasattr(self, "kcc_hdi_condition_sheet_var"):
            self.kcc_hdi_condition_sheet_var.set(path)
        save_config(self.config)
        show_operator_alert(self.root, "조건 시트", "TLB / KCC HDI 조건 시트 선택 완료", "info")

    def validate_tlb_paths(self) -> bool:
        ok, message = validate_process_paths(self.config, "TLB")
        if not ok:
            show_operator_alert(self.root, "경로 확인", message)
            self.set_tlb_status("dnc", "경로 확인 필요", False)
            return False
        if not str(self.config.get("tlb_condition_sheet", "")).strip():
            show_operator_alert(self.root, "TLB 조건 시트", "조건 시트 선택 필요")
            self.set_tlb_status("dnc", "조건 시트 선택 필요", False)
            return False
        return True

    def apply_tlb_work_time_defaults(self) -> None:
        period = get_work_period()
        if "work_date" in self.tlb_common_entries:
            self.tlb_common_entries["work_date"].set(period["work_date"])
        if "work_date" in self.common_entries:
            self.common_entries["work_date"].set(period["work_date"])
        if "shift" in self.tlb_common_entries:
            self.tlb_common_entries["shift"].set(period["shift"])
        if "shift" in self.common_entries:
            self.common_entries["shift"].set(period["shift"])
        if self.config.get("auto_shift_group_enabled", True) and "shift_group" in self.tlb_common_entries:
            auto_group = get_auto_shift_group(
                period["work_date"],
                period["shift"],
                str(self.config.get("a_group_day_start_date", "2026-05-17")),
            )
            if auto_group:
                self.tlb_common_entries["shift_group"].set(auto_group)
                if "shift_group" in self.common_entries:
                    self.common_entries["shift_group"].set(auto_group)

    def ensure_tlb_work_period_ready(self) -> bool:
        self.apply_tlb_work_time_defaults()
        missing = []
        if not self.tlb_common_entries["shift_group"].get():
            missing.append("조")
        if not self.tlb_common_entries["worker"].get():
            missing.append("작업자")
        if missing:
            show_operator_alert(
                self.root,
                "근무 정보 확인",
                "\n".join(missing_common_message(label) for label in missing),
            )
            self.set_tlb_status("dnc", "근무 정보 확인 필요", False)
            return False
        return True

    def load_tlb_condition_jig(self, lot_no: int = 1) -> bool:
        self.save_settings_from_ui_silent()
        entries = self.tlb_entries if lot_no == 1 else self.tlb_lot2_entries
        status_key = "condition1" if lot_no == 1 else "condition2"
        lot = {key: entry.get() for key, entry in entries.items()}
        missing = [label for key, label in (("manage_no", "Tool No"), ("round", "차수")) if not lot.get(key, "").strip()]
        if missing:
            self.set_tlb_run_gradient(False)
            show_operator_alert(self.root, "입력 확인", " / ".join(missing) + " 입력 필요")
            return False
        try:
            record = lookup_tlb_condition_record_from_sheet(self.config, lot["manage_no"], lot["round"])
            condition = record["condition"]
            jig = record["jig"]
        except Exception as exc:
            log_error("TLB 조건 시트 조회 실패", exc)
            show_operator_alert(self.root, "조건 시트 조회", format_excel_error_for_operator(exc, "TLB 조건 시트"), "error")
            self.show_judgement_card(self.tlb_status_labels[status_key], "조건 조회", "NG", False)
            self.set_tlb_run_gradient(False)
            return False
        entries["condition"].set(condition)
        entries["jig"].set(jig)
        self.tlb_condition_records[lot_no] = record
        self.show_judgement_card(self.tlb_status_labels[status_key], "조건 조회", "OK", True)
        self.set_tlb_run_gradient(self.is_tlb_run_ready_by_condition())
        self.update_tlb_match_status(self.get_tlb_used_lots())
        self.refresh_tlb_cycle_preview()
        self.set_tlb_status("dnc", f"LOT {lot_no} 조건 조회 완료", True)
        return True

    def validate_tlb_dnc(self, common: dict, lots: list[dict]) -> tuple[bool, list[str]]:
        errors = []
        for key, label in (("work_date", "작업일자"), ("machine", "트리밍 호기"), ("shift_group", "조"), ("shift", "근무"), ("worker", "작업자")):
            if not common.get(key, "").strip():
                errors.append(missing_common_message(label))
        if not lots:
            errors.append("LOT 1 입력 필요")
        for index, lot in enumerate(lots, start=1):
            for key, label in (("manage_no", "Tool No"), ("round", "차수"), ("lot_no", "LOT No"), ("qty", "매수"), ("condition", "조건"), ("jig", "지그")):
                if not lot.get(key, "").strip():
                    errors.append(f"LOT {index} {label} 입력 필요")
            ok, message = validate_positive_number(lot.get("qty", ""), f"LOT {index} 매수", required=True)
            if not ok:
                errors.append(message)
        if len(lots) == 2:
            if lots[0].get("lot_no", "").strip() == lots[1].get("lot_no", "").strip():
                errors.append("LOT 1 / LOT 2 LOT No 동일")
            if lots[0].get("condition", "").strip() != lots[1].get("condition", "").strip():
                errors.append("LOT 1 / LOT 2 조건 불일치")
            if lots[0].get("jig", "").strip() != lots[1].get("jig", "").strip():
                errors.append("LOT 1 / LOT 2 지그 불일치")
        return len(errors) == 0, errors

    def validate_tlb_condition_file(self, condition_name: str) -> Path | None:
        self.set_tlb_status("dnc", "조건 파일 검색중", None)
        source = Path(self.config.get("source_dnc_folders", {}).get("TLB", self.config.get("source_dnc_folder", "")))
        matches = search_condition_file(condition_name, source)
        if not matches:
            show_operator_alert(self.root, "조건 파일 없음", "DNC 파일 없음", "error")
            return None
        if len(matches) >= 2:
            show_operator_alert(self.root, "동일 DNC 파일", format_duplicate_condition_files(matches), "error")
            return None
        return matches[0]

    def run_tlb_dnc(self) -> None:
        if self.is_running:
            show_operator_alert(self.root, "진행 중", "DNC 실행중")
            return
        if not self.ensure_tlb_work_period_ready():
            return
        if not self.save_settings_from_ui_silent():
            return
        if not self.validate_tlb_paths():
            return
        common = self.get_tlb_common_data()
        lots = self.get_tlb_used_lots()
        self.update_tlb_match_status(lots)
        if not self.ensure_tlb_condition_confirmed(lots):
            return
        ok, errors = self.validate_tlb_dnc(common, lots)
        if not ok:
            show_operator_alert(self.root, "입력값 확인", format_operator_errors(errors))
            self.set_tlb_status("dnc", "입력값 NG", False)
            return
        condition_file = self.validate_tlb_condition_file(lots[0]["condition"])
        if not condition_file:
            self.set_tlb_status("dnc", "조건 파일 NG", False)
            return
        signature = self.make_tlb_cycle_signature(common, lots)
        state = self.tlb_cycle_state
        if state and state.get("remaining", 0) > 0:
            if state.get("signature") != signature:
                show_operator_alert(self.root, "사이클 확인", "진행 중인 LOT 정보와 다릅니다.\n입력값을 확인하세요.", "error")
                self.set_tlb_status("dnc", "사이클 정보 불일치", False)
                return
            stack = state["stack"]
            cycle_count = int(state["total"])
            cycle_detail = f"남은 {state['remaining']}사이클 / 총 {cycle_count}사이클"
        else:
            model_change = ask_system_yes_no(self.root, "기종교체 확인", "기종교체 입니까?")
            self.frequent_check_values = [""] * 12
            self.work_axis_values = [""] * 6
            machine_axes = get_machine_allowed_axes(common.get("machine", ""))
            if model_change:
                self.set_tlb_status("dnc", "하부 Pin 확인 대기중", None)
                if not self.open_frequent_check_popup("jig", allowed_axes=machine_axes):
                    self.set_tlb_status("dnc", "하부 Pin 확인 미완료", False)
                    return
            stack = ask_numeric_input(self.root, "Stack 수 입력", "Stack 수를 입력 하세요.")
            ok, message = validate_positive_number(stack or "", "Stack 수", required=True)
            if not ok:
                show_operator_alert(self.root, "Stack 수 확인", message)
                self.set_tlb_status("dnc", "Stack 수 확인 필요", False)
                return
            if model_change:
                capacity_values = ["OK" if value == "OK" else "" for value in self.frequent_check_values[6:]] + [""] * 6
                self.work_axis_values = capacity_values[:6]
            else:
                self.set_tlb_status("dnc", "작업 축 수 확인 대기중", None)
                self.frequent_check_values = [""] * 12
                if not self.open_frequent_check_popup("capacity", allowed_axes=machine_axes):
                    self.set_tlb_status("dnc", "작업 축 수 확인 미완료", False)
                    return
                capacity_values = self.frequent_check_values[:]
                self.work_axis_values = capacity_values[:6]
            try:
                cycle_count, cycle_detail = self.calculate_tlb_cycle_count(common, lots, stack, capacity_values)
            except Exception as exc:
                show_operator_alert(self.root, "사이클 확인", str(exc), "error")
                self.set_tlb_status("dnc", "사이클 확인 필요", False)
                return
            self.tlb_cycle_state = {
                "signature": signature,
                "stack": stack,
                "total": cycle_count,
                "remaining": cycle_count,
                "log_ids": None,
                "model_change": model_change,
                "work_axis_values": self.work_axis_values[:],
                "jig_axis_values": self.frequent_check_values[6:],
                "capacity_values": capacity_values[:],
                "first_check_done": False,
            }
        self.update_tlb_cycle_status(cycle_count, int(self.tlb_cycle_state["remaining"]))
        self.set_tlb_status("dnc", cycle_detail, True)
        self.set_running(True)
        threading.Thread(target=self.tlb_worker, args=(common, lots, stack, condition_file), daemon=True).start()

    def tlb_worker(self, common: dict, lots: list[dict], stack: str, condition_file: Path) -> None:
        try:
            state = self.tlb_cycle_state or {}
            total = int(state.get("total", 1))
            remaining_before = int(state.get("remaining", total))
            cycle_index = total - remaining_before + 1
            if total > 1 and remaining_before == 1:
                self.root.after(0, self.set_tlb_last_cycle_running)
            if not state.get("log_ids"):
                self.root.after(0, lambda: self.set_tlb_status("dnc", "DB 저장중", None))
                log_ids = insert_tlb_dnc_db(common, lots, stack, bool(state.get("model_change")))
                state["log_ids"] = log_ids
                self.tlb_cycle_state = state
            else:
                log_ids = list(state["log_ids"])
            transfer_folder = Path(self.config["transfer_dnc_folder"])
            self.root.after(0, lambda i=cycle_index, c=total: self.set_tlb_status("dnc", f"DNC 사이클 {i}/{c} 진행중", None))
            delete_existing_dnc_txt(transfer_folder)
            copied_file = copy_dnc_file(condition_file, transfer_folder)
            self.root.after(0, lambda i=cycle_index, c=total: self.set_tlb_status("dnc", f"DNC 파일 복사 완료 ({i}/{c})", True))
            delete_thread = threading.Thread(
                target=delete_after_delay,
                args=(copied_file, int(self.config["dnc_delete_seconds"]), lambda text, i=cycle_index, c=total: self.root.after(0, lambda t=text: self.set_tlb_status("dnc", f"{t} / {i}/{c}", None))),
                daemon=True,
            )
            delete_thread.start()
            if not state.get("first_check_done"):
                wait_seconds = int(self.config.get("first_article_wait_seconds", FIRST_ARTICLE_WAIT_SECONDS))
                for remain in range(wait_seconds, 0, -1):
                    self.root.after(0, lambda r=remain: self.set_tlb_status("dnc", f"초품 확인 대기중 ({r}초)", None))
                    time.sleep(1)
            self.root.after(0, lambda ids=log_ids, lt=lots, st=stack, dt=delete_thread: self.finish_tlb_cycle_dnc(ids, lt, st, dt))
        except Exception as exc:
            self.root.after(0, lambda error=exc: self.handle_run_error(error))

    def finish_tlb_cycle_dnc(self, log_ids: list[int], lots: list[dict], stack: str, delete_thread: threading.Thread | None = None) -> None:
        try:
            state = self.tlb_cycle_state or {}
            model_change = bool(state.get("model_change"))
            if not state.get("first_check_done"):
                while True:
                    if model_change:
                        allowed_axes = [
                            index
                            for index, value in enumerate(state.get("jig_axis_values", []))
                            if value == "OK"
                        ]
                        self.frequent_check_values = [""] * 6 + list(state.get("jig_axis_values", [""] * 6))
                    else:
                        allowed_axes = [
                            index
                            for index, value in enumerate(state.get("work_axis_values", []))
                            if value == "OK"
                        ]
                        self.frequent_check_values = [""] * 12
                    if not self.open_frequent_check_popup("first", allowed_axes=allowed_axes):
                        self.set_tlb_status("dnc", "초품 확인 미완료", False)
                        if delete_thread and delete_thread.is_alive():
                            def wait_and_release_after_first_cancel() -> None:
                                delete_thread.join()
                                self.root.after(0, lambda: self.set_running(False))

                            threading.Thread(target=wait_and_release_after_first_cancel, daemon=True).start()
                        else:
                            self.set_running(False)
                        return
                    if model_change:
                        first_axes = [
                            index
                            for index, value in enumerate(self.frequent_check_values[:6])
                            if value == "OK"
                        ]
                        if first_axes != allowed_axes:
                            show_operator_alert(self.root, "초품 4Point 확인", "하부 Pin 축과 초품 축 다름")
                            self.set_tlb_status("dnc", "초품 축 확인 NG", False)
                            continue
                    if count_frequent_check_axes(self.frequent_check_values) > 0:
                        self.set_tlb_status("dnc", "초품 확인 완료", True)
                        break
                    show_operator_alert(self.root, "초품 확인", "확인 축 선택 필요")
                    self.set_tlb_status("dnc", "초품 확인 NG", False)
                update_normal_frequent_check_db(log_ids, model_change, self.frequent_check_values, "TLB")
                state["first_check_done"] = True
            state["remaining"] = max(0, int(state.get("remaining", 1)) - 1)
            self.tlb_cycle_state = state
            if state["remaining"] > 0:
                self.hide_tlb_condition_cards()
                self.update_tlb_cycle_status(int(state.get("total", 1)), int(state["remaining"]))
                if delete_thread and delete_thread.is_alive():
                    self.set_tlb_status("dnc", "DNC 삭제 완료 대기중", None)

                    def wait_and_release() -> None:
                        delete_thread.join()
                        self.root.after(0, lambda r=state["remaining"]: self.set_tlb_status("dnc", f"이번 사이클 완료 / 남은 {r}회", True))
                        self.root.after(0, lambda: self.set_running(False))

                    threading.Thread(target=wait_and_release, daemon=True).start()
                    return
                self.set_tlb_status("dnc", f"이번 사이클 완료 / 남은 {state['remaining']}회", True)
                self.set_running(False)
                return
            if delete_thread and delete_thread.is_alive():
                self.set_tlb_status("dnc", "DNC 완료 대기중", None)

                def wait_and_finish() -> None:
                    delete_thread.join()
                    self.root.after(0, lambda: self.finish_tlb_dnc(log_ids))

                threading.Thread(target=wait_and_finish, daemon=True).start()
                return
            self.finish_tlb_dnc(log_ids)
        except Exception as exc:
            self.handle_run_error(exc)

    def finish_tlb_dnc(self, log_ids: list[int]) -> None:
        try:
            self.hide_tlb_cycle_status()
            self.hide_tlb_condition_cards()
            burr_ok = ask_system_yes_no(self.root, "Burr 확인", "4면 Burr 이상 없습니까?")
            update_normal_burr_db(log_ids, burr_ok, "TLB")
            pending = get_unexported_process_log_count("TLB")
            self.set_tlb_status("dnc", "DNC 완료", True)
            self.set_tlb_status("excel", f"DB 저장 완료 / Excel 미반영 {pending}건", True)
            self.auto_export_tlb_to_excel(parent=self.root)
            self.tlb_cycle_state = None
            self.clear_tlb_inputs(after_done=True)
        except Exception as exc:
            self.handle_run_error(exc)
        finally:
            self.set_running(False)

    def auto_export_tlb_to_excel(self, parent=None) -> bool:
        try:
            export_process_logs_to_excel(self.config, "TLB", "TLB")
            pending = get_unexported_process_log_count("TLB")
            self.set_tlb_status("excel", f"작업일보 반영 완료 / Excel 미반영 {pending}건", True)
            return True
        except Exception as exc:
            log_error("작업일보 자동 반영 실패", exc)
            self.set_tlb_status("excel", f"Excel 미반영 {get_unexported_process_log_count('TLB')}건", False)
            return False

    def export_tlb_to_excel_from_ui(self) -> None:
        try:
            count = export_process_logs_to_excel(self.config, "TLB", "TLB")
        except Exception as exc:
            log_error("TLB 작업일보 반영 실패", exc)
            show_operator_alert(self.root, "작업일보 반영 실패", format_excel_error_for_operator(exc, "작업일보 Excel"), "error")
            self.set_tlb_status("excel", "작업일보 반영 실패", False)
            return
        pending = get_unexported_process_log_count("TLB")
        show_operator_alert(self.root, "작업일보 반영", f"{count}건 반영 완료", "info")
        self.set_tlb_status("excel", f"Excel 미반영 {pending}건", True)

    def clear_tlb_inputs(self, after_done: bool = False) -> None:
        self.tlb_cycle_state = None
        self.tlb_condition_records.clear()
        for entries in (self.tlb_entries, self.tlb_lot2_entries):
            for entry in entries.values():
                entry.clear()
        self.hide_tlb_condition_cards()
        self.hide_tlb_cycle_status()
        self.set_tlb_run_gradient(False)
        self.set_tlb_status("dnc", "DNC 완료" if after_done else "대기중", True if after_done else None)

    def create_kcc_hdi_tab(self) -> None:
        font_name = "맑은 고딕"
        kcc_hdi_theme = PROCESS_COLORS["KCC HDI"]
        kcc_hdi_bg = kcc_hdi_theme["bg"]
        kcc_hdi_light = kcc_hdi_theme["light"]
        kcc_hdi_primary = kcc_hdi_theme["primary"]
        kcc_hdi_border = kcc_hdi_theme["border"]
        self.kcc_hdi_page.configure(bg=kcc_hdi_bg)
        self.kcc_hdi_page.columnconfigure(0, weight=1)
        self.kcc_hdi_page.rowconfigure(2, weight=0)
        self.kcc_hdi_page.rowconfigure(3, weight=0)
        title_wrap = tk.Frame(self.kcc_hdi_page, bg=kcc_hdi_light)
        title_wrap.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 8))
        title_wrap.columnconfigure(0, minsize=260)
        title_wrap.columnconfigure(1, weight=1)
        title_wrap.columnconfigure(2, minsize=360)
        logo_slot = tk.Frame(title_wrap, bg=kcc_hdi_light, width=220, height=1)
        logo_slot.grid(row=0, column=0, sticky="w", padx=(14, 8))
        logo_slot.grid_propagate(False)
        kcc_hdi_logo_path = BUNDLED_KCC_LOGO_FILE if BUNDLED_KCC_LOGO_FILE.exists() else KCC_LOGO_FILE
        if kcc_hdi_logo_path.exists():
            try:
                kcc_hdi_logo = tk.PhotoImage(file=str(kcc_hdi_logo_path))
                scale = max(1, kcc_hdi_logo.height() // 32)
                self.kcc_hdi_logo_image = kcc_hdi_logo.subsample(scale, scale)
                tk.Label(logo_slot, image=self.kcc_hdi_logo_image, bg=kcc_hdi_light, bd=0).pack(side=tk.LEFT, anchor="w")
            except tk.TclError:
                self.kcc_hdi_logo_image = None
        tk.Label(title_wrap, text="KCC HDI DNC", bg=kcc_hdi_light, fg=kcc_hdi_primary, font=(font_name, 14, "bold"), height=2).grid(row=0, column=1, sticky="ew")
        title_buttons = tk.Frame(title_wrap, bg=kcc_hdi_light)
        title_buttons.grid(row=0, column=2, sticky="e", padx=(8, 10))
        self.kcc_hdi_run_button = self.create_tlb_gradient_run_button(title_buttons, command=self.run_kcc_hdi_dnc, scheme="kcc")
        self.kcc_hdi_run_button.grid(row=0, column=0, padx=4, pady=4)
        self.add_normal_button(title_buttons, "입력 초기화", self.clear_kcc_hdi_inputs).grid(row=0, column=1, padx=4, pady=4)

        common = self.create_panel(self.kcc_hdi_page, "공통 입력", kcc_hdi_theme)
        common.configure(highlightbackground=kcc_hdi_border)
        common.winfo_children()[0].configure(bg=kcc_hdi_light, fg=kcc_hdi_primary)
        common.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 8))
        common_widgets = [
            ("machine", ComboField(common, "설비 호기", ["트리밍 1호기", "트리밍 2호기", "트리밍 3호기"], initial=self.config.get("machine", "트리밍 1호기"), width=12)),
            ("work_date", DateField(common, "작업일자", on_change=lambda key="work_date": self.handle_common_change(key, "kcc_hdi"))),
            ("shift_group", SegmentedField(common, "조", ["A", "B", "C"], allow_empty=True, on_change=lambda key="shift_group": self.handle_common_change(key, "kcc_hdi"))),
            ("shift", SegmentedField(common, "근무", ["주간", "야간"], on_change=lambda key="shift": self.handle_common_change(key, "kcc_hdi"))),
            ("worker", LabeledEntry(common, "작업자", width=12, on_change=lambda key="worker": self.handle_common_change(key, "kcc_hdi"), live_change=False)),
        ]
        for index, (key, entry) in enumerate(common_widgets):
            entry.grid(row=1, column=index, sticky="ew", padx=8, pady=8)
            self.kcc_hdi_common_entries[key] = entry
            if key == "machine":
                entry.combo.bind("<<ComboboxSelected>>", lambda _event, field=key: self.handle_common_change(field, "kcc_hdi"))
        common.columnconfigure(0, weight=0, minsize=320)
        common.columnconfigure(1, weight=1, minsize=430)
        common.columnconfigure(2, weight=1, minsize=330)
        common.columnconfigure(3, weight=1, minsize=360)
        common.columnconfigure(4, weight=0, minsize=320)

        body = tk.Frame(self.kcc_hdi_page, bg=kcc_hdi_bg)
        body.grid(row=2, column=0, sticky="nsew", padx=14, pady=0)
        body.columnconfigure(0, weight=1, uniform="kcc_hdi_body")
        body.columnconfigure(1, weight=1, uniform="kcc_hdi_body")
        fields = [
            ("manage_no", "\uad00\ub9ac\ubc88\ud638"),
            ("round", "차수"),
            ("lot_no", "LOT No"),
            ("qty", "매수"),
            ("condition", "조건(조회)"),
            ("jig", "지그(조회)"),
        ]

        def build_kcc_hdi_lot_panel(parent: tk.Frame, title: str, entries: dict, status_key: str, column: int) -> None:
            panel = self.create_panel(parent, title, kcc_hdi_theme)
            panel.configure(highlightbackground=kcc_hdi_border)
            panel.winfo_children()[0].configure(bg=kcc_hdi_light, fg=kcc_hdi_primary)
            panel.grid(row=0, column=column, sticky="nsew", padx=(0, 8) if column == 0 else (8, 0))
            lot_no = 1 if column == 0 else 2
            for index, (key, label) in enumerate(fields):
                row = index // 2 + 1
                col = index % 2
                if key == "round":
                    entry = RoundField(panel, label)
                elif key in {"condition", "jig"}:
                    entry = LabeledEntry(panel, label, width=32, style="Lookup.TEntry", readonly=True)
                elif key == "qty":
                    entry = LabeledEntry(panel, label, width=32, numeric_only=True)
                else:
                    entry = LabeledEntry(panel, label, width=32, uppercase=True)
                entry.grid(row=row, column=col, sticky="ew", padx=14, pady=9)
                panel.columnconfigure(col, weight=1)
                entries[key] = entry
                if key in {"manage_no", "round", "lot_no", "qty"}:
                    entry.var.trace_add("write", lambda *_args, lot_no=lot_no: self.handle_kcc_hdi_input_changed(lot_no))
            self.create_mes_lookup_button(
                panel,
                command=lambda lot_no=lot_no: self.load_kcc_hdi_condition_jig(lot_no),
                scheme="tlb",
                ready_check=lambda lot_no=lot_no: self.is_kcc_hdi_lookup_ready(lot_no),
            ).grid(row=4, column=0, columnspan=2, sticky="ew", padx=14, pady=(12, 8))
            status = tk.Frame(panel, bg=SURFACE_BG)
            status.grid(row=5, column=0, columnspan=2, sticky="ew", padx=14, pady=(0, 12))
            status.columnconfigure(0, weight=1, uniform="kcc_hdi_lot_status")
            status.columnconfigure(1, weight=1, uniform="kcc_hdi_lot_status")
            self.kcc_hdi_status_labels[status_key] = self.create_judgement_card(status, "조건 조회")
            self.kcc_hdi_status_labels[status_key].grid(row=0, column=0, sticky="ew", padx=(0, 6))
            self.hide_judgement_card(self.kcc_hdi_status_labels[status_key])
            if column == 0:
                self.kcc_hdi_status_labels["cycle"] = self.create_judgement_card(status, "완료 사이클")
                self.kcc_hdi_status_labels["cycle"].grid(row=0, column=1, sticky="ew", padx=(6, 0))
                self.hide_judgement_card(self.kcc_hdi_status_labels["cycle"])

        build_kcc_hdi_lot_panel(body, "LOT 1 입력", self.kcc_hdi_entries, "condition1", 0)
        build_kcc_hdi_lot_panel(body, "LOT 2 입력 (선택)", self.kcc_hdi_lot2_entries, "condition2", 1)
        self.kcc_hdi_status_labels["condition"] = self.kcc_hdi_status_labels["condition1"]

        bottom = tk.Frame(self.kcc_hdi_page, bg=kcc_hdi_bg)
        bottom.grid(row=3, column=0, sticky="ew", padx=14, pady=(8, 14))
        bottom.columnconfigure(0, weight=1)
        status_panel = tk.Frame(bottom, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        status_panel.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        status_panel.columnconfigure(1, weight=1)
        tk.Label(status_panel, text="2LOT 조건 일치 확인", bg=kcc_hdi_light, fg=kcc_hdi_primary, font=(font_name, 11, "bold"), width=22, height=2).grid(row=0, column=0, sticky="nsw")
        match_label = tk.Label(status_panel, text="LOT 2 미사용", bg=SURFACE_BG, fg=MUTED_TEXT, font=(font_name, 12, "bold"), anchor="w")
        match_label.grid(row=0, column=1, sticky="ew", padx=14)
        self.kcc_hdi_status_labels["match"] = match_label
        tk.Label(status_panel, text="DNC 진행 상태", bg=kcc_hdi_light, fg=kcc_hdi_primary, font=(font_name, 11, "bold"), width=22, height=2).grid(row=1, column=0, sticky="nsw")
        dnc_label = tk.Label(status_panel, text="대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=(font_name, 12, "bold"), anchor="w")
        dnc_label.grid(row=1, column=1, sticky="ew", padx=14)
        self.kcc_hdi_status_labels["dnc"] = dnc_label
        tk.Label(status_panel, text="작업일보 반영", bg=kcc_hdi_light, fg=kcc_hdi_primary, font=(font_name, 11, "bold"), width=22, height=2).grid(row=2, column=0, sticky="nsw")
        excel_label = tk.Label(status_panel, text="대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=(font_name, 12, "bold"), anchor="w")
        excel_label.grid(row=2, column=1, sticky="ew", padx=14)
        self.kcc_hdi_status_labels["excel"] = excel_label
        log_panel = tk.Frame(bottom, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        log_panel.grid(row=1, column=0, sticky="ew", padx=(0, 10), pady=(8, 0))
        log_panel.columnconfigure(0, weight=1)
        tk.Label(log_panel, text="KCC HDI DNC 작업 로그", bg=kcc_hdi_light, fg=kcc_hdi_primary, font=(font_name, 10, "bold"), height=1).grid(row=0, column=0, sticky="ew")
        self.kcc_hdi_log_text = scrolledtext.ScrolledText(log_panel, height=5, wrap=tk.WORD, state="disabled", bg=SURFACE_BG, fg=TEXT_COLOR, font=(font_name, 10), relief=tk.FLAT, padx=10, pady=8)
        self.kcc_hdi_log_text.grid(row=1, column=0, sticky="ew")
        button_panel = tk.Frame(bottom, bg=kcc_hdi_bg)
        button_panel.grid(row=0, column=1, rowspan=2, sticky="ne")
        button_panel.columnconfigure((0, 1), weight=1, uniform="kcc_hdi_side")
        self.add_side_button(button_panel, "조건 시트 선택", self.select_kcc_hdi_condition_sheet, "SidePrimary.TButton").grid(row=0, column=0, sticky="nsew", padx=4, pady=4)
        self.add_side_button(button_panel, "작업일보 반영", self.export_kcc_hdi_to_excel_from_ui).grid(row=1, column=0, sticky="nsew", padx=4, pady=4)
        self.add_side_button(button_panel, "작업일보 열기", self.open_log_excel_from_ui).grid(row=1, column=1, sticky="nsew", padx=4, pady=4)

    def get_kcc_hdi_common_data(self) -> dict:
        return {key: entry.get() for key, entry in self.kcc_hdi_common_entries.items()}

    def get_kcc_hdi_lot_data(self) -> dict:
        return {key: entry.get() for key, entry in self.kcc_hdi_entries.items()}

    def get_kcc_hdi_lot2_data(self) -> dict:
        return {key: entry.get() for key, entry in self.kcc_hdi_lot2_entries.items()}

    def is_kcc_hdi_lookup_ready(self, lot_no: int) -> bool:
        entries = self.kcc_hdi_entries if lot_no == 1 else self.kcc_hdi_lot2_entries
        required_keys = ("manage_no", "round", "lot_no", "qty")
        return all(entries[key].get().strip() for key in required_keys)

    def is_kcc_hdi_lot_used(self, lot: dict) -> bool:
        return any(str(lot.get(key, "")).strip() for key in ("manage_no", "round", "lot_no", "qty", "condition", "jig"))

    def handle_kcc_hdi_input_changed(self, lot_no: int) -> None:
        """TLB 입력값이 바뀌면 기존 조건 조회 결과를 즉시 무효화합니다."""
        entries = self.kcc_hdi_entries if lot_no == 1 else self.kcc_hdi_lot2_entries
        for key in ("condition", "jig"):
            if key in entries:
                entries[key].clear()
        self.kcc_hdi_condition_records.pop(lot_no, None)
        status_key = "condition1" if lot_no == 1 else "condition2"
        label = self.kcc_hdi_status_labels.get(status_key)
        if label is not None:
            self.hide_judgement_card(label)
        self.hide_kcc_hdi_cycle_status()
        self.set_kcc_hdi_run_gradient(False)
        self.update_kcc_hdi_match_status(self.get_kcc_hdi_used_lots())

    def get_kcc_hdi_used_lots(self) -> list[dict]:
        lots = [self.get_kcc_hdi_lot_data()]
        lot2 = self.get_kcc_hdi_lot2_data()
        if self.is_kcc_hdi_lot_used(lot2):
            lots.append(lot2)
        return lots

    def is_kcc_hdi_run_ready_by_condition(self) -> bool:
        lot1 = self.get_kcc_hdi_lot_data()
        if not self.is_kcc_hdi_lot_condition_confirmed(1, lot1):
            return False
        lot2 = self.get_kcc_hdi_lot2_data()
        if self.is_kcc_hdi_lot_used(lot2):
            return self.is_kcc_hdi_lot_condition_confirmed(2, lot2)
        return True

    def is_kcc_hdi_lot_condition_confirmed(self, lot_no: int, lot: dict) -> bool:
        """조건 시트 조회 버튼으로 확정된 LOT인지 확인합니다."""
        record = self.kcc_hdi_condition_records.get(lot_no)
        if not record:
            return False
        return (
            lot.get("condition", "").strip() == str(record.get("condition", "")).strip()
            and lot.get("jig", "").strip() == str(record.get("jig", "")).strip()
        )

    def ensure_kcc_hdi_condition_confirmed(self, lots: list[dict]) -> bool:
        missing_lots = [
            f"LOT {index}"
            for index, lot in enumerate(lots, start=1)
            if not self.is_kcc_hdi_lot_condition_confirmed(index, lot)
        ]
        if missing_lots:
            show_operator_alert(self.root, "DNC 조건 조회", "DNC 조건 조회 필요")
            self.set_kcc_hdi_status("dnc", "DNC 조건 조회 필요", False)
            return False
        return True

    def calculate_kcc_hdi_cycle_count(
        self,
        common: dict,
        lots: list[dict],
        stack: str,
        axis_values: list[str] | None = None,
    ) -> tuple[int, str]:
        total_qty = sum(int(lot.get("qty", "0") or 0) for lot in lots)
        axis_count = count_frequent_check_axes(axis_values) if axis_values is not None else get_machine_axis_count(common.get("machine", ""))
        stack_text = str(stack or "").strip()
        ok, message = validate_positive_number(stack_text, "Stack", required=True)
        if not ok:
            raise ValueError(message)
        stack_count = int(stack_text)
        capacity = axis_count * stack_count
        if capacity <= 0:
            raise ValueError("설비 축 수 / Stack 확인 필요")
        cycle_count = max(1, (total_qty + capacity - 1) // capacity)
        detail = f"총 {cycle_count}사이클 / {total_qty}매 / {axis_count}축 x {stack_count}Stack"
        return cycle_count, detail

    def make_kcc_hdi_cycle_signature(self, common: dict, lots: list[dict]) -> tuple:
        return (
            common.get("machine", "").strip(),
            common.get("work_date", "").strip(),
            common.get("shift_group", "").strip(),
            common.get("shift", "").strip(),
            common.get("worker", "").strip(),
            tuple(
                (
                    lot.get("manage_no", "").strip(),
                    lot.get("round", "").strip(),
                    lot.get("lot_no", "").strip(),
                    lot.get("qty", "").strip(),
                    lot.get("condition", "").strip(),
                    lot.get("jig", "").strip(),
                )
                for lot in lots
            ),
        )

    def update_kcc_hdi_cycle_status(self, cycle_count: int, remaining: int | None = None) -> None:
        label = self.kcc_hdi_status_labels.get("cycle")
        if label is None:
            return
        if cycle_count <= 1 or remaining is None or remaining <= 0:
            self.hide_kcc_hdi_cycle_status()
            return
        completed_cycle = cycle_count - remaining
        if completed_cycle <= 0:
            self.hide_kcc_hdi_cycle_status()
            return
        completed_cycle = min(cycle_count - 1, completed_cycle)
        text = f"완료 사이클\n{completed_cycle} / {cycle_count}회"
        condition_label = self.kcc_hdi_status_labels.get("condition1")
        if condition_label is not None:
            condition_label.grid_configure(row=0, column=0, columnspan=1, sticky="ew", padx=(0, 6))
        label.configure(text=text, fg="#075985", bg="#e0f2fe", highlightthickness=2, highlightbackground="#38bdf8")
        label.grid(row=0, column=1, sticky="ew", padx=(6, 0))

    def set_kcc_hdi_last_cycle_running(self) -> None:
        label = self.kcc_hdi_status_labels.get("cycle")
        if label is None:
            return
        condition_label = self.kcc_hdi_status_labels.get("condition1")
        if condition_label is not None:
            condition_label.grid_configure(row=0, column=0, columnspan=1, sticky="ew", padx=(0, 6))
        label.configure(text="마지막 Cycle\n진행 중", fg="#075985", bg="#e0f2fe", highlightthickness=2, highlightbackground="#38bdf8")
        label.grid(row=0, column=1, sticky="ew", padx=(6, 0))

    def hide_kcc_hdi_cycle_status(self) -> None:
        label = self.kcc_hdi_status_labels.get("cycle")
        if label is not None:
            self.hide_judgement_card(label)
            label.grid(row=0, column=1, sticky="ew", padx=(6, 0))
        condition_label = self.kcc_hdi_status_labels.get("condition1")
        if condition_label is not None:
            condition_label.grid_configure(row=0, column=0, columnspan=1, sticky="ew", padx=(0, 6))

    def hide_kcc_hdi_condition_cards(self) -> None:
        """KCC HDI 조건 조회 결과 카드는 조회/실행 직후에만 보이게 초기화합니다."""
        for key in ("condition1", "condition2"):
            label = self.kcc_hdi_status_labels.get(key)
            if label is not None:
                self.hide_judgement_card(label)

    def refresh_kcc_hdi_cycle_preview(self) -> None:
        # KCC HDI 사이클은 작업자가 DNC 실행 시 입력하는 Stack 기준으로 계산합니다.
        # 조건 조회 단계에서는 실제 Stack을 아직 모르므로 사이클을 미리 표시하지 않습니다.
        return

    def update_kcc_hdi_match_status(self, lots: list[dict]) -> None:
        label = self.kcc_hdi_status_labels.get("match")
        if label is None:
            return
        if len(lots) < 2:
            label.configure(text="LOT 2 미사용", fg=MUTED_TEXT)
            return
        lot1, lot2 = lots[0], lots[1]
        if lot1.get("condition", "").strip() != lot2.get("condition", "").strip():
            label.configure(text="NG - 조건 불일치", fg=NG_COLOR)
            return
        if lot1.get("jig", "").strip() != lot2.get("jig", "").strip():
            label.configure(text="NG - 지그 불일치", fg=NG_COLOR)
            return
        label.configure(text="OK - 2LOT 조건 일치", fg=OK_COLOR)

    def set_kcc_hdi_status(self, key: str, text: str, ok: bool | None = None) -> None:
        label = self.kcc_hdi_status_labels.get(key)
        if label is None:
            return
        fg = OK_COLOR if ok is True else NG_COLOR if ok is False else MUTED_TEXT
        label.configure(text=text, fg=fg)
        if key == "dnc":
            self.append_kcc_hdi_log(text)

    def draw_kcc_hdi_condition_preview(self, record: dict | None) -> None:
        canvas = self.kcc_hdi_preview_canvas
        if canvas is None:
            return
        canvas.delete("all")
        width = max(canvas.winfo_width(), 520)
        height = max(canvas.winfo_height(), 250)
        if not record:
            canvas.create_text(width / 2, height / 2, text="조건 시트 조회 후 표시", fill=MUTED_TEXT, font=("맑은 고딕", 12, "bold"))
            return

        line_color = "#2563eb"
        guide_color = "#93c5fd"
        value_bg = "#dff3f8"
        text_color = TEXT_COLOR
        red = "#ef4444"

        x1, y1 = width * 0.25, height * 0.22
        x2, y2 = width * 0.78, height * 0.78
        mid_y = height * 0.50
        hole_x = width * 0.245
        right_hole_x = width * 0.79

        def fmt(value: str) -> str:
            if value in ("", None):
                return ""
            try:
                return f"{float(value):.3f}"
            except Exception:
                return str(value)

        def value_box(x: float, y: float, value: str, fill: str = value_bg) -> None:
            text = fmt(value)
            if not text:
                return
            canvas.create_rectangle(x - 48, y - 14, x + 48, y + 14, fill=fill, outline="")
            canvas.create_text(x, y, text=text, fill=red if str(value).startswith("-") else text_color, font=("맑은 고딕", 10))

        canvas.create_rectangle(x1, y1, x2, y2, outline=line_color, width=2)
        canvas.create_line(x1, mid_y, x2 + 54, mid_y, fill=line_color, dash=(2, 2))
        canvas.create_line(hole_x, y1 - 26, hole_x, y2 + 26, fill=line_color, dash=(2, 2))
        canvas.create_line(right_hole_x, y1 - 26, right_hole_x, y2 + 26, fill=line_color, dash=(2, 2))
        canvas.create_line(x1 - 50, mid_y + 22, x2 + 68, mid_y + 22, fill=guide_color)
        canvas.create_line(x1 - 44, y1 - 14, x2 + 48, y1 - 14, fill=guide_color, dash=(2, 2))
        canvas.create_line(x1 - 44, y2 + 14, x2 + 48, y2 + 14, fill=guide_color, dash=(2, 2))

        for x, y in ((hole_x, mid_y), (hole_x, mid_y + 52), (right_hole_x, mid_y)):
            canvas.create_oval(x - 8, y - 8, x + 8, y + 8, outline="#64748b", width=3)
            canvas.create_oval(x - 3, y - 3, x + 3, y + 3, fill="#64748b", outline="")

        value_box((x1 + x2) / 2, y1 - 38, record.get("top", ""))
        value_box((x1 + x2) / 2, mid_y - 28, record.get("center", ""))
        value_box((x1 + x2) / 2, mid_y + 58, record.get("bottom_left", ""))
        value_box((x1 + x2) / 2, y2 + 38, record.get("bottom", ""))
        value_box(x1 - 92, mid_y - 40, record.get("left_mid", ""))

        left_gap = ""
        try:
            left_gap = f"{float(record.get('left_gap', '')) - float(record.get('left_mid', '')):.3f}"
        except Exception:
            left_gap = record.get("left_gap", "")
        value_box(x1 - 92, mid_y + 52, left_gap)
        value_box(x2 + 64, mid_y + 58, record.get("right_mid", ""))
        value_box(x2 + 88, mid_y, record.get("hole_shift", ""), "#dff3f8")

        product = record.get("product", "")
        stack = record.get("stack", "")
        if product or stack:
            canvas.create_text(14, 18, text=f"{product}  Stack {stack}".strip(), anchor="w", fill=MUTED_TEXT, font=("맑은 고딕", 9, "bold"))

    def append_kcc_hdi_log(self, text: str) -> None:
        if self.kcc_hdi_log_text is None or text.startswith("DNC 삭제 대기중"):
            return
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.kcc_hdi_log_text.configure(state="normal")
        self.kcc_hdi_log_text.insert(tk.END, f"[{timestamp}] {text}\n")
        self.kcc_hdi_log_text.see(tk.END)
        self.kcc_hdi_log_text.configure(state="disabled")

    def select_kcc_hdi_condition_sheet(self) -> None:
        path = filedialog.askopenfilename(parent=self.root, title="KCC HDI 조건 시트 선택", filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if not path:
            return
        self.config["tlb_condition_sheet"] = path
        self.config["kcc_hdi_condition_sheet"] = path
        if hasattr(self, "tlb_condition_sheet_var"):
            self.tlb_condition_sheet_var.set(path)
        if hasattr(self, "kcc_hdi_condition_sheet_var"):
            self.kcc_hdi_condition_sheet_var.set(path)
        save_config(self.config)
        show_operator_alert(self.root, "조건 시트", "TLB / KCC HDI 조건 시트 선택 완료", "info")

    def validate_kcc_hdi_paths(self) -> bool:
        ok, message = validate_process_paths(self.config, "KCC HDI")
        if not ok:
            show_operator_alert(self.root, "경로 확인", message)
            self.set_kcc_hdi_status("dnc", "경로 확인 필요", False)
            return False
        if not str(self.config.get("kcc_hdi_condition_sheet", "")).strip():
            show_operator_alert(self.root, "KCC HDI 조건 시트", "조건 시트 선택 필요")
            self.set_kcc_hdi_status("dnc", "조건 시트 선택 필요", False)
            return False
        return True
    def apply_kcc_hdi_work_time_defaults(self) -> None:
        period = get_work_period()
        if "work_date" in self.kcc_hdi_common_entries:
            self.kcc_hdi_common_entries["work_date"].set(period["work_date"])
        if "work_date" in self.common_entries:
            self.common_entries["work_date"].set(period["work_date"])
        if "shift" in self.kcc_hdi_common_entries:
            self.kcc_hdi_common_entries["shift"].set(period["shift"])
        if "shift" in self.common_entries:
            self.common_entries["shift"].set(period["shift"])
        if self.config.get("auto_shift_group_enabled", True) and "shift_group" in self.kcc_hdi_common_entries:
            auto_group = get_auto_shift_group(
                period["work_date"],
                period["shift"],
                str(self.config.get("a_group_day_start_date", "2026-05-17")),
            )
            if auto_group:
                self.kcc_hdi_common_entries["shift_group"].set(auto_group)
                if "shift_group" in self.common_entries:
                    self.common_entries["shift_group"].set(auto_group)

    def ensure_kcc_hdi_work_period_ready(self) -> bool:
        self.apply_kcc_hdi_work_time_defaults()
        missing = []
        if not self.kcc_hdi_common_entries["shift_group"].get():
            missing.append("조")
        if not self.kcc_hdi_common_entries["worker"].get():
            missing.append("작업자")
        if missing:
            show_operator_alert(
                self.root,
                "근무 정보 확인",
                "\n".join(missing_common_message(label) for label in missing),
            )
            self.set_kcc_hdi_status("dnc", "근무 정보 확인 필요", False)
            return False
        return True

    def load_kcc_hdi_condition_jig(self, lot_no: int = 1) -> bool:
        self.save_settings_from_ui_silent()
        entries = self.kcc_hdi_entries if lot_no == 1 else self.kcc_hdi_lot2_entries
        status_key = "condition1" if lot_no == 1 else "condition2"
        lot = {key: entry.get() for key, entry in entries.items()}
        missing = [label for key, label in (("manage_no", "\uad00\ub9ac\ubc88\ud638"), ("round", "차수")) if not lot.get(key, "").strip()]
        if missing:
            self.set_kcc_hdi_run_gradient(False)
            show_operator_alert(self.root, "입력 확인", " / ".join(missing) + " 입력 필요")
            return False
        try:
            record = lookup_kcc_hdi_condition_record_from_sheet(self.config, lot["manage_no"], lot["round"])
            condition = record["condition"]
            jig = record["jig"]
        except Exception as exc:
            log_error("KCC HDI 조건 시트 조회 실패", exc)
            show_operator_alert(self.root, "조건 시트 조회", format_excel_error_for_operator(exc, "KCC HDI 조건 시트"), "error")
            self.show_judgement_card(self.kcc_hdi_status_labels[status_key], "조건 조회", "NG", False)
            self.set_kcc_hdi_run_gradient(False)
            return False
        entries["condition"].set(condition)
        entries["jig"].set(jig)
        self.kcc_hdi_condition_records[lot_no] = record
        self.show_judgement_card(self.kcc_hdi_status_labels[status_key], "조건 조회", "OK", True)
        self.set_kcc_hdi_run_gradient(self.is_kcc_hdi_run_ready_by_condition())
        self.update_kcc_hdi_match_status(self.get_kcc_hdi_used_lots())
        self.refresh_kcc_hdi_cycle_preview()
        self.set_kcc_hdi_status("dnc", f"LOT {lot_no} 조건 조회 완료", True)
        return True

    def validate_kcc_hdi_dnc(self, common: dict, lots: list[dict]) -> tuple[bool, list[str]]:
        errors = []
        for key, label in (("work_date", "작업일자"), ("machine", "트리밍 호기"), ("shift_group", "조"), ("shift", "근무"), ("worker", "작업자")):
            if not common.get(key, "").strip():
                errors.append(missing_common_message(label))
        if not lots:
            errors.append("LOT 1 입력 필요")
        for index, lot in enumerate(lots, start=1):
            for key, label in (("manage_no", "\uad00\ub9ac\ubc88\ud638"), ("round", "차수"), ("lot_no", "LOT No"), ("qty", "매수"), ("condition", "조건"), ("jig", "지그")):
                if not lot.get(key, "").strip():
                    errors.append(f"LOT {index} {label} 입력 필요")
            ok, message = validate_positive_number(lot.get("qty", ""), f"LOT {index} 매수", required=True)
            if not ok:
                errors.append(message)
        if len(lots) == 2:
            if lots[0].get("lot_no", "").strip() == lots[1].get("lot_no", "").strip():
                errors.append("LOT 1 / LOT 2 LOT No 동일")
            if lots[0].get("condition", "").strip() != lots[1].get("condition", "").strip():
                errors.append("LOT 1 / LOT 2 조건 불일치")
            if lots[0].get("jig", "").strip() != lots[1].get("jig", "").strip():
                errors.append("LOT 1 / LOT 2 지그 불일치")
        return len(errors) == 0, errors

    def validate_kcc_hdi_condition_file(self, condition_name: str) -> Path | None:
        self.set_kcc_hdi_status("dnc", "조건 파일 검색중", None)
        source = Path(self.config.get("source_dnc_folders", {}).get("KCC HDI", self.config.get("source_dnc_folder", "")))
        matches = search_condition_file(condition_name, source)
        if not matches:
            show_operator_alert(self.root, "조건 파일 없음", "DNC 파일 없음", "error")
            return None
        if len(matches) >= 2:
            show_operator_alert(self.root, "동일 DNC 파일", format_duplicate_condition_files(matches), "error")
            return None
        return matches[0]

    def run_kcc_hdi_dnc(self) -> None:
        if self.is_running:
            show_operator_alert(self.root, "진행 중", "DNC 실행중")
            return
        if not self.ensure_kcc_hdi_work_period_ready():
            return
        if not self.save_settings_from_ui_silent():
            return
        if not self.validate_kcc_hdi_paths():
            return
        common = self.get_kcc_hdi_common_data()
        lots = self.get_kcc_hdi_used_lots()
        self.update_kcc_hdi_match_status(lots)
        if not self.ensure_kcc_hdi_condition_confirmed(lots):
            return
        ok, errors = self.validate_kcc_hdi_dnc(common, lots)
        if not ok:
            show_operator_alert(self.root, "입력값 확인", format_operator_errors(errors))
            self.set_kcc_hdi_status("dnc", "입력값 NG", False)
            return
        condition_file = self.validate_kcc_hdi_condition_file(lots[0]["condition"])
        if not condition_file:
            self.set_kcc_hdi_status("dnc", "조건 파일 NG", False)
            return
        signature = self.make_kcc_hdi_cycle_signature(common, lots)
        state = self.kcc_hdi_cycle_state
        if state and state.get("remaining", 0) > 0:
            if state.get("signature") != signature:
                show_operator_alert(self.root, "사이클 확인", "진행 중인 LOT 정보와 다릅니다.\n입력값을 확인하세요.", "error")
                self.set_kcc_hdi_status("dnc", "사이클 정보 불일치", False)
                return
            stack = state["stack"]
            cycle_count = int(state["total"])
            cycle_detail = f"남은 {state['remaining']}사이클 / 총 {cycle_count}사이클"
        else:
            model_change = ask_system_yes_no(self.root, "기종교체 확인", "기종교체 입니까?")
            self.frequent_check_values = [""] * 12
            self.work_axis_values = [""] * 6
            machine_axes = get_machine_allowed_axes(common.get("machine", ""))
            if model_change:
                self.set_kcc_hdi_status("dnc", "하부 Pin 확인 대기중", None)
                if not self.open_frequent_check_popup("jig", allowed_axes=machine_axes):
                    self.set_kcc_hdi_status("dnc", "하부 Pin 확인 미완료", False)
                    return
            stack = ask_numeric_input(self.root, "Stack 수 입력", "Stack 수를 입력 하세요.")
            ok, message = validate_positive_number(stack or "", "Stack 수", required=True)
            if not ok:
                show_operator_alert(self.root, "Stack 수 확인", message)
                self.set_kcc_hdi_status("dnc", "Stack 수 확인 필요", False)
                return
            if model_change:
                capacity_values = ["OK" if value == "OK" else "" for value in self.frequent_check_values[6:]] + [""] * 6
                self.work_axis_values = capacity_values[:6]
            else:
                self.set_kcc_hdi_status("dnc", "작업 축 수 확인 대기중", None)
                self.frequent_check_values = [""] * 12
                if not self.open_frequent_check_popup("capacity", allowed_axes=machine_axes):
                    self.set_kcc_hdi_status("dnc", "작업 축 수 확인 미완료", False)
                    return
                capacity_values = self.frequent_check_values[:]
                self.work_axis_values = capacity_values[:6]
            try:
                cycle_count, cycle_detail = self.calculate_kcc_hdi_cycle_count(common, lots, stack, capacity_values)
            except Exception as exc:
                show_operator_alert(self.root, "사이클 확인", str(exc), "error")
                self.set_kcc_hdi_status("dnc", "사이클 확인 필요", False)
                return
            self.kcc_hdi_cycle_state = {
                "signature": signature,
                "stack": stack,
                "total": cycle_count,
                "remaining": cycle_count,
                "log_ids": None,
                "model_change": model_change,
                "work_axis_values": self.work_axis_values[:],
                "jig_axis_values": self.frequent_check_values[6:],
                "capacity_values": capacity_values[:],
                "first_check_done": False,
            }
        self.update_kcc_hdi_cycle_status(cycle_count, int(self.kcc_hdi_cycle_state["remaining"]))
        self.set_kcc_hdi_status("dnc", cycle_detail, True)
        self.set_running(True)
        threading.Thread(target=self.kcc_hdi_worker, args=(common, lots, stack, condition_file), daemon=True).start()

    def kcc_hdi_worker(self, common: dict, lots: list[dict], stack: str, condition_file: Path) -> None:
        try:
            state = self.kcc_hdi_cycle_state or {}
            total = int(state.get("total", 1))
            remaining_before = int(state.get("remaining", total))
            cycle_index = total - remaining_before + 1
            if total > 1 and remaining_before == 1:
                self.root.after(0, self.set_kcc_hdi_last_cycle_running)
            if not state.get("log_ids"):
                self.root.after(0, lambda: self.set_kcc_hdi_status("dnc", "DB 저장중", None))
                log_ids = insert_kcc_hdi_dnc_db(common, lots, stack, bool(state.get("model_change")))
                state["log_ids"] = log_ids
                self.kcc_hdi_cycle_state = state
            else:
                log_ids = list(state["log_ids"])
            transfer_folder = Path(self.config["transfer_dnc_folder"])
            self.root.after(0, lambda i=cycle_index, c=total: self.set_kcc_hdi_status("dnc", f"DNC 사이클 {i}/{c} 진행중", None))
            delete_existing_dnc_txt(transfer_folder)
            copied_file = copy_dnc_file(condition_file, transfer_folder)
            self.root.after(0, lambda i=cycle_index, c=total: self.set_kcc_hdi_status("dnc", f"DNC 파일 복사 완료 ({i}/{c})", True))
            delete_thread = threading.Thread(
                target=delete_after_delay,
                args=(copied_file, int(self.config["dnc_delete_seconds"]), lambda text, i=cycle_index, c=total: self.root.after(0, lambda t=text: self.set_kcc_hdi_status("dnc", f"{t} / {i}/{c}", None))),
                daemon=True,
            )
            delete_thread.start()
            if not state.get("first_check_done"):
                wait_seconds = int(self.config.get("first_article_wait_seconds", FIRST_ARTICLE_WAIT_SECONDS))
                for remain in range(wait_seconds, 0, -1):
                    self.root.after(0, lambda r=remain: self.set_kcc_hdi_status("dnc", f"초품 확인 대기중 ({r}초)", None))
                    time.sleep(1)
            self.root.after(0, lambda ids=log_ids, lt=lots, st=stack, dt=delete_thread: self.finish_kcc_hdi_cycle_dnc(ids, lt, st, dt))
        except Exception as exc:
            self.root.after(0, lambda error=exc: self.handle_kcc_hdi_run_error(error))

    def finish_kcc_hdi_cycle_dnc(self, log_ids: list[int], lots: list[dict], stack: str, delete_thread: threading.Thread | None = None) -> None:
        try:
            state = self.kcc_hdi_cycle_state or {}
            model_change = bool(state.get("model_change"))
            if not state.get("first_check_done"):
                while True:
                    if model_change:
                        allowed_axes = [
                            index
                            for index, value in enumerate(state.get("jig_axis_values", []))
                            if value == "OK"
                        ]
                        self.frequent_check_values = [""] * 6 + list(state.get("jig_axis_values", [""] * 6))
                    else:
                        allowed_axes = [
                            index
                            for index, value in enumerate(state.get("work_axis_values", []))
                            if value == "OK"
                        ]
                        self.frequent_check_values = [""] * 12
                    if not self.open_frequent_check_popup("first", allowed_axes=allowed_axes):
                        self.set_kcc_hdi_status("dnc", "초품 확인 미완료", False)
                        if delete_thread and delete_thread.is_alive():
                            def wait_and_release_after_first_cancel() -> None:
                                delete_thread.join()
                                self.root.after(0, lambda: self.set_running(False))

                            threading.Thread(target=wait_and_release_after_first_cancel, daemon=True).start()
                        else:
                            self.set_running(False)
                        return
                    if model_change:
                        first_axes = [
                            index
                            for index, value in enumerate(self.frequent_check_values[:6])
                            if value == "OK"
                        ]
                        if first_axes != allowed_axes:
                            show_operator_alert(self.root, "초품 4Point 확인", "하부 Pin 축과 초품 축 다름")
                            self.set_kcc_hdi_status("dnc", "초품 축 확인 NG", False)
                            continue
                    if count_frequent_check_axes(self.frequent_check_values) > 0:
                        self.set_kcc_hdi_status("dnc", "초품 확인 완료", True)
                        break
                    show_operator_alert(self.root, "초품 확인", "확인 축 선택 필요")
                    self.set_kcc_hdi_status("dnc", "초품 확인 NG", False)
                update_normal_frequent_check_db(log_ids, model_change, self.frequent_check_values, "KCC HDI")
                state["first_check_done"] = True
            state["remaining"] = max(0, int(state.get("remaining", 1)) - 1)
            self.kcc_hdi_cycle_state = state
            if state["remaining"] > 0:
                self.hide_kcc_hdi_condition_cards()
                self.update_kcc_hdi_cycle_status(int(state.get("total", 1)), int(state["remaining"]))
                if delete_thread and delete_thread.is_alive():
                    self.set_kcc_hdi_status("dnc", "DNC 삭제 완료 대기중", None)

                    def wait_and_release() -> None:
                        delete_thread.join()
                        self.root.after(0, lambda r=state["remaining"]: self.set_kcc_hdi_status("dnc", f"이번 사이클 완료 / 남은 {r}회", True))
                        self.root.after(0, lambda: self.set_running(False))

                    threading.Thread(target=wait_and_release, daemon=True).start()
                    return
                self.set_kcc_hdi_status("dnc", f"이번 사이클 완료 / 남은 {state['remaining']}회", True)
                self.set_running(False)
                return
            if delete_thread and delete_thread.is_alive():
                self.set_kcc_hdi_status("dnc", "DNC 완료 대기중", None)

                def wait_and_finish() -> None:
                    delete_thread.join()
                    self.root.after(0, lambda: self.finish_kcc_hdi_dnc(log_ids))

                threading.Thread(target=wait_and_finish, daemon=True).start()
                return
            self.finish_kcc_hdi_dnc(log_ids)
        except Exception as exc:
            self.handle_kcc_hdi_run_error(exc)

    def finish_kcc_hdi_dnc(self, log_ids: list[int]) -> None:
        try:
            self.hide_kcc_hdi_cycle_status()
            self.hide_kcc_hdi_condition_cards()
            burr_ok = ask_system_yes_no(self.root, "Burr 확인", "4면 Burr 이상 없습니까?")
            update_normal_burr_db(log_ids, burr_ok, "KCC HDI")
            pending = get_unexported_process_log_count("KCC HDI")
            self.set_kcc_hdi_status("dnc", "DNC 완료", True)
            self.set_kcc_hdi_status("excel", f"DB 저장 완료 / Excel 미반영 {pending}건", True)
            self.auto_export_kcc_hdi_to_excel(parent=self.root)
            self.kcc_hdi_cycle_state = None
            self.clear_kcc_hdi_inputs(after_done=True)
        except Exception as exc:
            self.handle_kcc_hdi_run_error(exc)
        finally:
            self.set_running(False)

    def auto_export_kcc_hdi_to_excel(self, parent=None) -> bool:
        try:
            export_process_logs_to_excel(self.config, "KCC HDI", "KCC HDI")
            pending = get_unexported_process_log_count("KCC HDI")
            self.set_kcc_hdi_status("excel", f"작업일보 반영 완료 / Excel 미반영 {pending}건", True)
            return True
        except Exception as exc:
            log_error("작업일보 자동 반영 실패", exc)
            self.set_kcc_hdi_status("excel", f"Excel 미반영 {get_unexported_process_log_count('KCC HDI')}건", False)
            return False

    def export_kcc_hdi_to_excel_from_ui(self) -> None:
        try:
            count = export_process_logs_to_excel(self.config, "KCC HDI", "KCC HDI")
        except Exception as exc:
            log_error("KCC HDI 작업일보 반영 실패", exc)
            show_operator_alert(self.root, "작업일보 반영 실패", format_excel_error_for_operator(exc, "작업일보 Excel"), "error")
            self.set_kcc_hdi_status("excel", "작업일보 반영 실패", False)
            return
        pending = get_unexported_process_log_count("KCC HDI")
        show_operator_alert(self.root, "작업일보 반영", f"{count}건 반영 완료", "info")
        self.set_kcc_hdi_status("excel", f"Excel 미반영 {pending}건", True)

    def clear_kcc_hdi_inputs(self, after_done: bool = False) -> None:
        self.kcc_hdi_cycle_state = None
        self.kcc_hdi_condition_records.clear()
        for entries in (self.kcc_hdi_entries, self.kcc_hdi_lot2_entries):
            for entry in entries.values():
                entry.clear()
        self.hide_kcc_hdi_condition_cards()
        self.hide_kcc_hdi_cycle_status()
        self.set_kcc_hdi_run_gradient(False)
        self.set_kcc_hdi_status("dnc", "DNC 완료" if after_done else "대기중", True if after_done else None)



    def handle_kcc_hdi_run_error(self, exc: Exception) -> None:
        log_error("KCC HDI DNC run error", exc)
        show_operator_alert(self.root, "\uc624\ub958", str(exc), "error")
        self.set_kcc_hdi_status("dnc", "\uc624\ub958", False)
        self.set_running(False)

    def create_kcc_pkg_tab(self) -> None:
        kcc_theme = PROCESS_COLORS["KCC PKG"]
        kcc_bg = kcc_theme["bg"]
        kcc_light = kcc_theme["light"]
        kcc_primary = kcc_theme["primary"]
        kcc_border = kcc_theme["border"]
        self.kcc_pkg_page.configure(bg=kcc_bg)
        self.kcc_pkg_page.columnconfigure(0, weight=1)
        self.kcc_pkg_page.rowconfigure(2, weight=1)

        title_wrap = tk.Frame(self.kcc_pkg_page, bg=kcc_light)
        title_wrap.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 8))
        title_wrap.columnconfigure(0, minsize=260)
        title_wrap.columnconfigure(1, weight=1)
        title_wrap.columnconfigure(2, minsize=360)
        logo_slot = tk.Frame(title_wrap, bg=kcc_light, width=220, height=1)
        logo_slot.grid(row=0, column=0, sticky="w", padx=(14, 8))
        logo_slot.grid_propagate(False)
        kcc_logo_path = BUNDLED_KCC_LOGO_FILE if BUNDLED_KCC_LOGO_FILE.exists() else KCC_LOGO_FILE
        if kcc_logo_path.exists():
            try:
                kcc_logo = tk.PhotoImage(file=str(kcc_logo_path))
                scale = max(1, kcc_logo.height() // 32)
                self.kcc_logo_image = kcc_logo.subsample(scale, scale)
                tk.Label(logo_slot, image=self.kcc_logo_image, bg=kcc_light, bd=0).pack(side=tk.LEFT, anchor="w")
            except tk.TclError:
                self.kcc_logo_image = None
        title = tk.Label(
            title_wrap,
            text="KCC PKG DNC",
            bg=kcc_light,
            fg=kcc_primary,
            font=("맑은 고딕", 14, "bold"),
            height=2,
        )
        title.grid(row=0, column=1, sticky="ew")
        title_buttons = tk.Frame(title_wrap, bg=kcc_light)
        title_buttons.grid(row=0, column=2, sticky="e", padx=(8, 10))
        self.kcc_run_button = self.create_tlb_gradient_run_button(title_buttons, command=self.run_normal_dnc, scheme="kcc")
        self.kcc_run_button.grid(row=0, column=0, padx=4, pady=4)
        self.add_normal_button(title_buttons, "입력 초기화", self.clear_normal_inputs).grid(row=0, column=1, padx=4, pady=4)

        title_legacy = tk.Label(
            self.kcc_pkg_page,
            text="KCC PKG DNC",
            bg=kcc_light,
            fg=kcc_primary,
            font=("맑은 고딕", 14, "bold"),
            height=2,
        )

        common = self.create_panel(self.kcc_pkg_page, "공통 입력", kcc_theme)
        common.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 8))
        common_widgets = [
            ("machine", ComboField(common, "설비 호기", ["트리밍 1호기", "트리밍 2호기", "트리밍 3호기"], initial=self.config.get("machine", "트리밍 1호기"), width=12)),
            ("work_date", DateField(common, "작업일자", on_change=lambda key="work_date": self.handle_common_change(key, "kcc"))),
            ("shift_group", SegmentedField(common, "조", ["A", "B", "C"], allow_empty=True, on_change=lambda key="shift_group": self.handle_common_change(key, "kcc"))),
            ("shift", SegmentedField(common, "근무", ["주간", "야간"], on_change=lambda key="shift": self.handle_common_change(key, "kcc"))),
            ("worker", LabeledEntry(common, "작업자", width=12, on_change=lambda key="worker": self.handle_common_change(key, "kcc"), live_change=False)),
        ]
        for index, (key, entry) in enumerate(common_widgets):
            entry.grid(row=1, column=index, sticky="ew", padx=8, pady=8)
            self.common_entries[key] = entry
            if key == "machine":
                entry.combo.bind("<<ComboboxSelected>>", lambda _event, field=key: self.handle_common_change(field, "kcc"))
        common.columnconfigure(0, weight=0, minsize=320)
        common.columnconfigure(1, weight=1, minsize=430)
        common.columnconfigure(2, weight=1, minsize=330)
        common.columnconfigure(3, weight=1, minsize=360)
        common.columnconfigure(4, weight=0, minsize=320)

        lots = tk.Frame(self.kcc_pkg_page, bg=kcc_bg)
        lots.grid(row=2, column=0, sticky="nsew", padx=14, pady=0)
        lots.columnconfigure(0, weight=1)
        lots.columnconfigure(1, weight=1)

        lot1 = self.create_lot_panel(lots, "LOT 1 입력", self.lot1_entries, 1, kcc_theme)
        lot1.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=0)
        lot2 = self.create_lot_panel(lots, "LOT 2 입력 (선택)", self.lot2_entries, 2, kcc_theme)
        lot2.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=0)

        bottom = tk.Frame(self.kcc_pkg_page, bg=kcc_bg)
        bottom.grid(row=3, column=0, sticky="ew", padx=14, pady=(8, 14))
        bottom.columnconfigure(0, weight=1)

        status_panel = tk.Frame(bottom, bg=SURFACE_BG, highlightthickness=1, highlightbackground=kcc_border, bd=0)
        status_panel.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        status_panel.columnconfigure(1, weight=1)
        status_panel.columnconfigure(2, weight=0)
        tk.Label(status_panel, text="2LOT 조건 일치 확인", bg=kcc_light, fg=kcc_primary, font=("맑은 고딕", 11, "bold"), width=22, height=2).grid(row=0, column=0, sticky="nsw")
        self.lot_match_frame = tk.Frame(status_panel, bg=SURFACE_BG)
        self.lot_match_frame.grid(row=0, column=1, sticky="ew", padx=14)
        self.set_lot_match_segments([("LOT 2 미사용", MUTED_TEXT)])
        tk.Label(status_panel, text="DNC 진행 상태", bg=kcc_light, fg=kcc_primary, font=("맑은 고딕", 11, "bold"), width=22, height=2).grid(row=1, column=0, sticky="nsw")
        dnc_label = tk.Label(status_panel, text="대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 12, "bold"), anchor="w")
        dnc_label.grid(row=1, column=1, columnspan=2, sticky="ew", padx=14)
        self.status_labels["dnc"] = dnc_label
        tk.Label(status_panel, text="작업일보 반영", bg=kcc_light, fg=kcc_primary, font=("맑은 고딕", 11, "bold"), width=22, height=2).grid(row=2, column=0, sticky="nsw")
        excel_label = tk.Label(status_panel, text="대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 12, "bold"), anchor="w")
        excel_label.grid(row=2, column=1, columnspan=2, sticky="ew", padx=14)
        self.status_labels["excel"] = excel_label

        log_panel = tk.Frame(bottom, bg=SURFACE_BG, highlightthickness=1, highlightbackground=kcc_border, bd=0)
        log_panel.grid(row=1, column=0, sticky="ew", padx=(0, 10), pady=(8, 0))
        log_panel.columnconfigure(0, weight=1)
        tk.Label(log_panel, text="DNC 작업 로그", bg=kcc_light, fg=kcc_primary, font=("맑은 고딕", 10, "bold"), height=1).grid(row=0, column=0, sticky="ew")
        self.log_text = scrolledtext.ScrolledText(
            log_panel,
            height=5,
            wrap=tk.WORD,
            state="disabled",
            bg=SURFACE_BG,
            fg=TEXT_COLOR,
            font=("맑은 고딕", 10),
            relief=tk.FLAT,
            padx=10,
            pady=8,
        )
        self.log_text.grid(row=1, column=0, sticky="ew")

        button_panel = tk.Frame(bottom, bg=kcc_bg)
        button_panel.grid(row=0, column=1, rowspan=2, sticky="ne")
        for column in range(3):
            button_panel.columnconfigure(column, weight=1, uniform="side_buttons")
        self.add_side_button(button_panel, "작업일보 마스터 갱신", self.rebuild_condition_master, "SidePrimary.TButton").grid(row=0, column=0, sticky="nsew", padx=4, pady=4)
        self.new_model_button = self.add_side_button(button_panel, "신규 모델 검증 DNC", self.open_new_model_popup, "SidePrimary.TButton")
        self.new_model_button.grid(row=0, column=1, sticky="nsew", padx=4, pady=4)
        self.add_side_button(button_panel, "조건 마스터 관리", self.open_condition_master_popup, "SideDanger.TButton").grid(row=0, column=2, sticky="nsew", padx=4, pady=4)
        self.add_side_button(button_panel, "작업일보 반영", self.export_kcc_pkg_to_excel_from_ui).grid(row=1, column=0, sticky="nsew", padx=4, pady=4)
        self.add_side_button(button_panel, "작업일보 열기", self.open_log_excel_from_ui).grid(row=1, column=1, sticky="nsew", padx=4, pady=4)


        for entry in list(self.lot1_entries.values()) + list(self.lot2_entries.values()):
            entry.var.trace_add("write", lambda *_args: self.update_status_checks())
        for lot_number, entries in ((1, self.lot1_entries), (2, self.lot2_entries)):
            for key in ("step", "round", "manage_no", "process_code"):
                entries[key].var.trace_add("write", lambda *_args, n=lot_number: self.handle_lot_key_change(n))

    def create_panel(self, parent, title: str, process_theme: dict | None = None) -> tk.Frame:
        if process_theme is None:
            process_theme = {"light": PRIMARY_LIGHT, "primary": PRIMARY, "border": "#93c5fd"}
        panel = tk.Frame(parent, bg=SURFACE_BG, highlightthickness=1, highlightbackground=process_theme.get("border", "#93c5fd"), bd=0)
        tk.Label(
            panel,
            text=title,
            bg=process_theme.get("light", PRIMARY_LIGHT),
            fg=process_theme.get("primary", PRIMARY),
            font=("맑은 고딕", 10, "bold"),
            height=2,
        ).grid(row=0, column=0, columnspan=8, sticky="ew")
        return panel

    def create_lot_panel(self, parent, title: str, target: dict[str, LabeledEntry], lot_number: int, process_theme: dict | None = None) -> tk.Frame:
        panel = self.create_panel(parent, title, process_theme)
        fields = [
            ("step", "STEP"),
            ("round", "차수"),
            ("manage_no", "관리번호"),
            ("lot_no", "LOT No"),
            ("qty", "매수"),
            ("process_code", "공정코드"),
            ("condition", "조건(조회)"),
            ("jig", "지그(조회)"),
        ]
        for index, (key, label) in enumerate(fields):
            row = index // 2 + 1
            col = index % 2
            if key == "round":
                entry = RoundField(panel, label)
            elif key in {"condition", "jig"}:
                entry = LabeledEntry(panel, label, width=24, style="Lookup.TEntry", readonly=True)
            elif key in {"step", "qty"}:
                entry = LabeledEntry(panel, label, width=24, numeric_only=True)
            else:
                entry = LabeledEntry(panel, label, width=24, uppercase=True)
            entry.grid(row=row, column=col, sticky="ew", padx=10, pady=8)
            panel.columnconfigure(col, weight=1)
            target[key] = entry
        load_button = self.create_mes_lookup_button(
            panel,
            command=lambda: self.load_condition_jig_for_lot(lot_number),
            scheme="kcc",
            ready_check=lambda lot_number=lot_number: self.is_kcc_lookup_ready(lot_number),
        )
        load_button.grid(row=5, column=0, columnspan=2, sticky="ew", padx=10, pady=(8, 6))

        status = tk.Frame(panel, bg=SURFACE_BG)
        status.grid(row=6, column=0, columnspan=2, sticky="ew", padx=10, pady=(0, 12))
        status.columnconfigure(0, weight=1, uniform=f"lot{lot_number}_status")
        status.columnconfigure(1, weight=1, uniform=f"lot{lot_number}_status")
        mes_label = self.create_judgement_card(status, "MES Core")
        mes_label.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        condition_label = self.create_judgement_card(status, "조건 적용")
        condition_label.grid(row=0, column=1, sticky="ew", padx=(6, 0))
        self.lot_status_labels[f"lot{lot_number}_mes"] = mes_label
        self.lot_status_labels[f"lot{lot_number}_condition"] = condition_label
        return panel

    def create_judgement_card(self, parent, title: str) -> tk.Label:
        """LOT별 핵심 판정 상태를 크게 보여주는 카드형 라벨입니다."""
        return tk.Label(
            parent,
            text=f"{title}\n대기중",
            bg="#f8fafc",
            fg=MUTED_TEXT,
            font=("맑은 고딕", 12, "bold"),
            height=3,
            relief=tk.SOLID,
            bd=1,
            justify=tk.CENTER,
            anchor="center",
        )

    def hide_judgement_card(self, label: tk.Label) -> None:
        """판정 전에는 공간만 유지하고 카드 테두리/문구를 숨깁니다."""
        label.configure(
            text="",
            fg=SURFACE_BG,
            bg=SURFACE_BG,
            relief=tk.FLAT,
            bd=0,
            highlightthickness=0,
        )

    def show_judgement_card(self, label: tk.Label, title: str, text: str, ok: bool | None) -> None:
        """KCC/TLB 판정 카드 스타일을 한 가지 기준으로 맞춥니다."""
        if ok is True:
            bg = "#dcfce7"
            fg = OK_COLOR
            border = OK_COLOR
        elif ok is False:
            bg = "#fee2e2"
            fg = NG_COLOR
            border = NG_COLOR
        else:
            bg = "#f8fafc"
            fg = MUTED_TEXT
            border = BORDER_COLOR
        label.configure(
            text=f"{title}\n{text}",
            fg=fg,
            bg=bg,
            relief=tk.SOLID,
            bd=1,
            highlightthickness=2,
            highlightbackground=border,
        )

    def add_normal_button(self, parent, text: str, command, style: str = "TButton") -> ttk.Button:
        button = ttk.Button(parent, text=text, command=command, style=style, width=18)
        self.normal_buttons.append(button)
        return button

    def add_side_button(self, parent, text: str, command, style: str = "Side.TButton") -> ttk.Button:
        button = ttk.Button(parent, text=text, command=command, style=style, width=18)
        self.normal_buttons.append(button)
        return button


    def create_tlb_gradient_run_button(self, parent, command=None, scheme: str = "tlb") -> tk.Canvas:
        """조건 OK 후 톱니와 외곽 테두리가 도는 DNC 실행 버튼입니다."""
        width = 188
        height = 44
        if command is None:
            command = self.run_tlb_dnc
        process_theme = PROCESS_COLORS["KCC PKG"] if scheme == "kcc" else PROCESS_COLORS["TLB"]
        surface = process_theme["light"]
        surface_hover = process_theme["bg"]
        active_left = process_theme["light"]
        active_right = process_theme["bg"]
        border_color = process_theme["primary"]
        text_color_active = process_theme["primary"]
        parent_bg = process_theme["light"]
        canvas = tk.Canvas(parent, width=width, height=height, bd=0, highlightthickness=0, bg=parent_bg, cursor="hand2")
        canvas._run_state = "normal"
        canvas._gradient_on = False
        canvas._gradient_phase = 0
        canvas._hover = False
        canvas._pressed = False
        canvas._photo = None
        canvas._gear_photo = None
        canvas._gear_source = None

        def hex_to_rgb(color: str) -> tuple[int, int, int]:
            color = color.lstrip("#")
            return int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)

        def blend(left: tuple[int, int, int], right: tuple[int, int, int], ratio: float) -> tuple[int, int, int]:
            ratio = max(0.0, min(1.0, ratio))
            return tuple(int(left[i] + (right[i] - left[i]) * ratio) for i in range(3))

        def make_image(fill_left: str, fill_right: str, border: str, disabled: bool, pressed: bool) -> ImageTk.PhotoImage:
            scale = 3
            image = Image.new("RGBA", (width * scale, height * scale), (0, 0, 0, 0))
            draw_img = ImageDraw.Draw(image)
            offset = scale if pressed else 0
            x1, y1 = 2 * scale + offset, 2 * scale + offset
            x2, y2 = (width - 3) * scale + offset, (height - 4) * scale + offset
            radius = 4 * scale
            left_rgb = hex_to_rgb(fill_left)
            right_rgb = hex_to_rgb(fill_right)
            border_rgb = hex_to_rgb(border)

            draw_img.rounded_rectangle((x1 + 4 * scale, y1 + 5 * scale, x2 + 4 * scale, y2 + 5 * scale), radius=radius, fill=(80, 92, 110, 32))
            mask = Image.new("L", image.size, 0)
            mask_draw = ImageDraw.Draw(mask)
            mask_draw.rounded_rectangle((x1, y1, x2, y2), radius=radius, fill=255)

            gradient = Image.new("RGBA", image.size, (0, 0, 0, 0))
            grad_draw = ImageDraw.Draw(gradient)
            sweep_center = x1 + int(((canvas._gradient_phase % 120) / 119) * max(1, (x2 - x1)))
            sweep_width = 54 * scale
            for px in range(x1, max(x1 + 1, x2 + 1)):
                ratio = (px - x1) / max(1, (x2 - x1))
                rgb = blend(left_rgb, right_rgb, ratio)
                if canvas._gradient_on and not disabled:
                    distance = abs(px - sweep_center)
                    if distance < sweep_width:
                        shine = 1 - (distance / sweep_width)
                        rgb = blend(rgb, (255, 255, 255), 0.30 * shine)
                alpha = 150 if disabled else 245
                grad_draw.line((px, y1, px, y2), fill=(*rgb, alpha), width=1)
            image.alpha_composite(Image.composite(gradient, Image.new("RGBA", image.size, (0, 0, 0, 0)), mask))
            draw_img.rounded_rectangle((x1, y1, x2, y2), radius=radius, outline=(*border_rgb, 235 if not disabled else 110), width=scale)
            draw_img.line((x1 + 9 * scale, y1 + 6 * scale, x2 - 9 * scale, y1 + 6 * scale), fill=(255, 255, 255, 130 if not disabled else 55), width=scale)
            draw_img.line((x1 + 9 * scale, y2 - 6 * scale, x2 - 9 * scale, y2 - 6 * scale), fill=(75, 165, 160, 55 if canvas._gradient_on and not disabled else 30), width=scale)

            if canvas._gradient_on and not disabled:
                math = __import__("math")
                straight = max((x2 - x1) - (2 * radius), 1)
                arc_len = math.pi * radius
                perimeter = (straight * 2) + (arc_len * 2)
                pos = (canvas._gradient_phase % 180) / 180 * perimeter

                def point_at(distance: float) -> tuple[float, float]:
                    d = distance % perimeter
                    if d <= straight:
                        return x1 + radius + d, y1
                    d -= straight
                    if d <= arc_len:
                        angle = -90 + (d / arc_len) * 180
                        rad = math.radians(angle)
                        return x2 - radius + radius * math.cos(rad), y1 + radius + radius * math.sin(rad)
                    d -= arc_len
                    if d <= straight:
                        return x2 - radius - d, y2
                    d -= straight
                    angle = 90 + (d / arc_len) * 180
                    rad = math.radians(angle)
                    return x1 + radius + radius * math.cos(rad), y1 + radius + radius * math.sin(rad)

                def draw_runner(distance: float) -> None:
                    trail_alpha = (85, 140, 205) if scheme == "kcc" else (65, 105, 155)
                    for index, gap in enumerate((21, 13, 6)):
                        tx, ty = point_at(distance - gap * scale)
                        size = (1.05 + index * 0.32) * scale
                        draw_img.ellipse((tx - size, ty - size, tx + size, ty + size), fill=(255, 255, 255, trail_alpha[index]))
                    sx, sy = point_at(distance)
                    head_size = (2.75 if scheme == "kcc" else 2.35) * scale
                    outline_alpha = 230 if scheme == "kcc" else 190
                    draw_img.ellipse(
                        (sx - head_size, sy - head_size, sx + head_size, sy + head_size),
                        fill=(255, 255, 255, 245 if scheme == "kcc" else 220),
                        outline=(*border_rgb, outline_alpha),
                        width=scale,
                    )

                draw_runner(pos)
                draw_runner(pos + perimeter / 2)

            try:
                resample = Image.Resampling.LANCZOS
            except AttributeError:
                resample = Image.LANCZOS
            return ImageTk.PhotoImage(image.resize((width, height), resample))

        def draw() -> None:
            canvas.delete("all")
            disabled = canvas._run_state != "normal"
            if canvas._gradient_on:
                left = active_left if not canvas._hover else surface_hover
                right = active_right if not canvas._hover else active_right
                border = border_color
                text_color = text_color_active
            else:
                left = surface if not canvas._hover else surface_hover
                right = surface if not canvas._hover else surface_hover
                border = "#9fb4d0"
                text_color = text_color_active
            if disabled:
                text_color = MUTED_TEXT
            canvas._photo = make_image(left, right, border, disabled, canvas._pressed)
            canvas.create_image(0, 0, image=canvas._photo, anchor="nw")
            text_x = width / 2
            if canvas._gradient_on and not disabled:
                if canvas._gear_source is None:
                    gear_bytes = base64.b64decode(TLB_LOOKUP_GEAR_PNG_BASE64)
                    gear_image = Image.open(io.BytesIO(gear_bytes)).convert("RGBA")
                    bbox = gear_image.getchannel("A").getbbox()
                    if bbox:
                        gear_image = gear_image.crop(bbox)
                    canvas._gear_source = gear_image
                try:
                    rotate_resample = Image.Resampling.BICUBIC
                    resize_resample = Image.Resampling.LANCZOS
                except AttributeError:
                    rotate_resample = Image.BICUBIC
                    resize_resample = Image.LANCZOS
                gear_size = 15
                angle = -(canvas._gradient_phase / 180.0) * 360
                gear = canvas._gear_source.resize((gear_size, gear_size), resize_resample)
                gear = gear.rotate(angle, resample=rotate_resample, expand=True)
                gear_rgb = hex_to_rgb(text_color_active)
                alpha = gear.getchannel("A").point(lambda value: int(value * 0.72))
                tinted = Image.new("RGBA", gear.size, (*gear_rgb, 0))
                tinted.putalpha(alpha)
                canvas._gear_photo = ImageTk.PhotoImage(tinted)
                canvas.create_image(width / 2 - 47, height / 2 + (1 if canvas._pressed else 0), image=canvas._gear_photo)
                text_x = width / 2 + 8
            canvas.create_text(text_x, height / 2 + (1 if canvas._pressed else 0), text="DNC 실행", fill=text_color, font=("맑은 고딕", 11, "bold"))

        def animate() -> None:
            if not canvas.winfo_exists():
                return
            if canvas._gradient_on and canvas._run_state == "normal":
                canvas._gradient_phase = (canvas._gradient_phase + 2) % 180
                draw()
                canvas.after(70, animate)
            else:
                canvas.after(350, animate)

        def set_gradient(enabled: bool) -> None:
            canvas._gradient_on = bool(enabled)
            if not enabled:
                canvas._gradient_phase = 0
            draw()

        def set_state(state: str) -> None:
            canvas._run_state = state
            canvas._pressed = False
            original_configure(cursor="hand2" if state == "normal" else "arrow")
            draw()

        original_configure = canvas.configure

        def configure_proxy(*args, **kwargs):
            if "state" in kwargs:
                set_state(kwargs.pop("state"))
            if kwargs:
                return original_configure(*args, **kwargs)
            if args:
                return original_configure(*args)
            return None

        def on_enter(_event) -> None:
            if canvas._run_state == "normal":
                canvas._hover = True
                draw()

        def on_leave(_event) -> None:
            canvas._hover = False
            canvas._pressed = False
            draw()

        def on_press(_event) -> None:
            if canvas._run_state == "normal":
                canvas._pressed = True
                draw()

        def on_release(_event) -> None:
            if canvas._run_state == "normal":
                canvas._pressed = False
                draw()
                canvas.after(50, command)

        canvas.configure = configure_proxy
        canvas._set_gradient = set_gradient
        canvas.bind("<Enter>", on_enter)
        canvas.bind("<Leave>", on_leave)
        canvas.bind("<ButtonPress-1>", on_press)
        canvas.bind("<ButtonRelease-1>", on_release)
        draw()
        canvas.after(600, animate)
        self.normal_buttons.append(canvas)
        return canvas

    def set_tlb_run_gradient(self, enabled: bool) -> None:
        button = getattr(self, "tlb_run_button", None)
        if button is not None and hasattr(button, "_set_gradient"):
            button._set_gradient(enabled)

    def set_kcc_hdi_run_gradient(self, enabled: bool) -> None:
        button = getattr(self, "kcc_hdi_run_button", None)
        if button is not None and hasattr(button, "_set_gradient"):
            button._set_gradient(enabled)

    def set_kcc_run_animation(self, enabled: bool) -> None:
        button = getattr(self, "kcc_run_button", None)
        if button is not None and hasattr(button, "_set_gradient"):
            button._set_gradient(enabled)

    def is_kcc_run_ready_by_condition(self) -> bool:
        lot1 = self.get_lot_data(self.lot1_entries)
        lot2 = self.get_lot_data(self.lot2_entries)
        lot2_used = self.lot_has_any_value(lot2)

        lot1_mes_ok, _ = get_mes_core_message(lot1.get("lot_no", ""), lot1.get("process_code", ""))
        lot1_condition_ok, _ = get_single_condition_message(lot1)
        if not (lot1_mes_ok and lot1_condition_ok and self.is_kcc_lot_condition_confirmed(1, lot1)):
            return False

        if not lot2_used:
            return True

        lot2_mes_ok, _ = get_mes_core_message(lot2.get("lot_no", ""), lot2.get("process_code", ""))
        lot2_condition_ok, _ = get_single_condition_message(lot2)
        if not (lot2_mes_ok and lot2_condition_ok and self.is_kcc_lot_condition_confirmed(2, lot2)):
            return False

        lot1_lot_no = lot1.get("lot_no", "").strip()
        lot2_lot_no = lot2.get("lot_no", "").strip()
        if lot1_lot_no and lot1_lot_no == lot2_lot_no:
            return False
        return (
            lot1.get("condition", "").strip() == lot2.get("condition", "").strip()
            and lot1.get("jig", "").strip() == lot2.get("jig", "").strip()
        )

    def is_kcc_lot_condition_confirmed(self, lot_number: int, lot: dict) -> bool:
        """조건/지그 조회 버튼으로 확정된 LOT인지 확인합니다."""
        loaded_key = self.lot_condition_keys.get(lot_number, "")
        if not loaded_key:
            return False
        if loaded_key != self.make_lot_lookup_key(lot):
            return False
        return bool(lot.get("condition", "").strip() and lot.get("jig", "").strip())

    def ensure_kcc_condition_confirmed(self, lot1: dict, lot2: dict | None) -> bool:
        missing_lots = []
        if not self.is_kcc_lot_condition_confirmed(1, lot1):
            missing_lots.append("LOT 1")
        if lot2 and not self.is_kcc_lot_condition_confirmed(2, lot2):
            missing_lots.append("LOT 2")
        if missing_lots:
            show_operator_alert(self.root, "DNC 조건 조회", "DNC 조건 조회 필요")
            self.set_status("dnc", "DNC 조건 조회 필요", False)
            return False
        return True

    def create_mes_lookup_button(self, parent, command, scheme: str = "kcc", ready_check=None) -> tk.Canvas:
        """Smooth image-rendered lookup button."""
        process_theme = PROCESS_COLORS["KCC PKG"] if scheme == "kcc" else PROCESS_COLORS["TLB"]
        palette = {
            "surface": process_theme["light"],
            "surface_hover": process_theme["bg"],
            "surface_active": process_theme["light"],
            "fg": process_theme["primary"],
            "border": process_theme["primary"],
            "glow": process_theme["border"],
            "shine": "#ffffff",
        }
        surface = palette.get("surface", PRIMARY_LIGHT)
        hover = palette.get("surface_hover", PRIMARY_LIGHT)
        active = palette.get("surface_active", PRIMARY_LIGHT)
        fg = palette.get("fg", PRIMARY)
        border = palette.get("border", PRIMARY)
        glow = palette.get("glow", BORDER_COLOR)
        shine = palette.get("shine", "#ffffff")

        canvas = tk.Canvas(
            parent,
            height=62,
            bd=0,
            highlightthickness=0,
            bg=SURFACE_BG,
            cursor="hand2",
        )
        canvas._lookup_state = "normal"
        canvas._lookup_bg = surface
        canvas._lookup_anim = 0
        canvas._lookup_animating = False
        canvas._lookup_photo = None

        def hex_to_rgb(color: str) -> tuple[int, int, int]:
            color = color.lstrip("#")
            return int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)

        def make_button_image(width: int, height: int, fill: str, pressed: bool) -> ImageTk.PhotoImage:
            scale = 4
            image = Image.new("RGBA", (width * scale, height * scale), (0, 0, 0, 0))
            draw_img = ImageDraw.Draw(image)
            inset = 34 * scale
            y_pad = 6 * scale
            offset = scale if pressed else 0
            x1 = inset + offset
            y1 = y_pad + offset
            x2 = (width * scale) - inset + offset
            y2 = (height * scale) - y_pad + offset
            radius = (y2 - y1) // 2
            fill_rgb = hex_to_rgb(fill)
            border_rgb = hex_to_rgb(border)
            glow_rgb = hex_to_rgb(glow)
            shine_rgb = hex_to_rgb(shine)

            draw_img.rounded_rectangle((x1 + 8, y1 + 12, x2 + 8, y2 + 12), radius=radius, fill=(120, 132, 150, 42))
            draw_img.rounded_rectangle((x1, y1, x2, y2), radius=radius, fill=(*fill_rgb, 238), outline=(*border_rgb, 245), width=scale)
            draw_img.line((x1 + radius, y1 + 7 * scale, x2 - radius, y1 + 7 * scale), fill=(*shine_rgb, 210), width=scale)
            draw_img.line((x1 + radius, y2 - 7 * scale, x2 - radius, y2 - 7 * scale), fill=(*glow_rgb, 190), width=scale)

            if canvas._lookup_animating and canvas._lookup_state == "normal":
                math = __import__("math")
                straight = max((x2 - x1) - (2 * radius), 1)
                arc_len = math.pi * radius
                perimeter = (straight * 2) + (arc_len * 2)
                pos = (canvas._lookup_anim % 100) / 100 * perimeter

                def point_at(distance: float) -> tuple[float, float]:
                    d = distance % perimeter
                    if d <= straight:
                        return x1 + radius + d, y1
                    d -= straight
                    if d <= arc_len:
                        angle = -90 + (d / arc_len) * 180
                        rad = math.radians(angle)
                        return x2 - radius + radius * math.cos(rad), y1 + radius + radius * math.sin(rad)
                    d -= arc_len
                    if d <= straight:
                        return x2 - radius - d, y2
                    d -= straight
                    angle = 90 + (d / arc_len) * 180
                    rad = math.radians(angle)
                    return x1 + radius + radius * math.cos(rad), y1 + radius + radius * math.sin(rad)

                def draw_runner(distance: float) -> None:
                    for index, gap in enumerate((22, 15, 8)):
                        tx, ty = point_at(distance - gap * scale)
                        alpha = (80, 130, 185)[index]
                        size = (1.2 + index * 0.35) * scale
                        draw_img.ellipse((tx - size, ty - size, tx + size, ty + size), fill=(255, 255, 255, alpha))
                    sx, sy = point_at(distance)
                    size = 2.4 * scale
                    draw_img.ellipse((sx - size, sy - size, sx + size, sy + size), fill=(255, 255, 255, 235), outline=(148, 163, 184, 210), width=scale)

                draw_runner(pos)
                draw_runner(pos + perimeter / 2)

            try:
                resample = Image.Resampling.LANCZOS
            except AttributeError:
                resample = Image.LANCZOS
            return ImageTk.PhotoImage(image.resize((width, height), resample))

        def draw(fill: str, pressed: bool = False) -> None:
            canvas.delete("all")
            width = max(canvas.winfo_width(), 340)
            height = max(canvas.winfo_height(), 62)
            canvas._lookup_photo = make_button_image(width, height, fill, pressed)
            canvas.create_image(0, 0, image=canvas._lookup_photo, anchor="nw")
            text_color = fg if canvas._lookup_state == "normal" else MUTED_TEXT
            if canvas._lookup_animating:
                if not hasattr(canvas, "_lookup_gear_source"):
                    gear_bytes = base64.b64decode(TLB_LOOKUP_GEAR_PNG_BASE64)
                    gear_image = Image.open(io.BytesIO(gear_bytes)).convert("RGBA")
                    bbox = gear_image.getchannel("A").getbbox()
                    if bbox:
                        gear_image = gear_image.crop(bbox)
                    canvas._lookup_gear_source = gear_image

                try:
                    rotate_resample = Image.Resampling.BICUBIC
                    resize_resample = Image.Resampling.LANCZOS
                except AttributeError:
                    rotate_resample = Image.BICUBIC
                    resize_resample = Image.LANCZOS

                gear_size = 18
                angle = -(canvas._lookup_anim / 100.0) * 360
                gear = canvas._lookup_gear_source.resize((gear_size, gear_size), resize_resample)
                gear = gear.rotate(angle, resample=rotate_resample, expand=True)
                gear_rgb = hex_to_rgb(fg)
                alpha = gear.getchannel("A").point(lambda value: int(value * 0.74))
                tinted = Image.new("RGBA", gear.size, (*gear_rgb, 0))
                tinted.putalpha(alpha)
                canvas._lookup_gear_photo = ImageTk.PhotoImage(tinted)
                canvas.create_image(width / 2 - 92, height / 2 + (1 if pressed else 0), image=canvas._lookup_gear_photo)
            canvas.create_text(
                width / 2,
                height / 2 + (1 if pressed else 0),
                text="DNC 조건 조회",
                fill=text_color,
                font=("맑은 고딕", 15, "bold"),
            )

        def set_state(state: str) -> None:
            canvas._lookup_state = state
            canvas.configure(cursor="hand2" if state == "normal" else "arrow")
            draw(surface)

        original_configure = canvas.configure

        def configure_proxy(*args, **kwargs):
            if "state" in kwargs:
                set_state(kwargs.pop("state"))
            if kwargs:
                return original_configure(*args, **kwargs)
            if args:
                return original_configure(*args)
            return None

        canvas.configure = configure_proxy

        def is_ready() -> bool:
            if ready_check is None:
                return True
            try:
                return bool(ready_check())
            except Exception:
                return False

        def run_glow_cycle(step: int = 0) -> None:
            if not canvas.winfo_exists():
                return
            if canvas._lookup_state != "normal" or not is_ready():
                canvas._lookup_animating = False
                draw(canvas._lookup_bg)
                canvas.after(500, run_glow_cycle)
                return
            canvas._lookup_animating = True
            canvas._lookup_anim = step
            draw(canvas._lookup_bg)
            canvas.after(100, lambda: run_glow_cycle((step + 1) % 100))

        def on_enter(_event) -> None:
            if canvas._lookup_state == "normal":
                canvas._lookup_bg = hover
                draw(hover)

        def on_leave(_event) -> None:
            if canvas._lookup_state == "normal":
                canvas._lookup_bg = surface
                draw(surface)

        def on_press(_event) -> None:
            if canvas._lookup_state == "normal":
                draw(active, pressed=True)

        def on_release(_event) -> None:
            if canvas._lookup_state == "normal":
                if not is_ready():
                    canvas._lookup_bg = surface
                    draw(surface)
                    return
                canvas._lookup_bg = hover
                draw(hover)
                canvas.after(80, command)

        canvas.bind("<Configure>", lambda _event: draw(canvas._lookup_bg))
        canvas.bind("<Enter>", on_enter)
        canvas.bind("<Leave>", on_leave)
        canvas.bind("<ButtonPress-1>", on_press)
        canvas.bind("<ButtonRelease-1>", on_release)
        canvas.after(1200, run_glow_cycle)
        self.normal_buttons.append(canvas)
        return canvas
    def create_settings_tab(self) -> None:
        page = self.settings_page
        page.columnconfigure(0, weight=1)
        panel = self.create_panel(page, "설정")
        panel.grid(row=0, column=0, sticky="ew", padx=14, pady=14)
        panel.columnconfigure(1, weight=1)

        self.excel_var = tk.StringVar(value=self.config.get("excel_file", ""))
        source_folders = normalize_saved_source_folders(self.config)
        self.config["source_dnc_folders"] = source_folders
        self.source_var = tk.StringVar(value=source_folders.get("KCC PKG", self.config.get("source_dnc_folder", "")))
        self.source_vars = {
            process_name: tk.StringVar(value=source_folders.get(process_name, ""))
            for process_name in PROCESS_NAMES
        }
        self.transfer_var = tk.StringVar(value=self.config.get("transfer_dnc_folder", ""))
        self.delete_seconds_var = tk.StringVar(value=str(self.config.get("dnc_delete_seconds", DNC_DELETE_SECONDS)))
        self.first_article_wait_var = tk.StringVar(value=str(self.config.get("first_article_wait_seconds", FIRST_ARTICLE_WAIT_SECONDS)))
        self.clear_common_var = tk.BooleanVar(value=bool(self.config.get("clear_common_after_normal", False)))
        self.auto_shift_group_var = tk.BooleanVar(value=bool(self.config.get("auto_shift_group_enabled", True)))
        self.a_group_day_start_var = tk.StringVar(value=str(self.config.get("a_group_day_start_date", "2026-05-17")))
        settings_label_width = 16

        common_rows = [
            ("작업일보 경로", self.excel_var, lambda: select_excel_file(self.root, self.config, self.excel_var)),
            ("DNC 전송 폴더", self.transfer_var, lambda: self.select_folder_to_var(self.transfer_var, save_after=True)),
        ]
        for row, (label, var, command) in enumerate(common_rows, start=1):
            ttk.Label(panel, text=label, background=SURFACE_BG, width=settings_label_width, anchor="e").grid(row=row, column=0, sticky="e", padx=10, pady=8)
            entry = ttk.Entry(panel, textvariable=var, style="Wide.TEntry")
            entry.grid(row=row, column=1, sticky="ew", padx=8, pady=8)
            entry.bind("<FocusOut>", lambda _event: self.save_settings_from_ui_silent(show_error=False))
            ttk.Button(panel, text="선택", command=command, width=20).grid(row=row, column=2, padx=8, pady=8)

        ttk.Label(panel, text="DNC 조건 시트 폴더 설정", background=PRIMARY_LIGHT, foreground=PRIMARY, anchor="center", font=("맑은 고딕", 10, "bold")).grid(
            row=3, column=0, columnspan=3, sticky="ew", padx=10, pady=(16, 6)
        )
        for index, process_name in enumerate(PROCESS_NAMES):
            row = 5 + index
            process_theme = PROCESS_COLORS.get(process_name, PROCESS_COLORS["KCC PKG"])
            label_bg = process_theme["light"]
            ttk.Label(panel, text=process_name, background=label_bg, foreground=process_theme["primary"], width=settings_label_width, anchor="center").grid(row=row, column=0, sticky="ew", padx=10, pady=5)
            entry = ttk.Entry(panel, textvariable=self.source_vars[process_name], style="Wide.TEntry")
            entry.grid(row=row, column=1, sticky="ew", padx=8, pady=5)
            entry.bind("<FocusOut>", lambda _event: self.save_settings_from_ui_silent(show_error=False))
            ttk.Button(panel, text="폴더 선택", command=lambda name=process_name: self.select_source_folder(name), width=20).grid(row=row, column=2, padx=8, pady=5)

        shift_row = 10
        ttk.Label(panel, text="근무표 조 자동", background=PRIMARY_LIGHT, foreground=PRIMARY, anchor="center", font=("맑은 고딕", 10, "bold")).grid(
            row=shift_row, column=0, columnspan=3, sticky="ew", padx=10, pady=(16, 6)
        )
        auto_wrap = tk.Frame(panel, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        auto_wrap.grid(row=shift_row + 1, column=0, columnspan=3, sticky="ew", padx=10, pady=5)
        auto_wrap.columnconfigure(0, minsize=126)
        auto_wrap.columnconfigure(6, weight=1)
        tk.Label(auto_wrap, text="근무표 조 자동", bg=SURFACE_BG, fg=TEXT_COLOR, width=settings_label_width, anchor="center", font=("맑은 고딕", 10)).grid(row=0, column=0, sticky="ew", padx=10, pady=8)
        toggle_wrap = tk.Frame(auto_wrap, bg=APP_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        toggle_wrap.grid(row=0, column=1, sticky="w", padx=8, pady=8)
        self.auto_shift_buttons = {}
        for value, text in ((True, "사용"), (False, "미사용")):
            button = tk.Button(
                toggle_wrap,
                text=text,
                command=lambda selected=value: self.set_auto_shift_group_enabled(selected),
                relief=tk.FLAT,
                bd=0,
                width=10,
                padx=10,
                pady=5,
                cursor="hand2",
                font=("맑은 고딕", 10, "bold"),
                activebackground="#dbeafe",
                activeforeground=PRIMARY,
            )
            button.pack(side=tk.LEFT, fill=tk.X, expand=True)
            button.bind("<Enter>", lambda _event, selected=value: self.set_auto_shift_button_hover(selected, True))
            button.bind("<Leave>", lambda _event, selected=value: self.set_auto_shift_button_hover(selected, False))
            button.bind("<ButtonPress-1>", lambda _event, selected=value: self.set_auto_shift_button_hover(selected, True, pressed=True))
            button.bind("<ButtonRelease-1>", lambda _event, selected=value: self.set_auto_shift_button_hover(selected, True))
            self.auto_shift_buttons[value] = button
        self.update_auto_shift_buttons()
        tk.Label(auto_wrap, text="A조 주간 시작 기준일", bg=PRIMARY_LIGHT, fg=TEXT_COLOR, width=18, anchor="center").grid(row=0, column=2, sticky="w", padx=(14, 6), pady=8)
        date_pick_wrap = tk.Frame(auto_wrap, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0, cursor="hand2")
        date_pick_wrap.grid(row=0, column=3, sticky="w", padx=(0, 6), pady=8)
        a_entry = tk.Entry(
            date_pick_wrap,
            textvariable=self.a_group_day_start_var,
            width=14,
            justify="center",
            relief=tk.FLAT,
            bd=0,
            bg=SURFACE_BG,
            fg=TEXT_COLOR,
            readonlybackground=SURFACE_BG,
            cursor="hand2",
            font=("맑은 고딕", 10),
        )
        a_entry.pack(ipadx=4, ipady=5)
        a_entry.configure(state="readonly")
        date_pick_wrap.bind("<Button-1>", lambda _event: self.open_a_group_day_start_picker())
        a_entry.bind("<Button-1>", lambda _event: self.open_a_group_day_start_picker())
        date_pick_wrap.bind("<Enter>", lambda _event: date_pick_wrap.configure(bg="#f8fbff"))
        date_pick_wrap.bind("<Leave>", lambda _event: date_pick_wrap.configure(bg=SURFACE_BG))
        tk.Label(auto_wrap, text="예: 2026-05-17", bg=SURFACE_BG, fg=MUTED_TEXT).grid(row=0, column=4, sticky="w", padx=(0, 8), pady=8)
        apply_button = tk.Button(
            auto_wrap,
            text="적용",
            command=self.apply_auto_shift_settings,
            relief=tk.FLAT,
            bd=0,
            width=10,
            padx=10,
            pady=5,
            cursor="hand2",
            bg=PRIMARY_LIGHT,
            fg=PRIMARY,
            activebackground="#dbeafe",
            activeforeground=PRIMARY,
            font=("맑은 고딕", 10, "bold"),
            highlightthickness=1,
            highlightbackground="#93c5fd",
            highlightcolor="#60a5fa",
        )
        apply_button.grid(row=0, column=5, sticky="w", padx=(0, 10), pady=8)
        apply_button.bind("<Enter>", lambda _event: apply_button.configure(bg="#dbeafe"))
        apply_button.bind("<Leave>", lambda _event: apply_button.configure(bg=PRIMARY_LIGHT))
        apply_button.bind("<ButtonPress-1>", lambda _event: apply_button.configure(bg="#bfdbfe"))
        apply_button.bind("<ButtonRelease-1>", lambda _event: apply_button.configure(bg="#dbeafe"))

        ttk.Button(panel, text="마스터 설정", command=self.open_master_settings_popup, style="Primary.TButton", width=20).grid(
            row=shift_row + 1, column=2, sticky="e", padx=(8, 20), pady=5
        )

        manage_row = shift_row + 2
        ttk.Label(panel, text="공정별 관리 도구", background=PRIMARY_LIGHT, foreground=PRIMARY, anchor="center", font=("맑은 고딕", 10, "bold")).grid(
            row=manage_row, column=0, columnspan=3, sticky="ew", padx=10, pady=(16, 6)
        )
        manage_wrap = tk.Frame(panel, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        manage_wrap.grid(row=manage_row + 1, column=0, columnspan=3, sticky="ew", padx=10, pady=5)
        for index, process_name in enumerate(PROCESS_NAMES):
            manage_wrap.columnconfigure(index, weight=1, uniform="process_manage")
            process_theme = PROCESS_COLORS.get(process_name, PROCESS_COLORS["KCC PKG"])
            process_box = tk.Frame(manage_wrap, bg=SURFACE_BG, highlightthickness=1, highlightbackground=process_theme["border"], bd=0)
            process_box.grid(row=0, column=index, sticky="nsew", padx=5, pady=8)
            process_box.columnconfigure(0, weight=1)
            tk.Label(
                process_box,
                text=process_name,
                bg=process_theme["light"],
                fg=process_theme["primary"],
                font=("맑은 고딕", 9, "bold"),
                anchor="center",
                pady=7,
            ).grid(row=0, column=0, sticky="ew")
            process_buttons = [
                ("작업 이력 보기", lambda name=process_name: self.open_process_work_history(name)),
                ("작업 이력 Excel 내보내기", lambda name=process_name: self.export_process_work_history(name)),
                ("작업조건/지그 Excel 내보내기", lambda name=process_name: self.export_process_condition_master(name)),
                ("전체 data 백업 만들기", self.create_data_backup_from_settings),
            ]
            for button_index, (text, command) in enumerate(process_buttons, start=1):
                ttk.Button(process_box, text=text, command=command).grid(row=button_index, column=0, sticky="ew", padx=6, pady=4)

    def select_folder_to_var(self, var: tk.StringVar, save_after: bool = False) -> None:
        folder = filedialog.askdirectory(initialdir=var.get() or str(get_desktop_path()))
        if folder:
            var.set(folder)
            if save_after:
                self.save_settings_from_ui_silent(show_error=False)

    def select_source_folder(self, process_name: str) -> None:
        var = self.source_vars[process_name]
        folder = filedialog.askdirectory(initialdir=var.get() or str(get_desktop_path()))
        if folder:
            var.set(folder)
            self.save_settings_from_ui_silent(show_error=False)

    def set_auto_shift_group_enabled(self, enabled: bool) -> None:
        self.auto_shift_group_var.set(enabled)
        self.update_auto_shift_buttons()

    def apply_auto_shift_settings(self) -> None:
        if self.save_settings_from_ui_silent(show_error=True):
            self.apply_work_time_defaults(initial=False, schedule_next=False)
            show_operator_alert(self.root, "근무표 조 자동", "설정 적용", "info")

    def open_a_group_day_start_picker(self) -> None:
        picker = tk.Toplevel(self.root)
        picker.title("A조 주간 시작 기준일")
        picker.configure(bg=APP_BG)
        picker.resizable(False, False)
        picker.transient(self.root)
        picker.grab_set()

        today = datetime.now()
        try:
            selected = datetime.strptime(self.a_group_day_start_var.get().strip(), "%Y-%m-%d")
        except ValueError:
            selected = today

        year_var = tk.IntVar(value=selected.year)
        month_var = tk.IntVar(value=selected.month)

        header = ttk.Frame(picker, padding=(12, 12, 12, 6))
        header.pack(fill=tk.X)
        ttk.Spinbox(header, from_=today.year - 5, to=today.year + 5, textvariable=year_var, width=8, justify="center").pack(side=tk.LEFT, padx=(0, 6))
        ttk.Spinbox(header, from_=1, to=12, textvariable=month_var, width=5, justify="center").pack(side=tk.LEFT, padx=(0, 6))

        days_frame = ttk.Frame(picker, padding=(12, 6, 12, 12))
        days_frame.pack()

        def select_day(day: int) -> None:
            self.a_group_day_start_var.set(f"{year_var.get():04d}-{month_var.get():02d}-{day:02d}")
            picker.destroy()

        def refresh_days() -> None:
            for child in days_frame.winfo_children():
                child.destroy()
            last_day = calendar.monthrange(year_var.get(), month_var.get())[1]
            for day in range(1, last_day + 1):
                ttk.Button(days_frame, text=str(day), width=4, command=lambda value=day: select_day(value)).grid(
                    row=(day - 1) // 7,
                    column=(day - 1) % 7,
                    padx=2,
                    pady=2,
                )

        ttk.Button(header, text="변경", command=refresh_days, width=6).pack(side=tk.LEFT)
        refresh_days()

    def set_auto_shift_button_hover(self, value: bool, hover: bool, pressed: bool = False) -> None:
        if not hasattr(self, "auto_shift_buttons") or value not in self.auto_shift_buttons:
            return
        selected = bool(self.auto_shift_group_var.get()) == value
        if pressed:
            bg = "#bfdbfe"
        elif hover:
            bg = "#dbeafe" if selected else "#f8fbff"
        else:
            bg = PRIMARY_LIGHT if selected else SURFACE_BG
        self.auto_shift_buttons[value].configure(bg=bg)

    def update_auto_shift_buttons(self) -> None:
        if not hasattr(self, "auto_shift_buttons"):
            return
        for value, button in self.auto_shift_buttons.items():
            selected = bool(self.auto_shift_group_var.get()) == value
            button.configure(
                bg=PRIMARY_LIGHT if selected else SURFACE_BG,
                fg=PRIMARY if selected else TEXT_COLOR,
                activebackground="#dbeafe",
                activeforeground=PRIMARY,
            )

    def require_kcc_pkg_management(self, process_name: str) -> bool:
        if process_name in {"KCC PKG", "TLB", "KCC HDI"}:
            return True
        show_operator_alert(self.root, "추후 개발 예정", f"{process_name} 준비중", "info")
        return False

    def open_process_work_history(self, process_name: str) -> None:
        if not self.require_kcc_pkg_management(process_name):
            return
        self.open_work_history_popup(process_name)

    def export_process_work_history(self, process_name: str) -> None:
        if not self.require_kcc_pkg_management(process_name):
            return
        process_tag = process_name.replace(" ", "_").replace("/", "_")
        default_name = f"{process_tag}_작업이력_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        selected_path = filedialog.asksaveasfilename(
            parent=self.root,
            title="작업 이력 Excel 저장",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
        )
        if not selected_path:
            return
        try:
            path = export_process_work_history_excel(process_name, Path(selected_path))
        except Exception as exc:
            log_error("작업 이력 Excel 내보내기 실패", exc)
            show_operator_alert(self.root, "내보내기 실패", "작업 이력 저장 실패")
            return
        show_operator_alert(self.root, "내보내기 완료", f"{process_name} 작업 이력 생성\n{path}", "info")

    def export_process_condition_master(self, process_name: str) -> None:
        if not self.require_kcc_pkg_management(process_name):
            return
        process_tag = process_name.replace(" ", "_").replace("/", "_")
        default_name = f"{process_tag}_작업조건_지그_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        selected_path = filedialog.asksaveasfilename(
            parent=self.root,
            title="작업조건/지그 Excel 저장",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
        )
        if not selected_path:
            return
        try:
            path = export_process_condition_master_excel(process_name, self.config, Path(selected_path))
        except Exception as exc:
            log_error("작업조건/지그 Excel 내보내기 실패", exc)
            show_operator_alert(self.root, "내보내기 실패", "작업조건/지그 저장 실패")
            return
        show_operator_alert(self.root, "내보내기 완료", f"{process_name} 작업조건/지그 생성\n{path}", "info")

    def create_data_backup_from_settings(self) -> None:
        default_name = f"data_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        selected_path = filedialog.asksaveasfilename(
            parent=self.root,
            title="전체 data 백업 저장",
            initialfile=default_name,
            defaultextension=".zip",
            filetypes=[("ZIP 파일", "*.zip")],
        )
        if not selected_path:
            return
        try:
            path = create_full_data_backup_zip(Path(selected_path))
        except Exception as exc:
            log_error("전체 data 백업 실패", exc)
            show_operator_alert(self.root, "백업 실패", "data 백업 실패")
            return
        show_operator_alert(self.root, "백업 완료", f"data 백업 생성\n{path}", "info")

    def save_settings_from_ui(self) -> None:
        self.save_settings_from_ui_silent(show_error=True)

    def rebuild_condition_master(self) -> None:
        self.save_settings_from_ui_silent()
        try:
            count = rebuild_condition_master_from_log(self.config)
        except Exception as exc:
            show_operator_alert(self.root, "조건 마스터 갱신 실패", str(exc), "error")
            return
        show_operator_alert(self.root, "작업일보 마스터 갱신", f"신규 등록 {count}건", "info")

    def open_work_history_popup(self, process_name: str = "KCC PKG") -> None:
        WorkHistoryPopup(self, process_name)

    def focus_existing_popup(self, attr_name: str) -> bool:
        popup = getattr(self, attr_name, None)
        window = getattr(popup, "window", None) if popup is not None else None
        try:
            if window is not None and window.winfo_exists():
                window.deiconify()
                window.lift()
                window.focus_force()
                return True
        except tk.TclError:
            pass
        setattr(self, attr_name, None)
        return False

    def open_master_settings_popup(self) -> None:
        if self.focus_existing_popup("master_settings_popup"):
            return
        password = ask_system_input(self.root, "\uB9C8\uC2A4\uD130 \uC124\uC815", "\uBE44\uBC00\uBC88\uD638 \uC785\uB825", show="*")
        if password is None:
            return
        if password != str(self.config.get("master_password", MASTER_SETTINGS_PASSWORD)):
            show_operator_alert(self.root, "\uBE44\uBC00\uBC88\uD638 \uD655\uC778", "\uBE44\uBC00\uBC88\uD638 \uBD88\uC77C\uCE58")
            return

        def delayed_open() -> None:
            if self.focus_existing_popup("master_settings_popup"):
                return
            try:
                self.master_settings_popup = MasterSettingsPopup(self)
            except Exception as exc:
                self.master_settings_popup = None
                log_error("\uB9C8\uC2A4\uD130 \uC124\uC815 \uCC3D \uC5F4\uAE30 \uC2E4\uD328", exc)
                show_operator_alert(self.root, "\uB9C8\uC2A4\uD130 \uC124\uC815", "\uCC3D \uC5F4\uAE30 \uC2E4\uD328", "error")

        self.root.after(150, delayed_open)

    def get_common_data(self) -> dict:
        return {key: entry.get() for key, entry in self.common_entries.items()}

    def handle_common_change(self, key: str, source: str) -> None:
        self.sync_common_field(key, source)
        self.mark_common_manual_change()

    def sync_common_field(self, key: str, source: str) -> None:
        entry_groups = {
            "kcc": self.common_entries,
            "tlb": self.tlb_common_entries,
            "kcc_hdi": self.kcc_hdi_common_entries,
        }
        source_entries = entry_groups.get(source, self.common_entries)
        if key not in source_entries:
            return
        value = source_entries[key].get()
        for name, target_entries in entry_groups.items():
            if name != source and key in target_entries:
                target_entries[key].set(value)
        if key == "machine":
            self.config["machine"] = value

    def sync_common_defaults_to_tlb(self) -> None:
        for target_entries in (self.tlb_common_entries, self.kcc_hdi_common_entries):
            if not target_entries:
                continue
            for key in ("machine", "work_date", "shift_group", "shift", "worker"):
                if key in self.common_entries and key in target_entries:
                    target_entries[key].set(self.common_entries[key].get())

    def get_lot_data(self, entries: dict[str, LabeledEntry]) -> dict:
        return {key: entry.get() for key, entry in entries.items()}

    def is_kcc_lookup_ready(self, lot_number: int) -> bool:
        entries = self.lot1_entries if lot_number == 1 else self.lot2_entries
        required_keys = ("step", "round", "manage_no", "lot_no", "qty", "process_code")
        return all(entries[key].get().strip() for key in required_keys)

    def lot_has_any_value(self, lot: dict) -> bool:
        # LOT 2는 선택 입력이므로 차수 버튼만 눌린 상태는 사용으로 보지 않습니다.
        return any(value.strip() for key, value in lot.items() if key != "round")

    def has_new_model_target(self) -> bool:
        """조건/지그가 없는 신규 LOT가 있거나 입력 LOT가 없을 때만 신규 검증을 허용합니다."""
        used_lot_count = 0
        for entries in (self.lot1_entries, self.lot2_entries):
            lot = self.get_lot_data(entries)
            if not self.lot_has_any_value(lot):
                continue
            used_lot_count += 1
            condition_ok, _message = get_single_condition_message(lot)
            if not condition_ok:
                return True
        return used_lot_count == 0

    def update_new_model_button_state(self) -> None:
        if self.new_model_button is None:
            return
        state = "normal" if self.has_new_model_target() and not self.is_running else "disabled"
        self.new_model_button.configure(state=state)

    def mark_common_manual_change(self) -> None:
        self.last_common_manual_change_at = datetime.now()
        if not self.common_entries:
            return
        if self.common_entries["shift_group"].get() and self.common_entries["worker"].get():
            dnc_label = self.status_labels.get("dnc")
            if dnc_label and "근무" in str(dnc_label.cget("text")):
                self.set_status("dnc", "대기중", None)

    def get_work_prep_start(self, period: dict[str, str]) -> datetime:
        work_date = datetime.strptime(period["work_date"], "%Y-%m-%d")
        if period["shift"] == "주간":
            return work_date.replace(hour=8, minute=0, second=0, microsecond=0)
        return work_date.replace(hour=20, minute=0, second=0, microsecond=0)

    def has_prepared_next_shift(self, period: dict[str, str]) -> bool:
        changed_at = self.last_common_manual_change_at
        if changed_at is None:
            return False
        if changed_at < self.get_work_prep_start(period):
            return False
        return bool(self.common_entries["shift_group"].get() and self.common_entries["worker"].get())

    def apply_work_time_defaults(self, initial: bool = False, schedule_next: bool = True) -> None:
        """작업일자와 근무를 시간 기준으로 자동 적용하고 근무 전환 시 조/작업자를 다시 입력하게 합니다."""
        if not self.common_entries:
            return
        period = get_work_period()
        period_changed = bool(self.current_work_period_key and self.current_work_period_key != period["period_key"])
        prepared_next_shift = period_changed and self.has_prepared_next_shift(period)
        self.current_work_period_key = period["period_key"]
        self.common_entries["work_date"].set(period["work_date"])
        self.common_entries["shift"].set(period["shift"])
        auto_group = ""
        if self.config.get("auto_shift_group_enabled", True):
            auto_group = get_auto_shift_group(
                period["work_date"],
                period["shift"],
                str(self.config.get("a_group_day_start_date", "2026-05-17")),
            )
            if auto_group:
                self.common_entries["shift_group"].set(auto_group)
        if period_changed and not prepared_next_shift:
            if not auto_group:
                self.common_entries["shift_group"].clear()
            self.common_entries["worker"].clear()
            self.sync_common_defaults_to_tlb()
            self.set_status("dnc", "근무 전환 확인 필요", False)
            if not initial:
                show_operator_alert(
                    self.root,
                    "근무 전환 확인",
                    "작업일자/근무 변경됨\n작업자 입력 필요",
                )
        elif period_changed and prepared_next_shift:
            self.set_status("dnc", "근무 전환 확인 완료", True)
        self.sync_common_defaults_to_tlb()
        if schedule_next:
            self.root.after(60000, self.apply_work_time_defaults)

    def ensure_work_period_ready(self) -> bool:
        self.apply_work_time_defaults(initial=False, schedule_next=False)
        missing = []
        if not self.common_entries["shift_group"].get():
            missing.append("조")
        if not self.common_entries["worker"].get():
            missing.append("작업자")
        if missing:
            show_operator_alert(
                self.root,
                "근무 정보 확인",
                "\n".join(missing_common_message(label) for label in missing),
            )
            self.set_status("dnc", "근무 정보 확인 필요", False)
            return False
        dnc_label = self.status_labels.get("dnc")
        if dnc_label and "근무" in str(dnc_label.cget("text")):
            self.set_status("dnc", "대기중", None)
        return True

    def set_status(self, key: str, text: str, ok: bool | None = None) -> None:
        color = MUTED_TEXT if ok is None else (OK_COLOR if ok else NG_COLOR)
        self.status_labels[key].configure(text=text, fg=color)
        if key == "dnc":
            self.append_log(text)

    def set_lot_match_segments(self, segments: list[tuple[str, str]]) -> None:
        if self.lot_match_frame is None:
            return
        for child in self.lot_match_frame.winfo_children():
            child.destroy()
        for index, (text, color) in enumerate(segments):
            tk.Label(
                self.lot_match_frame,
                text=text,
                bg=SURFACE_BG,
                fg=color,
                font=("맑은 고딕", 12, "bold"),
                anchor="w",
            ).pack(side=tk.LEFT)
            if index < len(segments) - 1:
                tk.Label(
                    self.lot_match_frame,
                    text=" / ",
                    bg=SURFACE_BG,
                    fg=MUTED_TEXT,
                    font=("맑은 고딕", 12, "bold"),
                ).pack(side=tk.LEFT)

    def get_lot_short_status(self, label: str, lot: dict) -> tuple[str, str, bool]:
        missing = []
        if not lot.get("condition", "").strip():
            missing.append("조건")
        if not lot.get("jig", "").strip():
            missing.append("지그")
        if missing:
            return f"{label} {'·'.join(missing)} 없음", NG_COLOR, False
        return f"{label} 정상", TEXT_COLOR, True

    def update_lot_match_summary(self, lot1: dict, lot2: dict | None) -> None:
        if not lot2:
            self.set_lot_match_segments([("LOT 2 미사용", MUTED_TEXT)])
            return
        segments: list[tuple[str, str]] = []
        lot1_text, lot1_color, lot1_ok = self.get_lot_short_status("LOT 1", lot1)
        lot2_text, lot2_color, lot2_ok = self.get_lot_short_status("LOT 2", lot2)
        segments.append((lot1_text, lot1_color))
        segments.append((lot2_text, lot2_color))

        if lot1.get("lot_no", "").strip() and lot1.get("lot_no", "").strip() == lot2.get("lot_no", "").strip():
            segments.append(("LOT No 중복", NG_COLOR))
        if lot1_ok and lot2_ok:
            if lot1.get("condition", "").strip() != lot2.get("condition", "").strip():
                segments.append(("작업조건 다름", NG_COLOR))
            if lot1.get("jig", "").strip() != lot2.get("jig", "").strip():
                segments.append(("지그 다름", NG_COLOR))
            if len(segments) == 2:
                segments.append(("조건 일치", OK_COLOR))
        self.set_lot_match_segments(segments)

    def set_dnc_status(self, text: str) -> None:
        self.root.after(0, lambda: self.set_status("dnc", text, None))

    def append_log(self, text: str) -> None:
        if self.log_text is None:
            return
        if text.startswith("DNC 삭제 대기중"):
            return
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert(tk.END, f"[{timestamp}] {text}\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state="disabled")

    def update_status_checks(self) -> None:
        lot1 = self.get_lot_data(self.lot1_entries)
        lot2 = self.get_lot_data(self.lot2_entries)
        lot2_used = self.lot_has_any_value(lot2)
        self.update_lot_detail_status("lot1", lot1, waiting_when_empty=True)
        if lot2_used:
            self.update_lot_detail_status("lot2", lot2, waiting_when_empty=True)
            self.update_lot_match_summary(lot1, lot2)
        else:
            self.set_lot_status("lot2_mes", "", None)
            self.set_lot_status("lot2_condition", "", None)
            self.update_lot_match_summary(lot1, None)
        self.set_kcc_run_animation(self.is_kcc_run_ready_by_condition())
        self.update_new_model_button_state()

    def update_lot_detail_status(self, prefix: str, lot: dict, waiting_when_empty: bool) -> None:
        lot_no = lot.get("lot_no", "").strip()
        process_code = lot.get("process_code", "").strip()
        condition = lot.get("condition", "").strip()
        jig = lot.get("jig", "").strip()

        if waiting_when_empty and (not lot_no or not process_code):
            self.set_lot_status(f"{prefix}_mes", "", None)
        else:
            mes_ok, mes_message = get_mes_core_message(lot_no, process_code)
            if mes_ok:
                self.set_lot_status(f"{prefix}_mes", "OK", True)
            else:
                self.set_lot_status(f"{prefix}_mes", "NG", False)

        if waiting_when_empty and not condition and not jig:
            self.set_lot_status(f"{prefix}_condition", "", None)
        else:
            condition_ok, condition_message = get_single_condition_message(lot)
            self.set_lot_status(f"{prefix}_condition", "OK" if condition_ok else "NG", condition_ok)

    def set_lot_status(self, key: str, text: str, ok: bool | None = None) -> None:
        if key not in self.lot_status_labels:
            return
        label = self.lot_status_labels[key]
        title = "MES Core" if key.endswith("_mes") else "조건 적용"
        if ok is None and not str(text).strip():
            self.hide_judgement_card(label)
            return
        self.show_judgement_card(label, title, text, ok)

    def set_running(self, running: bool) -> None:
        self.is_running = running
        state = "disabled" if running else "normal"
        for button in self.normal_buttons:
            button.configure(state=state)
        if not running:
            self.update_new_model_button_state()

    def validate_condition_file(self, condition_name: str) -> Path | None:
        self.set_dnc_status("조건 파일 검색중")
        matches = search_condition_file(condition_name, Path(self.config["source_dnc_folder"]))
        if len(matches) == 0:
            show_operator_alert(self.root, "조건 파일 없음", "DNC 파일 없음", "error")
            return None
        if len(matches) >= 2:
            log_app(f"동일 조건 파일 차단: 조건={condition_name}, 검색수량={len(matches)}, 파일={matches}")
            show_operator_alert(
                self.root,
                "동일 DNC 파일",
                format_duplicate_condition_files(matches),
                "error",
            )
            return None
        return matches[0]

    def validate_kcc_pkg_process_paths(self) -> bool:
        ok, message = validate_process_paths(self.config, "KCC PKG")
        if not ok:
            log_app(f"KCC PKG 필수 경로 확인 실패: {message}")
            show_operator_alert(self.root, "경로 확인", message)
            self.set_status("dnc", "경로 확인 필요", False)
            return False
        return True

    def make_lot_lookup_key(self, lot: dict) -> str:
        """조건/지그를 불러온 기준 키입니다. 이 값이 바뀌면 기존 조건을 비웁니다."""
        return "|".join(
            [
                lot.get("step", "").strip(),
                lot.get("round", "").strip(),
                lot.get("process_code", "").strip(),
                lot.get("manage_no", "").strip(),
            ]
        )

    def handle_lot_key_change(self, lot_number: int) -> None:
        entries = self.lot1_entries if lot_number == 1 else self.lot2_entries
        loaded_key = self.lot_condition_keys.get(lot_number, "")
        if not loaded_key:
            return
        current_key = self.make_lot_lookup_key(self.get_lot_data(entries))
        if current_key == loaded_key:
            return
        entries["condition"].clear()
        entries["jig"].clear()
        self.lot_condition_keys[lot_number] = ""
        self.set_status("dnc", f"LOT {lot_number} 기준값 변경 - 작업조건/지그 초기화", None)

    def load_condition_jig_for_lot(self, lot_number: int) -> bool:
        """작업일보 이력에서 선택 LOT의 작업조건/지그를 불러와 화면에 채웁니다."""
        self.save_settings_from_ui_silent()
        entries = self.lot1_entries if lot_number == 1 else self.lot2_entries
        lot = self.get_lot_data(entries)
        missing = [
            label
            for key, label in (("step", "STEP"), ("round", "차수"), ("process_code", "공정코드"))
            if not lot.get(key, "").strip()
        ]
        if missing:
            show_operator_alert(self.root, "입력 확인", f"LOT {lot_number} {' / '.join(missing)} 입력 필요")
            return False
        try:
            condition, jig, source = lookup_condition_jig_from_history(self.config, lot)
        except ValueError as exc:
            show_operator_alert(self.root, "중복 조건", str(exc), "error")
            self.set_lot_status(f"lot{lot_number}_condition", "NG", False)
            self.set_status("dnc", "중복 조건 확인 필요", False)
            return False
        except Exception as exc:
            show_operator_alert(self.root, "이력 조회 실패", str(exc), "error")
            return False
        if not condition or not jig:
            detail = describe_condition_lookup_mismatch(lot)
            log_app(
                "조건 조회 실패: "
                f"LOT {lot_number}, STEP={lot.get('step')}, 차수={lot.get('round')}, "
                f"관리번호={lot.get('manage_no')}, 공정코드={lot.get('process_code')} / {detail}"
            )
            show_operator_alert(self.root, "이력 없음", f"LOT {lot_number} 조건/지그 없음")
            self.set_lot_status(f"lot{lot_number}_condition", "NG", False)
            return False
        entries["condition"].set(condition)
        entries["jig"].set(jig)
        refreshed_lot = self.get_lot_data(entries)
        self.lot_condition_keys[lot_number] = self.make_lot_lookup_key(refreshed_lot)
        self.update_status_checks()
        current_condition = lot.get("condition", "").strip()
        current_jig = lot.get("jig", "").strip()
        if current_condition != condition or current_jig != jig:
            self.set_status("dnc", f"LOT {lot_number} 조건/지그 불러옴: {source}", True)
        return True

    def run_normal_dnc(self) -> None:
        if self.is_running:
            show_operator_alert(self.root, "진행 중", "DNC 실행중")
            return
        if not self.ensure_work_period_ready():
            return
        if not self.save_settings_from_ui_silent():
            return
        if not self.validate_kcc_pkg_process_paths():
            return
        self.set_status("dnc", "입력값 확인중", None)
        common = self.get_common_data()
        lot1 = self.get_lot_data(self.lot1_entries)
        lot2_data = self.get_lot_data(self.lot2_entries)
        lot2 = lot2_data if self.lot_has_any_value(lot2_data) else None

        if not self.ensure_kcc_condition_confirmed(lot1, lot2):
            return
        ok, errors = validate_normal_dnc(common, lot1, lot2)
        if not ok:
            log_app("일반 DNC 입력값 NG: " + " / ".join(errors))
            show_operator_alert(self.root, "입력값 확인", format_operator_errors(errors))
            self.set_status("dnc", "입력값 NG", False)
            return
        model_change = ask_system_yes_no(self.root, "기종교체 확인", "기종교체 입니까?")
        self.frequent_check_values = [""] * 12
        self.work_axis_values = [""] * 6
        machine_axes = get_machine_allowed_axes(common.get("machine", ""))
        if model_change:
            self.set_status("dnc", "하부 Pin 확인 대기중", None)
            if not self.open_frequent_check_popup("jig", allowed_axes=machine_axes):
                self.set_status("dnc", "하부 Pin 확인 미완료", False)
                return
        stack = ask_numeric_input(self.root, "Stack 수 입력", "Stack 수를 입력 하세요.")
        ok, message = validate_positive_number(stack or "", "Stack 수", required=True)
        if not ok:
            show_operator_alert(self.root, "Stack 수 확인", message)
            return
        lots = [lot1] + ([lot2] if lot2 else [])
        if model_change:
            capacity_values = ["OK" if value == "OK" else "" for value in self.frequent_check_values[6:]] + [""] * 6
            self.work_axis_values = capacity_values[:6]
        else:
            self.set_status("dnc", "작업 축 수 확인 대기중", None)
            self.frequent_check_values = [""] * 12
            if not self.open_frequent_check_popup("capacity", allowed_axes=machine_axes):
                self.set_status("dnc", "작업 축 수 확인 미완료", False)
                return
            capacity_values = self.frequent_check_values[:]
            self.work_axis_values = capacity_values[:6]
        capacity_ok, capacity_message = validate_frequent_check_capacity(lots, stack, capacity_values)
        if not capacity_ok:
            show_operator_alert(self.root, "작업 수량 확인", capacity_message)
            self.set_status("dnc", "수량/Stack NG", False)
            return
        if model_change:
            self.set_status("dnc", "작업 수량 확인 완료", True)
        condition_file = self.validate_condition_file(lot1["condition"])
        if not condition_file:
            self.set_status("dnc", "조건 파일 NG", False)
            return

        log_app(f"일반 DNC 시작: LOT수={len(lots)}, 조건={lot1['condition']}, 기종교체={model_change}")
        self.set_running(True)
        threading.Thread(target=self.normal_worker, args=(common, lots, stack, model_change, condition_file), daemon=True).start()

    def normal_worker(self, common: dict, lots: list[dict], stack: str, model_change: bool, condition_file: Path) -> None:
        try:
            self.set_dnc_status("DB 저장중")
            log_ids = insert_normal_dnc_db(common, lots, stack, model_change)
            self.set_dnc_status("DB 저장 완료")
            log_app(f"일반 DNC 파일 처리 시작: ids={log_ids}")
            delete_existing_dnc_txt(Path(self.config["transfer_dnc_folder"]))
            log_app("DNC 전송 폴더 기존 txt 삭제 완료")
            copied_file = copy_dnc_file(condition_file, Path(self.config["transfer_dnc_folder"]))
            self.set_dnc_status("DNC 파일 복사 완료")
            log_app(f"DNC 파일 복사 완료: {copied_file}")
            delete_thread = threading.Thread(
                target=delete_after_delay,
                args=(copied_file, int(self.config["dnc_delete_seconds"]), self.set_dnc_status),
                daemon=True,
            )
            delete_thread.start()
            wait_seconds = int(self.config.get("first_article_wait_seconds", FIRST_ARTICLE_WAIT_SECONDS))
            log_app(f"초품 확인 대기 시작: {wait_seconds}초")
            for remain in range(wait_seconds, 0, -1):
                self.set_dnc_status(f"초품 확인 대기중 ({remain}초)")
                time.sleep(1)
            log_app("초품 확인 팝업 호출")
            self.root.after(0, lambda: self.finish_normal_dnc(log_ids, lots, stack, model_change, delete_thread))
        except Exception as exc:
            self.root.after(0, lambda error=exc: self.handle_run_error(error))

    def release_running_after_delete(self, delete_thread: threading.Thread | None) -> None:
        if delete_thread and delete_thread.is_alive():
            self.set_status("dnc", "DNC 삭제 완료 대기중", None)

            def wait_and_release() -> None:
                delete_thread.join()
                self.root.after(0, lambda: self.set_running(False))

            threading.Thread(target=wait_and_release, daemon=True).start()
            return
        self.set_running(False)

    def finish_normal_dnc(self, log_ids: list[int], lots: list[dict], stack: str, model_change: bool, delete_thread: threading.Thread | None = None) -> None:
        try:
            while True:
                allowed_axes = None
                if model_change:
                    allowed_axes = [
                        index
                        for index, value in enumerate(self.frequent_check_values[6:])
                        if value == "OK"
                    ]
                    for index in range(6):
                        self.frequent_check_values[index] = ""
                else:
                    allowed_axes = [
                        index
                        for index, value in enumerate(self.work_axis_values)
                        if value == "OK"
                    ]
                    self.frequent_check_values = [""] * 12
                if not self.open_frequent_check_popup("first", allowed_axes=allowed_axes):
                    self.set_status("dnc", "초품 확인 미완료", False)
                    return
                if model_change:
                    first_axes = [
                        index
                        for index, value in enumerate(self.frequent_check_values[:6])
                        if value == "OK"
                    ]
                    if first_axes != allowed_axes:
                        log_app(
                            "초품 축 확인 NG: "
                            f"하부Pin축={','.join(str(axis + 1) for axis in allowed_axes)}, "
                            f"초품축={','.join(str(axis + 1) for axis in first_axes) or '없음'}"
                        )
                        show_operator_alert(
                            self.root,
                            "초품 4Point 확인",
                            "하부 Pin 축과 초품 축 다름",
                        )
                        self.set_status("dnc", "초품 축 확인 NG", False)
                        continue
                ok, message = validate_frequent_check_capacity(lots, stack, self.frequent_check_values)
                if ok:
                    self.set_status("dnc", message, True)
                    break
                show_operator_alert(self.root, "초품 수량 확인", message)
                self.set_status("dnc", "초품 수량 NG", False)
            update_normal_frequent_check_db(log_ids, model_change, self.frequent_check_values)
            if delete_thread and delete_thread.is_alive():
                self.set_status("dnc", "DNC 완료 대기중", None)

                def wait_and_continue() -> None:
                    delete_thread.join()
                    self.root.after(0, lambda: self.finish_normal_after_delete(log_ids))

                threading.Thread(target=wait_and_continue, daemon=True).start()
                return
            self.finish_normal_after_delete(log_ids)
        except Exception as exc:
            self.handle_run_error(exc)

    def finish_normal_after_delete(self, log_ids: list[int]) -> None:
        try:
            burr_ok = ask_system_yes_no(self.root, "Burr 확인", "4면 Burr 이상 없습니까?")
            update_normal_burr_db(log_ids, burr_ok)
            pending_count = get_unexported_kcc_pkg_count()
            self.set_status("dnc", "DNC 완료", True)
            self.set_status("excel", f"DB 저장 완료 / Excel 미반영 {pending_count}건", True)
            log_app(f"일반 DNC 완료: ids={log_ids}, Excel 미반영={pending_count}건")
            self.auto_export_kcc_pkg_to_excel(parent=self.root)
            self.clear_normal_inputs(after_done=True)
        except Exception as exc:
            self.handle_run_error(exc)
        finally:
            self.set_running(False)

    def clear_normal_inputs(self, after_done: bool = False) -> None:
        for entry in self.lot1_entries.values():
            entry.clear()
        for entry in self.lot2_entries.values():
            entry.clear()
        self.frequent_check_values = [""] * 12
        self.work_axis_values = [""] * 6
        self.lot_condition_keys = {1: "", 2: ""}
        self.update_status_checks()
        if after_done:
            self.set_status("dnc", "DNC 완료", True)
        else:
            self.set_status("dnc", "대기중", None)

    def open_new_model_popup(self) -> None:
        if not self.has_new_model_target():
            show_operator_alert(
                self.root,
                "신규 모델 검증 DNC",
                "신규 검증 대상 없음\n일반 DNC 실행 사용",
                "info",
            )
            return
        if not self.ensure_work_period_ready():
            return
        open_new_model_popup(self)

    def open_log_excel_from_ui(self) -> None:
        self.save_settings_from_ui_silent()
        if ensure_excel_file_selected(self.root, self.config, self.excel_var if hasattr(self, "excel_var") else None):
            open_log_excel(self.config)

    def auto_export_kcc_pkg_to_excel(self, parent=None) -> bool:
        """DNC 완료 후 작업일보 반영을 자동으로 시도합니다. 실패해도 DB 미반영 상태는 유지됩니다."""
        pending_before = get_unexported_kcc_pkg_count()
        if pending_before == 0:
            self.set_status("excel", "Excel 미반영 0건", True)
            return True
        if not self.config.get("auto_export_after_dnc", True):
            self.set_status("excel", f"자동 반영 꺼짐 / Excel 미반영 {pending_before}건", False)
            return False
        excel_path = Path(self.config.get("excel_file", ""))
        if not excel_path.exists():
            self.set_status("excel", f"작업일보 경로 필요 / Excel 미반영 {pending_before}건", False)
            self.append_log(f"작업일보 자동 반영 실패: 작업일보 파일 없음 / 미반영 {pending_before}건")
            return False
        try:
            self.is_exporting_excel = True
            exported_count = export_kcc_pkg_db_to_excel(self.config)
        except Exception as exc:
            pending_after = get_unexported_kcc_pkg_count()
            self.set_status("excel", f"자동 반영 실패 / Excel 미반영 {pending_after}건", False)
            self.append_log(f"작업일보 자동 반영 실패: {exc}")
            alert_message = "다른 PC 반영 중\n나중에 작업일보 반영" if "다른 PC" in str(exc) else "DB 저장 완료\n나중에 작업일보 반영"
            show_operator_alert(
                parent or self.root,
                "작업일보 반영 실패",
                alert_message,
            )
            return False
        finally:
            self.is_exporting_excel = False
        pending_after = get_unexported_kcc_pkg_count()
        self.set_status("excel", f"작업일보 자동 반영 완료 / Excel 미반영 {pending_after}건", True)
        self.append_log(f"작업일보 자동 반영 완료: {exported_count}건 / 미반영 {pending_after}건")
        return True

    def export_kcc_pkg_to_excel_from_ui(self) -> None:
        """전체 공정 DB에 저장된 미반영 이력을 Excel 작업일보로 내보냅니다."""
        self.save_settings_from_ui_silent()
        if not ensure_excel_file_selected(self.root, self.config, self.excel_var if hasattr(self, "excel_var") else None):
            self.set_status("excel", "작업일보 반영 취소", False)
            return
        try:
            self.is_exporting_excel = True
            result = export_all_processes_to_excel(self.config)
        except FileNotFoundError as exc:
            show_operator_alert(self.root, "작업일보 반영 실패", format_excel_error_for_operator(exc, "작업일보 Excel"), "error")
            self.set_status("excel", "작업일보 파일 없음", False)
            return
        except PermissionError as exc:
            show_operator_alert(self.root, "작업일보 반영 실패", format_excel_error_for_operator(exc, "작업일보 Excel"), "error")
            self.set_status("excel", "작업일보 저장 실패", False)
            return
        except KeyError as exc:
            show_operator_alert(self.root, "작업일보 반영 실패", format_excel_error_for_operator(exc, "작업일보 Excel"), "error")
            self.set_status("excel", "작업일보 시트 없음", False)
            return
        except ValueError as exc:
            show_operator_alert(self.root, "작업일보 반영 실패", format_excel_error_for_operator(exc, "작업일보 Excel"), "error")
            self.set_status("excel", "작업일보 파일 오류", False)
            return
        except Exception as exc:
            log_error("작업일보 반영 실패", exc)
            show_operator_alert(self.root, "작업일보 반영 실패", format_excel_error_for_operator(exc, "작업일보 Excel"), "error")
            self.set_status("excel", "작업일보 반영 실패", False)
            return
        finally:
            self.is_exporting_excel = False
        total_count = sum(result.values())
        pending_total = sum(get_unexported_process_count(name) for name in PROCESS_NAMES)
        if total_count == 0:
            show_operator_alert(self.root, "작업일보 반영", "반영할 이력 없음", "info")
            self.set_status("excel", "Excel 미반영 0건", True)
            return
        detail = "\n".join(f"{name}: {count}건" for name, count in result.items() if count)
        show_operator_alert(self.root, "작업일보 반영 완료", f"{total_count}건 반영 완료", "info")
        self.set_status("excel", f"작업일보 반영 완료 / Excel 미반영 {pending_total}건", True)

    def open_condition_master_popup(self) -> None:
        if self.focus_existing_popup("condition_master_popup"):
            return
        password = ask_system_input(self.root, "\uC870\uAC74 \uB9C8\uC2A4\uD130 \uAD00\uB9AC", "\uBE44\uBC00\uBC88\uD638 \uC785\uB825", show="*")
        if password is None:
            return
        if password != str(self.config.get("condition_master_password", CONDITION_MASTER_PASSWORD)):
            show_operator_alert(self.root, "\uBE44\uBC00\uBC88\uD638 \uD655\uC778", "\uBE44\uBC00\uBC88\uD638 \uBD88\uC77C\uCE58")
            return

        def delayed_open() -> None:
            if self.focus_existing_popup("condition_master_popup"):
                return
            self.condition_master_popup = ConditionMasterPopup(self)

        self.root.after(150, delayed_open)

    def open_frequent_check_popup(self, mode: str, allowed_axes: list[int] | None = None) -> bool:
        if mode == "first":
            self.set_status("dnc", "초품 확인창 응답 대기중", None)
        elif mode == "jig":
            self.set_status("dnc", "하부 Pin 확인창 응답 대기중", None)
        elif mode == "capacity":
            self.set_status("dnc", "작업 축 수 확인창 응답 대기중", None)
        popup = FrequentCheckPopup(self, mode=mode, allowed_axes=allowed_axes)
        self.root.wait_window(popup.window)
        return popup.saved

    def has_frequent_check_completed(self, require_jig_check: bool = False) -> bool:
        return validate_frequent_check_values(self.frequent_check_values, check_mode="first")[0]

    def save_settings_from_ui_silent(self, show_error: bool = True) -> bool:
        if hasattr(self, "excel_var"):
            ok, message = validate_positive_number(self.delete_seconds_var.get(), "삭제 대기 시간", required=True)
            if not ok:
                if show_error:
                    show_operator_alert(self.root, "설정 확인", message)
                return False
            ok, message = validate_positive_number(self.first_article_wait_var.get(), "초품 알람 시간", required=True)
            if not ok:
                if show_error:
                    show_operator_alert(self.root, "설정 확인", message)
                return False
            excel_path = self.excel_var.get().strip()
            if excel_path:
                self.config["excel_file"] = excel_path
            elif self.config.get("excel_file", ""):
                self.excel_var.set(self.config["excel_file"])
            source_folders = dict(self.config.get("source_dnc_folders", {}))
            if hasattr(self, "source_vars"):
                for process_name, var in self.source_vars.items():
                    source_folders[process_name] = var.get().strip()
            self.config["source_dnc_folders"] = source_folders
            self.config["source_dnc_folder"] = source_folders.get("KCC PKG", self.source_var.get().strip())
            self.config["transfer_dnc_folder"] = self.transfer_var.get().strip()
            self.config["dnc_delete_seconds"] = int(self.delete_seconds_var.get().strip())
            self.config["first_article_wait_seconds"] = int(self.first_article_wait_var.get().strip())
            self.config["clear_common_after_normal"] = self.clear_common_var.get()
            if hasattr(self, "auto_shift_group_var"):
                self.config["auto_shift_group_enabled"] = bool(self.auto_shift_group_var.get())
            if hasattr(self, "a_group_day_start_var"):
                base_date = self.a_group_day_start_var.get().strip()
                try:
                    datetime.strptime(base_date, "%Y-%m-%d")
                    self.config["a_group_day_start_date"] = base_date
                except ValueError:
                    if show_error:
                        show_operator_alert(self.root, "설정 확인", "기준일 형식 확인")
                    return False
            if "machine" in self.common_entries:
                self.config["machine"] = self.common_entries["machine"].get()
            save_config(self.config)
        return True

    def handle_run_error(self, exc: Exception) -> None:
        log_error("DNC 실행 오류", exc)
        show_operator_alert(self.root, "오류", str(exc), "error")
        self.set_status("dnc", "오류", False)
        self.set_running(False)


class FrequentCheckPopup:
    """Q:AB에 기록할 초품/하부 Pin 확인 값을 클릭으로 입력하는 창입니다."""

    LABELS = [
        "1축",
        "2축",
        "3축",
        "4축",
        "5축",
        "6축",
        "1축",
        "2축",
        "3축",
        "4축",
        "5축",
        "6축",
    ]
    FIRST_OK_COLOR = "#0ea5e9"
    JIG_OK_COLOR = "#10b981"

    def __init__(self, app: JiinDncManager, mode: str, allowed_axes: list[int] | None = None):
        self.app = app
        self.mode = mode
        self.allowed_axes = set(allowed_axes) if allowed_axes is not None else None
        self.saved = False
        self.clicks_enabled = False
        if self.mode == "first":
            self.values = app.frequent_check_values[:]
        else:
            self.values = [""] * 12
        if len(self.values) < 12:
            self.values = (self.values + [""] * 12)[:12]
        self.buttons: list[tk.Button] = []
        self.visible_axes = self.get_visible_axes()
        # 팝업은 열릴 때마다 현재 확인 항목을 빈 상태로 시작합니다.
        # 이전 작업의 축 선택이 다음 하부 Pin/초품 확인에 남으면 현장 오판이 생깁니다.
        for axis_index in self.visible_axes:
            self.values[self.get_value_index(axis_index)] = ""
        self.window = tk.Toplevel(app.root)
        self.window.title(self.get_title())
        width = 560 if len(self.visible_axes) >= 6 else 420
        self.window.geometry(f"{width}x300")
        self.window.configure(bg=APP_BG)
        self.window.resizable(False, False)
        self.create_ui()
        self.window.transient(app.root)
        self.window.grab_set()
        self.center_on_parent()
        keep_modal_on_top(self.window, app.root)
        self.window.after(450, self.enable_clicks)

    def enable_clicks(self) -> None:
        self.clicks_enabled = True

    def center_on_parent(self) -> None:
        self.window.update_idletasks()
        parent_x = self.app.root.winfo_rootx()
        parent_y = self.app.root.winfo_rooty()
        parent_w = self.app.root.winfo_width()
        parent_h = self.app.root.winfo_height()
        width = self.window.winfo_width()
        height = self.window.winfo_height()
        x = parent_x + max((parent_w - width) // 2, 0)
        y = parent_y + max((parent_h - height) // 2, 0)
        self.window.geometry(f"{width}x{height}+{x}+{y}")

    def get_title(self) -> str:
        if self.mode == "capacity":
            return "작업 축 수 확인"
        return "초품 4Point 확인" if self.mode == "first" else "하부 Pin 4개 확인"

    def get_value_index(self, axis_index: int) -> int:
        return axis_index if self.mode in ("first", "capacity") else axis_index + 6

    def get_visible_axes(self) -> list[int]:
        if self.allowed_axes is None:
            return list(range(6))
        return [axis_index for axis_index in range(6) if axis_index in self.allowed_axes]

    def create_ui(self) -> None:
        title = tk.Label(
            self.window,
            text=self.get_title(),
            bg=PRIMARY_LIGHT,
            fg=PRIMARY,
            font=("맑은 고딕", 14, "bold"),
            height=2,
        )
        title.pack(fill=tk.X, padx=14, pady=(14, 8))

        body = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        body.pack(fill=tk.BOTH, expand=True, padx=14, pady=8)

        tk.Label(body, text=self.get_title(), bg="#d99a9a", fg="#111827", font=("맑은 고딕", 10, "bold"), height=2).grid(row=0, column=0, columnspan=len(self.visible_axes), sticky="ew")

        for column_index, axis_index in enumerate(self.visible_axes):
            label = self.LABELS[axis_index]
            value_index = self.get_value_index(axis_index)
            button = tk.Button(
                body,
                text=f"{label}\n{'√' if self.values[value_index] == 'OK' and self.mode in ('jig', 'capacity') else ('OK' if self.values[value_index] == 'OK' else '클릭')}",
                command=lambda i=axis_index: self.toggle(i),
                bg=self.get_ok_color() if self.values[value_index] == "OK" else SURFACE_BG,
                fg="#ffffff" if self.values[value_index] == "OK" else TEXT_COLOR,
                relief=tk.SOLID,
                bd=1,
                width=8,
                height=3,
                cursor="hand2",
                font=("맑은 고딕", 9, "bold"),
                highlightthickness=1,
                highlightbackground="#cbd5e1",
                takefocus=0,
            )
            button.grid(row=1, column=column_index, sticky="nsew", padx=2, pady=8)
            body.columnconfigure(column_index, weight=1)
            self.buttons.append(button)

        bottom = ttk.Frame(self.window, padding=(14, 4, 14, 14))
        bottom.pack(fill=tk.X)
        ttk.Button(bottom, text="초기화", command=self.clear).pack(side=tk.LEFT)
        ttk.Button(bottom, text="저장", command=self.save, style="Primary.TButton").pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(bottom, text="취소", command=self.window.destroy).pack(side=tk.RIGHT)

    def toggle(self, axis_index: int) -> None:
        if not self.clicks_enabled:
            return
        if self.allowed_axes is not None and axis_index not in self.allowed_axes:
            return
        value_index = self.get_value_index(axis_index)
        self.values[value_index] = "" if self.values[value_index] == "OK" else "OK"
        self.refresh_button(axis_index)

    def refresh_button(self, axis_index: int) -> None:
        if axis_index not in self.visible_axes:
            return
        button_index = self.visible_axes.index(axis_index)
        value_index = self.get_value_index(axis_index)
        ok = self.values[value_index] == "OK"
        ok_text = "√" if self.mode in ("jig", "capacity") else "OK"
        self.buttons[button_index].configure(
            text=f"{self.LABELS[axis_index]}\n{ok_text if ok else '클릭'}",
            bg=self.get_ok_color() if ok else SURFACE_BG,
            fg="#ffffff" if ok else TEXT_COLOR,
        )

    def get_ok_color(self) -> str:
        return self.FIRST_OK_COLOR if self.mode == "first" else self.JIG_OK_COLOR

    def clear(self) -> None:
        target_range = range(0, 6) if self.mode in ("first", "capacity") else range(6, 12)
        for index in target_range:
            self.values[index] = ""
        for axis_index in self.visible_axes:
            self.refresh_button(axis_index)

    def save(self) -> None:
        if self.mode == "first":
            checked_axes = [
                axis_index
                for axis_index in self.visible_axes
                if self.values[self.get_value_index(axis_index)] == "OK"
            ]
            if len(checked_axes) != len(self.visible_axes):
                show_operator_alert(self.window, self.get_title(), "표시된 축 모두 확인 필요")
                return
        check_mode = "first" if self.mode == "capacity" else self.mode
        ok, message = validate_frequent_check_values(self.values, check_mode=check_mode)
        if not ok:
            show_operator_alert(self.window, self.get_title(), message)
            return
        self.app.frequent_check_values = self.values[:]
        if self.mode == "capacity":
            self.app.set_status("dnc", "작업 축 수 확인 완료", True)
        else:
            self.app.set_status("dnc", message, True)
        self.saved = True
        self.window.destroy()


class WorkHistoryPopup:
    """DB Browser 없이 프로그램 안에서 공정별 작업 이력을 확인하는 보기 전용 창입니다."""

    def __init__(self, app: JiinDncManager, process_name: str = "KCC PKG"):
        self.app = app
        self.process_name = process_name
        self.window = tk.Toplevel(app.root)
        self.window.title(f"{process_name} 작업 이력 보기")
        self.window.geometry("1320x660")
        self.window.configure(bg=APP_BG)
        self.keyword_var = tk.StringVar()
        self.limit_var = tk.StringVar(value="500")
        self.only_unexported_var = tk.BooleanVar(value=False)
        self.only_incomplete_var = tk.BooleanVar(value=False)
        self.summary_var = tk.StringVar(value="")
        self.create_ui()
        self.refresh()

    def create_ui(self) -> None:
        top = ttk.Frame(self.window, padding=(12, 12, 12, 8))
        top.pack(fill=tk.X)
        ttk.Label(top, text="검색", width=6).pack(side=tk.LEFT)
        ttk.Entry(top, textvariable=self.keyword_var, style="Wide.TEntry", width=28).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Checkbutton(top, text="Excel 미반영만", variable=self.only_unexported_var, command=self.refresh).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Checkbutton(top, text="미완료만", variable=self.only_incomplete_var, command=self.refresh).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(top, text="표시 수").pack(side=tk.LEFT, padx=(10, 4))
        ttk.Combobox(top, textvariable=self.limit_var, values=["100", "500", "1000", "3000", "5000"], state="readonly", width=8).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(top, text="조회", command=self.refresh, style="Primary.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(top, text="닫기", command=self.window.destroy).pack(side=tk.RIGHT)

        ttk.Label(self.window, textvariable=self.summary_var, background=APP_BG, foreground=MUTED_TEXT).pack(fill=tk.X, padx=14, pady=(0, 6))

        body = ttk.Frame(self.window, padding=(12, 0, 12, 12))
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        columns = (
            "id", "dnc_type", "status", "exported", "machine", "work_date", "shift_group", "shift_name",
            "worker", "step", "round_no", "manage_no", "lot_no", "qty_text", "result_value",
            "condition_name", "jig", "stack", "model_change_text", "burr_result", "created_at", "record_time", "exported_at",
        )
        self.tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="browse")
        headings = {
            "id": "ID",
            "dnc_type": "구분",
            "status": "상태",
            "exported": "Excel",
            "machine": "호기",
            "work_date": "작업일자",
            "shift_group": "조",
            "shift_name": "근무",
            "worker": "작업자",
            "step": "STEP",
            "round_no": "차수",
            "manage_no": "관리번호",
            "lot_no": "LOT",
            "qty_text": "매수",
            "result_value": "실적",
            "condition_name": "작업조건",
            "jig": "지그",
            "stack": "Stack",
            "model_change_text": "기종/검증",
            "burr_result": "Burr",
            "created_at": "DNC 시작시간",
            "record_time": "DNC 완료시간",
            "exported_at": "엑셀 반영시간",
        }
        widths = {
            "id": 60,
            "dnc_type": 80,
            "status": 80,
            "exported": 80,
            "machine": 70,
            "work_date": 100,
            "shift_group": 50,
            "shift_name": 70,
            "worker": 90,
            "step": 70,
            "round_no": 70,
            "manage_no": 130,
            "lot_no": 150,
            "qty_text": 70,
            "result_value": 70,
            "condition_name": 240,
            "jig": 80,
            "stack": 70,
            "model_change_text": 90,
            "burr_result": 90,
            "created_at": 150,
            "record_time": 110,
            "exported_at": 150,
        }
        for column in columns:
            self.tree.heading(column, text=headings[column])
            self.tree.column(column, width=widths[column], anchor="center" if column not in ("condition_name", "lot_no", "manage_no") else "w")
        self.tree.grid(row=0, column=0, sticky="nsew")
        y_scroll = ttk.Scrollbar(body, orient=tk.VERTICAL, command=self.tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(body, orient=tk.HORIZONTAL, command=self.tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

    def refresh(self) -> None:
        try:
            limit = int(self.limit_var.get())
        except ValueError:
            limit = 500
            self.limit_var.set("500")
        rows = load_work_history(
            limit=limit,
            only_unexported=self.only_unexported_var.get(),
            only_incomplete=self.only_incomplete_var.get(),
            keyword=self.keyword_var.get(),
            process_name=self.process_name,
        )
        self.tree.delete(*self.tree.get_children())
        for row in rows:
            exported_text = "반영" if int(row["exported"] or 0) == 1 else "미반영"
            values = (
                row["id"],
                row["dnc_type"] or "",
                row["status"] or "",
                exported_text,
                row["machine"] or "",
                row["work_date"] or "",
                row["shift_group"] or "",
                row["shift_name"] or "",
                row["worker"] or "",
                row["step"] or "",
                row["round_no"] or "",
                row["manage_no"] or "",
                row["lot_no"] or "",
                row["qty_text"] or "",
                "" if row["result_value"] is None else row["result_value"],
                row["condition_name"] or "",
                row["jig"] or "",
                row["stack"] or "",
                row["model_change_text"] or "",
                row["burr_result"] or "",
                row["created_at"] or "",
                row["record_time"] or "",
                row["exported_at"] or "",
            )
            self.tree.insert("", "end", values=values)
        pending_count = get_unexported_process_count(self.process_name)
        incomplete_count = get_incomplete_process_count(self.process_name)
        self.summary_var.set(
            f"{self.process_name} 최근 {len(rows)}건 표시 / Excel 미반영 {pending_count}건 / 미완료 {incomplete_count}건"
        )


class MasterSettingsPopup:
    """작업자가 실수로 만지면 위험한 점검/관리 기능을 모아둔 관리자 전용 창입니다."""

    def __init__(self, app: JiinDncManager):
        self.app = app
        self.window = tk.Toplevel(app.root)
        self.window.title("마스터 설정")
        self.window.geometry("760x430")
        self.window.minsize(760, 430)
        self.window.configure(bg=APP_BG)
        self.delete_seconds_var = tk.StringVar(value=str(app.config.get("dnc_delete_seconds", DNC_DELETE_SECONDS)))
        self.first_article_wait_var = tk.StringVar(value=str(app.config.get("first_article_wait_seconds", FIRST_ARTICLE_WAIT_SECONDS)))
        self.create_ui()
        self.window.protocol("WM_DELETE_WINDOW", self.close)
        self.window.after_idle(self.window.focus_set)

    def create_ui(self) -> None:
        panel = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        panel.pack(fill=tk.BOTH, expand=True, padx=14, pady=14)
        panel.columnconfigure(1, weight=1)
        tk.Label(
            panel,
            text="마스터 설정",
            bg=PRIMARY_LIGHT,
            fg=PRIMARY,
            font=("맑은 고딕", 14, "bold"),
            height=2,
        ).grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 12))

        ttk.Label(panel, text="삭제 대기 시간", background=SURFACE_BG, width=16).grid(row=1, column=0, sticky="e", padx=10, pady=8)
        ttk.Entry(panel, textvariable=self.delete_seconds_var, style="Wide.TEntry").grid(row=1, column=1, sticky="ew", padx=8, pady=8)
        ttk.Label(panel, text="초", background=SURFACE_BG).grid(row=1, column=2, sticky="w", padx=(0, 10), pady=8)

        ttk.Label(panel, text="초품 알람 시간", background=SURFACE_BG, width=16).grid(row=2, column=0, sticky="e", padx=10, pady=8)
        ttk.Entry(panel, textvariable=self.first_article_wait_var, style="Wide.TEntry").grid(row=2, column=1, sticky="ew", padx=8, pady=8)
        ttk.Label(panel, text="초", background=SURFACE_BG).grid(row=2, column=2, sticky="w", padx=(0, 10), pady=8)

        buttons = ttk.Frame(panel)
        buttons.grid(row=3, column=0, columnspan=3, sticky="ew", padx=10, pady=(18, 8))
        buttons.columnconfigure((0, 1), weight=1)
        ttk.Button(buttons, text="작업 이력 보기", command=self.app.open_work_history_popup).grid(row=0, column=0, sticky="ew", padx=4, pady=4)
        ttk.Button(buttons, text="마스터 복구", command=self.rebuild_condition_master).grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        ttk.Button(buttons, text="마스터 설정 비번 변경", command=lambda: self.change_password("master_password", "마스터 설정", MASTER_SETTINGS_PASSWORD)).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
        ttk.Button(buttons, text="조건 마스터 비번 변경", command=lambda: self.change_password("condition_master_password", "조건 마스터 관리", CONDITION_MASTER_PASSWORD)).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
        tk.Button(
            buttons,
            text="라이선스 관리",
            command=self.open_license_settings,
            bg=NG_COLOR,
            fg="#ffffff",
            activebackground="#991b1b",
            activeforeground="#ffffff",
            relief=tk.SOLID,
            bd=1,
            font=("맑은 고딕", 10, "bold"),
            cursor="hand2",
        ).grid(row=2, column=0, columnspan=2, sticky="ew", padx=4, pady=(10, 4), ipady=6)

        bottom = ttk.Frame(panel)
        bottom.grid(row=4, column=0, columnspan=3, sticky="ew", padx=10, pady=(22, 10))
        bottom.columnconfigure(0, weight=1)
        ttk.Button(bottom, text="닫기", command=self.close, width=18).grid(row=0, column=1, sticky="e")

    def save_master_settings(self, show_message: bool = False) -> bool:
        ok, message = validate_positive_number(self.delete_seconds_var.get(), "삭제 대기 시간", required=True)
        if not ok:
            show_operator_alert(self.window, "마스터 설정", message)
            return False
        ok, message = validate_positive_number(self.first_article_wait_var.get(), "초품 알람 시간", required=True)
        if not ok:
            show_operator_alert(self.window, "마스터 설정", message)
            return False
        self.app.config["dnc_delete_seconds"] = int(self.delete_seconds_var.get().strip())
        self.app.config["first_article_wait_seconds"] = int(self.first_article_wait_var.get().strip())
        if hasattr(self.app, "delete_seconds_var"):
            self.app.delete_seconds_var.set(self.delete_seconds_var.get().strip())
        if hasattr(self.app, "first_article_wait_var"):
            self.app.first_article_wait_var.set(self.first_article_wait_var.get().strip())
        save_config(self.app.config)
        if show_message:
            show_operator_alert(self.window, "적용 완료", "마스터 설정 적용", "info")
        return True

    def close(self) -> None:
        if self.save_master_settings(show_message=False):
            self.app.master_settings_popup = None
            self.window.destroy()

    def rebuild_condition_master(self) -> None:
        if not self.save_master_settings(show_message=False):
            return
        try:
            count = rebuild_condition_master_from_log(self.app.config)
        except Exception as exc:
            show_operator_alert(self.window, "조건 복구 실패", str(exc), "error")
            return
        show_operator_alert(self.window, "조건 복구 완료", f"{count}개 조건 저장", "info")

    def open_license_settings(self) -> None:
        password = ask_system_input(self.window, "라이선스 관리", "라이선스 비밀번호 입력", show="*")
        if password is None:
            return
        if password != str(self.app.config.get("license_password", LICENSE_PASSWORD)):
            show_operator_alert(self.window, "라이선스 관리", "비밀번호 불일치")
            return
        LicenseSettingsPopup(self.app, self.window)

    def change_password(self, config_key: str, title: str, default_password: str) -> None:
        current_password = str(self.app.config.get(config_key, default_password))
        old_password = ask_system_input(self.window, f"{title} 비밀번호 변경", "현재 비밀번호 입력", show="*")
        if old_password is None:
            return
        if old_password != current_password:
            show_operator_alert(self.window, "비밀번호 확인", "현재 비밀번호 불일치")
            return
        new_password = ask_system_input(self.window, f"{title} 비밀번호 변경", "새 비밀번호 입력", show="*")
        if new_password is None:
            return
        new_password = new_password.strip()
        if not new_password:
            show_operator_alert(self.window, "비밀번호 확인", "새 비밀번호 입력 없음")
            return
        confirm_password = ask_system_input(self.window, f"{title} 비밀번호 변경", "새 비밀번호 재입력", show="*")
        if confirm_password is None:
            return
        if new_password != confirm_password.strip():
            show_operator_alert(self.window, "비밀번호 확인", "새 비밀번호 불일치")
            return
        self.app.config[config_key] = new_password
        save_config(self.app.config)
        show_operator_alert(self.window, "변경 완료", f"{title} 비밀번호 변경", "info")


class LicenseSettingsPopup:
    """마스터 PC와 허용 IP 대역을 관리하는 관리자 전용 창입니다."""

    def __init__(self, app: JiinDncManager, parent: tk.Toplevel):
        self.app = app
        self.parent = parent
        self.window = tk.Toplevel(parent)
        self.window.title("라이선스 관리")
        self.window.geometry("700x460")
        self.window.minsize(700, 460)
        self.window.configure(bg=APP_BG)
        self.master_pc_var = tk.StringVar(value=str(app.config.get("license_master_pc_name", DEFAULT_MASTER_PC_NAME)))
        self.create_ui()
        self.window.transient(parent)
        self.window.grab_set()

    def create_ui(self) -> None:
        panel = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground=NG_COLOR, bd=0)
        panel.pack(fill=tk.BOTH, expand=True, padx=14, pady=14)
        panel.columnconfigure(1, weight=1)
        tk.Label(
            panel,
            text="라이선스 관리",
            bg=NG_COLOR,
            fg="#ffffff",
            font=("맑은 고딕", 14, "bold"),
            height=2,
        ).grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 14))

        ttk.Label(panel, text="현재 PC 이름", background=SURFACE_BG, width=16).grid(row=1, column=0, sticky="e", padx=10, pady=7)
        ttk.Label(panel, text=get_current_pc_name(), background=SURFACE_BG, foreground=PRIMARY, font=("맑은 고딕", 10, "bold")).grid(row=1, column=1, sticky="w", padx=8, pady=7)

        ttk.Label(panel, text="현재 IP", background=SURFACE_BG, width=16).grid(row=2, column=0, sticky="e", padx=10, pady=7)
        current_ips = ", ".join(get_current_ip_addresses()) or "확인 안됨"
        ttk.Label(panel, text=current_ips, background=SURFACE_BG, foreground=MUTED_TEXT).grid(row=2, column=1, sticky="w", padx=8, pady=7)

        ttk.Label(panel, text="마스터 PC 이름", background=SURFACE_BG, width=16).grid(row=3, column=0, sticky="e", padx=10, pady=7)
        ttk.Entry(panel, textvariable=self.master_pc_var, style="Wide.TEntry").grid(row=3, column=1, sticky="ew", padx=8, pady=7)
        ttk.Label(panel, text="예: KUKJIN", background=SURFACE_BG, foreground=MUTED_TEXT).grid(row=3, column=2, sticky="w", padx=(0, 10), pady=7)

        ttk.Label(panel, text="허용 IP 대역", background=SURFACE_BG, width=16).grid(row=4, column=0, sticky="ne", padx=10, pady=7)
        self.ip_text = tk.Text(panel, height=7, font=("맑은 고딕", 10), relief=tk.SOLID, bd=1, highlightthickness=0)
        self.ip_text.grid(row=4, column=1, sticky="nsew", padx=8, pady=7)
        panel.rowconfigure(4, weight=1)
        prefixes = normalize_ip_prefixes(self.app.config.get("license_allowed_ip_prefixes", DEFAULT_ALLOWED_IP_PREFIXES))
        self.ip_text.insert("1.0", "\n".join(prefixes))
        ttk.Label(
            panel,
            text="한 줄에 하나씩\n빈칸이면 모든 PC 허용",
            background=SURFACE_BG,
            foreground=MUTED_TEXT,
        ).grid(row=4, column=2, sticky="nw", padx=(0, 10), pady=7)

        buttons = ttk.Frame(panel)
        buttons.grid(row=5, column=0, columnspan=3, sticky="ew", padx=10, pady=(14, 8))
        buttons.columnconfigure((0, 1, 2), weight=1)
        ttk.Button(buttons, text="저장", command=self.save).grid(row=0, column=0, sticky="ew", padx=4)
        ttk.Button(buttons, text="라이선스 비번 변경", command=self.change_license_password).grid(row=0, column=1, sticky="ew", padx=4)
        ttk.Button(buttons, text="닫기", command=self.window.destroy).grid(row=0, column=2, sticky="ew", padx=4)

    def save(self) -> None:
        master_pc = self.master_pc_var.get().strip().upper()
        ip_prefixes = normalize_ip_prefixes(self.ip_text.get("1.0", tk.END))
        self.app.config["license_master_pc_name"] = master_pc
        self.app.config["license_allowed_ip_prefixes"] = ip_prefixes
        save_config(self.app.config)
        show_operator_alert(self.window, "저장 완료", "라이선스 설정 저장", "info")

    def change_license_password(self) -> None:
        current_password = str(self.app.config.get("license_password", LICENSE_PASSWORD))
        old_password = ask_system_input(self.window, "라이선스 비번 변경", "현재 비밀번호 입력", show="*")
        if old_password is None:
            return
        if old_password != current_password:
            show_operator_alert(self.window, "라이선스 비번 변경", "현재 비밀번호 불일치")
            return
        new_password = ask_system_input(self.window, "라이선스 비번 변경", "새 비밀번호 입력", show="*")
        if new_password is None:
            return
        new_password = new_password.strip()
        if not new_password:
            show_operator_alert(self.window, "라이선스 비번 변경", "새 비밀번호 입력 없음")
            return
        confirm_password = ask_system_input(self.window, "라이선스 비번 변경", "새 비밀번호 재입력", show="*")
        if confirm_password is None:
            return
        if new_password != confirm_password.strip():
            show_operator_alert(self.window, "라이선스 비번 변경", "새 비밀번호 불일치")
            return
        self.app.config["license_password"] = new_password
        save_config(self.app.config)
        show_operator_alert(self.window, "변경 완료", "라이선스 비밀번호 변경", "info")


class ConditionMasterPopup:
    """작업일보에서 추출한 조건 마스터를 보고 수정하는 창입니다."""

    def __init__(self, app: JiinDncManager):
        self.app = app
        self.records = load_condition_master()
        self.search_var = tk.StringVar()
        self.show_duplicates_only = False
        self.duplicate_keys: set[str] = set()
        self.window = tk.Toplevel(app.root)
        self.window.title("KCC PKG 조건 마스터 관리")
        self.window.geometry("980x620")
        self.window.configure(bg=APP_BG)
        self.create_ui()
        self.refresh_tree()
        self.window.protocol("WM_DELETE_WINDOW", self.close)
        self.window.after_idle(self.window.focus_set)

    def close(self) -> None:
        self.app.condition_master_popup = None
        self.window.destroy()

    def create_ui(self) -> None:
        top = ttk.Frame(self.window, padding=(12, 12, 12, 8))
        top.pack(fill=tk.X)
        ttk.Button(top, text="작업일보 마스터 갱신", command=self.rebuild_from_log, style="Primary.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(top, text="선택 수정 저장", command=self.save_selected_edit).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(top, text="선택 삭제", command=self.delete_selected_record).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(top, text="닫기", command=self.window.destroy).pack(side=tk.RIGHT)

        search = ttk.Frame(self.window, padding=(12, 0, 12, 8))
        search.pack(fill=tk.X)
        ttk.Label(search, text="조회").pack(side=tk.LEFT, padx=(0, 6))
        search_entry = ttk.Entry(search, textvariable=self.search_var, style="Wide.TEntry", width=42)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(search, text="조회", command=self.refresh_tree).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(search, text="중복 조건만", command=self.toggle_duplicate_view).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(search, text="전체 보기", command=self.clear_search).pack(side=tk.LEFT)
        self.search_var.trace_add("write", lambda *_args: self.refresh_tree())

        body = ttk.Frame(self.window, padding=(12, 0, 12, 8))
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        columns = ("step", "round", "manage_no", "process_code", "condition", "jig", "source")
        self.tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="browse")
        headings = {
            "step": "STEP",
            "round": "차수",
            "manage_no": "관리번호",
            "process_code": "공정코드",
            "condition": "작업조건",
            "jig": "지그",
            "source": "출처",
        }
        widths = {
            "step": 80,
            "round": 70,
            "manage_no": 150,
            "process_code": 110,
            "condition": 220,
            "jig": 140,
            "source": 140,
        }
        for column in columns:
            self.tree.heading(column, text=headings[column])
            self.tree.column(column, width=widths[column], anchor="w")
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.tag_configure("duplicate", background="#fee2e2", foreground=NG_COLOR)
        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        scroll = ttk.Scrollbar(body, orient=tk.VERTICAL, command=self.tree.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scroll.set)

        edit = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        edit.pack(fill=tk.X, padx=12, pady=(0, 12))
        self.edit_vars = {
            "step": tk.StringVar(),
            "round": tk.StringVar(),
            "manage_no": tk.StringVar(),
            "process_code": tk.StringVar(),
            "condition": tk.StringVar(),
            "jig": tk.StringVar(),
        }
        edit_fields = [
            ("step", "STEP"),
            ("round", "차수"),
            ("manage_no", "관리번호"),
            ("process_code", "공정코드"),
            ("condition", "작업조건"),
            ("jig", "지그"),
        ]
        for index, (key, label) in enumerate(edit_fields):
            ttk.Label(edit, text=label, background=SURFACE_BG).grid(row=index // 3, column=(index % 3) * 2, sticky="e", padx=(10, 4), pady=8)
            ttk.Entry(edit, textvariable=self.edit_vars[key], style="Wide.TEntry", width=24).grid(row=index // 3, column=(index % 3) * 2 + 1, sticky="ew", padx=(0, 10), pady=8)
            edit.columnconfigure((index % 3) * 2 + 1, weight=1)

    def refresh_tree(self) -> None:
        self.tree.delete(*self.tree.get_children())
        self.duplicate_keys = get_duplicate_condition_keys(self.records)
        keyword = self.search_var.get().strip().lower()
        for index, record in enumerate(self.records):
            is_duplicate = make_condition_record_key(record) in self.duplicate_keys
            if self.show_duplicates_only and not is_duplicate:
                continue
            if keyword and not self.record_matches_keyword(record, keyword):
                continue
            self.tree.insert(
                "",
                "end",
                iid=str(index),
                values=(
                    record.get("step", ""),
                    record.get("round", ""),
                    record.get("manage_no", ""),
                    record.get("process_code", ""),
                    record.get("condition", ""),
                    record.get("jig", ""),
                    record.get("source", ""),
                ),
                tags=("duplicate",) if is_duplicate else (),
            )

    def toggle_duplicate_view(self) -> None:
        self.show_duplicates_only = not self.show_duplicates_only
        self.refresh_tree()

    def clear_search(self) -> None:
        self.search_var.set("")
        self.show_duplicates_only = False
        self.refresh_tree()

    def record_matches_keyword(self, record: dict, keyword: str) -> bool:
        searchable = [
            record.get("step", ""),
            record.get("round", ""),
            record.get("manage_no", ""),
            record.get("process_code", ""),
            record.get("condition", ""),
            record.get("jig", ""),
            record.get("source", ""),
        ]
        return any(keyword in str(value).lower() for value in searchable)

    def on_select(self, _event=None) -> None:
        selected = self.tree.selection()
        if not selected:
            return
        record = self.records[int(selected[0])]
        for key, var in self.edit_vars.items():
            var.set(record.get(key, ""))

    def save_selected_edit(self) -> None:
        selected = self.tree.selection()
        if not selected:
            show_operator_alert(self.window, "선택 필요", "수정할 조건 선택")
            return
        index = int(selected[0])
        record = self.records[index]
        for key, var in self.edit_vars.items():
            record[key] = var.get().strip()
        record["source"] = "사용자 수정"
        record["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_condition_master(self.records)
        self.refresh_tree()
        self.tree.selection_set(str(index))
        show_operator_alert(self.window, "저장 완료", "조건 마스터 수정", "info")

    def delete_selected_record(self) -> None:
        selected = self.tree.selection()
        if not selected:
            show_operator_alert(self.window, "선택 필요", "삭제할 조건 선택")
            return
        index = int(selected[0])
        record = self.records[index]
        ok = ask_system_yes_no(
            self.window,
            "조건 삭제 확인",
            f"조건 삭제\n{record.get('manage_no', '')}",
        )
        if not ok:
            return
        del self.records[index]
        save_condition_master(self.records)
        self.records = load_condition_master()
        self.refresh_tree()
        show_operator_alert(self.window, "삭제 완료", "조건 삭제 완료", "info")

    def rebuild_from_log(self) -> None:
        self.app.save_settings_from_ui_silent()
        try:
            count = rebuild_condition_master_from_log(self.app.config)
        except Exception as exc:
            show_operator_alert(self.window, "갱신 실패", str(exc), "error")
            return
        self.records = load_condition_master()
        self.refresh_tree()
        show_operator_alert(self.window, "작업일보 마스터 갱신", f"신규 등록 {count}건", "info")



class NewModelPopup:
    def __init__(self, app: JiinDncManager):
        self.app = app
        self.window = tk.Toplevel(app.root)
        self.window.title("KCC PKG 신규 모델 검증 DNC")
        self.window.geometry("1280x760")
        self.window.minsize(1180, 730)
        self.window.configure(bg=APP_BG)
        self.entries: dict[str, LabeledEntry] = {}
        self.both_entries: dict[str, dict[str, LabeledEntry]] = {}
        self.buttons: list[ttk.Button] = []
        self.mode_buttons: dict[str, tk.Button] = {}
        self.is_loading_fields = False
        self.selected_lot = tk.StringVar(value="lot1")
        self.run_mode = tk.StringVar(value="lot1")
        empty_lot = {
            key: ""
            for key in ["step", "round", "manage_no", "lot_no", "qty", "process_code", "condition", "jig"]
        }
        self.lot_drafts = {
            "lot1": dict(empty_lot),
            "lot2": dict(empty_lot),
        }
        self.target_lots = self.get_new_model_target_lots()
        self.lock_imported_fields = any(
            self.app.lot_has_any_value(self.app.get_lot_data(entries))
            for entries in (self.app.lot1_entries, self.app.lot2_entries)
        )
        if not self.target_lots:
            show_operator_alert(
                app.root,
                "신규 모델 검증 DNC",
                "신규 검증 대상 없음",
                "info",
            )
            self.window.destroy()
            return
        self.selected_lot.set(self.target_lots[0])
        self.run_mode.set(self.target_lots[0])
        self.is_running = False
        self.create_ui()
        self.load_lot_drafts_from_main()
        self.load_selected_lot()
        self.load_both_lot_panels()
        self.refresh_input_mode()
        self.update_checks()
        self.window.protocol("WM_DELETE_WINDOW", self.close)

    def close(self) -> None:
        if self.is_running or self.app.is_running:
            show_operator_alert(self.window, "DNC 진행중", "작업 완료 후 종료")
            return
        self.window.destroy()

    def get_new_model_target_lots(self) -> list[str]:
        """메인 화면에서 조건/지그가 없는 LOT만 신규 검증 대상으로 반환합니다."""
        targets: list[str] = []
        used_lot_count = 0
        for lot_key, entries in (("lot1", self.app.lot1_entries), ("lot2", self.app.lot2_entries)):
            lot = self.app.get_lot_data(entries)
            if not self.app.lot_has_any_value(lot):
                continue
            used_lot_count += 1
            condition_ok, _message = get_single_condition_message(lot)
            if not condition_ok:
                targets.append(lot_key)
        if targets:
            return targets
        if used_lot_count == 0:
            return ["lot1", "lot2"]
        return []

    def create_ui(self) -> None:
        title = tk.Label(self.window, text="KCC PKG 신규 모델 검증 DNC", bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 14, "bold"), height=2)
        title.pack(fill=tk.X, padx=14, pady=(14, 8))

        lot_select = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        lot_select.pack(fill=tk.X, padx=14, pady=(0, 8))
        if len(self.target_lots) == 1:
            lot_name = "LOT 1" if self.target_lots[0] == "lot1" else "LOT 2"
            tk.Label(
                lot_select,
                text=f"{lot_name} 입력",
                bg=SURFACE_BG,
                fg=PRIMARY,
                font=("맑은 고딕", 10, "bold"),
            ).pack(side=tk.LEFT, padx=14, pady=8)
        else:
            choices = [("lot1", "LOT 1 입력"), ("lot2", "LOT 2 입력"), ("both", "LOT 1 + LOT 2")]
            for value, label in choices:
                self.create_mode_button(lot_select, label, value).pack(side=tk.LEFT, padx=4, pady=8)

        self.input_area = tk.Frame(self.window, bg=APP_BG)
        self.input_area.pack(fill=tk.BOTH, expand=True, padx=14, pady=8)

        self.single_panel = tk.Frame(self.input_area, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        self.entries = self.create_lot_input_fields(self.single_panel, columns_per_row=2)

        self.both_panel = tk.Frame(self.input_area, bg=APP_BG)
        self.both_panel.columnconfigure(0, weight=1)
        self.both_panel.columnconfigure(1, weight=1)
        for column, (lot_key, title_text) in enumerate((("lot1", "LOT 1 신규 입력"), ("lot2", "LOT 2 신규 입력"))):
            lot_panel = tk.Frame(self.both_panel, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
            lot_panel.grid(row=0, column=column, sticky="nsew", padx=(0, 8) if column == 0 else (8, 0))
            tk.Label(lot_panel, text=title_text, bg=PRIMARY_LIGHT, fg=PRIMARY, font=("맑은 고딕", 11, "bold"), height=2).pack(fill=tk.X)
            fields_frame = tk.Frame(lot_panel, bg=SURFACE_BG)
            fields_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            self.both_entries[lot_key] = self.create_lot_input_fields(fields_frame, columns_per_row=1)

        status = tk.Frame(self.window, bg=SURFACE_BG, highlightthickness=1, highlightbackground="#93c5fd", bd=0)
        status.pack(fill=tk.X, padx=14, pady=8)
        self.mes_label = tk.Label(status, text="MES Core 일치화: 대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 10, "bold"))
        self.mes_label.pack(side=tk.LEFT, padx=16, pady=10)
        self.condition_label = tk.Label(status, text="조건 적용 확인: 대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 10, "bold"))
        self.condition_label.pack(side=tk.LEFT, padx=16, pady=10)
        self.dnc_label = tk.Label(status, text="DNC 진행 상태: 대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 10, "bold"))
        self.dnc_label.pack(side=tk.LEFT, padx=16, pady=10)
        self.excel_label = tk.Label(status, text="작업일보 반영: 대기중", bg=SURFACE_BG, fg=MUTED_TEXT, font=("맑은 고딕", 10, "bold"))
        self.excel_label.pack(side=tk.LEFT, padx=16, pady=10)

        buttons = ttk.Frame(self.window)
        buttons.pack(fill=tk.X, padx=14, pady=(4, 14))
        self.add_button(buttons, "신규 모델 DNC 실행", self.run_new_model_dnc, "Primary.TButton").pack(side=tk.LEFT, padx=4)
        self.add_button(buttons, "입력 초기화", self.clear_inputs).pack(side=tk.LEFT, padx=4)
        self.add_button(buttons, "닫기", self.window.destroy).pack(side=tk.RIGHT, padx=4)

    def add_button(self, parent, text, command, style="TButton") -> ttk.Button:
        button = ttk.Button(parent, text=text, command=command, style=style, width=20)
        self.buttons.append(button)
        return button

    def create_mode_button(self, parent, text: str, value: str) -> tk.Button:
        button = tk.Button(
            parent,
            text=text,
            command=lambda selected=value: self.set_run_mode(selected),
            relief=tk.FLAT,
            bd=0,
            padx=16,
            pady=6,
            cursor="hand2",
            font=("맑은 고딕", 10, "bold"),
        )
        self.mode_buttons[value] = button
        self.update_mode_buttons()
        return button

    def update_mode_buttons(self) -> None:
        for value, button in self.mode_buttons.items():
            selected = value == self.run_mode.get()
            button.configure(
                bg=PRIMARY_LIGHT if selected else SURFACE_BG,
                fg=PRIMARY if selected else TEXT_COLOR,
                highlightthickness=1,
                highlightbackground=PRIMARY if selected else BORDER_COLOR,
            )

    def create_lot_input_fields(self, parent, columns_per_row: int = 2) -> dict[str, LabeledEntry]:
        fields = [
            ("step", "STEP"),
            ("round", "차수"),
            ("manage_no", "관리번호"),
            ("lot_no", "LOT No"),
            ("process_code", "공정코드"),
            ("qty", "매수"),
            ("condition", "작업조건"),
            ("jig", "지그"),
        ]
        readonly_keys = {"step", "round", "manage_no", "lot_no", "process_code"} if self.lock_imported_fields else set()
        entries: dict[str, LabeledEntry] = {}
        for index, (key, label) in enumerate(fields):
            if key == "round":
                entry = RoundField(parent, label)
                entry.set_readonly(key in readonly_keys)
            elif key in {"step", "qty"}:
                entry = LabeledEntry(
                    parent,
                    label,
                    width=24,
                    numeric_only=True,
                    readonly=key in readonly_keys,
                    style="Lookup.TEntry" if key in readonly_keys else "Wide.TEntry",
                )
            elif key == "condition":
                entry = LabeledEntry(parent, label, width=24)
            else:
                entry = LabeledEntry(
                    parent,
                    label,
                    width=24,
                    uppercase=True,
                    readonly=key in readonly_keys,
                    style="Lookup.TEntry" if key in readonly_keys else "Wide.TEntry",
                )
            row = index // columns_per_row
            column = index % columns_per_row
            entry.grid(row=row, column=column, sticky="ew", padx=10, pady=8)
            parent.columnconfigure(column, weight=1)
            entries[key] = entry
            if hasattr(entry, "var"):
                entry.var.trace_add("write", lambda *_args: self.update_checks())
        return entries

    def read_entry_group(self, entries: dict[str, LabeledEntry]) -> dict:
        return {
            key: entries[key].get()
            for key in ["step", "round", "manage_no", "lot_no", "qty", "process_code", "condition", "jig"]
        }

    def fill_entry_group(self, entries: dict[str, LabeledEntry], draft: dict) -> None:
        self.is_loading_fields = True
        try:
            for key, entry in entries.items():
                entry.set(draft.get(key, ""))
        finally:
            self.is_loading_fields = False

    def save_current_lot_draft(self) -> None:
        if not self.entries or self.is_loading_fields:
            return
        if self.run_mode.get() == "both":
            for lot_key, entries in self.both_entries.items():
                self.lot_drafts[lot_key] = self.read_entry_group(entries)
            return
        self.lot_drafts[self.selected_lot.get()] = self.read_entry_group(self.entries)

    def load_selected_lot(self) -> None:
        draft = self.lot_drafts.get(self.selected_lot.get(), {})
        self.fill_entry_group(self.entries, draft)

    def load_both_lot_panels(self) -> None:
        for lot_key, entries in self.both_entries.items():
            self.fill_entry_group(entries, self.lot_drafts.get(lot_key, {}))

    def load_lot_drafts_from_main(self) -> None:
        """메인 일반 DNC 화면에 입력된 LOT 값을 신규 검증 팝업 초기값으로 가져옵니다.

        신규 검증에서는 매수/작업조건/지그를 다시 판단해야 하므로 비워둡니다.
        """
        for lot_key, entries in (("lot1", self.app.lot1_entries), ("lot2", self.app.lot2_entries)):
            lot = self.app.get_lot_data(entries)
            if not self.app.lot_has_any_value(lot):
                continue
            draft = dict(self.lot_drafts[lot_key])
            for key in ("step", "round", "manage_no", "lot_no", "process_code"):
                draft[key] = lot.get(key, "")
            draft["qty"] = ""
            draft["condition"] = ""
            draft["jig"] = ""
            self.lot_drafts[lot_key] = draft
        if not self.app.lot_has_any_value(self.app.get_lot_data(self.app.lot1_entries)):
            if self.app.lot_has_any_value(self.app.get_lot_data(self.app.lot2_entries)):
                self.selected_lot.set("lot2")

    def set_run_mode(self, mode: str) -> None:
        self.save_current_lot_draft()
        self.run_mode.set(mode)
        next_lot = "lot1" if mode == "both" else mode
        self.selected_lot.set(next_lot)
        self.load_selected_lot()
        self.load_both_lot_panels()
        self.refresh_input_mode()

    def refresh_input_mode(self) -> None:
        self.update_mode_buttons()
        if self.run_mode.get() == "both":
            self.single_panel.pack_forget()
            self.both_panel.pack(fill=tk.BOTH, expand=True)
        else:
            self.both_panel.pack_forget()
            self.single_panel.pack(fill=tk.BOTH, expand=True)
        self.update_checks()

    def get_data(self) -> tuple[dict, dict]:
        self.save_current_lot_draft()
        common = self.app.get_common_data()
        lot = dict(self.lot_drafts[self.selected_lot.get()])
        return common, lot

    def get_run_lots(self) -> tuple[dict, list[tuple[str, dict]]]:
        """신규 검증 실행 시 저장해야 할 LOT 목록을 반환합니다.

        둘 다 신규일 때만 작업자가 LOT 1만/LOT 2만/LOT 1+LOT 2 중 선택합니다.
        """
        self.save_current_lot_draft()
        common = self.app.get_common_data()
        mode = self.run_mode.get()
        run_keys = ["lot1", "lot2"] if mode == "both" else [mode]
        lots = [(key, dict(self.lot_drafts[key])) for key in run_keys]
        return common, lots

    def update_checks(self) -> None:
        if self.is_loading_fields:
            return
        _common, lot_items = self.get_run_lots()
        lots = [lot for _lot_key, lot in lot_items]

        if all(not lot.get("lot_no", "").strip() and not lot.get("process_code", "").strip() for lot in lots):
            self.mes_label.configure(text="MES Core 일치화: 대기중", fg=MUTED_TEXT)
        else:
            mes_results = [
                get_mes_core_message(lot.get("lot_no", ""), lot.get("process_code", ""))[0]
                for lot in lots
            ]
            if all(mes_results):
                self.mes_label.configure(text="MES Core 일치화: OK", fg=OK_COLOR)
            else:
                self.mes_label.configure(text="MES Core 일치화: NG", fg=NG_COLOR)

        if all(not lot.get("condition", "").strip() and not lot.get("jig", "").strip() for lot in lots):
            self.condition_label.configure(text="조건 적용 확인: 대기중", fg=MUTED_TEXT)
        else:
            condition_results = [get_single_condition_message(lot)[0] for lot in lots]
            self.condition_label.configure(
                text=f"조건 적용 확인: {'OK' if all(condition_results) else 'NG'}",
                fg=OK_COLOR if all(condition_results) else NG_COLOR,
            )

    def set_running(self, running: bool) -> None:
        self.is_running = running
        for button in self.buttons:
            button.configure(state="disabled" if running else "normal")

    def set_dnc_status(self, text: str) -> None:
        self.window.after(0, lambda: self.dnc_label.configure(text=f"DNC 진행 상태: {text}", fg=MUTED_TEXT))

    def run_new_model_dnc(self) -> None:
        run_new_model_dnc(self)

    def clear_inputs(self) -> None:
        run_keys = ["lot1", "lot2"] if self.run_mode.get() == "both" else [self.selected_lot.get()]
        for lot_key in run_keys:
            draft = dict(self.lot_drafts[lot_key])
            clear_keys = ["qty", "condition", "jig"]
            if not self.lock_imported_fields:
                clear_keys = ["step", "round", "manage_no", "lot_no", "qty", "process_code", "condition", "jig"]
            for key in clear_keys:
                draft[key] = ""
            self.lot_drafts[lot_key] = draft
        self.load_selected_lot()
        self.load_both_lot_panels()
        self.update_checks()
        self.dnc_label.configure(text="DNC 진행 상태: 대기중", fg=MUTED_TEXT)

    def clear_after_done(self) -> None:
        _common, lot_items = self.get_run_lots()
        for lot_key, draft in lot_items:
            for key in ["qty", "condition", "jig"]:
                draft[key] = ""
            self.lot_drafts[lot_key] = draft
        self.load_selected_lot()
        self.load_both_lot_panels()
        self.update_checks()


def open_new_model_popup(app: JiinDncManager) -> None:
    """KCC PKG 신규 모델 검증 DNC 팝업창을 엽니다."""
    NewModelPopup(app)


def run_new_model_dnc(popup: NewModelPopup) -> None:
    """신규 모델 DNC 실행 버튼 흐름입니다."""
    if popup.is_running or popup.app.is_running:
        show_operator_alert(popup.window, "진행 중", "DNC 실행중")
        return
    if not popup.app.save_settings_from_ui_silent():
        return
    if not popup.app.validate_kcc_pkg_process_paths():
        return
    common, lot_items = popup.get_run_lots()
    lots = [lot for _lot_key, lot in lot_items]
    all_errors = []
    for lot_key, lot in lot_items:
        lot_label = "LOT 1" if lot_key == "lot1" else "LOT 2"
        ok, errors = validate_new_model_dnc(common, lot)
        if not ok:
            all_errors.extend([f"{lot_label} - {error}" for error in errors])
    condition_names = {lot.get("condition", "").strip() for lot in lots if lot.get("condition", "").strip()}
    jig_names = {lot.get("jig", "").strip() for lot in lots if lot.get("jig", "").strip()}
    if len(lots) > 1 and len(condition_names) > 1:
        all_errors.append("LOT 1 / LOT 2 작업조건 다름")
    if len(lots) > 1 and len(jig_names) > 1:
        all_errors.append("LOT 1 / LOT 2 지그 다름")
    if all_errors:
        log_app("신규 모델 DNC 입력값 NG: " + " / ".join(all_errors))
        show_operator_alert(popup.window, "입력값 확인", format_operator_errors(all_errors))
        popup.dnc_label.configure(text="DNC 진행 상태: 입력값 NG", fg=NG_COLOR)
        return
    leader_name = ask_system_input(popup.window, "조장명 입력", "신규 모델 검증 조장명 입력")
    if not leader_name or not leader_name.strip():
        popup.dnc_label.configure(text="DNC 진행 상태: 취소", fg=MUTED_TEXT)
        return
    condition_file = popup.app.validate_condition_file(lots[0]["condition"])
    if not condition_file:
        popup.dnc_label.configure(text="DNC 진행 상태: 조건 파일 NG", fg=NG_COLOR)
        return
    log_app(f"신규 모델 DNC 시작: {len(lots)} LOT, 조건={lots[0]['condition']}")
    popup.set_running(True)
    popup.app.set_running(True)
    threading.Thread(target=new_model_worker, args=(popup, common, lots, leader_name.strip(), condition_file), daemon=True).start()


def run_normal_dnc(app: JiinDncManager) -> None:
    """요청 함수명 보존용 래퍼입니다. 실제 일반 DNC 실행은 JiinDncManager.run_normal_dnc에서 처리합니다."""
    app.run_normal_dnc()


def new_model_worker(popup: NewModelPopup, common: dict, lots: list[dict], leader_name: str, condition_file: Path) -> None:
    """신규 모델 DNC 백그라운드 작업입니다."""
    try:
        popup.set_dnc_status("DB 저장중")
        log_ids = [insert_new_model_db(common, lot, leader_name) for lot in lots]
        popup.set_dnc_status("DB 저장 완료")
        delete_existing_dnc_txt(Path(popup.app.config["transfer_dnc_folder"]))
        copied_file = copy_dnc_file(condition_file, Path(popup.app.config["transfer_dnc_folder"]))
        popup.set_dnc_status("DNC 파일 복사 완료")
        delete_after_delay(copied_file, int(popup.app.config["dnc_delete_seconds"]), popup.set_dnc_status)
        popup.window.after(0, lambda: finish_new_model_dnc(popup, log_ids, lots))
    except Exception as exc:
        popup.window.after(0, lambda error=exc: handle_popup_error(popup, error))


def finish_new_model_dnc(popup: NewModelPopup, log_ids: list[int], lots: list[dict]) -> None:
    """신규 모델 DNC 완료 후 초도품 확인 결과를 저장합니다."""
    try:
        first_article_ok = ask_system_yes_no(popup.window, "초도품 확인", "초도품 이상 없습니까?")
        for log_id, lot in zip(log_ids, lots):
            condition_name = lot["condition"]
            update_new_model_db(log_id, condition_name, first_article_ok)
            if first_article_ok:
                upsert_condition_master(lot, condition_name, lot["jig"], "신규 검증 완료")
        pending_count = get_unexported_kcc_pkg_count()
        popup.dnc_label.configure(text="DNC 진행 상태: DNC 완료", fg=OK_COLOR)
        popup.excel_label.configure(text=f"작업일보 반영: Excel 미반영 {pending_count}건", fg=OK_COLOR)
        log_app(f"신규 모델 DNC 완료: ids={log_ids}, Excel 미반영={pending_count}건")
        if popup.app.auto_export_kcc_pkg_to_excel(parent=popup.window):
            popup.excel_label.configure(text="작업일보 반영: 자동 반영 완료", fg=OK_COLOR)
        else:
            popup.excel_label.configure(text=f"작업일보 반영: Excel 미반영 {get_unexported_kcc_pkg_count()}건", fg=NG_COLOR)
        popup.clear_after_done()
        popup.set_running(False)
        popup.app.set_running(False)
        popup.window.after(200, popup.window.destroy)
    except Exception as exc:
        handle_popup_error(popup, exc)
    finally:
        if popup.window.winfo_exists():
            popup.set_running(False)
        popup.app.set_running(False)


def handle_popup_error(popup: NewModelPopup, exc: Exception) -> None:
    log_error("신규 모델 DNC 오류", exc)
    show_operator_alert(popup.window, "오류", str(exc), "error")
    popup.dnc_label.configure(text="DNC 진행 상태: 오류", fg=NG_COLOR)
    popup.set_running(False)
    popup.app.set_running(False)


def main() -> None:
    if not acquire_single_instance_lock():
        bring_existing_app_to_front()
        return
    config = load_config()
    if not is_license_allowed(config):
        show_license_block_message()
        return
    root = tk.Tk()
    app = JiinDncManager(root)
    root.mainloop()


if __name__ == "__main__":
    main()




































