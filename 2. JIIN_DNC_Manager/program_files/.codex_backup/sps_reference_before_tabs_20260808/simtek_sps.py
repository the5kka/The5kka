from __future__ import annotations

import json
import os
import shutil
import sqlite3
import threading
import time
from collections import Counter
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import messagebox, scrolledtext, simpledialog, ttk
import tkinter as tk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from simtek_sps_rules import (
    ParsedCondition,
    ProgramConditionCheck,
    Recommendation,
    SpsRuleBook,
    clean_text,
    extract_oms_program_shift,
    extract_program_shift,
    extract_program_size_key,
    find_exact_txt_files,
    format_shift,
    normalize_manage_no,
    parse_condition,
    parse_round,
    validate_program_condition,
    validate_steps,
)
from simtek_sps_program_catalog import (
    STATUS_AVAILABLE,
    STATUS_BLOCKED,
    STATUS_DUPLICATE_DISABLED,
    STATUS_MISSING_FILE,
    STATUS_REVIEW,
    SpsProgramCatalog,
    SpsProgramOption,
    condition_mismatch_reason,
)


PROCESS_NAME = "심텍 SPS"
WORKLOG_SHEET = "ST-PKG"
WORKLOG_MAPPING_READY = True
WORKLOG_ID_COLUMN = 33  # AG열: 중복 반영 방지용 DNC_LOG_ID
WORKLOG_HEADER_ROW = 8
WORKLOG_FIRST_DATA_ROW = 9


# V14-6월 ST-PKG 작업일보 M열의 정상 사용 이력입니다. 설정된 최신 작업일보와
# 추천 규칙 파일을 읽을 수 있으면 아래 목록에 자동으로 병합합니다.
SPS_PROGRAM_HISTORY = {
    "ST-BOC-500(411-512)2R": 1248,
    "ST-3Layer-500+4Hole(411-512)_2R_2_2024 T01 T02 test": 514,
    "ST-03L-500(415-515)2R": 36,
    "ST-03L2-500(415-515)_2R": 24,
    "ST-3Layer-500+4Hole(411-512)_2R_2_2024": 23,
    "ST-BOC-500(+10)(413-514)2R(New2)": 19,
    "06-OUT_(411-512)2R(New)": 14,
    "ST-04ETS-2-500+10(415-516)2R": 14,
    "ST-06-1-500+10(413-513)2R": 13,
    "ST-06-1-500(+10)(413-513)2R": 6,
    "ST-03L2-500(2ND)(415-515)_2R": 5,
    "ST-BOC-500(415-516)2R(New)": 4,
    "ST-03L2-500(415-515)2R": 4,
    "ST-03L-500(415-516-C)_2R": 1,
    "ST-BOC-500(+10)(415-516)2R(New)": 1,
    "ST-03R-500(414-515)_2R": 1,
    "ST-03L-500(415-516)_2R": 1,
}


@dataclass
class SpsLot:
    number: int
    manage_no: str
    lot_no: str
    start_step: int
    current_step: int
    round_no: int
    qty: int
    oms1: str
    oms2: str
    parsed: ParsedCondition
    program: str = ""
    master_id: int | None = None


def build_sps_program_catalog(
    program_counts: Counter[str] | dict[str, int],
    registered_programs: set[str] | tuple[str, ...] = (),
) -> list[SpsProgramOption]:
    counts = Counter({clean_text(name): int(count) for name, count in program_counts.items() if clean_text(name)})
    for program in registered_programs:
        name = clean_text(program)
        if name:
            counts.setdefault(name, 0)

    def sort_key(option: SpsProgramOption) -> tuple[tuple[int, int], int, int, str]:
        try:
            width, height = (int(value) for value in option.size_key.split("-", 1))
        except (TypeError, ValueError):
            width, height = 9999, 9999
        return (width, height), 0 if option.shift is None else 1, -option.history_count, option.program.casefold()

    options = [
        SpsProgramOption(
            program=name,
            size_key=extract_program_size_key(name),
            shift=extract_program_shift(name),
            history_count=count,
            status=STATUS_REVIEW,
            status_label="관리자 확인",
            file_verified=False,
            file_status="과거 이력",
        )
        for name, count in counts.items()
        if name not in {"작업 Program", "작업 P/G"}
    ]
    return sorted(options, key=sort_key)


class SpsRepository:
    def __init__(self, db_path: Path):
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.initialize()

    def connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path, timeout=10)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA busy_timeout=10000")
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        return conn

    def initialize(self) -> None:
        conn = self.connect()
        try:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS sps_condition_master (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    active INTEGER NOT NULL DEFAULT 1,
                    manage_no TEXT NOT NULL,
                    manage_key TEXT NOT NULL,
                    round_no INTEGER NOT NULL,
                    start_step INTEGER NOT NULL,
                    current_step INTEGER NOT NULL,
                    step_delta INTEGER NOT NULL,
                    oms1_raw TEXT NOT NULL,
                    oms1_norm TEXT NOT NULL,
                    oms2_raw TEXT NOT NULL,
                    oms2_norm TEXT NOT NULL,
                    size_key TEXT NOT NULL,
                    oms2_key TEXT NOT NULL,
                    features_json TEXT NOT NULL,
                    program_name TEXT NOT NULL,
                    leader_name TEXT,
                    source_text TEXT,
                    registered_at TEXT NOT NULL,
                    excluded_at TEXT,
                    excluded_reason TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_sps_master_lookup
                    ON sps_condition_master(manage_key, round_no, start_step, current_step, oms1_norm, oms2_norm, active);

                CREATE TABLE IF NOT EXISTS sps_dnc_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    dnc_type TEXT NOT NULL,
                    status TEXT NOT NULL,
                    machine TEXT,
                    work_date TEXT,
                    shift_group TEXT,
                    shift_name TEXT,
                    worker TEXT,
                    leader_name TEXT,
                    start_step INTEGER,
                    current_step INTEGER,
                    step_delta INTEGER,
                    round_no INTEGER,
                    manage_no TEXT,
                    lot_no TEXT,
                    qty INTEGER,
                    result_value REAL,
                    oms1_raw TEXT,
                    oms1_norm TEXT,
                    oms2_raw TEXT,
                    oms2_norm TEXT,
                    size_key TEXT,
                    oms2_key TEXT,
                    features_json TEXT,
                    program_name TEXT,
                    recommendation_basis TEXT,
                    first_article_result TEXT,
                    burr_result TEXT,
                    source_file TEXT,
                    copied_file TEXT,
                    record_time TEXT,
                    exported INTEGER NOT NULL DEFAULT 0,
                    exported_at TEXT,
                    error_text TEXT,
                    created_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_sps_logs_export ON sps_dnc_logs(exported, status, id);
                CREATE INDEX IF NOT EXISTS idx_sps_logs_condition
                    ON sps_dnc_logs(manage_no, round_no, start_step, current_step, oms1_norm, oms2_norm);
                """
            )
            conn.commit()
        finally:
            conn.close()

    def find_active_master(self, lot: SpsLot) -> list[sqlite3.Row]:
        conn = self.connect()
        try:
            return conn.execute(
                """
                SELECT * FROM sps_condition_master
                WHERE active=1 AND manage_key=? AND round_no=? AND start_step=? AND current_step=?
                  AND oms1_norm=? AND oms2_norm=?
                ORDER BY id DESC
                """,
                (
                    normalize_manage_no(lot.manage_no), lot.round_no, lot.start_step, lot.current_step,
                    lot.parsed.oms1_norm, lot.parsed.oms2_norm,
                ),
            ).fetchall()
        finally:
            conn.close()

    def history_candidates(self, lot: SpsLot) -> list[tuple[str, int]]:
        conn = self.connect()
        try:
            rows = conn.execute(
                """
                SELECT program_name, COUNT(*) AS used_count
                FROM sps_dnc_logs
                WHERE status='완료'
                  AND REPLACE(REPLACE(UPPER(manage_no), '-', ''), ' ', '')=?
                  AND round_no=? AND start_step=? AND current_step=?
                  AND oms1_norm=? AND oms2_norm=? AND first_article_result='이상 없음'
                  AND burr_result='이상 없음' AND program_name<>''
                GROUP BY program_name ORDER BY used_count DESC, program_name
                """,
                (
                    normalize_manage_no(lot.manage_no), lot.round_no, lot.start_step, lot.current_step,
                    lot.parsed.oms1_norm, lot.parsed.oms2_norm,
                ),
            ).fetchall()
            return [(str(row["program_name"]), int(row["used_count"])) for row in rows]
        finally:
            conn.close()

    def register_master(
        self,
        lot: SpsLot,
        program: str,
        leader_name: str,
        source_text: str,
        *,
        replace_existing: bool = False,
    ) -> int:
        existing = self.find_active_master(lot)
        if existing:
            programs = {clean_text(row["program_name"]) for row in existing}
            if programs == {clean_text(program)} and len(existing) == 1:
                return int(existing[0]["id"])
            if not replace_existing:
                raise ValueError("동일 조건의 활성 마스터가 이미 존재합니다. 조건 마스터 관리에서 확인하세요.")
        conn = self.connect()
        try:
            if existing:
                stamp = datetime.now().isoformat(timespec="seconds")
                conn.executemany(
                    "UPDATE sps_condition_master SET active=0, excluded_at=?, excluded_reason=? WHERE id=?",
                    [(stamp, "신규 재검증 정상 완료로 교체", int(row["id"])) for row in existing],
                )
            cursor = conn.execute(
                """
                INSERT INTO sps_condition_master (
                    active, manage_no, manage_key, round_no, start_step, current_step, step_delta,
                    oms1_raw, oms1_norm, oms2_raw, oms2_norm, size_key, oms2_key, features_json,
                    program_name, leader_name, source_text, registered_at
                ) VALUES (1,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    lot.manage_no, normalize_manage_no(lot.manage_no), lot.round_no,
                    lot.start_step, lot.current_step, lot.current_step - lot.start_step,
                    lot.parsed.oms1_raw, lot.parsed.oms1_norm, lot.parsed.oms2_raw, lot.parsed.oms2_norm,
                    lot.parsed.size_key, lot.parsed.oms2_key, lot.parsed.features.to_json(),
                    clean_text(program), clean_text(leader_name), clean_text(source_text), datetime.now().isoformat(timespec="seconds"),
                ),
            )
            conn.commit()
            return int(cursor.lastrowid)
        finally:
            conn.close()

    def list_masters(self) -> list[sqlite3.Row]:
        conn = self.connect()
        try:
            return conn.execute("SELECT * FROM sps_condition_master ORDER BY active DESC, id DESC").fetchall()
        finally:
            conn.close()

    def set_master_active(self, record_id: int, active: bool, reason: str = "") -> None:
        conn = self.connect()
        try:
            conn.execute(
                """UPDATE sps_condition_master
                   SET active=?, excluded_at=?, excluded_reason=? WHERE id=?""",
                (
                    1 if active else 0,
                    None if active else datetime.now().isoformat(timespec="seconds"),
                    "" if active else clean_text(reason),
                    record_id,
                ),
            )
            conn.commit()
        finally:
            conn.close()

    def insert_logs(self, common: dict[str, str], lots: list[SpsLot], dnc_type: str, leader_name: str, basis: str) -> list[int]:
        conn = self.connect()
        ids: list[int] = []
        try:
            now = datetime.now()
            for lot in lots:
                cursor = conn.execute(
                    """
                    INSERT INTO sps_dnc_logs (
                        dnc_type,status,machine,work_date,shift_group,shift_name,worker,leader_name,
                        start_step,current_step,step_delta,round_no,manage_no,lot_no,qty,result_value,
                        oms1_raw,oms1_norm,oms2_raw,oms2_norm,size_key,oms2_key,features_json,program_name,
                        recommendation_basis,record_time,created_at
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        dnc_type, "진행중", common["machine"], common["work_date"], common["shift_group"],
                        common["shift"], common["worker"], clean_text(leader_name), lot.start_step, lot.current_step,
                        lot.current_step - lot.start_step, lot.round_no, lot.manage_no, lot.lot_no, lot.qty,
                        float(lot.qty) * 0.2, lot.parsed.oms1_raw, lot.parsed.oms1_norm,
                        lot.parsed.oms2_raw, lot.parsed.oms2_norm, lot.parsed.size_key, lot.parsed.oms2_key,
                        lot.parsed.features.to_json(), lot.program, clean_text(basis), now.strftime("%H:%M:%S"),
                        now.isoformat(timespec="seconds"),
                    ),
                )
                ids.append(int(cursor.lastrowid))
            conn.commit()
            return ids
        finally:
            conn.close()

    def update_copy_paths(self, log_ids: list[int], source_file: Path, copied_file: Path) -> None:
        conn = self.connect()
        try:
            conn.executemany(
                "UPDATE sps_dnc_logs SET source_file=?, copied_file=? WHERE id=?",
                [(str(source_file), str(copied_file), log_id) for log_id in log_ids],
            )
            conn.commit()
        finally:
            conn.close()

    def finish_logs(self, log_ids: list[int], first_ok: bool, burr_ok: bool, error_text: str = "") -> None:
        status = "완료" if first_ok and burr_ok and not error_text else "NG" if not error_text else "오류"
        conn = self.connect()
        try:
            conn.executemany(
                """
                UPDATE sps_dnc_logs
                SET status=?, first_article_result=?, burr_result=?, error_text=? WHERE id=?
                """,
                [
                    (
                        status,
                        "이상 없음" if first_ok else "이상 있음",
                        "이상 없음" if burr_ok else "이상 있음",
                        clean_text(error_text),
                        log_id,
                    )
                    for log_id in log_ids
                ],
            )
            conn.commit()
        finally:
            conn.close()

    def fail_logs(self, log_ids: list[int], error_text: str) -> None:
        conn = self.connect()
        try:
            conn.executemany(
                "UPDATE sps_dnc_logs SET status='오류', error_text=? WHERE id=?",
                [(clean_text(error_text), log_id) for log_id in log_ids],
            )
            conn.commit()
        finally:
            conn.close()

    def unexported_logs(self) -> list[sqlite3.Row]:
        conn = self.connect()
        try:
            return conn.execute(
                "SELECT * FROM sps_dnc_logs WHERE exported=0 AND status IN ('완료','NG') ORDER BY id"
            ).fetchall()
        finally:
            conn.close()

    def mark_exported(self, log_ids: list[int]) -> None:
        if not log_ids:
            return
        conn = self.connect()
        try:
            stamp = datetime.now().isoformat(timespec="seconds")
            conn.executemany(
                "UPDATE sps_dnc_logs SET exported=1, exported_at=? WHERE id=?",
                [(stamp, log_id) for log_id in log_ids],
            )
            conn.commit()
        finally:
            conn.close()

    def recent_logs(self, limit: int = 300) -> list[sqlite3.Row]:
        conn = self.connect()
        try:
            return conn.execute("SELECT * FROM sps_dnc_logs ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
        finally:
            conn.close()


class SimtekSpsController:
    def __init__(
        self,
        app,
        page: tk.Frame,
        data_dir: Path,
        default_rule_file: Path | None = None,
        theme: dict | None = None,
        logo_path: Path | None = None,
        ui_components: dict | None = None,
        alert=None,
        log_callback=None,
        error_callback=None,
        ask_leader=None,
        save_workbook=None,
        acquire_excel_lock=None,
        release_excel_lock=None,
    ):
        self.app = app
        self.root = app.root
        self.page = page
        self.config = app.config
        self.data_dir = Path(data_dir) / "SIMTEK_SPS"
        self.repo = SpsRepository(self.data_dir / "sps.db")
        self.alert_callback = alert
        self.log_callback = log_callback or (lambda _text: None)
        self.error_callback = error_callback or (lambda _title, _exc: None)
        self.ask_leader_callback = ask_leader
        self.save_workbook_callback = save_workbook
        self.acquire_excel_lock_callback = acquire_excel_lock
        self.release_excel_lock_callback = release_excel_lock
        self.default_rule_file = Path(default_rule_file) if default_rule_file else None
        self.default_catalog_file = (
            self.default_rule_file.with_name("sps_program_catalog.json")
            if self.default_rule_file
            else None
        )
        self.theme = theme or {"bg": "#f3f6fb", "light": "#eaf2ff", "primary": "#0f5bff", "border": "#b8cdfa"}
        self.logo_path = Path(logo_path) if logo_path else None
        self.ui_components = ui_components or {}
        self.rule_book: SpsRuleBook | None = None
        self.program_catalog: SpsProgramCatalog | None = None
        self._program_catalog_cache: dict[str, SpsProgramOption] = {}
        self._excel_export_running = False
        self._worklog_cache_stamp: tuple[str, int, int] | None = None
        self._worklog_program_index: dict[tuple[str, int, int, str, str], Counter[str]] = {}
        self._worklog_program_counts: Counter[str] = Counter()
        self._worklog_program_recent: dict[str, str] = {}
        self._worklog_cache_lock = threading.Lock()
        self.master_by_lot: dict[int, sqlite3.Row] = {}
        self.status_cards: dict[int, tk.Label] = {}
        self.common_entries: dict[str, object] = {}
        self.lot_entries: dict[int, dict[str, object]] = {1: {}, 2: {}}
        self.common_vars = {
            "machine": tk.StringVar(value=self.config.get("machine", "트리밍 1호기")),
            "work_date": tk.StringVar(value=datetime.now().strftime("%Y-%m-%d")),
            "shift_group": tk.StringVar(value=""),
            "shift": tk.StringVar(value="주간"),
            "worker": tk.StringVar(value=""),
        }
        self.lot_vars: dict[int, dict[str, tk.StringVar]] = {1: self._new_lot_vars(), 2: self._new_lot_vars()}
        self.dnc_status_var = tk.StringVar(value="대기중")
        self.excel_status_var = tk.StringVar(value="대기중")
        self.match_status_var = tk.StringVar(value="LOT 2 미사용")
        self._load_program_catalog_policy()
        self._load_rule_book(silent=True)
        self.build_ui()

    @staticmethod
    def _new_lot_vars() -> dict[str, tk.StringVar]:
        return {
            "manage_no": tk.StringVar(), "lot_no": tk.StringVar(),
            "start_step": tk.StringVar(), "current_step": tk.StringVar(),
            "round": tk.StringVar(), "qty": tk.StringVar(),
            "oms1": tk.StringVar(), "oms2": tk.StringVar(),
            "program": tk.StringVar(), "detail": tk.StringVar(),
        }

    def _alert(self, title: str, message: str, kind: str = "warning", parent=None) -> None:
        parent = parent or self.root
        if self.alert_callback:
            self.alert_callback(parent, title, message, kind)
            return
        getattr(messagebox, "showerror" if kind == "error" else "showinfo")(title, message, parent=parent)

    def _append_log(self, text: str) -> None:
        stamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert(tk.END, f"[{stamp}] {text}\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state="disabled")
        self.log_callback(f"심텍 SPS: {text}")

    def _load_program_catalog_policy(self) -> None:
        path = self.default_catalog_file
        if path and path.is_file():
            try:
                self.program_catalog = SpsProgramCatalog.from_json(path)
                self.log_callback(f"심텍 SPS: P/G 기준 목록 로드 완료: {path}")
                return
            except Exception as exc:
                self.error_callback("심텍 SPS P/G 기준 목록 로드 실패", exc)
        else:
            self.log_callback(f"심텍 SPS: P/G 기준 목록 없음: {path or '미설정'}")
        self.program_catalog = None

    def _load_rule_book(self, silent: bool = False) -> None:
        configured = clean_text(self.config.get("simtek_sps_rule_file", ""))
        path = Path(configured) if configured else self.default_rule_file
        if path and path.exists():
            try:
                self.rule_book = SpsRuleBook(path)
                self.log_callback(f"심텍 SPS: 추천 규칙 파일 로드 완료: {path}")
                return
            except Exception as exc:
                self.error_callback("심텍 SPS 규칙 파일 로드 실패", exc)
        else:
            self.log_callback(f"심텍 SPS: 추천 규칙 파일 경로 없음: {path or '미설정'}")
        self.rule_book = None
        if not silent:
            self._alert("추천 규칙 확인", "SPS 추천 규칙 파일을 설정에서 선택하세요.", "warning")

    def build_ui(self) -> None:
        font_name = "맑은 고딕"
        theme = self.theme
        bg = theme["bg"]
        light = theme["light"]
        primary = theme["primary"]
        border = theme["border"]
        surface = "#ffffff"
        muted = "#687386"
        required_components = {"ComboField", "DateField", "SegmentedField", "LabeledEntry", "RoundField"}
        missing_components = required_components.difference(self.ui_components)
        if missing_components:
            raise RuntimeError(f"심텍 SPS 공용 UI 구성요소 누락: {', '.join(sorted(missing_components))}")

        ComboField = self.ui_components["ComboField"]
        DateField = self.ui_components["DateField"]
        SegmentedField = self.ui_components["SegmentedField"]
        LabeledEntry = self.ui_components["LabeledEntry"]

        self.page.configure(bg=theme["bg"])
        self.page.columnconfigure(0, weight=1)
        self.page.rowconfigure(2, weight=0)
        self.page.rowconfigure(3, weight=0)

        title = tk.Frame(self.page, bg=light)
        title.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 8))
        title.columnconfigure(0, minsize=260)
        title.columnconfigure(1, weight=1)
        title.columnconfigure(2, minsize=360)
        logo_slot = tk.Frame(title, bg=light, width=250, height=44)
        logo_slot.grid(row=0, column=0, sticky="w", padx=(14, 8))
        logo_slot.grid_propagate(False)
        if self.logo_path and self.logo_path.exists():
            try:
                logo = tk.PhotoImage(file=str(self.logo_path))
                scale = max(1, logo.height() // 32)
                self.logo_image = logo.subsample(scale, scale)
                tk.Label(logo_slot, image=self.logo_image, bg=light, bd=0).pack(side=tk.LEFT, anchor="w")
            except Exception:
                self.logo_image = None
                tk.Label(logo_slot, text="SIMMTECH", bg=light, fg=primary, font=(font_name, 18, "bold")).pack(side=tk.LEFT, anchor="w")
        else:
            self.logo_image = None
            tk.Label(logo_slot, text="SIMMTECH", bg=light, fg=primary, font=(font_name, 18, "bold")).pack(side=tk.LEFT, anchor="w")
        tk.Label(title, text="심텍 SPS DNC", bg=light, fg=primary, font=(font_name, 14, "bold"), height=2).grid(row=0, column=1, sticky="ew")
        title_buttons = tk.Frame(title, bg=light)
        title_buttons.grid(row=0, column=2, sticky="e", padx=(8, 10))
        self.run_button = self.app.create_tlb_gradient_run_button(title_buttons, command=self.run_normal_dnc, scheme="simtek_sps")
        self.run_button.grid(row=0, column=0, padx=4, pady=4)
        self.app.add_normal_button(title_buttons, "입력 초기화", self.clear_inputs).grid(row=0, column=1, padx=4, pady=4)

        common = self.app.create_panel(self.page, "공통 입력", theme)
        common.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 8))
        common_widgets = [
            ("machine", ComboField(common, "설비 호기", ["트리밍 1호기", "트리밍 2호기", "트리밍 3호기"], initial=self.config.get("machine", "트리밍 1호기"), width=12)),
            ("work_date", DateField(common, "작업일자", on_change=lambda: self.app.handle_common_change("work_date", "simtek_sps"))),
            ("shift_group", SegmentedField(common, "조", ["A", "B", "C"], allow_empty=True, on_change=lambda: self.app.handle_common_change("shift_group", "simtek_sps"))),
            ("shift", SegmentedField(common, "근무", ["주간", "야간"], on_change=lambda: self.app.handle_common_change("shift", "simtek_sps"))),
            ("worker", LabeledEntry(common, "작업자", width=12, on_change=lambda: self.app.handle_common_change("worker", "simtek_sps"), live_change=False)),
        ]
        for index, (key, entry) in enumerate(common_widgets):
            entry.grid(row=1, column=index, sticky="ew", padx=8, pady=8)
            self.common_entries[key] = entry
            self.common_vars[key] = entry.var
            if key == "machine":
                entry.combo.bind("<<ComboboxSelected>>", lambda _event: self.app.handle_common_change("machine", "simtek_sps"))
        common.columnconfigure(0, weight=0, minsize=320)
        common.columnconfigure(1, weight=1, minsize=430)
        common.columnconfigure(2, weight=1, minsize=330)
        common.columnconfigure(3, weight=1, minsize=360)
        common.columnconfigure(4, weight=0, minsize=320)

        lots = tk.Frame(self.page, bg=bg)
        lots.grid(row=2, column=0, sticky="nsew", padx=14)
        lots.columnconfigure(0, weight=1, uniform="sps_lots")
        lots.columnconfigure(1, weight=1, uniform="sps_lots")
        self._build_lot_panel(lots, 1, "LOT 1 입력", theme).grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        self._build_lot_panel(lots, 2, "LOT 2 입력 (선택)", theme).grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        bottom = tk.Frame(self.page, bg=bg)
        bottom.grid(row=3, column=0, sticky="ew", padx=14, pady=(8, 14))
        bottom.columnconfigure(0, weight=1)
        status = tk.Frame(bottom, bg=surface, highlightthickness=1, highlightbackground=border, bd=0)
        status.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        status.columnconfigure(1, weight=1)
        for row, (label, var) in enumerate((("2LOT 조건 일치 확인", self.match_status_var), ("DNC 진행 상태", self.dnc_status_var), ("작업일보 반영", self.excel_status_var))):
            tk.Label(status, text=label, bg=light, fg=primary, width=22, height=2, font=(font_name, 11, "bold")).grid(row=row, column=0, sticky="nsw")
            tk.Label(status, textvariable=var, bg=surface, fg=muted, anchor="w", font=(font_name, 12, "bold")).grid(row=row, column=1, sticky="ew", padx=14)

        log_wrap = tk.Frame(bottom, bg=surface, highlightthickness=1, highlightbackground=border, bd=0)
        log_wrap.grid(row=1, column=0, sticky="ew", padx=(0, 10), pady=(8, 0))
        log_wrap.columnconfigure(0, weight=1)
        tk.Label(log_wrap, text="심텍 SPS DNC 작업 로그", bg=light, fg=primary, font=(font_name, 10, "bold"), height=1).grid(row=0, column=0, sticky="ew")
        self.log_text = scrolledtext.ScrolledText(log_wrap, height=5, state="disabled", wrap=tk.WORD, bg=surface, fg="#172033", relief=tk.FLAT, font=(font_name, 10), padx=10, pady=8)
        self.log_text.grid(row=1, column=0, sticky="ew")

        actions = tk.Frame(bottom, bg=bg)
        actions.grid(row=0, column=1, rowspan=2, sticky="ne")
        for col in range(3):
            actions.columnconfigure(col, weight=1, uniform="sps_actions")
        self.app.add_side_button(actions, "추천 규칙 상세", self.open_rule_guide, "SidePrimary.TButton").grid(row=0, column=0, sticky="nsew", padx=4, pady=4)
        self.app.add_side_button(actions, "신규 모델 검증 DNC", self.open_new_model_popup, "SidePrimary.TButton").grid(row=0, column=1, sticky="nsew", padx=4, pady=4)
        self.app.add_side_button(actions, "조건 마스터 관리", self.open_master_popup, "SideDanger.TButton").grid(row=0, column=2, sticky="nsew", padx=4, pady=4)
        self.app.add_side_button(actions, "작업일보 반영", self.start_excel_export).grid(row=1, column=0, sticky="nsew", padx=4, pady=4)
        self.app.add_side_button(actions, "작업일보 열기", self.open_worklog).grid(row=1, column=1, sticky="nsew", padx=4, pady=4)

    def _build_lot_panel(self, parent, lot_number: int, title: str, theme: dict) -> tk.Frame:
        panel = self.app.create_panel(parent, title, theme)
        panel.columnconfigure(0, weight=1, uniform=f"sps{lot_number}")
        panel.columnconfigure(1, weight=1, uniform=f"sps{lot_number}")
        variables = self.lot_vars[lot_number]

        LabeledEntry = self.ui_components["LabeledEntry"]
        RoundField = self.ui_components["RoundField"]

        def make_lot_entry(key: str, label: str):
            if key == "round":
                entry = RoundField(panel, label)
            elif key == "program":
                entry = LabeledEntry(panel, label, width=24, style="Lookup.TEntry", readonly=True)
            elif key in {"start_step", "current_step", "qty"}:
                entry = LabeledEntry(panel, label, width=24, numeric_only=True)
            elif key in {"manage_no", "lot_no"}:
                entry = LabeledEntry(panel, label, width=24, uppercase=True)
            else:
                entry = LabeledEntry(panel, label, width=24)
            self.lot_entries[lot_number][key] = entry
            variables[key] = entry.var
            return entry

        fields = (
            ("start_step", "시작 STEP", 1, 0), ("current_step", "현 STEP", 1, 1),
            ("manage_no", "관리번호", 2, 0), ("lot_no", "LOT No", 2, 1),
            ("qty", "매수", 3, 0), ("round", "차수", 3, 1),
            ("oms1", "조건 1", 4, 0), ("oms2", "조건 2", 4, 1),
            ("program", "조건(조회)", 5, 0),
        )
        for key, label, row, col in fields:
            make_lot_entry(key, label).grid(row=row, column=col, sticky="ew", padx=10, pady=8)
        tk.Frame(panel, bg="#ffffff").grid(row=5, column=1, sticky="ew", padx=10, pady=8)

        for key in ("start_step", "current_step", "manage_no", "lot_no", "qty", "round", "oms1", "oms2"):
            variables[key].trace_add("write", lambda *_args, n=lot_number: self._invalidate_lot(n))

        self.app.create_mes_lookup_button(
            panel,
            command=lambda: self.lookup_lot(lot_number),
            scheme="simtek_sps",
            ready_check=lambda: self.is_lookup_ready(lot_number),
        ).grid(row=6, column=0, columnspan=2, sticky="ew", padx=10, pady=(8, 6))

        status = tk.Frame(panel, bg="#ffffff")
        status.grid(row=7, column=0, columnspan=2, sticky="ew", padx=10, pady=(0, 12))
        status.columnconfigure(0, weight=1, uniform=f"sps{lot_number}_status")
        status.columnconfigure(1, weight=1, uniform=f"sps{lot_number}_status")
        card = self.app.create_judgement_card(status, "조건 조회")
        card.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        self.app.hide_judgement_card(card)
        self.status_cards[lot_number] = card
        return panel

    def is_lookup_ready(self, lot_number: int) -> bool:
        variables = self.lot_vars[lot_number]
        required = ("start_step", "current_step", "manage_no", "lot_no", "qty", "round", "oms1", "oms2")
        return all(clean_text(variables[key].get()) for key in required)

    def _invalidate_lot(self, lot_number: int) -> None:
        self.master_by_lot.pop(lot_number, None)
        self.lot_vars[lot_number]["program"].set("")
        self.lot_vars[lot_number]["detail"].set("")
        card = self.status_cards.get(lot_number)
        if card:
            self.app.hide_judgement_card(card)
        self._update_match_status()

    def _lot_is_used(self, lot_number: int) -> bool:
        values = self.lot_vars[lot_number]
        return any(clean_text(values[key].get()) for key in ("manage_no", "lot_no", "start_step", "current_step", "oms1", "oms2"))

    def _collect_lot(self, lot_number: int) -> SpsLot:
        values = self.lot_vars[lot_number]
        required = {
            "관리번호": values["manage_no"].get(), "LOT No": values["lot_no"].get(),
            "시작 STEP": values["start_step"].get(), "현 STEP": values["current_step"].get(),
            "차수": values["round"].get(), "매수": values["qty"].get(),
            "조건 1": values["oms1"].get(), "조건 2": values["oms2"].get(),
        }
        missing = [label for label, value in required.items() if not clean_text(value)]
        if missing:
            raise ValueError(f"LOT {lot_number} 필수 입력 누락: {', '.join(missing)}")
        step = validate_steps(values["start_step"].get(), values["current_step"].get())
        if not step.ok:
            raise ValueError(f"LOT {lot_number}: {step.message}")
        round_no = parse_round(values["round"].get())
        if round_no is None:
            raise ValueError(f"LOT {lot_number}: 차수를 선택하세요.")
        try:
            qty = int(clean_text(values["qty"].get()))
        except ValueError as exc:
            raise ValueError(f"LOT {lot_number}: 매수는 숫자로 입력하세요.") from exc
        if qty < 0:
            raise ValueError(f"LOT {lot_number}: 매수는 0 이상이어야 합니다.")
        parsed = parse_condition(values["oms1"].get(), values["oms2"].get())
        if not parsed.size_key:
            raise ValueError(f"LOT {lot_number}: 조건 1에서 SIZE를 찾지 못했습니다.")
        return SpsLot(
            lot_number, clean_text(values["manage_no"].get()).upper(), clean_text(values["lot_no"].get()).upper(),
            int(step.start_step), int(step.current_step), int(round_no), qty,
            parsed.oms1_raw, parsed.oms2_raw, parsed,
            clean_text(values["program"].get()),
            int(self.master_by_lot[lot_number]["id"]) if lot_number in self.master_by_lot else None,
        )

    def _collect_lots(self) -> list[SpsLot]:
        if not self._lot_is_used(1):
            raise ValueError("LOT 1을 입력하세요.")
        lots = [self._collect_lot(1)]
        if self._lot_is_used(2):
            lots.append(self._collect_lot(2))
        if len(lots) == 2:
            if lots[0].lot_no == lots[1].lot_no:
                raise ValueError("LOT 1과 LOT 2의 LOT No가 같습니다. 동일 LOT는 함께 진행할 수 없습니다.")
            if (
                lots[0].start_step, lots[0].current_step, lots[0].round_no,
                lots[0].parsed.oms1_norm, lots[0].parsed.oms2_norm,
            ) != (
                lots[1].start_step, lots[1].current_step, lots[1].round_no,
                lots[1].parsed.oms1_norm, lots[1].parsed.oms2_norm,
            ):
                raise ValueError("LOT 1과 LOT 2의 STEP·차수·OMS 조건이 다릅니다. 함께 DNC할 수 없습니다.")
        return lots

    def _common_data(self) -> dict[str, str]:
        data = {key: clean_text(var.get()) for key, var in self.common_vars.items()}
        missing = [label for key, label in (("machine", "설비 호기"), ("work_date", "작업일자"), ("shift_group", "조"), ("shift", "근무"), ("worker", "작업자")) if not data[key]]
        if missing:
            raise ValueError("공통 입력 누락: " + ", ".join(missing))
        return data

    def _recommend(self, lot: SpsLot) -> Recommendation:
        if self.rule_book is None:
            self._load_rule_book(silent=True)
        if self.rule_book is None:
            return Recommendation("관리자 확인 필요", (), "추천 규칙 파일 없음", ("설정에서 SPS 추천 규칙 파일을 선택하세요.",))
        return self.rule_book.recommend(lot.manage_no, lot.round_no, lot.parsed)

    @staticmethod
    def _find_worklog_header_row(ws) -> int:
        for row_number in range(1, min(ws.max_row, 30) + 1):
            headers = {
                clean_text(ws.cell(row_number, column).value).replace(" ", "").upper()
                for column in range(1, min(ws.max_column, WORKLOG_ID_COLUMN) + 1)
            }
            if "STEP" in headers and "관리번호" in headers and any(value in headers for value in {"작업P/G", "작업PROGRAM"}):
                return row_number
        raise ValueError(f"{WORKLOG_SHEET} 시트에서 작업일보 제목 행을 찾지 못했습니다.")

    @staticmethod
    def _is_normal_worklog_result(value: object, *, trimming: bool = False) -> bool:
        normalized = clean_text(value).replace(" ", "").casefold()
        if trimming:
            return normalized in {"ok", "이상없음"}
        return normalized in {"이상없음", "ok"}

    @staticmethod
    def _worklog_step(value: object) -> int | None:
        try:
            return int(float(clean_text(value)))
        except (TypeError, ValueError):
            return None

    def _load_worklog_program_index(self, path: Path) -> dict[tuple[str, int, int, str, str], Counter[str]]:
        stat = path.stat()
        stamp = (str(path), int(stat.st_size), int(stat.st_mtime_ns))
        lock = getattr(self, "_worklog_cache_lock", None)
        if lock is None:
            lock = threading.Lock()
            self._worklog_cache_lock = lock
            self._worklog_cache_stamp = None
            self._worklog_program_index = {}
            self._worklog_program_counts = Counter()
            self._worklog_program_recent = {}
        with lock:
            if getattr(self, "_worklog_cache_stamp", None) == stamp:
                return self._worklog_program_index
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=True,
                keep_vba=path.suffix.lower() == ".xlsm",
                keep_links=False,
            )
            try:
                if WORKLOG_SHEET not in workbook.sheetnames:
                    raise KeyError(f"작업일보 시트 없음: {WORKLOG_SHEET}")
                ws = workbook[WORKLOG_SHEET]
                first_data_row = self._find_worklog_header_row(ws) + 1
                index: dict[tuple[str, int, int, str, str], Counter[str]] = {}
                program_counts: Counter[str] = Counter()
                program_recent: dict[str, str] = {}
                current_work_date = ""
                for values in ws.iter_rows(min_row=first_data_row, max_col=15, values_only=True):
                    if isinstance(values[0], datetime):
                        current_work_date = values[0].strftime("%Y-%m-%d")
                    elif clean_text(values[0]):
                        current_work_date = clean_text(values[0])[:10]
                    manage_no = normalize_manage_no(values[5])
                    current_step = self._worklog_step(values[4])
                    round_no = parse_round(values[9])
                    program = clean_text(values[12])
                    if not manage_no or current_step is None or round_no is None or not program:
                        continue
                    if not self._is_normal_worklog_result(values[13]):
                        continue
                    if not self._is_normal_worklog_result(values[14], trimming=True):
                        continue
                    parsed = parse_condition(values[10], values[11])
                    key = (manage_no, round_no, current_step, parsed.oms1_norm, parsed.oms2_norm)
                    index.setdefault(key, Counter())[program] += 1
                    program_counts[program] += 1
                    if current_work_date:
                        program_recent[program] = max(program_recent.get(program, ""), current_work_date)
            finally:
                workbook.close()
            self._worklog_cache_stamp = stamp
            self._worklog_program_index = index
            self._worklog_program_counts = program_counts
            self._worklog_program_recent = program_recent
            return index

    def load_program_catalog(self) -> list[SpsProgramOption]:
        if self.program_catalog is None:
            self._append_log("조건 목록: P/G 기준 목록이 없어 선택을 차단합니다.")
            self._program_catalog_cache = {}
            return []

        counts: Counter[str] = Counter()
        recent_dates: dict[str, str] = {}
        path_text = clean_text(self.config.get("simtek_sps_excel_file", ""))
        if path_text:
            path = Path(path_text)
            if path.is_file():
                try:
                    self._load_worklog_program_index(path)
                    counts.update(self._worklog_program_counts)
                    recent_dates.update(self._worklog_program_recent)
                except Exception as exc:
                    self.error_callback("심텍 SPS 조건 목록 작업일보 읽기 실패", exc)
                    self._append_log("조건 목록: 작업일보 읽기 실패 / 내장 이력 사용")
        registered = self.rule_book.registered_programs if self.rule_book is not None else set()
        source_text = clean_text(self.config.get("source_dnc_folders", {}).get(PROCESS_NAME, ""))
        source_folder = Path(source_text) if source_text else None
        options = self.program_catalog.build_options(
            counts,
            recent_dates,
            registered,
            source_folder,
        )
        self._program_catalog_cache = {option.program.casefold(): option for option in options}
        return options

    def resolve_program_name(self, value: object) -> str:
        if self.program_catalog is None:
            return clean_text(value)
        return self.program_catalog.resolve_name(value)

    def get_program_option(self, value: object) -> SpsProgramOption | None:
        resolved = self.resolve_program_name(value)
        option = self._program_catalog_cache.get(resolved.casefold())
        if option is not None or self._program_catalog_cache:
            return option
        self.load_program_catalog()
        return self._program_catalog_cache.get(resolved.casefold())

    def _worklog_candidates(self, lot: SpsLot) -> list[tuple[str, int]]:
        path_text = clean_text(self.config.get("simtek_sps_excel_file", ""))
        if not path_text:
            raise FileNotFoundError("심텍 SPS 작업일보 파일을 설정에서 선택하세요.")
        path = Path(path_text)
        if not path.is_file():
            raise FileNotFoundError(f"심텍 SPS 작업일보 파일을 확인하세요: {path}")
        key = (
            normalize_manage_no(lot.manage_no),
            lot.round_no,
            lot.current_step,
            lot.parsed.oms1_norm,
            lot.parsed.oms2_norm,
        )
        counter = self._load_worklog_program_index(path).get(key, Counter())
        return sorted(counter.items(), key=lambda item: (-item[1], item[0]))

    def _verified_existing_condition(self, lot: SpsLot) -> tuple[sqlite3.Row | None, str, str]:
        worklog_candidates = self._worklog_candidates(lot)
        if not worklog_candidates:
            return None, "", "작업일보 정상 이력 없음"
        # 과거 작업일보 1건 예외 PASS는 2026-08-08 최초 Rule 정리 때만 사용한다.
        # 운영 시작 후에는 새 P/G가 1건뿐이어도 조건 변경 신호이므로 즉시 차단한다.
        if len(worklog_candidates) > 1:
            programs = ", ".join(program for program, _count in worklog_candidates)
            return None, "", f"작업일보 P/G 복수: {programs}"
        worklog_program = clean_text(worklog_candidates[0][0])

        masters = self.repo.find_active_master(lot)
        if not masters:
            return None, "", "신규 검증 완료 DB 없음"
        if len(masters) > 1:
            return None, "", f"활성 DB 마스터 중복: {len(masters)}건"
        master = masters[0]
        db_program = clean_text(master["program_name"])
        if worklog_program != db_program:
            return None, "", f"작업일보 P/G({worklog_program})와 DB P/G({db_program}) 불일치"
        return master, worklog_program, "작업일보와 DB P/G 일치"

    def lookup_lot(self, lot_number: int) -> bool:
        try:
            lot = self._collect_lot(lot_number)
            master, program, reason = self._verified_existing_condition(lot)
            if master is not None:
                self.master_by_lot[lot_number] = master
                self.lot_vars[lot_number]["program"].set(program)
                self.lot_vars[lot_number]["detail"].set("")
                self._show_lot_card(lot_number, "조건 조회 OK", True)
                self._append_log(f"LOT {lot_number} 조건 조회 OK: {program} / {reason}")
                self._update_match_status()
                return True
            self.master_by_lot.pop(lot_number, None)
            self.lot_vars[lot_number]["program"].set("")
            self.lot_vars[lot_number]["detail"].set("")
            self._show_lot_card(lot_number, "조건 조회 NG", False)
            self._append_log(f"LOT {lot_number} 조건 조회 NG / 신규 검증 필요: {reason}")
            self._update_match_status()
            return False
        except Exception as exc:
            self.master_by_lot.pop(lot_number, None)
            self.lot_vars[lot_number]["program"].set("")
            self.lot_vars[lot_number]["detail"].set("")
            self._show_lot_card(lot_number, "조건 조회 NG", False)
            self._append_log(f"LOT {lot_number} 조건 조회 NG / 신규 검증 필요: {exc}")
            return False

    def _show_lot_card(self, lot_number: int, text: str, ok: bool) -> None:
        card = self.status_cards[lot_number]
        display_text = text.replace("조건 조회", "", 1).strip() or ("OK" if ok else "NG")
        self.app.show_judgement_card(card, "조건 조회", display_text, ok)

    def _update_match_status(self) -> None:
        if not self._lot_is_used(2):
            self.match_status_var.set("LOT 2 미사용")
            return
        try:
            lots = self._collect_lots()
        except Exception as exc:
            self.match_status_var.set(f"NG - {exc}")
            return
        if len(lots) == 2 and all(lot.master_id for lot in lots) and lots[0].program == lots[1].program:
            self.match_status_var.set("OK - 2LOT 동일 조건 / 동일 Program")
        else:
            self.match_status_var.set("조건 조회 필요")

    def _validate_paths_and_program(self, program: str) -> tuple[Path, Path]:
        source = Path(clean_text(self.config.get("source_dnc_folders", {}).get(PROCESS_NAME, "")))
        transfer = Path(clean_text(self.config.get("transfer_dnc_folder", "")))
        if not source.is_dir():
            raise FileNotFoundError(f"심텍 SPS DNC 조건 폴더를 확인하세요.\n{source}")
        if not transfer.is_dir():
            raise FileNotFoundError(f"DNC 전송 폴더를 확인하세요.\n{transfer}")
        matches = find_exact_txt_files(source, program)
        if not matches:
            raise FileNotFoundError(f"정확히 일치하는 DNC 파일이 없습니다.\n{program}.txt")
        if len(matches) > 1:
            details = "\n".join(str(path) for path in matches[:10])
            raise RuntimeError(f"동일 Program 파일이 {len(matches)}개입니다. DNC를 차단합니다.\n{details}")
        return matches[0], transfer

    def run_normal_dnc(self) -> None:
        if getattr(self.app, "is_running", False):
            self._alert("DNC 진행중", "다른 DNC 작업이 진행 중입니다.", "warning")
            return
        try:
            common = self._common_data()
            lots = self._collect_lots()
            for lot in lots:
                if lot.master_id is None:
                    raise ValueError(f"LOT {lot.number} 조건 조회가 완료되지 않았습니다.")
            programs = {lot.program for lot in lots}
            if len(programs) != 1:
                raise ValueError("2LOT Program이 서로 다릅니다. 함께 DNC할 수 없습니다.")
            source_file, transfer = self._validate_paths_and_program(lots[0].program)
            self._start_dnc(common, lots, source_file, transfer, "일반", "", "확정 조건 마스터", None)
        except Exception as exc:
            self.dnc_status_var.set("DNC 차단")
            self._alert("심텍 SPS DNC 실행 불가", str(exc), "error")

    def _start_dnc(self, common, lots, source_file, transfer, dnc_type, leader_name, basis, popup) -> None:
        self.app.is_running = True
        self.dnc_status_var.set("DNC 파일 복사중")
        self._append_log(f"{dnc_type} DNC 시작: {source_file.name}")
        thread = threading.Thread(
            target=self._dnc_worker,
            args=(common, lots, source_file, transfer, dnc_type, leader_name, basis, popup),
            daemon=True,
        )
        thread.start()

    def _dnc_worker(self, common, lots, source_file, transfer, dnc_type, leader_name, basis, popup) -> None:
        log_ids: list[int] = []
        try:
            log_ids = self.repo.insert_logs(common, lots, dnc_type, leader_name, basis)
            copied = transfer / source_file.name
            shutil.copy2(source_file, copied)
            self.repo.update_copy_paths(log_ids, source_file, copied)
            delete_seconds = max(1, int(self.config.get("dnc_delete_seconds", 60)))
            threading.Thread(target=self._delete_copied_file, args=(copied, delete_seconds), daemon=True).start()
            wait_seconds = max(0, int(self.config.get("first_article_wait_seconds", 5)))
            self.root.after(wait_seconds * 1000, lambda: self._finish_checks(log_ids, lots, dnc_type, leader_name, basis, popup))
        except Exception as exc:
            if log_ids:
                self.repo.fail_logs(log_ids, str(exc))
            self.root.after(0, lambda: self._handle_run_error(exc, popup))

    def _delete_copied_file(self, copied: Path, seconds: int) -> None:
        time.sleep(seconds)
        try:
            if copied.exists() and copied.is_file():
                copied.unlink()
                self.root.after(0, lambda: self._append_log(f"DNC 파일 삭제 완료: {copied.name}"))
        except Exception as exc:
            self.error_callback("심텍 SPS DNC 파일 삭제 실패", exc)
            self.root.after(0, lambda: self._append_log(f"DNC 파일 삭제 실패: {copied}"))

    def _finish_checks(self, log_ids, lots, dnc_type, leader_name, basis, popup) -> None:
        try:
            self.dnc_status_var.set("초품 확인 대기중")
            first_ok = messagebox.askyesno(
                "초품 4Point 확인", "초품 4Point 확인 결과 이상이 없습니까?\n\n이상이 없으면 [예], 이상이 있으면 [아니요]를 누르세요.",
                parent=popup.window if popup else self.root,
            )
            burr_ok = False
            if first_ok:
                burr_ok = messagebox.askyesno(
                    "4면 Burr 확인", "제품 4면 Burr 확인 결과 이상이 없습니까?\n\n이상이 없으면 [예], 이상이 있으면 [아니요]를 누르세요.",
                    parent=popup.window if popup else self.root,
                )
            self.repo.finish_logs(log_ids, first_ok, burr_ok)
            if dnc_type == "신규 검증" and first_ok and burr_ok:
                for lot in lots:
                    self.repo.register_master(lot, lot.program, leader_name, basis, replace_existing=True)
                self._append_log("신규 검증 OK / 조건 마스터 등록 완료")
            elif dnc_type == "신규 검증":
                self._append_log("신규 검증 NG / 조건 마스터 미등록")
            self.app.is_running = False
            self.dnc_status_var.set("DNC 완료" if first_ok and burr_ok else "검증 NG")
            worklog_path = Path(clean_text(self.config.get("simtek_sps_excel_file", "")))
            if WORKLOG_MAPPING_READY and worklog_path.is_file():
                self.start_excel_export(silent=True)
            else:
                self.excel_status_var.set("작업일보 새 시트 연결 전 / DB 이력 보관")
                self._append_log("작업일보 새 시트 연결 전: SPS DB에 미반영 이력 보관")
            if popup:
                popup.finish(first_ok and burr_ok)
            self._alert(
                "심텍 SPS DNC 완료" if first_ok and burr_ok else "심텍 SPS 검증 NG",
                "DNC와 확인이 완료되었습니다." if first_ok and burr_ok else "이상이 기록되었습니다. 신규 조건은 마스터에 등록하지 않았습니다.",
                "info" if first_ok and burr_ok else "warning",
                parent=popup.window if popup else self.root,
            )
        except Exception as exc:
            self._handle_run_error(exc, popup)

    def _handle_run_error(self, exc: Exception, popup=None) -> None:
        self.app.is_running = False
        self.dnc_status_var.set("오류")
        self.error_callback("심텍 SPS DNC 오류", exc)
        if popup:
            popup.set_running(False)
        self._alert("심텍 SPS DNC 오류", str(exc), "error", parent=popup.window if popup else self.root)

    def open_new_model_popup(self) -> None:
        if getattr(self.app, "is_running", False):
            self._alert("DNC 진행중", "다른 DNC 작업이 진행 중입니다.", "warning")
            return
        try:
            common = self._common_data()
            lots = self._collect_lots()
            if len(lots) == 2 and (
                lots[0].parsed.exact_key() != lots[1].parsed.exact_key()
                or lots[0].start_step != lots[1].start_step
                or lots[0].current_step != lots[1].current_step
                or lots[0].round_no != lots[1].round_no
            ):
                raise ValueError("LOT 1과 LOT 2 조건이 다릅니다. 신규 검증을 함께 할 수 없습니다.")
            for lot in lots:
                try:
                    master, _program, _reason = self._verified_existing_condition(lot)
                except Exception:
                    master = None
                if master is not None:
                    raise ValueError(f"LOT {lot.number}은 작업일보와 DB가 일치하는 기존 모델입니다. 일반 DNC 조건 조회를 사용하세요.")
            SpsNewModelPopup(self, common, lots)
        except Exception as exc:
            self._alert("신규 조건 검증 입력 확인", str(exc), "error")

    def run_new_model_from_popup(self, popup, program: str) -> None:
        try:
            program = popup.require_confirmed_program(program)
            checks = [validate_program_condition(lot.parsed, program) for lot in popup.lots]
            failed = [f"LOT {lot.number}: {check.summary}" for lot, check in zip(popup.lots, checks) if not check.ok]
            if failed:
                raise ValueError("사이즈·시프트 조건이 일치하지 않습니다.\n" + "\n".join(failed))
            source_file, transfer = self._validate_paths_and_program(program)
            leader = self.ask_leader_callback(popup.window) if self.ask_leader_callback else simpledialog.askstring("조장명 입력", "신규 조건 검증 조장명 입력", parent=popup.window)
            if not clean_text(leader):
                raise ValueError("조장명을 입력해야 신규 검증을 진행할 수 있습니다.")
            for lot in popup.lots:
                lot.program = program
            popup.set_running(True)
            basis = popup.registration_basis()
            self._start_dnc(popup.common, popup.lots, source_file, transfer, "신규 검증", clean_text(leader), basis, popup)
        except Exception as exc:
            popup.set_running(False)
            self._alert("신규 조건 검증 실행 불가", str(exc), "error", parent=popup.window)

    def start_excel_export(self, silent: bool = False) -> None:
        if not WORKLOG_MAPPING_READY:
            self.excel_status_var.set("작업일보 새 시트 연결 전 / DB 이력 보관")
            if not silent:
                self._alert(
                    "작업일보 연결 전",
                    "심텍 SPS 새 작업일보 시트의 열 위치가 확정되지 않았습니다.\n"
                    "현재 이력은 SPS DB에 안전하게 보관됩니다.",
                    "info",
                )
            return
        if self._excel_export_running:
            if not silent:
                self._alert("작업일보 반영중", "심텍 SPS 작업일보를 반영하고 있습니다.", "info")
            return
        self._excel_export_running = True
        self.excel_status_var.set("작업일보 반영중 / 화면 사용 가능")
        threading.Thread(target=self._excel_export_worker, args=(silent,), daemon=True).start()

    def _excel_export_worker(self, silent: bool) -> None:
        try:
            count = self.export_logs_to_excel()
            self.root.after(0, lambda: self._finish_excel_export(count, silent))
        except Exception as exc:
            self.root.after(0, lambda exc=exc: self._fail_excel_export(exc, silent))

    def _finish_excel_export(self, count: int, silent: bool) -> None:
        self._excel_export_running = False
        if count:
            self.excel_status_var.set(f"작업일보 반영 완료 / {count}건")
            self._append_log(f"작업일보 반영 완료: {count}건")
            if not silent:
                self._alert("작업일보 반영 완료", f"심텍 SPS {count}건을 반영했습니다.", "info")
        else:
            self.excel_status_var.set("작업일보 미반영 0건")
            if not silent:
                self._alert("작업일보 반영", "반영할 심텍 SPS 이력이 없습니다.", "info")

    def _fail_excel_export(self, exc: Exception, silent: bool) -> None:
        self._excel_export_running = False
        self.excel_status_var.set("작업일보 반영 실패 / DB 미반영 유지")
        self.error_callback("심텍 SPS 작업일보 자동 반영 실패", exc)
        if not silent:
            self._alert("작업일보 반영 실패", str(exc), "error")

    def export_logs_to_excel(self) -> int:
        if not WORKLOG_MAPPING_READY:
            raise RuntimeError(
                "심텍 SPS 새 작업일보 시트의 열 위치가 아직 확정되지 않았습니다. "
                "시트 확인 전에는 DB 이력만 보관합니다."
            )
        logs = self.repo.unexported_logs()
        if not logs:
            return 0
        path = Path(clean_text(self.config.get("simtek_sps_excel_file", "")))
        if not path.is_file():
            raise FileNotFoundError(f"심텍 SPS 작업일보 파일을 확인하세요.\n{path}")
        lock_path = None
        workbook = None
        exported_ids: list[int] = []
        try:
            if self.acquire_excel_lock_callback:
                lock_path = self.acquire_excel_lock_callback(path)
            workbook = load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm", keep_links=False)
            if WORKLOG_SHEET not in workbook.sheetnames:
                raise KeyError(f"작업일보 시트 없음: {WORKLOG_SHEET}")
            ws = workbook[WORKLOG_SHEET]
            existing_ids = {
                clean_text(ws.cell(row=row, column=WORKLOG_ID_COLUMN).value)
                for row in range(WORKLOG_FIRST_DATA_ROW, ws.max_row + 1)
                if clean_text(ws.cell(row=row, column=WORKLOG_ID_COLUMN).value)
            }
            if not clean_text(ws.cell(WORKLOG_HEADER_ROW, WORKLOG_ID_COLUMN).value):
                ws.cell(WORKLOG_HEADER_ROW, WORKLOG_ID_COLUMN).value = "DNC_LOG_ID"
                ws.column_dimensions["AG"].hidden = True
            row = self._next_worklog_row(ws)
            template_row = max(WORKLOG_FIRST_DATA_ROW, row - 1)
            for log in logs:
                marker = f"SPS:{int(log['id'])}"
                if marker in existing_ids:
                    exported_ids.append(int(log["id"]))
                    continue
                self._copy_row_style(ws, template_row, row, 1, WORKLOG_ID_COLUMN)
                self._write_worklog_row(ws, row, log, marker)
                exported_ids.append(int(log["id"]))
                existing_ids.add(marker)
                template_row = row
                row += 1
            if self.save_workbook_callback:
                self.save_workbook_callback(workbook, path)
            else:
                workbook.save(path)
            self._worklog_cache_stamp = None
            self.repo.mark_exported(exported_ids)
            return len(exported_ids)
        finally:
            if workbook is not None:
                workbook.close()
            if self.release_excel_lock_callback:
                self.release_excel_lock_callback(lock_path)

    @staticmethod
    def _next_worklog_row(ws) -> int:
        for row in range(WORKLOG_FIRST_DATA_ROW, ws.max_row + 2):
            if all(ws.cell(row=row, column=col).value in (None, "") for col in (5, 6, 7)):
                return row
        return max(WORKLOG_FIRST_DATA_ROW, ws.max_row + 1)

    @staticmethod
    def _copy_row_style(ws, source_row: int, target_row: int, start_col: int, end_col: int) -> None:
        if source_row == target_row:
            return
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
        for col in range(start_col, end_col + 1):
            src = ws.cell(source_row, col)
            dst = ws.cell(target_row, col)
            if src.has_style:
                dst._style = copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
            dst.protection = copy(src.protection)

    @staticmethod
    def _write_worklog_row(ws, row: int, log: sqlite3.Row, marker: str) -> None:
        values = {
            1: log["work_date"] or "", 2: log["shift_group"] or "", 3: log["shift_name"] or "",
            4: log["worker"] or "", 5: log["current_step"] or "", 6: log["manage_no"] or "",
            7: log["lot_no"] or "", 8: log["qty"] if log["qty"] is not None else "",
            10: f"{log['round_no']}차" if log["round_no"] else "", 11: log["oms1_raw"] or "",
            12: log["oms2_raw"] or "", 13: log["program_name"] or "",
            14: "이상없음" if log["status"] == "완료" else "이상 있음",
            15: "OK" if log["status"] == "완료" else "NG",
            18: SimtekSpsController._excel_time(log["record_time"]),
            WORKLOG_ID_COLUMN: marker,
        }
        for col, value in values.items():
            ws.cell(row, col).value = value
        if isinstance(values[8], (int, float)):
            ws.cell(row, 9).value = f"=H{row}*0.2"
        else:
            ws.cell(row, 9).value = ""
        ws.cell(row, 18).number_format = "hh:mm:ss"

    @staticmethod
    def _excel_time(value: object):
        text = clean_text(value)
        if not text:
            return ""
        try:
            return datetime.strptime(text, "%H:%M:%S").time()
        except ValueError:
            return text

    def open_worklog(self) -> None:
        path = Path(clean_text(self.config.get("simtek_sps_excel_file", "")))
        if not path.is_file():
            self._alert("작업일보 열기", f"심텍 SPS 작업일보 파일을 확인하세요.\n{path}", "error")
            return
        os.startfile(path)

    def clear_inputs(self) -> None:
        for lot_number in (1, 2):
            for key, var in self.lot_vars[lot_number].items():
                entry = self.lot_entries[lot_number].get(key)
                if entry is not None and hasattr(entry, "clear"):
                    entry.clear()
                else:
                    var.set("")
            self.master_by_lot.pop(lot_number, None)
            self._show_blank_card(lot_number)
        self.match_status_var.set("LOT 2 미사용")
        self.dnc_status_var.set("대기중")

    def _show_blank_card(self, lot_number: int) -> None:
        self.app.hide_judgement_card(self.status_cards[lot_number])

    def open_rule_guide(self) -> None:
        window = tk.Toplevel(self.root)
        window.title("심텍 SPS 추천 규칙 상세")
        window.geometry("980x700")
        window.transient(self.root)
        text = scrolledtext.ScrolledText(window, wrap=tk.WORD, font=("맑은 고딕", 10), padx=16, pady=14)
        text.pack(fill=tk.BOTH, expand=True)
        text.insert(tk.END, self.rule_guide_text())
        text.configure(state="disabled")
        ttk.Button(window, text="닫기", command=window.destroy, width=18).pack(pady=10)

    def rule_guide_text(self) -> str:
        path = clean_text(self.config.get("simtek_sps_rule_file", "")) or str(self.default_rule_file or "")
        return (
            "심텍 SPS 추천 규칙 운영 기준\n\n"
            "1. 추천과 확정은 다릅니다.\n"
            "규칙 Excel은 신규 조건에서 참고할 Program을 보여줍니다. 추천만으로 일반 DNC는 실행되지 않습니다.\n"
            "초품 4Point와 4면 Burr 확인을 통과하고 조장명이 기록된 조건만 확정 마스터가 됩니다.\n\n"
            "2. 기존 모델 자동 조회\n"
            "작업일보 정상 이력의 P/G와 신규 검증 DB의 활성 마스터 P/G가 정확히 같을 때만 조건 조회 OK가 됩니다.\n"
            "한쪽 이력이 없거나 P/G가 다르면 조건 조회 NG이며 신규 조건 검증 DNC가 필요합니다. "
            "Rule 구축 이후 작업일보에 다른 P/G가 1건이라도 추가되면 변경 신호로 차단합니다.\n"
            "과거 작업일보의 다른 P/G 1건 PASS는 2026-08-08 최초 Rule 정리 때만 적용한 기준입니다.\n\n"
            "3. 신규 검증 추천 순서\n"
            "① 관리번호 전용 규칙\n② 제품군 + 차수 + OMS1 SIZE + OMS2 조건\n"
            "③ 제품군 일반 규칙(ANY)\n④ 과거 INPUT 이력 후보(자동 확정 금지)\n⑤ 추천 규칙 조건 없음\n\n"
            "4. OMS 조건 비교\n"
            "공백과 표시용 특수문자의 차이는 정리하지만 숫자, SIZE, CENTER, 방향홀, 우측 10mm, 후가공 차수, "
            "뽕따기 금지, STACKING, X/Y SHIFT, 좌우 Edge R은 보존합니다. 원문도 DB에 함께 저장합니다.\n\n"
            "5. STEP 규칙\n"
            "제공된 V15 작업일보 VBA 기준으로 현 STEP은 시작 STEP +60 또는 +70만 허용합니다. "
            "기준 밖이면 추천과 DNC를 차단합니다.\n\n"
            "6. 자동 DNC 차단 조건\n"
            "활성 마스터 없음, 동일 조건 마스터 2개 이상, LOT 중복, 2LOT 조건 불일치, OMS1 SIZE 미추출, "
            "OMS2 누락, STEP 불일치, Program 후보 2개 이상, 실제 .txt 파일 0개 또는 2개 이상, 작업일보/전송 경로 오류.\n\n"
            "7. 실제 파일 검증\n"
            "Program 이름과 확장자를 제외한 .txt 파일명이 정확히 일치해야 합니다. 부분 일치와 .dnc 파일은 사용하지 않습니다.\n\n"
            f"현재 규칙 파일: {path or '미설정'}\n"
        )

    def open_master_popup(self) -> None:
        window = tk.Toplevel(self.root)
        window.title("심텍 SPS 조건 마스터 관리")
        window.geometry("1250x650")
        window.transient(self.root)
        columns = ("id", "state", "manage", "round", "start", "current", "size", "oms2", "program", "leader", "date")
        tree = ttk.Treeview(window, columns=columns, show="headings")
        headings = ("ID", "상태", "관리번호", "차수", "시작", "현", "SIZE", "OMS2 KEY", "Program", "조장", "등록일")
        widths = (55, 70, 150, 55, 60, 60, 80, 100, 320, 90, 145)
        for col, heading, width in zip(columns, headings, widths):
            tree.heading(col, text=heading)
            tree.column(col, width=width, anchor="center" if col != "program" else "w")
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        def refresh():
            tree.delete(*tree.get_children())
            for row in self.repo.list_masters():
                tree.insert("", tk.END, values=(
                    row["id"], "사용" if row["active"] else "제외", row["manage_no"], f"{row['round_no']}차",
                    row["start_step"], row["current_step"], row["size_key"], row["oms2_key"], row["program_name"],
                    row["leader_name"] or "", row["registered_at"],
                ))

        def set_active(active: bool):
            selected = tree.selection()
            if not selected:
                self._alert("조건 마스터 관리", "대상을 선택하세요.", "warning", window)
                return
            record_id = int(tree.item(selected[0], "values")[0])
            reason = ""
            if not active:
                reason = simpledialog.askstring("마스터 등록 제외", "제외 사유를 입력하세요.", parent=window) or ""
                if not reason:
                    return
            self.repo.set_master_active(record_id, active, reason)
            refresh()

        buttons = tk.Frame(window)
        buttons.pack(fill=tk.X, padx=10, pady=(0, 10))
        ttk.Button(buttons, text="마스터 등록 제외", command=lambda: set_active(False), style="SideDanger.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons, text="제외 복구", command=lambda: set_active(True)).pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons, text="새로고침", command=refresh).pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons, text="닫기", command=window.destroy).pack(side=tk.RIGHT, padx=5)
        refresh()

    def open_history_popup(self) -> None:
        window = tk.Toplevel(self.root)
        window.title("심텍 SPS 작업 이력")
        window.geometry("1300x650")
        window.transient(self.root)
        columns = ("id", "type", "status", "date", "worker", "manage", "lot", "step", "round", "qty", "program", "excel")
        tree = ttk.Treeview(window, columns=columns, show="headings")
        headings = ("ID", "구분", "상태", "작업일", "작업자", "관리번호", "LOT", "STEP", "차수", "매수", "Program", "Excel")
        widths = (50, 80, 65, 90, 80, 150, 120, 85, 55, 55, 320, 60)
        for col, heading, width in zip(columns, headings, widths):
            tree.heading(col, text=heading)
            tree.column(col, width=width, anchor="center" if col != "program" else "w")
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        for row in self.repo.recent_logs():
            tree.insert("", tk.END, values=(
                row["id"], row["dnc_type"], row["status"], row["work_date"], row["worker"], row["manage_no"],
                row["lot_no"], f"{row['start_step']}→{row['current_step']}", f"{row['round_no']}차", row["qty"],
                row["program_name"], "완료" if row["exported"] else "미반영",
            ))
        ttk.Button(window, text="닫기", command=window.destroy, width=18).pack(pady=(0, 10))

    def export_history_excel(self, path: Path) -> Path:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "심텍 SPS 작업이력"
        headers = [
            "ID", "구분", "상태", "설비", "작업일", "조", "근무", "작업자", "조장",
            "시작 STEP", "현 STEP", "차수", "관리번호", "LOT", "매수", "OMS 조건1",
            "OMS 조건2", "SIZE", "OMS2 KEY", "Program", "초품", "Burr", "Excel 반영", "생성일시",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in reversed(self.repo.recent_logs(limit=100000)):
            ws.append([
                row["id"], row["dnc_type"], row["status"], row["machine"], row["work_date"],
                row["shift_group"], row["shift_name"], row["worker"], row["leader_name"],
                row["start_step"], row["current_step"], row["round_no"], row["manage_no"], row["lot_no"],
                row["qty"], row["oms1_raw"], row["oms2_raw"], row["size_key"], row["oms2_key"],
                row["program_name"], row["first_article_result"], row["burr_result"],
                "완료" if row["exported"] else "미반영", row["created_at"],
            ])
        workbook.save(path)
        workbook.close()
        return Path(path)

    def export_master_excel(self, path: Path) -> Path:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "심텍 SPS 조건마스터"
        headers = [
            "ID", "상태", "관리번호", "차수", "시작 STEP", "현 STEP", "STEP 차이",
            "OMS 조건1 원문", "OMS 조건2 원문", "SIZE", "OMS2 KEY", "조건2 분석 JSON",
            "Program", "조장", "등록 근거", "등록일", "제외일", "제외 사유",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in reversed(self.repo.list_masters()):
            ws.append([
                row["id"], "사용" if row["active"] else "제외", row["manage_no"], row["round_no"],
                row["start_step"], row["current_step"], row["step_delta"], row["oms1_raw"], row["oms2_raw"],
                row["size_key"], row["oms2_key"], row["features_json"], row["program_name"],
                row["leader_name"], row["source_text"], row["registered_at"], row["excluded_at"], row["excluded_reason"],
            ])
        workbook.save(path)
        workbook.close()
        return Path(path)


class SpsNewModelPopup:
    def __init__(self, controller: SimtekSpsController, common: dict[str, str], lots: list[SpsLot]):
        self.controller = controller
        self.common = common
        self.lots = lots
        self.window = tk.Toplevel(controller.root)
        self.window.title("심텍 SPS 신규 조건 검증 DNC")
        self.window.geometry("1280x850")
        self.window.minsize(1180, 800)
        self.window.transient(controller.root)
        self.program_var = tk.StringVar()
        self.recommendation_basis = tk.StringVar()
        self.status_var = tk.StringVar(value="최종 조건 확인: 대기")
        self.validation_var = tk.StringVar()
        self.status_label: tk.Label | None = None
        self.validation_label: tk.Label | None = None
        self.run_button: ttk.Button | None = None
        self.can_run = False
        self.confirmed_program = ""
        self.recommendation_text: dict[int, str] = {}
        self._prepare_recommendations()
        self._build()
        self.program_var.trace_add("write", self._on_program_changed)
        self._refresh_validation_preview()

    def _prepare_recommendations(self) -> None:
        self.can_run = False
        selected_programs: list[str] = []
        bases: list[str] = []
        for lot in self.lots:
            recommendation = self.controller._recommend(lot)
            programs = list(
                dict.fromkeys(
                    self.controller.resolve_program_name(value)
                    for value in recommendation.programs
                    if clean_text(value)
                )
            )
            program_log = ", ".join(programs) if programs else "추천 없음"
            warning_log = " / ".join(recommendation.warnings)
            append_log = getattr(self.controller, "_append_log", None)
            if callable(append_log):
                append_log(
                    f"LOT {lot.number} 추천 판정: {recommendation.basis} / {program_log}"
                    + (f" / {warning_log}" if warning_log else "")
                )
            if len(programs) == 1:
                self.recommendation_text[lot.number] = programs[0]
                selected_programs.append(programs[0])
            elif not programs:
                self.recommendation_text[lot.number] = "추천 규칙 조건 없음"
            else:
                self.recommendation_text[lot.number] = "추천 규칙 조건 복수 - 관리자 확인 필요"
            bases.append(f"LOT {lot.number}: {recommendation.basis}")

        unique_programs = list(dict.fromkeys(selected_programs))
        if len(selected_programs) == len(self.lots) and len(unique_programs) == 1:
            self.program_var.set(unique_programs[0])
        elif len(unique_programs) > 1:
            for lot in self.lots:
                self.recommendation_text[lot.number] = "LOT별 추천 조건 불일치"
        self.recommendation_basis.set(" / ".join(bases) if bases else "추천 규칙 조건 없음")

    def _build_lot_panel(self, parent, lot: SpsLot, column: int) -> None:
        theme = self.controller.theme
        panel = tk.Frame(parent, bg="#ffffff", highlightthickness=1, highlightbackground=theme["border"], bd=0)
        panel.grid(row=0, column=column, sticky="nsew", padx=(0, 8) if column == 0 else (8, 0), pady=8)
        first_row = 0
        if len(self.lots) == 2:
            tk.Label(
                panel,
                text=f"LOT {lot.number} 신규 입력",
                bg=theme["light"],
                fg=theme["primary"],
                font=("맑은 고딕", 11, "bold"),
                height=2,
            ).grid(row=0, column=0, columnspan=2, sticky="ew")
            first_row = 1
        panel.columnconfigure(0, weight=1, uniform=f"sps_new_{lot.number}")
        panel.columnconfigure(1, weight=1, uniform=f"sps_new_{lot.number}")
        LabeledEntry = self.controller.ui_components["LabeledEntry"]
        fields = (
            ("시작 STEP", str(lot.start_step), first_row, 0, 1),
            ("현 STEP", str(lot.current_step), first_row, 1, 1),
            ("관리번호", lot.manage_no, first_row + 1, 0, 1),
            ("LOT No", lot.lot_no, first_row + 1, 1, 1),
            ("매수", str(lot.qty), first_row + 2, 0, 1),
            ("차수", f"{lot.round_no}차", first_row + 2, 1, 1),
            ("조건 1", lot.oms1, first_row + 3, 0, 1),
            ("조건 2", lot.oms2, first_row + 3, 1, 1),
            ("추천 조건", self.recommendation_text[lot.number], first_row + 4, 0, 2),
        )
        for label, value, row, field_column, span in fields:
            entry = LabeledEntry(panel, label, width=24, readonly=True, style="Lookup.TEntry")
            entry.set(value)
            entry.grid(row=row, column=field_column, columnspan=span, sticky="ew", padx=10, pady=8)

    def _build(self) -> None:
        theme = self.controller.theme
        self.window.configure(bg=theme["bg"])
        tk.Label(
            self.window,
            text="심텍 SPS 신규 조건 검증 DNC",
            bg=theme["light"],
            fg=theme["primary"],
            font=("맑은 고딕", 14, "bold"),
            height=2,
        ).pack(fill=tk.X, padx=14, pady=(14, 8))
        selector = tk.Frame(self.window, bg="#ffffff", highlightthickness=1, highlightbackground=theme["border"], bd=0)
        selector.pack(fill=tk.X, padx=14, pady=(0, 6))
        selector_text = "LOT 1 + LOT 2" if len(self.lots) == 2 else f"LOT {self.lots[0].number} 입력"
        tk.Label(selector, text=selector_text, bg="#ffffff", fg=theme["primary"], font=("맑은 고딕", 10, "bold")).pack(side=tk.LEFT, padx=14, pady=8)

        body = tk.Frame(self.window, bg=theme["bg"])
        body.pack(fill=tk.X, padx=14, pady=4)
        for column in range(len(self.lots)):
            body.columnconfigure(column, weight=1, uniform="sps_new_lots")
        for column, lot in enumerate(self.lots):
            self._build_lot_panel(body, lot, column)

        final_frame = tk.Frame(
            self.window,
            bg="#ffffff",
            highlightthickness=1,
            highlightbackground=theme["border"],
            bd=0,
        )
        final_frame.pack(fill=tk.X, padx=14, pady=6)
        final_frame.columnconfigure(1, weight=1)
        tk.Label(
            final_frame,
            text="최종 Trim P/G",
            bg=theme["light"],
            fg=theme["primary"],
            width=16,
            font=("맑은 고딕", 10, "bold"),
        ).grid(row=0, column=0, sticky="nsew", padx=(10, 0), pady=(10, 6))
        ttk.Entry(final_frame, textvariable=self.program_var, style="Lookup.TEntry").grid(
            row=0, column=1, sticky="ew", padx=8, pady=(10, 6), ipady=4
        )
        ttk.Button(final_frame, text="조건 목록 선택", command=self.open_program_picker, width=18).grid(
            row=0, column=2, padx=(0, 8), pady=(10, 6)
        )
        ttk.Button(final_frame, text="조건 확인·확정", command=self.confirm_program, style="Primary.TButton", width=18).grid(
            row=0, column=3, padx=(0, 10), pady=(10, 6)
        )
        expected_size = self.lots[0].parsed.size_key or "확인 불가"
        expected_shift = format_shift(extract_oms_program_shift(self.lots[0].oms2))
        tk.Label(
            final_frame,
            text=f"필수 조건  |  사이즈 {expected_size}  |  시프트 {expected_shift}",
            bg="#ffffff",
            fg="#1f2937",
            anchor="w",
            font=("맑은 고딕", 10, "bold"),
        ).grid(row=1, column=0, columnspan=4, sticky="ew", padx=16, pady=(2, 2))
        self.validation_label = tk.Label(
            final_frame,
            textvariable=self.validation_var,
            bg="#ffffff",
            fg="#dc2626",
            anchor="w",
            font=("맑은 고딕", 10),
        )
        self.validation_label.grid(row=2, column=0, columnspan=4, sticky="ew", padx=16, pady=(2, 10))

        status = tk.Frame(self.window, bg="#ffffff", highlightthickness=1, highlightbackground=theme["border"], bd=0)
        status.pack(fill=tk.X, padx=14, pady=6)
        self.status_label = tk.Label(
            status,
            textvariable=self.status_var,
            bg="#ffffff",
            fg="#dc2626",
            anchor="w",
            font=("맑은 고딕", 10, "bold"),
        )
        self.status_label.pack(side=tk.LEFT, padx=16, pady=10)

        buttons = tk.Frame(self.window, bg=theme["bg"])
        buttons.pack(fill=tk.X, padx=14, pady=(4, 14))
        self.run_button = ttk.Button(
            buttons,
            text="신규 조건 DNC 실행",
            command=lambda: self.controller.run_new_model_from_popup(self, self.program_var.get()),
            style="Primary.TButton",
            width=22,
            state="disabled",
        )
        self.run_button.pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons, text="닫기", command=self.window.destroy, width=18).pack(side=tk.RIGHT, padx=5)

    def _set_status(self, text: str, color: str) -> None:
        self.status_var.set(text)
        if self.status_label:
            self.status_label.configure(fg=color)

    def _validation_checks(self, program: str):
        controller = getattr(self, "controller", None)
        if controller is None:
            return [validate_program_condition(lot.parsed, program) for lot in self.lots]
        resolved = controller.resolve_program_name(program)
        option = controller.get_program_option(resolved)
        checks: list[ProgramConditionCheck] = []
        for lot in self.lots:
            if option is None:
                checks.append(
                    ProgramConditionCheck(
                        False,
                        lot.parsed.size_key,
                        "",
                        extract_oms_program_shift(lot.oms2),
                        None,
                        ("실제 P/G 기준 목록에 등록되지 않은 조건입니다.",),
                    )
                )
                continue
            reason = condition_mismatch_reason(
                option,
                lot.parsed.size_key,
                extract_oms_program_shift(lot.oms2),
            )
            checks.append(
                ProgramConditionCheck(
                    not reason,
                    lot.parsed.size_key,
                    option.size_key,
                    extract_oms_program_shift(lot.oms2),
                    option.shift,
                    (reason,) if reason else (),
                )
            )
        return checks

    def _refresh_validation_preview(self) -> None:
        program = clean_text(self.program_var.get())
        if not program:
            self.validation_var.set("조건 목록에서 선택하거나 Trim P/G를 직접 입력하세요.")
            if self.validation_label:
                self.validation_label.configure(fg="#dc2626")
            return
        checks = self._validation_checks(program)
        failures = [f"LOT {lot.number}: {check.summary}" for lot, check in zip(self.lots, checks) if not check.ok]
        if failures:
            self.validation_var.set(" / ".join(failures))
            if self.validation_label:
                self.validation_label.configure(fg="#dc2626")
            return
        check = checks[0]
        self.validation_var.set(
            f"사이즈 {check.program_size} 일치 / 시프트 {format_shift(check.program_shift)} 일치 / 조장 확정 필요"
        )
        if self.validation_label:
            self.validation_label.configure(fg="#2563eb")

    def _on_program_changed(self, *_args) -> None:
        program = clean_text(self.program_var.get())
        if program != self.confirmed_program:
            self.confirmed_program = ""
            self.can_run = False
            if self.run_button:
                self.run_button.configure(state="disabled")
            self._set_status("최종 조건 확인: 대기", "#dc2626")
        self._refresh_validation_preview()

    def open_program_picker(self) -> None:
        catalog = self.controller.load_program_catalog()
        if not catalog:
            self.controller._alert(
                "조건 목록 확인",
                "심텍 SPS P/G 기준 목록을 불러오지 못했습니다. 관리자에게 확인하세요.",
                "error",
                parent=self.window,
            )
            return
        SpsProgramPicker(self, catalog)

    def apply_program_selection(self, program: str) -> None:
        self.program_var.set(self.controller.resolve_program_name(program))
        self._set_status("조건 선택 완료 / 조건 확인·확정을 누르세요.", "#2563eb")

    def confirm_program(self) -> bool:
        program = self.controller.resolve_program_name(self.program_var.get())
        if program != clean_text(self.program_var.get()):
            self.program_var.set(program)
        checks = self._validation_checks(program)
        failures = [f"LOT {lot.number}: {check.summary}" for lot, check in zip(self.lots, checks) if not check.ok]
        if failures:
            self.confirmed_program = ""
            self.can_run = False
            if self.run_button:
                self.run_button.configure(state="disabled")
            message = "사이즈·시프트가 맞지 않아 확정할 수 없습니다.\n\n" + "\n".join(failures)
            self._set_status("최종 조건 확인: NG", "#dc2626")
            self.controller._alert("조건 선택 불가", message, "error", parent=self.window)
            return False
        self.confirmed_program = program
        self.can_run = True
        if self.run_button:
            self.run_button.configure(state="normal")
        self.validation_var.set(checks[0].summary)
        if self.validation_label:
            self.validation_label.configure(fg="#059669")
        self._set_status("최종 조건 확인: OK / 신규 조건 DNC 실행 가능", "#059669")
        return True

    def require_confirmed_program(self, program: str) -> str:
        controller = getattr(self, "controller", None)
        if controller is not None:
            program = controller.resolve_program_name(program)
        else:
            program = clean_text(program)
        if not self.can_run or not self.confirmed_program or program != self.confirmed_program:
            raise ValueError("최종 Trim P/G를 선택한 뒤 조건 확인·확정을 먼저 누르세요.")
        failures = [check.summary for check in self._validation_checks(program) if not check.ok]
        if failures:
            self.can_run = False
            self.confirmed_program = ""
            if self.run_button:
                self.run_button.configure(state="disabled")
            raise ValueError("사이즈·시프트 재확인 NG: " + " / ".join(failures))
        return program

    def registration_basis(self) -> str:
        selected = self.controller.resolve_program_name(self.program_var.get())
        recommended = {
            self.controller.resolve_program_name(value)
            for value in self.recommendation_text.values()
            if clean_text(value)
        }
        selection = "추천 조건 조장 확정" if recommended == {selected} else "조장 직접 선택"
        return f"{self.recommendation_basis.get()} / {selection} / 사이즈·시프트 일치 확인"

    def set_running(self, running: bool) -> None:
        if self.run_button:
            self.run_button.configure(state="disabled" if running or not self.can_run else "normal")
        if running:
            self._set_status("신규 검증 진행중", "#2563eb")
        elif self.can_run:
            self._set_status("최종 조건 확인: OK / 신규 조건 DNC 실행 가능", "#059669")
        else:
            self._set_status("최종 조건 확인: 대기", "#dc2626")

    def finish(self, ok: bool) -> None:
        self.set_running(False)
        self._set_status("조건 마스터 등록: OK" if ok else "조건 마스터 미등록: 검증 NG", "#059669" if ok else "#dc2626")


class SpsProgramPicker:
    def __init__(self, owner: SpsNewModelPopup, catalog: list[SpsProgramOption]):
        self.owner = owner
        self.catalog = catalog
        self.window = tk.Toplevel(owner.window)
        self.window.title("심텍 SPS Trim P/G 조건 선택")
        self.window.geometry("1240x780")
        self.window.minsize(1080, 680)
        self.window.transient(owner.window)
        self.search_var = tk.StringVar()
        self.count_var = tk.StringVar()
        self.selected_name_var = tk.StringVar(value="선택된 Trim P/G 없음")
        self.selected_status_var = tk.StringVar(value="선택 가능 여부: 대기")
        self.selected_reason_var = tk.StringVar(value="목록에서 조건을 한 번 클릭하세요.")
        self.notebook: ttk.Notebook | None = None
        self.trees: dict[str, ttk.Treeview] = {}
        self.options_by_item: dict[tuple[str, str], SpsProgramOption] = {}
        self.selected_option: SpsProgramOption | None = None
        self._build()
        self.search_var.trace_add("write", self._filter_catalog)
        self._select_current_program()
        self.window.lift()

    @property
    def expected_size(self) -> str:
        return self.owner.lots[0].parsed.size_key

    @property
    def expected_shift(self) -> float | None:
        return extract_oms_program_shift(self.owner.lots[0].oms2)

    @staticmethod
    def _same_shift(left: float | None, right: float | None) -> bool:
        if left is None or right is None:
            return left is right
        return abs(left - right) <= 0.001

    def _ordered_sizes(self) -> list[str]:
        sizes = sorted({option.size_key for option in self.catalog if option.size_key})
        if self.expected_size and self.expected_size not in sizes:
            sizes.append(self.expected_size)
            sizes.sort()
        if self.expected_size in sizes:
            sizes.remove(self.expected_size)
            sizes.insert(0, self.expected_size)
        if any(not option.size_key for option in self.catalog):
            sizes.append("기타")
        return sizes

    def _recommended_programs(self) -> set[str]:
        programs: set[str] = set()
        for value in self.owner.recommendation_text.values():
            resolved = self.owner.controller.resolve_program_name(value)
            if self.owner.controller.get_program_option(resolved) is not None:
                programs.add(resolved)
        return programs

    def _build(self) -> None:
        theme = self.owner.controller.theme
        self.window.configure(bg=theme["bg"])

        header = tk.Frame(self.window, bg=theme["light"], highlightthickness=1, highlightbackground=theme["border"])
        header.pack(fill=tk.X, padx=14, pady=(14, 8))
        tk.Label(
            header,
            text="심텍 SPS Trim P/G 조건 선택",
            bg=theme["light"],
            fg=theme["primary"],
            font=("맑은 고딕", 15, "bold"),
        ).pack(side=tk.LEFT, padx=16, pady=14)
        self.count_label = tk.Label(
            header,
            textvariable=self.count_var,
            bg="#ffffff",
            fg="#047857",
            font=("맑은 고딕", 10, "bold"),
            padx=12,
            pady=6,
            relief="solid",
            bd=1,
        )
        self.count_label.pack(side=tk.RIGHT, padx=(6, 16), pady=10)
        tk.Label(
            header,
            text=f"필수 SHIFT  {format_shift(self.expected_shift)}",
            bg="#ffffff",
            fg="#b45309" if self.expected_shift is not None else "#374151",
            font=("맑은 고딕", 10, "bold"),
            padx=12,
            pady=6,
            relief="solid",
            bd=1,
        ).pack(side=tk.RIGHT, padx=6, pady=10)
        tk.Label(
            header,
            text=f"필수 사이즈  {self.expected_size or '확인 불가'}",
            bg="#dbeafe",
            fg="#1d4ed8",
            font=("맑은 고딕", 10, "bold"),
            padx=12,
            pady=6,
            relief="solid",
            bd=1,
        ).pack(side=tk.RIGHT, padx=6, pady=10)

        search = tk.Frame(self.window, bg="#ffffff", highlightthickness=1, highlightbackground=theme["border"], bd=0)
        search.pack(fill=tk.X, padx=14, pady=(0, 8))
        tk.Label(search, text="P/G 검색", bg="#ffffff", fg="#1f2937", font=("맑은 고딕", 10, "bold")).pack(side=tk.LEFT, padx=(14, 8), pady=10)
        ttk.Entry(search, textvariable=self.search_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8), pady=10, ipady=3)
        ttk.Button(search, text="검색 초기화", command=lambda: self.search_var.set(""), width=14).pack(side=tk.RIGHT, padx=(0, 8), pady=8)
        ttk.Button(search, text="다른 사이즈 참고 보기", command=self._open_reference, width=22).pack(side=tk.RIGHT, padx=(0, 8), pady=8)

        self.notebook = ttk.Notebook(self.window)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=14, pady=4)
        self.notebook.bind("<Control-Tab>", lambda _event: "break")
        self.notebook.bind("<Control-Shift-Tab>", lambda _event: "break")
        for size_key in self._ordered_sizes():
            frame = tk.Frame(self.notebook, bg="#ffffff")
            frame.rowconfigure(0, weight=1)
            frame.columnconfigure(0, weight=1)
            label = f"{size_key} (필수)" if size_key == self.expected_size else size_key
            self.notebook.add(frame, text=label)
            tab_index = self.notebook.index("end") - 1
            if size_key != self.expected_size:
                self.notebook.tab(tab_index, state="disabled")
            tree = self._build_tree(frame, selectable=size_key == self.expected_size)
            self.trees[size_key] = tree

        selected = tk.Frame(self.window, bg="#ffffff", highlightthickness=1, highlightbackground=theme["border"], bd=0)
        selected.pack(fill=tk.X, padx=14, pady=(8, 4))
        tk.Label(selected, text="선택한 Trim P/G", bg=theme["light"], fg=theme["primary"], width=18, font=("맑은 고딕", 10, "bold")).grid(row=0, column=0, rowspan=3, sticky="nsew")
        tk.Label(selected, textvariable=self.selected_name_var, bg="#ffffff", fg="#111827", anchor="w", font=("맑은 고딕", 10, "bold"), wraplength=940, justify="left").grid(row=0, column=1, sticky="ew", padx=12, pady=(8, 2))
        self.selected_status_label = tk.Label(selected, textvariable=self.selected_status_var, bg="#ffffff", fg="#374151", anchor="w", font=("맑은 고딕", 10, "bold"))
        self.selected_status_label.grid(row=1, column=1, sticky="ew", padx=12, pady=2)
        tk.Label(selected, textvariable=self.selected_reason_var, bg="#ffffff", fg="#6b7280", anchor="w", font=("맑은 고딕", 9), wraplength=940, justify="left").grid(row=2, column=1, sticky="ew", padx=12, pady=(2, 8))
        selected.columnconfigure(1, weight=1)

        footer = tk.Frame(self.window, bg=theme["bg"])
        footer.pack(fill=tk.X, padx=14, pady=(4, 14))
        tk.Label(
            footer,
            text="한 번 클릭은 선택만 합니다. 최종 전달은 '이 조건 적용' 버튼으로 진행합니다.",
            bg=theme["bg"],
            fg="#4b5563",
            font=("맑은 고딕", 9, "bold"),
        ).pack(side=tk.LEFT, padx=4)
        ttk.Button(footer, text="이 조건 적용", command=self._apply_selected, style="Primary.TButton", width=18).pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(footer, text="닫기", command=self.window.destroy, width=14).pack(side=tk.RIGHT)
        self._populate_trees("")

    def _build_tree(self, frame: tk.Frame, selectable: bool) -> ttk.Treeview:
        columns = ("status", "program", "size", "shift", "count", "recent", "file")
        tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode="browse" if selectable else "none")
        headings = {
            "status": "상태",
            "program": "Trim P/G",
            "size": "사이즈",
            "shift": "SHIFT",
            "count": "사용 횟수",
            "recent": "최근 사용일",
            "file": "실제 파일",
        }
        widths = {"status": 115, "program": 570, "size": 90, "shift": 80, "count": 90, "recent": 105, "file": 125}
        for column in columns:
            tree.heading(column, text=headings[column])
            tree.column(column, width=widths[column], minwidth=70, anchor="w" if column == "program" else "center", stretch=column == "program")
        tree.tag_configure("available", foreground="#047857", background="#ecfdf5")
        tree.tag_configure("review", foreground="#b45309", background="#fff7ed")
        tree.tag_configure("blocked", foreground="#b91c1c", background="#fef2f2")
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        if selectable:
            tree.bind("<<TreeviewSelect>>", lambda _event, current=tree: self._on_tree_select(current))
            tree.bind("<Double-1>", lambda _event: "break")
        return tree

    def _option_tag(self, option: SpsProgramOption) -> str:
        if option.status == STATUS_AVAILABLE:
            return "available"
        if option.status == STATUS_REVIEW:
            return "review"
        return "blocked"

    def _eligible_options(self) -> list[SpsProgramOption]:
        return [
            option
            for option in self.catalog
            if option.size_key == self.expected_size and self._same_shift(option.shift, self.expected_shift)
        ]

    def _display_sort_key(self, option: SpsProgramOption):
        recommended = option.program in self._recommended_programs()
        status_order = {STATUS_AVAILABLE: 0, STATUS_REVIEW: 1, STATUS_DUPLICATE_DISABLED: 2, STATUS_MISSING_FILE: 3, STATUS_BLOCKED: 4}
        recent_number = int(option.recent_date.replace("-", "") or 0)
        return (
            0 if recommended else 1,
            status_order.get(option.status, 9),
            0 if option.file_verified else 1,
            -option.history_count,
            -recent_number,
            option.program.casefold(),
        )

    def _populate_trees(self, query: str) -> None:
        query = clean_text(query).casefold()
        self.options_by_item.clear()
        self.selected_option = None
        for tree in self.trees.values():
            tree.delete(*tree.get_children())
        displayed = 0
        selectable_count = 0
        tree = self.trees.get(self.expected_size)
        if tree is not None:
            for option in sorted(self._eligible_options(), key=self._display_sort_key):
                if query and query not in option.program.casefold():
                    continue
                item_id = tree.insert(
                    "",
                    "end",
                    values=(
                        option.status_label,
                        option.program,
                        option.size_key,
                        format_shift(option.shift),
                        f"{option.history_count}회" if option.history_count else "이력 없음",
                        option.recent_date or "-",
                        option.file_status,
                    ),
                    tags=(self._option_tag(option),),
                )
                self.options_by_item[(self.expected_size, item_id)] = option
                displayed += 1
                if option.selectable:
                    selectable_count += 1
        self.count_var.set(f"선택 가능 {selectable_count}개 / 표시 {displayed}개")
        self._clear_selected()

    def _filter_catalog(self, *_args) -> None:
        self._populate_trees(self.search_var.get())

    def _clear_selected(self) -> None:
        self.selected_option = None
        self.selected_name_var.set("선택된 Trim P/G 없음")
        self.selected_status_var.set("선택 가능 여부: 대기")
        self.selected_reason_var.set("목록에서 조건을 한 번 클릭하세요.")
        self.selected_status_label.configure(fg="#374151")

    def _on_tree_select(self, tree: ttk.Treeview) -> None:
        selection = tree.selection()
        if not selection:
            self._clear_selected()
            return
        option = self.options_by_item.get((self.expected_size, selection[0]))
        if option is None:
            self._clear_selected()
            return
        self.selected_option = option
        reason = condition_mismatch_reason(option, self.expected_size, self.expected_shift)
        self.selected_name_var.set(option.program)
        if reason:
            self.selected_status_var.set(f"선택 가능 여부: 적용 불가 ({option.status_label})")
            self.selected_reason_var.set(reason)
            self.selected_status_label.configure(fg="#b91c1c")
        elif option.status == STATUS_REVIEW:
            self.selected_status_var.set("선택 가능 여부: 관리자 확인 후 가능")
            self.selected_reason_var.set(option.note or "실제 파일은 있으나 작업일보 이력이 부족합니다.")
            self.selected_status_label.configure(fg="#b45309")
        else:
            self.selected_status_var.set("선택 가능 여부: 사용 가능")
            self.selected_reason_var.set("필수 사이즈와 SHIFT가 정확히 일치합니다.")
            self.selected_status_label.configure(fg="#047857")

    def _select_current_program(self) -> None:
        current = self.owner.controller.resolve_program_name(self.owner.program_var.get())
        tree = self.trees.get(self.expected_size)
        if not current or tree is None:
            return
        for item_id in tree.get_children():
            option = self.options_by_item.get((self.expected_size, item_id))
            if option and option.program == current:
                tree.selection_set(item_id)
                tree.focus(item_id)
                tree.see(item_id)
                self._on_tree_select(tree)
                return

    def _apply_selected(self) -> None:
        option = self.selected_option
        if option is None:
            self.owner.controller._alert("조건 선택", "적용할 Trim P/G를 먼저 선택하세요.", "warning", parent=self.window)
            return
        reason = condition_mismatch_reason(option, self.expected_size, self.expected_shift)
        if reason:
            self.owner.controller._alert(
                "조건 선택 불가",
                "현재 신규 모델의 필수 조건과 일치하지 않아 적용할 수 없습니다.\n\n" + reason,
                "error",
                parent=self.window,
            )
            return
        self.owner.apply_program_selection(option.program)
        self.window.destroy()

    def _open_reference(self) -> None:
        SpsProgramReferencePopup(self)


class SpsProgramReferencePopup:
    def __init__(self, picker: SpsProgramPicker):
        self.picker = picker
        self.window = tk.Toplevel(picker.window)
        self.window.title("심텍 SPS 다른 사이즈 참고 보기")
        self.window.geometry("1160x700")
        self.window.minsize(980, 600)
        self.window.transient(picker.window)
        self.search_var = tk.StringVar()
        self.tree: ttk.Treeview | None = None
        self._build()
        self.search_var.trace_add("write", self._filter)
        self.window.lift()

    def _build(self) -> None:
        theme = self.picker.owner.controller.theme
        self.window.configure(bg="#eef1f5")
        header = tk.Frame(self.window, bg="#e5e7eb", highlightthickness=1, highlightbackground="#cbd5e1")
        header.pack(fill=tk.X, padx=14, pady=(14, 8))
        tk.Label(header, text="다른 사이즈 참고 보기", bg="#e5e7eb", fg="#374151", font=("맑은 고딕", 14, "bold")).pack(anchor="w", padx=16, pady=(12, 2))
        tk.Label(header, text="참고용 조건입니다. 현재 모델에는 적용할 수 없습니다.", bg="#e5e7eb", fg="#b91c1c", font=("맑은 고딕", 10, "bold")).pack(anchor="w", padx=16, pady=(2, 12))

        search = tk.Frame(self.window, bg="#f3f4f6", highlightthickness=1, highlightbackground="#cbd5e1")
        search.pack(fill=tk.X, padx=14, pady=(0, 8))
        tk.Label(search, text="P/G 검색", bg="#f3f4f6", fg="#374151", font=("맑은 고딕", 10, "bold")).pack(side=tk.LEFT, padx=(14, 8), pady=10)
        ttk.Entry(search, textvariable=self.search_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8), pady=10, ipady=3)
        ttk.Button(search, text="검색 초기화", command=lambda: self.search_var.set(""), width=14).pack(side=tk.RIGHT, padx=14, pady=8)

        body = tk.Frame(self.window, bg="#eef1f5")
        body.pack(fill=tk.BOTH, expand=True, padx=14, pady=4)
        body.rowconfigure(0, weight=1)
        body.columnconfigure(0, weight=1)
        columns = ("status", "program", "size", "shift", "count", "recent", "file")
        self.tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="none")
        headings = ("상태", "Trim P/G", "사이즈", "SHIFT", "사용 횟수", "최근 사용일", "실제 파일")
        widths = (120, 570, 90, 80, 90, 105, 125)
        for column, heading, width in zip(columns, headings, widths):
            self.tree.heading(column, text=heading)
            self.tree.column(column, width=width, minwidth=70, anchor="w" if column == "program" else "center", stretch=column == "program")
        self.tree.tag_configure("reference", foreground="#6b7280", background="#f3f4f6")
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind("<Double-1>", lambda _event: "break")
        y_scroll = ttk.Scrollbar(body, orient="vertical", command=self.tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(body, orient="horizontal", command=self.tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self._populate("")

        footer = tk.Frame(self.window, bg="#eef1f5")
        footer.pack(fill=tk.X, padx=14, pady=(8, 14))
        ttk.Button(footer, text="닫기", command=self.window.destroy, width=16).pack(side=tk.RIGHT)

    def _reference_options(self) -> list[SpsProgramOption]:
        return [
            option
            for option in self.picker.catalog
            if option.size_key != self.picker.expected_size
            or not self.picker._same_shift(option.shift, self.picker.expected_shift)
            or not option.selectable
        ]

    def _populate(self, query: str) -> None:
        if self.tree is None:
            return
        self.tree.delete(*self.tree.get_children())
        query = clean_text(query).casefold()
        for option in self._reference_options():
            if query and query not in option.program.casefold():
                continue
            self.tree.insert(
                "",
                "end",
                values=(
                    "참고용" if option.selectable else option.status_label,
                    option.program,
                    option.size_key or "확인 불가",
                    format_shift(option.shift),
                    f"{option.history_count}회" if option.history_count else "이력 없음",
                    option.recent_date or "-",
                    option.file_status,
                ),
                tags=("reference",),
            )

    def _filter(self, *_args) -> None:
        self._populate(self.search_var.get())
